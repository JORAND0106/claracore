"""Curva S: agregación mensual y costos de presupuesto."""
from __future__ import annotations

import io
from datetime import date, datetime
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from prog_obra_curva_s import (
    _build_curva_s_chart_block,
    _nodes_with_ppto_costs,
    _xl_curva_s_sumif_detalle,
    _xl_detalle_pk_month_formula,
    build_curva_s,
    build_curva_s_xlsx_bytes,
)


def _sample_nodes():
    return {
        "n1": {
            "pk_id": "PK1",
            "capitulo": "01",
            "agrupador_id": 10,
            "label": "2.1",
            "fecha_inicio": date(2026, 6, 1),
            "fecha_fin": date(2026, 6, 30),
            "costo_programado": 999.0,
        }
    }


@patch("prog_obra_curva_s.apply_ppto_cost_overlay")
@patch("prog_obra_curva_s.build_cost_overlay_maps")
@patch("prog_obra_curva_s.fetch_compare_nodes")
def test_nodes_with_ppto_costs_no_strict_zero(mock_fetch, mock_maps, mock_apply):
    mock_fetch.return_value = _sample_nodes()
    mock_maps.return_value = ({("PK1", "01", 10): 500_000.0}, {})
    mock_apply.side_effect = lambda nodes, ag, it: {
        k: {**v, "costo_programado": ag.get(("PK1", "01", 10), v.get("costo_programado"))}
        for k, v in nodes.items()
    }

    out = _nodes_with_ppto_costs(MagicMock(), "vid", 1, "ppto-1")
    assert float(out["n1"]["costo_programado"]) == 500_000.0
    mock_apply.assert_called_once()
    assert mock_apply.call_args.kwargs.get("strict") is None
    assert len(mock_apply.call_args.args) == 3


@patch("prog_obra_curva_s.build_brecha_presupuesto_programacion")
@patch("prog_obra_curva_s._fetch_ejecutado_mensual")
@patch("prog_obra_curva_s._apply_ppto_scope_total_scale")
@patch("prog_obra_curva_s._aggregate_version_monthly")
@patch("prog_obra_curva_s.resolve_ppto_vigente_version_id")
@patch("prog_obra_curva_s.fetch_baseline_version_id")
@patch("prog_obra_curva_s.fetch_vigente_meta")
@patch("prog_obra_curva_s.ppto_scope_direct_total")
def test_build_curva_s_nonzero_totals(
    mock_scope, mock_vigente, mock_baseline, mock_ppto, mock_agg, mock_scale, mock_ej, mock_brecha
):
    mock_vigente.return_value = ("target-vid", {})
    mock_baseline.return_value = "baseline-vid"
    mock_ppto.return_value = "ppto-vigente-id"
    mock_agg.return_value = ({"2026-06": 250_000.0}, 250_000.0)
    mock_scale.side_effect = lambda _sb, _c, m, t, *_a, **_k: (m, t)
    mock_ej.return_value = ({"2026-06": 50_000.0}, 50_000.0)
    mock_scope.return_value = 250_000.0
    mock_brecha.return_value = {
        "presupuesto_total": 250_000.0,
        "programado_total": 250_000.0,
        "diferencia": 0.0,
        "tiene_brecha": False,
    }

    data = build_curva_s(MagicMock(), 1, version_ppto_id="ppto-sellada")

    assert data["indicadores"]["presupuesto_total"] == 250_000.0
    assert data["meses"][0]["baseline_mes"] == 250_000.0
    assert data["meses"][0]["vigente_mes"] == 250_000.0
    assert data["version_ppto_id"] == "ppto-vigente-id"
    mock_ppto.assert_called_once()
    assert mock_ppto.call_args.kwargs.get("force_vigente") is True
    assert mock_scale.call_count == 1


def test_xl_formula_helpers():
    assert "SUMIF" in _xl_curva_s_sumif_detalle("Detalle PK", "Vigente", "H", 5, 10)
    assert "EOMONTH" in _xl_detalle_pk_month_formula(5, "H", hdr_row=4)
    assert "$F5" in _xl_detalle_pk_month_formula(5, "H", hdr_row=4)


def test_build_curva_s_xlsx_uses_formulas():
    contrato = {"numero": "C-1", "id": 3, "objeto": "Test", "contratista": "X", "interventoria": "Y"}
    data = {
        "meses": [
            {
                "mes": "2026-06",
                "mes_label": "Jun 2026",
                "baseline_mes": 100,
                "baseline_acum": 100,
                "vigente_mes": 100,
                "vigente_acum": 100,
                "ejecutado_mes": 50,
                "ejecutado_acum": 50,
                "delta_vigente_pct": 0,
                "delta_ejecutado_pct": -50,
            }
        ],
        "detalle_pk": {
            "vigente": [
                {
                    "pk_id": "PK1",
                    "capitulo": "01",
                    "agrupador_id": 1,
                    "label": "2.A",
                    "agrupador_nombre": "Excavacion",
                    "fecha_inicio": date(2026, 6, 1),
                    "fecha_fin": date(2026, 6, 30),
                    "costo_total": 1000.0,
                    "distribucion_mensual": {"2026-06": 1000.0},
                }
            ]
        },
        "baseline_id": "v1",
        "target_id": "v2",
        "indicadores": {"presupuesto_total": 1000.0},
    }
    raw = build_curva_s_xlsx_bytes(
        contrato,
        data,
        prog_meta={"numero_version": 1, "tipo": "borrador"},
        fecha_generacion="2026-06-08",
    )
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Curva S"]
    ws_det = wb["Detalle PK"]

    assert str(ws["D21"].value).startswith("=SUMIF")
    assert str(ws["C21"].value).startswith("=B21") or str(ws["C21"].value).startswith("=SUM")
    assert str(ws["E21"].value).startswith("=D21") or str(ws["E21"].value).startswith("=SUM")
    assert str(ws["H21"].value).startswith("=IF")

    assert str(ws_det["H5"].value).startswith("=IF")
    assert ws_det["E5"].value == 1000.0
    assert ws_det["F5"].value in (date(2026, 6, 1), datetime(2026, 6, 1))
    assert str(ws["A6"].value).startswith("=SUMIF")
    assert str(ws["B6"].value).startswith("=E")
    assert "image/svg+xml" in _build_curva_s_chart_block(
        [{"mes": "2026-06", "mes_label": "Jun", "baseline_acum": 1, "vigente_acum": 1, "ejecutado_acum": 0}]
    )

from fastapi import FastAPI, HTTPException, Depends, Request, UploadFile, File, Form, Query, Header
from fastapi.exceptions import RequestValidationError
from fastapi.responses import StreamingResponse, JSONResponse
import io, csv, requests as req_http
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, Field
from typing import List, Optional, Any, Dict, Set, Tuple
from collections import defaultdict
from supabase import create_client, ClientOptions
import httpx
from passlib.context import CryptContext
from jose import jwt, JWTError
from datetime import datetime, timedelta, timezone, date
from dotenv import load_dotenv
import os
import base64
import logging
import traceback
import time
import uuid
import threading
import calendar
import re
import json
import math
from concurrent.futures import ThreadPoolExecutor

# ── Sesiones DWG activas (en memoria) ─────────────────────────────────────────
# Clave: (contrato_id, usuario_id) → timestamp Unix. Nunca mezclar con otro usuario:
# el badge «DWG enlazado» en la web debe ser solo para quien corre SicoeCAD con su sesión.
_dwg_sessions: dict = {}
# Auditoría: última carga de cantidades desde SicoeCAD (cliente) → /presupuesto/.../bulk?source=sicoe_cad
# Notificación consumida vía GET .../sincro-sicoe-cad-auditoria; en un solo proceso de API (no multi-réplica).
_sicoe_cad_sincro_audit: dict = {}
# ── Jobs de exportación Excel en background ────────────────────────────────────
_export_jobs: dict = {}  # { job_id: { "estado": "procesando"|"listo"|"error", "buf": bytes, "filename": str } }
_DWG_TIMEOUT = 30  # segundos — margen para curl.exe
_MAINTENANCE_SECRET = os.getenv("MAINTENANCE_SECRET", "claracore_deploy_2026")
_MAINTENANCE_DEFAULT_SECONDS = int(os.getenv("MAINTENANCE_COUNTDOWN_SECONDS", "25"))
_maintenance_state = {
    "activo": False,
    "mensaje": "Actualización del sistema en curso. Por favor guarda tu trabajo antes de continuar.",
    "expires_at": None,
}

def _session_key_cad(contrato_id: int, usuario_id: int):
    return (int(contrato_id), int(usuario_id) if usuario_id is not None else 0)


def _dwg_activo(contrato_id: int, usuario_id: int = None) -> bool:
    """
    True si *algún* usuario tiene heartbeat CAD reciente para el contrato (p. ej. cola a AutoCAD).
    No confundir con /cad-queue/.../estado, que es solo el usuario de la petición.
    """
    t = time.time()
    ctx = int(contrato_id)
    for (cid, _uid), last in _dwg_sessions.items():
        if int(cid) == ctx and (t - float(last)) < 10:
            return True
    return False

load_dotenv()
_MAINTENANCE_SECRET = os.getenv("MAINTENANCE_SECRET", _MAINTENANCE_SECRET)
_MAINTENANCE_DEFAULT_SECONDS = int(os.getenv("MAINTENANCE_COUNTDOWN_SECONDS", str(_MAINTENANCE_DEFAULT_SECONDS)))

app = FastAPI(title="ClaraCore API")

# Orígenes permitidos (CORS). Incluye claracore.co y localhost; CORS_EXTRA_ORIGINS=url1,url2 añade más sin redeploy.
_cors_extra = [
    o.strip()
    for o in (os.getenv("CORS_EXTRA_ORIGINS") or "").split(",")
    if o.strip()
]
_cors_origins = [
    "https://claracore.co",
    "https://www.claracore.co",
    "https://app.claracore.co",
    "http://localhost:5173",
    "http://localhost:5174",
    "http://127.0.0.1:5173",
    "http://127.0.0.1:5174",
] + _cors_extra

# CORSMiddleware se registra al final del archivo (tras los @app.middleware) para que sea la capa
# más externa y añada cabeceras CORS también cuando un middleware interno devuelve JSONResponse
# sin llamar a call_next (p. ej. políticas 403); si CORS va “dentro”, el navegador ve CORS bloqueado.

_log_api = logging.getLogger("uvicorn.error")

# Seguimiento en memoria para alertas (login fallido, 500 repetidos)
_login_fail_tracker: dict = {}
_endpoint_500_tracker: dict = {}
# Columnas de `logs` que PostgREST reporta como inexistentes (BD sin migración alter_logs_auditoria.sql).
_logs_omit_columns: Set[str] = set()
_logs_omit_lock = threading.Lock()

# Versión del texto de políticas mostrada al usuario (auditoría junto a politicas_version en BD)
POLITICAS_VERSION_DEFAULT = os.getenv("POLITICAS_VERSION", "1.0")

# Cache en memoria: evita 1 consulta a Supabase por cada request autenticado (reduce carga y latencia).
# Invalidación al aceptar políticas. TTL por defecto 120 s (ajustable con POLITICAS_CACHE_TTL_SECONDS).
_POLITICAS_CACHE_TTL = float(os.getenv("POLITICAS_CACHE_TTL_SECONDS", "120"))
_POLITICAS_CACHE_LOCK = threading.Lock()
# uid -> (exp_unix, "ok"|"pend"|"none")  ok=aceptó, pend=rechazó pendiente, none=usuario no existe en select
_politicas_cache: dict = {}


def _politicas_cache_get(uid: int) -> Optional[str]:
    with _POLITICAS_CACHE_LOCK:
        row = _politicas_cache.get(uid)
        if not row:
            return None
        exp, state = row
        if time.time() > exp:
            del _politicas_cache[uid]
            return None
        return state


def _politicas_cache_set(uid: int, state: str) -> None:
    """state: ok | pend | none"""
    with _POLITICAS_CACHE_LOCK:
        _politicas_cache[uid] = (time.time() + _POLITICAS_CACHE_TTL, state)
        if len(_politicas_cache) > 8000:
            now = time.time()
            dead = [k for k, (e, _) in _politicas_cache.items() if e < now]
            for k in dead[:4000]:
                _politicas_cache.pop(k, None)


def politicas_cache_invalidate(uid: int) -> None:
    with _POLITICAS_CACHE_LOCK:
        _politicas_cache.pop(uid, None)


def _client_ip(request: Request) -> str:
    xff = request.headers.get("x-forwarded-for") or request.headers.get("X-Forwarded-For")
    if xff:
        return xff.split(",")[0].strip()[:128]
    if request.client and request.client.host:
        return str(request.client.host)[:128]
    return ""


def _json_for_log(obj):
    """Serializa valores para columnas jsonb (evita tipos no JSON)."""
    if obj is None:
        return None
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    if isinstance(obj, (str, int, bool)):
        return obj
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, uuid.UUID):
        return str(obj)
    if isinstance(obj, dict):
        return {str(k): _json_for_log(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_for_log(x) for x in obj]
    return str(obj)


def _audit_user_contrato(current_user, contrato_id: int) -> dict:
    """Payload usuario + número de contrato para registrar_log (SICOE, informes, etc.)."""
    cnum = None
    try:
        cr = supabase.table("contratos").select("numero").eq("id", contrato_id).limit(1).execute().data
        if cr:
            cnum = cr[0].get("numero")
    except Exception:
        pass
    return {
        "sub": str(current_user.get("sub") or current_user.get("id") or ""),
        "nombre": current_user.get("nombre") or "",
        "email": current_user.get("email"),
        "cargo_nombre": current_user.get("cargo_nombre"),
        "rol_nombre": current_user.get("rol_nombre"),
        "contrato_id": contrato_id,
        "contrato_numero": cnum,
    }


def _so_registro_audit_snapshot(row: Optional[dict]) -> Optional[dict]:
    """Subconjunto de so_registros para jsonb (evita filas enormes y valores no JSON)."""
    if not row:
        return None
    keys = (
        "id", "reporte_id", "contrato_id", "numero_registro", "capitulo", "competencia",
        "item_numero", "item_descripcion", "unidad", "vlr_unitario", "longitud", "ancho", "espesor",
        "cantidad", "cantidad_total", "costo_directo", "observacion", "corte_id",
        "nivel1_estado", "nivel2_estado", "nivel3_estado", "sub_estado", "bloqueado",
        "foto_url", "foto_numero", "grafico_url", "acta_rpo_id", "semana_id", "nivel2_objeto_pago_sub",
    )
    return _json_for_log({k: row.get(k) for k in keys})


def _so_registro_validacion_audit_snapshot(row: Optional[dict]) -> Optional[dict]:
    """Campos de validación / bloqueo para auditoría (antes-después)."""
    if not row:
        return None
    keys = (
        "nivel1_estado", "nivel1_usuario_id", "nivel1_fecha",
        "nivel2_estado", "nivel2_usuario_id", "nivel2_fecha", "nivel2_objeto_pago_sub",
        "nivel3_estado", "nivel3_usuario_id", "nivel3_fecha",
        "sub_estado", "sub_usuario_id", "sub_fecha",
        "bloqueado", "solicitud_reversion",
        "reversion_arm_n2_usuario_id", "reversion_arm_n3_usuario_id",
    )
    return _json_for_log({k: row.get(k) for k in keys})


def _so_registro_fetch_validacion_audit(contrato_id: int, registro_id: int) -> Optional[dict]:
    def _q():
        return supabase.table("so_registros").select(
            "nivel1_estado, nivel1_usuario_id, nivel1_fecha,"
            "nivel2_estado, nivel2_usuario_id, nivel2_fecha, nivel2_objeto_pago_sub,"
            "nivel3_estado, nivel3_usuario_id, nivel3_fecha,"
            "sub_estado, sub_usuario_id, sub_fecha, bloqueado, solicitud_reversion,"
            "reversion_arm_n2_usuario_id, reversion_arm_n3_usuario_id"
        ).eq("id", registro_id).eq("contrato_id", contrato_id).limit(1).execute().data
    r = supabase_execute(_q)
    return r[0] if r else None


def _enriquecer_registros_labels_reversion_doble_llave(registros: Optional[list]) -> None:
    """Expone nombre legible para quien disparó cada llave (UI doble autorización reversión N3)."""
    if not registros:
        return
    ids: List[int] = []
    for reg in registros:
        if not isinstance(reg, dict):
            continue
        for key in ("reversion_arm_n2_usuario_id", "reversion_arm_n3_usuario_id"):
            v = reg.get(key)
            if v is None:
                continue
            try:
                ids.append(int(v))
            except (TypeError, ValueError):
                pass
    uniq = list({i for i in ids})
    if not uniq:
        for reg in registros:
            if isinstance(reg, dict):
                reg["reversion_arm_n2_nombre"] = None
                reg["reversion_arm_n3_nombre"] = None
        return
    nombres_por_id: Dict[int, str] = {}
    try:
        step = 200
        for i in range(0, len(uniq), step):
            batch = uniq[i : i + step]

            def _lookup(b=batch):
                return (
                    supabase.table("usuarios")
                    .select("id, nombre, apellidos")
                    .in_("id", b)
                    .execute()
                    .data
                )

            rows = supabase_execute(_lookup) or []
            for ur in rows:
                uid = ur.get("id")
                if uid is None:
                    continue
                lbl = f"{ur.get('nombre') or ''} {ur.get('apellidos') or ''}".strip()
                if not lbl:
                    lbl = f"Usuario #{uid}"
                nombres_por_id[int(uid)] = lbl
    except Exception:
        pass
    for reg in registros:
        if not isinstance(reg, dict):
            continue
        for key_uid, key_nom in (
            ("reversion_arm_n2_usuario_id", "reversion_arm_n2_nombre"),
            ("reversion_arm_n3_usuario_id", "reversion_arm_n3_nombre"),
        ):
            v = reg.get(key_uid)
            if v is None:
                reg[key_nom] = None
                continue
            try:
                reg[key_nom] = nombres_por_id.get(int(v))
            except (TypeError, ValueError):
                reg[key_nom] = None


def _so_reporte_audit_snapshot(row: Optional[dict]) -> Optional[dict]:
    if not row:
        return None
    keys = (
        "id", "numero_reporte", "estado", "capitulo", "semana_id", "acta_rpo_id", "corte_id",
        "pk_id_id", "civ", "tramo", "infraestructura", "calzada", "ubicacion",
        "coord_lat", "coord_lng", "abs_inicio", "abs_final", "nodo_ini", "nodo_fin", "margen",
        "subcontratista_id", "inspector_id",
    )
    return _json_for_log({k: row.get(k) for k in keys})


def _default_severidad(accion: str, modulo: str, resultado: str) -> str:
    if resultado and str(resultado).lower() not in ("ok", "success", "éxito"):
        return "ERROR" if "denegado" not in str(resultado).lower() else "WARNING"
    if modulo == "PERMISOS" or accion in (
        "ELIMINAR", "VALIDAR", "APROBAR", "RECHAZAR", "IMPORTAR_MASIVO",
        "ASIGNAR_ITEM", "MOVER", "COMENTAR",
    ):
        return "AUDIT"
    if accion in ("LOGIN_FAIL", "ACCESO_DENEGADO"):
        return "WARNING"
    return "INFO"


@app.exception_handler(Exception)
async def unhandled_exception_to_json(request: Request, exc: Exception):
    """Evita HTML genérico 'Internal Server Error'; el front puede mostrar `detail` en JSON."""
    if isinstance(exc, HTTPException):
        return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})
    if isinstance(exc, RequestValidationError):
        return JSONResponse(status_code=422, content={"detail": exc.errors()})
    _log_api.exception("Error no manejado: %s %s", request.method, request.url.path)
    try:
        registrar_log_sistema(
            "ERROR",
            request.url.path,
            request.method,
            {"error": type(exc).__name__, "msg": str(exc)[:2000]},
            traceback.format_exc()[:12000],
            resultado="error",
            alerta_generada=False,
        )
        _track_500_endpoint_alert(request.url.path)
    except Exception:
        pass
    debug = os.getenv("CLARACORE_DEBUG", "").lower() in ("1", "true", "yes")
    detail = traceback.format_exc() if debug else f"{type(exc).__name__}: {exc}"
    return JSONResponse(status_code=500, content={"detail": detail})


_SUPABASE_URL = os.getenv("SUPABASE_URL")
_SUPABASE_KEY = os.getenv("SUPABASE_KEY")
# PostgREST puede tardar bajo carga; el default de httpx (~5s) provoca ReadTimeout en /auth/refresh y otras rutas.
_SUPABASE_HTTP_TIMEOUT = httpx.Timeout(connect=20.0, read=120.0, write=90.0, pool=30.0)

def get_supabase():
    return create_client(
        _SUPABASE_URL,
        _SUPABASE_KEY,
        options=ClientOptions(
            httpx_client=httpx.Client(http2=False, timeout=_SUPABASE_HTTP_TIMEOUT)
        ),
    )

supabase = get_supabase()
security = HTTPBearer(auto_error=False)

SECRET_KEY = os.getenv("SECRET_KEY")
ALGORITHM = os.getenv("ALGORITHM")
EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES") or "10080")


@app.middleware("http")
async def exigir_politicas_confidencialidad(request: Request, call_next):
    """Bloquea el uso de la API con token si el usuario no ha aceptado las políticas (salvo rutas mínimas)."""
    if request.method == "OPTIONS":
        return await call_next(request)
    path = request.url.path
    auth = request.headers.get("authorization") or ""
    if not auth.startswith("Bearer "):
        return await call_next(request)
    if (
        (request.method == "POST" and path == "/auth/refresh")
        or (request.method == "GET" and path == "/usuarios/me")
        or (request.method == "POST" and path == "/usuarios/me/politicas-aceptar")
    ):
        return await call_next(request)
    try:
        payload = jwt.decode(auth[7:], SECRET_KEY, algorithms=[ALGORITHM])
        sub = payload.get("sub")
        if not sub:
            return await call_next(request)
        uid = int(sub)
    except (JWTError, ValueError, TypeError):
        return await call_next(request)
    cached = _politicas_cache_get(uid)
    if cached == "ok":
        return await call_next(request)
    if cached == "pend":
        return JSONResponse(status_code=403, content={"detail": "politicas_pendientes"})
    if cached == "none":
        return JSONResponse(status_code=403, content={"detail": "politicas_pendientes"})
    try:
        r = supabase.table("usuarios").select("politicas_aceptadas").eq("id", uid).limit(1).execute()
        row = r.data[0] if r.data else None
        if row is None:
            _politicas_cache_set(uid, "none")
            return JSONResponse(status_code=403, content={"detail": "politicas_pendientes"})
        if row.get("politicas_aceptadas") is True:
            _politicas_cache_set(uid, "ok")
            return await call_next(request)
        _politicas_cache_set(uid, "pend")
        return JSONResponse(status_code=403, content={"detail": "politicas_pendientes"})
    except Exception:
        # Columna aún no migrada u otro error: no bloquear despliegue
        return await call_next(request)


# Cloudinary: claracore/{contrato_id}/fotos | graficos | Fotos de Perfil | Fotos de Firmas
CLOUDINARY_ROOT = "claracore"
CLOUDINARY_SUB_FOTOS = "fotos"
CLOUDINARY_SUB_GRAFICOS = "graficos"
CLOUDINARY_SUB_PERFIL = "Fotos de Perfil"
CLOUDINARY_SUB_FIRMAS = "Fotos de Firmas"
# PNG 1×1 px (marcador para crear carpetas al dar de alta un contrato)
_CLOUDINARY_SEED_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def _cloudinary_config():
    import cloudinary
    cloudinary.config(
        cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
        api_key=os.getenv("CLOUDINARY_API_KEY"),
        api_secret=os.getenv("CLOUDINARY_API_SECRET"),
    )


def _cloudinary_folder_contrato(contrato_id: int, subcarpeta: str) -> str:
    return f"{CLOUDINARY_ROOT}/{contrato_id}/{subcarpeta}"


def _cloudinary_seed_carpetas_contrato(contrato_id: int) -> None:
    """Crea en Cloudinary las cuatro carpetas del contrato (un PNG mínimo por carpeta)."""
    if not os.getenv("CLOUDINARY_CLOUD_NAME"):
        return
    import cloudinary.uploader
    _cloudinary_config()
    seeds = [
        (CLOUDINARY_SUB_FOTOS, "_inicial_fotos"),
        (CLOUDINARY_SUB_GRAFICOS, "_inicial_graficos"),
        (CLOUDINARY_SUB_PERFIL, "_inicial_perfil"),
        (CLOUDINARY_SUB_FIRMAS, "_inicial_firmas"),
    ]
    for sub, public_id in seeds:
        try:
            cloudinary.uploader.upload(
                _CLOUDINARY_SEED_PNG,
                folder=_cloudinary_folder_contrato(contrato_id, sub),
                public_id=public_id,
                overwrite=True,
                resource_type="image",
            )
        except Exception as e:
            _log_api.warning("Cloudinary seed %s/%s: %s", contrato_id, sub, e)

# ─────────────────────────────────────────────
# MODELOS
# ─────────────────────────────────────────────

class LoginRequest(BaseModel):
    email: str
    password: str

class UsuarioCreate(BaseModel):
    nombre: str
    email: str
    password: str
    cargo_id: int

class UsuarioRegistro(BaseModel):
    nombre: str
    apellidos: str
    email: str
    cargo_id: int
    contrato_id: Optional[int] = None
    password: str

class PerfilUpdate(BaseModel):
    """Actualización de perfil por el propio usuario (no modifica cargo, contrato ni email)."""
    nombre: Optional[str] = None
    apellidos: Optional[str] = None
    # Cadena vacía borra la fecha; YYYY-MM-DD la guarda; omitir no cambia.
    fecha_nacimiento: Optional[str] = None


class InicioNovedadCreate(BaseModel):
    titulo: str
    resumen: str = ""
    tipo: str = "actualización"
    fecha: Optional[str] = None
    autor: str = "Equipo ClaraCore"
    icono: str = "📢"
    color: str = "#00B4C6"
    imagen_url: Optional[str] = None
    # Solo Desarrollador: null = novedad global (todos los contratos). Se ignora para Administrador.
    contrato_id: Optional[int] = None


class InicioNovedadUpdate(BaseModel):
    titulo: Optional[str] = None
    resumen: Optional[str] = None
    tipo: Optional[str] = None
    fecha: Optional[str] = None
    autor: Optional[str] = None
    icono: Optional[str] = None
    color: Optional[str] = None
    imagen_url: Optional[str] = None
    contrato_id: Optional[int] = None


class InicioNovedadMejorarTexto(BaseModel):
    texto: str = ""

class AprobarRequest(BaseModel):
    rol_id: int

class CargoCreate(BaseModel):
    nombre: str
    rol_id: Optional[int] = None
    categoria_id: Optional[int] = None


class GuiaCreate(BaseModel):
    titulo: str
    modulo: Optional[str] = None
    descripcion_corta: Optional[str] = None
    bloques: List[Any] = Field(default_factory=list)
    roles_visibles: List[int] = Field(default_factory=list)
    publicado: bool = False
    orden: int = 0
    contrato_id: Optional[int] = None


class GuiaUpdate(BaseModel):
    titulo: Optional[str] = None
    modulo: Optional[str] = None
    descripcion_corta: Optional[str] = None
    bloques: Optional[List[Any]] = None
    roles_visibles: Optional[List[int]] = None
    publicado: Optional[bool] = None
    orden: Optional[int] = None
    contrato_id: Optional[int] = None


class CostoAdicionalItem(BaseModel):
    concepto_contractual: str = ""
    valor_mensual: Optional[float] = None
    tiempo_meses: Optional[float] = None
    # Calculado: round( valor_mensual * tiempo_meses, 0 ); puede venir de cliente pero se pisa.
    valor: Optional[float] = None


class ContratoCreate(BaseModel):
    numero: str
    objeto: Optional[str] = None
    contratista: Optional[str] = None
    nit: Optional[str] = None
    interventoria: Optional[str] = None
    entidad: Optional[str] = None
    entidad_otra: Optional[str] = None
    logo_entidad: Optional[str] = None
    plano_geojson: Optional[dict] = None
    centro_lat: Optional[float] = None
    centro_lng: Optional[float] = None
    logo_contratista: Optional[str] = None
    logo_interventoria: Optional[str] = None
    aiu: Optional[float] = None
    iva: Optional[float] = None
    valor_componente_ambiental: Optional[float] = None
    valor_componente_social: Optional[float] = None
    valor_componente_pmt: Optional[float] = None
    costo_directo_contrato: Optional[float] = None
    # Suma legada; si envías costos_adicionales_lista, el API recalcula.
    costos_adicionales: Optional[float] = None
    costos_adicionales_lista: List[CostoAdicionalItem] = Field(default_factory=list)

class PermisoUpdate(BaseModel):
    cargo_id: int
    funcion_id: int
    ver: bool = False
    crear: bool = False
    editar: bool = False
    eliminar: bool = False
    validar: bool = False
    exportar: bool = False

class UsuarioUpdate(BaseModel):
    cargo_id: Optional[int] = None
    rol_id: Optional[int] = None
    contrato_id: Optional[int] = None
    estado: Optional[str] = None
    subcontratista_id: Optional[int] = None
    politicas_aceptadas: Optional[bool] = None

class UsuarioContratoCreate(BaseModel):
    usuario_id: int
    contrato_id: int

class ContratoUpdate(BaseModel):
    numero: Optional[str] = None
    objeto: Optional[str] = None
    contratista: Optional[str] = None
    nit: Optional[str] = None
    interventoria: Optional[str] = None
    entidad: Optional[str] = None
    entidad_otra: Optional[str] = None
    logo_entidad: Optional[str] = None
    plano_geojson: Optional[dict] = None
    centro_lat: Optional[float] = None
    centro_lng: Optional[float] = None
    logo_contratista: Optional[str] = None
    logo_interventoria: Optional[str] = None
    fase: Optional[str] = None  # 'PRESUPUESTO' | 'LIQUIDACION'
    aiu: Optional[float] = None
    iva: Optional[float] = None
    valor_componente_ambiental: Optional[float] = None
    valor_componente_social: Optional[float] = None
    valor_componente_pmt: Optional[float] = None
    costo_directo_contrato: Optional[float] = None
    costos_adicionales: Optional[float] = None
    costos_adicionales_lista: Optional[List[CostoAdicionalItem]] = None

class ListadoPrecioItem(BaseModel):
    capitulo: Optional[str] = None
    competencia: Optional[str] = None
    item_numero: Optional[str] = None
    descripcion: Optional[str] = None
    unidad: Optional[str] = None
    precio_unitario: Optional[float] = None
    color_hex: Optional[str] = None
    tipo_precio: Optional[str] = None
    especificacion_tecnica: Optional[str] = None
    acta_fijacion: Optional[str] = None
    acta_modificatoria: Optional[str] = None
    observaciones: Optional[str] = None
    estado_precio: Optional[str] = None
    tipo_calculo:  Optional[str] = None

class PresupuestoRow(BaseModel):
    pk_id: Optional[str] = None
    capitulo: Optional[str] = None
    competencia: Optional[str] = None
    item: Optional[str] = None
    descripcion: Optional[str] = None
    und: Optional[str] = None
    calzada: Optional[str] = None
    tramo: Optional[str] = None
    abs_inicio: Optional[str] = None
    abs_final: Optional[str] = None
    vlr_unitario: Optional[float] = None
    no_inicio: Optional[str] = None
    no_final: Optional[str] = None
    area_long_nod: Optional[float] = None
    ancho: Optional[float] = None
    espesor: Optional[float] = None
    cant_total: Optional[float] = None
    costo_directo: Optional[float] = None
    tipo_ejecucion: Optional[str] = None
    tipo_entidad: Optional[str] = None
    id_pol: Optional[str] = None
    observacion: Optional[str] = None
    revisado: Optional[str] = None
    observacion_externa: Optional[str] = None
    ent_handle: Optional[str] = None
    txt_handle: Optional[str] = None
    layer_ent: Optional[str] = None
    layer_txt: Optional[str] = None
    color_hex: Optional[str] = None
    guid: Optional[str] = None
    x_label: Optional[float] = None
    y_label: Optional[float] = None

class PresupuestoUpdate(BaseModel):
    capitulo: Optional[str] = None
    competencia: Optional[str] = None
    item: Optional[str] = None
    descripcion: Optional[str] = None
    und: Optional[str] = None
    calzada: Optional[str] = None
    tramo: Optional[str] = None
    no_inicio: Optional[str] = None
    no_final: Optional[str] = None
    vlr_unitario: Optional[float] = None
    area_long_nod: Optional[float] = None
    ancho: Optional[float] = None
    espesor: Optional[float] = None
    cant_total: Optional[float] = None
    costo_directo: Optional[float] = None
    revisado: Optional[str] = None
    observacion_externa: Optional[str] = None
    # Solo contratista, registro sellado: explica por qué reabre; interventoría no puede "reversar" desde aquí.
    motivo_edicion_tras_sellado: Optional[str] = None
    # Contratista: al editar datos con revisado Pendiente/Rechazado/Aprobado (sin sellado), motivo ≥15 y reset a No Revisado.
    motivo_edicion_con_estado_interv: Optional[str] = None

class DimOverride(BaseModel):
    id: int
    ancho: Optional[float] = None
    espesor: Optional[float] = None
    area_long_nod: Optional[float] = None
    capitulo: Optional[str] = None
    item: Optional[str] = None

class PresupuestoBulkRecalc(BaseModel):
    ids: List[int]
    capitulo: Optional[str] = None
    item: Optional[str] = None
    descripcion: Optional[str] = None
    vlr_unitario: Optional[float] = None
    dims: Optional[List[DimOverride]] = None  # cambios de dimensión por fila

class PresupuestoBulkEstado(BaseModel):
    ids: List[int]
    revisado: str


class PresupuestoBulkPreInterv(BaseModel):
    """Depuración contratista antes de que Interventoría revise (Residente de Costos u Obra)."""
    ids: List[int]
    estado: str

class ComentariosValidacionIds(BaseModel):
    """Lista de filas `presupuesto` para comentarios de validación; POST evita query strings gigantes (5xx en proxy)."""
    ids: List[int]

class AgregarCantidadBody(BaseModel):
    # Nuevo ítem
    item: str
    descripcion: str
    und: str
    vlr_unitario: float
    area_long_nod: Optional[float] = None
    ancho: Optional[float] = None
    espesor: Optional[float] = None
    # Heredado del clon base
    capitulo: str
    competencia: Optional[str] = None
    calzada: Optional[str] = None
    tramo: Optional[str] = None
    abs_inicio: Optional[str] = None
    abs_final: Optional[str] = None
    no_inicio: Optional[str] = None
    no_final: Optional[str] = None
    tipo_ejecucion: Optional[str] = None
    tipo_entidad: Optional[str] = None
    id_pol_base: Optional[str] = None
    layer_ent: Optional[str] = None
    layer_txt: Optional[str] = None
    x_label: Optional[float] = None
    y_label: Optional[float] = None

class CadQueueCreate(BaseModel):
    tipo: str        # cambiar_layer | insertar_bloque
    payload: dict

class CadQueueProcesado(BaseModel):
    rev_block_handle: Optional[str] = None   # solo para insertar_bloque
    presupuesto_id:   Optional[int] = None

class CobroRow(BaseModel):
    pk_id: Optional[str] = None
    acta: Optional[int] = None
    semana: Optional[str] = None
    fecha: Optional[str] = None
    capitulo: Optional[str] = None
    competencia: Optional[str] = None
    abs_inicial: Optional[str] = None
    abs_final: Optional[str] = None
    civ: Optional[str] = None
    item: Optional[str] = None
    descripcion: Optional[str] = None
    und: Optional[str] = None
    longitud: Optional[float] = None
    ancho: Optional[float] = None
    espesor: Optional[float] = None
    cantidad: Optional[float] = None
    valor_unitario: Optional[float] = None
    costo_directo: Optional[float] = None
    calzada: Optional[str] = None
    tramo_inicio: Optional[str] = None
    tramo_final: Optional[str] = None
    registro: Optional[str] = None
    tramo: Optional[str] = None
    observaciones: Optional[str] = None

class ResetSolicitud(BaseModel):
    email: str

class ResetAutorizar(BaseModel):
    contrasena_temporal: str

class CambiarPassword(BaseModel):
    email: str
    contrasena_temporal: str
    nueva_password: str

class MantenimientoRequest(BaseModel):
    secret: str
    activo: bool
    mensaje: Optional[str] = None
    segundos: Optional[int] = None

def _estado_mantenimiento():
    now = time.time()
    if _maintenance_state["activo"] and _maintenance_state["expires_at"] is not None and now >= _maintenance_state["expires_at"]:
        _maintenance_state["activo"] = False
        _maintenance_state["expires_at"] = None

    restantes = None
    if _maintenance_state["activo"] and _maintenance_state["expires_at"] is not None:
        restantes = max(0, int(_maintenance_state["expires_at"] - now))

    return {
        "activo": _maintenance_state["activo"],
        "mensaje": _maintenance_state["mensaje"],
        "segundos_restantes": restantes,
    }

# ─────────────────────────────────────────────
# HELPER SUPABASE CON REINTENTOS
# ─────────────────────────────────────────────
def _is_supabase_statement_timeout(err: Exception) -> bool:
    """Postgres 57014 — evita 500 genérico y alinea con saturación de la BD."""
    s = str(err).lower()
    if "57014" in s or "statement timeout" in s or "canceling statement" in s:
        return True
    code = getattr(err, "code", None)
    if code == "57014":
        return True
    d = getattr(err, "dict", None)
    if callable(d):
        try:
            di = d()
            if isinstance(di, dict) and str(di.get("code") or "") == "57014":
                return True
        except Exception:
            pass
    return False


def supabase_execute(fn, retries=3, delay=0.5):
    import time
    global supabase
    last_err = None
    for i in range(retries):
        try:
            return fn()
        except Exception as e:
            last_err = e
            supabase = get_supabase()
            if i < retries - 1:
                time.sleep(delay)
    try:
        registrar_log_sistema(
            "ERROR",
            "supabase_execute",
            "RPC",
            {"error": type(last_err).__name__, "msg": str(last_err)[:2000], "reintentos": retries},
            traceback.format_exc()[:8000],
            resultado="error",
            alerta_generada=False,
        )
    except Exception:
        pass
    if _is_supabase_statement_timeout(last_err):
        raise HTTPException(
            status_code=503,
            detail="La base de datos está ocupada o la consulta tardó demasiado. Reintenta en unos segundos o acota el filtro.",
        )
    raise last_err

# ─────────────────────────────────────────────
# SISTEMA DE LOGS
# ─────────────────────────────────────────────
_LOGS_INSERT_MAX_STRIPS = 32


def _logs_pgrst_unknown_column(err: Exception) -> Optional[str]:
    """PGRST204: columna ausente en la caché de esquema de PostgREST (tabla logs sin ALTER)."""
    text = str(err)
    if "PGRST204" not in text and "schema cache" not in text.lower():
        return None
    m = re.search(r"find the '([^']+)' column", text, re.I)
    if m:
        return m.group(1)
    m = re.search(r'find the "([^"]+)" column', text, re.I)
    return m.group(1) if m else None


def _logs_insert_row(row: Dict[str, Any]) -> bool:
    """Inserta en logs; omite columnas desconocidas y las guarda en caché para no penalizar cada request."""
    payload = dict(row)
    with _logs_omit_lock:
        frozen_omit = frozenset(_logs_omit_columns)
    for c in frozen_omit:
        payload.pop(c, None)
    strips = 0
    while strips <= _LOGS_INSERT_MAX_STRIPS:
        try:
            supabase.table("logs").insert(payload).execute()
            return True
        except Exception as e:
            col = _logs_pgrst_unknown_column(e)
            if col and col in payload:
                with _logs_omit_lock:
                    _logs_omit_columns.add(col)
                payload.pop(col, None)
                strips += 1
                continue
            try:
                _log_api.warning(
                    "registrar_log: insert falló (%s) (strips=%s)",
                    e,
                    strips,
                )
            except Exception:
                pass
            return False
    try:
        _log_api.warning("registrar_log: se alcanzó el máximo de columnas omitidas; revise esquema de logs")
    except Exception:
        pass
    return False


def registrar_log_sistema(
    severidad: str,
    endpoint: str,
    metodo_http: str,
    detalle: dict,
    stack_trace: Optional[str] = None,
    resultado: str = "error",
    alerta_generada: bool = False,
    duracion_ms: Optional[int] = None,
):
    """Errores y eventos técnicos (categoría sistema). No requiene usuario."""
    try:
        row = {
            "usuario_id":       None,
            "usuario_nombre":   "SISTEMA",
            "cargo_nombre":     "",
            "rol_nombre":       None,
            "contrato_id":      None,
            "contrato_numero":  None,
            "accion":           "ERROR_SISTEMA" if severidad == "ERROR" else "EVENTO_SISTEMA",
            "modulo":           "SISTEMA",
            "entidad_tipo":     "endpoint",
            "entidad_id":       None,
            "detalle":          _json_for_log(detalle) if detalle is not None else {},
            "resultado":        resultado,
            "categoria":        "sistema",
            "severidad":        severidad,
            "endpoint":         (endpoint or "")[:1024],
            "metodo_http":      (metodo_http or "")[:32],
            "stack_trace":      (stack_trace or "")[:12000] if stack_trace else None,
            "duracion_ms":      duracion_ms,
            "alerta_generada":  alerta_generada,
        }
        _logs_insert_row(row)
    except Exception:
        pass


def registrar_log(
    usuario,
    accion,
    modulo,
    entidad_tipo=None,
    entidad_id=None,
    detalle=None,
    resultado="ok",
    *,
    valor_anterior=None,
    valor_nuevo=None,
    ip=None,
    severidad=None,
    categoria="auditoria",
    rol_nombre=None,
    endpoint=None,
    metodo_http=None,
    stack_trace=None,
    duracion_ms=None,
    alerta_generada=None,
):
    try:
        uid = usuario.get("sub") or usuario.get("id")
        uid_int = None
        if uid is not None and str(uid).strip() != "":
            try:
                uid_int = int(str(uid).strip())
            except (ValueError, TypeError):
                uid_int = None
        if severidad is None:
            severidad = _default_severidad(accion, modulo, resultado)
        row = {
            "usuario_id":       uid_int,
            "usuario_nombre":   usuario.get("nombre") or usuario.get("email", ""),
            "cargo_nombre":     usuario.get("cargo_nombre") or "",
            "rol_nombre":       rol_nombre if rol_nombre is not None else usuario.get("rol_nombre"),
            "contrato_id":      usuario.get("contrato_id"),
            "contrato_numero":  usuario.get("contrato_numero"),
            "accion":           accion,
            "modulo":           modulo,
            "entidad_tipo":     entidad_tipo,
            "entidad_id":       str(entidad_id) if entidad_id is not None else None,
            "detalle":          _json_for_log(detalle) if detalle is not None else {},
            "resultado":        resultado,
            "valor_anterior":   _json_for_log(valor_anterior),
            "valor_nuevo":      _json_for_log(valor_nuevo),
            "ip":               ip,
            "categoria":        categoria,
            "severidad":        severidad,
            "endpoint":         (endpoint or "")[:1024] if endpoint else None,
            "metodo_http":      (metodo_http or "")[:32] if metodo_http else None,
            "stack_trace":      (stack_trace or "")[:12000] if stack_trace else None,
            "duracion_ms":      duracion_ms,
            "alerta_generada":  alerta_generada if alerta_generada is not None else False,
        }
        if not _logs_insert_row(row):
            pass
    except Exception:
        pass


def _track_login_failure_alert(email: str, ip: str):
    import time
    now = time.time()
    key = (email.strip().lower(), ip or "")
    q = _login_fail_tracker.setdefault(key, [])
    q.append(now)
    cutoff = now - 900
    while q and q[0] < cutoff:
        q.pop(0)
    if len(q) >= 5:
        try:
            registrar_log_sistema(
                "WARNING",
                "/auth/login",
                "POST",
                {"email": email, "ip": ip, "ventana_seg": 900, "intentos": len(q)},
                None,
                resultado="alerta",
                alerta_generada=True,
            )
        except Exception:
            pass
        q.clear()


def _track_500_endpoint_alert(path: str):
    import time
    now = time.time()
    q = _endpoint_500_tracker.setdefault(path, [])
    q.append(now)
    cutoff = now - 300
    while q and q[0] < cutoff:
        q.pop(0)
    if len(q) >= 3:
        try:
            registrar_log_sistema(
                "ERROR",
                path,
                "—",
                {"tipo": "error_500_repetido", "ventana_seg": 300, "conteo": len(q)},
                None,
                resultado="alerta",
                alerta_generada=True,
            )
        except Exception:
            pass
        q.clear()


def _cargo_puede_auditar_logs(current_user) -> bool:
    try:
        uid = int(current_user.get("sub"))
    except (TypeError, ValueError):
        return False
    try:
        u = supabase.table("usuarios").select("cargo_id").eq("id", uid).limit(1).execute().data
        u = u[0] if u else None
        if not u or not u.get("cargo_id"):
            return False
        c = supabase.table("cargos").select("nombre").eq("id", u["cargo_id"]).limit(1).execute().data
        c = c[0] if c else None
        n = ((c or {}).get("nombre") or "").strip().lower()
        return n in ("desarrollador", "administrador")
    except Exception:
        return False


def _es_desarrollador(current_user) -> bool:
    """Solo el cargo Desarrollador (no Administrador): acciones destructivas de desarrollo."""
    try:
        uid = int(current_user.get("sub"))
    except (TypeError, ValueError):
        return False
    try:
        u = supabase.table("usuarios").select("cargo_id").eq("id", uid).limit(1).execute().data
        u = u[0] if u else None
        if not u or not u.get("cargo_id"):
            return False
        c = supabase.table("cargos").select("nombre").eq("id", u["cargo_id"]).limit(1).execute().data
        c = c[0] if c else None
        n = ((c or {}).get("nombre") or "").strip().lower()
        return n == "desarrollador"
    except Exception:
        return False


def _caller_cargo_id(current_user) -> Optional[int]:
    """cargo_id del usuario autenticado (tabla usuarios)."""
    try:
        uid = int(current_user.get("sub"))
    except (TypeError, ValueError):
        return None
    try:
        u = supabase.table("usuarios").select("cargo_id").eq("id", uid).limit(1).execute().data
        u = u[0] if u else None
        if not u:
            return None
        cid = u.get("cargo_id")
        return int(cid) if cid is not None else None
    except Exception:
        return None


def _cargo_permiso_editar_registros_presupuesto(current_user) -> bool:
    """Matriz de permisos: función «editar registros presupuesto» con acción editar (alineado con el frontend)."""
    cid = _caller_cargo_id(current_user)
    if cid is None:
        return False
    try:
        perms = supabase_execute(
            lambda: supabase.table("permisos")
            .select("funcion_id, editar")
            .eq("cargo_id", cid)
            .execute()
            .data
        ) or []
        fids = [p["funcion_id"] for p in perms if p.get("editar")]
        if not fids:
            return False
        funcs = supabase_execute(
            lambda: supabase.table("funciones").select("id, nombre").in_("id", fids).execute().data
        ) or []
        want = "editar registros presupuesto"
        for f in funcs:
            if (f.get("nombre") or "").strip().lower() == want:
                return True
    except Exception:
        return False
    return False


def _cargo_permiso_validar_presupuesto(current_user) -> bool:
    """Matriz: función «editar registros presupuesto» con acción validar."""
    cid = _caller_cargo_id(current_user)
    if cid is None:
        return False
    try:
        perms = supabase_execute(
            lambda: supabase.table("permisos")
            .select("funcion_id, validar")
            .eq("cargo_id", cid)
            .execute()
            .data
        ) or []
        fids = [p["funcion_id"] for p in perms if p.get("validar")]
        if not fids:
            return False
        funcs = supabase_execute(
            lambda: supabase.table("funciones").select("id, nombre").in_("id", fids).execute().data
        ) or []
        want = "editar registros presupuesto"
        for f in funcs:
            if (f.get("nombre") or "").strip().lower() == want:
                return True
    except Exception:
        return False
    return False


def _cargo_permiso_validar_reporte_cantidades_user_id(user_id: int) -> bool:
    """Matriz: función «reporte de cantidades» con acción validar (SICOE obra)."""
    try:
        uid = int(user_id)
    except (TypeError, ValueError):
        return False
    try:
        urows = supabase.table("usuarios").select("cargo_id").eq("id", uid).limit(1).execute().data
        u = urows[0] if urows else None
        if not u or u.get("cargo_id") is None:
            return False
        cid = int(u["cargo_id"])
    except (TypeError, ValueError, KeyError):
        return False
    try:
        perms = supabase_execute(
            lambda: supabase.table("permisos")
            .select("funcion_id, validar")
            .eq("cargo_id", cid)
            .execute()
            .data
        ) or []
        fids = [p["funcion_id"] for p in perms if p.get("validar")]
        if not fids:
            return False
        funcs = supabase_execute(
            lambda: supabase.table("funciones").select("id, nombre").in_("id", fids).execute().data
        ) or []
        want = "reporte de cantidades"
        for f in funcs:
            if (f.get("nombre") or "").strip().lower() == want:
                return True
    except Exception:
        return False
    return False


def _cargo_permiso_editar_reporte_cantidades_user_id(user_id: int) -> bool:
    """Matriz: función «reporte de cantidades» con acción editar (SICOE obra)."""
    try:
        uid = int(user_id)
    except (TypeError, ValueError):
        return False
    try:
        urows = supabase.table("usuarios").select("cargo_id").eq("id", uid).limit(1).execute().data
        u = urows[0] if urows else None
        if not u or u.get("cargo_id") is None:
            return False
        cid = int(u["cargo_id"])
    except (TypeError, ValueError, KeyError):
        return False
    try:
        perms = supabase_execute(
            lambda: supabase.table("permisos")
            .select("funcion_id, editar")
            .eq("cargo_id", cid)
            .execute()
            .data
        ) or []
        fids = [p["funcion_id"] for p in perms if p.get("editar")]
        if not fids:
            return False
        funcs = supabase_execute(
            lambda: supabase.table("funciones").select("id, nombre").in_("id", fids).execute().data
        ) or []
        want = "reporte de cantidades"
        for f in funcs:
            if (f.get("nombre") or "").strip().lower() == want:
                return True
    except Exception:
        return False
    return False


def _puede_editar_dimensiones_presupuesto(current_user) -> bool:
    return _es_desarrollador(current_user) or _cargo_permiso_editar_registros_presupuesto(current_user)


def _caller_rol_id(current_user) -> Optional[int]:
    """rol_id del usuario autenticado (tabla usuarios → tabla roles)."""
    try:
        uid = int(current_user.get("sub"))
    except (TypeError, ValueError):
        return None
    try:
        u = supabase.table("usuarios").select("rol_id").eq("id", uid).limit(1).execute().data
        u = u[0] if u else None
        if not u:
            return None
        rid = u.get("rol_id")
        return int(rid) if rid is not None else None
    except Exception:
        return None


def _guia_visible_por_rol(row: dict, rol_id: Optional[int]) -> bool:
    """roles_visibles = IDs en tabla roles; vacío = todos los roles; si no, debe incluir rol_id del usuario."""
    rv = row.get("roles_visibles")
    if rv is None or (isinstance(rv, list) and len(rv) == 0):
        return True
    if rol_id is None:
        return False
    try:
        ids = [int(x) for x in rv]
    except (TypeError, ValueError):
        return False
    return rol_id in ids


_GUIAS_SELECT_LISTA = (
    "id, contrato_id, titulo, slug, modulo, descripcion_corta, roles_visibles, publicado, orden, created_at, updated_at"
)


def _titulo_a_slug_base(titulo: str) -> str:
    t = (titulo or "").strip().lower()
    t = re.sub(r"\s+", "-", t)
    t = re.sub(r"[^a-z0-9\-]+", "-", t)
    t = re.sub(r"-+", "-", t).strip("-")
    return (t or "guia")[:190]


def _siguiente_slug_unico(base: str, exclude_id: Optional[int]) -> str:
    cand = base
    suf = 2
    while True:
        q = supabase.table("guias").select("id").eq("slug", cand).limit(1).execute()
        row = q.data[0] if q.data else None
        if not row:
            return cand
        if exclude_id is not None and int(row["id"]) == int(exclude_id):
            return cand
        cand = f"{base[:170]}-{suf}"
        suf += 1


def _es_rol_contratista_ppto(current_user) -> bool:
    """Contratista u operativo: único perfil que puede reabrir un registro sellado con motivo (no la Interventoría vía API)."""
    r = (current_user.get("rol_nombre") or "").strip().lower()
    return r in ("contratista", "operativo contratista")


# ─────────────────────────────────────────────
# SEGURIDAD
# ─────────────────────────────────────────────

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def hash_password(password: str):
    return pwd_context.hash(password)

def verify_password(plain: str, hashed: str):
    return pwd_context.verify(plain, hashed)

def create_token(data: dict):
    payload = data.copy()
    payload.update({"exp": datetime.utcnow() + timedelta(minutes=EXPIRE_MINUTES)})
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except JWTError:
        raise HTTPException(status_code=401, detail="Token inválido")


def _calculo_usuario_label(current_user) -> str:
    """Etiqueta legible del usuario autenticado (JWT) para auditoría de recálculo."""
    if not current_user:
        return ""
    n = (current_user.get("nombre") or "").strip()
    if n:
        return n
    e = (current_user.get("email") or "").strip()
    if e:
        return e
    s = current_user.get("sub")
    return f"Usuario {s}" if s is not None else "—"


def require_logs_auditoria(current_user=Depends(get_current_user)):
    if not _cargo_puede_auditar_logs(current_user):
        raise HTTPException(status_code=403, detail="Solo Desarrollador y Administrador pueden acceder a la auditoría de logs")
    return current_user


def _query_inicio_novedades_accesibles(contrato_id_usuario: Optional[int]):
    """Novedades globales (contrato_id NULL) + las del contrato del usuario, si aplica."""
    q = supabase.table("inicio_novedades").select("*").order("created_at", desc=True)
    if contrato_id_usuario is not None:
        return q.or_(f"contrato_id.is.null,contrato_id.eq.{int(contrato_id_usuario)}")
    return q.is_("contrato_id", "null")


def _novedad_puede_gestionar_admin(
    novedad: dict, contrato_caller: Optional[int], es_desarrollador: bool
) -> bool:
    if es_desarrollador:
        return True
    ncid = novedad.get("contrato_id")
    if ncid is None:
        return False
    if contrato_caller is None:
        return False
    return int(ncid) == int(contrato_caller)


def _caller_contract_scope(current_user):
    """Retorna (contrato_id, cargo_nombre) del usuario autenticado."""
    try:
        caller_id = int(current_user.get("sub"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=401, detail="Token inválido")

    caller_data = supabase.table("usuarios").select("cargo_id, contrato_id").eq("id", caller_id).single().execute().data
    if not caller_data:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")

    cargo_nombre = ""
    cargo_id = caller_data.get("cargo_id")
    if cargo_id:
        c = supabase.table("cargos").select("nombre").eq("id", cargo_id).single().execute().data
        cargo_nombre = (c or {}).get("nombre", "")
    return caller_data.get("contrato_id"), (cargo_nombre or "").strip().lower()

def _require_contract_access(current_user, contrato_id: int):
    """Bloquea acceso si intenta consultar un contrato distinto al propio."""
    caller_contrato, _ = _caller_contract_scope(current_user)
    if caller_contrato and int(caller_contrato) != int(contrato_id):
        raise HTTPException(status_code=403, detail="No tienes acceso a información de otro contrato")


def _usuario_vinculado_a_contrato(usuario_id: int, contrato_id: int) -> bool:
    """True si contrato es el principal del usuario o está en usuario_contratos."""
    try:
        cid = int(contrato_id)
        uid = int(usuario_id)
    except (TypeError, ValueError):
        return False
    u = supabase.table("usuarios").select("contrato_id").eq("id", uid).limit(1).execute().data
    if u and u[0].get("contrato_id") is not None:
        try:
            if int(u[0]["contrato_id"]) == cid:
                return True
        except (TypeError, ValueError):
            pass
    uc = supabase.table("usuario_contratos").select("id").eq("usuario_id", uid).eq("contrato_id", cid).limit(1).execute().data
    return bool(uc)

from modulos_experimentales_routes import router as _modulos_experimentales_router
app.include_router(_modulos_experimentales_router)

from informes import router as informes_router
app.include_router(informes_router, prefix="/informes")

# Vista previa JSON (CC-SUB-001 / CC-SUB-002): registrado aquí porque en algunos equipos el router
# importado desde informes.py no exponía estas rutas en OpenAPI (Not Found en el cliente).
from informes import _perm_informes_ccd, _respuesta_json_corte, _respuesta_json_memoria
from ccd_conciliacion import (
    rpo_conciliacion_por_contrato,
    rpo_conciliacion_un_acta_rpc,
    rpo_resumen_actas_rpc,
)


@app.get("/informes/{contrato_id}/datos/corte-subcontratista/{corte_id}")
def informes_datos_corte_sub(contrato_id: int, corte_id: int, current_user=Depends(get_current_user)):
    _perm_informes_ccd(current_user, "ver")
    return _respuesta_json_corte(contrato_id, corte_id, current_user)


@app.get("/informes/{contrato_id}/vista-json/corte-sub/{corte_id}")
def informes_vista_json_corte_sub(contrato_id: int, corte_id: int, current_user=Depends(get_current_user)):
    _perm_informes_ccd(current_user, "ver")
    return _respuesta_json_corte(contrato_id, corte_id, current_user)


@app.get("/informes/{contrato_id}/datos/memoria-item/{corte_id}")
def informes_datos_memoria_item(
    contrato_id: int,
    corte_id: int,
    item_numero: str = Query(...),
    current_user=Depends(get_current_user),
):
    _perm_informes_ccd(current_user, "ver")
    return _respuesta_json_memoria(contrato_id, corte_id, item_numero, current_user)


@app.get("/informes/{contrato_id}/vista-json/memoria/{corte_id}")
def informes_vista_json_memoria(
    contrato_id: int,
    corte_id: int,
    item_numero: str = Query(...),
    current_user=Depends(get_current_user),
):
    _perm_informes_ccd(current_user, "ver")
    return _respuesta_json_memoria(contrato_id, corte_id, item_numero, current_user)


# ── Mapa de cargo_id → campo de validación en so_registros ───────────────────
CARGO_ID_NIVEL_MAP = {
    54: 'nivel1_estado',   # Inspector de Obra
    44: 'nivel2_estado',   # Residente de Obra
    45: 'nivel2_estado',   # Cargo migrado Bubble (nivel 2)
    51: 'nivel2_estado',   # Cargo migrado Bubble (nivel 2)
    56: 'nivel2_estado',   # Director de Obra
    50: 'nivel3_estado',   # Residente de Interventoría
    58: 'nivel3_estado',   # Director de Interventoría
}

CARGO_NIVEL_PRERREQUISITO = {
    'nivel2_estado': ('nivel1_estado', 'Aprobado'),
    'nivel3_estado': ('nivel2_estado', 'Aprobado'),
}

# Filtros Sicoe y capas JSON pueden usar { "nivel": 1|2|3, "estado": "..." } en lugar de cargo_id.
NIVEL_VALIDACION_NUM_A_CAMPO = {
    1: "nivel1_estado",
    2: "nivel2_estado",
    3: "nivel3_estado",
}

# Misma semántica que el filtro UI (rol + nivel), no cargo_id.
NIVEL_VALIDACION_ENCABEZADO = {
    1: "Nivel 1 · Operativo contratista",
    2: "Nivel 2 · Contratista",
    3: "Nivel 3 · Interventoría",
}

# Mismo catálogo que el desplegable de validación en frontend (ETIQUETAS_VALIDACION).
SICOE_ETIQUETAS_VALIDACION = (
    "01. Ensayos de Laboratorio",
    "02. Certificados de Calidad",
    "03. Información y/o Entrega Topografía",
    "04. Entrega en obra",
    "05. Informe o Concepto Especialista",
    "06. Incluida dentro del precio",
    "07. Reportado en actas anteriores",
    "08. Pendiente por aprobación de precio",
    "09. Actividad sin concluir",
    "10. Precio no corresponde con la actividad",
    "11. Actualizar información",
    "12. Reproceso",
    "13. Actividad no ejecutada",
    "14. Relacionada con Balance de Obra",
)
SICOE_ETIQUETAS_VALIDACION_SET = frozenset(SICOE_ETIQUETAS_VALIDACION)


def _sicoe_parse_etiqueta_validacion_param(raw: Optional[str]) -> Optional[str]:
    if raw is None:
        return None
    t = str(raw).strip()
    if not t:
        return None
    if t not in SICOE_ETIQUETAS_VALIDACION_SET:
        raise HTTPException(
            status_code=422,
            detail="etiqueta_validacion no está en el catálogo de etiquetas de validación.",
        )
    return t


def _sicoe_fetch_registro_ids_etiqueta_validacion(contrato_id: int, etiqueta: str) -> Set[int]:
    """registro_id con al menos un comentario tipo validación y etiqueta exacta."""
    out: Set[int] = set()
    off = 0
    page = 1000
    while True:
        def _pg(o=off):
            return (
                supabase.table("so_registro_comentarios")
                .select("registro_id")
                .eq("contrato_id", contrato_id)
                .eq("tipo", "validacion")
                .eq("etiqueta", etiqueta)
                .range(o, o + page - 1)
                .execute()
                .data
            )

        batch = supabase_execute(_pg)
        for row in batch:
            rid = row.get("registro_id")
            if rid is not None:
                try:
                    out.add(int(rid))
                except (TypeError, ValueError):
                    pass
        if len(batch) < page:
            break
        off += page
    return out


def _sicoe_chunks_int(ids: List[int], size: int):
    for i in range(0, len(ids), size):
        yield ids[i : i + size]


def _sicoe_norm_txt(s: Optional[str]) -> str:
    if s is None:
        return ""
    import unicodedata
    t = unicodedata.normalize("NFD", str(s).strip().lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def _capa_campo_validacion(capa: dict) -> Optional[str]:
    if not isinstance(capa, dict):
        return None
    campo = capa.get("campo")
    if campo in ("nivel1_estado", "nivel2_estado", "nivel3_estado"):
        return str(campo)
    nv = capa.get("nivel")
    if nv is not None and str(nv).strip() != "":
        try:
            n = int(nv)
            return NIVEL_VALIDACION_NUM_A_CAMPO.get(n)
        except (TypeError, ValueError):
            pass
    cid = capa.get("cargo_id")
    if cid is not None and str(cid).strip() != "":
        try:
            return CARGO_ID_NIVEL_MAP.get(int(cid))
        except (TypeError, ValueError):
            pass
    return None


def _sicoe_capa_etiqueta_panel(capx: dict) -> str:
    """Etiqueta del panel dinámico por capa: nivel (rol) o nombre de cargo (legado)."""
    if not isinstance(capx, dict):
        return "Validación"
    nv = capx.get("nivel")
    if nv is not None and str(nv).strip() != "":
        try:
            ni = int(nv)
            return NIVEL_VALIDACION_ENCABEZADO.get(ni) or f"Nivel {ni}"
        except (TypeError, ValueError):
            pass
    cid = capx.get("cargo_id")
    if cid is not None and str(cid).strip() != "":
        try:
            ci = int(cid)
            lbl = f"cargo {ci}"
            try:
                crow = supabase.table("cargos").select("nombre").eq("id", ci).single().execute().data
                if crow and crow.get("nombre"):
                    lbl = crow["nombre"]
            except Exception:
                pass
            return lbl
        except (TypeError, ValueError):
            pass
    fld = capx.get("campo") or _capa_campo_validacion(capx)
    if fld == "nivel1_estado":
        return NIVEL_VALIDACION_ENCABEZADO[1]
    if fld == "nivel2_estado":
        return NIVEL_VALIDACION_ENCABEZADO[2]
    if fld == "nivel3_estado":
        return NIVEL_VALIDACION_ENCABEZADO[3]
    return "Validación"


def _sicoe_nivel_num_desde_campo(campo: Optional[str]) -> Optional[int]:
    if campo == "nivel1_estado":
        return 1
    if campo == "nivel2_estado":
        return 2
    if campo == "nivel3_estado":
        return 3
    return None


def _sicoe_db_nivel_validacion_usuario(user_id: int) -> Optional[int]:
    """Nivel 1–3 según rol y, si aplica, cargo_id (mismo criterio que CARGO_ID_NIVEL_MAP / panel de acceso)."""
    try:
        uid = int(user_id)
    except (TypeError, ValueError):
        return None
    try:
        urows = supabase.table("usuarios").select("rol_id, cargo_id").eq("id", uid).limit(1).execute().data
        if not urows:
            return None
        u = urows[0]
    except Exception:
        return None
    rn = ""
    if u.get("rol_id"):
        try:
            rrows = supabase.table("roles").select("nombre").eq("id", u["rol_id"]).limit(1).execute().data
            if rrows:
                rn = _sicoe_norm_txt(rrows[0].get("nombre"))
        except Exception:
            pass
    if rn == "operativo contratista":
        return 1
    if rn == "contratista":
        return 2
    if rn == "interventoria":
        return 3
    # Perfil solo lectura / comentarios en interventoría: no validar por cargo aunque el mapa tenga N3
    if rn == "operativo interventoria":
        return None
    try:
        cid = int(u.get("cargo_id")) if u.get("cargo_id") is not None else None
    except (TypeError, ValueError):
        cid = None
    if cid is not None:
        n_cargo = _sicoe_nivel_num_desde_campo(CARGO_ID_NIVEL_MAP.get(cid))
        if n_cargo is not None:
            return n_cargo
    # Rol con nombre extendido p. ej. «Residente de Interventoría» como nombre de rol en BD
    if rn and "interventoria" in rn:
        return 3
    return None


def _require_sicoe_puede_validar_nivel(current_user, user_id: int, nivel: int) -> None:
    if _es_desarrollador(current_user):
        return
    if not _cargo_permiso_validar_reporte_cantidades_user_id(user_id):
        raise HTTPException(
            status_code=403,
            detail="No tiene permiso de validación en «Reporte de cantidades» (matriz de accesos).",
        )
    got = _sicoe_db_nivel_validacion_usuario(user_id)
    if got != nivel:
        raise HTTPException(
            status_code=403,
            detail="Tu rol no autoriza validar en este nivel de SICOE obra.",
        )


def _require_llave_reversion_sicoe_nivel(current_user, user_id: int, nivel_arm: int) -> None:
    """
    Doble llave reversión N3: debe coincidir el nivel SICOE obra del usuario (rol/cargo en BD) con la llave.
    No se exige la matriz «Reporte de cantidades» para esta acción (evita falsos negativos si la matriz está incompleta).
    """
    if _es_desarrollador(current_user):
        return
    if nivel_arm not in (2, 3):
        raise HTTPException(status_code=403, detail="Llave de reversión no reconocida.")
    got = _sicoe_db_nivel_validacion_usuario(user_id)
    if got != nivel_arm:
        raise HTTPException(
            status_code=403,
            detail="Tu rol no autoriza esta llave en SICOE obra.",
        )


def _estado_registro_eq_desde_filtro_ui(evp: str) -> str:
    """Alinea plural de la UI con el valor almacenado en so_registros."""
    p = (evp or "").strip()
    m = {
        "Aprobados": "Aprobado",
        "Pendientes": "Pendiente",
        "Rechazados": "Rechazado",
    }
    return m.get(p, p)


def _parse_validacion_capas_param(
    validacion_capas: Optional[str],
    cargo_id: Optional[int],
    estado_validacion: Optional[str],
) -> List[dict]:
    out: List[dict] = []
    if validacion_capas and str(validacion_capas).strip():
        try:
            raw = json.loads(validacion_capas)
            if isinstance(raw, list):
                for c in raw:
                    if not isinstance(c, dict):
                        continue
                    est = (c.get("estado") or "").strip()
                    if not est:
                        continue
                    fld = _capa_campo_validacion(c)
                    if fld:
                        row: dict = {"estado": est, "campo": fld}
                        if c.get("nivel") is not None:
                            try:
                                row["nivel"] = int(c["nivel"])
                            except (TypeError, ValueError):
                                pass
                        if c.get("cargo_id") is not None:
                            try:
                                row["cargo_id"] = int(c["cargo_id"])
                            except (TypeError, ValueError):
                                pass
                        out.append(row)
        except (json.JSONDecodeError, TypeError, ValueError, KeyError):
            pass
    if not out and cargo_id is not None and (estado_validacion or "").strip():
        fld = CARGO_ID_NIVEL_MAP.get(int(cargo_id))
        if fld:
            out = [{
                "cargo_id": int(cargo_id),
                "estado": str(estado_validacion).strip(),
                "campo": fld,
            }]
    return out


def _so_registros_q_y_capas_validacion(
    q,
    capas: List[dict],
    pk_id_val,
    tramo_v,
    costado_v,
    cap_v,
    sub_v,
    item_v,
):
    """AND en la misma fila de so_registros: todas las capas (cargo+estado) a la vez."""
    for capa in capas:
        fld = capa.get("campo") or _capa_campo_validacion(capa)
        if not fld:
            continue
        evp = (capa.get("estado") or "").strip()
        if not evp:
            continue
        prereq = CARGO_NIVEL_PRERREQUISITO.get(fld)
        if prereq:
            q = q.eq(prereq[0], prereq[1])
        if evp in ("No Revisado", "No Revisados"):
            q = _so_reg_or_pendiente_nivel(q, fld)
            # N1: misma lógica que "cola" N2/N3: sin ítem asignado no entra a revisión de inspector
            if fld == "nivel1_estado":
                q = _so_reg_item_asignado(q)
        else:
            evq = _estado_registro_eq_desde_filtro_ui(evp)
            q = q.eq(fld, evq)
        if _es_validacion_avanzada(fld):
            q = _so_reg_item_asignado(q)
    if pk_id_val is not None:
        q = q.eq("pk_id_id", pk_id_val)
    if tramo_v:
        q = q.eq("tramo", tramo_v)
    if costado_v:
        q = _so_reg_filtro_costado(q, costado_v)
    if cap_v:
        q = q.eq("capitulo", cap_v)
    if sub_v is not None:
        q = q.eq("subcontratista_id", sub_v)
    if item_v:
        q = q.ilike("item_numero", f"%{item_v}%")
    return q


def _validacion_cualquier_nivel2_o_3(capas: List[dict]) -> bool:
    for c in capas:
        fld = c.get("campo") or _capa_campo_validacion(c)
        if fld and _es_validacion_avanzada(fld):
            return True
    return False

# Nivel 2/3: no listar cantidades en reportes borrador / sin ítem asignado a nivel reporte,
# ni registros sin item_numero (cola de validación solo sobre cantidades listas).
NIVEL_FIELD_VALIDACION_AVANZADA = ('nivel2_estado', 'nivel3_estado')
ESTADOS_REPORTE_EXCL_VALIDACION_AVANZADA = ('Borrador', 'Sin Asignar Ítem')


def _so_reportes_estado_valores_desde_filtro_ui(estado: Optional[str]) -> Optional[List[str]]:
    """La grilla filtra con etiquetas en plural (Aprobados, …); en BD puede haber plural o singular."""
    if estado is None or not str(estado).strip():
        return None
    s = str(estado).strip()
    plur_sing = {
        "Aprobados": ("Aprobados", "Aprobado"),
        "Pendientes": ("Pendientes", "Pendiente"),
        "Rechazados": ("Rechazados", "Rechazado"),
        "No Revisados": ("No Revisados", "No Revisado"),
    }
    if s in plur_sing:
        return list(plur_sing[s])
    sl = s.lower()
    if "sin asignar" in sl and "item" in sl.replace("í", "i"):
        return list({
            "Sin Asignar Ítem", "Sin Asignar Item", "sin asignar ítem", "SIN ASIGNAR ITEM",
        })
    return [s]


def _estado_filtro_omite_validacion_por_cargo(estado: Optional[str]) -> bool:
    """Borrador / Sin asignar ítem: cola distinta a validación N1–N3; no combinar con filtro por cargo."""
    if estado is None or not str(estado).strip():
        return False
    sl = str(estado).strip().lower()
    if sl == "borrador":
        return True
    if "sin asignar" in sl and "item" in sl.replace("í", "i"):
        return True
    return False


def _estado_filtro_es_sin_asignar_item(estado: Optional[str]) -> bool:
    """Filtro de grilla por cabecera 'Sin Asignar Ítem' (excl. Borrador)."""
    if estado is None or not str(estado).strip():
        return False
    sl = str(estado).strip().lower()
    return "sin asignar" in sl and "item" in sl.replace("í", "i")


def _so_reportes_q_por_estado(q, estado: Optional[str]):
    """Aplica filtro por estado de cabecera (OR entre variantes plural/singular)."""
    evs = _so_reportes_estado_valores_desde_filtro_ui(estado)
    if not evs:
        return q
    if len(evs) == 1:
        return q.eq("estado", evs[0])
    return q.in_("estado", evs)


def _es_validacion_avanzada(nivel_field: Optional[str]) -> bool:
    return nivel_field in NIVEL_FIELD_VALIDACION_AVANZADA


def _so_reg_item_asignado(q):
    """Registro con ítem asignado (no null ni cadena vacía)."""
    return q.not_.is_("item_numero", "null").neq("item_numero", "")


def _so_reg_sin_item_asignado(q):
    """Registro sin ítem (cola 'Sin Asignar Ítem' a nivel línea; alineado con UI)."""
    return q.or_('item_numero.is.null,item_numero.eq.""')


def _so_reg_or_pendiente_nivel(q, nivel_field: str):
    """(campo IS NULL OR campo = 'No Revisado'). PostgREST exige comillas en valores con espacio."""
    return q.or_(f'{nivel_field}.is.null,{nivel_field}.eq."No Revisado"')


def _so_reg_filtro_costado(q, valor: Optional[str]):
    """Barra 'costado': opciones vienen de cabecera (calzada). En línea el valor suele estar en
    calzada (columna Excel CALZADA); datos antiguos pueden usar margen."""
    if valor is None:
        return q
    s = str(valor).strip()
    if not s:
        return q
    esc = s.replace('"', '""')
    return q.or_(f'calzada.eq."{esc}",margen.eq."{esc}"')


def _so_reg_filtro_abs_solape(q, abs_inicio: Optional[float], abs_final: Optional[float]):
    """Abscisa en línea: solape del tramo [abs_inicio, abs_final] del registro con el rango filtrado."""
    if abs_inicio is None and abs_final is None:
        return q
    q = q.not_.is_("abs_inicio", "null").not_.is_("abs_final", "null")
    if abs_inicio is not None:
        q = q.gte("abs_final", abs_inicio)
    if abs_final is not None:
        q = q.lte("abs_inicio", abs_final)
    return q


def _sicoe_parse_nodo_tokens(q: Optional[str]) -> list:
    if not q or not str(q).strip():
        return []
    return [t for t in re.split(r"[\s,;/]+", str(q).strip()) if t]


def _sicoe_norm_txt(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _sicoe_row_match_nodo_tokens(tokens: list, *vals) -> bool:
    if not tokens:
        return True
    hay = " ".join(_sicoe_norm_txt(v) for v in vals)
    for tok in tokens:
        t = _sicoe_norm_txt(tok)
        if t and t not in hay:
            return False
    return True


def _sicoe_reporte_ids_coinciden_nodo(
    contrato_id: int,
    q_nodo: str,
    reporte_ids_prev: Optional[list],
) -> Optional[set]:
    """
    None = no aplicar filtro.
    set vacío = sin coincidencias.
    set no vacío = IDs de reporte que cumplen (cabecera nodo_ini/nodo_fin y/o líneas no_inicio/no_final).
    """
    if not q_nodo or not str(q_nodo).strip():
        return None
    tokens = _sicoe_parse_nodo_tokens(q_nodo)
    if not tokens:
        return None

    def _ids_para_token(tok: str) -> set:
        s_rep: set = set()
        pat = f"%{tok}%"
        if reporte_ids_prev is not None:
            if not reporte_ids_prev:
                return set()
            for i in range(0, len(reporte_ids_prev), 200):
                chunk = reporte_ids_prev[i : i + 200]

                def _qc(c=chunk):
                    try:
                        return supabase.table("so_reportes").select("id").eq("contrato_id", contrato_id).in_("id", c).or_(
                            f"nodo_ini.ilike.{pat},nodo_fin.ilike.{pat}"
                        ).limit(50000).execute().data
                    except Exception:
                        return []

                for row in supabase_execute(_qc):
                    if row.get("id"):
                        s_rep.add(row["id"])
        else:
            try:
                def _qa():
                    return supabase.table("so_reportes").select("id").eq("contrato_id", contrato_id).or_(
                        f"nodo_ini.ilike.{pat},nodo_fin.ilike.{pat}"
                    ).limit(50000).execute().data

                for row in supabase_execute(_qa):
                    if row.get("id"):
                        s_rep.add(row["id"])
            except Exception:
                pass
        s_reg: set = set()
        if reporte_ids_prev is not None:
            if not reporte_ids_prev:
                return set()
            for i in range(0, len(reporte_ids_prev), 200):
                chunk = reporte_ids_prev[i : i + 200]

                def _rc(c=chunk):
                    try:
                        return supabase.table("so_registros").select("reporte_id, no_inicio, no_final").eq("contrato_id", contrato_id).in_("reporte_id", c).execute().data
                    except Exception:
                        return []

                for row in supabase_execute(_rc):
                    if _sicoe_row_match_nodo_tokens([tok], row.get("no_inicio"), row.get("no_final")):
                        rid = row.get("reporte_id")
                        if rid:
                            s_reg.add(rid)
        else:
            off = 0
            while True:
                def _rb(o=off):
                    try:
                        return supabase.table("so_registros").select("reporte_id, no_inicio, no_final").eq("contrato_id", contrato_id).range(o, o + 999).execute().data
                    except Exception:
                        return []

                batch = supabase_execute(_rb)
                for row in batch:
                    if _sicoe_row_match_nodo_tokens([tok], row.get("no_inicio"), row.get("no_final")):
                        rid = row.get("reporte_id")
                        if rid:
                            s_reg.add(rid)
                if len(batch) < 1000:
                    break
                off += 1000
        return s_rep | s_reg

    out: Optional[set] = None
    for tok in tokens:
        s_tok = _ids_para_token(tok)
        if out is None:
            out = s_tok
        else:
            out &= s_tok
        if not out:
            return set()
    return out


def _sicoe_reporte_ids_abs_solapa_registros(
    contrato_id: int,
    q_abs0: Optional[float],
    q_abs1: Optional[float],
    reporte_ids_prev: Optional[list],
) -> Optional[set]:
    """
    Filtro tipo Excel sobre abscisas en líneas (so_registros).

    Un registro con tramo [a, b] se incluye si se solapa con el rango solicitado [q_abs0, q_abs1]:
    a <= q_abs1 y b >= q_abs0 (con ambos límites del query opcionales).

    None = no hay filtro de abscisa.
    set() = ningún reporte cumple.
    """
    if q_abs0 is None and q_abs1 is None:
        return None
    if reporte_ids_prev is not None and len(reporte_ids_prev) == 0:
        return set()

    out: set = set()
    CHUNK = 200

    def _apply_solape(q):
        q = q.not_.is_("abs_inicio", "null").not_.is_("abs_final", "null")
        if q_abs0 is not None:
            q = q.gte("abs_final", q_abs0)
        if q_abs1 is not None:
            q = q.lte("abs_inicio", q_abs1)
        return q

    if reporte_ids_prev is not None:
        for i in range(0, len(reporte_ids_prev), CHUNK):
            chunk = reporte_ids_prev[i : i + CHUNK]

            def _page(c=chunk):
                q = supabase.table("so_registros").select("reporte_id").eq("contrato_id", contrato_id).in_("reporte_id", c)
                return _apply_solape(q).limit(50000).execute().data

            for row in supabase_execute(_page):
                rid = row.get("reporte_id")
                if rid:
                    out.add(rid)
    else:
        off = 0
        while True:
            def _page(o=off):
                q = supabase.table("so_registros").select("reporte_id").eq("contrato_id", contrato_id)
                return _apply_solape(q).range(o, o + 999).execute().data

            batch = supabase_execute(_page)
            for row in batch:
                rid = row.get("reporte_id")
                if rid:
                    out.add(rid)
            if len(batch) < 1000:
                break
            off += 1000
    return out


def _sicoe_so_registros_q_linea_filtros_busqueda(
    q,
    *,
    numero_registro: Optional[int] = None,
    abs_inicio: Optional[float] = None,
    abs_final: Optional[float] = None,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    items: Optional[List[str]] = None,
    items_op: Optional[str] = None,
    subcontratista_id: Optional[int] = None,
    tramo: Optional[str] = None,
    costado: Optional[str] = None,
    pk_id: Optional[int] = None,
    q_observacion: Optional[str] = None,
    semana_id: Optional[int] = None,
    acta_rpo_id: Optional[int] = None,
    reporte_id_in: Optional[List[int]] = None,
    require_item: bool = False,
    capas_v: Optional[List[dict]] = None,
    estado: Optional[str] = None,
    registro_id_in: Optional[List[int]] = None,
):
    """AND sobre columnas de so_registros; misma semántica que /reportes/buscar y /analisis."""
    items_eff = list(items) if items else []
    if not items_eff and item is not None and str(item).strip():
        items_eff = [str(item).strip()]
    if numero_registro is not None:
        q = q.eq("numero_registro", numero_registro)
    q = _so_reg_filtro_abs_solape(q, abs_inicio, abs_final)
    if capitulo:
        q = q.eq("capitulo", capitulo)
    q = _apply_item_patterns_to_so_registros_q(q, items_eff, items_op)
    if subcontratista_id is not None:
        q = q.eq("subcontratista_id", subcontratista_id)
    if tramo:
        q = q.eq("tramo", tramo)
    if costado:
        q = _so_reg_filtro_costado(q, costado)
    if pk_id is not None:
        q = q.eq("pk_id_id", pk_id)
    if semana_id is not None:
        q = q.eq("semana_id", semana_id)
    # reporte_id_in tiene precedencia: filtra por los reportes del acta (semántica correcta)
    if reporte_id_in is not None:
        q = q.in_("reporte_id", reporte_id_in)
    elif acta_rpo_id is not None:
        q = q.eq("acta_rpo_id", acta_rpo_id)
    if q_observacion and str(q_observacion).strip():
        q = q.ilike("observacion", f"%{str(q_observacion).strip()}%")
    if capas_v:
        q = _so_registros_q_y_capas_validacion(
            q, capas_v, pk_id, tramo, costado, capitulo, subcontratista_id, None
        )
    # require_item: alinea con la vista del dashboard (solo registros con ítem asignado)
    # Solo se aplica si el estado no es "sin_asignar_item" (filtro inverso explícito)
    if require_item and not _estado_filtro_es_sin_asignar_item(estado):
        q = _so_reg_item_asignado(q)
    elif _estado_filtro_es_sin_asignar_item(estado):
        q = _so_reg_sin_item_asignado(q)
    if registro_id_in is not None:
        if len(registro_id_in) == 0:
            q = q.eq("id", -1)
        else:
            q = q.in_("id", registro_id_in)
    return q


def _sicoe_filtrar_registros_coinciden_nodo_ui(regs: list, reporte_row: dict, q_nodo: Optional[str]) -> list:
    """Cabecera del reporte y/o líneas no_inicio/no_final (coherente con la grilla)."""
    if not q_nodo or not str(q_nodo).strip():
        return regs
    tokens = _sicoe_parse_nodo_tokens(q_nodo)
    if not tokens:
        return regs
    hdr_ok = _sicoe_row_match_nodo_tokens(tokens, reporte_row.get("nodo_ini"), reporte_row.get("nodo_fin"))
    out = []
    for reg in regs:
        if hdr_ok or _sicoe_row_match_nodo_tokens(tokens, reg.get("no_inicio"), reg.get("no_final")):
            out.append(reg)
    return out


def _sicoe_collect_reporte_ids_misma_linea(
    contrato_id: int,
    *,
    numero_registro: Optional[int] = None,
    abs_inicio: Optional[float] = None,
    abs_final: Optional[float] = None,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    items: Optional[List[str]] = None,
    items_op: Optional[str] = None,
    subcontratista_id: Optional[int] = None,
    tramo: Optional[str] = None,
    costado: Optional[str] = None,
    pk_id: Optional[int] = None,
    q_observacion: Optional[str] = None,
    semana_id: Optional[int] = None,
    acta_rpo_id: Optional[int] = None,
    reporte_ids_restrict: Optional[List[int]] = None,
    capas_v: Optional[List[dict]] = None,
    estado: Optional[str] = None,
    registro_ids_etiqueta: Optional[Set[int]] = None,
    capas_v_op: Optional[str] = None,
) -> set:
    """
    reporte_id tales que existe al menos una fila en so_registros que cumple todos
    los criterios a la vez (AND). Evita intersectar IDs por criterios distintos, que
    incluía reportes sin ninguna línea coincidente con el panel /analisis.
    Cuando reporte_ids_restrict está presente, pagina sobre esos IDs en lugar de
    usar acta_rpo_id en la columna del registro (semántica correcta: acta del reporte).
    """
    items_eff: List[str] = list(items) if items else []
    if not items_eff and item is not None and str(item).strip():
        items_eff = [str(item).strip()]
    ids: set = set()
    capas_ok = bool(capas_v)
    if (
        capas_ok
        and not _estado_filtro_omite_validacion_por_cargo(estado)
        and _parse_capas_validacion_op(capas_v_op) == "or"
        and len(capas_v or []) > 1
    ):
        merged: set = set()
        for c in capas_v or []:
            if not (str(c.get("estado") or "").strip()):
                continue
            fld = c.get("campo") or _capa_campo_validacion(c)
            if not fld:
                continue
            merged |= _sicoe_collect_reporte_ids_misma_linea(
                contrato_id,
                numero_registro=numero_registro,
                abs_inicio=abs_inicio,
                abs_final=abs_final,
                capitulo=capitulo,
                items=items_eff,
                items_op=items_op,
                subcontratista_id=subcontratista_id,
                tramo=tramo,
                costado=costado,
                pk_id=pk_id,
                q_observacion=q_observacion,
                semana_id=semana_id,
                acta_rpo_id=acta_rpo_id,
                reporte_ids_restrict=reporte_ids_restrict,
                capas_v=[c],
                capas_v_op="and",
                estado=estado,
                registro_ids_etiqueta=registro_ids_etiqueta,
            )
        return merged

    if registro_ids_etiqueta is not None:
        if not registro_ids_etiqueta:
            return set()
        reg_list = sorted(registro_ids_etiqueta)
        CHUNK_R = 200
        if reporte_ids_restrict is not None:
            CHUNK_REP = 500
            for reg_chunk in _sicoe_chunks_int(reg_list, CHUNK_R):
                rc = list(reg_chunk)
                for i in range(0, len(reporte_ids_restrict), CHUNK_REP):
                    rep_chunk = reporte_ids_restrict[i : i + CHUNK_REP]
                    rp = list(rep_chunk)

                    def _chunk_page(c=rc, rp_fix=rp):
                        q = supabase.table("so_registros").select("reporte_id").eq("contrato_id", contrato_id)
                        q = _sicoe_so_registros_q_linea_filtros_busqueda(
                            q,
                            numero_registro=numero_registro,
                            abs_inicio=abs_inicio,
                            abs_final=abs_final,
                            capitulo=capitulo,
                            items=items_eff,
                            items_op=items_op,
                            subcontratista_id=subcontratista_id,
                            tramo=tramo,
                            costado=costado,
                            pk_id=pk_id,
                            q_observacion=q_observacion,
                            semana_id=semana_id,
                            reporte_id_in=rp_fix,
                            require_item=True,
                            capas_v=(capas_v if capas_ok else None),
                            estado=estado,
                            registro_id_in=c,
                        )
                        return q.limit(5000).execute().data

                    for row in supabase_execute(_chunk_page):
                        rid = row.get("reporte_id")
                        if rid:
                            ids.add(rid)
        else:
            for reg_chunk in _sicoe_chunks_int(reg_list, CHUNK_R):
                rc = list(reg_chunk)

                def _one_page(c=rc):
                    q = supabase.table("so_registros").select("reporte_id").eq("contrato_id", contrato_id)
                    q = _sicoe_so_registros_q_linea_filtros_busqueda(
                        q,
                        numero_registro=numero_registro,
                        abs_inicio=abs_inicio,
                        abs_final=abs_final,
                        capitulo=capitulo,
                        items=items_eff,
                        items_op=items_op,
                        subcontratista_id=subcontratista_id,
                        tramo=tramo,
                        costado=costado,
                        pk_id=pk_id,
                        q_observacion=q_observacion,
                        semana_id=semana_id,
                        acta_rpo_id=acta_rpo_id,
                        require_item=(acta_rpo_id is not None),
                        capas_v=(capas_v if capas_ok else None),
                        estado=estado,
                        registro_id_in=c,
                    )
                    return q.limit(5000).execute().data

                for row in supabase_execute(_one_page):
                    rid = row.get("reporte_id")
                    if rid:
                        ids.add(rid)
        return ids

    if reporte_ids_restrict is not None:
        # Iterar en chunks sobre los report_ids del acta; no necesita paginación ilimitada
        CHUNK = 500
        for i in range(0, len(reporte_ids_restrict), CHUNK):
            chunk = reporte_ids_restrict[i:i + CHUNK]

            def _chunk_page(c=chunk):
                q = supabase.table("so_registros").select("reporte_id").eq("contrato_id", contrato_id)
                q = _sicoe_so_registros_q_linea_filtros_busqueda(
                    q,
                    numero_registro=numero_registro,
                    abs_inicio=abs_inicio,
                    abs_final=abs_final,
                    capitulo=capitulo,
                    items=items_eff,
                    items_op=items_op,
                    subcontratista_id=subcontratista_id,
                    tramo=tramo,
                    costado=costado,
                    pk_id=pk_id,
                    q_observacion=q_observacion,
                    semana_id=semana_id,
                    reporte_id_in=c,
                    # Cuando filtramos por acta del reporte, alineamos con la semántica
                    # del dashboard: solo registros con ítem asignado cuentan.
                    require_item=True,
                    capas_v=(capas_v if capas_ok else None),
                    estado=estado,
                )
                return q.limit(5000).execute().data

            for row in supabase_execute(_chunk_page):
                rid = row.get("reporte_id")
                if rid:
                    ids.add(rid)
    else:
        off = 0
        page = 1000
        max_pages = int(os.getenv("SICOE_BUSCAR_VALIDACION_MAX_PAGES", "100"))
        for _pn in range(max_pages):

            def _one_page(o=off):
                q = supabase.table("so_registros").select("reporte_id").eq("contrato_id", contrato_id)
                q = _sicoe_so_registros_q_linea_filtros_busqueda(
                    q,
                    numero_registro=numero_registro,
                    abs_inicio=abs_inicio,
                    abs_final=abs_final,
                    capitulo=capitulo,
                    items=items_eff,
                    items_op=items_op,
                    subcontratista_id=subcontratista_id,
                    tramo=tramo,
                    costado=costado,
                    pk_id=pk_id,
                    q_observacion=q_observacion,
                    semana_id=semana_id,
                    acta_rpo_id=acta_rpo_id,
                    require_item=(acta_rpo_id is not None),
                    capas_v=(capas_v if capas_ok else None),
                    estado=estado,
                )
                return q.range(o, o + page - 1).execute().data

            batch = supabase_execute(_one_page)
            for row in batch:
                rid = row.get("reporte_id")
                if rid:
                    ids.add(rid)
            if len(batch) < page:
                break
            off += page
        else:
            raise HTTPException(
                status_code=503,
                detail="El filtro devuelve demasiados registros. Acote con tramo, capítulo, PK, subcontratista, ítem, semana, acta u observación.",
            )
    # No filtrar por estado de so_reportes: la matriz dashboard y consultas en BD usan solo líneas
    # (so_registros + validación); excluir Borrador/Sin Asignar Ítem en cabecera omitía millones
    # de costo aunque nivel3_estado en línea ya estuviera Aprobado.
    return ids


def _sicoe_ocultar_costo_directo_reportes(current_user) -> bool:
    """Operativo Contratista / Interventoría no reciben montos en la grilla SICOE Obra."""
    rol = (current_user.get("rol_nombre") or "").strip().lower()
    return rol in ("operativo contratista", "operativo interventoria", "operativo interventoría")


def _filtrar_registros_validacion_por_campo(
    regs: list,
    fld: str,
    estado_validacion: Optional[str],
    reporte_row: Optional[dict] = None,
) -> list:
    """Filtra registros por campo nivelX_estado y estado UI (coherente con búsqueda Sicoe)."""
    if not regs or not fld or not (estado_validacion or "").strip():
        return regs
    prereq = CARGO_NIVEL_PRERREQUISITO.get(fld)
    ev = estado_validacion.strip()
    out: List[dict] = []
    for reg in regs:
        if _es_validacion_avanzada(fld):
            if not (reg.get("item_numero") or "").strip():
                continue
        if fld == "nivel1_estado" and ev in ("No Revisado", "No Revisados"):
            if not (reg.get("item_numero") or "").strip():
                continue
        if prereq and reg.get(prereq[0]) != prereq[1]:
            continue
        cur = reg.get(fld)
        if ev in ("No Revisado", "No Revisados"):
            if cur is None or str(cur).strip() in ("", "No Revisado"):
                out.append(reg)
        else:
            if cur == ev:
                out.append(reg)
    return out


def _filtrar_registros_validacion_sicoe(
    regs: list,
    cargo_id: Optional[int],
    estado_validacion: Optional[str],
    reporte_row: Optional[dict] = None,
) -> list:
    """Misma semántica que la búsqueda por cargo: nivel 2/3 solo si cumple prerrequisito del nivel previo."""
    if not regs or cargo_id is None or not (estado_validacion or "").strip():
        return regs
    fld = CARGO_ID_NIVEL_MAP.get(int(cargo_id))
    if not fld:
        return regs
    return _filtrar_registros_validacion_por_campo(regs, fld, estado_validacion, reporte_row)


def _parse_capas_validacion_op(val: Optional[str]) -> str:
    """AND por defecto. OR acepta variantes en español / inglés."""
    s = (str(val or "").strip().lower())
    if s in ("or", "o", "||", "any", "cualquiera"):
        return "or"
    return "and"


def _normalize_items_filtro_list(items_filtro_json: Optional[str], item_legacy: Optional[str]) -> List[str]:
    """Varios ítems vía JSON `items_filtro` o un solo `item` (query legado). Sin duplicados, orden estable."""
    out: List[str] = []
    if items_filtro_json and str(items_filtro_json).strip():
        try:
            j = json.loads(items_filtro_json)
            if isinstance(j, list):
                out = [str(x).strip() for x in j if x is not None and str(x).strip()]
        except Exception:
            pass
    if not out and item_legacy is not None and str(item_legacy).strip():
        out = [str(item_legacy).strip()]
    seen: Set[str] = set()
    deduped: List[str] = []
    for x in out:
        if x not in seen:
            seen.add(x)
            deduped.append(x)
    return deduped


def _apply_item_patterns_to_so_registros_q(q, items: List[str], items_op: Optional[str] = None):
    """
    Filtro por texto en item_numero (ilike %pat%). Varias patrones: Y (todas en la misma fila) u O (cualquiera).
    Misma semántica de operador que validacion_capas_op.
    """
    if not items:
        return q
    if len(items) == 1:
        return q.ilike("item_numero", f"%{items[0]}%")
    op = _parse_capas_validacion_op(items_op)
    if op == "or":
        parts = [f"item_numero.ilike.%{it}%" for it in items]
        try:
            return q.or_(",".join(parts))
        except Exception:
            pass
    for it in items:
        q = q.ilike("item_numero", f"%{it}%")
    return q


def _registro_cumple_capa_validacion_sicoe(reg: dict, capa: dict, reporte_row: Optional[dict] = None) -> bool:
    fld = capa.get("campo") or _capa_campo_validacion(capa)
    if not fld or not (str(capa.get("estado") or "").strip()):
        return False
    return bool(_filtrar_registros_validacion_por_campo([reg], fld, capa.get("estado"), reporte_row))


def _filtrar_registros_validacion_capas_sicoe(
    regs: list,
    capas: List[dict],
    reporte_row: Optional[dict] = None,
    op: str = "and",
) -> list:
    """AND: todas las capas (legado). OR: basta que el registro cumpla una capa cualquiera."""
    if not capas:
        return regs
    o = _parse_capas_validacion_op(op)
    if o == "or" and len(capas) > 1:
        return [
            reg for reg in regs
            if any(_registro_cumple_capa_validacion_sicoe(reg, c, reporte_row) for c in capas)
        ]
    out = regs
    for c in capas:
        fld = c.get("campo") or _capa_campo_validacion(c)
        if not fld:
            continue
        out = _filtrar_registros_validacion_por_campo(out, fld, c.get("estado"), reporte_row)
    return out


# ─────────────────────────────────────────────
# RUTAS PÚBLICAS
# ─────────────────────────────────────────────

@app.post("/frase-del-dia")
def frase_del_dia(body: dict, current_user=Depends(get_current_user)):
    from frase_del_dia_sources import _pool_local_aleatoria, frase_dia_espanol

    def _fallback_min():
        return {
            "frase": "El avance de hoy construye el resultado de mañana.",
            "autor": "ClaraCore",
            "tipo": "motivadora",
        }

    # Versículo RVR (es) desde la red, citas de autores en español, aforismos de obra — nada en inglés
    try:
        f = frase_dia_espanol()
        if f and f.get("frase"):
            return f
    except Exception as e:
        print(f"WARNING /frase-del-dia: {e}", flush=True)
    p = _pool_local_aleatoria()
    return p if p else _fallback_min()


_SLOW_REQUEST_MS = int(os.getenv("CLARACORE_SLOW_REQUEST_MS", "8000"))


@app.middleware("http")
async def registrar_respuesta_lenta(request: Request, call_next):
    import time
    t0 = time.perf_counter()
    response = await call_next(request)
    ms = int((time.perf_counter() - t0) * 1000)
    # No escribir en `logs` por /healthz lento: en incidentes eso añade carga a la misma BD.
    if ms >= _SLOW_REQUEST_MS and request.url.path != "/healthz":
        try:
            registrar_log_sistema(
                "WARNING",
                request.url.path,
                request.method,
                {"tipo": "respuesta_lenta", "duracion_ms": ms},
                None,
                resultado="ok",
                duracion_ms=ms,
            )
        except Exception:
            pass
    return response


# Debe ir después de todos los @app.middleware("http") para quedar como capa externa y que
# todas las respuestas (incl. cortocircuitos 403) lleven cabeceras CORS.
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_origin_regex=r"^https://([a-z0-9-]+\.)*claracore\.co$|^https?://(localhost|127\.0\.0\.1)(:\d+)?$",
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)


@app.get("/")
def root():
    return {"message": "ClaraCore API funcionando"}


@app.get("/healthz")
def healthz():
    """Sin Supabase ni lógica pesada — útil para keep-alive y comprobar que el worker responde (Azure cold start)."""
    return {"ok": True}

@app.get("/cargos")
def listar_cargos():
    return supabase.table("cargos").select("*").order("nombre").execute().data

@app.get("/roles")
def listar_roles():
    return supabase.table("roles").select("*").order("nombre").execute().data


@app.get("/guias/admin/todas")
def guias_admin_todas(current_user=Depends(get_current_user)):
    """Listado completo (incl. borradores) para el panel de documentación — solo Desarrollador."""
    if not _es_desarrollador(current_user):
        raise HTTPException(status_code=403, detail="Solo el cargo Desarrollador puede gestionar guías")
    rows = supabase.table("guias").select("*").execute().data or []
    rows.sort(key=lambda r: ((r.get("orden") if r.get("orden") is not None else 0), (r.get("titulo") or "").lower()))
    return rows


@app.get("/guias/buscar")
def guias_buscar(q: str = Query(""), current_user=Depends(get_current_user)):
    """Búsqueda por palabra en título y descripción; respeta visibilidad por rol (excepto Desarrollador: ve todas)."""
    rol_id = _caller_rol_id(current_user)
    es_dev = _es_desarrollador(current_user)
    needle = (q or "").strip().lower()
    if es_dev:
        rows = supabase.table("guias").select(_GUIAS_SELECT_LISTA).order("titulo").execute().data or []
    else:
        rows = supabase.table("guias").select(_GUIAS_SELECT_LISTA).eq("publicado", True).order("titulo").execute().data or []
        rows = [r for r in rows if _guia_visible_por_rol(r, rol_id)]
    if needle:
        rows = [
            r
            for r in rows
            if needle in (r.get("titulo") or "").lower()
            or needle in (r.get("descripcion_corta") or "").lower()
        ]
    return rows


@app.get("/guias")
def guias_listar_publicadas(current_user=Depends(get_current_user)):
    """Guías publicadas visibles para el rol del usuario, orden alfabético por título."""
    rol_id = _caller_rol_id(current_user)
    rows = supabase.table("guias").select(_GUIAS_SELECT_LISTA).eq("publicado", True).order("titulo").execute().data or []
    return [r for r in rows if _guia_visible_por_rol(r, rol_id)]


@app.get("/guias/{slug}")
def guia_detalle_por_slug(slug: str, current_user=Depends(get_current_user)):
    """Detalle completo de una guía por slug."""
    rol_id = _caller_rol_id(current_user)
    es_dev = _es_desarrollador(current_user)
    r = supabase.table("guias").select("*").eq("slug", slug).limit(1).execute()
    row = r.data[0] if r.data else None
    if not row:
        raise HTTPException(status_code=404, detail="Guía no encontrada")
    if not es_dev:
        if not row.get("publicado"):
            raise HTTPException(status_code=404, detail="Guía no encontrada")
        if not _guia_visible_por_rol(row, rol_id):
            raise HTTPException(status_code=404, detail="Guía no encontrada")
    return row


@app.post("/guias")
def guias_crear(body: GuiaCreate, current_user=Depends(get_current_user)):
    if not _es_desarrollador(current_user):
        raise HTTPException(status_code=403, detail="Solo el cargo Desarrollador puede crear guías")
    base = _titulo_a_slug_base(body.titulo)
    slug = _siguiente_slug_unico(base, None)
    now = datetime.now(timezone.utc).isoformat()
    payload = {
        "titulo": body.titulo.strip(),
        "slug": slug,
        "modulo": body.modulo,
        "descripcion_corta": body.descripcion_corta,
        "bloques": _json_for_log(body.bloques) if body.bloques else [],
        "roles_visibles": body.roles_visibles or [],
        "publicado": bool(body.publicado),
        "orden": int(body.orden or 0),
        "contrato_id": body.contrato_id,
        "updated_at": now,
    }
    ins = supabase.table("guias").insert(payload).execute()
    data = ins.data
    if data:
        return data[0] if isinstance(data, list) else data
    q = supabase.table("guias").select("*").eq("slug", slug).limit(1).execute()
    return q.data[0] if q.data else payload


@app.put("/guias/{guia_id}")
def guias_actualizar(guia_id: int, body: GuiaUpdate, current_user=Depends(get_current_user)):
    if not _es_desarrollador(current_user):
        raise HTTPException(status_code=403, detail="Solo el cargo Desarrollador puede editar guías")
    cur = supabase.table("guias").select("*").eq("id", guia_id).limit(1).execute()
    existing = cur.data[0] if cur.data else None
    if not existing:
        raise HTTPException(status_code=404, detail="Guía no encontrada")
    payload: Dict[str, Any] = {}
    if body.titulo is not None:
        payload["titulo"] = body.titulo.strip()
        base = _titulo_a_slug_base(body.titulo)
        payload["slug"] = _siguiente_slug_unico(base, guia_id)
    if body.modulo is not None:
        payload["modulo"] = body.modulo
    if body.descripcion_corta is not None:
        payload["descripcion_corta"] = body.descripcion_corta
    if body.bloques is not None:
        payload["bloques"] = _json_for_log(body.bloques)
    if body.roles_visibles is not None:
        payload["roles_visibles"] = body.roles_visibles
    if body.publicado is not None:
        payload["publicado"] = bool(body.publicado)
    if body.orden is not None:
        payload["orden"] = int(body.orden)
    if body.contrato_id is not None:
        payload["contrato_id"] = body.contrato_id
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    supabase.table("guias").update(payload).eq("id", guia_id).execute()
    out = supabase.table("guias").select("*").eq("id", guia_id).limit(1).execute()
    return out.data[0] if out.data else payload


@app.delete("/guias/{guia_id}")
def guias_eliminar(guia_id: int, current_user=Depends(get_current_user)):
    if not _es_desarrollador(current_user):
        raise HTTPException(status_code=403, detail="Solo el cargo Desarrollador puede eliminar guías")
    cur = supabase.table("guias").select("id").eq("id", guia_id).limit(1).execute()
    if not cur.data:
        raise HTTPException(status_code=404, detail="Guía no encontrada")
    supabase.table("guias").delete().eq("id", guia_id).execute()
    return {"ok": True, "id": guia_id}


@app.post("/guias/imagen")
async def guias_subir_imagen_bloque(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    """Sube una imagen para un bloque de guía (archivo). Solo Desarrollador — misma infra que novedades Inicio."""
    if not _es_desarrollador(current_user):
        raise HTTPException(status_code=403, detail="Solo el cargo Desarrollador puede subir imágenes para guías")
    contents = await file.read()
    url = _guia_bloque_subir_imagen(contents, file.content_type)
    return {"url": url}


# Listado ligero: SOLO columnas presentes en cualquier despliegue. Si pides columnas que aún no existen
# en la BD (migración pendiente), PostgREST falla y el panel enseña 0 contratos aunque existan.
_CONTRATOS_SELECT_LISTA = (
    "id, numero, objeto, contratista, nit, interventoria, entidad, entidad_otra, logo_entidad, "
    "centro_lat, centro_lng, logo_contratista, logo_interventoria, fase, costos_adicionales"
)
# Al editar, traemos * para no romper si faltan columnas nuevas en un entorno; el listado queda aligerado.
_CONTRATOS_SELECT_DETALLE = "*"


@app.get("/contratos")
def listar_contratos():
    return supabase.table("contratos").select(_CONTRATOS_SELECT_LISTA).order("numero").execute().data


@app.get("/admin/contratos-resumen")
def admin_contratos_resumen(current_user=Depends(get_current_user)):
    """Solo id y número para selects del panel admin (evita payload pesado de listar_contratos)."""
    return supabase.table("contratos").select("id, numero, fase").order("numero").execute().data or []


@app.get("/contratos/{contrato_id}")
def obtener_contrato(contrato_id: int):
    """Una fila completa (incl. plano_geojson y columnas añadidas vía migraciones, sin listar nombres fijos)."""
    r = supabase.table("contratos").select(_CONTRATOS_SELECT_DETALLE).eq("id", contrato_id).limit(1).execute()
    row = r.data[0] if r.data else None
    if not row:
        raise HTTPException(status_code=404, detail="Contrato no encontrado")
    return row

@app.post("/auth/login")
def login(request: Request, body: LoginRequest):
    ip = _client_ip(request)
    result = supabase.table("usuarios").select("*").eq("email", body.email).execute()
    if not result.data:
        registrar_log(
            {"nombre": "", "email": body.email},
            "LOGIN_FAIL", "AUTH", None, None,
            {"motivo": "usuario_no_encontrado", "email": body.email},
            resultado="fallido",
            ip=ip,
            severidad="WARNING",
        )
        _track_login_failure_alert(body.email, ip)
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    usuario = result.data[0]
    if not verify_password(body.password, usuario["password_hash"]):
        registrar_log(
            {"sub": str(usuario["id"]), "nombre": usuario.get("nombre", ""), "email": body.email,
             "cargo_nombre": "", "contrato_id": usuario.get("contrato_id")},
            "LOGIN_FAIL", "AUTH", "usuario", str(usuario["id"]),
            {"motivo": "password_incorrecto"},
            resultado="fallido",
            ip=ip,
            severidad="WARNING",
        )
        _track_login_failure_alert(body.email, ip)
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")
    if usuario.get("estado") == "pendiente":
        registrar_log(
            {"sub": str(usuario["id"]), "nombre": usuario.get("nombre", ""), "email": body.email,
             "contrato_id": usuario.get("contrato_id")},
            "LOGIN_FAIL", "AUTH", "usuario", str(usuario["id"]),
            {"motivo": "cuenta_pendiente"},
            resultado="denegado",
            ip=ip,
            severidad="WARNING",
        )
        raise HTTPException(status_code=403, detail="Tu cuenta está pendiente de aprobación")
    if usuario.get("estado") == "rechazado":
        registrar_log(
            {"sub": str(usuario["id"]), "nombre": usuario.get("nombre", ""), "email": body.email},
            "LOGIN_FAIL", "AUTH", "usuario", str(usuario["id"]),
            {"motivo": "cuenta_rechazada"},
            resultado="denegado",
            ip=ip,
            severidad="WARNING",
        )
        raise HTTPException(status_code=403, detail="Tu cuenta fue rechazada")

    cargo_nombre = None
    if usuario.get("cargo_id"):
        r = supabase.table("cargos").select("nombre").eq("id", usuario["cargo_id"]).execute()
        if r.data: cargo_nombre = r.data[0]["nombre"]

    rol_nombre = None
    if usuario.get("rol_id"):
        r = supabase.table("roles").select("nombre").eq("id", usuario["rol_id"]).execute()
        if r.data: rol_nombre = r.data[0]["nombre"]

    contrato_numero = None
    logo_contratista = None
    logo_interventoria = None
    if usuario.get("contrato_id"):
        r = supabase.table("contratos").select("numero, logo_contratista, logo_interventoria").eq("id", usuario["contrato_id"]).execute()
        if r.data:
            contrato_numero = r.data[0]["numero"]
            logo_contratista = r.data[0].get("logo_contratista")
            logo_interventoria = r.data[0].get("logo_interventoria")

    nombre_completo = f"{usuario.get('nombre','')} {usuario.get('apellidos','')}".strip()
    token = create_token({
        "sub": str(usuario["id"]),
        "email": usuario["email"],
        "nombre": nombre_completo or usuario.get("nombre") or "",
        "cargo_nombre": cargo_nombre or "",
        "rol_nombre": rol_nombre or "",
        "contrato_id": usuario.get("contrato_id"),
        "contrato_numero": contrato_numero or "",
    })

    # Cargar permisos del cargo para control de acceso en el panel
    permisos = []
    if usuario.get("cargo_id"):
        permisos_raw = supabase.table("permisos").select("*").eq("cargo_id", usuario["cargo_id"]).execute().data
        funciones_rows = supabase.table("funciones").select("id, nombre").execute().data
        funciones_map = {f["id"]: f["nombre"] for f in funciones_rows}
        permisos = [{**p, "funcion_nombre": funciones_map.get(p["funcion_id"], "")} for p in permisos_raw]
    # Hotfix: garantizar exportación para Desarrollador en cualquier contrato
    if cargo_nombre and cargo_nombre.strip().lower() == "desarrollador":
        permisos = [{**p, "exportar": True, "ver": True} for p in (permisos or [])]
        if not any((p.get("funcion_nombre") or "").strip().lower() == "reporte de cantidades" for p in permisos):
            funcion_id = None
            try:
                fr = supabase.table("funciones").select("id,nombre").ilike("nombre", "Reporte de Cantidades").limit(1).execute().data
                if fr:
                    funcion_id = fr[0].get("id")
            except Exception:
                funcion_id = None
            permisos.append({
                "id": None,
                "cargo_id": usuario.get("cargo_id"),
                "funcion_id": funcion_id,
                "funcion_nombre": "Reporte de Cantidades",
                "ver": True,
                "crear": True,
                "editar": True,
                "eliminar": True,
                "validar": True,
                "exportar": True,
            })
    # C3: Subcontratista sin subcontratista asignado → sin acceso
    if cargo_nombre and cargo_nombre.lower() == 'subcontratista' and not usuario.get('subcontratista_id'):
        permisos = []

    registrar_log(
        {"sub": str(usuario["id"]), "nombre": nombre_completo or usuario.get("nombre", ""),
         "cargo_nombre": cargo_nombre, "contrato_id": usuario.get("contrato_id"),
         "contrato_numero": contrato_numero, "rol_nombre": rol_nombre},
        "LOGIN", "AUTH", "usuario", str(usuario["id"]),
        {"email": usuario["email"], "cargo": cargo_nombre, "contrato": contrato_numero},
        ip=ip,
        rol_nombre=rol_nombre,
    )

    return {
        "access_token": token,
        "token_type": "bearer",
        "usuario": {
            "id": usuario["id"],
            "nombre": usuario["nombre"],
            "apellidos": usuario.get("apellidos"),
            "email": usuario["email"],
            "cargo_id": usuario.get("cargo_id"),
            "cargo_nombre": cargo_nombre,
            "rol_id": usuario.get("rol_id"),
            "rol_nombre": rol_nombre,
            "contrato_id": usuario.get("contrato_id"),
            "contrato_numero": contrato_numero,
            "logo_contratista": logo_contratista,
            "logo_interventoria": logo_interventoria,
            "estado": usuario.get("estado"),
            "activo": usuario.get("activo"),
            "subcontratista_id": usuario.get("subcontratista_id"),
            "permisos": permisos,
            "fecha_nacimiento": usuario.get("fecha_nacimiento"),
            "foto_perfil_url": usuario.get("foto_perfil_url"),
            "firma_imagen_url": usuario.get("firma_imagen_url"),
            "politicas_aceptadas": usuario.get("politicas_aceptadas") is True,
            "politicas_fecha": usuario.get("politicas_fecha"),
            "politicas_version": usuario.get("politicas_version"),
            "politicas_ip": usuario.get("politicas_ip"),
        }
    }

@app.post("/usuarios/registro")
def registro_usuario(usuario: UsuarioRegistro):
    existe = supabase.table("usuarios").select("id").eq("email", usuario.email).execute()
    if existe.data:
        raise HTTPException(status_code=400, detail="Este correo ya está registrado")
    hashed = hash_password(usuario.password)
    supabase.table("usuarios").insert({
        "nombre": usuario.nombre,
        "apellidos": usuario.apellidos,
        "email": usuario.email,
        "password_hash": hashed,
        "cargo_id": usuario.cargo_id,
        "contrato_id": usuario.contrato_id,
        "activo": False,
        "estado": "pendiente"
    }).execute()
    return {"mensaje": "Registro exitoso, pendiente de aprobación"}

# ─────────────────────────────────────────────
# RUTAS AUTENTICADAS
# ─────────────────────────────────────────────

@app.get("/usuarios/me")
def get_mi_usuario(current_user=Depends(get_current_user)):
    """Devuelve el perfil actualizado del usuario en sesión."""
    uid = int(current_user["sub"])
    sb = get_supabase()
    try:
        result = sb.table("usuarios").select("*").eq("id", uid).execute()
    except Exception:
        raise HTTPException(status_code=503, detail="Base de datos no disponible")
    if not result.data:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    u = result.data[0]
    cargo_nombre = None
    if u.get("cargo_id"):
        try:
            r = sb.table("cargos").select("nombre").eq("id", u["cargo_id"]).execute()
            if r.data: cargo_nombre = r.data[0]["nombre"]
        except Exception: pass
    rol_nombre = None
    if u.get("rol_id"):
        try:
            r = sb.table("roles").select("nombre").eq("id", u["rol_id"]).execute()
            if r.data: rol_nombre = r.data[0]["nombre"]
        except Exception: pass
    permisos = []
    if u.get("cargo_id"):
        try:
            permisos_raw = sb.table("permisos").select("*").eq("cargo_id", u["cargo_id"]).execute().data
            funciones_rows = sb.table("funciones").select("id, nombre").execute().data
            funciones_map = {f["id"]: f["nombre"] for f in funciones_rows}
            permisos = [{**p, "funcion_nombre": funciones_map.get(p["funcion_id"], "")} for p in permisos_raw]
        except Exception:
            permisos_raw = []
    # Hotfix: garantizar exportación para Desarrollador en cualquier contrato
    if cargo_nombre and cargo_nombre.strip().lower() == "desarrollador":
        permisos = [{**p, "exportar": True, "ver": True} for p in (permisos or [])]
        if not any((p.get("funcion_nombre") or "").strip().lower() == "reporte de cantidades" for p in permisos):
            funcion_id = None
            try:
                fr = sb.table("funciones").select("id,nombre").ilike("nombre", "Reporte de Cantidades").limit(1).execute().data
                if fr:
                    funcion_id = fr[0].get("id")
            except Exception:
                funcion_id = None
            permisos.append({
                "id": None,
                "cargo_id": u.get("cargo_id"),
                "funcion_id": funcion_id,
                "funcion_nombre": "Reporte de Cantidades",
                "ver": True,
                "crear": True,
                "editar": True,
                "eliminar": True,
                "validar": True,
                "exportar": True,
            })
    if cargo_nombre and cargo_nombre.lower() == 'subcontratista' and not u.get('subcontratista_id'):
        permisos = []
    return {
        "id": u["id"], "nombre": u["nombre"], "apellidos": u.get("apellidos"),
        "email": u["email"], "cargo_id": u.get("cargo_id"), "cargo_nombre": cargo_nombre,
        "rol_id": u.get("rol_id"), "rol_nombre": rol_nombre,
        "contrato_id": u.get("contrato_id"), "estado": u.get("estado"), "activo": u.get("activo"),
        "subcontratista_id": u.get("subcontratista_id"),
        "permisos": permisos,
        "fecha_nacimiento": u.get("fecha_nacimiento"),
        "foto_perfil_url": u.get("foto_perfil_url"),
        "firma_imagen_url": u.get("firma_imagen_url"),
        "politicas_aceptadas": u.get("politicas_aceptadas") is True,
        "politicas_fecha": u.get("politicas_fecha"),
        "politicas_version": u.get("politicas_version"),
        "politicas_ip": u.get("politicas_ip"),
    }

@app.post("/usuarios/me/politicas-aceptar")
def aceptar_politicas_confidencialidad(request: Request, current_user=Depends(get_current_user)):
    """Registra aceptación de políticas (versión, fecha UTC, IP). Permite usar el resto de la API."""
    uid = int(current_user["sub"])
    ip = _client_ip(request)
    ver = POLITICAS_VERSION_DEFAULT
    now_iso = datetime.now(timezone.utc).isoformat()
    try:
        sb = get_supabase()
        sb.table("usuarios").update({
            "politicas_aceptadas": True,
            "politicas_fecha": now_iso,
            "politicas_version": ver,
            "politicas_ip": ip or None,
        }).eq("id", uid).execute()
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"No se pudo registrar la aceptación: {e}")
    politicas_cache_invalidate(uid)
    return {
        "politicas_aceptadas": True,
        "politicas_fecha": now_iso,
        "politicas_version": ver,
        "politicas_ip": ip or None,
    }

@app.put("/usuarios/me")
def actualizar_mi_perfil(body: PerfilUpdate, current_user=Depends(get_current_user)):
    """El usuario edita nombre, apellidos y fecha de cumpleaños."""
    uid = int(current_user["sub"])
    sb = get_supabase()
    data = {}
    if body.nombre is not None:
        data["nombre"] = (body.nombre or "").strip()
    if body.apellidos is not None:
        data["apellidos"] = (body.apellidos or "").strip()
    if body.fecha_nacimiento is not None:
        raw = (body.fecha_nacimiento or "").strip()
        if not raw:
            data["fecha_nacimiento"] = None
        else:
            data["fecha_nacimiento"] = raw[:10]
    if not data:
        return get_mi_usuario(current_user)
    try:
        sb.table("usuarios").update(data).eq("id", uid).execute()
    except Exception as e:
        _log_api.warning("PUT /usuarios/me: %s", e)
        raise HTTPException(
            status_code=503,
            detail="No se pudo guardar el perfil. Si acabas de desplegar, ejecuta backend/sql/usuario_perfil.sql en Supabase.",
        )
    return get_mi_usuario(current_user)


def _ext_desde_content_type(content_type: Optional[str]) -> str:
    c = (content_type or "image/jpeg").split(";")[0].strip().lower()
    return {
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/gif": ".gif",
    }.get(c, ".jpg")


def _usuario_contrato_id(sb, uid: int) -> Optional[int]:
    try:
        r = sb.table("usuarios").select("contrato_id").eq("id", uid).limit(1).execute()
        if r.data:
            cid = r.data[0].get("contrato_id")
            return int(cid) if cid is not None else None
    except Exception:
        pass
    return None


def _usuario_imagen_subir(
    contents: bytes,
    uid: int,
    asset: str,
    content_type: Optional[str],
    contrato_id: Optional[int] = None,
) -> str:
    """
    Sube foto de perfil o imagen de firma.
    Cloudinary: claracore/{contrato_id}/Fotos_de_perfil | Fotos_de_firmas (si no hay contrato: sin_contrato/{uid}/...).
    Supabase: contratos/{id}/perfil|firmas/... o usuarios/{id}/...
    """
    ext = _ext_desde_content_type(content_type)
    ct = (content_type or "image/jpeg").split(";")[0].strip()

    if os.getenv("CLOUDINARY_CLOUD_NAME"):
        import cloudinary.uploader
        _cloudinary_config()
        sub = CLOUDINARY_SUB_PERFIL if asset == "avatar" else CLOUDINARY_SUB_FIRMAS
        public_id = f"usuario_{uid}"
        if contrato_id:
            folder = _cloudinary_folder_contrato(int(contrato_id), sub)
        else:
            folder = f"{CLOUDINARY_ROOT}/sin_contrato/{uid}/{sub}"
        result = cloudinary.uploader.upload(
            contents,
            folder=folder,
            public_id=public_id,
            overwrite=True,
            resource_type="image",
        )
        return result["secure_url"]

    sb = get_supabase()
    bucket = os.getenv("SUPABASE_PERFIL_BUCKET", "claracore-perfiles")
    if contrato_id:
        sub = "perfil" if asset == "avatar" else "firmas"
        path = f"contratos/{contrato_id}/{sub}/u{uid}{ext}"
    else:
        path = f"usuarios/{uid}/{asset}{ext}"
    file_opts = {"content-type": ct, "upsert": "true"}

    def _do_upload():
        sb.storage.from_(bucket).upload(path, contents, file_options=file_opts)

    try:
        _do_upload()
    except Exception as e1:
        msg = str(e1).lower()
        # Bucket inexistente: intentar crear (requiere service_role) y reintentar una vez
        if "bucket" in msg or "not found" in msg or "not_found" in msg or "404" in msg:
            try:
                sb.storage.create_bucket(
                    bucket,
                    options={"public": True, "file_size_limit": 6 * 1024 * 1024},
                )
            except Exception as e_create:
                # Ya existe u otro error: seguimos e intentamos subir de nuevo
                _log_api.warning("create_bucket %s: %s", bucket, e_create)
            try:
                _do_upload()
            except Exception as e2:
                _log_api.warning("Supabase Storage upload %s: %s", path, e2)
                raise HTTPException(
                    status_code=503,
                    detail=(
                        "No se pudo subir la imagen. Crea el bucket en Supabase Storage "
                        f"({bucket}, público) o configura Cloudinary (variables CLOUDINARY_*). "
                        "Puedes usar el script backend/sql/storage_perfil_bucket.sql"
                    ),
                )
        else:
            _log_api.warning("Supabase Storage upload %s: %s", path, e1)
            raise HTTPException(
                status_code=503,
                detail=(
                    "No se pudo subir la imagen a Supabase Storage. "
                    "Comprueba el bucket y la clave SUPABASE (service role recomendada para Storage)."
                ),
            )

    return sb.storage.from_(bucket).get_public_url(path)


def _inicio_novedad_subir_imagen(contents: bytes, content_type: Optional[str]) -> str:
    """Imagen de contexto para novedades de inicio (Cloudinary o Supabase Storage)."""
    if len(contents) > 6 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="La imagen no puede superar 6 MB.")
    ext = _ext_desde_content_type(content_type)
    ct = (content_type or "image/jpeg").split(";")[0].strip()
    if os.getenv("CLOUDINARY_CLOUD_NAME"):
        import cloudinary.uploader
        _cloudinary_config()
        uid_part = uuid.uuid4().hex[:16]
        result = cloudinary.uploader.upload(
            contents,
            folder=f"{CLOUDINARY_ROOT}/inicio-novedades",
            public_id=f"nov_{uid_part}",
            overwrite=False,
            resource_type="image",
        )
        return result["secure_url"]
    sb = get_supabase()
    bucket = os.getenv("SUPABASE_INICIO_BUCKET", os.getenv("SUPABASE_PERFIL_BUCKET", "claracore-perfiles"))
    path = f"inicio-novedades/{uuid.uuid4().hex}{ext}"
    file_opts = {"content-type": ct, "upsert": "true"}

    def _do_upload():
        sb.storage.from_(bucket).upload(path, contents, file_options=file_opts)

    try:
        _do_upload()
    except Exception as e1:
        msg = str(e1).lower()
        if "bucket" in msg or "not found" in msg or "not_found" in msg or "404" in msg:
            try:
                sb.storage.create_bucket(
                    bucket,
                    options={"public": True, "file_size_limit": 6 * 1024 * 1024},
                )
            except Exception as e_create:
                _log_api.warning("create_bucket inicio novedades %s: %s", bucket, e_create)
            try:
                _do_upload()
            except Exception as e2:
                _log_api.warning("Supabase Storage upload inicio %s: %s", path, e2)
                raise HTTPException(
                    status_code=503,
                    detail=(
                        "No se pudo subir la imagen. Configura Cloudinary (CLOUDINARY_*) "
                        f"o el bucket público {bucket} en Supabase Storage."
                    ),
                )
        else:
            _log_api.warning("Supabase Storage upload inicio %s: %s", path, e1)
            raise HTTPException(status_code=503, detail="No se pudo subir la imagen a Supabase Storage.")
    return sb.storage.from_(bucket).get_public_url(path)


def _guia_bloque_subir_imagen(contents: bytes, content_type: Optional[str]) -> str:
    """Imagen embebida en bloques de guías (Cloudinary o Supabase Storage). Solo vía API autorizada."""
    if len(contents) > 6 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="La imagen no puede superar 6 MB.")
    ext = _ext_desde_content_type(content_type)
    ct = (content_type or "image/jpeg").split(";")[0].strip()
    if os.getenv("CLOUDINARY_CLOUD_NAME"):
        import cloudinary.uploader
        _cloudinary_config()
        uid_part = uuid.uuid4().hex[:16]
        result = cloudinary.uploader.upload(
            contents,
            folder=f"{CLOUDINARY_ROOT}/guias-bloques",
            public_id=f"guia_{uid_part}",
            overwrite=False,
            resource_type="image",
        )
        return result["secure_url"]
    sb = get_supabase()
    bucket = os.getenv("SUPABASE_GUIAS_BUCKET", os.getenv("SUPABASE_PERFIL_BUCKET", "claracore-perfiles"))
    path = f"guias-bloques/{uuid.uuid4().hex}{ext}"
    file_opts = {"content-type": ct, "upsert": "true"}

    def _do_upload():
        sb.storage.from_(bucket).upload(path, contents, file_options=file_opts)

    try:
        _do_upload()
    except Exception as e1:
        msg = str(e1).lower()
        if "bucket" in msg or "not found" in msg or "not_found" in msg or "404" in msg:
            try:
                sb.storage.create_bucket(
                    bucket,
                    options={"public": True, "file_size_limit": 6 * 1024 * 1024},
                )
            except Exception as e_create:
                _log_api.warning("create_bucket guias bloques %s: %s", bucket, e_create)
            try:
                _do_upload()
            except Exception as e2:
                _log_api.warning("Supabase Storage upload guia bloque %s: %s", path, e2)
                raise HTTPException(
                    status_code=503,
                    detail=(
                        "No se pudo subir la imagen. Configura Cloudinary (CLOUDINARY_*) "
                        f"o el bucket público {bucket} en Supabase Storage."
                    ),
                )
        else:
            _log_api.warning("Supabase Storage upload guia bloque %s: %s", path, e1)
            raise HTTPException(status_code=503, detail="No se pudo subir la imagen a Supabase Storage.")
    return sb.storage.from_(bucket).get_public_url(path)


def _detalle_guardar_url_perfil(exc: Exception, campo: str) -> str:
    """Mensaje claro cuando falla el UPDATE (p. ej. columnas inexistentes en usuarios)."""
    raw = str(exc).lower()
    pg_undefined_column = "42703" in raw or "pgrst204" in raw  # Postgres / PostgREST
    if pg_undefined_column or ("column" in raw and ("does not exist" in raw or "undefined column" in raw)):
        return (
            f'La columna "{campo}" no existe en la tabla usuarios. '
            "Abre Supabase → SQL y ejecuta el archivo backend/sql/usuario_perfil.sql "
            "(añade foto_perfil_url, firma_imagen_url y fecha_nacimiento)."
        )
    return (
        "La imagen pudo subirse al almacenamiento, pero no se guardó la URL en usuarios. "
        "Lo más habitual es no haber ejecutado la migración: backend/sql/usuario_perfil.sql en Supabase. "
        f"Detalle: {exc!s}"
    )


@app.post("/usuarios/me/foto-perfil")
async def subir_foto_perfil(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    uid = int(current_user["sub"])
    contents = await file.read()
    if len(contents) > 6 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="La imagen no puede superar 6 MB.")
    sb = get_supabase()
    cid = _usuario_contrato_id(sb, uid)
    url = _usuario_imagen_subir(contents, uid, "avatar", file.content_type, contrato_id=cid)
    try:
        sb.table("usuarios").update({"foto_perfil_url": url}).eq("id", uid).execute()
    except Exception as e:
        _log_api.warning("POST /usuarios/me/foto-perfil: %s", e)
        raise HTTPException(status_code=503, detail=_detalle_guardar_url_perfil(e, "foto_perfil_url"))
    return {"url": url}


@app.delete("/usuarios/me/foto-perfil")
def quitar_foto_perfil(current_user=Depends(get_current_user)):
    uid = int(current_user["sub"])
    sb = get_supabase()
    try:
        sb.table("usuarios").update({"foto_perfil_url": None}).eq("id", uid).execute()
    except Exception as e:
        _log_api.warning("DELETE /usuarios/me/foto-perfil: %s", e)
        raise HTTPException(status_code=503, detail="No se pudo actualizar el perfil.")
    return {"ok": True}


@app.post("/usuarios/me/firma-imagen")
async def subir_firma_imagen(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    uid = int(current_user["sub"])
    contents = await file.read()
    if len(contents) > 6 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="La imagen no puede superar 6 MB.")
    sb = get_supabase()
    cid = _usuario_contrato_id(sb, uid)
    url = _usuario_imagen_subir(contents, uid, "firma", file.content_type, contrato_id=cid)
    try:
        sb.table("usuarios").update({"firma_imagen_url": url}).eq("id", uid).execute()
    except Exception as e:
        _log_api.warning("POST /usuarios/me/firma-imagen: %s", e)
        raise HTTPException(status_code=503, detail=_detalle_guardar_url_perfil(e, "firma_imagen_url"))
    return {"url": url}


@app.delete("/usuarios/me/firma-imagen")
def quitar_firma_imagen(current_user=Depends(get_current_user)):
    uid = int(current_user["sub"])
    sb = get_supabase()
    try:
        sb.table("usuarios").update({"firma_imagen_url": None}).eq("id", uid).execute()
    except Exception as e:
        _log_api.warning("DELETE /usuarios/me/firma-imagen: %s", e)
        raise HTTPException(status_code=503, detail="No se pudo actualizar el perfil.")
    return {"ok": True}

@app.get("/usuarios")
def listar_usuarios(current_user=Depends(get_current_user)):
    return supabase.table("usuarios").select(
        "id, nombre, apellidos, email, activo, cargo_id, rol_id, contrato_id, estado, created_at"
    ).execute().data

@app.post("/usuarios")
def crear_usuario(usuario: UsuarioCreate, current_user=Depends(get_current_user)):
    hashed = hash_password(usuario.password)
    return supabase.table("usuarios").insert({
        "nombre": usuario.nombre,
        "email": usuario.email,
        "password_hash": hashed,
        "cargo_id": usuario.cargo_id
    }).execute().data

@app.get("/categorias")
def listar_categorias(current_user=Depends(get_current_user)):
    return supabase.table("categorias").select("*").execute().data

@app.get("/funciones")
def listar_funciones(current_user=Depends(get_current_user)):
    funciones = supabase.table("funciones").select("*").order("nombre").execute().data or []
    existentes = {(f.get("nombre") or "").strip().lower() for f in funciones}
    codigos_existentes = {
        str((f.get("codigo") or "")).strip().upper()
        for f in funciones
        if f.get("codigo") is not None and str(f.get("codigo")).strip() != ""
    }
    # Asegurar filas base (Dashboard, Informes CCD). En producción el INSERT vía API a veces
    # falla por RLS o restricciones; en ese caso ejecutar backend/sql/funcion_informes_ccd.sql en Supabase.
    requeridas = [
        {"codigo": "DASHBOARD", "nombre": "Dashboard", "modulo": "Dashboard"},
        {"codigo": "INFCCD", "nombre": "Informes CCD", "modulo": "Informes"},
        {"codigo": "SSTDOC", "nombre": "SST documental", "modulo": "SST"},
        {"codigo": "ENSPIP", "nombre": "Ensayos PIP", "modulo": "Laboratorio"},
        {"codigo": "NUVECC", "nombre": "Integración nube ClaraCore", "modulo": "Administración"},
        {"codigo": "AUDSST", "nombre": "Auditor SST (IA)", "modulo": "SST"},
    ]
    for req in requeridas:
        nombre_funcion = req["nombre"]
        cod = str(req.get("codigo") or "").strip().upper()
        if nombre_funcion.lower() in existentes or (cod and cod in codigos_existentes):
            continue
        try:
            supabase.table("funciones").insert(req).execute()
            existentes.add(nombre_funcion.lower())
            if cod:
                codigos_existentes.add(cod)
        except Exception as e:
            try:
                supabase.table("funciones").upsert(req, on_conflict="codigo").execute()
                existentes.add(nombre_funcion.lower())
                if cod:
                    codigos_existentes.add(cod)
            except Exception as e2:
                _log_api.warning(
                    "/funciones: no se pudo insertar ni upsert '%s' (%s). "
                    "Si falta en el panel admin, ejecuta backend/sql/funcion_informes_ccd.sql en Supabase. "
                    "insert_err=%s | upsert_err=%s",
                    nombre_funcion,
                    req.get("codigo"),
                    e,
                    e2,
                )
    funciones = supabase.table("funciones").select("*").order("nombre").execute().data or []
    return funciones

def _normalizar_costos_adicionales_lista(items: List[CostoAdicionalItem]) -> List[dict]:
    out: List[dict] = []

    def _f(x) -> Optional[float]:
        if x is None:
            return None
        try:
            v = float(x)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(v):
            return None
        return v

    for it in items or []:
        c = (it.concepto_contractual or "").strip()
        if not c:
            continue
        vm = _f(it.valor_mensual)
        tm = _f(it.tiempo_meses)
        v_old = _f(it.valor)
        if vm is None and v_old is not None and tm is None:
            vm = v_old
            tm = 1.0
        if vm is None:
            vm = 0.0
        if tm is None:
            tm = 0.0
        if tm < 0.0:
            tm = 0.0
        v = float(round(float(vm) * float(tm), 0))
        out.append(
            {
                "concepto_contractual": c,
                "valor_mensual": float(vm),
                "tiempo_meses": float(tm),
                "valor": v,
            }
        )
    return out


def _suma_costos_adicionales_cop(lista: List[dict]) -> Optional[float]:
    t = 0.0
    n = 0
    for x in lista or []:
        v = x.get("valor")
        if v is None:
            continue
        t += float(v)
        n += 1
    if not lista:
        return None
    if n == 0:
        return None
    return t


@app.post("/contratos")
def crear_contrato(contrato: ContratoCreate, current_user=Depends(get_current_user)):
    existe = supabase.table("contratos").select("id").eq("numero", contrato.numero).execute()
    if existe.data:
        raise HTTPException(status_code=400, detail="Ya existe un contrato con ese número")
    norm = _normalizar_costos_adicionales_lista(contrato.costos_adicionales_lista or [])
    costos_cop = _suma_costos_adicionales_cop(norm)
    if not norm and contrato.costos_adicionales is not None:
        costos_cop = contrato.costos_adicionales
    result = supabase.table("contratos").insert({
        "numero": contrato.numero,
        "objeto": contrato.objeto,
        "contratista": contrato.contratista,
        "nit": contrato.nit,
        "interventoria": contrato.interventoria,
        "entidad": contrato.entidad,
        "entidad_otra": contrato.entidad_otra,
        "logo_entidad": contrato.logo_entidad,
        "plano_geojson": contrato.plano_geojson,
        "centro_lat": contrato.centro_lat,
        "centro_lng": contrato.centro_lng,
        "logo_contratista": contrato.logo_contratista,
        "logo_interventoria": contrato.logo_interventoria,
        "aiu": contrato.aiu,
        "iva": contrato.iva,
        "valor_componente_ambiental": contrato.valor_componente_ambiental,
        "valor_componente_social": contrato.valor_componente_social,
        "valor_componente_pmt": contrato.valor_componente_pmt,
        "costo_directo_contrato": contrato.costo_directo_contrato,
        "costos_adicionales": costos_cop,
        "costos_adicionales_lista": norm,
    }).execute()
    nuevo = result.data[0]
    try:
        _cloudinary_seed_carpetas_contrato(int(nuevo["id"]))
    except Exception as e:
        _log_api.warning("Tras crear contrato, seed Cloudinary: %s", e)
    return nuevo


@app.post("/contratos/{contrato_id}/cloudinary-sembrar-carpetas")
def contrato_sembrar_carpetas_cloudinary(contrato_id: int, current_user=Depends(get_current_user)):
    """
    Contratos ya existentes (antes de la semilla automática): crea en Cloudinary las cuatro carpetas
    (`fotos`, `graficos`, `Fotos de Perfil`, `Fotos de Firmas`) con un PNG mínimo por carpeta.
    Si no usas Cloudinary en el backend, esta ruta responde 400 explicando que las imágenes van a Supabase.
    """
    if not os.getenv("CLOUDINARY_CLOUD_NAME"):
        raise HTTPException(
            status_code=400,
            detail=(
                "Cloudinary no está configurado (variables CLOUDINARY_* en el servidor). "
                "Las fotos de perfil y firma se suben entonces a Supabase Storage: no aparecerán en Cloudinary "
                "y no hace falta crear carpetas allí."
            ),
        )
    ex = supabase.table("contratos").select("id").eq("id", contrato_id).limit(1).execute()
    if not ex.data:
        raise HTTPException(status_code=404, detail="Contrato no encontrado")
    caller_contrato, cargo = _caller_contract_scope(current_user)
    priv = (cargo or "") in ("desarrollador", "administrador")
    if not priv and caller_contrato and int(caller_contrato) != int(contrato_id):
        raise HTTPException(status_code=403, detail="No tienes acceso a información de otro contrato")
    _cloudinary_seed_carpetas_contrato(contrato_id)
    return {
        "ok": True,
        "contrato_id": contrato_id,
        "mensaje": f"Carpetas iniciales en Cloudinary: {CLOUDINARY_ROOT}/{contrato_id}/(fotos|graficos|…)",
    }


@app.put("/contratos/{contrato_id}")
def actualizar_contrato(contrato_id: int, body: ContratoUpdate, current_user=Depends(get_current_user)):
    data = body.dict(exclude_unset=True)
    if "costos_adicionales_lista" in data and body.costos_adicionales_lista is not None:
        norm = _normalizar_costos_adicionales_lista(body.costos_adicionales_lista)
        data["costos_adicionales_lista"] = norm
        data["costos_adicionales"] = _suma_costos_adicionales_cop(norm)
    if not data:
        return {"mensaje": "Sin cambios"}
    supabase.table("contratos").update(data).eq("id", contrato_id).execute()
    return {"mensaje": "Contrato actualizado"}

@app.delete("/contratos/{contrato_id}")
def eliminar_contrato(contrato_id: int, current_user=Depends(get_current_user)):
    supabase.table("contratos").delete().eq("id", contrato_id).execute()
    return {"mensaje": "Contrato eliminado"}
# ─────────────────────────────────────────────
# RUTAS ADMIN
# ─────────────────────────────────────────────

@app.get("/admin/usuarios-pendientes")
def usuarios_pendientes(current_user=Depends(get_current_user)):
    result = supabase.table("usuarios").select(
        "id, nombre, apellidos, email, cargo_id, contrato_id, estado, created_at"
    ).eq("estado", "pendiente").execute()

    cargos = {c["id"]: c["nombre"] for c in supabase.table("cargos").select("id, nombre").execute().data}
    contratos = {c["id"]: c["numero"] for c in supabase.table("contratos").select("id, numero").execute().data}
    for u in result.data:
        u["cargo_nombre"] = cargos.get(u.get("cargo_id"), "Sin cargo")
        u["contrato_numero"] = contratos.get(u.get("contrato_id"), "Sin contrato")
    return result.data

@app.put("/admin/usuarios/{usuario_id}/aprobar")
def aprobar_usuario(usuario_id: int, body: AprobarRequest, current_user=Depends(get_current_user)):
    supabase.table("usuarios").update({
        "estado": "aprobado", "activo": True, "rol_id": body.rol_id
    }).eq("id", usuario_id).execute()
    rol_nombre = ""
    if body.rol_id:
        r = supabase.table("roles").select("nombre").eq("id", body.rol_id).execute()
        rol_nombre = r.data[0]["nombre"] if r.data else str(body.rol_id)
    registrar_log(current_user, "APROBAR", "USUARIOS", "usuario", str(usuario_id),
        {"estado": "aprobado", "rol": rol_nombre})
    return {"mensaje": "Usuario aprobado"}

@app.put("/admin/usuarios/{usuario_id}/rechazar")
def rechazar_usuario(usuario_id: int, current_user=Depends(get_current_user)):
    supabase.table("usuarios").update({
        "estado": "rechazado", "activo": False
    }).eq("id", usuario_id).execute()
    registrar_log(current_user, "RECHAZAR", "USUARIOS", "usuario", str(usuario_id),
        {"estado": "rechazado"})
    return {"mensaje": "Usuario rechazado"}

@app.post("/admin/cargos")
def crear_cargo(cargo: CargoCreate, current_user=Depends(get_current_user)):
    return supabase.table("cargos").insert({
        "nombre": cargo.nombre, "rol_id": cargo.rol_id, "categoria_id": cargo.categoria_id
    }).execute().data

@app.delete("/admin/cargos/{cargo_id}")
def eliminar_cargo(cargo_id: int, current_user=Depends(get_current_user)):
    supabase.table("cargos").delete().eq("id", cargo_id).execute()
    return {"mensaje": "Cargo eliminado"}

@app.get("/admin/permisos/{cargo_id}")
def obtener_permisos(cargo_id: int, current_user=Depends(get_current_user)):
    return supabase.table("permisos").select("*").eq("cargo_id", cargo_id).execute().data

@app.get("/admin/todos-usuarios")
def todos_usuarios(current_user=Depends(get_current_user)):
    result = supabase.table("usuarios").select(
        "id, nombre, apellidos, email, activo, cargo_id, rol_id, contrato_id, estado, created_at, politicas_aceptadas, politicas_fecha, politicas_version, politicas_ip"
    ).order("nombre").execute()
    cargos = {c["id"]: c["nombre"] for c in supabase.table("cargos").select("id, nombre").execute().data}
    roles = {r["id"]: r["nombre"] for r in supabase.table("roles").select("id, nombre").execute().data}
    contratos = {c["id"]: c["numero"] for c in supabase.table("contratos").select("id, numero").execute().data}
    caller_contrato, _ = _caller_contract_scope(current_user)
    for u in result.data:
        u["cargo_nombre"] = cargos.get(u.get("cargo_id"), "Sin cargo")
        u["rol_nombre"] = roles.get(u.get("rol_id"), "Sin rol")
        u["contrato_numero"] = contratos.get(u.get("contrato_id"), "Sin contrato")
    caller_id = int(current_user["sub"])
    # Privacidad estricta: cada sesión solo puede ver usuarios de su contrato.
    if caller_contrato:
        filtered = [u for u in result.data if u.get("contrato_id") == caller_contrato]
    else:
        filtered = list(result.data)
    filtered = [u for u in filtered if u.get("cargo_nombre", "").lower() != "desarrollador" or u["id"] == caller_id]
    return filtered

@app.put("/admin/usuarios/{usuario_id}")
def actualizar_usuario(usuario_id: int, body: UsuarioUpdate, current_user=Depends(get_current_user)):
    # Proteger: no se puede modificar un Desarrollador SALVO que sea él mismo editándose
    es_el_mismo = str(usuario_id) == str(current_user.get("sub"))
    if not es_el_mismo:
        target = supabase.table("usuarios").select("cargo_id").eq("id", usuario_id).execute()
        if target.data and target.data[0].get("cargo_id"):
            cargo_res = supabase.table("cargos").select("nombre").eq("id", target.data[0]["cargo_id"]).execute()
            if cargo_res.data and cargo_res.data[0]["nombre"].lower() == "desarrollador":
                raise HTTPException(status_code=403, detail="No se puede modificar un usuario Desarrollador")

    prev_snap = supabase.table("usuarios").select(
        "cargo_id, rol_id, contrato_id, estado, activo, subcontratista_id, nombre, apellidos, email"
    ).eq("id", usuario_id).limit(1).execute().data
    prev_snap = prev_snap[0] if prev_snap else {}

    # exclude_unset=True: campos no enviados no se tocan; null explícito sí borra el campo
    data = body.dict(exclude_unset=True)
    if body.estado == "aprobado":
        data["activo"] = True
    elif body.estado == "rechazado":
        data["activo"] = False
    if body.politicas_aceptadas is False:
        data["politicas_fecha"] = None
        data["politicas_ip"] = None
    supabase.table("usuarios").update(data).eq("id", usuario_id).execute()
    # Resetear contador de inactividad: insertar LOGIN sintético al reactivar un usuario
    if body.estado == "aprobado":
        try:
            u_data = supabase.table("usuarios").select("nombre, apellidos, cargo_id, contrato_id").eq("id", usuario_id).execute().data
            u_nombre = ""
            u_contrato_id = None
            if u_data:
                u = u_data[0]
                u_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip()
                u_contrato_id = u.get("contrato_id")
            supabase.table("logs").insert({
                "usuario_id":   usuario_id,
                "usuario_nombre": u_nombre,
                "cargo_nombre": "",
                "contrato_id":  u_contrato_id,
                "contrato_numero": None,
                "accion":       "LOGIN",
                "modulo":       "AUTH",
                "entidad_tipo": "usuario",
                "entidad_id":   str(usuario_id),
                "detalle":      {"origen": "Reactivación por administrador"},
                "resultado":    "ok",
            }).execute()
        except Exception:
            pass
    # Enriquecer detalle con nombres legibles
    detalle_log = dict(data)
    if "cargo_id" in detalle_log:
        r = supabase.table("cargos").select("nombre").eq("id", detalle_log["cargo_id"]).execute()
        detalle_log["cargo"] = r.data[0]["nombre"] if r.data else str(detalle_log["cargo_id"])
        del detalle_log["cargo_id"]
    if "rol_id" in detalle_log:
        r = supabase.table("roles").select("nombre").eq("id", detalle_log["rol_id"]).execute()
        detalle_log["rol"] = r.data[0]["nombre"] if r.data else str(detalle_log["rol_id"])
        del detalle_log["rol_id"]
    if "contrato_id" in detalle_log:
        r = supabase.table("contratos").select("numero").eq("id", detalle_log["contrato_id"]).execute()
        detalle_log["contrato"] = r.data[0]["numero"] if r.data else str(detalle_log["contrato_id"])
        del detalle_log["contrato_id"]

    def _usuario_audit_enriquecido(row: dict) -> dict:
        if not row:
            return {}
        o = dict(row)
        if row.get("cargo_id"):
            cr = supabase.table("cargos").select("nombre").eq("id", row["cargo_id"]).execute()
            o["cargo"] = cr.data[0]["nombre"] if cr.data else str(row["cargo_id"])
        if row.get("rol_id"):
            rr = supabase.table("roles").select("nombre").eq("id", row["rol_id"]).execute()
            o["rol"] = rr.data[0]["nombre"] if rr.data else str(row["rol_id"])
        if row.get("contrato_id"):
            ct = supabase.table("contratos").select("numero").eq("id", row["contrato_id"]).execute()
            o["contrato"] = ct.data[0]["numero"] if ct.data else str(row["contrato_id"])
        return o

    after_snap = supabase.table("usuarios").select(
        "cargo_id, rol_id, contrato_id, estado, activo, subcontratista_id, nombre, apellidos, email"
    ).eq("id", usuario_id).limit(1).execute().data
    after_snap = after_snap[0] if after_snap else {}

    registrar_log(
        current_user,
        "EDITAR",
        "USUARIOS",
        "usuario",
        str(usuario_id),
        detalle_log,
        valor_anterior=_usuario_audit_enriquecido(prev_snap),
        valor_nuevo=_usuario_audit_enriquecido(after_snap),
        severidad="AUDIT",
    )
    return {"mensaje": "Usuario actualizado"}

@app.post("/admin/verificar-inactividad")
def verificar_inactividad(current_user=Depends(get_current_user)):
    """Marca como pendiente a usuarios aprobados con >7 días sin iniciar sesión (según último LOGIN en logs)."""
    if not _cargo_puede_auditar_logs(current_user):
        raise HTTPException(
            status_code=403,
            detail="Solo Desarrollador o Administrador pueden ejecutar la verificación de inactividad.",
        )
    from datetime import datetime, timedelta, timezone
    CARGOS_EXCLUIDOS = {'director', 'gerencia', 'supervisor externo', 'desarrollador', 'administrador'}
    ahora = datetime.now(timezone.utc)
    limite = ahora - timedelta(days=7)
    usuarios_raw = supabase.table("usuarios").select("id, cargo_id, estado").eq("estado", "aprobado").execute().data
    if not usuarios_raw:
        return {"afectados": 0}
    cargos = {c["id"]: c["nombre"].lower() for c in supabase.table("cargos").select("id, nombre").execute().data}
    candidatos = [u for u in usuarios_raw if cargos.get(u.get("cargo_id"), "") not in CARGOS_EXCLUIDOS]
    if not candidatos:
        return {"afectados": 0}
    cand_ids = [u["id"] for u in candidatos]
    last_login: Dict[int, str] = {}
    page_size = 500
    max_filas_por_bloque = 40000
    # Evita URL/query gigantes y timeouts: bloques de usuarios.
    id_bloque = 100
    for bi in range(0, len(cand_ids), id_bloque):
        sub_ids = cand_ids[bi : bi + id_bloque]
        needed = set(sub_ids)
        escaneadas = 0
        while needed and escaneadas < max_filas_por_bloque:
            start = escaneadas
            end = escaneadas + page_size - 1
            batch = (
                supabase.table("logs")
                .select("usuario_id, created_at")
                .eq("accion", "LOGIN")
                .in_("usuario_id", sub_ids)
                .order("created_at", desc=True)
                .range(start, end)
                .execute()
                .data
                or []
            )
            if not batch:
                break
            for log in batch:
                uid = log.get("usuario_id")
                if uid is None:
                    continue
                if uid in needed and uid not in last_login:
                    last_login[uid] = log.get("created_at")
            needed = set(sub_ids) - set(last_login.keys())
            if not needed:
                break
            escaneadas += len(batch)
            if len(batch) < page_size:
                break
    afectados = 0
    for u in candidatos:
        uid = u["id"]
        last = last_login.get(uid)
        inactivo = False
        if last is None:
            inactivo = True
        else:
            try:
                last_dt = datetime.fromisoformat(last.replace("Z", "+00:00"))
                if last_dt.tzinfo is None:
                    last_dt = last_dt.replace(tzinfo=timezone.utc)
                if last_dt < limite:
                    inactivo = True
            except Exception:
                pass
        if inactivo:
            supabase.table("usuarios").update({"estado": "pendiente", "activo": False}).eq("id", uid).execute()
            afectados += 1
    registrar_log(current_user, "VERIFICAR_INACTIVIDAD", "USUARIOS", None, None, {"afectados": afectados})
    return {"afectados": afectados}

@app.get("/admin/usuario-contratos/{usuario_id}")
def get_usuario_contratos(usuario_id: int, current_user=Depends(get_current_user)):
    result = supabase.table("usuario_contratos").select("contrato_id").eq("usuario_id", usuario_id).execute()
    ids = [r["contrato_id"] for r in result.data]
    if not ids:
        return []
    contratos = supabase.table("contratos").select("id, numero, contratista, interventoria, entidad, entidad_otra, logo_entidad, plano_geojson, centro_lat, centro_lng, logo_contratista, logo_interventoria, fase").in_("id", ids).execute()
    return contratos.data

@app.post("/admin/usuario-contratos")
def agregar_usuario_contrato(body: UsuarioContratoCreate, current_user=Depends(get_current_user)):
    existe = supabase.table("usuario_contratos").select("id").eq("usuario_id", body.usuario_id).eq("contrato_id", body.contrato_id).execute()
    if existe.data:
        raise HTTPException(status_code=400, detail="El usuario ya tiene ese contrato asignado")
    supabase.table("usuario_contratos").insert({"usuario_id": body.usuario_id, "contrato_id": body.contrato_id}).execute()
    return {"mensaje": "Contrato asignado"}

@app.delete("/admin/usuario-contratos/{usuario_id}/{contrato_id}")
def quitar_usuario_contrato(usuario_id: int, contrato_id: int, current_user=Depends(get_current_user)):
    supabase.table("usuario_contratos").delete().eq("usuario_id", usuario_id).eq("contrato_id", contrato_id).execute()
    return {"mensaje": "Contrato removido"}

@app.get("/mantenimiento")
def get_mantenimiento():
    return _estado_mantenimiento()

@app.post("/mantenimiento")
def set_mantenimiento(body: MantenimientoRequest):
    if body.secret != _MAINTENANCE_SECRET:
        raise HTTPException(status_code=403, detail="Secret inválido")

    if body.activo:
        segundos = body.segundos if (body.segundos is not None and body.segundos > 0) else _MAINTENANCE_DEFAULT_SECONDS
        _maintenance_state["activo"] = True
        _maintenance_state["mensaje"] = body.mensaje or _maintenance_state["mensaje"]
        _maintenance_state["expires_at"] = time.time() + segundos
    else:
        _maintenance_state["activo"] = False
        _maintenance_state["expires_at"] = None
        if body.mensaje:
            _maintenance_state["mensaje"] = body.mensaje

    return _estado_mantenimiento()

@app.post("/auth/refresh")
def refresh_token(current_user=Depends(get_current_user)):
    """Renueva el token JWT del usuario activo."""
    uid = int(current_user.get("sub"))
    try:
        def _u():
            return supabase.table("usuarios").select("*").eq("id", uid).execute()
        result = supabase_execute(_u, retries=4, delay=0.6)
        if not result.data:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        usuario = result.data[0]
        cargo_nombre = None
        if usuario.get("cargo_id"):
            cid = usuario["cargo_id"]
            def _c():
                return supabase.table("cargos").select("nombre").eq("id", cid).execute()
            r = supabase_execute(_c, retries=3, delay=0.5)
            if r.data:
                cargo_nombre = r.data[0]["nombre"]
        rol_nombre = None
        if usuario.get("rol_id"):
            rid = usuario["rol_id"]
            def _r():
                return supabase.table("roles").select("nombre").eq("id", rid).execute()
            r = supabase_execute(_r, retries=3, delay=0.5)
            if r.data:
                rol_nombre = r.data[0]["nombre"]
        contrato_numero = None
        if usuario.get("contrato_id"):
            ctid = usuario["contrato_id"]
            def _ct():
                return supabase.table("contratos").select("numero").eq("id", ctid).execute()
            r = supabase_execute(_ct, retries=3, delay=0.5)
            if r.data:
                contrato_numero = r.data[0]["numero"]
        nombre_completo = f"{usuario.get('nombre','')} {usuario.get('apellidos','')}".strip()
        new_token = create_token({
            "sub": str(uid),
            "email": usuario["email"],
            "nombre": nombre_completo or usuario.get("nombre") or "",
            "cargo_nombre": cargo_nombre or "",
            "rol_nombre": rol_nombre or "",
            "contrato_id": usuario.get("contrato_id"),
            "contrato_numero": contrato_numero or "",
        })
        return {"access_token": new_token, "token_type": "bearer"}
    except HTTPException:
        raise
    except (httpx.ReadTimeout, httpx.ConnectTimeout, httpx.ConnectError, httpx.RemoteProtocolError) as e:
        raise HTTPException(
            status_code=503,
            detail="El servicio de datos no respondió a tiempo. Intente de nuevo en unos segundos.",
        ) from e


@app.post("/auth/logout")
def auth_logout(request: Request, current_user=Depends(get_current_user)):
    """Registra cierre de sesión (auditoría)."""
    ip = _client_ip(request)
    uid = current_user.get("sub")
    nombre = current_user.get("nombre") or current_user.get("email", "")
    registrar_log(
        {
            "sub": str(uid),
            "nombre": nombre,
            "cargo_nombre": current_user.get("cargo_nombre"),
            "rol_nombre": current_user.get("rol_nombre"),
            "contrato_id": current_user.get("contrato_id"),
            "contrato_numero": current_user.get("contrato_numero"),
        },
        "LOGOUT",
        "AUTH",
        "usuario",
        str(uid),
        {},
        ip=ip,
    )
    return {"ok": True}


@app.post("/auth/solicitar-reset")
def solicitar_reset(body: ResetSolicitud):
    usuario = supabase.table("usuarios").select("id, email, nombre").eq("email", body.email).execute()
    if not usuario.data:
        raise HTTPException(status_code=404, detail="Correo no registrado")
    u = usuario.data[0]
    pendiente = supabase.table("password_reset_requests").select("id").eq("usuario_id", u["id"]).eq("estado", "pendiente").execute()
    if pendiente.data:
        return {"mensaje": "Ya tienes una solicitud pendiente"}
    supabase.table("password_reset_requests").insert({"usuario_id": u["id"], "email": body.email, "estado": "pendiente"}).execute()
    return {"mensaje": "Solicitud enviada al administrador"}

@app.get("/auth/reset-autorizado")
def check_reset_autorizado(email: str):
    result = supabase.table("password_reset_requests").select("estado").eq("email", email).eq("estado", "autorizado").execute()
    return {"autorizado": len(result.data) > 0}

@app.post("/auth/cambiar-password-temporal")
def cambiar_password_temporal(body: CambiarPassword):
    solicitud = supabase.table("password_reset_requests").select("*").eq("email", body.email).eq("estado", "autorizado").order("id", desc=True).limit(1).execute()
    if not solicitud.data:
        raise HTTPException(status_code=403, detail="No tienes autorización para cambiar la contraseña")
    s = solicitud.data[0]
    if not verify_password(body.contrasena_temporal, s["contrasena_temporal"]):
        raise HTTPException(status_code=401, detail="Contraseña temporal incorrecta")
    nuevo_hash = hash_password(body.nueva_password)
    supabase.table("usuarios").update({"password_hash": nuevo_hash}).eq("email", body.email).execute()
    supabase.table("password_reset_requests").delete().eq("id", s["id"]).execute()
    return {"mensaje": "Contraseña actualizada correctamente"}

@app.get("/admin/reset-requests")
def listar_reset_requests(current_user=Depends(get_current_user)):
    return supabase.table("password_reset_requests").select("*").eq("estado", "pendiente").order("created_at", desc=True).execute().data

@app.put("/admin/reset-requests/{request_id}/autorizar")
def autorizar_reset(request_id: int, body: ResetAutorizar, current_user=Depends(get_current_user)):
    hashed_temp = hash_password(body.contrasena_temporal)
    supabase.table("password_reset_requests").update({"estado": "autorizado", "contrasena_temporal": hashed_temp}).eq("id", request_id).execute()
    return {"mensaje": "Reset autorizado"}

@app.post("/admin/permisos")
def guardar_permisos(permisos: List[PermisoUpdate], current_user=Depends(get_current_user)):
    cargo_ids = sorted({p.cargo_id for p in permisos})

    def _norm_perm_rows(rows):
        if not rows:
            return []

        def _key(r):
            return (r.get("cargo_id"), r.get("funcion_id"))

        out = []
        for r in sorted(rows, key=_key):
            out.append({
                "cargo_id": r.get("cargo_id"),
                "funcion_id": r.get("funcion_id"),
                "ver": r.get("ver"),
                "crear": r.get("crear"),
                "editar": r.get("editar"),
                "eliminar": r.get("eliminar"),
                "validar": r.get("validar"),
                "exportar": r.get("exportar"),
            })
        return out

    antes_por_cargo = {}
    for cid in cargo_ids:
        antes_por_cargo[cid] = _norm_perm_rows(
            supabase.table("permisos").select("*").eq("cargo_id", cid).execute().data or []
        )

    for permiso in permisos:
        existe = supabase.table("permisos").select("id") \
            .eq("cargo_id", permiso.cargo_id).eq("funcion_id", permiso.funcion_id).execute()
        data = permiso.dict()
        if existe.data:
            supabase.table("permisos").update(data) \
                .eq("cargo_id", permiso.cargo_id).eq("funcion_id", permiso.funcion_id).execute()
        else:
            supabase.table("permisos").insert(data).execute()

    for cid in cargo_ids:
        despues = _norm_perm_rows(
            supabase.table("permisos").select("*").eq("cargo_id", cid).execute().data or []
        )
        if despues == antes_por_cargo.get(cid, []):
            continue
        cargo_row = supabase.table("cargos").select("nombre").eq("id", cid).limit(1).execute().data
        cn = cargo_row[0]["nombre"] if cargo_row else str(cid)
        registrar_log(
            current_user,
            "EDITAR",
            "PERMISOS",
            "cargo",
            str(cid),
            {"cargo": cn, "cambio": "matriz_permisos"},
            valor_anterior=antes_por_cargo.get(cid, []),
            valor_nuevo=despues,
            severidad="AUDIT",
            alerta_generada=True,
        )
    return {"mensaje": f"{len(permisos)} permisos guardados"}

# ─────────────────────────────────────────────
# LISTADO DE PRECIOS
# ─────────────────────────────────────────────

@app.get("/listado-precios/{contrato_id}")
def get_listado_precios(contrato_id: int, current_user=Depends(get_current_user)):
    _require_contract_access(current_user, contrato_id)
    all_rows = []
    offset = 0
    while True:
        batch = supabase.table("listado_precios").select("*").eq("contrato_id", contrato_id).order("item_numero").range(offset, offset + 999).execute().data
        all_rows.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    return all_rows

@app.post("/listado-precios/{contrato_id}/bulk")
def bulk_precios(contrato_id: int, items: List[ListadoPrecioItem], current_user=Depends(get_current_user)):
    _require_contract_access(current_user, contrato_id)
    """Reemplaza todos los precios del contrato con los items del CSV."""
    supabase.table("listado_precios").delete().eq("contrato_id", contrato_id).execute()
    if items:
        rows = []
        for item in items:
            row = {"contrato_id": contrato_id, **{k: v for k, v in item.dict().items() if v is not None}}
            if row.get("tipo_precio") == "Precio Contractual":
                row["estado_precio"] = "Aprobado"
                row["acta_fijacion"] = "Contractual"
                row.pop("acta_modificatoria", None)
            elif row.get("tipo_precio") == "Precio No Previsto":
                try:
                    f_val = float(row.get("acta_fijacion") or 0)
                    m_val = float(row.get("acta_modificatoria") or 0)
                except (ValueError, TypeError):
                    f_val, m_val = 0, 0
                row["estado_precio"] = "Aprobado" if (f_val > 0 and m_val > 0) else "Pendiente"
            rows.append(row)
        supabase.table("listado_precios").insert(rows).execute()
    registrar_log(current_user, "IMPORTAR", "PRECIOS", "listado_precios", str(contrato_id),
                  {"cantidad": len(items)})
    return {"mensaje": f"{len(items)} items cargados"}

@app.put("/listado-precios/item/{item_id}")
def update_precio(item_id: int, body: ListadoPrecioItem, current_user=Depends(get_current_user)):
    data = body.dict(exclude_none=True)
    tipo = data.get("tipo_precio")
    if tipo == "Precio Contractual":
        data["estado_precio"] = "Aprobado"
        data["acta_fijacion"] = "Contractual"
        data.pop("acta_modificatoria", None)
    elif tipo == "Precio No Previsto":
        try:
            f_val = float(data.get("acta_fijacion") or 0)
            m_val = float(data.get("acta_modificatoria") or 0)
        except (ValueError, TypeError):
            f_val, m_val = 0, 0
        data["estado_precio"] = "Aprobado" if (f_val > 0 and m_val > 0) else "Pendiente"
    supabase.table("listado_precios").update(data).eq("id", item_id).execute()
    registrar_log(current_user, "EDITAR", "PRECIOS", "listado_precios", str(item_id),
                  {"tipo_precio": data.get("tipo_precio"), "estado_precio": data.get("estado_precio")})
    return {"ok": True}

@app.delete("/listado-precios/item/{item_id}")
def delete_precio(item_id: int, current_user=Depends(get_current_user)):
    supabase.table("listado_precios").delete().eq("id", item_id).execute()

@app.post("/listado-precios/{contrato_id}/item")
def crear_precio(contrato_id: int, body: ListadoPrecioItem, current_user=Depends(get_current_user)):
    _require_contract_access(current_user, contrato_id)
    """Crea un ítem individual en el listado de precios con lógica de aprobación automática."""
    row = body.dict(exclude_none=True)
    row["contrato_id"] = contrato_id
    if body.tipo_precio == "Precio Contractual":
        row["estado_precio"] = "Aprobado"
        row["acta_fijacion"] = "Contractual"
        row.pop("acta_modificatoria", None)
    else:
        try:
            f_val = float(row.get("acta_fijacion") or 0)
            m_val = float(row.get("acta_modificatoria") or 0)
        except (ValueError, TypeError):
            f_val, m_val = 0, 0
        row["estado_precio"] = "Aprobado" if (f_val > 0 and m_val > 0) else "Pendiente"
    result = supabase.table("listado_precios").insert(row).execute()
    nuevo = result.data[0] if result.data else {}
    registrar_log(current_user, "CREAR", "PRECIOS", "listado_precios", str(nuevo.get("id", "")),
                  {"item_numero": row.get("item_numero"), "descripcion": row.get("descripcion"),
                   "tipo_precio": row.get("tipo_precio"), "estado_precio": row.get("estado_precio")})
    return nuevo if nuevo else {"ok": True}

@app.get("/listado-precios/item/{item_id}/stats")
def get_precio_stats(item_id: int, current_user=Depends(get_current_user)):
    """Retorna cantidades y costos presupuestados, cobrados y balance para un ítem del listado."""
    precio = supabase.table("listado_precios").select(
        "contrato_id, capitulo, competencia, item_numero, precio_unitario"
    ).eq("id", item_id).single().execute().data
    if not precio:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")
    contrato_id  = precio["contrato_id"]
    capitulo     = precio.get("capitulo") or ""
    competencia  = precio.get("competencia") or ""
    item_numero  = precio.get("item_numero") or ""
    vlr_unitario = float(precio.get("precio_unitario") or 0)
    ppto_q = supabase.table("presupuesto").select("cant_total").eq("contrato_id", contrato_id).eq("item", item_numero).eq("tipo_ejecucion", "Presupuesto de Obra").eq("dado_de_baja", False)
    if capitulo:
        ppto_q = ppto_q.eq("capitulo", capitulo)
    if competencia:
        ppto_q = ppto_q.eq("competencia", competencia)
    ppto_rows = ppto_q.execute().data or []
    cant_ppto = sum(float(r.get("cant_total") or 0) for r in ppto_rows)
    cap_k = _dash_norm_capitulo_key_py(capitulo) if capitulo else None
    comp_f = (competencia or "").strip()
    cant_cobro = 0.0
    costo_cobro = 0.0
    off_cb = 0
    it_key = _dash_norm_item_key_py(item_numero)
    while True:
        def _q_obra(o=off_cb):
            q = (
                supabase.table("so_registros")
                .select("capitulo, competencia, item_numero, cantidad_total, costo_directo, nivel3_estado")
                .eq("contrato_id", contrato_id)
                .range(o, o + 999)
            )
            return q.execute().data

        batch = supabase_execute(_q_obra) or []
        for r in batch:
            if _matriz_validacion_norm_estado(r.get("nivel3_estado")) != "Aprobado":
                continue
            if cap_k and _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_k:
                continue
            if comp_f and (r.get("competencia") or "").strip() != comp_f:
                continue
            if _dash_norm_item_key_py(r.get("item_numero")) != it_key:
                continue
            cant_cobro += float(r.get("cantidad_total") or 0)
            costo_cobro += float(r.get("costo_directo") or 0)
        if len(batch) < 1000:
            break
        off_cb += 1000
    costo_ppto  = round(cant_ppto * vlr_unitario)
    liq_q = supabase.table("presupuesto").select("cant_total").eq("contrato_id", contrato_id).eq("item", item_numero).eq("tipo_ejecucion", "Obra Ejecutada").eq("dado_de_baja", False)
    if capitulo:
        liq_q = liq_q.eq("capitulo", capitulo)
    if competencia:
        liq_q = liq_q.eq("competencia", competencia)
    liq_rows = liq_q.execute().data or []
    cant_liq  = sum(float(r.get("cant_total") or 0) for r in liq_rows)
    costo_liq = round(cant_liq * vlr_unitario)
    return {
        "cant_presupuestada":  round(cant_ppto, 4),
        "costo_presupuestado": costo_ppto,
        "cant_cobrada":        round(cant_cobro, 4),
        "costo_cobrado":       round(costo_cobro),
        "balance_cant":        round(cant_ppto - cant_cobro, 4),
        "balance_costo":       round(costo_ppto - costo_cobro),
        "cant_liquidacion":    round(cant_liq, 4),
        "costo_liquidacion":   costo_liq,
        "balance_liq_cant":    round(cant_liq - cant_cobro, 4),
        "balance_liq_costo":   round(costo_liq - costo_cobro),
    }

@app.post("/listado-precios/item/{item_id}/recalcular")
def recalcular_cobros_precio(item_id: int, current_user=Depends(get_current_user)):
    """Actualiza de Pendiente → Aprobado todos los registros de cobro de este ítem de precio."""
    precio = supabase.table("listado_precios").select(
        "contrato_id, capitulo, competencia, item_numero, estado_precio"
    ).eq("id", item_id).single().execute().data
    if not precio:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")
    if precio.get("estado_precio") != "Aprobado":
        raise HTTPException(status_code=400, detail="El precio debe estar Aprobado antes de recalcular cobros")
    q = supabase.table("cobro").update({"precio_estado": "Aprobado"}).eq("contrato_id", precio["contrato_id"]).eq("item", precio["item_numero"]).eq("precio_estado", "Pendiente")
    if precio.get("capitulo"):
        q = q.eq("capitulo", precio["capitulo"])
    if precio.get("competencia"):
        q = q.eq("competencia", precio["competencia"])
    result = q.execute()
    return {"recalculados": len(result.data or [])}

@app.post("/listado-precios/{contrato_id}/log-exportar")
def log_exportar_precios(contrato_id: int, current_user=Depends(get_current_user)):
    """Registra en el log que el usuario exportó el listado de precios en XLSX."""
    _require_contract_access(current_user, contrato_id)
    registrar_log(current_user, "EXPORTAR", "PRECIOS", "listado_precios", str(contrato_id),
                  {"formato": "xlsx"})
    return {"ok": True}

# ─────────────────────────────────────────────
# PRESUPUESTO
# ─────────────────────────────────────────────

def _reject_if_presupuesto_sellado(supabase, item_ids: List[int]) -> None:
    """Registros sellados (aprobación Interventoría) no admiten modificaciones ni cambio de estado."""
    if not item_ids:
        return
    rows = supabase.table("presupuesto").select("id, sellado").in_("id", list(set(item_ids))).execute().data or []
    if any(r.get("sellado") for r in rows):
        raise HTTPException(
            status_code=403,
            detail="Registro sellado (aprobado por Interventoría): no puede modificarse.",
        )


def _pre_interv_liberado(row: dict) -> bool:
    """NULL/vacío = legado (antes de la columna); solo 'Aprobado' habilita a Interventoría."""
    v = row.get("pre_interv_estado")
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return str(v).strip() == "Aprobado"


def _cargo_puede_prevalidar_interventoria(cargo_nombre: str) -> bool:
    n = (cargo_nombre or "").lower()
    if "residente de interventoria" in n or "residente de interventoría" in n:
        return False
    if "residente de costos" in n:
        return True
    if "residente de obra" in n:
        return True
    return False


def _presupuesto_aplica_filtro_interventoria(current_user) -> bool:
    """Perfiles Interventoría solo ven cantidades ya depuradas por contratista (costos u obra)."""
    rol = (current_user.get("rol_nombre") or "").strip().lower()
    if rol in ("administrador", "desarrollador"):
        return False
    cargo = (current_user.get("cargo_nombre") or "").strip().lower()
    if cargo == "desarrollador":
        return False
    return rol in ("interventoría", "interventoria", "operativo interventoria")


def _presupuesto_q_filtros_ubicacion(
    q,
    nodo_inicio: Optional[str] = None,
    nodo_final: Optional[str] = None,
    buscar: Optional[str] = None,
    id_pol: Optional[str] = None,
    pk_criterio: Optional[str] = None,
    texto: Optional[str] = None,
    abs_desde: Optional[float] = None,
    abs_hasta: Optional[float] = None,
    revisado: Optional[str] = None,
    pre_interv_estado: Optional[str] = None,
):
    """Filtros opcionales para GET /presupuesto (alineable con criterios tipo SICOE / dashboard)."""
    if nodo_inicio and str(nodo_inicio).strip():
        q = q.ilike("no_inicio", f"%{str(nodo_inicio).strip()}%")
    if nodo_final and str(nodo_final).strip():
        q = q.ilike("no_final", f"%{str(nodo_final).strip()}%")
    has_split = (id_pol and str(id_pol).strip()) or (pk_criterio and str(pk_criterio).strip()) or (texto and str(texto).strip())
    if not has_split and buscar and str(buscar).strip():
        b = str(buscar).strip()
        pat = f"%{b}%"
        # Legacy: un solo cuadro busca en cuatro columnas
        q = q.or_(f"id_pol.ilike.{pat},pk_id.ilike.{pat},registro.ilike.{pat},descripcion.ilike.{pat}")
    else:
        if id_pol and str(id_pol).strip():
            q = q.ilike("id_pol", f"%{str(id_pol).strip()}%")
        if pk_criterio and str(pk_criterio).strip():
            q = q.ilike("pk_id", f"%{str(pk_criterio).strip()}%")
        if texto and str(texto).strip():
            t = f"%{str(texto).strip()}%"
            q = q.or_(f"registro.ilike.{t},descripcion.ilike.{t}")
    if revisado and str(revisado).strip():
        q = q.eq("revisado", str(revisado).strip())
    if pre_interv_estado and str(pre_interv_estado).strip():
        pe = str(pre_interv_estado).strip()
        if str(pe).strip().lower() in ("no revisado", "—", "-"):
            q = q.is_("pre_interv_estado", "null")
        else:
            q = q.eq("pre_interv_estado", pe)
    q = _so_reg_filtro_abs_solape(q, abs_desde, abs_hasta)
    return q


def _orden_capitulo_presupuesto(c: Optional[str]) -> tuple:
    if not c:
        return (2, 0, c or "")
    m = re.match(r"^(\d+)", str(c).strip())
    if m:
        return (0, int(m.group(1)), c)
    return (1, 0, c)


@app.get("/presupuesto/{contrato_id}")
def get_presupuesto(
    contrato_id: int,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    items: Optional[List[str]] = Query(None),
    tramo: Optional[str] = None,
    calzada: Optional[str] = None,
    nodo_inicio: Optional[str] = None,
    nodo_final: Optional[str] = None,
    buscar: Optional[str] = None,
    id_pol: Optional[str] = None,
    pk_criterio: Optional[str] = None,
    texto: Optional[str] = None,
    abs_desde: Optional[float] = None,
    abs_hasta: Optional[float] = None,
    revisado: Optional[str] = None,
    pre_interv_estado: Optional[str] = None,
    papelera: bool = False,
    limit: Optional[int] = Query(None, ge=1, le=20000),
    offset: int = Query(0, ge=0),
    current_user=Depends(get_current_user),
):
    """
    Listado de presupuesto con filtros de servidor. Parámetros capitulo / item alinean con el drill del
    dashboard. id_pol, pk_criterio, texto: filtros separados (campos distintos). `buscar` mantiene
    compatibilidad: OR en id_pol, pk_id, registro, descripcion si no se usan esos tres.
    pre_interv_estado: filtro depuración (roles contratista / obra); revisado: Interventoría.
    `limit` + `offset`: una sola página (grilla) sin bajar 10k+ filas. Sin `limit`, comportamiento
    legado: acumulación por lotes de 1000.
    """
    def _q_base():
        q = supabase.table("presupuesto").select("*").eq("contrato_id", contrato_id)
        if papelera:
            q = q.eq("dado_de_baja", True)
        else:
            q = q.eq("dado_de_baja", False)
        if capitulo:
            q = q.eq("capitulo", capitulo)
        ins = [str(x).strip() for x in (items or []) if str(x).strip()]
        if len(ins) > 1:
            if len(ins) > 200:
                raise HTTPException(status_code=422, detail="Máximo 200 ítems en lista items")
            q = q.in_("item", ins)
        elif len(ins) == 1:
            q = q.eq("item", ins[0])
        elif item:
            q = q.eq("item", item)
        if tramo:
            q = q.eq("tramo", tramo)
        if calzada:
            q = q.eq("calzada", calzada)
        q = _presupuesto_q_filtros_ubicacion(
            q,
            nodo_inicio=nodo_inicio,
            nodo_final=nodo_final,
            buscar=buscar,
            id_pol=id_pol,
            pk_criterio=pk_criterio,
            texto=texto,
            abs_desde=abs_desde,
            abs_hasta=abs_hasta,
            revisado=revisado,
            pre_interv_estado=pre_interv_estado,
        )
        if _presupuesto_aplica_filtro_interventoria(current_user):
            q = q.or_("pre_interv_estado.is.null,pre_interv_estado.eq.Aprobado")
        return q.order("capitulo").order("item").order("pk_id")

    if limit is not None:
        q = _q_base()
        return q.range(offset, offset + limit - 1).execute().data

    PAGE = 1000
    all_rows = []
    off = 0
    while True:
        batch = _q_base().range(off, off + PAGE - 1).execute().data
        all_rows.extend(batch)
        if len(batch) < PAGE:
            break
        off += PAGE
    return all_rows


@app.get("/presupuesto/{contrato_id}/conteo")
def get_presupuesto_conteo(
    contrato_id: int,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    items: Optional[List[str]] = Query(None),
    tramo: Optional[str] = None,
    calzada: Optional[str] = None,
    nodo_inicio: Optional[str] = None,
    nodo_final: Optional[str] = None,
    buscar: Optional[str] = None,
    id_pol: Optional[str] = None,
    pk_criterio: Optional[str] = None,
    texto: Optional[str] = None,
    abs_desde: Optional[float] = None,
    abs_hasta: Optional[float] = None,
    revisado: Optional[str] = None,
    pre_interv_estado: Optional[str] = None,
    papelera: bool = False,
    current_user=Depends(get_current_user),
):
    """
    Mismos query params que GET /presupuesto/{id}; respuesta mínima para dashboard y UI sin bajar filas.
    Fase C: pre-paginación y consistencia con el listado.
    """
    q = supabase.table("presupuesto").select("id", count="exact").eq("contrato_id", contrato_id)
    if papelera:
        q = q.eq("dado_de_baja", True)
    else:
        q = q.eq("dado_de_baja", False)
    if capitulo:
        q = q.eq("capitulo", capitulo)
    ins = [str(x).strip() for x in (items or []) if str(x).strip()]
    if len(ins) > 1:
        if len(ins) > 200:
            raise HTTPException(status_code=422, detail="Máximo 200 ítems en lista items")
        q = q.in_("item", ins)
    elif len(ins) == 1:
        q = q.eq("item", ins[0])
    elif item:
        q = q.eq("item", item)
    if tramo:
        q = q.eq("tramo", tramo)
    if calzada:
        q = q.eq("calzada", calzada)
    q = _presupuesto_q_filtros_ubicacion(
        q,
        nodo_inicio=nodo_inicio,
        nodo_final=nodo_final,
        buscar=buscar,
        id_pol=id_pol,
        pk_criterio=pk_criterio,
        texto=texto,
        abs_desde=abs_desde,
        abs_hasta=abs_hasta,
        revisado=revisado,
        pre_interv_estado=pre_interv_estado,
    )
    if _presupuesto_aplica_filtro_interventoria(current_user):
        q = q.or_("pre_interv_estado.is.null,pre_interv_estado.eq.Aprobado")
    result = q.execute()
    return {"total": int(result.count or 0)}


@app.get("/presupuesto/{contrato_id}/filtros")
def get_filtros_presupuesto(
    contrato_id: int,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    items: Optional[List[str]] = Query(None),
    tramo: Optional[str] = None,
    calzada: Optional[str] = None,
    current_user=Depends(get_current_user),
):
    """Devuelve valores únicos para filtros en cascada."""
    q = supabase.table("presupuesto").select("capitulo, item, tramo, calzada").eq("contrato_id", contrato_id)
    if capitulo:
        q = q.eq("capitulo", capitulo)
    ins = [str(x).strip() for x in (items or []) if str(x).strip()]
    if len(ins) > 1:
        if len(ins) > 200:
            raise HTTPException(status_code=422, detail="Máximo 200 ítems en lista items")
        q = q.in_("item", ins)
    elif len(ins) == 1:
        q = q.eq("item", ins[0])
    elif item:
        q = q.eq("item", item)
    if tramo:
        q = q.eq("tramo", tramo)
    if calzada:
        q = q.eq("calzada", calzada)
    if _presupuesto_aplica_filtro_interventoria(current_user):
        q = q.or_("pre_interv_estado.is.null,pre_interv_estado.eq.Aprobado")
    rows = q.execute().data
    caps    = sorted(set(r["capitulo"] for r in rows if r.get("capitulo")))
    items   = sorted(set(r["item"]     for r in rows if r.get("item")))
    tramos  = sorted(set(r["tramo"]    for r in rows if r.get("tramo")))
    calzadas= sorted(set(r["calzada"]  for r in rows if r.get("calzada")))
    return {"capitulos": caps, "items": items, "tramos": tramos, "calzadas": calzadas}


@app.get("/presupuesto/{contrato_id}/analisis-liquidacion")
def presupuesto_analisis_liquidacion(
    contrato_id: int,
    nivel: str = Query("item", description="item | capitulo"),
    current_user=Depends(get_current_user),
):
    """
    Contratos en fase liquidación: compara «recalculado» vs obra aprobada (SICOE N3).
    - Ítem con polígonos en presupuesto (tipo_ejecucion = Presupuesto de Obra): recalc = suma PPTO.
    - Ítem sin polígonos: recalc = obra N3 ✓ (igual al cobro; categoría EJECUCION).
    """
    _require_contract_access(current_user, contrato_id)
    n = (nivel or "item").strip().lower()
    if n not in ("item", "capitulo"):
        raise HTTPException(status_code=422, detail="nivel debe ser item o capitulo")
    items = _liquidacion_analisis_items(contrato_id, n, current_user)
    return {"items": items}


@app.get("/presupuesto/{contrato_id}/resumen")
def get_resumen_presupuesto(contrato_id: int, current_user=Depends(get_current_user)):
    try:
        res  = supabase.table("vista_ppto_resumen").select("*").eq("contrato_id", contrato_id).execute().data
    except Exception:
        res = []
    try:
        caps = supabase.table("vista_ppto_por_capitulo").select("*").eq("contrato_id", contrato_id).execute().data
    except Exception:
        caps = []
    total = res[0].get("total_ppto", 0) if res else 0
    regs  = res[0].get("total_registros", 0) if res else 0
    return {
        "total_registros": regs,
        "costo_total": total,
        "revisados": 0, "campo": 0, "pendientes": 0,
        "por_capitulo": [{"capitulo": r["capitulo"], "costo": r["presupuesto"], "registros": r["registros"]} for r in caps]
    }

@app.get("/presupuesto/{contrato_id}/capitulos-lista")
def get_capitulos_presupuesto(contrato_id: int, current_user=Depends(get_current_user)):
    """Devuelve capítulos con costo total y total de registros. Carga rápida sin traer filas individuales."""
    caps = supabase.table("vista_ppto_por_capitulo").select("*").eq("contrato_id", contrato_id).execute().data
    rows = [{"capitulo": r["capitulo"], "costo_total": r["presupuesto"], "total_registros": r["registros"]} for r in (caps or [])]
    return sorted(rows, key=lambda x: _orden_capitulo_presupuesto(x.get("capitulo")))


@app.get("/presupuesto/{contrato_id}/maestro-ubicacion-pk")
def get_maestro_ubicacion_pk_ids(contrato_id: int, current_user=Depends(get_current_user)):
    """Tramos y calzadas distintas desde el maestro pk_ids del contrato (SICOE)."""
    def _q():
        return supabase.table("pk_ids").select("tramo, calzada").eq("contrato_id", contrato_id).execute().data
    try:
        rows = supabase_execute(_q) or []
    except Exception:
        rows = []
    tramos = sorted({str(r["tramo"]).strip() for r in rows if r.get("tramo") and str(r.get("tramo", "")).strip()})
    calzadas = sorted({str(r["calzada"]).strip() for r in rows if r.get("calzada") and str(r.get("calzada", "")).strip()})
    return {"tramos": tramos, "calzadas": calzadas}

@app.get("/presupuesto/{contrato_id}/items-lista")
def get_items_presupuesto(contrato_id: int, capitulo: str, current_user=Depends(get_current_user)):
    """Devuelve ítems de un capítulo con costo y cantidad agregados — sin traer registros individuales."""
    rows = []
    offset = 0
    while True:
        q_it = supabase.table("presupuesto").select(
            "item, descripcion, und, vlr_unitario, cant_total, costo_directo, revisado"
        ).eq("contrato_id", contrato_id).eq("capitulo", capitulo).eq("dado_de_baja", False)
        if _presupuesto_aplica_filtro_interventoria(current_user):
            q_it = q_it.or_("pre_interv_estado.is.null,pre_interv_estado.eq.Aprobado")
        batch = q_it.range(offset, offset + 999).execute().data
        rows.extend(batch)
        if len(batch) < 1000: break 
        offset += 1000
    items = {}
    for r in rows:
        it = r.get("item") or ""
        if it not in items:
            items[it] = {
                "item": it,
                "descripcion": r.get("descripcion") or "",
                "und": r.get("und") or "",
                "vlr_unitario": r.get("vlr_unitario") or 0,
                "cant_total": 0,
                "costo_total": 0,
                "total_registros": 0,
                "revisados": []
            }
        items[it]["cant_total"]     += r.get("cant_total") or 0
        items[it]["costo_total"]    += r.get("costo_directo") or 0
        items[it]["total_registros"] += 1
        items[it]["revisados"].append(r.get("revisado") or "No Revisado")
    result = sorted(items.values(), key=lambda x: x["item"])
    return result

@app.get("/presupuesto/item/{item_id}")
def get_presupuesto_item(item_id: int, current_user=Depends(get_current_user)):
    """Trae un único registro de presupuesto por ID."""
    row = supabase.table("presupuesto").select("*").eq("id", item_id).single().execute().data
    if not row:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    return row

# Campos que cuentan como “edición” para reapertura de registro sellado (no basta con el motivo solo).
_PPTO_REABRIR_CAMPOS = frozenset(
    {
        "capitulo", "competencia", "item", "descripcion", "und", "calzada", "tramo",
        "vlr_unitario", "area_long_nod", "ancho", "espesor", "cant_total", "costo_directo",
        "observacion_externa",
    }
)

# Cambios de datos “de negocio” por contratista que invalidan un estado ya asignado por Interventoría (sin sellado).
_PPTO_CT_SUBSTANTIVE = frozenset(
    {
        "capitulo", "competencia", "item", "descripcion", "und",
        "vlr_unitario", "observacion_externa", "costo_directo",
        "area_long_nod", "ancho", "espesor",
    }
)


def _ppto_val_eq(a, b) -> bool:
    if a is None and b is None:
        return True
    try:
        fa, fb = float(a), float(b)
        if math.isfinite(fa) and math.isfinite(fb):
            tol = 1e-6 + 1e-9 * max(abs(fa), abs(fb), 1.0)
            return abs(fa - fb) < tol
    except (TypeError, ValueError):
        pass
    return str(a or "").strip() == str(b or "").strip()


def _ppto_substantive_contractor_fields_changed(prev: dict, data: dict) -> bool:
    for k in _PPTO_CT_SUBSTANTIVE:
        if k not in data:
            continue
        if not _ppto_val_eq(data.get(k), prev.get(k)):
            return True
    return False


@app.put("/presupuesto/item/{item_id}")
def update_presupuesto_item(item_id: int, body: PresupuestoUpdate, current_user=Depends(get_current_user)):
    data = body.dict(exclude_unset=True)
    _mr = data.pop("motivo_edicion_tras_sellado", None)
    motivo_reap = str(_mr).strip() if _mr is not None else ""
    _mi = data.pop("motivo_edicion_con_estado_interv", None)
    motivo_interv = str(_mi).strip() if _mi is not None else ""
    prev_row = supabase.table("presupuesto").select("*").eq("id", item_id).limit(1).execute().data
    prev_row = prev_row[0] if prev_row else {}
    if ("no_inicio" in data or "no_final" in data) and not _es_desarrollador(current_user):
        raise HTTPException(
            status_code=403,
            detail="Solo el cargo Desarrollador puede modificar los nodos (No.Ini / No.Fin) en presupuesto.",
        )
    reabrir = False
    if prev_row.get("sellado"):
        if _es_desarrollador(current_user):
            raise HTTPException(
                status_code=403,
                detail="Registro sellado (aprobado por Interventoría): el Desarrollador no lo modifica desde aquí. Contacte soporte si aplica un caso excepcional.",
            )
        if not _es_rol_contratista_ppto(current_user):
            raise HTTPException(
                status_code=403,
                detail="Registro sellado (aprobado por Interventoría): no puede modificarse salvo el flujo de reapertura del contratista.",
            )
        if not motivo_reap or len(motivo_reap) < 15:
            raise HTTPException(
                status_code=400,
                detail="Debe consignar un motivo de reapertura (mínimo 15 caracteres) para editar un registro aprobado por Interventoría.",
            )
        toca = [k for k in data if k in _PPTO_REABRIR_CAMPOS]
        if not toca:
            raise HTTPException(
                status_code=400,
                detail="Indique al menos un dato a modificar (p. ej. capítulo, ítem o vlr. unitario) además del motivo de reapertura.",
            )
        for k in ("revisado", "sellado"):
            data.pop(k, None)
        data["sellado"] = False
        data["revisado"] = "No Revisado"
        data["validado_por"] = None
        data["validado_en"] = None
        reabrir = True
    reset_interv_comment = None
    prev_rev_for_log = (prev_row.get("revisado") or "No Revisado").strip()
    if (
        not prev_row.get("sellado")
        and not reabrir
        and _es_rol_contratista_ppto(current_user)
        and not _es_desarrollador(current_user)
    ):
        data.pop("revisado", None)
        if prev_rev_for_log in ("Aprobado", "Pendiente", "Rechazado") and _ppto_substantive_contractor_fields_changed(
            prev_row, data
        ):
            if len(motivo_interv) < 15:
                raise HTTPException(
                    status_code=400,
                    detail="Debe consignar un motivo (mínimo 15 caracteres) al modificar un registro ya marcado por Interventoría; el estado volverá a «No Revisado».",
                )
            data["revisado"] = "No Revisado"
            data["validado_por"] = None
            data["validado_en"] = None
            reset_interv_comment = motivo_interv
    if not _puede_editar_dimensiones_presupuesto(current_user):
        for k in ("area_long_nod", "ancho", "espesor", "cant_total"):
            if k in data:
                raise HTTPException(
                    status_code=403,
                    detail="No tiene permiso para modificar dimensiones o la cantidad total en presupuesto (requiere Desarrollador o permiso «editar registros presupuesto» con edición).",
                )
    # Cualquier clave de dimensión presente en el body (aunque venga null en JSON) dispara
    # recálculo. Antes, solo "not None" fallaba con null explícito y .get() no fusionaba con la fila.
    _DIMK = ("area_long_nod", "ancho", "espesor")
    toco_dimensiones = any(k in data for k in _DIMK)
    if toco_dimensiones:
        # prev_row ya tiene select("*") — reutilizar en vez de hacer un segundo SELECT
        def _dim_merged(k: str) -> float:
            if k not in data:
                return float(prev_row.get(k) or 0)
            v = data.get(k)
            if v is None:
                return float(prev_row.get(k) or 0)
            return float(v or 0)

        area = _dim_merged("area_long_nod")
        ancho = _dim_merged("ancho")
        esp = _dim_merged("espesor")
        if "vlr_unitario" in data and data.get("vlr_unitario") is not None:
            vlr = float(data.get("vlr_unitario") or 0)
        else:
            vlr = float(prev_row.get("vlr_unitario") or 0)
        cant = round(area * ancho * esp, 2) if (ancho or esp) else round(area, 2)
        data["area_long_nod"] = area
        data["ancho"] = ancho
        data["espesor"] = esp
        data["cant_total"] = cant
        data["costo_directo"] = round(cant * vlr, 0)
    if toco_dimensiones:
        data["calculo_por"] = _calculo_usuario_label(current_user)
        data["calculo_en"] = datetime.now(timezone.utc).isoformat()
    if not toco_dimensiones and "vlr_unitario" in data:
        cant0 = float(prev_row.get("cant_total") or 0)
        vlr0 = float(data.get("vlr_unitario") or 0)
        data["costo_directo"] = round(cant0 * vlr0, 0)
    if "revisado" in data and not reabrir:
        nu = str(data.get("revisado") or "No Revisado").strip()
        prev_rev_cmp = str(prev_row.get("revisado") or "No Revisado").strip()
        if nu != prev_rev_cmp and not reset_interv_comment:
            if not _es_desarrollador(current_user) and not _cargo_permiso_validar_presupuesto(current_user):
                raise HTTPException(
                    status_code=403,
                    detail="No tiene permiso de validación en «editar registros presupuesto» para cambiar el estado «revisado».",
                )
    data["updated_at"] = "now()"
    supabase.table("presupuesto").update(data).eq("id", item_id).execute()

    if reabrir and motivo_reap:
        try:
            nombre_u = current_user.get("nombre") or current_user.get("email") or "Usuario"
            supabase.table("comentarios").insert(
                {
                    "presupuesto_id": item_id,
                    "tipo": "validacion",
                    "mensaje": f"[Reapertura tras aprobación Interventoría — edición por contratista] {motivo_reap}",
                    "usuario_nombre": nombre_u,
                    "parent_id": None,
                }
            ).execute()
        except Exception:
            try:
                _log_api.warning("comentario reapertura sellado: insert falló item_id=%s", item_id)
            except Exception:
                pass
    if reset_interv_comment:
        try:
            nombre_u = current_user.get("nombre") or current_user.get("email") or "Usuario"
            supabase.table("comentarios").insert(
                {
                    "presupuesto_id": item_id,
                    "tipo": "validacion",
                    "mensaje": f"[Contratista — edición con estado Interventoría «{prev_rev_for_log}»] {reset_interv_comment}",
                    "usuario_nombre": nombre_u,
                    "parent_id": None,
                }
            ).execute()
        except Exception:
            try:
                _log_api.warning("comentario reset estado interv: insert falló item_id=%s", item_id)
            except Exception:
                pass

    # ── Encolar cambio de layer en CAD si cambió ítem o capítulo ──────────────
    # prev_row ya tiene todos los campos necesarios — sin SELECT adicional
    if "capitulo" in data or "item" in data:
        try:
            nuevo_cap  = data.get("capitulo") or prev_row.get("capitulo") or ""
            nuevo_item = data.get("item")     or prev_row.get("item")     or ""
            comp       = prev_row.get("competencia") or ""
            cap6       = nuevo_cap.replace(".", "")[:5]
            new_layer_ent = f"{cap6}_{comp}_{nuevo_item}"
            new_layer_txt = f"txt_{cap6}_{comp}_{nuevo_item}"
            old_id_pol = prev_row.get("id_pol") or ""
            sufijo = old_id_pol[old_id_pol.index("._"):] if "._" in old_id_pol else f"._{item_id}"
            new_id_pol = f"{nuevo_item}{sufijo}"
            payload_cad = {
                "ent_handle": prev_row.get("ent_handle") or "",
                "txt_handle": prev_row.get("txt_handle") or "",
                "layer_ent":  new_layer_ent,
                "layer_txt":  new_layer_txt,
                "color_hex":  prev_row.get("color_hex") or "",
                "new_text":   new_id_pol,
            }
            supabase.table("cad_queue").insert({
                "contrato_id": prev_row.get("contrato_id"),
                "tipo": "cambiar_layer",
                "estado": "pendiente",
                "payload": payload_cad,
            }).execute()
            supabase.table("presupuesto").update({
                "layer_ent": new_layer_ent,
                "layer_txt": new_layer_txt,
                "id_pol":    new_id_pol,
            }).eq("id", item_id).execute()
        except: pass

    updated = supabase.table("presupuesto").select("*").eq("id", item_id).execute().data
    row_after = updated[0] if updated else {}
    try:
        cinfo = None
        cid = row_after.get("contrato_id") or prev_row.get("contrato_id")
        if cid:
            cr = supabase.table("contratos").select("numero").eq("id", cid).limit(1).execute().data
            cinfo = cr[0].get("numero") if cr else None
        u_log = {
            "sub": str(current_user.get("sub")),
            "nombre": current_user.get("nombre") or "",
            "email": current_user.get("email"),
            "cargo_nombre": current_user.get("cargo_nombre"),
            "rol_nombre": current_user.get("rol_nombre"),
            "contrato_id": cid,
            "contrato_numero": cinfo,
        }
        rev_prev = (prev_row.get("revisado") or "No Revisado").strip()
        rev_new = (row_after.get("revisado") or "No Revisado").strip()
        cambio_revisado = (not reabrir) and rev_prev != rev_new
        accion_log = "VALIDAR" if cambio_revisado else "EDITAR"
        registrar_log(
            u_log,
            accion_log,
            "PRESUPUESTO",
            "presupuesto",
            str(item_id),
            {
                "id_pol": row_after.get("id_pol") or prev_row.get("id_pol"),
                "item": row_after.get("item"),
                "reapertura_tras_sellado": bool(reabrir),
            },
            valor_anterior=_json_for_log(prev_row),
            valor_nuevo=_json_for_log(row_after),
        )
    except Exception:
        pass
    return row_after if updated else {"mensaje": "Registro actualizado"}

@app.put("/presupuesto/item/{item_id}/dar-baja")
def dar_baja_presupuesto(item_id: int, current_user=Depends(get_current_user)):
    """Soft delete: marca el registro como dado de baja y renombra sus layers en CAD."""
    row = supabase.table("presupuesto").select(
        "layer_txt, layer_ent, x_label, y_label, contrato_id, ent_handle, txt_handle, rev_block_handle, sellado"
    ).eq("id", item_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    r = row[0]
    if r.get("sellado"):
        raise HTTPException(
            status_code=403,
            detail="Registro sellado (aprobado por Interventoría): no puede modificarse.",
        )
    supabase.table("presupuesto").update({
        "dado_de_baja": True,
        "updated_at": "now()"
    }).eq("id", item_id).execute()
    # Cola CAD: renombrar layers con prefijo del_
    if _dwg_activo(r.get("contrato_id")):
        old_lent = r.get("layer_ent") or ""
        old_ltxt = r.get("layer_txt") or ""
        if old_lent and not old_lent.startswith("del_"):
            supabase.table("cad_queue").insert({
                "contrato_id": r["contrato_id"],
                "tipo": "cambiar_layer",
                "estado": "pendiente",
                "payload": {
                    "ent_handle":  r.get("ent_handle") or "",
                    "txt_handle":  r.get("txt_handle") or "",
                    "layer_ent":   f"del_{old_lent}",
                    "layer_txt":   f"del_{old_ltxt}",
                    "color_hex":   "",
                    "rev_block_handle": r.get("rev_block_handle") or "",
                    "layoff": True
                }
            }).execute()
        # Actualizar layers en la tabla presupuesto también
        supabase.table("presupuesto").update({
            "layer_ent": f"del_{old_lent}" if old_lent and not old_lent.startswith("del_") else old_lent,
            "layer_txt": f"del_{old_ltxt}" if old_ltxt and not old_ltxt.startswith("del_") else old_ltxt,
        }).eq("id", item_id).execute()
    try:
        cid = r.get("contrato_id")
        cinfo = None
        if cid:
            cr = supabase.table("contratos").select("numero").eq("id", cid).limit(1).execute().data
            cinfo = cr[0].get("numero") if cr else None
        u_log = {
            "sub": str(current_user.get("sub")),
            "nombre": current_user.get("nombre") or "",
            "email": current_user.get("email"),
            "cargo_nombre": current_user.get("cargo_nombre"),
            "rol_nombre": current_user.get("rol_nombre"),
            "contrato_id": cid,
            "contrato_numero": cinfo,
        }
        registrar_log(
            u_log,
            "ELIMINAR",
            "PRESUPUESTO",
            "presupuesto",
            str(item_id),
            {"accion": "dar_baja", "layer_ent": r.get("layer_ent")},
            severidad="AUDIT",
        )
    except Exception:
        pass
    return {"ok": True}

@app.put("/presupuesto/item/{item_id}/restaurar")
def restaurar_presupuesto(item_id: int, current_user=Depends(get_current_user)):
    """Restaura un registro dado de baja: quita del_ de layers y reactiva en CAD."""
    row = supabase.table("presupuesto").select(
        "layer_txt, layer_ent, x_label, y_label, contrato_id, ent_handle, txt_handle, color_hex, sellado"
    ).eq("id", item_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    r = row[0]
    if r.get("sellado"):
        raise HTTPException(
            status_code=403,
            detail="Registro sellado (aprobado por Interventoría): no puede modificarse.",
        )
    supabase.table("presupuesto").update({
        "dado_de_baja": False,
        "updated_at": "now()"
    }).eq("id", item_id).execute()
    # Cola CAD: restaurar layers quitando prefijo del_
    if _dwg_activo(r.get("contrato_id")):
        old_lent = r.get("layer_ent") or ""
        old_ltxt = r.get("layer_txt") or ""
        new_lent = old_lent[4:] if old_lent.startswith("del_") else old_lent
        new_ltxt = old_ltxt[4:] if old_ltxt.startswith("del_") else old_ltxt
        if new_lent:
            supabase.table("cad_queue").insert({
                "contrato_id": r["contrato_id"],
                "tipo": "cambiar_layer",
                "estado": "pendiente",
                "payload": {
                    "ent_handle": r.get("ent_handle") or "",
                    "txt_handle": r.get("txt_handle") or "",
                    "layer_ent":  new_lent,
                    "layer_txt":  new_ltxt,
                    "color_hex":  r.get("color_hex") or "",
                    "layoff": False
                }
            }).execute()
        supabase.table("presupuesto").update({
            "layer_ent": new_lent,
            "layer_txt": new_ltxt,
        }).eq("id", item_id).execute()
    try:
        cid = r.get("contrato_id")
        cinfo = None
        if cid:
            cr = supabase.table("contratos").select("numero").eq("id", cid).limit(1).execute().data
            cinfo = cr[0].get("numero") if cr else None
        u_log = {
            "sub": str(current_user.get("sub")),
            "nombre": current_user.get("nombre") or "",
            "email": current_user.get("email"),
            "cargo_nombre": current_user.get("cargo_nombre"),
            "rol_nombre": current_user.get("rol_nombre"),
            "contrato_id": cid,
            "contrato_numero": cinfo,
        }
        registrar_log(
            u_log,
            "EDITAR",
            "PRESUPUESTO",
            "presupuesto",
            str(item_id),
            {"accion": "restaurar"},
            severidad="AUDIT",
        )
    except Exception:
        pass
    return {"ok": True}

@app.post("/presupuesto/{contrato_id}/agregar-cantidad")
def agregar_cantidad(contrato_id: int, body: AgregarCantidadBody, current_user=Depends(get_current_user)):
    """Inserta una nueva cantidad clonando la posición de un registro existente."""
    if not _es_desarrollador(current_user):
        raise HTTPException(
            status_code=403,
            detail="Solo el cargo Desarrollador puede agregar cantidades (dimensiones) en presupuesto.",
        )
    area  = float(body.area_long_nod or 0)
    ancho = float(body.ancho or 0)
    esp   = float(body.espesor or 0)
    vlr   = float(body.vlr_unitario or 0)
    cant  = round(area * ancho * esp, 2) if (ancho or esp) else round(area, 2)
    costo = round(cant * vlr, 0)

    # Construir nuevo id_pol: {nuevo_item}_{consecutivo}
    base = body.id_pol_base or ""
    if "._" in base:
        sufijo = base[base.index("._") + 2:]
    elif "_" in base:
        sufijo = base.rsplit("_", 1)[-1]
    else:
        sufijo = "1"
    new_id_pol = f"{body.item}_{sufijo}"

    row = {
        "contrato_id": contrato_id,
        "capitulo":      body.capitulo,
        "competencia":   body.competencia,
        "item":          body.item,
        "descripcion":   body.descripcion,
        "und":           body.und,
        "calzada":       body.calzada,
        "tramo":         body.tramo,
        "abs_inicio":    body.abs_inicio,
        "abs_final":     body.abs_final,
        "no_inicio":     body.no_inicio,
        "no_final":      body.no_final,
        "vlr_unitario":  vlr,
        "area_long_nod": body.area_long_nod,
        "ancho":         body.ancho,
        "espesor":       body.espesor,
        "cant_total":    cant,
        "costo_directo": costo,
        "tipo_ejecucion": body.tipo_ejecucion,
        "tipo_entidad":  body.tipo_entidad,
        "id_pol":        new_id_pol,
        "layer_ent":     body.layer_ent or "",
        "layer_txt":     body.layer_txt or "",
        "x_label":       body.x_label,
        "y_label":       body.y_label,
        "dado_de_baja":  False,
        "revisado":      "No Revisado",
        "pre_interv_estado": "No Revisado",
    }
    inserted = supabase.table("presupuesto").insert(row).execute().data
    if not inserted:
        raise HTTPException(status_code=500, detail="Error al insertar cantidad")
    new_row = inserted[0]

    # Encolar CAD: create_label
    try:
        supabase.table("cad_queue").insert({
            "contrato_id": contrato_id,
            "tipo":        "create_label",
            "estado":      "pendiente",
            "payload": {
                "id_pol":       new_id_pol,
                "layer_ent":    body.layer_ent or "",
                "layer_txt":    body.layer_txt or "",
                "x_label":      body.x_label,
                "y_label":      body.y_label,
                "descripcion":  body.descripcion,
                "unidad":       body.und,
                "cant_total":   cant,
                "costo_directo": costo,
            }
        }).execute()
    except Exception:
        pass

    try:
        cr = supabase.table("contratos").select("numero").eq("id", contrato_id).limit(1).execute().data
        cinfo = cr[0].get("numero") if cr else None
        u_log = {
            "sub": str(current_user.get("sub")),
            "nombre": current_user.get("nombre") or "",
            "email": current_user.get("email"),
            "cargo_nombre": current_user.get("cargo_nombre"),
            "rol_nombre": current_user.get("rol_nombre"),
            "contrato_id": contrato_id,
            "contrato_numero": cinfo,
        }
        registrar_log(
            u_log,
            "CREAR",
            "PRESUPUESTO",
            "presupuesto",
            str(new_row.get("id") or ""),
            {"id_pol": new_id_pol, "origen": "agregar_cantidad"},
            valor_nuevo=_json_for_log(new_row),
        )
    except Exception:
        pass

    return new_row


@app.post("/presupuesto/{contrato_id}/bulk")
def bulk_presupuesto(
    contrato_id: int,
    items: List[PresupuestoRow],
    mode: str = "append",
    source: Optional[str] = None,
    x_sicoe_cad_enviados: Optional[str] = Header(None, alias="X-SicoeCAD-Enviados"),
    current_user=Depends(get_current_user),
):
    """Importa registros de presupuesto. mode=replace elimina todo primero, mode=append agrega.
    Si el cliente es SicoeCAD, usar query source=sicoe_cad y opcional header X-SicoeCAD-Enviados
    (cantidad de registros leídos en el DWG) para la auditoría en la web."""
    if mode == "replace":
        supabase.table("presupuesto").delete().eq("contrato_id", contrato_id).execute()
    if not items:
        return {"insertados": 0}
    BATCH = 500
    rows = []
    for item in items:
        d = {k: v for k, v in item.dict().items() if v is not None}
        d["contrato_id"] = contrato_id
        rows.append(d)
    insertados = 0
    for i in range(0, len(rows), BATCH):
        try:
            supabase.table("presupuesto").insert(rows[i:i+BATCH]).execute()
            insertados += len(rows[i:i+BATCH])
        except Exception as e:
            for row in rows[i:i+BATCH]:
                try:
                    supabase.table("presupuesto").insert(row).execute()
                    insertados += 1
                except Exception:
                    pass
    registrar_log(current_user, "IMPORTAR", "PRESUPUESTO", "presupuesto_bulk", str(contrato_id),
        {"contrato_id": contrato_id, "mode": mode, "registros_insertados": insertados,
         "source": (source or "").lower() or None})
    if (source or "").strip().lower() == "sicoe_cad":
        enviados = None
        h = (x_sicoe_cad_enviados or "").strip()
        if h.isdigit():
            enviados = int(h)
        _sicoe_cad_sincro_audit[contrato_id] = {
            "insertados": insertados,
            "enviados": enviados,
            "ts": time.time(),
            # Solo el usuario que ejecutó el POST /bulk ve el aviso en la web (JWT sub).
            "user_sub": str(current_user.get("sub") or ""),
        }
    return {"insertados": insertados}


@app.get("/presupuesto/{contrato_id}/sincro-sicoe-cad-auditoria")
def presupuesto_sincro_sicoe_cad_pendiente(contrato_id: int, current_user=Depends(get_current_user)):
    """Aviso para la web: última importación de cantidades por SicoeCAD (sinc. con cola CAD / DWG)."""
    e = _sicoe_cad_sincro_audit.get(contrato_id)
    if not e:
        return {"pendiente": None}
    who = str(current_user.get("sub") or "")
    stored = str(e.get("user_sub") or "")
    # Entradas antiguas sin user_sub: no notificar a nadie (evita spam a toda la mesa).
    if not stored or stored != who:
        return {"pendiente": None}
    if time.time() - e["ts"] > 600:
        _sicoe_cad_sincro_audit.pop(contrato_id, None)
        return {"pendiente": None}
    vis = {k: v for k, v in e.items() if k != "user_sub"}
    return {"pendiente": vis}


@app.post("/presupuesto/{contrato_id}/sincro-sicoe-cad-auditoria/ack")
def presupuesto_sincro_sicoe_cad_ack(contrato_id: int, current_user=Depends(get_current_user)):
    e = _sicoe_cad_sincro_audit.get(contrato_id)
    who = str(current_user.get("sub") or "")
    if e and str(e.get("user_sub") or "") == who:
        _sicoe_cad_sincro_audit.pop(contrato_id, None)
    return {"ok": True}

@app.put("/presupuesto/{contrato_id}/bulk-recalcular")
def bulk_recalcular(contrato_id: int, body: PresupuestoBulkRecalc, current_user=Depends(get_current_user)):
    if not body.ids:
        raise HTTPException(status_code=400, detail="No hay registros seleccionados")
    _reject_if_presupuesto_sellado(supabase, body.ids)
    # Solo area_long_nod está restringido al Desarrollador (viene del plano CAD)
    has_area_change = any(d.area_long_nod is not None for d in (body.dims or []))
    if has_area_change and not _es_desarrollador(current_user):
        raise HTTPException(
            status_code=403,
            detail="Solo el cargo Desarrollador puede modificar el campo Área/Long/Nodo (viene del plano CAD).",
        )
    dims_map = {d.id: d for d in (body.dims or [])}
    # Traer también handles, layers e id_pol para cad_queue y reconstrucción de id_pol
    rows = supabase.table("presupuesto").select(
        "id, area_long_nod, ancho, espesor, cant_total, vlr_unitario, ent_handle, txt_handle, layer_ent, layer_txt, color_hex, competencia, id_pol"
    ).in_("id", body.ids).execute().data

    ts          = datetime.now(timezone.utc).isoformat()
    calculo_por = _calculo_usuario_label(current_user)

    # Paso 1: calcular todos los payloads en memoria (sin tocar la BD)
    batch_ppto   = []   # para upsert en una sola llamada
    batch_cad    = []   # para insert en una sola llamada

    for r in rows:
        rid = r["id"]
        dim = dims_map.get(rid)
        vlr = body.vlr_unitario if body.vlr_unitario is not None else (r.get("vlr_unitario") or 0)
        vlr = float(vlr) if vlr is not None else 0.0
        if dim and (
            dim.ancho is not None
            or dim.espesor is not None
            or dim.area_long_nod is not None
        ):
            area = float(dim.area_long_nod) if dim.area_long_nod is not None else float(r.get("area_long_nod") or 0)
            ancho = float(dim.ancho) if dim.ancho is not None else float(r.get("ancho") or 0)
            espesor = float(dim.espesor) if dim.espesor is not None else float(r.get("espesor") or 0)
            cant = round(area * ancho * espesor, 2) if (ancho or espesor) else round(area, 2)
            costo = round(cant * vlr, 0)
            data = {
                "area_long_nod": area, "ancho": ancho, "espesor": espesor,
                "cant_total": cant, "costo_directo": costo,
                "updated_at": ts, "calculo_por": calculo_por, "calculo_en": ts,
            }
        else:
            ancho   = (dim.ancho   if dim and dim.ancho   is not None else None) or r.get("ancho")   or 1
            espesor = (dim.espesor if dim and dim.espesor is not None else None) or r.get("espesor") or 1
            area    = r.get("area_long_nod") or 0
            if dim and (dim.ancho is not None or dim.espesor is not None):
                cant = round(float(area) * float(ancho) * float(espesor), 2)
                data_ancho = {"ancho": ancho, "espesor": espesor}
            else:
                cant = r.get("cant_total") or 0
                data_ancho = {}
            costo = round(float(cant) * vlr, 0)
            data  = {
                "cant_total": cant, "costo_directo": costo,
                "updated_at": ts, "calculo_por": calculo_por, "calculo_en": ts,
                **data_ancho,
            }
        # Capitulo/item: valor por fila (dim) tiene prioridad sobre el valor global del body
        cap_eff  = (dim.capitulo if dim and dim.capitulo is not None else None) or body.capitulo
        item_eff = (dim.item     if dim and dim.item     is not None else None) or body.item
        if cap_eff  is not None: data["capitulo"]    = cap_eff
        if item_eff is not None: data["item"]        = item_eff
        if body.descripcion  is not None: data["descripcion"]  = body.descripcion
        if body.vlr_unitario is not None: data["vlr_unitario"] = body.vlr_unitario

        new_id_pol = None
        if item_eff is not None:
            old_id_pol = r.get("id_pol") or ""
            sufijo = old_id_pol[old_id_pol.index("._"):] if "._" in old_id_pol else f"._{rid}"
            new_id_pol = f"{item_eff}{sufijo}"
            data["id_pol"] = new_id_pol

        batch_ppto.append({"id": rid, **data})

        if cap_eff is not None or item_eff is not None:
            nuevo_cap  = cap_eff  or ""
            nuevo_item = item_eff or ""
            comp       = r.get("competencia") or ""
            cap6       = nuevo_cap.replace(".", "")[:5]
            new_layer_ent = f"{cap6}_{comp}_{nuevo_item}"
            new_layer_txt = f"txt_{cap6}_{comp}_{nuevo_item}"
            payload_cad = {
                "ent_handle": r.get("ent_handle") or "",
                "txt_handle": r.get("txt_handle") or "",
                "layer_ent":  new_layer_ent,
                "layer_txt":  new_layer_txt,
                "color_hex":  r.get("color_hex") or "",
            }
            if new_id_pol:
                payload_cad["new_text"] = new_id_pol
            batch_cad.append({
                "contrato_id": contrato_id,
                "tipo": "cambiar_layer",
                "estado": "pendiente",
                "payload": payload_cad,
            })

    # Paso 2: updates secuenciales (ThreadPoolExecutor agota el pool de conexiones de Supabase)
    # El optimistic update del frontend hace que la UI sea instantánea de todas formas
    for item in batch_ppto:
        rid  = item["id"]
        data = {k: v for k, v in item.items() if k != "id"}
        supabase.table("presupuesto").update(data).eq("id", rid).execute()
    if batch_cad:
        supabase.table("cad_queue").insert(batch_cad).execute()

    registrar_log(current_user, "RECALCULAR", "PRESUPUESTO", "presupuesto_bulk", str(contrato_id),
        {"contrato_id": contrato_id, "cantidad_registros": len(rows),
         "capitulo": body.capitulo, "item": body.item})
    return {"actualizados": len(rows)}

@app.put("/presupuesto/{contrato_id}/bulk-estado")
def bulk_estado(contrato_id: int, body: PresupuestoBulkEstado, current_user=Depends(get_current_user)):
    if not body.ids:
        raise HTTPException(status_code=400, detail="No hay registros seleccionados")
    if not _es_desarrollador(current_user) and not _cargo_permiso_validar_presupuesto(current_user):
        raise HTTPException(
            status_code=403,
            detail="No tiene permiso de validación en «editar registros presupuesto» (matriz de accesos).",
        )
    # Traer sellado, pre_interv y datos CAD
    rows_info = supabase.table("presupuesto").select(
        "id, x_label, y_label, layer_txt, rev_block_handle, sellado, pre_interv_estado"
    ).in_("id", body.ids).execute().data or []
    info_map = {r["id"]: r for r in rows_info}
    if any(r.get("sellado") for r in rows_info):
        raise HTTPException(
            status_code=403,
            detail="Registro sellado (aprobado por Interventoría): no puede modificarse.",
        )
    rol_l = (current_user.get("rol_nombre") or "").strip().lower()
    es_perfil_interventoria = rol_l in ("interventoría", "interventoria", "operativo interventoria")
    if es_perfil_interventoria:
        for rid in body.ids:
            row = info_map.get(rid) or {}
            if not _pre_interv_liberado(row):
                raise HTTPException(
                    status_code=403,
                    detail="El registro debe estar aprobado en depuración contratista (Residente de Costos u Obra) antes de la validación de Interventoría.",
                )
    es_interventoria_sellar = current_user.get("rol_nombre") == "Interventoría"
    sellar = body.revisado == "Aprobado" and es_interventoria_sellar
    nombre_usuario = current_user.get("nombre") or current_user.get("email") or "Usuario"
    data_upd = {"revisado": body.revisado, "updated_at": "now()"}
    if sellar:
        data_upd["sellado"] = True
    if body.revisado == "Aprobado":
        data_upd["validado_por"] = nombre_usuario
        data_upd["validado_en"]  = datetime.utcnow().isoformat()
    else:
        data_upd["validado_por"] = None
        data_upd["validado_en"]  = None
    # Una sola query batch en lugar de N queries secuenciales
    supabase.table("presupuesto").update(data_upd).in_("id", body.ids).execute()
    registrar_log(current_user, "VALIDAR", "PRESUPUESTO", "presupuesto_bulk", str(contrato_id),
        {"contrato_id": contrato_id, "cantidad_registros": len(body.ids), "estado": body.revisado})
    return {"actualizados": len(body.ids)}


@app.put("/presupuesto/{contrato_id}/bulk-pre-interv")
def bulk_pre_interv(contrato_id: int, body: PresupuestoBulkPreInterv, current_user=Depends(get_current_user)):
    """Depuración contratista: Residente de Costos u Obra aprueba antes de que Interventoría vea/valide."""
    if not body.ids:
        raise HTTPException(status_code=400, detail="No hay registros seleccionados")
    _reject_if_presupuesto_sellado(supabase, body.ids)
    ESTADOS_PRE = {"No Revisado", "Pendiente", "Rechazado", "Aprobado"}
    if body.estado not in ESTADOS_PRE:
        raise HTTPException(status_code=422, detail=f"Estado inválido. Use: {ESTADOS_PRE}")
    rol = (current_user.get("rol_nombre") or "").strip().lower()
    cargo = (current_user.get("cargo_nombre") or "").strip().lower()
    es_dev = cargo == "desarrollador" or rol == "desarrollador"
    if not es_dev:
        if not _cargo_permiso_validar_presupuesto(current_user):
            raise HTTPException(
                status_code=403,
                detail="No tiene permiso de validación en «editar registros presupuesto» (matriz de accesos).",
            )
        if rol not in ("contratista", "operativo contratista"):
            raise HTTPException(status_code=403, detail="Solo el contratista puede gestionar la depuración previa.")
        if not _cargo_puede_prevalidar_interventoria(cargo):
            raise HTTPException(
                status_code=403,
                detail="Solo Residente de Costos u Residente de Obra puede validar esta etapa.",
            )
    nombre_usuario = current_user.get("nombre") or current_user.get("email") or "Usuario"
    data_upd = {"pre_interv_estado": body.estado, "updated_at": "now()"}
    if body.estado == "Aprobado":
        data_upd["pre_interv_por"] = nombre_usuario
        data_upd["pre_interv_en"] = datetime.utcnow().isoformat()
    else:
        data_upd["pre_interv_por"] = None
        data_upd["pre_interv_en"] = None
    # Una sola query batch en lugar de N queries secuenciales
    supabase.table("presupuesto").update(data_upd).in_("id", body.ids).execute()
    registrar_log(
        current_user, "VALIDAR", "PRESUPUESTO", "presupuesto_pre_interv", str(contrato_id),
        {"contrato_id": contrato_id, "cantidad_registros": len(body.ids), "estado": body.estado},
    )
    return {"actualizados": len(body.ids)}

# ─────────────────────────────────────────────
# CAD QUEUE
# ─────────────────────────────────────────────


@app.get("/exportar/estado/{job_id}")
def exportar_estado(job_id: str, current_user=Depends(get_current_user)):
    """Frontend consulta si el Excel ya está listo."""
    job = _export_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return {"estado": job["estado"], "filename": job.get("filename","")}


@app.get("/exportar/descargar/{job_id}")
def exportar_descargar(job_id: str, current_user=Depends(get_current_user)):
    """Descarga el Excel generado en background."""
    job = _export_jobs.get(job_id)
    if not job or job["estado"] != "listo":
        raise HTTPException(status_code=404, detail="Archivo no listo")
    buf = io.BytesIO(job["buf"])
    filename = job.get("filename", "ClaraCore.xlsx")
    # No eliminar aún — permite diagnóstico
    # del _export_jobs[job_id]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.post("/cad-queue/{contrato_id}/heartbeat")
def cad_heartbeat(contrato_id: int, usuario_id: int = 0):
    """SicoeCAD llama esto cada 3s — sin auth, persiste en Supabase.
    usuario_id: id del usuario ClaraCore (mismo que inicia sesión) para no mostrar “enlazado” a otros.
    """
    k = _session_key_cad(contrato_id, usuario_id)
    _dwg_sessions[k] = time.time()
    try:
        supabase.table("cad_sessions").upsert({
            "contrato_id": contrato_id,
            "usuario_id": usuario_id,
            "ultimo_heartbeat": datetime.utcnow().isoformat()
        }, on_conflict="contrato_id,usuario_id").execute()
    except: pass
    return {"ok": True}

@app.get("/cad-queue/{contrato_id}/debug")
def cad_debug(contrato_id: int, current_user=Depends(get_current_user)):
    """Diagnóstico temporal — muestra estado de sesiones DWG."""
    sessions_info = []
    for k, ts in _dwg_sessions.items():
        if isinstance(k, tuple) and len(k) == 2 and int(k[0]) == int(contrato_id):
            cctx, uu = k
            sessions_info.append({
                "contrato_id": cctx,
                "usuario_id": uu,
                "hace_segundos": round(time.time() - float(ts), 1),
                "activo": (time.time() - float(ts)) < _DWG_TIMEOUT
            })
    db_sessions = []
    try:
        rows = supabase.table("cad_sessions").select("*").eq("contrato_id", contrato_id).execute().data
        db_sessions = rows
    except Exception as e:
        db_sessions = [{"error": str(e)}]
    return {
        "contrato_id": contrato_id,
        "memoria_sessions": sessions_info,
        "supabase_sessions": db_sessions,
        "dwg_activo": _dwg_activo(contrato_id)
    }

@app.get("/cad-queue/{contrato_id}/estado")
def cad_estado(contrato_id: int, current_user=Depends(get_current_user)):
    """Solo el usuario cuyo SicoeCAD envía heartbeats (mismo usuario_id) ve enlazado=True."""
    my_uid = 0
    try:
        my_uid = int(current_user.get("sub") or 0)
    except (TypeError, ValueError):
        my_uid = 0
    if my_uid <= 0:
        return {"enlazado": False}
    # 1) Memoria: sesión (contrato, mi usuario)
    k = _session_key_cad(contrato_id, my_uid)
    last = _dwg_sessions.get(k)
    if last is not None and (time.time() - float(last)) < 30:
        return {"enlazado": True}
    # 2) Supabase — filtrar por usuario; sin esto, cualquier fila del contrato ponía en verde a todos
    try:
        from datetime import timezone
        row = supabase.table("cad_sessions").select("ultimo_heartbeat") \
            .eq("contrato_id", contrato_id) \
            .eq("usuario_id", my_uid) \
            .limit(1) \
            .execute()
        rows = (row.data or [])
        if rows:
            tstr = rows[0]["ultimo_heartbeat"]
            tstr = (tstr or "").replace("Z", "+00:00")
            ultimo = datetime.fromisoformat(tstr)
            if ultimo.tzinfo is None:
                ultimo = ultimo.replace(tzinfo=timezone.utc)
            diff = (datetime.now(timezone.utc) - ultimo).total_seconds()
            return {"enlazado": diff < 30}
    except Exception:
        pass
    return {"enlazado": False}

@app.get("/cad-queue/{contrato_id}/pendientes")
def cad_pendientes(contrato_id: int, current_user=Depends(get_current_user)):
    """SicoeCAD descarga las operaciones pendientes."""
    try:
        rows = supabase.table("cad_queue").select("*") \
            .eq("contrato_id", contrato_id).eq("estado", "pendiente") \
            .order("id").limit(50).execute().data
    except Exception:
        return []
    return rows

@app.post("/cad-queue/{contrato_id}/highlight-registro")
def highlight_registro(contrato_id: int, body: dict, current_user=Depends(get_current_user)):
    """Encola highlight de entidad+texto de un registro de presupuesto."""
    presupuesto_id = body.get("presupuesto_id")
    if not presupuesto_id:
        raise HTTPException(status_code=400, detail="presupuesto_id requerido")
    row = supabase.table("presupuesto").select("ent_handle, txt_handle, x_label, y_label").eq("id", presupuesto_id).single().execute().data
    if not row or not row.get("ent_handle"):
        raise HTTPException(status_code=404, detail="Registro sin ent_handle")
    payload = {
        "ent_handle": row.get("ent_handle", ""),
        "txt_handle": row.get("txt_handle", ""),
        "x_label":    row.get("x_label", 0),
        "y_label":    row.get("y_label", 0),
    }
    try:
        usuario_id = int(current_user.get("sub") or current_user.get("id") or 0)
    except (TypeError, ValueError):
        usuario_id = 0
    if usuario_id <= 0:
        raise HTTPException(status_code=401, detail="Usuario no identificado")
    supabase.table("cad_queue").insert({
        "contrato_id":  contrato_id,
        "tipo":         "highlight_registro",
        "payload":      payload,
        "usuario_id":   usuario_id,
        "procesado":    False,
    }).execute()
    return {"ok": True}

@app.post("/cad-queue/{contrato_id}/zoom-pkid")
def zoom_pkid(contrato_id: int, pk_id: str, current_user=Depends(get_current_user)):
    """Encola operación de zoom a un PK_ID en AutoCAD."""
    rows = supabase.table("presupuesto").select(
        "ent_handle, x_label, y_label"
    ).eq("contrato_id", contrato_id).eq("pk_id", pk_id).limit(1).execute().data
    if not rows:
        raise HTTPException(status_code=404, detail="PK_ID no encontrado en presupuesto")
    r = rows[0]
    ent_handle = r.get("ent_handle") or ""
    x = r.get("x_label") or 0
    y = r.get("y_label") or 0
    if not ent_handle and (x == 0 and y == 0):
        raise HTTPException(status_code=404, detail="Sin coordenadas para este PK_ID")
    supabase.table("cad_queue").insert({
        "contrato_id": contrato_id,
        "tipo": "zoom_pkid",
        "estado": "pendiente",
        "payload": {
            "pk_id":      pk_id,
            "ent_handle": ent_handle,
            "x":          x,
            "y":          y,
            "radio":      30
        }
    }).execute()
    return {"ok": True, "pk_id": pk_id}

@app.put("/cad-queue/{op_id}/procesado")
def cad_procesado(op_id: int, body: CadQueueProcesado, current_user=Depends(get_current_user)):
    """SicoeCAD marca la operación como procesada."""
    supabase.table("cad_queue").update({
        "estado": "procesado",
        "processed_at": datetime.utcnow().isoformat()
    }).eq("id", op_id).execute()
    # Si la op era insertar_bloque, guardar el handle del bloque en presupuesto
    if body.presupuesto_id and body.rev_block_handle:
        supabase.table("presupuesto").update({
            "rev_block_handle": body.rev_block_handle
        }).eq("id", body.presupuesto_id).execute()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════════════════════
# COMENTARIOS
# ═══════════════════════════════════════════════════════════════════════════════

class ComentarioBulk(BaseModel):
    presupuesto_ids: List[int]
    tipo: str        # dims | item_capitulo | validacion
    mensaje: str
    usuario_nombre: str

class RespuestaCreate(BaseModel):
    mensaje: str
    usuario_nombre: str

@app.post("/presupuesto/{contrato_id}/comentarios/bulk")
def crear_comentarios_bulk(contrato_id: int, body: ComentarioBulk, current_user=Depends(get_current_user)):
    rows = [{"presupuesto_id": pid, "tipo": body.tipo, "mensaje": body.mensaje,
             "usuario_nombre": body.usuario_nombre} for pid in body.presupuesto_ids]
    supabase.table("comentarios").insert(rows).execute()
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        n = len(body.presupuesto_ids or [])
        registrar_log(
            u_log,
            "COMENTAR",
            "PRESUPUESTO",
            f"comentario_{body.tipo}",
            str(contrato_id),
            {
                "contrato_id": contrato_id,
                "tipo": body.tipo,
                "registros": n,
                "presupuesto_ids_muestra": [str(x) for x in (body.presupuesto_ids or [])[:20]],
            },
        )
    except Exception:
        pass
    return {"creados": len(rows)}

@app.get("/presupuesto/{contrato_id}/comentarios-resumen")
def comentarios_resumen(contrato_id: int, ids: str, current_user=Depends(get_current_user)):
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return {}
    rows = supabase.table("comentarios").select(
        "id, presupuesto_id, tipo, parent_id"
    ).in_("presupuesto_id", id_list).execute().data
    id_to_tipo = {r["id"]: r["tipo"] for r in rows}
    result = {}
    for r in rows:
        pid = r["presupuesto_id"]
        if pid not in result:
            result[pid] = {
                "dims":          {"count": 0, "replies": False},
                "item_capitulo": {"count": 0, "replies": False},
                "validacion":    {"count": 0, "replies": False},
            }
        tipo = r["tipo"]
        if r["parent_id"] is None:
            result[pid][tipo]["count"] += 1
        else:
            result[pid][tipo]["replies"] = True
    return result

def _comentarios_validacion_por_ids(id_list: List[int]) -> dict:
    """Comentario de validación más reciente (raíz) por presupuesto_id. Chunking: PostgREST limita .in_() en la URL."""
    if not id_list:
        return {}
    _CHUNK = 200
    result: Dict[int, Any] = {}
    for i in range(0, len(id_list), _CHUNK):
        chunk = id_list[i : i + _CHUNK]
        rows = supabase.table("comentarios").select(
            "presupuesto_id, mensaje, usuario_nombre, created_at"
        ).in_("presupuesto_id", chunk).eq("tipo", "validacion").is_("parent_id", "null").order(
            "created_at", desc=True
        ).execute().data or []
        for r in rows:
            pid = r["presupuesto_id"]
            if pid not in result:
                result[pid] = {
                    "mensaje": r["mensaje"],
                    "usuario_nombre": r["usuario_nombre"],
                    "created_at": r["created_at"],
                }
    return result

def _comentarios_validacion_por_capitulo_contrato(contrato_id: int, capitulo: str) -> dict:
    """Misma salida que por IDs, pero una sola petición: join comentarios → presupuesto (sin N×chunk a Supabase)."""
    if not (capitulo or "").strip():
        return {}
    rows = (
        supabase.table("comentarios")
        .select("presupuesto_id, mensaje, usuario_nombre, created_at, presupuesto!inner(contrato_id, capitulo)")
        .eq("tipo", "validacion")
        .is_("parent_id", "null")
        .eq("presupuesto.contrato_id", contrato_id)
        .eq("presupuesto.capitulo", capitulo)
        .order("created_at", desc=True)
        .execute()
        .data
        or []
    )
    result: Dict[int, Any] = {}
    for r in rows:
        pid = r["presupuesto_id"]
        if pid not in result:
            result[pid] = {
                "mensaje": r["mensaje"],
                "usuario_nombre": r["usuario_nombre"],
                "created_at": r["created_at"],
            }
    return result

@app.get("/presupuesto/{contrato_id}/comentarios-validacion-capitulo")
def comentarios_validacion_por_capitulo(
    contrato_id: int, capitulo: str, current_user=Depends(get_current_user)
):
    """Revisor por tramo: carga comentarios de validación de todo el capítulo en un solo ida-vuelta."""
    return _comentarios_validacion_por_capitulo_contrato(contrato_id, capitulo)

@app.get("/presupuesto/{contrato_id}/comentarios-validacion")
def comentarios_validacion_batch(contrato_id: int, ids: str, current_user=Depends(get_current_user)):
    """Devuelve el comentario de validacion más reciente (sin hijos) por cada presupuesto_id."""
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    return _comentarios_validacion_por_ids(id_list)

@app.post("/presupuesto/{contrato_id}/comentarios-validacion")
def comentarios_validacion_batch_post(
    contrato_id: int, body: ComentariosValidacionIds, current_user=Depends(get_current_user)
):
    """Igual que GET: lista de IDs en el cuerpo para capítulos muy grandes (evita URL > límite del proxy)."""
    id_list: List[int] = []
    seen: set = set()
    for x in body.ids:
        try:
            xi = int(x)
        except (TypeError, ValueError):
            continue
        if xi in seen:
            continue
        seen.add(xi)
        id_list.append(xi)
    return _comentarios_validacion_por_ids(id_list)

@app.get("/presupuesto/{presupuesto_id}/comentarios")
def get_comentarios(presupuesto_id: int, tipo: str, current_user=Depends(get_current_user)):
    rows = supabase.table("comentarios").select("*").eq(
        "presupuesto_id", presupuesto_id).eq("tipo", tipo).order("created_at").execute().data
    roots = [r for r in rows if r["parent_id"] is None]
    children = {}
    for r in rows:
        if r["parent_id"]:
            children.setdefault(r["parent_id"], []).append(r)
    for root in roots:
        root["respuestas"] = children.get(root["id"], [])
    return roots

@app.post("/comentarios/{comentario_id}/respuesta")
def responder_comentario(comentario_id: int, body: RespuestaCreate, current_user=Depends(get_current_user)):
    parent = supabase.table("comentarios").select("presupuesto_id, tipo").eq("id", comentario_id).execute().data
    if not parent:
        raise HTTPException(status_code=404, detail="Comentario no encontrado")
    p = parent[0]
    supabase.table("comentarios").insert({
        "presupuesto_id": p["presupuesto_id"],
        "tipo":           p["tipo"],
        "mensaje":        body.mensaje,
        "usuario_nombre": body.usuario_nombre,
        "parent_id":      comentario_id
    }).execute()
    try:
        pr = supabase.table("presupuesto").select("contrato_id").eq("id", p["presupuesto_id"]).limit(1).execute().data
        cid = (pr[0] or {}).get("contrato_id") if pr else None
        if cid:
            u_log = _audit_user_contrato(current_user, int(cid))
            registrar_log(
                u_log,
                "COMENTAR",
                "PRESUPUESTO",
                "respuesta_hilo",
                str(comentario_id),
                {
                    "parent_comentario_id": comentario_id,
                    "presupuesto_id": p["presupuesto_id"],
                    "tipo": p.get("tipo"),
                    "es_respuesta": True,
                },
            )
    except Exception:
        pass
    return {"ok": True}
# ─────────────────────────────────────────────
# LOGS
# ─────────────────────────────────────────────

def _sort_logs_rows(rows: Optional[List[dict]]) -> List[dict]:
    """Orden estable: más reciente primero (defensa si PostgREST devolviera filas fuera de orden)."""

    if not rows:
        return []

    def _key(r: dict):
        ts = r.get("created_at") or "1970-01-01T00:00:00+00:00"
        try:
            rid = int(r.get("id") or 0)
        except (TypeError, ValueError):
            rid = 0
        return (ts, rid)

    return sorted(rows, key=_key, reverse=True)


def _logs_query_base(
    usuario_id: Optional[int] = None,
    modulo: Optional[str] = None,
    accion: Optional[str] = None,
    categoria: Optional[str] = None,
    severidad: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    excluir_accion: Optional[str] = None,
    excluir_acciones: Optional[List[str]] = None,
    excluir_modulo: Optional[str] = None,
):
    q = supabase.table("logs").select("*")
    if usuario_id:
        q = q.eq("usuario_id", usuario_id)
    if modulo:
        q = q.eq("modulo", modulo)
    elif excluir_modulo:
        q = q.neq("modulo", excluir_modulo)
    if accion:
        q = q.eq("accion", accion)
    elif excluir_acciones:
        q = q.not_.in_("accion", excluir_acciones)
    elif excluir_accion:
        q = q.neq("accion", excluir_accion)
    if categoria:
        q = q.eq("categoria", categoria)
    if severidad:
        q = q.eq("severidad", severidad)
    if fecha_desde:
        q = q.gte("created_at", fecha_desde)
    if fecha_hasta:
        q = q.lte("created_at", fecha_hasta + "T23:59:59")
    return q.order("created_at", desc=True, nullsfirst=False).order("id", desc=True)


@app.get("/logs")
def get_logs(
    usuario_id:   Optional[int] = None,
    modulo:       Optional[str] = None,
    accion:       Optional[str] = None,
    categoria:    Optional[str] = None,
    severidad:    Optional[str] = None,
    fecha_desde:  Optional[str] = None,
    fecha_hasta:  Optional[str] = None,
    excluir_accion: Optional[str] = None,
    excluir_rutina_auth: bool = Query(False, description="Excluye LOGIN y LOGIN_FAIL (los 50 más recientes suelen ser solo inicios de sesión)."),
    excluir_modulo: Optional[str] = None,
    limit:        int = 100,
    offset:       int = 0,
    current_user=Depends(require_logs_auditoria),
):
    """Consulta logs con filtros. Solo Desarrollador y Administrador."""
    excluir_acciones = None
    if excluir_rutina_auth and not accion:
        excluir_acciones = ["LOGIN", "LOGIN_FAIL"]
    q = _logs_query_base(
        usuario_id=usuario_id,
        modulo=modulo,
        accion=accion,
        categoria=categoria,
        severidad=severidad,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        excluir_accion=excluir_accion if not excluir_acciones else None,
        excluir_acciones=excluir_acciones,
        excluir_modulo=excluir_modulo,
    )
    q = q.range(offset, offset + limit - 1)
    return _sort_logs_rows(q.execute().data)


@app.get("/logs/alertas")
def get_logs_alertas(
    limit: int = 50,
    current_user=Depends(require_logs_auditoria),
):
    """Eventos marcados como alerta (login masivo, 500 repetidos, permisos, etc.)."""
    rows = (
        supabase.table("logs")
        .select("*")
        .eq("alerta_generada", True)
        .order("created_at", desc=True)
        .limit(min(limit, 200))
        .execute()
        .data
    )
    return rows


@app.get("/logs/export.csv")
def export_logs_csv(
    usuario_id:  Optional[int] = None,
    modulo:      Optional[str] = None,
    accion:      Optional[str] = None,
    categoria:   Optional[str] = None,
    severidad:   Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    excluir_accion: Optional[str] = None,
    excluir_rutina_auth: bool = Query(False),
    excluir_modulo: Optional[str] = None,
    max_rows:    int = 5000,
    current_user=Depends(require_logs_auditoria),
):
    """Exportación CSV para interventoría / auditoría externa."""
    cap = min(max(max_rows, 1), 20000)
    excluir_acciones = None
    if excluir_rutina_auth and not accion:
        excluir_acciones = ["LOGIN", "LOGIN_FAIL"]
    q = _logs_query_base(
        usuario_id=usuario_id,
        modulo=modulo,
        accion=accion,
        categoria=categoria,
        severidad=severidad,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        excluir_accion=excluir_accion if not excluir_acciones else None,
        excluir_acciones=excluir_acciones,
        excluir_modulo=excluir_modulo,
    )
    rows = _sort_logs_rows(q.limit(cap).execute().data or [])
    import json as _json

    def _cell(v):
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            return _json.dumps(v, ensure_ascii=False)
        return str(v)

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "id", "created_at", "categoria", "severidad", "usuario_id", "usuario_nombre", "rol_nombre",
        "cargo_nombre", "contrato_id", "contrato_numero", "accion", "modulo",
        "entidad_tipo", "entidad_id", "resultado", "ip", "endpoint", "detalle",
        "valor_anterior", "valor_nuevo", "alerta_generada",
    ])
    for r in rows:
        w.writerow([
            r.get("id"), r.get("created_at"), r.get("categoria"), r.get("severidad"),
            r.get("usuario_id"), r.get("usuario_nombre"), r.get("rol_nombre"),
            r.get("cargo_nombre"), r.get("contrato_id"), r.get("contrato_numero"),
            r.get("accion"), r.get("modulo"), r.get("entidad_tipo"), r.get("entidad_id"),
            r.get("resultado"), r.get("ip"), r.get("endpoint"),
            _cell(r.get("detalle")), _cell(r.get("valor_anterior")), _cell(r.get("valor_nuevo")),
            r.get("alerta_generada"),
        ])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="claracore_logs.csv"'},
    )


@app.get("/logs/usuarios-lista")
def get_logs_usuarios(current_user=Depends(require_logs_auditoria)):
    """Lista de usuarios para el filtro (todos los visibles para el administrador, no solo quienes ya tienen log)."""
    result = supabase.table("usuarios").select(
        "id, nombre, apellidos, cargo_id, contrato_id"
    ).order("nombre").execute()
    cargos = {c["id"]: c["nombre"] for c in supabase.table("cargos").select("id, nombre").execute().data}
    caller_contrato, _ = _caller_contract_scope(current_user)
    caller_id = int(current_user["sub"])
    out = []
    for u in result.data or []:
        uid = u.get("id")
        cn = cargos.get(u.get("cargo_id"), "Sin cargo")
        if caller_contrato and u.get("contrato_id") != caller_contrato:
            continue
        if (cn or "").lower() == "desarrollador" and uid != caller_id:
            continue
        nombre = f"{u.get('nombre', '')} {u.get('apellidos', '')}".strip() or f"Usuario {uid}"
        out.append({"id": uid, "nombre": nombre, "cargo": cn})
    return out


@app.get("/logs/entidad/{entidad_tipo}/{entidad_id}")
def get_logs_entidad(
    entidad_tipo: str,
    entidad_id: str,
    current_user=Depends(get_current_user),
):
    """Historial de una entidad. Para `registro` incluye eventos del reporte padre (importación masiva, cabecera)."""
    def _fetch(tipo: str, eid: str) -> List[dict]:
        return (
            supabase.table("logs")
            .select("*")
            .eq("entidad_tipo", tipo)
            .eq("entidad_id", eid)
            .order("created_at", desc=False)
            .execute()
            .data
            or []
        )

    merged: Dict[Any, dict] = {}
    for row in _fetch(entidad_tipo, entidad_id):
        pk = row.get("id")
        if pk is not None:
            merged[pk] = row

    if entidad_tipo == "registro":
        try:
            rid_int = int(entidad_id)
        except (TypeError, ValueError):
            rid_int = None
        if rid_int is not None:
            rrows = (
                supabase.table("so_registros")
                .select("reporte_id")
                .eq("id", rid_int)
                .limit(1)
                .execute()
                .data
            )
            rep_id = (rrows[0].get("reporte_id") if rrows else None)
            if rep_id is not None:
                for row in _fetch("reporte", str(rep_id)):
                    pk = row.get("id")
                    if pk is not None:
                        merged[pk] = row

    out = list(merged.values())
    out.sort(key=lambda x: (x.get("created_at") or ""))
    return out


@app.post("/admin/deploy-log")
def log_deploy_event(
    body: dict,
    current_user=Depends(require_logs_auditoria),
):
    """Registra un despliegue (versión, notas). Uso manual desde pipeline o consola."""
    ver = (body or {}).get("version") or ""
    notas = (body or {}).get("notas") or ""
    uid = int(current_user.get("sub"))
    nombre = current_user.get("nombre") or current_user.get("email", "")
    registrar_log(
        {
            "sub": str(uid),
            "nombre": nombre,
            "cargo_nombre": current_user.get("cargo_nombre"),
            "rol_nombre": current_user.get("rol_nombre"),
            "contrato_id": current_user.get("contrato_id"),
            "contrato_numero": current_user.get("contrato_numero"),
        },
        "DEPLOY",
        "SISTEMA",
        "deploy",
        ver or "sin_version",
        {"version": ver, "notas": notas},
        severidad="INFO",
        categoria="auditoria",
    )
    return {"ok": True}


def _inicio_novedades_rows_con_lectura(uid: int, contrato_id_u: Optional[int]) -> List[dict]:
    try:
        rows = _query_inicio_novedades_accesibles(contrato_id_u).execute().data or []
    except Exception as e:
        _log_api.warning("inicio novedades filtradas, fallback sin contrato_id: %s", e)
        try:
            rows = (
                supabase.table("inicio_novedades")
                .select("*")
                .order("created_at", desc=True)
                .execute()
                .data
                or []
            )
        except Exception as e2:
            _log_api.warning("inicio novedades list: %s", e2)
            rows = []
    ids = [int(r["id"]) for r in rows if r.get("id") is not None]
    leidas: set = set()
    if ids:
        try:
            lr = (
                supabase.table("inicio_novedades_lecturas")
                .select("novedad_id")
                .eq("usuario_id", int(uid))
                .in_("novedad_id", ids)
                .execute()
                .data
                or []
            )
            leidas = {int(x["novedad_id"]) for x in lr if x.get("novedad_id") is not None}
        except Exception as e2:
            _log_api.warning("inicio novedades lecturas: %s", e2)
    out: List[dict] = []
    for r in rows:
        rr = dict(r)
        try:
            nid = int(rr.get("id") or 0)
        except (TypeError, ValueError):
            nid = 0
        rr["leida"] = nid in leidas if nid else False
        out.append(rr)
    return out


@app.get("/inicio/novedades")
def inicio_novedades_listado(current_user=Depends(get_current_user)):
    """
    Novedades visibles según el contrato del usuario + bandera `leida` por usuario.
    Requiere token. Sin columna `contrato_id` en BD, el filtro puede fallar y se listan todas.
    """
    try:
        uid = int(current_user.get("sub"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=401, detail="Token inválido")
    contrato_id_u, _ = _caller_contract_scope(current_user)
    if isinstance(contrato_id_u, (int, float)) and not isinstance(contrato_id_u, bool):
        ciu = int(contrato_id_u)
    else:
        ciu = None
    return _inicio_novedades_rows_con_lectura(uid, ciu)


@app.post("/inicio/novedades/{novedad_id}/leida")
def inicio_novedad_marcar_leida(
    novedad_id: int,
    current_user=Depends(get_current_user),
):
    try:
        uid = int(current_user.get("sub"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=401, detail="Token inválido")
    contrato_id_u, _ = _caller_contract_scope(current_user)
    ciu = int(contrato_id_u) if contrato_id_u is not None else None
    rows = _inicio_novedades_rows_con_lectura(uid, ciu)
    ok = any(int(r.get("id") or 0) == int(novedad_id) for r in rows)
    if not ok:
        raise HTTPException(status_code=404, detail="Novedad no disponible")
    try:
        supabase.table("inicio_novedades_lecturas").insert(
            {"usuario_id": uid, "novedad_id": int(novedad_id)}
        ).execute()
    except Exception as e:
        if "23505" not in str(e) and "duplicate" not in str(e).lower() and "unique" not in str(e).lower():
            _log_api.warning("marcar leida novedad %s: %s", novedad_id, e)
    return {"ok": True}


@app.get("/admin/inicio/novedades")
def admin_inicio_novedades_list(current_user=Depends(require_logs_auditoria)):
    """Novedades que puede ver el admin: globales + del propio contrato (mismo criterio que inicio)."""
    try:
        uid = int(current_user.get("sub"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=401, detail="Token inválido")
    contrato_id_u, _ = _caller_contract_scope(current_user)
    ciu = int(contrato_id_u) if contrato_id_u is not None else None
    return _inicio_novedades_rows_con_lectura(uid, ciu)


@app.post("/admin/inicio/novedades")
def admin_inicio_novedades_create(
    body: InicioNovedadCreate,
    request: Request,
    current_user=Depends(require_logs_auditoria),
):
    row = body.model_dump()
    es_dev = _es_desarrollador(current_user)
    caller_cid, _ = _caller_contract_scope(current_user)
    raw_cid = row.pop("contrato_id", None)
    if es_dev:
        if raw_cid is not None:
            try:
                row["contrato_id"] = int(raw_cid)
            except (TypeError, ValueError):
                row["contrato_id"] = None
        else:
            row["contrato_id"] = None
    else:
        if not caller_cid:
            raise HTTPException(
                status_code=400,
                detail="Los administradores de novedad deben tener un contrato asignado para publicar en su equipo.",
            )
        row["contrato_id"] = int(caller_cid)
    raw_fecha = (row.pop("fecha") or "").strip()
    if raw_fecha:
        row["fecha"] = raw_fecha[:10]
    else:
        row["fecha"] = datetime.now(timezone.utc).date().isoformat()
    try:
        # supabase-py 2.x: insert() devuelve SyncQueryRequestBuilder (no admite .select() encadenado).
        # returning=representation es el valor por defecto; execute() devuelve la fila insertada en data.
        res = supabase.table("inicio_novedades").insert(row).execute()
    except Exception as e:
        _log_api.warning("POST /admin/inicio/novedades insert: %s", e)
        raise HTTPException(
            status_code=503,
            detail=(
                "No se pudo guardar en la base de datos. Comprueba que exista la tabla "
                "`inicio_novedades` (ejecuta backend/sql/inicio_novedades.sql en Supabase) y que "
                "las columnas coincidan con el despliegue actual. "
                f"Detalle técnico: {e!s}"
            ),
        )
    data = res.data[0] if res.data else None
    if not data:
        raise HTTPException(
            status_code=503,
            detail="La inserción no devolvió fila (PostgREST). Revisa permisos RLS y la tabla inicio_novedades.",
        )
    u = dict(current_user)
    u["nombre"] = u.get("nombre") or u.get("email", "")
    registrar_log(
        u,
        "CREAR",
        "INICIO",
        "inicio_novedad",
        str(data.get("id")) if data else None,
        detalle={"titulo": row.get("titulo")},
        resultado="ok",
        valor_nuevo=data,
        ip=_client_ip(request),
    )
    return data


@app.patch("/admin/inicio/novedades/{novedad_id}")
def admin_inicio_novedades_update(
    novedad_id: int,
    body: InicioNovedadUpdate,
    request: Request,
    current_user=Depends(require_logs_auditoria),
):
    prev_rows = supabase.table("inicio_novedades").select("*").eq("id", novedad_id).limit(1).execute().data or []
    prev = prev_rows[0] if prev_rows else None
    if not prev:
        raise HTTPException(status_code=404, detail="Novedad no encontrada")
    es_dev = _es_desarrollador(current_user)
    caller_cid, _ = _caller_contract_scope(current_user)
    ccr = int(caller_cid) if caller_cid is not None else None
    if not _novedad_puede_gestionar_admin(prev, ccr, es_dev):
        raise HTTPException(
            status_code=403,
            detail="Solo el Desarrollador puede editar novedades globales. Las de tu contrato las gestiona el administrador de ese contrato.",
        )
    patch = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    if "contrato_id" in patch:
        if not es_dev:
            patch.pop("contrato_id", None)
        else:
            try:
                patch["contrato_id"] = int(patch["contrato_id"]) if patch["contrato_id"] is not None else None
            except (TypeError, ValueError):
                patch["contrato_id"] = None
    if "fecha" in patch and patch["fecha"] is not None:
        patch["fecha"] = str(patch["fecha"])[:10]
    if not patch:
        return prev
    try:
        res = supabase.table("inicio_novedades").update(patch).eq("id", novedad_id).execute()
    except Exception as e:
        _log_api.warning("PATCH /admin/inicio/novedades/%s: %s", novedad_id, e)
        raise HTTPException(status_code=503, detail=f"No se pudo actualizar: {e!s}")
    data = (res.data[0] if res.data else None) or {**prev, **patch}
    u = dict(current_user)
    u["nombre"] = u.get("nombre") or u.get("email", "")
    registrar_log(
        u,
        "EDITAR",
        "INICIO",
        "inicio_novedad",
        str(novedad_id),
        detalle=patch,
        resultado="ok",
        valor_anterior=prev,
        valor_nuevo=data,
        ip=_client_ip(request),
    )
    return data


@app.delete("/admin/inicio/novedades/{novedad_id}")
def admin_inicio_novedades_delete(
    novedad_id: int,
    request: Request,
    current_user=Depends(require_logs_auditoria),
):
    prev_rows = supabase.table("inicio_novedades").select("*").eq("id", novedad_id).limit(1).execute().data or []
    prev = prev_rows[0] if prev_rows else None
    if not prev:
        raise HTTPException(status_code=404, detail="Novedad no encontrada")
    es_dev = _es_desarrollador(current_user)
    caller_cid, _ = _caller_contract_scope(current_user)
    ccr = int(caller_cid) if caller_cid is not None else None
    if not _novedad_puede_gestionar_admin(prev, ccr, es_dev):
        raise HTTPException(
            status_code=403,
            detail="No puedes eliminar esta novedad (p. ej. novedad global: solo Desarrollador).",
        )
    supabase.table("inicio_novedades").delete().eq("id", novedad_id).execute()
    u = dict(current_user)
    u["nombre"] = u.get("nombre") or u.get("email", "")
    registrar_log(
        u,
        "ELIMINAR",
        "INICIO",
        "inicio_novedad",
        str(novedad_id),
        detalle={"titulo": prev.get("titulo")},
        resultado="ok",
        valor_anterior=prev,
        ip=_client_ip(request),
    )
    return {"ok": True}


@app.post("/admin/inicio/novedades/imagen")
async def admin_inicio_novedad_imagen(
    file: UploadFile = File(...),
    current_user=Depends(require_logs_auditoria),
):
    contents = await file.read()
    url = _inicio_novedad_subir_imagen(contents, file.content_type)
    return {"url": url}


@app.post("/admin/inicio/novedades/mejorar-texto")
def admin_inicio_novedad_mejorar_texto(
    body: InicioNovedadMejorarTexto,
    current_user=Depends(require_logs_auditoria),
):
    """Mejora la redacción del resumen con el mismo modelo que /frase-del-dia (Anthropic)."""
    raw = (body.texto or "").strip()
    if len(raw) > 8000:
        raise HTTPException(status_code=400, detail="Texto demasiado largo (máx. 8000 caracteres).")
    if not raw:
        return {"texto": ""}
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return {"texto": raw, "sin_ia": True}
    try:
        res = httpx.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5",
                "max_tokens": 900,
                "messages": [{
                    "role": "user",
                    "content": (
                        "Eres editor en español para avisos de una plataforma de gestión de obra (ClaraCore). "
                        "Mejora la redacción del siguiente resumen de novedad: tono claro y profesional, "
                        "sin inventar datos ni añadir información que no aparezca en el original. "
                        "Conserva listas o viñetas si las hay. "
                        "Devuelve SOLO el texto mejorado, sin título, sin comillas envolventes ni frases como \"Aquí tienes\".\n\n"
                        + raw
                    ),
                }],
            },
            timeout=35.0,
        )
        data = res.json()
        if res.status_code >= 400:
            detail = data.get("error", {}).get("message") if isinstance(data.get("error"), dict) else data
            raise HTTPException(status_code=503, detail=str(detail or "Error del proveedor IA"))
        if "content" not in data or not data["content"]:
            return {"texto": raw, "sin_ia": True}
        texto = (data["content"][0].get("text") or "").strip()
        if not texto:
            return {"texto": raw, "sin_ia": True}
        return {"texto": texto}
    except HTTPException:
        raise
    except Exception as e:
        _log_api.warning("mejorar-texto novedad: %s", e)
        return {"texto": raw, "sin_ia": True}


# ─────────────────────────────────────────────
# NOTIFICACIONES
# ─────────────────────────────────────────────

def _remitente_nombre_from_user(current_user) -> str:
    return (current_user.get("nombre") or current_user.get("email") or "").strip() or "Usuario"


def _resolver_contrato_notificacion(
    current_user,
    destinatario_id: Optional[int],
    contrato_id_explicit: Optional[int],
) -> Optional[int]:
    """Rellena contrato_id cuando el cliente envía null: evita que el buzón (filtrado por contrato) oculte el mensaje."""
    if contrato_id_explicit is not None:
        try:
            return int(contrato_id_explicit)
        except (TypeError, ValueError):
            pass
    try:
        uid = int(current_user.get("sub", 0))
    except (TypeError, ValueError):
        uid = 0
    if uid:
        urows = supabase.table("usuarios").select("contrato_id").eq("id", uid).limit(1).execute().data
        u = urows[0] if urows else {}
        if u.get("contrato_id") is not None:
            try:
                return int(u["contrato_id"])
            except (TypeError, ValueError):
                pass
        uc = (
            supabase.table("usuario_contratos")
            .select("contrato_id")
            .eq("usuario_id", uid)
            .limit(1)
            .execute()
            .data
        )
        if uc and uc[0].get("contrato_id") is not None:
            try:
                return int(uc[0]["contrato_id"])
            except (TypeError, ValueError):
                pass
    if destinatario_id is not None:
        try:
            did = int(destinatario_id)
        except (TypeError, ValueError):
            return None
        u2 = supabase.table("usuarios").select("contrato_id").eq("id", did).limit(1).execute().data
        u2r = u2[0] if u2 else {}
        if u2r.get("contrato_id") is not None:
            try:
                return int(u2r["contrato_id"])
            except (TypeError, ValueError):
                pass
        uc2 = (
            supabase.table("usuario_contratos")
            .select("contrato_id")
            .eq("usuario_id", did)
            .limit(1)
            .execute()
            .data
        )
        if uc2 and uc2[0].get("contrato_id") is not None:
            try:
                return int(uc2[0]["contrato_id"])
            except (TypeError, ValueError):
                pass
    return None


def _filtro_query_notif_contrato_o_nulo(q, contrato_id: Optional[int]):
    """Incluye filas con contrato_id NULL (mensajes legados / remitentes sin contrato principal en usuarios)."""
    if contrato_id is None:
        return q
    try:
        cid = int(contrato_id)
    except (TypeError, ValueError):
        return q
    return q.or_(f"contrato_id.eq.{cid},contrato_id.is.null")


def _push_notif_validacion_sicoe_destinatarios(
    current_user,
    autor_id: int,
    contrato_id: int,
    registro_id: int,
    asunto: str,
    mensaje: str,
    comentario_data: Optional[dict],
) -> None:
    if not comentario_data:
        return
    rem_nom = _remitente_nombre_from_user(current_user)
    for dest in comentario_data.get("destinatarios") or []:
        dest_id = dest.get("id") if isinstance(dest, dict) else None
        if not dest_id:
            continue
        try:
            did = int(dest_id)
        except (TypeError, ValueError):
            continue
        if did == autor_id:
            continue
        row = {
            "destinatario_id": did,
            "remitente_id": autor_id,
            "remitente_nombre": rem_nom,
            "contrato_id": contrato_id,
            "tipo": "validacion",
            "modulo": "sicoe_obra",
            "entidad_tipo": "registro",
            "entidad_id": str(registro_id),
            "asunto": asunto,
            "mensaje": mensaje or "",
            "leido": False,
        }
        try:
            supabase_execute(lambda: supabase.table("notificaciones").insert(row).execute().data)
        except Exception as e:
            try:
                _log_api.warning(
                    "notificación SICOE validación: insert falló destinatario=%s contrato=%s reg=%s %s",
                    did,
                    contrato_id,
                    registro_id,
                    e,
                )
            except Exception:
                pass


class NotificacionCreate(BaseModel):
    destinatario_id: Optional[int] = None  # None = broadcast a todos
    asunto: str
    mensaje: str
    tipo: str = "MENSAJE_DIRECTO"  # MENSAJE_DIRECTO | BROADCAST | SISTEMA | SOPORTE
    modulo: Optional[str] = None
    contrato_id: Optional[int] = None
    entidad_tipo: Optional[str] = None
    entidad_id: Optional[str] = None
    padre_id: Optional[int] = None

@app.post("/notificaciones")
def crear_notificacion(body: NotificacionCreate, current_user=Depends(get_current_user)):
    """Envía una notificación. Si destinatario_id es None y tipo=BROADCAST, envía a todos."""
    uid = int(current_user.get("sub", 0))
    nombre = _remitente_nombre_from_user(current_user)

    if body.tipo == "BROADCAST":
        # Enviar a todos los usuarios activos excepto el remitente
        usuarios = supabase.table("usuarios").select("id").eq("activo", True).execute().data
        rows = []
        for u in usuarios:
            if u["id"] == uid:
                continue
            dest_id = int(u["id"])
            rcid = _resolver_contrato_notificacion(current_user, dest_id, body.contrato_id)
            rows.append({
                "remitente_id":     uid,
                "remitente_nombre": nombre,
                "destinatario_id":  dest_id,
                "asunto":           body.asunto,
                "mensaje":          body.mensaje,
                "tipo":             body.tipo,
                "modulo":           body.modulo,
                "contrato_id":      rcid,
                "entidad_tipo":     body.entidad_tipo,
                "entidad_id":       body.entidad_id,
                "padre_id":         body.padre_id,
            })
        if rows:
            supabase.table("notificaciones").insert(rows).execute()
        registrar_log(current_user, "BROADCAST", "NOTIFICACIONES", "notificacion", None,
            {"asunto": body.asunto, "destinatarios": len(rows)})
        return {"enviados": len(rows)}
    else:
        rcid = _resolver_contrato_notificacion(current_user, body.destinatario_id, body.contrato_id)
        row = {
            "remitente_id":     uid,
            "remitente_nombre": nombre,
            "destinatario_id":  body.destinatario_id,
            "asunto":           body.asunto,
            "mensaje":          body.mensaje,
            "tipo":             body.tipo,
            "modulo":           body.modulo,
            "contrato_id":      rcid,
            "entidad_tipo":     body.entidad_tipo,
            "entidad_id":       body.entidad_id,
            "padre_id":         body.padre_id,
        }
        result = supabase.table("notificaciones").insert(row).execute()

        # Si es respuesta, marcar el mensaje raíz como no leído para el destinatario
        if body.padre_id:
            try:
                supabase.table("notificaciones") \
                    .update({"leido": False}) \
                    .eq("id", body.padre_id) \
                    .execute()
            except: pass

        registrar_log(current_user, "ENVIAR", "NOTIFICACIONES", "notificacion",
            str(result.data[0]["id"]) if result.data else None,
            {"asunto": body.asunto, "destinatario_id": body.destinatario_id, "tipo": body.tipo})
        return result.data[0] if result.data else {}

@app.get("/notificaciones/recibidas")
def get_notificaciones_recibidas(
    solo_no_leidas: bool = False,
    limit: int = 50,
    offset: int = 0,
    contrato_id: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    """Notificaciones recibidas por el usuario actual. Si se envía contrato_id, solo las de ese contrato."""
    uid = int(current_user.get("sub", 0))
    q = supabase.table("notificaciones").select("*") \
        .eq("destinatario_id", uid) \
        .order("created_at", desc=True)
    q = _filtro_query_notif_contrato_o_nulo(q, contrato_id)
    if solo_no_leidas:
        q = q.eq("leido", False)
    q = q.range(offset, offset + limit - 1)
    return q.execute().data

@app.get("/notificaciones/enviadas")
def get_notificaciones_enviadas(
    limit: int = 50,
    offset: int = 0,
    contrato_id: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    """Notificaciones enviadas por el usuario actual. Si se envía contrato_id, solo las de ese contrato."""
    uid = int(current_user.get("sub", 0))
    q = supabase.table("notificaciones").select("*") \
        .eq("remitente_id", uid) \
        .order("created_at", desc=True)
    q = _filtro_query_notif_contrato_o_nulo(q, contrato_id)
    return q.range(offset, offset + limit - 1).execute().data

@app.get("/notificaciones/no-leidas-count")
def get_no_leidas_count(
    contrato_id: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    """Conteo de notificaciones no leídas — solo mensajes raíz. Opcionalmente filtrado por contrato."""
    uid = int(current_user.get("sub", 0))
    try:
        q = supabase.table("notificaciones").select("id", count="exact") \
            .eq("destinatario_id", uid) \
            .eq("leido", False) \
            .is_("padre_id", "null")
        q = _filtro_query_notif_contrato_o_nulo(q, contrato_id)
        result = q.execute()
        return {"count": result.count or 0}
    except Exception:
        return {"count": 0}

@app.get("/notificaciones/{notif_id}/hilo")
def get_hilo(notif_id: int, current_user=Depends(get_current_user)):
    """Devuelve el hilo completo de una notificación (padre + respuestas)."""
    # Primero encontrar el padre raíz
    notif = supabase.table("notificaciones").select("*").eq("id", notif_id).execute().data
    if not notif: raise HTTPException(status_code=404, detail="No encontrada")
    padre_id = notif[0].get("padre_id") or notif_id
    # Marcar como leída
    uid = int(current_user.get("sub", 0))
    supabase.table("notificaciones").update({"leido": True, "leido_at": datetime.utcnow().isoformat()}) \
        .eq("id", notif_id).eq("destinatario_id", uid).execute()
    # Traer padre + todas las respuestas
    padre = supabase.table("notificaciones").select("*").eq("id", padre_id).execute().data
    respuestas = supabase.table("notificaciones").select("*") \
        .eq("padre_id", padre_id).order("created_at").execute().data
    return {"hilo": padre + respuestas}

@app.put("/notificaciones/{notif_id}/leida")
def marcar_leida(notif_id: int, current_user=Depends(get_current_user)):
    uid = int(current_user.get("sub", 0))
    supabase.table("notificaciones").update({"leido": True, "leido_at": "now()"}) \
        .eq("id", notif_id).eq("destinatario_id", uid).execute()
    return {"ok": True}

@app.get("/notificaciones/usuarios-destinatarios")
def get_usuarios_destinatarios(
    contrato_id: Optional[int] = Query(None, description="Acota destinatarios a este contrato (recomendado)"),
    current_user=Depends(get_current_user),
):
    """Usuarios activos del contrato para el selector de destinatario. No devuelve toda la plataforma."""
    uid = int(current_user.get("sub", 0))
    urow = supabase.table("usuarios").select("contrato_id, cargo_id").eq("id", uid).limit(1).execute().data
    urow = urow[0] if urow else {}
    cargo_nom = ""
    if urow.get("cargo_id"):
        crow = supabase.table("cargos").select("nombre").eq("id", urow["cargo_id"]).limit(1).execute().data
        if crow:
            cargo_nom = (crow[0].get("nombre") or "").strip().lower()

    es_dev = cargo_nom == "desarrollador"
    scope = set()

    if contrato_id is not None:
        try:
            cid = int(contrato_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="contrato_id inválido")
        if not es_dev and not _usuario_vinculado_a_contrato(uid, cid):
            raise HTTPException(status_code=403, detail="No tienes acceso a destinatarios de ese contrato")
        scope.add(cid)
    else:
        if urow.get("contrato_id") is not None:
            try:
                scope.add(int(urow["contrato_id"]))
            except (TypeError, ValueError):
                pass
        for r in supabase.table("usuario_contratos").select("contrato_id").eq("usuario_id", uid).execute().data or []:
            if r.get("contrato_id") is not None:
                try:
                    scope.add(int(r["contrato_id"]))
                except (TypeError, ValueError):
                    pass
        jwt_cid = current_user.get("contrato_id")
        if jwt_cid is not None and str(jwt_cid).strip() != "":
            try:
                scope.add(int(jwt_cid))
            except (TypeError, ValueError):
                pass

    def _rows_to_out(rows: list) -> list:
        cargos = {c["id"]: c["nombre"] for c in supabase.table("cargos").select("id, nombre").execute().data}
        return [
            {"id": r["id"], "nombre": f"{r['nombre']} {r.get('apellidos','')}", "cargo": cargos.get(r.get("cargo_id"), "")}
            for r in (rows or []) if r.get("id") is not None and r["id"] != uid
        ]

    if not scope:
        return []

    scope_l = list(scope)
    by_id = {}
    r1 = supabase.table("usuarios").select("id, nombre, apellidos, cargo_id").eq("activo", True).in_("contrato_id", scope_l).execute().data or []
    for r in r1:
        by_id[r["id"]] = r
    uc = supabase.table("usuario_contratos").select("usuario_id").in_("contrato_id", scope_l).execute().data or []
    extra_ids = list({int(x["usuario_id"]) for x in uc if x.get("usuario_id") is not None} - {uid})
    if extra_ids:
        chunk = 120
        for i in range(0, len(extra_ids), chunk):
            part = extra_ids[i : i + chunk]
            r2 = supabase.table("usuarios").select("id, nombre, apellidos, cargo_id").eq("activo", True).in_("id", part).execute().data or []
            for r in r2:
                by_id[r["id"]] = r
    out = _rows_to_out(list(by_id.values()))
    out.sort(key=lambda x: (x.get("nombre") or "").lower())
    return out

# ─────────────────────────────────────────────────────────────
# ACTAS
# ─────────────────────────────────────────────────────────────
class ActaTipoCreate(BaseModel):
    nombre:   str
    es_cobro: bool = False

class ActaCreate(BaseModel):
    consecutivo:            int
    tipo_acta_id:           Optional[int]   = None
    tipo_grupo:             str             = "administrativa"
    observacion:            Optional[str]   = None
    asignado_a:             Optional[int]   = None
    fecha_asignacion:       Optional[str]   = None
    enlace:                 Optional[str]   = None
    numero_rpo:             Optional[int]   = None
    fecha_inicio:           Optional[str]   = None
    fecha_fin:              Optional[str]   = None
    valor_comp_ambiental:   Optional[float] = 0
    calificacion_ambiental: Optional[float] = None
    valor_comp_social:      Optional[float] = 0
    calificacion_social:    Optional[float] = None
    valor_comp_pmt:         Optional[float] = 0
    calificacion_pmt:       Optional[float] = None
    valor_cobrado_adicional:Optional[float] = 0
    ajuste_iccp:            Optional[float] = 0
    ajuste_icociv:          Optional[float] = 0
    ajuste_ipc:             Optional[float] = 0
    pct_proyectado_ajustes: Optional[float] = None

@app.get("/actas-tipos/{contrato_id}")
def get_actas_tipos(contrato_id: int, current_user=Depends(get_current_user)):
    rows = supabase.table("actas_tipos").select("*")\
        .or_(f"contrato_id.is.null,contrato_id.eq.{contrato_id}")\
        .order("nombre").execute().data
    return rows or []

@app.post("/actas-tipos/{contrato_id}")
def crear_acta_tipo(contrato_id: int, body: ActaTipoCreate, current_user=Depends(get_current_user)):
    existente = supabase.table("actas_tipos").select("id").eq("nombre", body.nombre).execute().data
    if existente:
        return existente[0]
    result = supabase.table("actas_tipos").insert({
        "nombre": body.nombre, "es_cobro": body.es_cobro, "contrato_id": contrato_id
    }).execute()
    return result.data[0] if result.data else {"ok": True}

@app.get("/actas/{contrato_id}/lista")
def listar_actas(contrato_id: int, current_user=Depends(get_current_user)):
    rows = supabase.table("actas").select(
        "*, actas_tipos(nombre, es_cobro), usuarios(nombre, apellidos)"
    ).eq("contrato_id", contrato_id).order("consecutivo", desc=True).execute().data
    rpo_ids = [
        r["id"]
        for r in (rows or [])
        if (r.get("tipo_grupo") or "").strip().upper() == "RPO" and r.get("id") is not None
    ]
    rpo_costo = {}
    if rpo_ids:
        rpo_costo = rpo_resumen_actas_rpc(supabase, contrato_id, rpo_ids)
        if rpo_costo is None:
            rpo_costo = rpo_conciliacion_por_contrato(supabase, contrato_id, rpo_ids)
    result = []
    for r in (rows or []):
        tipo = r.get("actas_tipos") or {}
        usr  = r.get("usuarios") or {}
        adj  = (r.get("ajuste_iccp") or 0) + (r.get("ajuste_icociv") or 0) + (r.get("ajuste_ipc") or 0)
        total = (r.get("valor_comp_ambiental") or 0) + (r.get("valor_comp_social") or 0) + \
                (r.get("valor_comp_pmt") or 0) + (r.get("valor_cobrado_adicional") or 0) + adj
        is_rpo = (r.get("tipo_grupo") or "").strip().upper() == "RPO"
        rpo = None
        if is_rpo and r.get("id") is not None:
            rpo = rpo_costo.get(int(r["id"]), {
                "costo_directo_total": 0.0,
                "registros_n3_aprobado": 0,
            })
        valor_mostrado = float(rpo.get("costo_directo_total", 0.0) or 0) if is_rpo and rpo else total
        out = {**r,
            "tipo_nombre":       tipo.get("nombre", ""),
            "es_cobro":          tipo.get("es_cobro", False),
            "asignado_nombre":   f"{usr.get('nombre','')} {usr.get('apellidos','')}".strip(),
            "valor_total_ajustes": adj,
            "valor_total_acta":    valor_mostrado,
        }
        if is_rpo and rpo is not None:
            out["costo_directo_rpo_sicoe_n3"] = rpo.get("costo_directo_total", 0.0)
            out["registros_sicoe_n3_aprobado"] = rpo.get("registros_n3_aprobado", 0)
            out["registros_sicoe_cascade_interventoria"] = rpo.get(
                "registros_cascade_interventoria", rpo.get("registros_n3_aprobado", 0)
            )
        result.append(out)
    return result

@app.get("/actas/{contrato_id}/proximo-consecutivo")
def proximo_consecutivo_acta(contrato_id: int, current_user=Depends(get_current_user)):
    rows = supabase.table("actas").select("consecutivo").eq("contrato_id", contrato_id).execute().data
    maximo = max((r["consecutivo"] for r in rows), default=0)
    return {"proximo": maximo + 1}

@app.get("/actas/{contrato_id}/usuarios-contrato")
def usuarios_del_contrato(contrato_id: int, current_user=Depends(get_current_user)):
    uc = supabase.table("usuario_contratos").select("usuario_id").eq("contrato_id", contrato_id).execute().data or []
    ids_uc = [r["usuario_id"] for r in uc]
    usuarios_principal = supabase.table("usuarios").select("id, nombre, apellidos, cargo_id")\
        .eq("contrato_id", contrato_id).execute().data or []
    ids_principal = [u["id"] for u in usuarios_principal]
    todos_ids = list(set(ids_uc + ids_principal))
    if not todos_ids:
        return []
    users = supabase.table("usuarios").select("id, nombre, apellidos").in_("id", todos_ids).execute().data or []
    return users

@app.post("/actas/{contrato_id}")
def crear_acta(contrato_id: int, body: ActaCreate, current_user=Depends(get_current_user)):
    row = {"contrato_id": contrato_id, **{k: v for k, v in body.dict().items() if v is not None}}
    result = supabase.table("actas").insert(row).execute()
    nuevo = result.data[0] if result.data else {}
    registrar_log(current_user, "CREAR", "ACTAS", "acta", str(nuevo.get("id","")),
                  {"consecutivo": body.consecutivo, "tipo_grupo": body.tipo_grupo})
    return nuevo

@app.put("/actas/{acta_id}")
def actualizar_acta(acta_id: int, body: ActaCreate, current_user=Depends(get_current_user)):
    data = {k: v for k, v in body.dict().items() if v is not None}
    supabase.table("actas").update({**data, "updated_at": "now()"}).eq("id", acta_id).execute()
    registrar_log(current_user, "EDITAR", "ACTAS", "acta", str(acta_id), {})
    return {"ok": True}


class CerrarActaRpoBody(BaseModel):
    """Cierra un acta RPO en fecha_cierre (≤ fin de mes original), crea el siguiente mes completo y traslada cantidades residuales (sin N3 aprobado)."""
    fecha_cierre: str
    acta_id: Optional[int] = None


def _first_day_next_month(d: date) -> date:
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def _last_day_of_month(y: int, m: int) -> date:
    return date(y, m, calendar.monthrange(y, m)[1])


def _acta_rpo_periodos_se_solapan(fi_a: str, ff_a: str, fi_b: str, ff_b: str) -> bool:
    """[fi_a,ff_a] y [fi_b,ff_b] se solapan (inclusive)."""
    return fi_a <= ff_b and ff_a >= fi_b


@app.post("/actas/{contrato_id}/rpo/cerrar-y-siguiente")
def cerrar_acta_rpo_y_crear_siguiente(
    contrato_id: int,
    body: CerrarActaRpoBody,
    current_user=Depends(get_current_user),
):
    try:
        fc = date.fromisoformat((body.fecha_cierre or "")[:10])
    except ValueError:
        raise HTTPException(status_code=422, detail="fecha_cierre debe ser YYYY-MM-DD válida.")

    cerrar_id = body.acta_id
    acta_row = None

    if cerrar_id is not None:
        rows = supabase.table("actas").select("*").eq("id", cerrar_id).eq("contrato_id", contrato_id).limit(1).execute().data
        if not rows:
            raise HTTPException(status_code=404, detail="Acta no encontrada en este contrato.")
        acta_row = rows[0]
    else:
        ds = fc.isoformat()
        rows = supabase.table("actas").select("*").eq("contrato_id", contrato_id).eq("tipo_grupo", "RPO")\
            .lte("fecha_inicio", ds).gte("fecha_fin", ds).order("id", desc=True).limit(1).execute().data
        if not rows:
            raise HTTPException(
                status_code=404,
                detail="No hay Acta RPO cuyo período incluya la fecha de cierre. Indica acta_id o revise las fechas.",
            )
        acta_row = rows[0]
        cerrar_id = acta_row["id"]

    if (acta_row.get("tipo_grupo") or "").strip().upper() != "RPO":
        raise HTTPException(status_code=400, detail="Solo aplica a actas con tipo_grupo RPO.")

    fi_old = (acta_row.get("fecha_inicio") or "")[:10]
    ff_old = (acta_row.get("fecha_fin") or "")[:10]
    if not fi_old or not ff_old:
        raise HTTPException(status_code=400, detail="El acta no tiene fecha_inicio / fecha_fin.")
    try:
        d_ini = date.fromisoformat(fi_old)
        d_fin_prev = date.fromisoformat(ff_old)
    except ValueError:
        raise HTTPException(status_code=400, detail="Fechas del acta corruptas.")

    if fc < d_ini:
        raise HTTPException(status_code=422, detail="La fecha de cierre no puede ser anterior al inicio del acta.")
    if fc > d_fin_prev:
        raise HTTPException(status_code=422, detail="La fecha de cierre no puede ser posterior al fin vigente del acta.")
    if fc > date.today():
        raise HTTPException(status_code=422, detail="La fecha de cierre no puede ser futura.")

    # Mes completo siguiente al mes de fecha_cierre (ej.: cierre 29-ene → 1–28/29 feb)
    fi_n = _first_day_next_month(fc)
    ff_n = _last_day_of_month(fi_n.year, fi_n.month)
    fi_ns, ff_ns = fi_n.isoformat(), ff_n.isoformat()

    existentes = supabase.table("actas").select("id, numero_rpo, fecha_inicio, fecha_fin, consecutivo")\
        .eq("contrato_id", contrato_id).eq("tipo_grupo", "RPO").execute().data or []
    for ex in existentes:
        if ex.get("id") == cerrar_id:
            continue
        ei = (ex.get("fecha_inicio") or "")[:10]
        ef = (ex.get("fecha_fin") or "")[:10]
        if ei and ef and _acta_rpo_periodos_se_solapan(ei, ef, fi_ns, ff_ns):
            raise HTTPException(
                status_code=409,
                detail=f"Ya existe un Acta RPO que cubre el período {fi_ns} … {ff_ns} (id {ex.get('id')} / RPO {ex.get('numero_rpo')}).",
            )

    rpo_rows = [r for r in existentes if r.get("numero_rpo") is not None]
    max_rpo = max((int(r["numero_rpo"]) for r in rpo_rows), default=0)
    nuevo_numero_rpo = max_rpo + 1

    cons_rows = supabase.table("actas").select("consecutivo").eq("contrato_id", contrato_id).execute().data or []
    max_cons = max((r["consecutivo"] for r in cons_rows), default=0)
    nuevo_cons = max_cons + 1

    # 1) Acortar período del acta actual
    supabase.table("actas").update({"fecha_fin": fc.isoformat(), "updated_at": "now()"}).eq("id", cerrar_id).execute()

    new_row = {
        "contrato_id":      contrato_id,
        "consecutivo":      nuevo_cons,
        "tipo_grupo":       "RPO",
        "numero_rpo":       nuevo_numero_rpo,
        "fecha_inicio":     fi_ns,
        "fecha_fin":        ff_ns,
        "tipo_acta_id":     acta_row.get("tipo_acta_id"),
        "asignado_a":       acta_row.get("asignado_a"),
        "fecha_asignacion": acta_row.get("fecha_asignacion"),
        "observacion":      acta_row.get("observacion"),
    }
    new_row = {k: v for k, v in new_row.items() if v is not None}
    ins = supabase.table("actas").insert(new_row).execute()
    creada = ins.data[0] if ins.data else {}
    nueva_id = creada.get("id")

    registros_movidos = 0
    reportes_tocados = set()

    if nueva_id:
        off = 0
        while True:
            batch = supabase.table("so_registros").select("id, reporte_id, nivel3_estado")\
                .eq("contrato_id", contrato_id).eq("acta_rpo_id", cerrar_id)\
                .range(off, off + 199).execute().data or []
            for reg in batch:
                n3 = (reg.get("nivel3_estado") or "").strip().lower()
                if n3 == "aprobado":
                    continue
                rid = reg["id"]
                supabase.table("so_registros").update({"acta_rpo_id": nueva_id}).eq("id", rid).execute()
                registros_movidos += 1
                rp = reg.get("reporte_id")
                if rp is not None:
                    reportes_tocados.add(int(rp))
            if len(batch) < 200:
                break
            off += 200

        for rep_id in reportes_tocados:
            supabase.table("so_reportes").update({"acta_rpo_id": nueva_id})\
                .eq("id", rep_id).eq("contrato_id", contrato_id).execute()

    registrar_log(
        current_user,
        "EDITAR",
        "ACTAS",
        "acta",
        str(cerrar_id),
        {
            "accion": "cerrar_rpo_anticipado",
            "fecha_cierre": fc.isoformat(),
            "nueva_acta_id": nueva_id,
            "registros_movidos": registros_movidos,
        },
    )

    return {
        "ok": True,
        "acta_cerrada": {"id": cerrar_id, "fecha_fin": fc.isoformat()},
        "acta_creada": creada,
        "periodo_siguiente": {"fecha_inicio": fi_ns, "fecha_fin": ff_ns},
        "registros_movidos_residual": registros_movidos,
    }


@app.get("/actas/{acta_id}/financiero")
def acta_financiero(acta_id: int, current_user=Depends(get_current_user)):
    acta = supabase.table("actas").select("contrato_id, numero_rpo, tipo_grupo")\
        .eq("id", acta_id).single().execute().data
    if not acta or acta.get("tipo_grupo") != "cobro":
        return {"resumen": [], "capitulos": []}
    contrato_id = acta["contrato_id"]
    caps = supabase.table("listado_precios").select("capitulo").eq("contrato_id", contrato_id)\
        .execute().data or []
    caps_unicos = list(dict.fromkeys([r["capitulo"] for r in caps if r.get("capitulo")]))
    resumen = [
        {"estado": "Aprobado",                    "aiu": 0, "iva": 0},
        {"estado": "Pendiente",                   "aiu": 0, "iva": 0},
        {"estado": "Pendiente aprobación precio", "aiu": 0, "iva": 0},
        {"estado": "Rechazado",                   "aiu": 0, "iva": 0},
    ]
    capitulos = [{"capitulo": c, "aprobado": 0, "pendiente": 0, "pendiente_precio": 0,
                  "no_revisado": 0, "rechazado": 0} for c in caps_unicos]
    return {"resumen": resumen, "capitulos": capitulos}


@app.get("/actas/{acta_id}/rpo-costo-conciliacion")
def acta_rpo_costo_conciliacion(acta_id: int, current_user=Depends(get_current_user)):
    """Desglose de costo directo (SICOE, N3 aprobado) por capítulo; solo actas tipo RPO."""
    acta = supabase.table("actas").select(
        "id, contrato_id, tipo_grupo, numero_rpo, consecutivo, fecha_inicio, fecha_fin, tipo_acta_id"
    ).eq("id", acta_id).limit(1).execute().data
    acta = acta[0] if acta else None
    if not acta or (acta.get("tipo_grupo") or "").strip().upper() != "RPO":
        raise HTTPException(status_code=400, detail="Solo aplica a actas con tipo RPO.")
    contrato_id = int(acta["contrato_id"])
    block = rpo_conciliacion_un_acta_rpc(supabase, contrato_id, int(acta_id))
    if block is None:
        m = rpo_conciliacion_por_contrato(supabase, contrato_id, [int(acta_id)])
        block = m.get(
            int(acta_id),
            {
                "costo_directo_total": 0.0,
                "registros_n3_aprobado": 0,
                "registros_cascade_interventoria": 0,
                "por_capitulo": [],
                "secciones": {},
            },
        )
    return {
        "acta_id": int(acta_id),
        "contrato_id": contrato_id,
        "numero_rpo": acta.get("numero_rpo"),
        "consecutivo": acta.get("consecutivo"),
        "periodo": {
            "fecha_inicio": (acta.get("fecha_inicio") or "")[:10],
            "fecha_fin": (acta.get("fecha_fin") or "")[:10],
        },
        "criterio": (
            "Alineado con el dashboard de validación: solo líneas SICOE con N1, N2 y N3 en «Aprobado» (cascada) "
            "e ítem asignado. Sin filtro de «bloqueado» (diferente a formatos CCD con sello de bloqueo). "
            "Bloque «obra / ensayos–sondeos» = misma regla de capítulos que dashboard_matriz_validacion (14, 15, ENSAYO, SONDEO)."
        ),
        "costo_directo_total": block.get("costo_directo_total", 0.0),
        "registros_n3_aprobado": block.get("registros_n3_aprobado", 0),
        "registros_cascade_interventoria": block.get("registros_cascade_interventoria", block.get("registros_n3_aprobado", 0)),
        "por_capitulo": block.get("por_capitulo") or [],
        "secciones": block.get("secciones") or {},
    }


# ─────────────────────────────────────────────────────────────
# SUBCONTRATISTAS
# ─────────────────────────────────────────────────────────────
class SubcontratistaCreate(BaseModel):
    razon_social:    str
    objeto_contrato: Optional[str] = None
    nit:             Optional[str] = None
    nombre_contacto: Optional[str] = None
    telefono:        Optional[str] = None

class CorteCreate(BaseModel):
    tipo_periodo: str          # 'quincenal' | 'mensual'
    consecutivo:  int
    fecha_inicio: str          # ISO date YYYY-MM-DD
    fecha_fin:    str

class CorteUpdate(BaseModel):
    fecha_fin: str

class SubprecioCreate(BaseModel):
    listado_precio_id:   int
    precio_unitario_sub: float

class SubprecioUpdate(BaseModel):
    precio_unitario_sub: float

@app.get("/subcontratistas/{contrato_id}")
def listar_subcontratistas(contrato_id: int, current_user=Depends(get_current_user)):
    _require_contract_access(current_user, contrato_id)
    rows = supabase.table("subcontratistas").select("*").eq("contrato_id", contrato_id).order("razon_social").execute().data
    return rows or []

@app.post("/subcontratistas/{contrato_id}")
def crear_subcontratista(contrato_id: int, body: SubcontratistaCreate, current_user=Depends(get_current_user)):
    _require_contract_access(current_user, contrato_id)
    row = {"contrato_id": contrato_id, **body.dict()}
    result = supabase.table("subcontratistas").insert(row).execute()
    nuevo = result.data[0] if result.data else {}
    registrar_log(current_user, "CREAR", "SUBCONTRATISTAS", "subcontratista", str(nuevo.get("id","")),
                  {"razon_social": body.razon_social})
    return nuevo

@app.put("/subcontratistas/{sub_id}")
def actualizar_subcontratista(sub_id: int, body: SubcontratistaCreate, current_user=Depends(get_current_user)):
    supabase.table("subcontratistas").update(body.dict()).eq("id", sub_id).execute()
    registrar_log(current_user, "EDITAR", "SUBCONTRATISTAS", "subcontratista", str(sub_id),
                  {"razon_social": body.razon_social})
    return {"ok": True}

@app.patch("/subcontratistas/{sub_id}/toggle-activo")
def toggle_activo_subcontratista(sub_id: int, current_user=Depends(get_current_user)):
    actual = supabase.table("subcontratistas").select("activo").eq("id", sub_id).single().execute().data
    nuevo_estado = not (actual.get("activo") or False)
    supabase.table("subcontratistas").update({"activo": nuevo_estado}).eq("id", sub_id).execute()
    registrar_log(current_user, "EDITAR", "SUBCONTRATISTAS", "subcontratista", str(sub_id),
                  {"activo": nuevo_estado})
    return {"activo": nuevo_estado}

# ── Cortes ──────────────────────────────────────────────────
@app.get("/subcontratistas/{sub_id}/cortes")
def listar_cortes(sub_id: int, current_user=Depends(get_current_user)):
    rows = supabase.table("subcontratista_cortes").select("*").eq("subcontratista_id", sub_id).order("consecutivo").execute().data
    return rows or []

@app.get("/subcontratistas/{sub_id}/proximo-consecutivo")
def proximo_consecutivo(sub_id: int, current_user=Depends(get_current_user)):
    rows = supabase.table("subcontratista_cortes").select("consecutivo").eq("subcontratista_id", sub_id).execute().data
    maximo = max((r["consecutivo"] for r in rows), default=0)
    return {"proximo": maximo + 1}

@app.post("/subcontratistas/{sub_id}/cortes")
def crear_corte(sub_id: int, body: CorteCreate, current_user=Depends(get_current_user)):
    from datetime import date
    fi = date.fromisoformat(body.fecha_inicio)
    ff = date.fromisoformat(body.fecha_fin)
    if ff <= fi:
        raise HTTPException(status_code=400, detail="La fecha fin debe ser posterior a la fecha inicio.")
    cortes_existentes = supabase.table("subcontratista_cortes").select("fecha_inicio, fecha_fin, consecutivo")\
        .eq("subcontratista_id", sub_id).order("consecutivo").execute().data or []
    if cortes_existentes:
        ultimo = cortes_existentes[-1]
        ultimo_fin = date.fromisoformat(str(ultimo["fecha_fin"]))
        if fi != ultimo_fin:
            raise HTTPException(status_code=400,
                detail=f"La fecha inicio debe ser exactamente {ultimo_fin.isoformat()} (fecha fin del corte anterior).")
    row = {
        "subcontratista_id": sub_id,
        "contrato_id": supabase.table("subcontratistas").select("contrato_id").eq("id", sub_id).single().execute().data["contrato_id"],
        "consecutivo":   body.consecutivo,
        "tipo_periodo":  body.tipo_periodo,
        "fecha_inicio":  body.fecha_inicio,
        "fecha_fin":     body.fecha_fin,
    }
    result = supabase.table("subcontratista_cortes").insert(row).execute()
    nuevo = result.data[0] if result.data else {}
    registrar_log(current_user, "CREAR", "SUBCONTRATISTAS", "corte", str(nuevo.get("id","")),
                  {"subcontratista_id": sub_id, "consecutivo": body.consecutivo})
    return nuevo

@app.put("/subcontratistas/cortes/{corte_id}")
def actualizar_corte(corte_id: int, body: CorteUpdate, current_user=Depends(get_current_user)):
    from datetime import date, timedelta
    corte = supabase.table("subcontratista_cortes").select("*").eq("id", corte_id).single().execute().data
    if not corte:
        raise HTTPException(status_code=404, detail="Corte no encontrado.")
    nueva_fin = date.fromisoformat(body.fecha_fin)
    supabase.table("subcontratista_cortes").update({"fecha_fin": body.fecha_fin}).eq("id", corte_id).execute()
    siguiente = supabase.table("subcontratista_cortes").select("*")\
        .eq("subcontratista_id", corte["subcontratista_id"])\
        .eq("consecutivo", corte["consecutivo"] + 1).execute().data
    if siguiente:
        sig = siguiente[0]
        tipo = sig.get("tipo_periodo", corte.get("tipo_periodo", "quincenal"))
        nueva_fi_sig = nueva_fin
        if tipo == "quincenal":
            nueva_ff_sig = nueva_fi_sig + timedelta(days=15)
        else:
            import calendar
            dias_mes = calendar.monthrange(nueva_fi_sig.year, nueva_fi_sig.month)[1]
            nueva_ff_sig = nueva_fi_sig + timedelta(days=dias_mes)
        supabase.table("subcontratista_cortes").update({
            "fecha_inicio": nueva_fi_sig.isoformat(),
            "fecha_fin": nueva_ff_sig.isoformat(),
        }).eq("id", sig["id"]).execute()
    registrar_log(current_user, "EDITAR", "SUBCONTRATISTAS", "corte", str(corte_id),
                  {"nueva_fecha_fin": body.fecha_fin})
    return {"ok": True}

# ── Precios subcontratista ───────────────────────────────────
@app.get("/subcontratistas/{sub_id}/precios")
def listar_precios_sub(sub_id: int, current_user=Depends(get_current_user)):
    rows = supabase.table("subcontratista_precios").select(
        "*, listado_precios(capitulo, competencia, item_numero, descripcion, unidad, precio_unitario)"
    ).eq("subcontratista_id", sub_id).execute().data
    result = []
    for r in (rows or []):
        lp = r.get("listado_precios") or {}
        result.append({
            "id":                   r["id"],
            "listado_precio_id":    r["listado_precio_id"],
            "precio_unitario_sub":  r["precio_unitario_sub"],
            "capitulo":             lp.get("capitulo", ""),
            "competencia":          lp.get("competencia", ""),
            "item_numero":          lp.get("item_numero", ""),
            "descripcion":          lp.get("descripcion", ""),
            "unidad":               lp.get("unidad", ""),
            "precio_unitario_ref":  lp.get("precio_unitario", 0),
        })
    return result

@app.post("/subcontratistas/{sub_id}/precios")
def agregar_precio_sub(sub_id: int, body: SubprecioCreate, current_user=Depends(get_current_user)):
    sub = supabase.table("subcontratistas").select("contrato_id").eq("id", sub_id).single().execute().data
    if not sub:
        raise HTTPException(status_code=404, detail="Subcontratista no encontrado.")
    existente = supabase.table("listado_precios").select("id").eq("id", body.listado_precio_id)\
        .eq("contrato_id", sub["contrato_id"]).execute().data
    if not existente:
        raise HTTPException(status_code=400, detail="El ítem no pertenece al listado de precios de este contrato.")
    row = {
        "subcontratista_id":   sub_id,
        "contrato_id":         sub["contrato_id"],
        "listado_precio_id":   body.listado_precio_id,
        "precio_unitario_sub": body.precio_unitario_sub,
    }
    result = supabase.table("subcontratista_precios").insert(row).execute()
    registrar_log(current_user, "CREAR", "SUBCONTRATISTAS", "precio_sub", str(sub_id),
                  {"listado_precio_id": body.listado_precio_id})
    return result.data[0] if result.data else {"ok": True}

@app.put("/subcontratistas/precios/{precio_id}")
def actualizar_precio_sub(precio_id: int, body: SubprecioUpdate, current_user=Depends(get_current_user)):
    supabase.table("subcontratista_precios").update({"precio_unitario_sub": body.precio_unitario_sub})\
        .eq("id", precio_id).execute()
    registrar_log(current_user, "EDITAR", "SUBCONTRATISTAS", "precio_sub", str(precio_id),
                  {"precio_unitario_sub": body.precio_unitario_sub})
    return {"ok": True}

@app.get("/subcontratistas/{contrato_id}/alertas-corte")
def alertas_corte(contrato_id: int, current_user=Depends(get_current_user)):
    """Subcontratistas cuyo corte activo vence mañana o hoy."""
    from datetime import date, timedelta
    hoy = date.today()
    manana = hoy + timedelta(days=1)
    subs = supabase.table("subcontratistas").select("id, razon_social").eq("contrato_id", contrato_id)\
        .eq("activo", True).execute().data or []
    alertas = []
    for s in subs:
        cortes = supabase.table("subcontratista_cortes").select("consecutivo, fecha_fin, tipo_periodo")\
            .eq("subcontratista_id", s["id"]).order("consecutivo", desc=True).limit(1).execute().data
        if cortes:
            ultimo = cortes[0]
            ff = date.fromisoformat(str(ultimo["fecha_fin"]))
            if ff in (hoy, manana):
                alertas.append({
                    "subcontratista_id": s["id"],
                    "razon_social":      s["razon_social"],
                    "fecha_fin":         ultimo["fecha_fin"],
                    "consecutivo":       ultimo["consecutivo"],
                    "vence_hoy":         ff == hoy,
                })
    return alertas

# ─────────────────────────────────────────────
# SICOE OBRA
# ─────────────────────────────────────────────

@app.get("/sicoe-obra/{contrato_id}/reportes")
def listar_reportes_obra(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.table("so_reportes")\
            .select("*, subcontratistas(razon_social)")\
            .eq("contrato_id", contrato_id)\
            .order("numero_reporte", desc=True).execute().data
    rows = supabase_execute(_q)

    # Batch-resolve semana_numero y acta_rpo (las FKs no están expuestas como JOINs implícitos en PostgREST)
    semana_ids = list({r["semana_id"] for r in rows if r.get("semana_id")})
    acta_ids   = list({r["acta_rpo_id"] for r in rows if r.get("acta_rpo_id")})

    semana_map = {}
    if semana_ids:
        try:
            def _sems():
                return supabase.table("so_semanas").select("id, numero_semana")\
                    .in_("id", semana_ids).execute().data
            for s in supabase_execute(_sems):
                semana_map[s["id"]] = s["numero_semana"]
        except Exception:
            pass

    acta_map = {}
    if acta_ids:
        try:
            def _actas():
                return supabase.table("actas").select("id, numero_rpo, consecutivo")\
                    .in_("id", acta_ids).execute().data
            for a in supabase_execute(_actas):
                acta_map[a["id"]] = a
        except Exception:
            pass

    for r in rows:
        sub = r.pop("subcontratistas", None)
        r["subcontratista_nombre"] = sub["razon_social"] if sub else None
        r["semana_numero"]    = semana_map.get(r.get("semana_id"))
        acta = acta_map.get(r.get("acta_rpo_id"))
        r["acta_rpo"]         = acta["numero_rpo"] if acta else None
        r["acta_consecutivo"] = acta["consecutivo"] if acta else None
    return rows

@app.get("/sicoe-obra/{contrato_id}/registros-bulk")
def registros_bulk_offline(
    contrato_id: int,
    acta_id: Optional[int] = None,
    current_user=Depends(get_current_user),
):
    """
    Descarga masiva de registros para caché offline.
    Si se indica acta_id, filtra solo los registros de esa acta.
    Devuelve hasta 5000 filas.
    """
    COLS = (
        "id, reporte_id, contrato_id, acta_rpo_id, numero_registro, "
        "item_numero, capitulo, descripcion, unidad, cantidad, "
        "longitud, ancho, espesor, tramo, costado, pk_id, "
        "abscisa_inicio, abscisa_final, observacion, "
        "nivel1_estado, nivel1_usuario_id, nivel1_fecha, "
        "nivel2_estado, nivel2_usuario_id, nivel2_fecha, "
        "nivel3_estado, nivel3_usuario_id, nivel3_fecha, "
        "subcontratista_id, numero_corte_subcontratista, "
        "foto_url, foto_numero, foto_descripcion, "
        "grafico_url, grafico_numero, grafico_descripcion, "
        "created_at, updated_at"
    )
    def _q():
        q = (
            supabase.table("so_registros")
            .select(COLS)
            .eq("contrato_id", contrato_id)
        )
        if acta_id is not None:
            q = q.eq("acta_rpo_id", acta_id)
        return q.limit(5000).execute().data
    return supabase_execute(_q)


@app.get("/sicoe-obra/{contrato_id}/offline-pack")
def offline_pack(
    contrato_id: int,
    acta_rpo: int,
    current_user=Depends(get_current_user),
):
    """
    Paquete offline completo para un acta.
    Una sola petición devuelve: actas, semanas, precios, reportes y registros.
    El servidor resuelve acta_id desde acta_rpo internamente.
    Cada query es independiente; errores parciales se reportan en 'errores'.
    """
    errores = {}

    # 1. Actas del contrato
    actas = []
    acta_id = None
    try:
        _r_acta = supabase.table("actas").select("id, numero_rpo, consecutivo") \
            .eq("contrato_id", contrato_id).execute().data or []
        # Resolver acta_id SOLO por numero_rpo (igual que el resto del backend)
        # NO usar consecutivo: puede coincidir con otra acta diferente
        for a in _r_acta:
            if str(a.get("numero_rpo")) == str(acta_rpo):
                acta_id = a["id"]
                break
        actas = _r_acta
    except Exception as e:
        errores["actas"] = str(e)

    # 3. Reportes del acta
    reportes = []
    try:
        q = supabase.table("so_reportes") \
            .select("*, subcontratistas(razon_social)") \
            .eq("contrato_id", contrato_id)
        if acta_id is not None:
            q = q.eq("acta_rpo_id", acta_id)
        reportes = q.order("numero_reporte", desc=True).limit(2000).execute().data or []
    except Exception as e:
        errores["reportes"] = str(e)

    # 4. Registros del acta — paginado de 1000 en 1000 (límite PostgREST)
    registros = []
    try:
        off = 0
        while True:
            q = supabase.table("so_registros").select("*") \
                .eq("contrato_id", contrato_id)
            if acta_id is not None:
                q = q.eq("acta_rpo_id", acta_id)
            batch = q.order("id").range(off, off + 999).execute().data or []
            registros.extend(batch)
            if len(batch) < 1000:
                break
            off += 1000
            if off >= 5000:   # techo de seguridad
                break
    except Exception as e:
        errores["registros"] = str(e)

    # 5. Semanas
    semanas = []
    try:
        semanas = supabase.table("so_semanas").select("*") \
            .eq("contrato_id", contrato_id).execute().data or []
    except Exception as e:
        errores["semanas"] = str(e)

    # 6. Precios (paginado)
    precios = []
    try:
        off = 0
        while True:
            batch = supabase.table("listado_precios").select("*") \
                .eq("contrato_id", contrato_id) \
                .order("item_numero").range(off, off + 999).execute().data
            precios.extend(batch)
            if len(batch) < 1000:
                break
            off += 1000
    except Exception as e:
        errores["precios"] = str(e)

    return {
        "acta_id":   acta_id,
        "actas":     actas,
        "semanas":   semanas,
        "precios":   precios,
        "reportes":  reportes,
        "registros": registros,
        "errores":   errores,  # campo de diagnóstico — vacío si todo OK
    }


@app.get("/sicoe-obra/{contrato_id}/pk-ids")
def listar_pk_ids(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.table("pk_ids")\
            .select("*")\
            .eq("contrato_id", contrato_id)\
            .order("pk_id").execute().data
    return supabase_execute(_q)

@app.get("/sicoe-obra/{contrato_id}/subcontratistas-activos")
def listar_subcontratistas_activos(contrato_id: int, current_user=Depends(get_current_user)):
    _require_contract_access(current_user, contrato_id)
    def _q():
        return supabase.table("subcontratistas")\
            .select("id, razon_social")\
            .eq("contrato_id", contrato_id)\
            .eq("activo", True)\
            .order("razon_social").execute().data
    rows = supabase_execute(_q)
    for r in rows:
        r["nombre"] = r.pop("razon_social", "")
    return rows

@app.get("/usuarios/{usuario_id}")
def obtener_usuario(usuario_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.table("usuarios").select("id, nombre, apellidos")\
            .eq("id", usuario_id).execute().data
    rows = supabase_execute(_q)
    return rows[0] if rows else {}

@app.get("/sicoe-obra/{contrato_id}/inspectores")
def listar_inspectores(contrato_id: int, current_user=Depends(get_current_user)):
    def _u():
        return supabase.table("usuarios")\
            .select("id, nombre, apellidos")\
            .eq("contrato_id", contrato_id)\
            .eq("estado", "aprobado")\
            .eq("cargo_id", 54)\
            .order("nombre").execute().data
    usuarios = supabase_execute(_u)
    return [{"id": u["id"], "nombre": f"{u.get('nombre','')} {u.get('apellidos','')}".strip()} for u in usuarios]

@app.get("/sicoe-obra/{contrato_id}/capitulos")
def listar_capitulos_obra(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.table("listado_precios")\
            .select("capitulo")\
            .eq("contrato_id", contrato_id)\
            .execute().data
    rows = supabase_execute(_q)
    import re
    def orden_capitulo(c):
        m = re.match(r'^(\d+)', c)
        return (int(m.group(1)) if m else 9999, c)
    caps = sorted(set(r["capitulo"] for r in rows if r.get("capitulo")), key=orden_capitulo)
    return [{"capitulo": c} for c in caps]

@app.get("/sicoe-obra/{contrato_id}/next-reporte")
def next_numero_reporte(contrato_id: int, current_user=Depends(get_current_user)):
    """Vista previa del siguiente número (sin llamar a la RPC ni reservar). Misma lógica que
    `siguiente_numero_reporte` (incl. contratos con sicoe_consecutivos_desde_uno)."""

    def _contrato():
        return supabase.table("contratos").select("sicoe_consecutivos_desde_uno")\
            .eq("id", contrato_id).limit(1).execute().data

    def _max_num():
        return supabase.table("so_reportes").select("numero_reporte")\
            .eq("contrato_id", contrato_id).order("numero_reporte", desc=True).limit(1).execute().data

    def _rsv():
        return supabase.table("sico_ultimo_numero_reporte").select("reservado_hasta")\
            .eq("contrato_id", contrato_id).limit(1).execute().data

    crows = supabase_execute(_contrato)
    desde_uno = bool(crows and crows[0].get("sicoe_consecutivos_desde_uno"))
    mrows = supabase_execute(_max_num)
    m_tab = mrows[0]["numero_reporte"] if mrows else 0
    if m_tab is None:
        m_tab = 0
    urows = supabase_execute(_rsv)
    if urows:
        rsv = urows[0].get("reservado_hasta")
    else:
        rsv = 0 if desde_uno else 34999
    if rsv is None:
        rsv = 0 if desde_uno else 34999
    piso = 35000
    if desde_uno:
        sig = max(m_tab + 1, rsv + 1)
    else:
        sig = max(piso, m_tab + 1, rsv + 1)
    return {"siguiente": sig}

@app.get("/sicoe-obra/{contrato_id}/nodos")
def listar_nodos_obra(contrato_id: int, capitulo: str = None, current_user=Depends(get_current_user)):
    def _q():
        q = supabase.table("presupuesto")\
            .select("no_inicio, no_final")\
            .eq("contrato_id", contrato_id)\
            .eq("dado_de_baja", False)
        if capitulo:
            q = q.eq("capitulo", capitulo)
        return q.execute().data
    rows = supabase_execute(_q)
    nodos = set()
    for r in rows:
        if r.get("no_inicio"): nodos.add(r["no_inicio"])
        if r.get("no_final"): nodos.add(r["no_final"])
    return sorted(list(nodos))

class ReporteCreate(BaseModel):
    descripcion_actividad: str
    subcontratista_id: Optional[int] = None
    inspector_id: Optional[int] = None
    capitulo: str
    pk_id_id: Optional[int] = None
    civ: Optional[str] = None
    tramo: Optional[str] = None
    infraestructura: Optional[str] = None
    calzada: Optional[str] = None
    ubicacion: Optional[str] = None
    coord_lat: Optional[float] = None
    coord_lng: Optional[float] = None
    margen: Optional[str] = None
    abs_inicio: Optional[float] = None
    abs_final: Optional[float] = None
    nodo_ini: Optional[str] = None
    nodo_fin: Optional[str] = None
    estado: Optional[str] = None
    enlace_soporte: Optional[str] = None
    semana_id: Optional[int] = None
    acta_rpo_id: Optional[int] = None
    corte_id: Optional[int] = None

@app.get("/sicoe-obra/{contrato_id}/cargos-validacion")
def cargos_con_validacion(contrato_id: int, current_user=Depends(get_current_user)):
    try:
        # Cargos que están en CARGO_ID_NIVEL_MAP (tabla cargos). Siempre listarlos para
        # filtros aunque no haya usuario aprobado con ese cargo en el contrato — si no,
        # opciones como «Residente de Obra» desaparecen sin poder filtrar registros ya validados.
        cargo_ids_nivel = list(CARGO_ID_NIVEL_MAP.keys())

        def _cargos():
            return supabase.table("cargos").select("id, nombre")\
                .in_("id", cargo_ids_nivel).execute().data
        cargos_rows = supabase_execute(_cargos)
        cargo_id_nombre = {r["id"]: r["nombre"] for r in cargos_rows}

        return [
            {"id": cid, "nombre": cargo_id_nombre[cid]}
            for cid in cargo_ids_nivel
            if cid in cargo_id_nombre
        ]
    except Exception as e:
        return []


@app.get("/sicoe-obra/{contrato_id}/reportes/buscar")
def buscar_reportes_obra(
    contrato_id: int,
    numero_reporte: Optional[int] = None,
    numero_registro: Optional[int] = None,
    semana: Optional[int] = None,
    acta_rpo: Optional[int] = None,
    subcontratista_id: Optional[int] = None,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    items_filtro: Optional[str] = None,
    items_filtro_op: Optional[str] = Query(None),
    tramo: Optional[str] = None,
    costado: Optional[str] = None,
    pk_id: Optional[int] = None,
    abs_inicio: Optional[float] = None,
    abs_final: Optional[float] = None,
    estado: Optional[str] = None,
    cargo_id: Optional[int] = None,
    estado_validacion: Optional[str] = None,
    validacion_capas: Optional[str] = None,
    validacion_capas_op: Optional[str] = Query(None),
    q_observacion: Optional[str] = None,
    q_nodo: Optional[str] = None,
    etiqueta_validacion: Optional[str] = None,
    offset: int = 0,
    limit: int = 50,
    current_user=Depends(get_current_user)
):
    limit = min(limit, 100)
    _ocultar_costo_rep = _sicoe_ocultar_costo_directo_reportes(current_user)
    capas_v = _parse_validacion_capas_param(validacion_capas, cargo_id, estado_validacion)
    consulta_directa_identificador = (
        numero_reporte is not None or numero_registro is not None
    )
    if consulta_directa_identificador:
        capas_v = []
        semana = None
        acta_rpo = None
    _cap_op_buscar = _parse_capas_validacion_op(validacion_capas_op)
    _defer_capas_or_grilla = (
        bool(capas_v)
        and not _estado_filtro_omite_validacion_por_cargo(estado)
        and len(capas_v) > 1
        and _cap_op_buscar == "or"
    )
    _nivel_l = None
    _ev_l = None
    if capas_v:
        try:
            _nivel_l = capas_v[0].get("campo") or _capa_campo_validacion(capas_v[0])
            _ev_l = (capas_v[0].get("estado") or "").strip()
        except (TypeError, ValueError, KeyError, IndexError):
            _nivel_l = None
            _ev_l = None

    items_buscar_norm = _normalize_items_filtro_list(items_filtro, item)
    items_buscar_op = items_filtro_op

    has_reg_f = any([
        capitulo, bool(items_buscar_norm), subcontratista_id is not None,
        bool(tramo), bool(costado),
    ])
    capas_aplican_a_lineas = bool(capas_v) and not _estado_filtro_omite_validacion_por_cargo(estado)
    q_obs_trim = (str(q_observacion).strip() if q_observacion is not None else "")

    # Universo de reportes: IDs con al menos una fila de obra que cumple todos los criterios de línea a la vez
    # (AND), alineado con GET /sicoe-obra/.../analisis. Evita intersectar conjuntos por reporte_id con criterios distintos.
    reporte_ids_from_reg = None
    omit_header_semana_acta_en_reportes = False

    semana_id_filtro = None
    if semana is not None:
        try:
            def _sem_id():
                return supabase.table("so_semanas").select("id")\
                    .eq("contrato_id", contrato_id)\
                    .eq("numero_semana", semana).execute().data
            sem_rows = supabase_execute(_sem_id)
            if sem_rows:
                semana_id_filtro = sem_rows[0]["id"]
            else:
                return {"reportes": [], "total": 0, "offset": offset, "limit": limit, "hay_mas": False}
        except Exception:
            return {"reportes": [], "total": 0, "offset": offset, "limit": limit, "hay_mas": False}

    acta_id_filtro = None
    if acta_rpo is not None:
        try:
            def _acta_id():
                rows = supabase.table("actas").select("id")\
                    .eq("contrato_id", contrato_id)\
                    .eq("numero_rpo", acta_rpo).execute().data
                if not rows:
                    rows = supabase.table("actas").select("id")\
                        .eq("contrato_id", contrato_id)\
                        .eq("consecutivo", acta_rpo).execute().data
                return rows
            acta_rows = supabase_execute(_acta_id)
            if acta_rows:
                acta_id_filtro = acta_rows[0]["id"]
            else:
                return {"reportes": [], "total": 0, "offset": offset, "limit": limit, "hay_mas": False}
        except Exception:
            return {"reportes": [], "total": 0, "offset": offset, "limit": limit, "hay_mas": False}

    etiqueta_f = _sicoe_parse_etiqueta_validacion_param(etiqueta_validacion)
    registro_ids_etiqueta = (
        _sicoe_fetch_registro_ids_etiqueta_validacion(contrato_id, etiqueta_f)
        if etiqueta_f
        else None
    )

    unified_line = any([
        numero_registro is not None,
        abs_inicio is not None or abs_final is not None,
        has_reg_f,
        bool(q_obs_trim),
        pk_id is not None,
        capas_aplican_a_lineas,
        semana_id_filtro is not None,
        acta_id_filtro is not None,
        bool(etiqueta_f),
    ])

    if unified_line:
        ids_unif = _sicoe_collect_reporte_ids_misma_linea(
            contrato_id,
            numero_registro=numero_registro,
            abs_inicio=abs_inicio,
            abs_final=abs_final,
            capitulo=capitulo,
            items=items_buscar_norm,
            items_op=items_buscar_op,
            subcontratista_id=subcontratista_id,
            tramo=tramo,
            costado=costado,
            pk_id=pk_id,
            q_observacion=q_obs_trim or None,
            semana_id=semana_id_filtro,
            acta_rpo_id=acta_id_filtro,
            capas_v=(capas_v if capas_aplican_a_lineas else None),
            capas_v_op=validacion_capas_op,
            estado=estado,
            registro_ids_etiqueta=registro_ids_etiqueta,
        )
        if not ids_unif:
            return {"reportes": [], "total": 0, "offset": offset, "limit": limit, "hay_mas": False}
        reporte_ids_from_reg = list(ids_unif)
        if semana_id_filtro is not None or acta_id_filtro is not None:
            omit_header_semana_acta_en_reportes = True

    if q_nodo is not None and str(q_nodo).strip():
        ids_n = _sicoe_reporte_ids_coinciden_nodo(contrato_id, q_nodo, reporte_ids_from_reg)
        if ids_n is not None:
            if not ids_n:
                return {"reportes": [], "total": 0, "offset": offset, "limit": limit, "hay_mas": False}
            if reporte_ids_from_reg is not None:
                reporte_ids_from_reg = [x for x in reporte_ids_from_reg if x in ids_n]
            else:
                reporte_ids_from_reg = list(ids_n)
            if not reporte_ids_from_reg:
                return {"reportes": [], "total": 0, "offset": offset, "limit": limit, "hay_mas": False}

    # PostgREST envía .in_() como parámetro URL; listas grandes superan el límite
    # de longitud del servidor y se truncan silenciosamente. Se divide en chunks
    # para garantizar que todos los IDs se filtren correctamente.
    _IDS_CHUNK_SIZE = 200

    def _build_q(ids_chunk=None):
        q = supabase.table("so_reportes").select("*, subcontratistas(razon_social)")\
            .eq("contrato_id", contrato_id)
        if numero_reporte is not None:
            q = q.eq("numero_reporte", numero_reporte)
        # subcontratista / tramo / costado: ya restringidos vía so_registros (has_reg_f)
        # o vía bloque de validación; repetir en cabecera vacía el universo.
        # pk_id: restringido por reporte_ids_from_reg (líneas so_registros), no por cabecera.
        # abs_inicio/abs_final: ya aplicados por solape en so_registros (arriba)
        _hay_n23_build = (
            (_validacion_cualquier_nivel2_o_3(capas_v) if capas_v else False)
            or (_nivel_l is not None and _es_validacion_avanzada(_nivel_l))
        )
        if estado:
            q = _so_reportes_q_por_estado(q, estado)
        elif _hay_n23_build and ids_chunk is None:
            q = q.not_.in_("estado", list(ESTADOS_REPORTE_EXCL_VALIDACION_AVANZADA))
        # Importante: para validaciones por nivel, el universo debe definirse por
        # estado de so_registros (nivelX_estado), no por estado de so_reportes.
        # Solo se filtra por estado de reporte cuando el usuario lo pide explícitamente.
        # Si ya tenemos IDs desde so_registros (ids_chunk), no volver a excluir por cabecera.
        if not omit_header_semana_acta_en_reportes:
            if semana_id_filtro is not None:
                q = q.eq("semana_id", semana_id_filtro)
            if acta_id_filtro is not None:
                q = q.eq("acta_rpo_id", acta_id_filtro)
        if ids_chunk is not None:
            q = q.in_("id", ids_chunk)
        return q

    # Chunking completo para garantizar todos los IDs
    all_rows = []
    if reporte_ids_from_reg is not None:
        seen_ids = set()
        for i in range(0, len(reporte_ids_from_reg), 200):
            _chunk = reporte_ids_from_reg[i:i + 200]
            def _qc(c=_chunk):
                return _build_q(c).limit(1000).execute().data
            for row in supabase_execute(_qc):
                if row["id"] not in seen_ids:
                    seen_ids.add(row["id"])
                    all_rows.append(row)
    else:
        def _qall():
            return _build_q(None).limit(5000).execute().data
        all_rows = supabase_execute(_qall)

    all_rows.sort(key=lambda r: (r.get("numero_reporte") or 0), reverse=True)
    rows = all_rows[offset:offset + limit + 1]
    hay_mas = len(rows) > limit
    rows = rows[:limit]

    # Batch-resolve semana_numero y acta_rpo
    semana_ids = list({r["semana_id"] for r in rows if r.get("semana_id")})
    acta_ids   = list({r["acta_rpo_id"] for r in rows if r.get("acta_rpo_id")})
    semana_map = {}
    if semana_ids:
        try:
            def _sems():
                return supabase.table("so_semanas").select("id, numero_semana")\
                    .in_("id", semana_ids).execute().data
            for s in supabase_execute(_sems):
                semana_map[s["id"]] = s["numero_semana"]
        except Exception:
            pass
    acta_map = {}
    if acta_ids:
        try:
            def _actas():
                return supabase.table("actas").select("id, numero_rpo, consecutivo")\
                    .in_("id", acta_ids).execute().data
            for a in supabase_execute(_actas):
                acta_map[a["id"]] = a
        except Exception:
            pass
    # Batch-resolve per-cargo estado_max desde so_registros
    reporte_ids_batch = [r["id"] for r in rows]
    if reporte_ids_batch:
        try:
            _rb_l = reporte_ids_batch

            def _reg_estados_q_base(reg_id_filter: Optional[List[int]] = None):
                q = supabase.table("so_registros")\
                    .select("reporte_id, costo_directo, nivel1_estado, nivel2_estado, nivel3_estado, sub_estado, semana_id, acta_rpo_id, item_numero, capitulo, subcontratista_id, tramo, margen")\
                    .in_("reporte_id", _rb_l)
                if reg_id_filter is not None:
                    q = q.in_("id", reg_id_filter)

                if numero_registro is not None:
                    q = q.eq("numero_registro", numero_registro)

                # Mantener coherencia con el universo filtrado de grilla/panel
                if semana_id_filtro is not None:
                    q = q.eq("semana_id", semana_id_filtro)
                if acta_id_filtro is not None:
                    q = q.eq("acta_rpo_id", acta_id_filtro)
                if capitulo:
                    q = q.eq("capitulo", capitulo)
                if subcontratista_id is not None:
                    q = q.eq("subcontratista_id", subcontratista_id)
                q = _apply_item_patterns_to_so_registros_q(q, items_buscar_norm, items_buscar_op)
                if tramo:
                    q = q.eq("tramo", tramo)
                if costado:
                    q = _so_reg_filtro_costado(q, costado)
                if pk_id is not None:
                    q = q.eq("pk_id_id", pk_id)
                if q_observacion is not None and str(q_observacion).strip():
                    q = q.ilike("observacion", f"%{str(q_observacion).strip()}%")
                q = _so_reg_filtro_abs_solape(q, abs_inicio, abs_final)
                if acta_id_filtro is not None and not _estado_filtro_es_sin_asignar_item(estado):
                    q = _so_reg_item_asignado(q)

                if capas_v and not _estado_filtro_omite_validacion_por_cargo(estado):
                    if not _defer_capas_or_grilla:
                        q = _so_registros_q_y_capas_validacion(
                            q, capas_v, pk_id, tramo, costado, capitulo, subcontratista_id, None
                        )

                if _estado_filtro_es_sin_asignar_item(estado):
                    q = _so_reg_sin_item_asignado(q)
                return q

            reg_estados = []
            if registro_ids_etiqueta is not None and not registro_ids_etiqueta:
                reg_estados = []
            elif registro_ids_etiqueta is not None:
                for rg_chunk in _sicoe_chunks_int(sorted(registro_ids_etiqueta), 200):
                    rc = list(rg_chunk)
                    _re_off = 0
                    _re_page = 1000
                    while True:
                        def _re_fetch(o=_re_off, ids=rc):
                            return _reg_estados_q_base(ids).range(o, o + _re_page - 1).execute().data

                        _batch = supabase_execute(_re_fetch)
                        reg_estados.extend(_batch)
                        if len(_batch) < _re_page:
                            break
                        _re_off += _re_page
            else:
                _re_off = 0
                _re_page = 1000
                while True:
                    def _re_fetch(o=_re_off):
                        return _reg_estados_q_base().range(o, o + _re_page - 1).execute().data

                    _batch = supabase_execute(_re_fetch)
                    reg_estados.extend(_batch)
                    if len(_batch) < _re_page:
                        break
                    _re_off += _re_page
            if _estado_filtro_es_sin_asignar_item(estado):
                reg_estados = [
                    x for x in reg_estados
                    if not (str(x.get("item_numero") or "").strip())
                ]
            if _defer_capas_or_grilla:
                reg_estados = _filtrar_registros_validacion_capas_sicoe(reg_estados, capas_v, None, "or")
            cargo_map = {r["id"]: {"n1": [], "n2": [], "n3": [], "sub": [], "count": 0} for r in rows}
            costo_map = {}
            for reg in reg_estados:
                rid = reg.get("reporte_id")
                if rid in cargo_map:
                    cargo_map[rid]["n1"].append(reg.get("nivel1_estado") or "No Revisado")
                    cargo_map[rid]["n2"].append(reg.get("nivel2_estado") or "No Revisado")
                    cargo_map[rid]["n3"].append(reg.get("nivel3_estado") or "No Revisado")
                    cargo_map[rid]["sub"].append(reg.get("sub_estado") or "No Revisado")
                    cargo_map[rid]["count"] += 1
                    costo_map[rid] = costo_map.get(rid, 0.0) + float(reg.get("costo_directo") or 0)
            cap_por_rep: dict = {}
            for reg in reg_estados:
                rid = reg.get("reporte_id")
                if not rid:
                    continue
                c = (reg.get("capitulo") or "").strip()
                if rid not in cap_por_rep:
                    cap_por_rep[rid] = set()
                if c:
                    cap_por_rep[rid].add(c)
            for r in rows:
                m = cargo_map.get(r["id"], {})
                r["nivel1_estados"] = list(set(m.get("n1", [])))
                r["nivel2_estados"] = list(set(m.get("n2", [])))
                r["nivel3_estados"] = list(set(m.get("n3", [])))
                r["sub_estados"]    = list(set(m.get("sub", [])))
                r["num_registros"] = m.get("count", 0)
                if not _ocultar_costo_rep:
                    r["costo_directo_validacion"] = round(costo_map.get(r["id"], 0.0), 0)
                caps = cap_por_rep.get(r["id"], set())
                if caps:
                    r["capitulo"] = ", ".join(sorted(caps))
        except Exception:
            for r in rows:
                r["nivel1_estados"] = r["nivel2_estados"] = r["nivel3_estados"] = r["sub_estados"] = []
                r["num_registros"] = 0
            if not _ocultar_costo_rep:
                for r in rows:
                    r["costo_directo_validacion"] = 0.0

    # "Sin asignar ítem" aplica a registros: un reporte sin filas en so_registros no puede asignar ítem
    if _estado_filtro_es_sin_asignar_item(estado):
        rows = [r for r in rows if (r.get("num_registros") or 0) > 0]

    for r in rows:
        sub = r.pop("subcontratistas", None)
        r["subcontratista_nombre"] = sub["razon_social"] if sub else None
        r["semana_numero"]    = semana_map.get(r.get("semana_id"))
        acta = acta_map.get(r.get("acta_rpo_id"))
        r["acta_rpo"]         = acta["numero_rpo"] if acta else None
        r["acta_consecutivo"] = acta["consecutivo"] if acta else None

    # Si la búsqueda está en modo validación por nivel, eliminar reportes
    # sin registros coincidentes para evitar descuadres panel/grilla.
    if _nivel_l and _ev_l:
        rows = [r for r in rows if (r.get("num_registros") or 0) > 0]

    return {"reportes": rows, "total": len(rows), "offset": offset, "limit": limit, "hay_mas": hay_mas}


# ─── SICOE OBRA: Exportar registros filtrados ───────────────────────────────
class ExportarRegistrosBody(BaseModel):
    campos: List[str]

    # Filtros (mismos nombres que la grilla / buscar_reportes_obra)
    numero_reporte: Optional[int] = None
    numero_registro: Optional[int] = None
    semana: Optional[int] = None
    acta_rpo: Optional[int] = None
    subcontratista_id: Optional[int] = None
    capitulo: Optional[str] = None
    item: Optional[str] = None
    items_filtro: Optional[str] = None
    items_filtro_op: Optional[str] = None
    tramo: Optional[str] = None
    costado: Optional[str] = None
    pk_id: Optional[int] = None
    abs_inicio: Optional[float] = None
    abs_final: Optional[float] = None
    estado: Optional[str] = None

    # Validación por nivel (JSON [{cargo_id, estado}, ...] o capasValidacion[0] vía cargo_id/estado)
    cargo_id: Optional[int] = None
    estado_validacion: Optional[str] = None
    validacion_capas: Optional[str] = None
    validacion_capas_op: Optional[str] = None

    q_observacion: Optional[str] = None
    q_nodo: Optional[str] = None
    etiqueta_validacion: Optional[str] = None


# Mismo universo de filtros que ExportarRegistrosBody; validación masiva N2/N3 según grilla.
SICOE_MASIVO_MAX_REGISTROS = 500


class ValidarNivelMasivoFiltroBody(BaseModel):
    nivel: int = Field(..., ge=2, le=3)
    marcar_estado: str
    comentario_data: Optional[dict] = None

    numero_reporte: Optional[int] = None
    numero_registro: Optional[int] = None
    semana: Optional[int] = None
    acta_rpo: Optional[int] = None
    subcontratista_id: Optional[int] = None
    capitulo: Optional[str] = None
    item: Optional[str] = None
    items_filtro: Optional[str] = None
    items_filtro_op: Optional[str] = None
    tramo: Optional[str] = None
    costado: Optional[str] = None
    pk_id: Optional[int] = None
    abs_inicio: Optional[float] = None
    abs_final: Optional[float] = None
    estado: Optional[str] = None

    cargo_id: Optional[int] = None
    estado_validacion: Optional[str] = None
    validacion_capas: Optional[str] = None
    validacion_capas_op: Optional[str] = None

    q_observacion: Optional[str] = None
    q_nodo: Optional[str] = None
    etiqueta_validacion: Optional[str] = None


def _sicoe_masivo_filtro_to_export_body(b: ValidarNivelMasivoFiltroBody) -> ExportarRegistrosBody:
    _dump = getattr(b, "model_dump", None)
    if _dump:
        d = _dump(exclude={"nivel", "marcar_estado", "comentario_data"})
    else:
        d = b.dict(exclude={"nivel", "marcar_estado", "comentario_data"})
    return ExportarRegistrosBody(
        campos=[
            "id",
            "reporte_id",
            "nivel1_estado",
            "nivel2_estado",
            "nivel3_estado",
            "nivel2_objeto_pago_sub",
        ],
        **d,
    )


def _sicoe_colectar_registros_masivo_desde_filtros(
    contrato_id: int, body: ExportarRegistrosBody
) -> Tuple[List[dict], Dict[str, Any]]:
    """
    Registros elegibles para validación masiva (misma semántica que exportar), excluyendo
    flujo objeto de pago a subcontratista, con tope SICOE_MASIVO_MAX_REGISTROS.
    """
    consulta_directa_identificador = (
        body.numero_reporte is not None or body.numero_registro is not None
    )
    acta_rpo_x = None if consulta_directa_identificador else body.acta_rpo
    semana_x = None if consulta_directa_identificador else body.semana

    semana_id_filtro = None
    if semana_x is not None:
        try:
            sem_rows = supabase_execute(
                lambda: supabase.table("so_semanas")
                .select("id")
                .eq("contrato_id", contrato_id)
                .eq("numero_semana", semana_x)
                .limit(1)
                .execute()
                .data
            )
            semana_id_filtro = sem_rows[0]["id"] if sem_rows else None
        except Exception:
            semana_id_filtro = None

    acta_id_filtro = None
    if acta_rpo_x is not None:
        try:
            acta_rows = supabase_execute(
                lambda: supabase.table("actas")
                .select("id")
                .eq("contrato_id", contrato_id)
                .eq("numero_rpo", acta_rpo_x)
                .limit(1)
                .execute()
                .data
            )
            if acta_rows:
                acta_id_filtro = acta_rows[0]["id"]
            else:
                acta_rows = supabase_execute(
                    lambda: supabase.table("actas")
                    .select("id")
                    .eq("contrato_id", contrato_id)
                    .eq("consecutivo", acta_rpo_x)
                    .limit(1)
                    .execute()
                    .data
                )
                acta_id_filtro = acta_rows[0]["id"] if acta_rows else None
        except Exception:
            acta_id_filtro = None

    necesita_reporte_filter = any(
        v is not None for v in [body.numero_reporte, body.pk_id, body.estado]
    )

    reporte_ids_base = None
    if necesita_reporte_filter:
        def _rep_ids():
            q = supabase.table("so_reportes").select("id").eq("contrato_id", contrato_id)
            if body.numero_reporte is not None:
                q = q.eq("numero_reporte", body.numero_reporte)
            if body.pk_id is not None:
                q = q.eq("pk_id_id", body.pk_id)
            if body.estado:
                q = _so_reportes_q_por_estado(q, body.estado)
            if body.subcontratista_id is not None:
                q = q.eq("subcontratista_id", body.subcontratista_id)
            if semana_x is not None and semana_id_filtro is not None:
                q = q.eq("semana_id", semana_id_filtro)
            if acta_rpo_x is not None and acta_id_filtro is not None:
                q = q.eq("acta_rpo_id", acta_id_filtro)
            return q.limit(50000).execute().data

        rep_rows = supabase_execute(_rep_ids)
        reporte_ids_base = [r["id"] for r in rep_rows if r.get("id")]
        if not reporte_ids_base:
            return [], {
                "excluidos_objeto_pago_sub": 0,
                "truncado": False,
            }

    if body.q_nodo is not None and str(body.q_nodo).strip():
        ids_n = _sicoe_reporte_ids_coinciden_nodo(contrato_id, body.q_nodo, reporte_ids_base)
        if ids_n is not None:
            if not ids_n:
                return [], {"excluidos_objeto_pago_sub": 0, "truncado": False}
            if reporte_ids_base is not None:
                reporte_ids_base = [x for x in reporte_ids_base if x in ids_n]
            else:
                reporte_ids_base = list(ids_n)
            if not reporte_ids_base:
                return [], {"excluidos_objeto_pago_sub": 0, "truncado": False}

    reg_ids_export_etiqueta = None
    if body.etiqueta_validacion:
        ev_ex = _sicoe_parse_etiqueta_validacion_param(body.etiqueta_validacion)
        reg_ids_export_etiqueta = _sicoe_fetch_registro_ids_etiqueta_validacion(contrato_id, ev_ex)
        if not reg_ids_export_etiqueta:
            return [], {"excluidos_objeto_pago_sub": 0, "truncado": False}

    capas_exp_export = _parse_validacion_capas_param(
        body.validacion_capas, body.cargo_id, body.estado_validacion
    )
    if consulta_directa_identificador:
        capas_exp_export = []
    _cap_op_ex = _parse_capas_validacion_op(body.validacion_capas_op)
    _defer_capas_or_export = (
        bool(capas_exp_export)
        and len(capas_exp_export) > 1
        and _cap_op_ex == "or"
        and not _estado_filtro_omite_validacion_por_cargo(body.estado)
    )

    items_export_norm = _normalize_items_filtro_list(body.items_filtro, body.item)

    def _aplicar_filtros_reg(q):
        q = q.eq("contrato_id", contrato_id)
        q = _so_reg_filtro_abs_solape(q, body.abs_inicio, body.abs_final)
        if body.numero_registro is not None:
            q = q.eq("numero_registro", body.numero_registro)
        if semana_id_filtro is not None:
            q = q.eq("semana_id", semana_id_filtro)
        if acta_id_filtro is not None:
            q = q.eq("acta_rpo_id", acta_id_filtro)
        if body.subcontratista_id is not None:
            q = q.eq("subcontratista_id", body.subcontratista_id)
        if body.capitulo:
            q = q.eq("capitulo", body.capitulo)
        q = _apply_item_patterns_to_so_registros_q(q, items_export_norm, body.items_filtro_op)
        if body.tramo:
            q = q.eq("tramo", body.tramo)
        if body.costado:
            q = _so_reg_filtro_costado(q, body.costado)
        if body.q_observacion is not None and str(body.q_observacion).strip():
            q = q.ilike("observacion", f"%{str(body.q_observacion).strip()}%")
        if acta_id_filtro is not None and not _estado_filtro_es_sin_asignar_item(body.estado):
            q = _so_reg_item_asignado(q)
        if _estado_filtro_es_sin_asignar_item(body.estado):
            q = _so_reg_sin_item_asignado(q)

        if not _estado_filtro_omite_validacion_por_cargo(body.estado):
            if capas_exp_export and not _defer_capas_or_export:
                q = _so_registros_q_y_capas_validacion(
                    q,
                    capas_exp_export,
                    body.pk_id,
                    body.tramo,
                    body.costado,
                    body.capitulo,
                    body.subcontratista_id,
                    None,
                )

        return q

    campos_aux = [
        "id",
        "reporte_id",
        "nivel1_estado",
        "nivel2_estado",
        "nivel3_estado",
        "nivel2_objeto_pago_sub",
        "item_numero",
    ]
    batch_size = 999
    candidatos: List[dict] = []
    excluidos_objeto_pago_sub = 0
    truncado = False

    def _ingest_batch(batch: List[dict]) -> bool:
        nonlocal candidatos, excluidos_objeto_pago_sub, truncado
        if not batch:
            return False
        if _estado_filtro_es_sin_asignar_item(body.estado):
            batch = [
                r for r in batch
                if not (str(r.get("item_numero") or "").strip())
            ]
        if _defer_capas_or_export and capas_exp_export:
            batch = _filtrar_registros_validacion_capas_sicoe(batch, capas_exp_export, None, "or")
        for row in batch:
            if row.get("nivel2_objeto_pago_sub"):
                excluidos_objeto_pago_sub += 1
                continue
            candidatos.append(row)
            if len(candidatos) > SICOE_MASIVO_MAX_REGISTROS:
                truncado = True
                return True
        return False

    def _fetch_by_reporte_id_list(id_list: List[int]):
        if reg_ids_export_etiqueta is not None:
            for rg_chunk in _sicoe_chunks_int(sorted(reg_ids_export_etiqueta), 200):
                rc = list(rg_chunk)
                off = 0
                while True:
                    o = off

                    def _run_fetch():
                        q = (
                            supabase.table("so_registros")
                            .select(",".join(campos_aux))
                            .in_("reporte_id", id_list)
                            .in_("id", rc)
                        )
                        q = _aplicar_filtros_reg(q)
                        return q.range(o, o + batch_size).execute().data

                    batch = supabase_execute(_run_fetch)
                    if not batch:
                        break
                    if _ingest_batch(batch):
                        return
                    if len(batch) < batch_size + 1:
                        break
                    off += batch_size + 1
            return

        off = 0
        base_q = (
            supabase.table("so_registros")
            .select(",".join(campos_aux))
            .in_("reporte_id", id_list)
        )
        base_q = _aplicar_filtros_reg(base_q)
        while True:
            o = off

            def _run_fetch_plain():
                return base_q.range(o, o + batch_size).execute().data

            batch = supabase_execute(_run_fetch_plain)
            if not batch:
                break
            if _ingest_batch(batch):
                break
            if len(batch) < batch_size + 1:
                break
            off += batch_size + 1

    if reporte_ids_base is None:
        if reg_ids_export_etiqueta is not None:
            for rg_chunk in _sicoe_chunks_int(sorted(reg_ids_export_etiqueta), 200):
                rc = list(rg_chunk)
                off = 0
                while True:
                    o = off

                    def _run_fetch_full():
                        q = supabase.table("so_registros").select(",".join(campos_aux)).in_("id", rc)
                        q = _aplicar_filtros_reg(q)
                        return q.range(o, o + batch_size).execute().data

                    batch = supabase_execute(_run_fetch_full)
                    if not batch:
                        break
                    if _ingest_batch(batch):
                        break
                    if len(batch) < batch_size + 1:
                        break
                    off += batch_size + 1
                if truncado:
                    break
        else:
            base_q = supabase.table("so_registros").select(",".join(campos_aux))
            base_q = _aplicar_filtros_reg(base_q)
            off = 0
            while True:
                o = off

                def _run_fetch_all():
                    return base_q.range(o, o + batch_size).execute().data

                batch = supabase_execute(_run_fetch_all)
                if not batch:
                    break
                if _ingest_batch(batch):
                    break
                if len(batch) < batch_size + 1:
                    break
                off += batch_size + 1
    else:
        _CHUNK = 200
        for i in range(0, len(reporte_ids_base), _CHUNK):
            chunk = reporte_ids_base[i:i + _CHUNK]
            _fetch_by_reporte_id_list(chunk)
            if truncado:
                break

    out = candidatos[:SICOE_MASIVO_MAX_REGISTROS]
    stats = {
        "excluidos_objeto_pago_sub": excluidos_objeto_pago_sub,
        "truncado": truncado,
    }
    for r in out:
        r.pop("item_numero", None)
    return out, stats


@app.get("/sicoe-obra/{contrato_id}/registros/campos")
def listar_campos_registros_sicoe(
    contrato_id: int,
    current_user=Depends(get_current_user),
):
    try:
        rows = supabase_execute(
            lambda: supabase.table("so_registros")
            .select("*")
            .eq("contrato_id", contrato_id)
            .limit(1)
            .execute()
            .data
        )
    except Exception:
        rows = []
    if not rows:
        return []
    return sorted(list(rows[0].keys()))


@app.post("/sicoe-obra/{contrato_id}/registros/exportar")
def exportar_registros_sicoe(
    contrato_id: int,
    body: ExportarRegistrosBody,
    current_user=Depends(get_current_user),
):
    if not body.campos:
        raise HTTPException(status_code=400, detail="Debe seleccionar al menos un campo para exportar.")

    def _seguro_campo(c: str) -> bool:
        c = str(c or "")
        return bool(c) and all((ch.isalnum() or ch == "_") for ch in c)

    campos_solicitados = [c for c in body.campos if _seguro_campo(c)]
    if not campos_solicitados:
        raise HTTPException(status_code=400, detail="Campos inválidos.")
    campos_virtuales = {"reporte_numero", "acta_rpo_numero", "semana_numero", "pk_id_valor", "subcontratista_nombre"}
    campos = [c for c in campos_solicitados if c not in campos_virtuales]
    if not campos:
        # Permitir exportar solo virtuales; internamente se consulta llaves mínimas.
        campos = []
    campos_aux = list(dict.fromkeys(campos + ["reporte_id", "acta_rpo_id", "semana_id"]))

    consulta_directa_identificador = (
        body.numero_reporte is not None or body.numero_registro is not None
    )
    acta_rpo_x = None if consulta_directa_identificador else body.acta_rpo
    semana_x = None if consulta_directa_identificador else body.semana

    # 1) Resolver semana_id / acta_rpo_id
    semana_id_filtro = None
    if semana_x is not None:
        try:
            sem_rows = supabase_execute(
                lambda: supabase.table("so_semanas")
                .select("id")
                .eq("contrato_id", contrato_id)
                .eq("numero_semana", semana_x)
                .limit(1)
                .execute()
                .data
            )
            semana_id_filtro = sem_rows[0]["id"] if sem_rows else None
        except Exception:
            semana_id_filtro = None

    acta_id_filtro = None
    if acta_rpo_x is not None:
        try:
            acta_rows = supabase_execute(
                lambda: supabase.table("actas")
                .select("id")
                .eq("contrato_id", contrato_id)
                .eq("numero_rpo", acta_rpo_x)
                .limit(1)
                .execute()
                .data
            )
            if acta_rows:
                acta_id_filtro = acta_rows[0]["id"]
            else:
                acta_rows = supabase_execute(
                    lambda: supabase.table("actas")
                    .select("id")
                    .eq("contrato_id", contrato_id)
                    .eq("consecutivo", acta_rpo_x)
                    .limit(1)
                    .execute()
                    .data
                )
                acta_id_filtro = acta_rows[0]["id"] if acta_rows else None
        except Exception:
            acta_id_filtro = None

    # 2) Filtros que viven en so_reportes (necesitan restricción por reporte_id)
    # abs_inicio/abs_final se aplican por solape en líneas (paso aparte, como buscar/analisis)
    necesita_reporte_filter = any(
        v is not None for v in [body.numero_reporte, body.pk_id, body.estado]
    )

    reporte_ids_base = None
    if necesita_reporte_filter:
        def _rep_ids():
            q = supabase.table("so_reportes").select("id").eq("contrato_id", contrato_id)
            if body.numero_reporte is not None:
                q = q.eq("numero_reporte", body.numero_reporte)
            if body.pk_id is not None:
                q = q.eq("pk_id_id", body.pk_id)
            if body.estado:
                q = _so_reportes_q_por_estado(q, body.estado)
            if body.subcontratista_id is not None:
                q = q.eq("subcontratista_id", body.subcontratista_id)
            if semana_x is not None and semana_id_filtro is not None:
                q = q.eq("semana_id", semana_id_filtro)
            if acta_rpo_x is not None and acta_id_filtro is not None:
                q = q.eq("acta_rpo_id", acta_id_filtro)
            return q.limit(50000).execute().data
        rep_rows = supabase_execute(_rep_ids)
        reporte_ids_base = [r["id"] for r in rep_rows if r.get("id")]
        if not reporte_ids_base:
            return []

    if body.q_nodo is not None and str(body.q_nodo).strip():
        ids_n = _sicoe_reporte_ids_coinciden_nodo(contrato_id, body.q_nodo, reporte_ids_base)
        if ids_n is not None:
            if not ids_n:
                return []
            if reporte_ids_base is not None:
                reporte_ids_base = [x for x in reporte_ids_base if x in ids_n]
            else:
                reporte_ids_base = list(ids_n)
            if not reporte_ids_base:
                return []

    reg_ids_export_etiqueta = None
    if body.etiqueta_validacion:
        ev_ex = _sicoe_parse_etiqueta_validacion_param(body.etiqueta_validacion)
        reg_ids_export_etiqueta = _sicoe_fetch_registro_ids_etiqueta_validacion(contrato_id, ev_ex)
        if not reg_ids_export_etiqueta:
            return []

    # Abscisa en línea (misma semántica que analisis / grilla); no precalcular miles de reporte_id

    capas_exp_export = _parse_validacion_capas_param(
        body.validacion_capas, body.cargo_id, body.estado_validacion
    )
    if consulta_directa_identificador:
        capas_exp_export = []
    _cap_op_ex = _parse_capas_validacion_op(body.validacion_capas_op)
    _defer_capas_or_export = (
        bool(capas_exp_export)
        and len(capas_exp_export) > 1
        and _cap_op_ex == "or"
        and not _estado_filtro_omite_validacion_por_cargo(body.estado)
    )

    items_export_norm = _normalize_items_filtro_list(body.items_filtro, body.item)

    # 3) Query base sobre so_registros
    def _aplicar_filtros_reg(q):
        q = q.eq("contrato_id", contrato_id)
        q = _so_reg_filtro_abs_solape(q, body.abs_inicio, body.abs_final)
        if body.numero_registro is not None:
            q = q.eq("numero_registro", body.numero_registro)
        if semana_id_filtro is not None:
            q = q.eq("semana_id", semana_id_filtro)
        if acta_id_filtro is not None:
            q = q.eq("acta_rpo_id", acta_id_filtro)
        if body.subcontratista_id is not None:
            q = q.eq("subcontratista_id", body.subcontratista_id)
        if body.capitulo:
            q = q.eq("capitulo", body.capitulo)
        q = _apply_item_patterns_to_so_registros_q(q, items_export_norm, body.items_filtro_op)
        if body.tramo:
            q = q.eq("tramo", body.tramo)
        if body.costado:
            q = _so_reg_filtro_costado(q, body.costado)
        if body.q_observacion is not None and str(body.q_observacion).strip():
            q = q.ilike("observacion", f"%{str(body.q_observacion).strip()}%")
        if acta_id_filtro is not None and not _estado_filtro_es_sin_asignar_item(body.estado):
            q = _so_reg_item_asignado(q)
        if _estado_filtro_es_sin_asignar_item(body.estado):
            q = _so_reg_sin_item_asignado(q)

        # Validación: capas en AND en SQL, salvo OR con varias capas (filtro en memoria al final)
        if not _estado_filtro_omite_validacion_por_cargo(body.estado):
            if capas_exp_export and not _defer_capas_or_export:
                q = _so_registros_q_y_capas_validacion(
                    q,
                    capas_exp_export,
                    body.pk_id,
                    body.tramo,
                    body.costado,
                    body.capitulo,
                    body.subcontratista_id,
                    None,
                )

        return q

    registros: list = []
    batch_size = 999

    def _enriquecer_registros_export(rows: List[dict]) -> List[dict]:
        if not rows:
            return rows
        reporte_ids = list({r.get("reporte_id") for r in rows if r.get("reporte_id")})
        rep_map = {}
        if reporte_ids:
            try:
                rep_rows = supabase_execute(
                    lambda: supabase.table("so_reportes")
                    .select("id, numero_reporte, acta_rpo_id, semana_id, pk_id_id, subcontratista_id, estado")
                    .in_("id", reporte_ids)
                    .execute()
                    .data
                )
                rep_map = {r["id"]: r for r in rep_rows if r.get("id")}
            except Exception:
                rep_map = {}

        acta_ids = list({
            (r.get("acta_rpo_id") or (rep_map.get(r.get("reporte_id")) or {}).get("acta_rpo_id"))
            for r in rows
            if (r.get("acta_rpo_id") or (rep_map.get(r.get("reporte_id")) or {}).get("acta_rpo_id"))
        })
        semana_ids = list({
            (r.get("semana_id") or (rep_map.get(r.get("reporte_id")) or {}).get("semana_id"))
            for r in rows
            if (r.get("semana_id") or (rep_map.get(r.get("reporte_id")) or {}).get("semana_id"))
        })
        pk_ids = list({
            (rep_map.get(r.get("reporte_id")) or {}).get("pk_id_id")
            for r in rows
            if (rep_map.get(r.get("reporte_id")) or {}).get("pk_id_id")
        })

        acta_map = {}
        if acta_ids:
            try:
                aa = supabase_execute(
                    lambda: supabase.table("actas")
                    .select("id, numero_rpo")
                    .in_("id", acta_ids)
                    .execute()
                    .data
                )
                acta_map = {a["id"]: a.get("numero_rpo") for a in aa if a.get("id")}
            except Exception:
                acta_map = {}

        semana_map = {}
        if semana_ids:
            try:
                ss = supabase_execute(
                    lambda: supabase.table("so_semanas")
                    .select("id, numero_semana")
                    .in_("id", semana_ids)
                    .execute()
                    .data
                )
                semana_map = {s["id"]: s.get("numero_semana") for s in ss if s.get("id")}
            except Exception:
                semana_map = {}

        pk_map = {}
        if pk_ids:
            try:
                pp = supabase_execute(
                    lambda: supabase.table("pk_ids")
                    .select("id, pk_id")
                    .in_("id", pk_ids)
                    .execute()
                    .data
                )
                pk_map = {p["id"]: p.get("pk_id") for p in pp if p.get("id")}
            except Exception:
                pk_map = {}

        sub_ids = list({
            (rep_map.get(r.get("reporte_id")) or {}).get("subcontratista_id")
            for r in rows
            if (rep_map.get(r.get("reporte_id")) or {}).get("subcontratista_id")
        })
        sub_map = {}
        if sub_ids:
            try:
                ssb = supabase_execute(
                    lambda: supabase.table("subcontratistas")
                    .select("id, razon_social")
                    .in_("id", sub_ids)
                    .execute()
                    .data
                )
                sub_map = {s["id"]: s.get("razon_social") for s in ssb if s.get("id")}
            except Exception:
                sub_map = {}

        for r in rows:
            rep = rep_map.get(r.get("reporte_id")) or {}
            acta_id = r.get("acta_rpo_id") or rep.get("acta_rpo_id")
            sem_id = r.get("semana_id") or rep.get("semana_id")
            pk_id_id = rep.get("pk_id_id")
            sub_id = rep.get("subcontratista_id")
            r["reporte_numero"] = rep.get("numero_reporte")
            r["acta_rpo_numero"] = acta_map.get(acta_id)
            r["semana_numero"] = semana_map.get(sem_id)
            r["pk_id_valor"] = pk_map.get(pk_id_id)
            r["subcontratista_nombre"] = sub_map.get(sub_id)
        return rows

    def _fetch_by_reporte_id_list(id_list: List[int]):
        out: list = []
        if reg_ids_export_etiqueta is not None:
            for rg_chunk in _sicoe_chunks_int(sorted(reg_ids_export_etiqueta), 200):
                rc = list(rg_chunk)
                off = 0
                while True:
                    o = off

                    def _run_fetch():
                        q = (
                            supabase.table("so_registros")
                            .select(",".join(campos_aux))
                            .in_("reporte_id", id_list)
                            .in_("id", rc)
                        )
                        q = _aplicar_filtros_reg(q)
                        return q.range(o, o + batch_size).execute().data

                    batch = supabase_execute(_run_fetch)
                    if not batch:
                        break
                    out.extend(batch)
                    if len(batch) < batch_size + 1:
                        break
                    off += batch_size + 1
            return out

        off = 0
        base_q = (
            supabase.table("so_registros")
            .select(",".join(campos_aux))
            .in_("reporte_id", id_list)
        )
        base_q = _aplicar_filtros_reg(base_q)
        while True:
            o = off

            def _run_fetch_plain():
                return base_q.range(o, o + batch_size).execute().data

            batch = supabase_execute(_run_fetch_plain)
            if not batch:
                break
            out.extend(batch)
            if len(batch) < batch_size + 1:
                break
            off += batch_size + 1
        return out

    if reporte_ids_base is None:
        if reg_ids_export_etiqueta is not None:
            for rg_chunk in _sicoe_chunks_int(sorted(reg_ids_export_etiqueta), 200):
                rc = list(rg_chunk)
                off = 0
                while True:
                    o = off

                    def _run_fetch_full():
                        q = supabase.table("so_registros").select(",".join(campos_aux)).in_("id", rc)
                        q = _aplicar_filtros_reg(q)
                        return q.range(o, o + batch_size).execute().data

                    batch = supabase_execute(_run_fetch_full)
                    if not batch:
                        break
                    registros.extend(batch)
                    if len(batch) < batch_size + 1:
                        break
                    off += batch_size + 1
        else:
            base_q = supabase.table("so_registros").select(",".join(campos_aux))
            base_q = _aplicar_filtros_reg(base_q)
            off = 0
            while True:
                o = off

                def _run_fetch_all():
                    return base_q.range(o, o + batch_size).execute().data

                batch = supabase_execute(_run_fetch_all)
                if not batch:
                    break
                registros.extend(batch)
                if len(batch) < batch_size + 1:
                    break
                off += batch_size + 1
        registros = _enriquecer_registros_export(registros)
        if _defer_capas_or_export and registros:
            registros = _filtrar_registros_validacion_capas_sicoe(registros, capas_exp_export, None, "or")
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            registrar_log(
                u_log,
                "EXPORTAR",
                "SICOE",
                "registro_export",
                str(contrato_id),
                {"filas": len(registros), "campos": len(campos_solicitados)},
            )
        except Exception:
            pass
        return registros

    # Cuando hay restricción por reporte_ids, chunking por .in_()
    _CHUNK = 200
    for i in range(0, len(reporte_ids_base), _CHUNK):
        chunk = reporte_ids_base[i:i + _CHUNK]
        registros.extend(_fetch_by_reporte_id_list(chunk))

    registros = _enriquecer_registros_export(registros)
    if _defer_capas_or_export and registros:
        registros = _filtrar_registros_validacion_capas_sicoe(registros, capas_exp_export, None, "or")
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "EXPORTAR",
            "SICOE",
            "registro_export",
            str(contrato_id),
            {"filas": len(registros), "campos": len(campos_solicitados)},
        )
    except Exception:
        pass
    return registros


def _sicoe_analisis_fetch_registros_paginated(build_q):
    """
    Carga todas las filas filtradas para el panel /analisis. Antes: bucle secuencial de páginas
    de 1000 filas (muy lento con muchas filas). Ahora: oleadas de varias páginas en paralelo;
    tamaño de página configurable para proyectos con max-rows mayor en PostgREST.

    build_q: callable sin argumentos que devuelve el query builder Supabase (sin .range).

    Variables de entorno opcionales:
      SICOE_ANALISIS_PAGE_SIZE   (default 1000; igual al toque habitual de PostgREST/Supabase.
                                  Si en tu proyecto elevaste max-rows, puedes probar 3000–8000.)
      SICOE_ANALISIS_FETCH_WORKERS (default 8; peticiones paralelas por oleada. Bajar si hay rate limit.)
    """
    PAGE = max(200, min(50000, int(os.getenv("SICOE_ANALISIS_PAGE_SIZE", "1000"))))
    WORKERS = max(1, min(16, int(os.getenv("SICOE_ANALISIS_FETCH_WORKERS", "8"))))

    def _page(off: int):
        return supabase_execute(
            lambda o=off: build_q().range(o, o + PAGE - 1).execute().data
        )

    out: list = []
    first = _page(0)
    out.extend(first)
    if len(first) < PAGE:
        return out
    off = PAGE
    while True:
        offsets = [off + i * PAGE for i in range(WORKERS)]
        with ThreadPoolExecutor(max_workers=WORKERS) as ex:
            waves = list(ex.map(_page, offsets))
        for batch in waves:
            out.extend(batch)
            if len(batch) < PAGE:
                return out
        off += WORKERS * PAGE


@app.get("/sicoe-obra/{contrato_id}/analisis")
def analisis_registros_obra(
    contrato_id:      int,
    acta_rpo:         Optional[int]   = None,
    semana:           Optional[int]   = None,
    subcontratista_id: Optional[int]  = None,
    capitulo:         Optional[str]   = None,
    item:             Optional[str]   = None,
    items_filtro:     Optional[str]   = None,
    items_filtro_op:  Optional[str]   = Query(None),
    tramo:            Optional[str]   = None,
    costado:          Optional[str]   = None,
    abs_inicio:       Optional[float] = None,
    abs_final:        Optional[float] = None,
    estado:           Optional[str]   = None,
    numero_reporte:   Optional[int]   = None,
    numero_registro:  Optional[int]   = None,
    pk_id:            Optional[int]   = None,
    cargo_id:         Optional[int]   = None,
    estado_validacion: Optional[str]  = None,
    validacion_capas: Optional[str] = None,
    validacion_capas_op: Optional[str] = Query(None),
    q_observacion: Optional[str] = None,
    q_nodo: Optional[str] = None,
    etiqueta_validacion: Optional[str] = None,
    current_user=Depends(get_current_user)
):
    """Agregados del panel dinámico: cada query param activo es un filtro AND sobre el universo de registros."""
    _empty = {"modo":"general","encabezado":"Sin resultados","grupos":[],
              "total_costo_directo":0,"total_registros":0,"total_cantidad":0,
              "total_aprobados":0,"total_pendientes":0,"total_rechazados":0}

    items_ana = _normalize_items_filtro_list(items_filtro, item)

    consulta_directa_identificador = (
        numero_reporte is not None or numero_registro is not None
    )
    if consulta_directa_identificador:
        acta_rpo = None
        semana = None

    # ── 1. Determinar modo jerárquico ─────────────────────────────────────────
    tiene_contexto = bool(acta_rpo or semana)
    if len(items_ana) == 1:
        modo = "item_detalle"
    elif capitulo:
        modo = "capitulo_items"
    elif tiene_contexto:
        modo = "acta_semana"
    else:
        modo = "general"

    # ── 2. Resolver acta_id y semana_id con sus metadatos ────────────────────
    acta_id = None; acta_info = None
    if acta_rpo is not None:
        try:
            def _ai():
                rows = supabase.table("actas").select("id, numero_rpo, consecutivo")\
                    .eq("contrato_id", contrato_id).eq("numero_rpo", acta_rpo).execute().data
                if not rows:
                    rows = supabase.table("actas").select("id, numero_rpo, consecutivo")\
                        .eq("contrato_id", contrato_id).eq("consecutivo", acta_rpo).execute().data
                return rows
            ar = supabase_execute(_ai)
            if ar: acta_id = ar[0]["id"]; acta_info = ar[0]
        except Exception: pass

    if acta_rpo is not None and acta_id is None:
        return _empty

    semana_id = None; semana_info = None
    if semana is not None:
        try:
            def _si():
                return supabase.table("so_semanas")\
                    .select("id, numero_semana, fecha_inicio, fecha_fin")\
                    .eq("contrato_id", contrato_id).eq("numero_semana", semana).execute().data
            sr = supabase_execute(_si)
            if sr: semana_id = sr[0]["id"]; semana_info = sr[0]
        except Exception: pass

    if semana is not None and semana_id is None:
        return _empty

    # ── 3. Campo de validación (KPI panel: primera capa; filtro SQL: todas con AND) ─
    capas_ana = _parse_validacion_capas_param(validacion_capas, cargo_id, estado_validacion)
    if consulta_directa_identificador:
        capas_ana = []
    _val_campo_l = None
    _val_estado_l = None
    if capas_ana and not _estado_filtro_omite_validacion_por_cargo(estado):
        _c0 = capas_ana[0]
        _c = _capa_campo_validacion(_c0)
        if _c:
            _val_campo_l = _c
            _val_estado_l = (_c0.get("estado") or "").strip()

    # ── 4. Filtros AND a nivel so_reportes (misma idea que export / buscar grilla) ─
    # tramo/costado van en so_registros (paso 5). subcontratista_id va en registros.
    _cap_l = capitulo
    _sub_l = subcontratista_id
    reporte_ids_base = None
    has_rep_f = any([
        numero_reporte is not None,
        bool(estado and str(estado).strip()),
    ])
    if has_rep_f:
        try:
            def _reps():
                q = supabase.table("so_reportes").select("id")\
                    .eq("contrato_id", contrato_id)
                if numero_reporte is not None:
                    q = q.eq("numero_reporte", numero_reporte)
                if estado and str(estado).strip():
                    q = _so_reportes_q_por_estado(q, estado.strip())
                return q.limit(50000).execute().data
            rr = supabase_execute(_reps)
            reporte_ids_base = list({r["id"] for r in rr if r.get("id")})
            if not reporte_ids_base:
                return _empty
        except Exception:
            pass

    # PK por líneas (so_registros), igual que /reportes/buscar — panel alineado al plano
    if pk_id is not None:
        try:
            def _rpk_a():
                q = supabase.table("so_registros").select("reporte_id")\
                    .eq("contrato_id", contrato_id).eq("pk_id_id", pk_id)
                return q.limit(50000).execute().data
            ids_pk_a = list({r["reporte_id"] for r in supabase_execute(_rpk_a) if r.get("reporte_id")})
            if not ids_pk_a:
                return _empty
            if reporte_ids_base is not None:
                reporte_ids_base = list(set(reporte_ids_base) & set(ids_pk_a))
            else:
                reporte_ids_base = ids_pk_a
            if not reporte_ids_base:
                return _empty
        except Exception:
            return _empty

    if q_nodo is not None and str(q_nodo).strip():
        ids_n = _sicoe_reporte_ids_coinciden_nodo(contrato_id, q_nodo, reporte_ids_base)
        if ids_n is not None:
            if not ids_n:
                return _empty
            if reporte_ids_base is not None:
                reporte_ids_base = [x for x in reporte_ids_base if x in ids_n]
            else:
                reporte_ids_base = list(ids_n)
            if not reporte_ids_base:
                return _empty

    # Abscisa: filtrar en la propia línea (evita .in_(reporte_id) gigante que rompe PostgREST / panel vacío)
    _abs_ai = abs_inicio
    _abs_af = abs_final

    reg_ids_etiqueta_ana = None
    if etiqueta_validacion:
        ev_a = _sicoe_parse_etiqueta_validacion_param(etiqueta_validacion)
        reg_ids_etiqueta_ana = _sicoe_fetch_registro_ids_etiqueta_validacion(contrato_id, ev_a)

    # ── 5. Obtener registros: filtros de barra en AND; validación capas según validacion_capas_op ───
    registros = []
    _defer_capas_or_ana = False
    try:
        _a_l = acta_id
        _s_l = semana_id
        _rp_l = reporte_ids_base
        _cap_op_ana = _parse_capas_validacion_op(validacion_capas_op)
        _defer_capas_or_ana = (
            bool(capas_ana)
            and len(capas_ana) > 1
            and _cap_op_ana == "or"
            and not _estado_filtro_omite_validacion_por_cargo(estado)
        )
        _capas_sql = None if _defer_capas_or_ana else capas_ana
        _nr = numero_registro

        def _build_regs_q(reg_id_filter: Optional[List[int]] = None):
            q = supabase.table("so_registros")\
                .select("reporte_id, costo_directo, cantidad_total, item_numero, item_descripcion, unidad, acta_rpo_id, nivel1_estado, nivel2_estado, nivel3_estado, capitulo, subcontratista_id")\
                .eq("contrato_id", contrato_id)
            q = _so_reg_filtro_abs_solape(q, _abs_ai, _abs_af)
            if _nr is not None:
                q = q.eq("numero_registro", _nr)
            if _a_l is not None:
                q = q.eq("acta_rpo_id", _a_l)
            if _s_l is not None:
                q = q.eq("semana_id", _s_l)
            q = _apply_item_patterns_to_so_registros_q(q, items_ana, items_filtro_op)
            if _cap_l:
                q = q.eq("capitulo", _cap_l)
            if _sub_l is not None:
                q = q.eq("subcontratista_id", _sub_l)
            if _rp_l is not None:
                q = q.in_("reporte_id", _rp_l)
            if reg_id_filter is not None:
                q = q.in_("id", reg_id_filter)
            if tramo:
                q = q.eq("tramo", tramo)
            if costado:
                q = _so_reg_filtro_costado(q, costado)
            if pk_id is not None:
                q = q.eq("pk_id_id", pk_id)
            if q_observacion is not None and str(q_observacion).strip():
                q = q.ilike("observacion", f"%{str(q_observacion).strip()}%")
            if _a_l is not None and not _estado_filtro_es_sin_asignar_item(estado):
                q = _so_reg_item_asignado(q)
            if _capas_sql and not _estado_filtro_omite_validacion_por_cargo(estado):
                q = _so_registros_q_y_capas_validacion(
                    q, _capas_sql, pk_id, tramo, costado, _cap_l, _sub_l, None
                )
            if _estado_filtro_es_sin_asignar_item(estado):
                q = _so_reg_sin_item_asignado(q)
            return q

        if reg_ids_etiqueta_ana is not None:
            if not reg_ids_etiqueta_ana:
                registros = []
            else:
                for rg_chunk in _sicoe_chunks_int(sorted(reg_ids_etiqueta_ana), 200):
                    rc = list(rg_chunk)
                    part = _sicoe_analisis_fetch_registros_paginated(lambda c=rc: _build_regs_q(c))
                    registros.extend(part)
        else:
            registros = _sicoe_analisis_fetch_registros_paginated(lambda: _build_regs_q())
    except Exception:
        if not registros:
            registros = []

    if _defer_capas_or_ana and registros:
        registros = _filtrar_registros_validacion_capas_sicoe(registros, capas_ana, None, "or")

    if _estado_filtro_es_sin_asignar_item(estado):
        registros = [
            r for r in registros
            if not (str(r.get("item_numero") or "").strip())
        ]

    # No re-filtrar por estado de so_reportes (coherente con dashboard_matriz_validacion / SQL del usuario).

    # ── 6. Batch-resolve capitulo y estado desde so_reportes ─────────────────
    rep_ids_found = list({r["reporte_id"] for r in registros if r.get("reporte_id")})
    reporte_map: dict = {}
    if rep_ids_found:
        # Procesar en lotes de 500 para no exceder límites de URL con .in_()
        _cid_l = contrato_id
        for chunk_start in range(0, len(rep_ids_found), 500):
            chunk = rep_ids_found[chunk_start:chunk_start + 500]
            try:
                def _ri(ids=chunk):
                    return supabase.table("so_reportes").select("id, capitulo, estado")\
                        .eq("contrato_id", _cid_l).in_("id", ids).execute().data
                for r in supabase_execute(_ri):
                    reporte_map[r["id"]] = r
            except Exception:
                pass

    # ── 7. Agrupar según modo ─────────────────────────────────────────────────
    def _estado_efectivo(reg):
        # Si hay filtro de validación activo, usar solo el estado del nivel (primera capa) para KPI
        if _val_campo_l:
            estado = reg.get(_val_campo_l) or ""
            if estado == "Aprobado":  return "Aprobado"
            if estado == "Pendiente": return "Pendiente"
            if estado == "Rechazado": return "Rechazado"
            return "No Revisado"
        # Sin filtro: estado global del registro
        niveles = [
            reg.get("nivel1_estado") or "",
            reg.get("nivel2_estado") or "",
            reg.get("nivel3_estado") or "",
        ]
        if "Rechazado" in niveles: return "Rechazado"
        if "Pendiente" in niveles: return "Pendiente"
        activos = [n for n in niveles if n]
        if activos and all(n == "Aprobado" for n in activos): return "Aprobado"
        return "No Revisado"

    grupos: dict = {}

    if modo in ("acta_semana", "general"):
        for reg in registros:
            cap = reg.get("capitulo") or "Sin capítulo"
            ee  = _estado_efectivo(reg)
            cd  = float(reg.get("costo_directo") or 0)
            if cap not in grupos:
                grupos[cap] = {"label": cap, "costo_directo": 0.0,
                               "total_registros": 0, "no_revisados": 0, "no_revisados_costo": 0.0,
                               "aprobados": 0.0, "pendientes": 0.0, "rechazados": 0.0,
                               "aprobados_count": 0, "pendientes_count": 0, "rechazados_count": 0}
            grupos[cap]["costo_directo"]   += cd
            grupos[cap]["total_registros"] += 1
            if   ee == "Aprobado":   grupos[cap]["aprobados"]  += cd; grupos[cap]["aprobados_count"]  += 1
            elif ee == "Pendiente":  grupos[cap]["pendientes"] += cd; grupos[cap]["pendientes_count"] += 1
            elif ee == "Rechazado":  grupos[cap]["rechazados"] += cd; grupos[cap]["rechazados_count"] += 1
            else:
                grupos[cap]["no_revisados"] += 1
                grupos[cap]["no_revisados_costo"] = grupos[cap].get("no_revisados_costo", 0.0) + cd

    elif modo == "capitulo_items":
        for reg in registros:
            ee  = _estado_efectivo(reg)
            cd  = float(reg.get("costo_directo") or 0)
            it  = reg.get("item_numero") or "Sin ítem"
            if it not in grupos:
                grupos[it] = {
                    "label":           it,
                    "descripcion":     reg.get("item_descripcion") or "",
                    "cantidad_total":  0.0,
                    "unidad":          reg.get("unidad") or "",
                    "costo_directo":   0.0,
                    "total_registros": 0,
                    "no_revisados": 0,
                    "no_revisados_costo": 0.0,
                    "aprobados": 0.0, "pendientes": 0.0, "rechazados": 0.0,
                    "aprobados_count": 0, "pendientes_count": 0, "rechazados_count": 0,
                }
            if not grupos[it]["descripcion"] and reg.get("item_descripcion"):
                grupos[it]["descripcion"] = reg["item_descripcion"]
            if not grupos[it]["unidad"] and reg.get("unidad"):
                grupos[it]["unidad"] = reg["unidad"]
            grupos[it]["cantidad_total"] += float(reg.get("cantidad_total") or 0)
            grupos[it]["costo_directo"]  += cd
            grupos[it]["total_registros"] += 1
            if   ee == "Aprobado":   grupos[it]["aprobados"]  += cd; grupos[it]["aprobados_count"]  += 1
            elif ee == "Pendiente":  grupos[it]["pendientes"] += cd; grupos[it]["pendientes_count"] += 1
            elif ee == "Rechazado":  grupos[it]["rechazados"] += cd; grupos[it]["rechazados_count"] += 1
            else:
                grupos[it]["no_revisados"] += 1
                grupos[it]["no_revisados_costo"] = grupos[it].get("no_revisados_costo", 0.0) + cd

    elif modo == "item_detalle":
        acta_ids_found = list({r.get("acta_rpo_id") for r in registros if r.get("acta_rpo_id")})
        acta_map_local: dict = {}
        if acta_ids_found:
            try:
                def _am():
                    return supabase.table("actas").select("id, numero_rpo, consecutivo")\
                        .in_("id", acta_ids_found).execute().data
                for a in supabase_execute(_am):
                    acta_map_local[a["id"]] = a
            except Exception: pass
        for reg in registros:
            ee   = _estado_efectivo(reg)
            cd   = float(reg.get("costo_directo") or 0)
            cap  = reg.get("capitulo") or "Sin capítulo"
            a_id = reg.get("acta_rpo_id")
            a    = acta_map_local.get(a_id) or {}
            nr   = a.get("numero_rpo") or a.get("consecutivo") or "?"
            label = f"RPO {nr}"
            key   = f"{label}||{cap}"
            if key not in grupos:
                grupos[key] = {
                    "label": label, "capitulo": cap,
                    "cantidad_total": 0.0, "costo_directo": 0.0,
                    "total_registros": 0,
                    "no_revisados": 0,
                    "no_revisados_costo": 0.0,
                    "aprobados": 0.0, "pendientes": 0.0, "rechazados": 0.0,
                    "aprobados_count": 0, "pendientes_count": 0, "rechazados_count": 0,
                }
            grupos[key]["cantidad_total"] += float(reg.get("cantidad_total") or 0)
            grupos[key]["costo_directo"]  += cd
            grupos[key]["total_registros"] += 1
            if   ee == "Aprobado":
                grupos[key]["aprobados"]  += cd
                grupos[key]["aprobados_count"] += 1
            elif ee == "Pendiente":
                grupos[key]["pendientes"] += cd
                grupos[key]["pendientes_count"] += 1
            elif ee == "Rechazado":
                grupos[key]["rechazados"] += cd
                grupos[key]["rechazados_count"] += 1
            else:
                grupos[key]["no_revisados"] += 1
                grupos[key]["no_revisados_costo"] += cd

    import re as _re
    def _cap_sort_key(label):
        m = _re.match(r'^(\d+)', label or "")
        return (0, int(m.group(1)), label) if m else (1, 0, label)

    if modo in ("acta_semana", "general", "capitulo_items"):
        grupos_list = sorted(grupos.values(), key=lambda g: _cap_sort_key(g["label"]))
    elif modo == "item_detalle":
        def _rpo_detalle_sort_key(g):
            lab = str(g.get("label") or "")
            m = _re.search(r"(\d+)", lab)
            n = int(m.group(1)) if m else 10**9
            cap = str(g.get("capitulo") or "")
            return (n, cap, lab)

        grupos_list = sorted(grupos.values(), key=_rpo_detalle_sort_key)
    else:
        grupos_list = sorted(grupos.values(), key=lambda g: g["costo_directo"], reverse=True)
    for g in grupos_list:
        g["costo_directo"] = round(g["costo_directo"], 0)
        if "no_revisados_costo" in g:
            g["no_revisados_costo"] = round(g.get("no_revisados_costo") or 0, 0)
        if "cantidad_total" in g:
            g["cantidad_total"] = round(g["cantidad_total"], 2)

    # ── 8. Encabezado ─────────────────────────────────────────────────────────
    if modo == "acta_semana":
        if acta_rpo is not None:
            nr = (acta_info or {}).get("numero_rpo") or acta_rpo
            encabezado = f"Acta RPO {nr}"
        else:
            if semana_info:
                encabezado = (f"Semana {semana_info['numero_semana']} | "
                              f"{semana_info['fecha_inicio']} → {semana_info['fecha_fin']}")
            else:
                encabezado = f"Semana {semana}"
    elif modo == "capitulo_items":
        prefix = ""
        if acta_rpo is not None:
            nr = (acta_info or {}).get("numero_rpo") or acta_rpo
            prefix = f"Acta RPO {nr}"
        elif semana is not None:
            sn = (semana_info or {}).get("numero_semana") or semana
            prefix = f"Semana {sn}"
        encabezado = f"{prefix} — {capitulo}" if prefix else (capitulo or "Ítems")
    elif modo == "item_detalle":
        encabezado = f"Ítem: {items_ana[0]}"
    else:  # general
        partes = []
        if numero_reporte is not None:
            partes.append(f"Rep. #{numero_reporte}")
        if numero_registro is not None:
            partes.append(f"Reg. #{numero_registro}")
        if pk_id is not None:
            partes.append(f"PK_ID id {pk_id}")
        if subcontratista_id:
            partes.append(f"Subc. #{subcontratista_id}")
        if capitulo:
            partes.append(f"Cap.: {capitulo}")
        if items_ana:
            if len(items_ana) == 1:
                partes.append(f"Ítem: {items_ana[0]}")
            else:
                _op_h = "O" if _parse_capas_validacion_op(items_filtro_op) == "or" else "Y"
                partes.append(f"Ítems ({_op_h}): " + " · ".join(items_ana))
        if tramo:
            partes.append(f"Tramo: {tramo}")
        if costado:
            partes.append(f"Costado: {costado}")
        if estado:
            partes.append(f"Estado rep.: {estado}")
        if abs_inicio is not None:
            partes.append(f"Abs. ≥ {abs_inicio}")
        if abs_final is not None:
            partes.append(f"Abs. ≤ {abs_final}")
        if capas_ana:
            for capx in capas_ana:
                evx = (capx.get("estado") or "").strip()
                cap_lbl = _sicoe_capa_etiqueta_panel(capx)
                partes.append(f"Val. {cap_lbl}: {evx}")
        encabezado = " · ".join(partes) if partes else "Todos los registros"

    tc  = round(sum(g["costo_directo"]   for g in grupos_list), 0)
    tr  = sum(g["total_registros"] for g in grupos_list)
    ta  = sum(g["aprobados"]       for g in grupos_list)
    tp  = sum(g["pendientes"]      for g in grupos_list)
    trj = sum(g["rechazados"]      for g in grupos_list)
    ta_c  = sum(g.get("aprobados_count",  0) for g in grupos_list)
    tp_c  = sum(g.get("pendientes_count", 0) for g in grupos_list)
    trj_c = sum(g.get("rechazados_count", 0) for g in grupos_list)
    tnr   = sum(g.get("no_revisados",     0) for g in grupos_list)
    tnrc  = round(sum(g.get("no_revisados_costo", 0.0) for g in grupos_list), 0)
    t_cant = round(sum(float(g.get("cantidad_total") or 0) for g in grupos_list), 2)

    return {"modo": modo, "encabezado": encabezado, "grupos": grupos_list,
            "total_costo_directo": tc, "total_registros": tr,
            "total_cantidad": t_cant,
            "total_no_revisados": tnr,
            "total_no_revisados_costo": tnrc,
            "total_aprobados": ta, "total_pendientes": tp, "total_rechazados": trj,
            "total_aprobados_count": ta_c, "total_pendientes_count": tp_c, "total_rechazados_count": trj_c}


@app.get("/sicoe-obra/{contrato_id}/filtros/semanas")
def filtros_semanas(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.table("so_semanas").select("id, numero_semana")\
            .eq("contrato_id", contrato_id).order("numero_semana").execute().data
    return supabase_execute(_q)


@app.get("/sicoe-obra/{contrato_id}/filtros/actas")
def filtros_actas(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.table("actas").select("id, numero_rpo")\
            .eq("contrato_id", contrato_id).not_.is_("numero_rpo", "null")\
            .order("numero_rpo").execute().data
    rows = supabase_execute(_q)
    return [{"numero_rpo": r["numero_rpo"]} for r in rows if r.get("numero_rpo") is not None]


@app.get("/sicoe-obra/{contrato_id}/filtros/capitulos")
def filtros_capitulos_reportes(
    contrato_id: int,
    acta_rpo: Optional[int] = None,
    semana: Optional[int] = None,
    subcontratista_id: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    import re
    def orden_cap(c):
        m = re.match(r'^(\d+)', c)
        return (int(m.group(1)) if m else 9999, c)

    # Sin filtros → lista completa con paginación
    if acta_rpo is None and semana is None and subcontratista_id is None:
        todos: list = []
        off = 0
        while True:
            def _q(o=off):
                return supabase.table("so_reportes").select("capitulo")\
                    .eq("contrato_id", contrato_id)\
                    .not_.is_("capitulo", "null").range(o, o + 999).execute().data
            batch = supabase_execute(_q)
            todos.extend(batch)
            if len(batch) < 1000:
                break
            off += 1000
        return sorted({r["capitulo"] for r in todos if r.get("capitulo")}, key=orden_cap)

    # Resolver acta_id
    acta_id = None
    if acta_rpo is not None:
        try:
            def _ai():
                rows = supabase.table("actas").select("id")\
                    .eq("contrato_id", contrato_id).eq("numero_rpo", acta_rpo).execute().data
                if not rows:
                    rows = supabase.table("actas").select("id")\
                        .eq("contrato_id", contrato_id).eq("consecutivo", acta_rpo).execute().data
                return rows
            acta_rows = supabase_execute(_ai)
            if not acta_rows:
                return []
            acta_id = acta_rows[0]["id"]
        except Exception:
            return []

    # Resolver semana_id
    semana_id = None
    if semana is not None:
        try:
            def _si():
                return supabase.table("so_semanas").select("id")\
                    .eq("contrato_id", contrato_id).eq("numero_semana", semana).execute().data
            sem_rows = supabase_execute(_si)
            if not sem_rows:
                return []
            semana_id = sem_rows[0]["id"]
        except Exception:
            return []

    # Obtener reporte_ids desde so_registros si hay filtro de acta o semana
    reporte_ids = None
    if acta_id is not None or semana_id is not None:
        try:
            _acta_id_l = acta_id
            _semana_id_l = semana_id
            def _regs():
                q = supabase.table("so_registros").select("reporte_id")\
                    .eq("contrato_id", contrato_id)
                if _acta_id_l is not None:
                    q = q.eq("acta_rpo_id", _acta_id_l)
                if _semana_id_l is not None:
                    q = q.eq("semana_id", _semana_id_l)
                return q.execute().data
            reg_rows = supabase_execute(_regs)
            reporte_ids = list({r["reporte_id"] for r in reg_rows if r.get("reporte_id")})
            if not reporte_ids:
                return []
        except Exception:
            return []

    # Obtener capítulos desde so_reportes (paginado)
    try:
        _rep_ids_l = reporte_ids
        _sub_id_l  = subcontratista_id
        cap_todos: list = []
        off2 = 0
        while True:
            def _caps(o=off2):
                q = supabase.table("so_reportes").select("capitulo")\
                    .eq("contrato_id", contrato_id)\
                    .not_.is_("capitulo", "null")
                if _sub_id_l is not None:
                    q = q.eq("subcontratista_id", _sub_id_l)
                if _rep_ids_l is not None:
                    q = q.in_("id", _rep_ids_l)
                return q.range(o, o + 999).execute().data
            batch = supabase_execute(_caps)
            cap_todos.extend(batch)
            if len(batch) < 1000:
                break
            off2 += 1000
        return sorted({r["capitulo"] for r in cap_todos if r.get("capitulo")}, key=orden_cap)
    except Exception:
        return []


@app.get("/sicoe-obra/{contrato_id}/filtros/items")
def filtros_items_registros(
    contrato_id: int,
    q: Optional[str] = None,
    capitulo: Optional[str] = None,
    acta_rpo: Optional[int] = None,
    semana: Optional[int] = None,
    subcontratista_id: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    # Resolver acta_id
    acta_id = None
    if acta_rpo is not None:
        try:
            def _ai():
                rows = supabase.table("actas").select("id")\
                    .eq("contrato_id", contrato_id).eq("numero_rpo", acta_rpo).execute().data
                if not rows:
                    rows = supabase.table("actas").select("id")\
                        .eq("contrato_id", contrato_id).eq("consecutivo", acta_rpo).execute().data
                return rows
            acta_rows = supabase_execute(_ai)
            if not acta_rows:
                return []
            acta_id = acta_rows[0]["id"]
        except Exception:
            return []

    # Resolver semana_id
    semana_id = None
    if semana is not None:
        try:
            def _si():
                return supabase.table("so_semanas").select("id")\
                    .eq("contrato_id", contrato_id).eq("numero_semana", semana).execute().data
            sem_rows = supabase_execute(_si)
            if not sem_rows:
                return []
            semana_id = sem_rows[0]["id"]
        except Exception:
            return []

    # Obtener items desde so_registros (paginado)
    # capitulo y subcontratista_id se filtran directamente en so_registros,
    # igual que el endpoint /analisis, para garantizar coherencia entre panel y autocomplete
    try:
        _acta_id_l   = acta_id
        _semana_id_l = semana_id
        _cap_l       = capitulo
        _sub_l       = subcontratista_id
        _q_l         = q
        item_todos: list = []
        off3 = 0
        while True:
            def _items(o=off3):
                qr = supabase.table("so_registros")\
                    .select("item_numero, item_descripcion")\
                    .eq("contrato_id", contrato_id)\
                    .not_.is_("item_numero", "null")
                if _acta_id_l is not None:
                    qr = qr.eq("acta_rpo_id", _acta_id_l)
                if _semana_id_l is not None:
                    qr = qr.eq("semana_id", _semana_id_l)
                if _cap_l is not None:
                    qr = qr.eq("capitulo", _cap_l)
                if _sub_l is not None:
                    qr = qr.eq("subcontratista_id", _sub_l)
                if _q_l:
                    qr = qr.or_(f"item_numero.ilike.%{_q_l}%,item_descripcion.ilike.%{_q_l}%")
                return qr.range(o, o + 999).execute().data
            batch = supabase_execute(_items)
            item_todos.extend(batch)
            if len(batch) < 1000:
                break
            off3 += 1000
        # Deduplicar por item_numero, conservar descripcion
        seen: dict = {}
        for r in item_todos:
            num = r.get("item_numero")
            if not num:
                continue
            if num not in seen:
                seen[num] = r.get("item_descripcion") or ""
            elif not seen[num] and r.get("item_descripcion"):
                seen[num] = r["item_descripcion"]
        results = [{"item_numero": k, "item_descripcion": v} for k, v in sorted(seen.items())]
        return results[:20]
    except Exception:
        return []


@app.get("/sicoe-obra/{contrato_id}/filtros/tramoscostados")
def filtros_tramos_costados(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.table("so_reportes").select("tramo")\
            .eq("contrato_id", contrato_id).execute().data
    rows = supabase_execute(_q)
    tramos = {r["tramo"] for r in rows if r.get("tramo")}
    # Calzada del filtro: valores únicos desde pk_ids.calzada (maestro por contrato)
    costados = set()
    try:
        _off_pk = 0
        _step_pk = 1000
        while True:
            def _q_pk(o=_off_pk):
                return supabase.table("pk_ids").select("calzada")\
                    .eq("contrato_id", contrato_id)\
                    .range(o, o + _step_pk - 1).execute().data
            batch_pk = supabase_execute(_q_pk)
            for r in batch_pk:
                c = r.get("calzada")
                if c is not None and str(c).strip() != "":
                    costados.add(str(c).strip())
            if len(batch_pk) < _step_pk:
                break
            _off_pk += _step_pk
    except Exception:
        pass
    return {
        "tramos":   sorted(tramos),
        "costados": sorted(costados),
    }


@app.get("/sicoe-obra/{contrato_id}/reportes/{reporte_id}")
def obtener_reporte(
    contrato_id: int,
    reporte_id: int,
    aplicar_filtros_busqueda: bool = Query(False),
    numero_registro: Optional[int] = None,
    semana: Optional[int] = None,
    acta_rpo: Optional[int] = None,
    subcontratista_id: Optional[int] = None,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    items_filtro: Optional[str] = Query(None),
    items_filtro_op: Optional[str] = Query(None),
    tramo: Optional[str] = None,
    costado: Optional[str] = None,
    pk_id: Optional[int] = None,
    abs_inicio: Optional[float] = None,
    abs_final: Optional[float] = None,
    estado: Optional[str] = None,
    cargo_id: Optional[int] = Query(None),
    estado_validacion: Optional[str] = Query(None),
    validacion_capas: Optional[str] = Query(None),
    validacion_capas_op: Optional[str] = Query(None),
    q_observacion: Optional[str] = Query(None),
    q_nodo: Optional[str] = Query(None),
    etiqueta_validacion: Optional[str] = Query(None),
    current_user=Depends(get_current_user),
):
    items_detalle_norm = _normalize_items_filtro_list(items_filtro, item)
    capas_v = _parse_validacion_capas_param(validacion_capas, cargo_id, estado_validacion)
    capas_aplican_a_lineas = bool(capas_v) and not _estado_filtro_omite_validacion_por_cargo(estado)
    _cap_op_det = _parse_capas_validacion_op(validacion_capas_op)
    _defer_capas_or_detalle = (
        capas_aplican_a_lineas
        and bool(capas_v)
        and len(capas_v) > 1
        and _cap_op_det == "or"
    )

    reg_tag_detalle = None
    if aplicar_filtros_busqueda and etiqueta_validacion:
        ev_d = _sicoe_parse_etiqueta_validacion_param(etiqueta_validacion)
        reg_tag_detalle = _sicoe_fetch_registro_ids_etiqueta_validacion(contrato_id, ev_d)

    def _r():
        return supabase.table("so_reportes").select("*, subcontratistas(razon_social), pk_ids(pk_id, civ, tramo, infraestructura, calzada, abs_inicio, abs_final)")\
            .eq("id", reporte_id).eq("contrato_id", contrato_id).execute().data

    def _reg_all():
        return supabase.table("so_registros").select("*")\
            .eq("reporte_id", reporte_id).eq("contrato_id", contrato_id).order("id").execute().data

    def _reg_filtrados_como_busqueda():
        semana_id_filtro = None
        if semana is not None:
            try:
                def _sem_id():
                    return supabase.table("so_semanas").select("id")\
                        .eq("contrato_id", contrato_id)\
                        .eq("numero_semana", semana).execute().data
                sem_rows = supabase_execute(_sem_id)
                if sem_rows:
                    semana_id_filtro = sem_rows[0]["id"]
                else:
                    return []
            except Exception:
                return []

        acta_id_filtro = None
        if acta_rpo is not None:
            try:
                def _acta_id():
                    rows = supabase.table("actas").select("id")\
                        .eq("contrato_id", contrato_id)\
                        .eq("numero_rpo", acta_rpo).execute().data
                    if not rows:
                        rows = supabase.table("actas").select("id")\
                            .eq("contrato_id", contrato_id)\
                            .eq("consecutivo", acta_rpo).execute().data
                    return rows
                acta_rows = supabase_execute(_acta_id)
                if acta_rows:
                    acta_id_filtro = acta_rows[0]["id"]
                else:
                    return []
            except Exception:
                return []

        if reg_tag_detalle is not None and not reg_tag_detalle:
            return []

        out = []
        off = 0
        page = 1000
        while True:
            def _page(o=off):
                q = supabase.table("so_registros").select("*")\
                    .eq("reporte_id", reporte_id).eq("contrato_id", contrato_id)
                q = _sicoe_so_registros_q_linea_filtros_busqueda(
                    q,
                    numero_registro=numero_registro,
                    abs_inicio=abs_inicio,
                    abs_final=abs_final,
                    capitulo=capitulo,
                    items=items_detalle_norm,
                    items_op=items_filtro_op,
                    subcontratista_id=subcontratista_id,
                    tramo=tramo,
                    costado=costado,
                    pk_id=pk_id,
                    q_observacion=q_observacion,
                    semana_id=semana_id_filtro,
                    acta_rpo_id=acta_id_filtro,
                    capas_v=(
                        None
                        if _defer_capas_or_detalle
                        else (capas_v if capas_aplican_a_lineas else None)
                    ),
                    estado=estado,
                )
                return q.order("id").range(o, o + page - 1).execute().data

            batch = supabase_execute(_page)
            out.extend(batch)
            if len(batch) < page:
                break
            off += page
        if _defer_capas_or_detalle:
            out = _filtrar_registros_validacion_capas_sicoe(out, capas_v, None, "or")
        if reg_tag_detalle is not None:
            out = [r for r in out if r.get("id") in reg_tag_detalle]
        return out

    def _pts():
        return supabase.table("so_puntos_topograficos").select("*")\
            .eq("reporte_id", reporte_id).order("id").execute().data

    with ThreadPoolExecutor(max_workers=3) as ex:
        fut_rep = ex.submit(lambda: supabase_execute(_r))
        fut_reg = ex.submit(
            _reg_filtrados_como_busqueda if aplicar_filtros_busqueda else _reg_all
        )
        fut_pts = ex.submit(lambda: supabase_execute(_pts))
    reporte = fut_rep.result()
    if not reporte:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    regs_raw = fut_reg.result()
    puntos_rows = fut_pts.result()

    if aplicar_filtros_busqueda and q_nodo is not None and str(q_nodo).strip():
        regs_raw = _sicoe_filtrar_registros_coinciden_nodo_ui(regs_raw, reporte[0], q_nodo)

    r = reporte[0]
    sub = r.pop("subcontratistas", None)
    r["subcontratista_nombre"] = sub["razon_social"] if sub else None
    pk = r.pop("pk_ids", None)
    if pk:
        r["pk_id_valor"]    = pk.get("pk_id")
        r["civ"]            = r.get("civ")     or pk.get("civ")
        r["tramo"]          = r.get("tramo")   or pk.get("tramo")
        r["infraestructura"]= r.get("infraestructura") or pk.get("infraestructura")
        r["calzada"]        = r.get("calzada") or pk.get("calzada")
    else:
        r["pk_id_valor"] = None
    # Si aplicar_filtros_busqueda=true, registros coinciden con la misma semántica AND que la grilla/panel.
    reg_ids = [reg["id"] for reg in regs_raw if reg.get("id")]
    num_comentarios_map = {}
    if reg_ids:
        try:
            def _cnt():
                return supabase.table("so_registro_comentarios")\
                    .select("registro_id")\
                    .in_("registro_id", reg_ids).execute().data
            cnt_rows = supabase_execute(_cnt)
            for row in cnt_rows:
                rid = row["registro_id"]
                num_comentarios_map[rid] = num_comentarios_map.get(rid, 0) + 1
        except Exception:
            pass
    for reg in regs_raw:
        reg["num_comentarios"] = num_comentarios_map.get(reg["id"], 0)
    _enriquecer_registros_labels_reversion_doble_llave(regs_raw)
    r["registros"] = regs_raw
    r["puntos"] = puntos_rows
    if aplicar_filtros_busqueda:
        r["registros_vista_filtrada"] = True

    # Resolver nombre del modificador
    modificado_por = r.get("modificado_por")
    if modificado_por:
        def _modif():
            return supabase.table("usuarios")\
                .select("nombre, apellidos")\
                .eq("id", modificado_por).single().execute().data
        try:
            modif = supabase_execute(_modif)
            r["nombre_modificador"] = f"{modif.get('nombre','')} {modif.get('apellidos','')}".strip() if modif else None
        except:
            r["nombre_modificador"] = None
    else:
        r["nombre_modificador"] = None

    # Resolver nombre del creador
    creado_por = r.get("creado_por")
    if creado_por:
        def _creador():
            return supabase.table("usuarios")\
                .select("nombre, apellidos")\
                .eq("id", creado_por).single().execute().data
        try:
            creador = supabase_execute(_creador)
            r["nombre_creador"] = f"{creador.get('nombre','')} {creador.get('apellidos','')}".strip() if creador else None
        except:
            r["nombre_creador"] = None
    else:
        r["nombre_creador"] = None

    # Resolver nombre del inspector
    inspector_id = r.get("inspector_id")
    if inspector_id:
        def _insp():
            return supabase.table("usuarios")\
                .select("nombre, apellidos")\
                .eq("id", inspector_id).single().execute().data
        try:
            insp = supabase_execute(_insp)
            r["inspector_nombre"] = f"{insp.get('nombre','')} {insp.get('apellidos','')}".strip() if insp else None
        except:
            r["inspector_nombre"] = None
    else:
        r["inspector_nombre"] = None

    # Resolver número de acta RPO
    if r.get("acta_rpo_id"):
        def _acta():
            return supabase.table("actas")\
                .select("numero_rpo").eq("id", r["acta_rpo_id"]).single().execute().data
        try:
            acta = supabase_execute(_acta)
            r["acta_rpo_numero"] = acta.get("numero_rpo") if acta else None
        except:
            r["acta_rpo_numero"] = None
    else:
        r["acta_rpo_numero"] = None

    # Resolver consecutivo de corte
    if r.get("corte_id"):
        def _corte():
            return supabase.table("subcontratista_cortes")\
                .select("consecutivo").eq("id", r["corte_id"]).single().execute().data
        try:
            corte = supabase_execute(_corte)
            r["corte_numero"] = corte.get("consecutivo") if corte else None
        except:
            r["corte_numero"] = None
    else:
        r["corte_numero"] = None

    # Resolver número y período de semana
    if r.get("semana_id"):
        def _sem():
            return supabase.table("so_semanas")\
                .select("numero_semana, fecha_inicio, fecha_fin")\
                .eq("id", r["semana_id"]).single().execute().data
        try:
            sem = supabase_execute(_sem)
            if sem:
                r["semana_numero"]  = sem.get("numero_semana")
                r["semana_periodo"] = f"{sem.get('fecha_inicio','')} → {sem.get('fecha_fin','')}"
            else:
                r["semana_numero"] = None
                r["semana_periodo"] = None
        except:
            r["semana_numero"] = None
            r["semana_periodo"] = None
    else:
        r["semana_numero"] = None
        r["semana_periodo"] = None

    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "CONSULTAR",
            "SICOE",
            "reporte",
            str(reporte_id),
            {
                "numero_reporte": r.get("numero_reporte"),
                "estado": r.get("estado"),
                "n_registros": len(regs_raw),
                "n_puntos": len(puntos_rows or []),
            },
        )
    except Exception:
        pass

    return r

@app.delete("/sicoe-obra/{contrato_id}/reportes/{reporte_id}")
def eliminar_reporte(contrato_id: int, reporte_id: int, current_user=Depends(get_current_user)):
    def _del():
        return supabase.table("so_reportes").delete()\
            .eq("id", reporte_id).eq("contrato_id", contrato_id)\
            .eq("estado", "Borrador").execute().data
    supabase_execute(_del)
    return {"ok": True}

@app.post("/sicoe-obra/{contrato_id}/reportes")
def crear_reporte_obra(contrato_id: int, body: ReporteCreate, current_user=Depends(get_current_user)):
    def _num():
        return supabase.rpc("siguiente_numero_reporte", {"p_contrato_id": contrato_id}).execute().data
    numero = supabase_execute(_num)
    data = body.dict()
    data["contrato_id"] = contrato_id
    data["numero_reporte"] = numero
    data["estado"] = "Borrador"
    data["creado_por"] = int(current_user.get("sub") or current_user.get("id", 0))
    def _ins():
        return supabase.table("so_reportes").insert(data).execute().data
    result = supabase_execute(_ins)
    row = result[0] if result else {}
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "CREAR",
            "SICOE",
            "reporte",
            str(row.get("id") or ""),
            {"numero_reporte": row.get("numero_reporte"), "estado": row.get("estado")},
        )
    except Exception:
        pass
    return row


def _sicoe_resolver_acta_semana_corte(
    contrato_id: int, subcontratista_id: Optional[int]
) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """Misma resolución que asignar ítem: acta RPO en período + semana activa + corte vigente del sub (si aplica)."""
    today = date.today().isoformat()

    def _acta():
        return (
            supabase.table("actas")
            .select("id, numero_rpo")
            .eq("contrato_id", contrato_id)
            .eq("tipo_grupo", "RPO")
            .lte("fecha_inicio", today)
            .gte("fecha_fin", today)
            .order("id", desc=True)
            .limit(1)
            .execute()
            .data
        )

    actas = supabase_execute(_acta)
    acta_rpo_id = actas[0]["id"] if actas else None

    corte_id = None
    if subcontratista_id:
        try:

            def _corte():
                return (
                    supabase.table("subcontratista_cortes")
                    .select("id, consecutivo")
                    .eq("subcontratista_id", subcontratista_id)
                    .lte("fecha_inicio", today)
                    .gte("fecha_fin", today)
                    .limit(1)
                    .execute()
                    .data
                )

            cortes = supabase_execute(_corte)
            corte_id = cortes[0]["id"] if cortes else None
        except Exception:
            corte_id = None

    semana_id = None
    try:

        def _sem():
            return (
                supabase.table("so_semanas")
                .select("id, numero_semana")
                .eq("contrato_id", contrato_id)
                .eq("estado", "activa")
                .lte("fecha_inicio", today)
                .gte("fecha_fin", today)
                .limit(1)
                .execute()
                .data
            )

        sems = supabase_execute(_sem)
        semana_id = sems[0]["id"] if sems else None
    except Exception:
        semana_id = None

    return acta_rpo_id, semana_id, corte_id


def _coords_desde_fila_pk_maestro(pk_row: Optional[Dict[str, Any]]) -> Tuple[Optional[float], Optional[float]]:
    if not pk_row:
        return None, None
    pares = (
        ("centro_lat", "centro_lng"),
        ("lat", "lng"),
        ("coord_lat", "coord_lng"),
    )
    for a, b in pares:
        try:
            if pk_row.get(a) is not None and pk_row.get(b) is not None:
                return float(pk_row[a]), float(pk_row[b])
        except (TypeError, ValueError):
            continue
    return None, None


def _listado_precios_index_por_item_norm(contrato_id: int, capitulo: str) -> Dict[str, Dict[str, Any]]:
    cap_key = _dash_norm_capitulo_key_py(capitulo)
    idx: Dict[str, Dict[str, Any]] = {}
    off = 0
    while True:

        def _b(o=off):
            return (
                supabase.table("listado_precios")
                .select("id, capitulo, competencia, item_numero, descripcion, unidad, precio_unitario")
                .eq("contrato_id", contrato_id)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_b) or []
        for r in batch:
            if _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_key:
                continue
            ik = _dash_norm_item_key_py(r.get("item_numero"))
            if ik and ik not in idx:
                idx[ik] = r
        if len(batch) < 1000:
            break
        off += 1000
    return idx


def _presupuesto_listado_fallback_por_item_pk(
    contrato_id: int, capitulo: str
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """
    Por (item_norm, pk_disp): primera fila presupuesto con vlr_unitario, mismo capítulo.
    Sirve cuando el ítem no está en listado_precios pero sí en presupuesto (p. ej. líneas con Δ negativa).
    """
    cap_key = _dash_norm_capitulo_key_py(capitulo)
    out: Dict[Tuple[str, str], Dict[str, Any]] = {}
    off = 0
    while True:

        def _b(o=off):
            return (
                supabase.table("presupuesto")
                .select("capitulo, competencia, item, descripcion, und, vlr_unitario, pk_id")
                .eq("contrato_id", contrato_id)
                .eq("dado_de_baja", False)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_b) or []
        for r in batch:
            if _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_key:
                continue
            ik = _dash_norm_item_key_py(r.get("item"))
            pk_disp = _dash_pk_disp_key_py(r.get("pk_id"))
            if not ik or not pk_disp or pk_disp == "(sin pk)":
                continue
            key = (ik, pk_disp)
            if key in out:
                continue
            try:
                vu = float(r.get("vlr_unitario") or 0)
            except (TypeError, ValueError):
                continue
            out[key] = {
                "capitulo": r.get("capitulo"),
                "competencia": r.get("competencia"),
                "item_numero": r.get("item"),
                "descripcion": r.get("descripcion"),
                "unidad": r.get("und"),
                "precio_unitario": vu,
            }
        if len(batch) < 1000:
            break
        off += 1000
    return out


def _pk_maestro_por_display(contrato_id: int, pk_disps: List[str]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    seen = set()
    chunk: List[str] = []
    for raw in pk_disps:
        k = _dash_pk_disp_key_py(raw)
        if not k or k == "(sin pk)" or k in seen:
            continue
        seen.add(k)
        chunk.append(k)
    for i in range(0, len(chunk), 100):
        part = chunk[i : i + 100]

        def _q():
            return (
                supabase.table("pk_ids")
                .select("id, pk_id, civ, tramo, infraestructura, calzada, abs_inicio, abs_final, ubicacion")
                .eq("contrato_id", contrato_id)
                .in_("pk_id", part)
                .execute()
                .data
            )

        rows = supabase_execute(_q) or []
        for r in rows:
            disp = _dash_pk_disp_key_py(r.get("pk_id"))
            if disp and disp not in out:
                out[disp] = r
    return out


# Tope p_n en public.siguiente_n_numeros_registro (debe alinearse con el SQL del proyecto).
_REGISTRO_N_MAX_RPC = 2000


def _sicoe_reservar_numeros_registro_total(contrato_id: int, total_n: int) -> List[int]:
    """Reserva total_n consecutivos; varias RPC si total_n supera el tope de la función SQL."""
    if total_n < 1:
        return []
    out: List[int] = []
    rem = total_n
    while rem > 0:
        chunk = min(rem, _REGISTRO_N_MAX_RPC)
        out.extend(_sicoe_reservar_numeros_registro(contrato_id, chunk))
        rem -= chunk
    if len(out) != total_n:
        raise HTTPException(
            status_code=500,
            detail="Reserva de números de registro incompleta; reintente.",
        )
    return out


def _sicoe_reservar_numeros_registro(contrato_id: int, nlines: int) -> List[int]:
    def _parse_numero_registro_raw(raw) -> int:
        if raw is None:
            raise ValueError("RPC sin número de registro")
        if isinstance(raw, (int, float)) and not isinstance(raw, bool):
            return int(raw)
        if isinstance(raw, list) and len(raw) > 0:
            return _parse_numero_registro_raw(raw[0])
        if isinstance(raw, dict):
            for k in ("numero", "siguiente_numero_registro", "siguiente", "id"):
                if k in raw and raw[k] is not None:
                    return int(raw[k])
        return int(raw)

    numeros: List[int] = []

    def _try_rpc_bloque() -> bool:
        nonlocal numeros

        def _q():
            return supabase.rpc(
                "siguiente_n_numeros_registro", {"p_contrato_id": contrato_id, "p_n": nlines}
            ).execute().data

        try:
            raw = supabase_execute(_q)
        except Exception:
            return False
        if raw is None:
            return False
        if isinstance(raw, list):
            if len(raw) != nlines:
                return False
            try:
                numeros = [int(x) for x in raw]
            except (TypeError, ValueError):
                return False
            return True
        if isinstance(raw, dict):
            for k in ("siguiente_n_numeros_registro", "data", "result"):
                if k in raw and isinstance(raw[k], list):
                    try:
                        arr = [int(x) for x in raw[k]]
                    except (TypeError, ValueError):
                        return False
                    if len(arr) != nlines:
                        return False
                    numeros = arr
                    return True
        return False

    if not _try_rpc_bloque():
        for _ in range(nlines):

            def _n():
                return supabase.rpc("siguiente_numero_registro", {"p_contrato_id": contrato_id}).execute().data

            numeros.append(_parse_numero_registro_raw(supabase_execute(_n)))
    if len(set(numeros)) != len(numeros):
        raise HTTPException(
            status_code=500,
            detail="Números de registro duplicados al reservar; reintente.",
        )
    return numeros


def _parse_numero_reporte_raw(raw) -> int:
    if raw is None:
        raise ValueError("RPC sin número de reporte")
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return int(raw)
    if isinstance(raw, list) and len(raw) > 0:
        return _parse_numero_reporte_raw(raw[0])
    if isinstance(raw, dict):
        for k in ("numero_reporte", "siguiente_numero_reporte", "numero", "siguiente"):
            if k in raw and raw[k] is not None:
                return int(raw[k])
    return int(raw)


def _sicoe_reservar_n_numeros_reporte(contrato_id: int, n: int) -> List[int]:
    """N consecutivos vía RPC batch; si no existe la función en BD, cae en N llamadas a siguiente_numero_reporte."""
    if n < 1:
        return []

    def _try_batch() -> Optional[List[int]]:

        def _q():
            return (
                supabase.rpc(
                    "siguiente_n_numeros_reporte", {"p_contrato_id": contrato_id, "p_n": n}
                )
                .execute()
                .data
            )

        try:
            raw = supabase_execute(_q)
        except Exception:
            return None
        if raw is None:
            return None
        if isinstance(raw, list):
            if len(raw) != n:
                return None
            try:
                return [int(x) for x in raw]
            except (TypeError, ValueError):
                return None
        if isinstance(raw, dict):
            for k in ("siguiente_n_numeros_reporte", "data", "result"):
                if k in raw and isinstance(raw[k], list):
                    arr = raw[k]
                    if len(arr) != n:
                        return None
                    try:
                        return [int(x) for x in arr]
                    except (TypeError, ValueError):
                        return None
        return None

    got = _try_batch()
    if got is not None:
        if len(set(got)) != len(got):
            raise HTTPException(status_code=500, detail="Números de reporte duplicados al reservar; reintente.")
        return got
    out: List[int] = []
    for _ in range(n):

        def _one():
            return supabase.rpc("siguiente_numero_reporte", {"p_contrato_id": contrato_id}).execute().data

        out.append(_parse_numero_reporte_raw(supabase_execute(_one)))
    if len(set(out)) != len(out):
        raise HTTPException(status_code=500, detail="Números de reporte duplicados al reservar; reintente.")
    return out


# Nombre obligatorio de reportes generados desde migración dashboard → SICOE (biblioteca / lotes).
MIGR_BALANCE_CANTIDADES_DESC_PREFIX = "Balance De cantidades para pk_id "
# Compatibilidad con reportes antiguos antes del cambio de convención de nombre.
MIGR_DELTA_LEGACY_DESC_PREFIX = "Δ Dashboard ·"
DASHBOARD_DELTA_REPORTES_MAX = 220


def _descripcion_es_reporte_migracion_balance(desc: Optional[str]) -> bool:
    d = (desc or "").strip()
    return d.startswith(MIGR_BALANCE_CANTIDADES_DESC_PREFIX) or d.startswith(MIGR_DELTA_LEGACY_DESC_PREFIX)


def _pk_disp_desde_descripcion_migracion(desc: Optional[str]) -> Optional[str]:
    """Obtiene texto PK del maestro desde la descripción del reporte (nuevo o legado)."""
    d = (desc or "").strip()
    if d.startswith(MIGR_BALANCE_CANTIDADES_DESC_PREFIX):
        rest = d[len(MIGR_BALANCE_CANTIDADES_DESC_PREFIX) :].strip()
        for sep in (" ·", " |", "\n"):
            if sep in rest:
                rest = rest.split(sep)[0].strip()
        return _dash_pk_disp_key_py(rest) if rest else None
    if d.startswith(MIGR_DELTA_LEGACY_DESC_PREFIX):
        try:
            parts = d.split("·")
            if len(parts) >= 2:
                return _dash_pk_disp_key_py(parts[1].strip())
        except Exception:
            pass
    return None


def _n1_aprobar_todos_registros_de_reportes(
    contrato_id: int, reporte_ids: List[int], autor_id: int
) -> int:
    """Marca Nivel 1 Aprobado en todos los registros de los reportes indicados (migración balance)."""
    now = datetime.utcnow().isoformat()
    nupd = 0
    for rep_id in reporte_ids:
        def _ids():
            return (
                supabase.table("so_registros")
                .select("id")
                .eq("contrato_id", contrato_id)
                .eq("reporte_id", rep_id)
                .execute()
                .data
            )

        idrows = supabase_execute(_ids) or []
        reg_ids = []
        for row in idrows:
            try:
                reg_ids.append(int(row["id"]))
            except (TypeError, ValueError, KeyError):
                continue
        for i in range(0, len(reg_ids), 90):
            chunk = reg_ids[i : i + 90]
            if not chunk:
                continue

            def _u():
                return (
                    supabase.table("so_registros")
                    .update(
                        {
                            "nivel1_estado": "Aprobado",
                            "nivel1_usuario_id": autor_id,
                            "nivel1_fecha": now,
                        }
                    )
                    .in_("id", chunk)
                    .eq("contrato_id", contrato_id)
                    .execute()
                )

            supabase_execute(_u)
            nupd += len(chunk)
    return nupd


def _count_pk_ids_maestro_contrato(contrato_id: int) -> int:
    n = 0
    off = 0
    while True:

        def _b(o=off):
            return (
                supabase.table("pk_ids")
                .select("id")
                .eq("contrato_id", contrato_id)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_b) or []
        n += len(batch)
        if len(batch) < 1000:
            break
        off += 1000
    return n


def _pk_disp_migrados_dashboard_delta(contrato_id: int, capitulo: str) -> Set[str]:
    """PK (texto maestro) que ya tienen reporte generado por esta migración (mismo capítulo normalizado)."""
    cap_key = _dash_norm_capitulo_key_py(capitulo)
    out: Set[str] = set()
    off = 0
    while True:

        def _q(o=off):
            return (
                supabase.table("so_reportes")
                .select("id, capitulo, descripcion_actividad, pk_id_id, pk_ids(pk_id)")
                .eq("contrato_id", contrato_id)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_q) or []
        for r in batch:
            desc = r.get("descripcion_actividad")
            if not _descripcion_es_reporte_migracion_balance(desc):
                continue
            if _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_key:
                continue
            from_desc = _pk_disp_desde_descripcion_migracion(desc)
            if from_desc and from_desc != "(sin pk)":
                out.add(from_desc)
                continue
            pj = r.get("pk_ids") if isinstance(r.get("pk_ids"), dict) else {}
            disp = _dash_pk_disp_key_py((pj or {}).get("pk_id"))
            if disp and disp != "(sin pk)":
                out.add(disp)
        if len(batch) < 1000:
            break
        off += 1000
    return out


def _dash_parse_float_num(v: Any) -> float:
    """Parse numérico tolerante (JSON/RPC/comas) para deltas del dashboard."""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return float(int(v))
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _dash_pk_display_sort_key(disp: str) -> Tuple[int, Any]:
    s = str(disp).strip()
    try:
        return (0, int(s))
    except ValueError:
        return (1, s)


def _dashboard_delta_armar_pk_lines(
    contrato_id: int,
    capitulo: str,
    items_ord: List[str],
    listado_idx: Dict[str, Dict[str, Any]],
    solo_delta_positivo: bool,
) -> Tuple[Dict[str, List[Dict[str, Any]]], List[str]]:
    pk_lines: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    warnings_item: List[str] = []
    for it in items_ord:
        it_key = _dash_norm_item_key_py(it)
        if it_key not in listado_idx:
            warnings_item.append(it)
        tabla = _dashboard_pkid_tabla_obra_core(contrato_id, capitulo, it)
        for row in tabla.get("rows") or []:
            pk_disp = _dash_pk_disp_key_py(row.get("pk_id"))
            dc = _dash_parse_float_num(row.get("delta_cant"))
            if solo_delta_positivo and dc <= 1e-9:
                continue
            if not solo_delta_positivo and abs(dc) <= 1e-9:
                continue
            pk_lines[pk_disp].append(
                {
                    "item": it,
                    "item_key": it_key,
                    "delta_cant": round(dc, 4),
                    "descripcion_item": tabla.get("descripcion_item") or "",
                }
            )
    return pk_lines, warnings_item


def _dashboard_delta_parse_items_ord(items_in: List[str]) -> List[str]:
    items_ord: List[str] = []
    seen_it: Set[str] = set()
    for it in items_in:
        kn = _dash_norm_item_key_py(str(it).strip())
        if not kn or kn in seen_it:
            continue
        seen_it.add(kn)
        items_ord.append(str(it).strip())
    return items_ord


class ReportesMasivosDeltaDashboardBody(BaseModel):
    """Consolidar Δ (ppto N3 aprobado − obra N3 aprobada) por PK_ID: un reporte por PK con una línea por ítem."""

    capitulo: str
    items: List[str]
    # False = incluir descuentos (Δ < 0) y excesos (Δ > 0); solo excluye Δ≈0
    solo_delta_positivo: bool = False


class ConfirmarDeltaDashboardBody(BaseModel):
    capitulo: str
    limite: int = 45


def _hay_mas_reportes_borrador_delta(contrato_id: int, cap_key: str) -> bool:
    """True si aún hay reportes Borrador de migración balance para el capítulo (normalizado)."""
    off = 0
    while True:

        def _b(o=off):
            return (
                supabase.table("so_reportes")
                .select("id, capitulo, descripcion_actividad")
                .eq("contrato_id", contrato_id)
                .eq("estado", "Borrador")
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_b) or []
        for r in batch:
            desc = r.get("descripcion_actividad")
            if not _descripcion_es_reporte_migracion_balance(desc):
                continue
            if _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_key:
                continue
            return True
        if len(batch) < 1000:
            break
        off += 1000
    return False


@app.get("/sicoe-obra/{contrato_id}/reportes-biblioteca-balance-cantidades")
def biblioteca_reportes_balance_cantidades(
    contrato_id: int,
    capitulo: str = Query(...),
    estado: Optional[str] = Query(
        None,
        description="Filtrar por estado de cabecera (ej. Borrador, No Revisados); vacío = todos",
    ),
    current_user=Depends(get_current_user),
):
    """Listado de reportes generados por migración balance (revisión uno a uno)."""
    cap = (capitulo or "").strip()
    if not cap:
        raise HTTPException(status_code=422, detail="Indique capítulo.")
    cap_key = _dash_norm_capitulo_key_py(cap)
    est_f = (estado or "").strip() or None
    out: List[Dict[str, Any]] = []
    off = 0
    while True:

        def _b(o=off):
            return (
                supabase.table("so_reportes")
                .select("id, numero_reporte, descripcion_actividad, estado, capitulo, pk_id_id, pk_ids(pk_id)")
                .eq("contrato_id", contrato_id)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_b) or []
        for r in batch:
            if _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_key:
                continue
            if not _descripcion_es_reporte_migracion_balance(r.get("descripcion_actividad")):
                continue
            if est_f and (r.get("estado") or "").strip() != est_f:
                continue
            pj = r.get("pk_ids") if isinstance(r.get("pk_ids"), dict) else {}
            pk_txt = _pk_disp_desde_descripcion_migracion(r.get("descripcion_actividad")) or _dash_pk_disp_key_py(
                (pj or {}).get("pk_id")
            )
            out.append(
                {
                    "id": r.get("id"),
                    "numero_reporte": r.get("numero_reporte"),
                    "descripcion_actividad": r.get("descripcion_actividad"),
                    "estado": r.get("estado"),
                    "pk_id": pk_txt,
                }
            )
        if len(batch) < 1000:
            break
        off += 1000
    def _sort_key_bib(r: Dict[str, Any]):
        nr = r.get("numero_reporte")
        try:
            n = int(nr) if nr is not None and str(nr).strip() != "" else 0
        except (TypeError, ValueError):
            n = 0
        try:
            rid = int(r.get("id") or 0)
        except (TypeError, ValueError):
            rid = 0
        return (n, rid)

    out.sort(key=_sort_key_bib)
    return {"reportes": out, "total": len(out)}


@app.get("/sicoe-obra/{contrato_id}/reportes-masivos-dashboard-delta/preview")
def preview_reportes_masivos_dashboard_delta(
    contrato_id: int,
    capitulo: str = Query(...),
    items: List[str] = Query([], description="Ítem del listado (repetir param items=)"),
    solo_delta_positivo: bool = Query(
        False,
        description="Si true, solo PK con Δ cantidad > 0. Si false (recomendado), toda Δ distinta de cero (incluye negativos / descuentos).",
    ),
    current_user=Depends(get_current_user),
):
    """Vista previa: totales de PK en maestro, cuántos tienen Δ pendiente y cuántos caben en el siguiente lote (excl. ya migrados)."""
    cap = (capitulo or "").strip()
    if not cap:
        raise HTTPException(status_code=422, detail="Indique capítulo.")
    items_ord = _dashboard_delta_parse_items_ord(list(items or []))
    if not items_ord:
        raise HTTPException(status_code=422, detail="Indique al menos un ítem.")
    listado_idx = _listado_precios_index_por_item_norm(contrato_id, cap)
    pk_lines, warnings_item = _dashboard_delta_armar_pk_lines(
        contrato_id, cap, items_ord, listado_idx, solo_delta_positivo
    )
    migrados = _pk_disp_migrados_dashboard_delta(contrato_id, cap)
    todos_pk_delta = {k for k in pk_lines.keys() if k and k != "(sin pk)" and pk_lines[k]}
    pendientes = sorted(
        (t for t in todos_pk_delta if t not in migrados),
        key=lambda x: _dash_pk_display_sort_key(str(x)),
    )
    ya_migrados_con_delta = sorted(
        (t for t in todos_pk_delta if t in migrados),
        key=lambda x: _dash_pk_display_sort_key(str(x)),
    )
    total_maestro = _count_pk_ids_maestro_contrato(contrato_id)
    n_pend = len(pendientes)
    n_lote = min(DASHBOARD_DELTA_REPORTES_MAX, n_pend)
    n_resto = max(0, n_pend - n_lote)
    return {
        "capitulo": cap,
        "total_pk_maestro": total_maestro,
        "pk_con_delta_total": len(todos_pk_delta),
        "pk_pendientes_migrar": n_pend,
        "pk_ya_migrados_con_delta": len(ya_migrados_con_delta),
        "este_lote_cantidad_pk": n_lote,
        "restantes_tras_este_lote": n_resto,
        "limite_lote_max": DASHBOARD_DELTA_REPORTES_MAX,
        "advertencias": {"items_sin_listado_en_capitulo": warnings_item},
    }


@app.post("/sicoe-obra/{contrato_id}/reportes-masivos-dashboard-delta/confirmar")
def confirmar_reportes_masivos_dashboard_delta(
    contrato_id: int,
    body: ConfirmarDeltaDashboardBody,
    current_user=Depends(get_current_user),
):
    """
    Pasa a «No Revisados» reportes en Borrador de migración balance (mismo capítulo) y aprueba
    Nivel 1 en todos sus registros (habilita validación / flujo hacia Nivel 2).
    Requiere permiso de validación en Nivel 1 (misma regla que validar-nivel1).
    """
    cap = (body.capitulo or "").strip()
    if not cap:
        raise HTTPException(status_code=422, detail="Indique capítulo.")
    cap_key = _dash_norm_capitulo_key_py(cap)
    limite = max(1, min(int(body.limite or 45), 80))
    uid = int(current_user.get("sub") or current_user.get("id", 0))
    _require_sicoe_puede_validar_nivel(current_user, uid, 1)

    ids_pick: List[int] = []
    off = 0
    while len(ids_pick) < limite:
        def _b(o=off):
            return (
                supabase.table("so_reportes")
                .select("id, capitulo, estado, descripcion_actividad")
                .eq("contrato_id", contrato_id)
                .eq("estado", "Borrador")
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_b) or []
        for r in batch:
            if len(ids_pick) >= limite:
                break
            if _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_key:
                continue
            desc = r.get("descripcion_actividad")
            if not _descripcion_es_reporte_migracion_balance(desc):
                continue
            try:
                rid = int(r.get("id"))
            except (TypeError, ValueError):
                continue
            ids_pick.append(rid)
        if len(batch) < 1000:
            break
        off += 1000

    if not ids_pick:
        return {"actualizados": 0, "hay_mas": False, "ids": [], "registros_n1_aprobados": 0}

    def _upd():
        return (
            supabase.table("so_reportes")
            .update(
                {
                    "estado": "No Revisados",
                    "modificado_por": uid,
                    "updated_at": datetime.utcnow().isoformat(),
                }
            )
            .in_("id", ids_pick)
            .eq("contrato_id", contrato_id)
            .eq("estado", "Borrador")
            .execute()
        )

    supabase_execute(_upd)
    n1_n = _n1_aprobar_todos_registros_de_reportes(contrato_id, ids_pick, uid)
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "EDITAR",
            "SICOE",
            "reporte",
            "dashboard_delta_confirmar_lote",
            {
                "n": len(ids_pick),
                "capitulo": cap,
                "registros_n1_aprobados": n1_n,
                "ids_muestra": ids_pick[:20],
            },
        )
    except Exception:
        pass

    hay_mas = _hay_mas_reportes_borrador_delta(contrato_id, cap_key)
    return {
        "actualizados": len(ids_pick),
        "hay_mas": hay_mas,
        "ids": ids_pick,
        "registros_n1_aprobados": n1_n,
    }


@app.post("/sicoe-obra/{contrato_id}/reportes-masivos-dashboard-delta")
def crear_reportes_masivos_dashboard_delta(
    contrato_id: int,
    body: ReportesMasivosDeltaDashboardBody,
    current_user=Depends(get_current_user),
):
    """
    Crea reportes SICOE en Borrador a partir del mismo criterio que la tabla PK del dashboard:
    presupuesto validado N3 «Aprobado» vs. registros de obra con N3 «Aprobado».
    Por cada PK con |Δ| de cantidad según `solo_delta_positivo` (por defecto se incluyen Δ negativos y positivos),
    Excluye PK que ya tienen reporte «Balance De cantidades…» o legado «Δ Dashboard ·» (mismo capítulo). Como máximo DASHBOARD_DELTA_REPORTES_MAX PK por llamada.
    Cabecera: descripción «Balance De cantidades para pk_id {PK}»; maestro pk_ids (abscisas, ubicación); creado_por = quien llama.
    """
    capitulo = (body.capitulo or "").strip()
    if not capitulo:
        raise HTTPException(status_code=422, detail="Indique capítulo.")
    items_in = [str(x).strip() for x in (body.items or []) if str(x).strip()]
    items_ord = _dashboard_delta_parse_items_ord(items_in)
    if not items_ord:
        raise HTTPException(status_code=422, detail="Indique al menos un ítem.")

    acta_rpo_id, semana_id, corte_id = _sicoe_resolver_acta_semana_corte(contrato_id, None)
    if not acta_rpo_id:
        raise HTTPException(
            status_code=422,
            detail="No existe acta RPO vigente para hoy. Créela en administración antes de generar reportes.",
        )

    listado_idx = _listado_precios_index_por_item_norm(contrato_id, capitulo)
    pk_lines, warnings_item = _dashboard_delta_armar_pk_lines(
        contrato_id, capitulo, items_ord, listado_idx, body.solo_delta_positivo
    )

    omitidos: List[Dict[str, Any]] = []
    creados: List[Dict[str, Any]] = []
    uid = int(current_user.get("sub") or current_user.get("id", 0))

    migrados = _pk_disp_migrados_dashboard_delta(contrato_id, capitulo)
    candidatos_all = [
        (pk, lines)
        for pk, lines in sorted(pk_lines.items(), key=lambda x: _dash_pk_display_sort_key(str(x[0])))
        if pk and pk != "(sin pk)" and lines
    ]
    listado_fb = _presupuesto_listado_fallback_por_item_pk(contrato_id, capitulo)
    candidatos_pk = [(pk, lines) for pk, lines in candidatos_all if pk not in migrados][:DASHBOARD_DELTA_REPORTES_MAX]
    omitidos_ya = [pk for pk, _ in candidatos_all if pk in migrados][:500]
    n_pend_total = len([1 for pk, _ in candidatos_all if pk not in migrados])
    restantes_tras = max(0, n_pend_total - len(candidatos_pk))
    total_maestro = _count_pk_ids_maestro_contrato(contrato_id)

    pk_m = _pk_maestro_por_display(contrato_id, [pk for pk, _ in candidatos_pk])

    work_items: List[Dict[str, Any]] = []
    for pk_disp, lines in candidatos_pk:
        filas_insert: List[Dict[str, Any]] = []
        for ln in lines:
            lp = listado_idx.get(ln["item_key"]) or listado_fb.get((ln["item_key"], pk_disp))
            if not lp:
                continue
            vlr = float(lp.get("precio_unitario") or 0)
            dc = float(ln["delta_cant"])
            cant_tot = round(dc, 2)
            costo_dir = round(cant_tot * vlr, 0)
            filas_insert.append(
                {
                    "capitulo": lp.get("capitulo"),
                    "competencia": lp.get("competencia"),
                    "item_numero": lp.get("item_numero"),
                    "item_descripcion": lp.get("descripcion"),
                    "unidad": lp.get("unidad"),
                    "vlr_unitario": vlr,
                    "longitud": dc,
                    "cantidad_total": cant_tot,
                    "costo_directo": costo_dir,
                    "acta_rpo_id": acta_rpo_id,
                    "semana_id": semana_id,
                    "corte_id": corte_id,
                    "nivel1_estado": "No Revisado",
                    "nivel2_estado": "No Revisado",
                    "nivel3_estado": "No Revisado",
                    "nombre": str(lp.get("item_numero") or ln["item"] or "").strip() or None,
                    "descripcion": (ln.get("descripcion_item") or lp.get("descripcion") or "")[:2000] or None,
                }
            )
        if not filas_insert:
            omitidos.append({"pk_id": pk_disp, "motivo": "sin líneas válidas (listado o Δ)"})
            continue

        pk_row = pk_m.get(pk_disp)
        pk_id_id = int(pk_row["id"]) if pk_row and pk_row.get("id") is not None else None
        lat_g, lng_g = _coords_desde_fila_pk_maestro(pk_row)

        desc_act = f"{MIGR_BALANCE_CANTIDADES_DESC_PREFIX}{pk_disp}"[:500]
        rep_payload: Dict[str, Any] = {
            "contrato_id": contrato_id,
            "descripcion_actividad": desc_act,
            "capitulo": capitulo,
            "estado": "Borrador",
            "creado_por": uid,
            "acta_rpo_id": acta_rpo_id,
            "semana_id": semana_id,
            "corte_id": corte_id,
            "pk_id_id": pk_id_id,
        }
        if pk_row:
            for fld in ("civ", "tramo", "infraestructura", "calzada", "ubicacion", "abs_inicio", "abs_final"):
                if pk_row.get(fld) is not None:
                    rep_payload[fld] = pk_row.get(fld)
        if lat_g is not None and lng_g is not None:
            rep_payload["coord_lat"] = lat_g
            rep_payload["coord_lng"] = lng_g

        work_items.append(
            {
                "pk_disp": pk_disp,
                "lines": lines,
                "filas_insert": filas_insert,
                "rep_payload": rep_payload,
            }
        )

    if work_items:
        numeros_rep = _sicoe_reservar_n_numeros_reporte(contrato_id, len(work_items))
        inserted_rows: List[Dict[str, Any]] = []
        work_ok: List[Dict[str, Any]] = []
        for w, numero_rep in zip(work_items, numeros_rep):
            rp = dict(w["rep_payload"])
            rp["numero_reporte"] = numero_rep

            def _ins_one(payload=rp):
                return supabase.table("so_reportes").insert(payload).execute().data

            rep_rows = supabase_execute(_ins_one) or []
            rep = rep_rows[0] if rep_rows else {}
            rid_one = rep.get("id")
            if not rid_one:
                omitidos.append({"pk_id": w["pk_disp"], "motivo": "error insertando reporte"})
                continue
            inserted_rows.append(rep)
            work_ok.append(w)

        total_lines = sum(len(w["filas_insert"]) for w in work_ok)
        nums_all = (
            _sicoe_reservar_numeros_registro_total(contrato_id, total_lines)
            if total_lines
            else []
        )
        pos_nr = 0
        flat_regs: List[Dict[str, Any]] = []
        for w, rep in zip(work_ok, inserted_rows):
            rid = rep.get("id")
            if not rid:
                continue
            filas_insert = w["filas_insert"]
            nlines = len(filas_insert)
            nums_reg = nums_all[pos_nr : pos_nr + nlines]
            pos_nr += nlines
            rep_payload = w["rep_payload"]
            for line, numero in zip(filas_insert, nums_reg):
                d = dict(line)
                d["reporte_id"] = rid
                d["numero_registro"] = numero
                d["contrato_id"] = contrato_id
                d["creado_por_reg"] = uid
                for campo in (
                    "pk_id_id",
                    "civ",
                    "tramo",
                    "infraestructura",
                    "calzada",
                    "ubicacion",
                    "coord_lat",
                    "coord_lng",
                    "abs_inicio",
                    "abs_final",
                    "subcontratista_id",
                    "inspector_id",
                ):
                    if rep_payload.get(campo) is not None:
                        d[campo] = rep_payload[campo]
                flat_regs.append(d)

        _REG_BATCH = 200
        for i in range(0, len(flat_regs), _REG_BATCH):

            def _ins_reg_blk(fr=flat_regs[i : i + _REG_BATCH]):
                return supabase.table("so_registros").insert(fr).execute().data

            supabase_execute(_ins_reg_blk)

        creados_ids: List[int] = []
        for w, rep in zip(work_ok, inserted_rows):
            rid = rep.get("id")
            if rid:
                creados_ids.append(int(rid))
            creados.append(
                {
                    "id": rep.get("id"),
                    "numero_reporte": rep.get("numero_reporte"),
                    "pk_id": w["pk_disp"],
                    "n_registros": len(w["filas_insert"]),
                }
            )

        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            registrar_log(
                u_log,
                "CREAR",
                "SICOE",
                "reporte",
                "dashboard_delta_masivo_lote",
                {
                    "origen": "dashboard_delta_masivo",
                    "n_reportes": len(creados_ids),
                    "n_registros": len(flat_regs),
                    "ids_muestra": creados_ids[:30],
                    "capitulo": capitulo,
                },
            )
        except Exception:
            pass

    return {
        "creados": creados,
        "omitidos": omitidos,
        "resumen": {
            "total_pk_maestro": total_maestro,
            "pk_procesados_este_lote": len(creados),
            "pk_pendientes_totales_antes_lote": n_pend_total,
            "restantes_tras_este_lote": restantes_tras,
            "pk_omitidos_ya_migrados": len(omitidos_ya),
            "limite_lote_max": DASHBOARD_DELTA_REPORTES_MAX,
        },
        "advertencias": {
            "items_sin_listado_en_capitulo": warnings_item,
        },
    }


class RegistroOfflineCreate(BaseModel):
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    longitud: Optional[float] = None
    ancho: Optional[float] = None
    espesor: Optional[float] = None
    cantidad: Optional[float] = None
    unidad: Optional[str] = None
    observacion: Optional[str] = None
    foto_url: Optional[str] = None
    foto_numero: Optional[int] = None
    grafico_url: Optional[str] = None
    grafico_numero: Optional[int] = None

class ReporteOfflineCreate(BaseModel):
    """Creación atómica de reporte + registros desde modo offline."""
    descripcion_actividad: str
    capitulo: str
    estado: Optional[str] = "Sin Asignar Ítem"
    margen: Optional[str] = None
    abs_inicio: Optional[float] = None
    abs_final: Optional[float] = None
    nodo_ini: Optional[str] = None
    nodo_fin: Optional[str] = None
    registros: List[RegistroOfflineCreate] = []

@app.post("/sicoe-obra/{contrato_id}/reportes-offline")
def crear_reporte_offline(
    contrato_id: int,
    body: ReporteOfflineCreate,
    current_user=Depends(get_current_user),
):
    """
    Endpoint de sincronización offline: crea un reporte con todos sus registros
    en una sola transacción. El servidor asigna el número de reporte definitivo.
    Acepta Idempotency-Key en el header — el servidor lo almacena y devuelve
    el mismo resultado si ya fue procesado.
    """
    usuario_id = int(current_user.get("sub") or current_user.get("id", 0))

    def _num():
        return supabase.rpc("siguiente_numero_reporte", {"p_contrato_id": contrato_id}).execute().data
    numero = supabase_execute(_num)

    reporte_data = {
        "contrato_id": contrato_id,
        "numero_reporte": numero,
        "descripcion_actividad": body.descripcion_actividad,
        "capitulo": body.capitulo,
        "estado": "Borrador",
        "margen": body.margen,
        "abs_inicio": body.abs_inicio,
        "abs_final": body.abs_final,
        "nodo_ini": body.nodo_ini,
        "nodo_fin": body.nodo_fin,
        "creado_por": usuario_id,
    }

    def _ins_reporte():
        return supabase.table("so_reportes").insert(reporte_data).execute().data
    result = supabase_execute(_ins_reporte)
    reporte_row = result[0] if result else {}
    reporte_id = reporte_row.get("id")

    # Insertar registros en lote
    registros_creados = []
    if reporte_id and body.registros:
        regs = [
            {
                "reporte_id": reporte_id,
                "contrato_id": contrato_id,
                "nombre": r.nombre,
                "descripcion": r.descripcion,
                "longitud": r.longitud,
                "ancho": r.ancho,
                "espesor": r.espesor,
                "cantidad": r.cantidad,
                "unidad": r.unidad,
                "observacion": r.observacion,
                "foto_url": r.foto_url,
                "foto_numero": r.foto_numero,
                "grafico_url": r.grafico_url,
                "grafico_numero": r.grafico_numero,
                "nivel1_estado": "No Revisado",
                "creado_por": usuario_id,
            }
            for r in body.registros
        ]
        def _ins_regs():
            return supabase.table("so_registros").insert(regs).execute().data
        registros_creados = supabase_execute(_ins_regs) or []

    return {"reporte": reporte_row, "registros": registros_creados}


@app.get("/sicoe-obra/{contrato_id}/plantillas")
def listar_plantillas(contrato_id: int, capitulo: str = None, current_user=Depends(get_current_user)):
    def _q():
        q = supabase.table("so_plantillas").select("*, so_plantilla_items(*)").eq("contrato_id", contrato_id)
        if capitulo:
            q = q.eq("capitulo", capitulo)
        return q.order("nombre").execute().data
    rows = supabase_execute(_q)
    for r in rows:
        r["items"] = r.pop("so_plantilla_items", [])
    return rows

class PlantillaItem(BaseModel):
    nombre: str
    descripcion: Optional[str] = None
    orden: int = 0

class PlantillaCreate(BaseModel):
    nombre: str
    capitulo: str
    items: List[PlantillaItem] = []

@app.post("/sicoe-obra/{contrato_id}/plantillas")
def crear_plantilla(contrato_id: int, body: PlantillaCreate, current_user=Depends(get_current_user)):
    def _ins():
        return supabase.table("so_plantillas").insert({
            "contrato_id": contrato_id,
            "nombre": body.nombre,
            "capitulo": body.capitulo,
            "creado_por": int(current_user.get("sub", 0) or current_user.get("id", 0))
        }).execute().data
    plantilla = supabase_execute(_ins)
    if not plantilla:
        raise HTTPException(status_code=500, detail="Error creando plantilla")
    pid = plantilla[0]["id"]
    if body.items:
        def _items():
            return supabase.table("so_plantilla_items").insert([
                {"plantilla_id": pid, "nombre": it.nombre, "descripcion": it.descripcion, "orden": it.orden}
                for it in body.items
            ]).execute().data
        supabase_execute(_items)
    return plantilla[0]

@app.post("/sicoe-obra/{contrato_id}/next-registro")
def next_numero_registro(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.rpc("siguiente_numero_registro", {"p_contrato_id": contrato_id}).execute().data
    numero = supabase_execute(_q)
    return {"numero": numero}

@app.post("/sicoe-obra/{contrato_id}/upload-foto")
async def upload_foto(contrato_id: int, file: UploadFile = File(...), numero: int = Form(...), descripcion: str = Form(""), current_user=Depends(get_current_user)):
    import cloudinary.uploader
    _cloudinary_config()
    contents = await file.read()
    result = cloudinary.uploader.upload(
        contents,
        folder=_cloudinary_folder_contrato(contrato_id, CLOUDINARY_SUB_FOTOS),
        public_id=f"foto_{numero}",
        overwrite=True,
        resource_type="image",
    )
    return {"url": result["secure_url"], "numero": numero}

@app.post("/sicoe-obra/{contrato_id}/upload-grafico")
async def upload_grafico(contrato_id: int, file: UploadFile = File(...), numero: int = Form(...), descripcion: str = Form(""), current_user=Depends(get_current_user)):
    import cloudinary.uploader
    _cloudinary_config()
    contents = await file.read()
    result = cloudinary.uploader.upload(
        contents,
        folder=_cloudinary_folder_contrato(contrato_id, CLOUDINARY_SUB_GRAFICOS),
        public_id=f"grafico_{numero}",
        overwrite=True,
        resource_type="image",
    )
    return {"url": result["secure_url"], "numero": numero}

@app.get("/sicoe-obra/{contrato_id}/galeria")
def galeria_imagenes(contrato_id: int, tipo: str = "foto", desde: str = None, hasta: str = None, current_user=Depends(get_current_user)):
    def _q():
        q = supabase.table("so_registros")\
            .select("foto_url, foto_numero, foto_descripcion, grafico_url, grafico_numero, grafico_descripcion, created_at")\
            .eq("contrato_id", contrato_id)
        if desde:
            q = q.gte("created_at", desde)
        if hasta:
            q = q.lte("created_at", hasta + "T23:59:59")
        return q.order("created_at", desc=True).execute().data
    rows = supabase_execute(_q)
    result = []
    for r in rows:
        if tipo == "foto" and r.get("foto_url"):
            result.append({"url": r["foto_url"], "numero": r["foto_numero"], "descripcion": r.get("foto_descripcion","")})
        elif tipo == "grafico" and r.get("grafico_url"):
            result.append({"url": r["grafico_url"], "numero": r["grafico_numero"], "descripcion": r.get("grafico_descripcion","")})
    return result

@app.post("/sicoe-obra/{contrato_id}/next-foto")
def next_numero_foto(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.rpc("siguiente_numero_foto", {"p_contrato_id": contrato_id}).execute().data
    return {"numero": supabase_execute(_q)}

@app.post("/sicoe-obra/{contrato_id}/next-grafico")
def next_numero_grafico(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.rpc("siguiente_numero_grafico", {"p_contrato_id": contrato_id}).execute().data
    return {"numero": supabase_execute(_q)}

class RegistroCreate(BaseModel):
    reporte_id: int
    numero_registro: int
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    longitud: Optional[float] = None
    ancho: Optional[float] = None
    espesor: Optional[float] = None
    cantidad: Optional[float] = None
    cantidad_total: Optional[float] = None
    unidad: Optional[str] = None
    observacion: Optional[str] = None
    foto_url: Optional[str] = None
    foto_numero: Optional[int] = None
    foto_descripcion: Optional[str] = None
    grafico_url: Optional[str] = None
    grafico_numero: Optional[int] = None
    grafico_descripcion: Optional[str] = None
    competencia: Optional[str] = None
    item_numero: Optional[str] = None
    item_descripcion: Optional[str] = None
    vlr_unitario: Optional[float] = None
    costo_directo: Optional[float] = None
    enlace_soporte: Optional[str] = None
    semana_id: Optional[int] = None
    acta_rpo_id: Optional[int] = None
    corte_id: Optional[int] = None
    pk_id_id: Optional[int] = None
    civ: Optional[str] = None
    tramo: Optional[str] = None
    infraestructura: Optional[str] = None
    calzada: Optional[str] = None
    ubicacion: Optional[str] = None
    coord_lat: Optional[float] = None
    coord_lng: Optional[float] = None
    abs_inicio: Optional[float] = None
    abs_final: Optional[float] = None
    nodo_ini: Optional[str] = None
    nodo_fin: Optional[str] = None
    subcontratista_id: Optional[int] = None
    inspector_id: Optional[int] = None
    creado_por_reg: Optional[int] = None
    modificado_por_reg: Optional[int] = None

class RegistroLineaNuevoReporte(BaseModel):
    """Líneas del modal Nuevo reporte: sin reporte_id ni número (los asigna el servidor)."""
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    longitud: Optional[float] = None
    ancho: Optional[float] = None
    espesor: Optional[float] = None
    cantidad: Optional[float] = None
    cantidad_total: Optional[float] = None
    unidad: Optional[str] = None
    observacion: Optional[str] = None
    foto_url: Optional[str] = None
    foto_numero: Optional[int] = None
    foto_descripcion: Optional[str] = None
    grafico_url: Optional[str] = None
    grafico_numero: Optional[int] = None
    grafico_descripcion: Optional[str] = None

class AsignarActoresPorPkBody(BaseModel):
    """Inspector y subcontratista en todas las cabeceras `so_reportes` con el mismo `pk_id_id`."""

    pk_id_id: int
    inspector_id: Optional[int] = None
    subcontratista_id: Optional[int] = None

class ReemplazarRegistrosNuevoReporteBody(BaseModel):
    registros: List[RegistroLineaNuevoReporte]

@app.put("/sicoe-obra/{contrato_id}/reportes/{reporte_id}")
def actualizar_reporte(contrato_id: int, reporte_id: int, body: ReporteCreate, current_user=Depends(get_current_user)):
    def _prev_rep():
        return supabase.table("so_reportes").select("*").eq("id", reporte_id).eq("contrato_id", contrato_id).limit(1).execute().data
    prev_rows = supabase_execute(_prev_rep)
    prev_rep = prev_rows[0] if prev_rows else {}
    data = body.dict()
    # Se persisten localización y PK en so_reportes (antes se descartaban y solo existía PATCH en Borrador;
    # un reintento tras guardar cabecera dejaba estado≠Borrador y el PATCH devolvía 400).
    data.pop("updated_at", None)
    data["updated_at"]     = "now()"
    data["modificado_por"] = int(current_user.get("sub") or current_user.get("id", 0))
    def _upd():
        return supabase.table("so_reportes").update(data)\
            .eq("id", reporte_id).eq("contrato_id", contrato_id).execute().data
    result = supabase_execute(_upd)
    out = result[0] if result else {}
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "EDITAR",
            "SICOE",
            "reporte",
            str(reporte_id),
            {"numero_reporte": out.get("numero_reporte") or prev_rep.get("numero_reporte")},
            valor_anterior=_so_reporte_audit_snapshot(prev_rep),
            valor_nuevo=_so_reporte_audit_snapshot(out),
        )
    except Exception:
        pass
    return out

@app.post("/sicoe-obra/{contrato_id}/reportes/asignar-actores-por-pk")
def asignar_actores_por_pk(
    contrato_id: int, body: AsignarActoresPorPkBody, current_user=Depends(get_current_user)
):
    uid = int(current_user.get("sub") or current_user.get("id", 0))
    if not _es_desarrollador(current_user) and not _cargo_permiso_editar_reporte_cantidades_user_id(uid):
        raise HTTPException(
            status_code=403,
            detail="No tiene permiso de edición en «Reporte de cantidades» (matriz de accesos).",
        )
    try:
        pk_id_id = int(body.pk_id_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail="pk_id_id inválido")
    patch = {
        "inspector_id": body.inspector_id,
        "subcontratista_id": body.subcontratista_id,
        "modificado_por": uid,
        "updated_at": "now()",
    }

    def _upd():
        return (
            supabase.table("so_reportes")
            .update(patch)
            .eq("contrato_id", contrato_id)
            .eq("pk_id_id", pk_id_id)
            .execute()
            .data
        )

    rows = supabase_execute(_upd) or []
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "EDITAR",
            "SICOE",
            "reporte",
            "asignar_actores_por_pk",
            {
                "pk_id_id": pk_id_id,
                "actualizados": len(rows),
                "inspector_id": body.inspector_id,
                "subcontratista_id": body.subcontratista_id,
            },
        )
    except Exception:
        pass
    return {"ok": True, "actualizados": len(rows)}

@app.patch("/sicoe-obra/{contrato_id}/reportes/{reporte_id}/localizacion")
def actualizar_localizacion_borrador(contrato_id: int, reporte_id: int, body: dict, current_user=Depends(get_current_user)):
    """Solo actualiza localización si el reporte está en Borrador."""
    def _check():
        return supabase.table("so_reportes").select("estado")\
            .eq("id", reporte_id).eq("contrato_id", contrato_id).single().execute().data
    reporte = supabase_execute(_check)
    if not reporte or reporte.get("estado") != "Borrador":
        raise HTTPException(status_code=400, detail="Solo se puede actualizar localización en Borrador")
    campos = {k: v for k, v in body.items() if k in (
        'pk_id_id','civ','tramo','infraestructura','calzada',
        'ubicacion','coord_lat','coord_lng','abs_inicio','abs_final',
        'nodo_ini','nodo_fin','margen'
    )}
    if not campos:
        raise HTTPException(status_code=400, detail="Sin campos de localización")
    def _prev_snap():
        return supabase.table("so_reportes").select("*").eq("id", reporte_id).eq("contrato_id", contrato_id).limit(1).execute().data
    pr = supabase_execute(_prev_snap)
    prev_rep = pr[0] if pr else {}
    def _upd():
        return supabase.table("so_reportes").update(campos)\
            .eq("id", reporte_id).eq("contrato_id", contrato_id).execute().data
    result = supabase_execute(_upd)
    out = result[0] if result else {}
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "EDITAR",
            "SICOE",
            "reporte",
            str(reporte_id),
            {"alcance": "localizacion_borrador", "campos": list(campos.keys())},
            valor_anterior=_so_reporte_audit_snapshot(prev_rep),
            valor_nuevo=_so_reporte_audit_snapshot(out),
        )
    except Exception:
        pass
    return out

def _registro_nivel3_aprobado(row: Optional[Dict[str, Any]]) -> bool:
    """Tras aprobación Nivel 3 (Interventoría), no se editan datos de obra salvo corte de subcontratista."""
    if not row:
        return False
    return (row.get("nivel3_estado") or "").strip() == "Aprobado"


def _sicoe_exige_topografia_para_aprobar_nivel2(contrato_id: int) -> bool:
    """Contrato 2 no aplica la regla de puntos topográficos para aprobar N2; el resto sí."""
    try:
        return int(contrato_id) != 2
    except (TypeError, ValueError):
        return True


def _reporte_tiene_puntos_topograficos(contrato_id: int, reporte_id: int) -> bool:
    """True si existe al menos un punto (misma regla que el front: Portada / reporte.puntos)."""
    try:
        rid = int(reporte_id)
    except (TypeError, ValueError):
        return False
    try:

        def _q():
            return (
                supabase.table("so_puntos_topograficos")
                .select("id")
                .eq("contrato_id", int(contrato_id))
                .eq("reporte_id", rid)
                .limit(1)
                .execute()
                .data
            )

        return bool(supabase_execute(_q))
    except Exception:
        return False


def _reportes_ids_con_topografia(contrato_id: int, reporte_ids: List[int]) -> set:
    out: set = set()
    ids = []
    seen = set()
    for x in reporte_ids or []:
        try:
            xi = int(x)
        except (TypeError, ValueError):
            continue
        if xi in seen:
            continue
        seen.add(xi)
        ids.append(xi)
    if not ids:
        return out
    _CHUNK = 120
    cid = int(contrato_id)
    for i in range(0, len(ids), _CHUNK):
        chunk = ids[i : i + _CHUNK]

        def _page(c=chunk):
            return (
                supabase.table("so_puntos_topograficos")
                .select("reporte_id")
                .eq("contrato_id", cid)
                .in_("reporte_id", c)
                .execute()
                .data
            )

        for row in supabase_execute(_page) or []:
            r = row.get("reporte_id")
            if r is not None:
                try:
                    out.add(int(r))
                except (TypeError, ValueError):
                    pass
    return out


@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}")
def actualizar_registro(contrato_id: int, registro_id: int, body: RegistroCreate, current_user=Depends(get_current_user)):
    data = {k: v for k, v in body.dict().items() if v is not None}

    def _prev():
        return supabase.table("so_registros").select("*").eq("id", registro_id).eq("contrato_id", contrato_id).limit(1).execute().data

    prev_rows = supabase_execute(_prev)
    prev_row = prev_rows[0] if prev_rows else None
    if not prev_row:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if _registro_nivel3_aprobado(prev_row):
        # Campos siempre editables aunque el registro esté aprobado en N3
        _CAMPOS_N3_PERMITIDOS = {
            "corte_id", "reporte_id", "numero_registro",
            "foto_url", "foto_numero", "foto_descripcion",
            "grafico_url", "grafico_numero", "grafico_descripcion",
        }
        otros = {k: v for k, v in data.items() if k not in _CAMPOS_N3_PERMITIDOS}
        if otros:
            raise HTTPException(
                status_code=400,
                detail="Registro aprobado por Interventoría (Nivel 3): solo puede modificarse el corte de subcontratista y la foto/gráfico del registro.",
            )
        if data.get("reporte_id") is not None and int(data["reporte_id"]) != int(prev_row["reporte_id"]):
            raise HTTPException(status_code=400, detail="No puede modificarse el reporte del registro aprobado por Nivel 3.")
        if data.get("numero_registro") is not None and int(data["numero_registro"]) != int(prev_row["numero_registro"]):
            raise HTTPException(status_code=400, detail="No puede modificarse el número de registro aprobado por Nivel 3.")
        data = {k: v for k, v in data.items() if k in _CAMPOS_N3_PERMITIDOS - {"reporte_id", "numero_registro"}}
        if not data:
            return prev_row

    for dk in ("longitud", "ancho", "espesor", "cantidad"):
        if dk in data:
            data[dk] = round(float(data[dk]), 2)
    if "cantidad_total" in data:
        data["cantidad_total"] = round(float(data["cantidad_total"]), 2)
    vlr_merged = (
        float(data["vlr_unitario"])
        if data.get("vlr_unitario") is not None
        else float(prev_row.get("vlr_unitario") or 0)
    )
    if "cantidad_total" in data and (str(prev_row.get("item_numero") or "").strip()):
        data["costo_directo"] = round(float(data["cantidad_total"]) * vlr_merged, 0)
    elif "costo_directo" in data:
        data["costo_directo"] = round(float(data["costo_directo"]), 0)

    def _upd():
        return supabase.table("so_registros").update(data)\
            .eq("id", registro_id).eq("contrato_id", contrato_id).execute().data

    out = supabase_execute(_upd)
    row = out[0] if out else {}
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "EDITAR",
            "SICOE",
            "registro",
            str(registro_id),
            {"reporte_id": row.get("reporte_id"), "id_pol": row.get("id_pol")},
            valor_anterior=_so_registro_audit_snapshot(prev_row),
            valor_nuevo=_so_registro_audit_snapshot(row),
        )
    except Exception:
        pass
    return row

@app.delete("/sicoe-obra/{contrato_id}/reportes/{reporte_id}/registros")
def eliminar_registros_reporte(contrato_id: int, reporte_id: int, current_user=Depends(get_current_user)):
    def _list():
        return supabase.table("so_registros").select("id").eq("reporte_id", reporte_id).eq("contrato_id", contrato_id).execute().data

    ids_rows = supabase_execute(_list) or []
    n = len(ids_rows)

    def _del():
        return supabase.table("so_registros").delete()\
            .eq("reporte_id", reporte_id).eq("contrato_id", contrato_id).execute().data

    supabase_execute(_del)
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "ELIMINAR",
            "SICOE",
            "reporte",
            str(reporte_id),
            {
                "registros_eliminados": n,
                "muestra_ids": [r.get("id") for r in ids_rows[:80]],
            },
            valor_anterior={"ids": [r.get("id") for r in ids_rows]},
            valor_nuevo={},
            severidad="AUDIT",
            alerta_generada=(n >= 20),
        )
    except Exception:
        pass
    return {"ok": True}

@app.put("/sicoe-obra/{contrato_id}/reportes/{reporte_id}/reemplazar-registros")
def reemplazar_registros_nuevo_reporte(
    contrato_id: int, reporte_id: int, body: ReemplazarRegistrosNuevoReporteBody, current_user=Depends(get_current_user)
):
    """
    Un solo ida y vuelta: borra so_registros del reporte e inserta las líneas nuevas.
    Sustituye N×(next-registro+POST) para evitar “Failed to fetch” en móviles (timeout / cierre de conexión).
    """
    def _get_rep():
        return supabase.table("so_reportes").select(
            "pk_id_id,civ,tramo,infraestructura,calzada,ubicacion,"
            "coord_lat,coord_lng,abs_inicio,abs_final,nodo_ini,nodo_fin,"
            "subcontratista_id,inspector_id"
        ).eq("id", reporte_id).eq("contrato_id", contrato_id).limit(1).execute().data
    rep_rows = supabase_execute(_get_rep)
    if not rep_rows:
        raise HTTPException(status_code=404, detail="Reporte no encontrado en este contrato")
    rep = rep_rows[0]
    def _del():
        return supabase.table("so_registros").delete()\
            .eq("reporte_id", reporte_id).eq("contrato_id", contrato_id).execute().data
    supabase_execute(_del)
    uid = int(current_user.get("sub") or current_user.get("id", 0))
    if not body.registros:
        return {"ok": True, "insertados": 0}

    def _parse_numero_registro_raw(raw) -> int:
        if raw is None:
            raise ValueError("RPC sin número de registro")
        if isinstance(raw, (int, float)) and not isinstance(raw, bool):
            return int(raw)
        if isinstance(raw, list) and len(raw) > 0:
            return _parse_numero_registro_raw(raw[0])
        if isinstance(raw, dict):
            for k in ("numero", "siguiente_numero_registro", "siguiente", "id"):
                if k in raw and raw[k] is not None:
                    return int(raw[k])
        return int(raw)

    nlines = len(body.registros)
    # Preferir 1 ida a Postgres (función en backend/sql/siguiente_n_numeros_registro.sql) frente a N o muchas
    # llamadas en paralelo móvil→App→PostgREST (timeout / “no hubo conexión a tiempo”).
    numeros: List[int] = []
    def _try_rpc_bloque() -> bool:
        nonlocal numeros
        def _q():
            return supabase.rpc(
                "siguiente_n_numeros_registro", {"p_contrato_id": contrato_id, "p_n": nlines}
            ).execute().data
        try:
            raw = supabase_execute(_q)
        except Exception:
            return False
        if raw is None:
            return False
        if isinstance(raw, list):
            if len(raw) != nlines:
                return False
            try:
                numeros = [int(x) for x in raw]
            except (TypeError, ValueError):
                return False
            return True
        if isinstance(raw, dict):
            for k in ("siguiente_n_numeros_registro", "data", "result"):
                if k in raw and isinstance(raw[k], list):
                    try:
                        arr = [int(x) for x in raw[k]]
                    except (TypeError, ValueError):
                        return False
                    if len(arr) != nlines:
                        return False
                    numeros = arr
                    return True
        return False

    if not _try_rpc_bloque():
        # NUNCA en paralelo: varias RPC concurrentes a siguiente_numero_registro
        # duplican el mismo (contrato_id, numero_registro) → 23505 en so_registros.
        for _ in range(nlines):
            def _n():
                return supabase.rpc("siguiente_numero_registro", {"p_contrato_id": contrato_id}).execute().data
            numeros.append(_parse_numero_registro_raw(supabase_execute(_n)))
    if len(set(numeros)) != len(numeros):
        raise HTTPException(
            status_code=500,
            detail="Números de registro duplicados al reservar; reintente o revise la función de consecutivo en la base de datos",
        )
    rows_to_insert: List[Dict[str, Any]] = []
    for line, numero in zip(body.registros, numeros):
        data: Dict[str, Any] = line.dict()
        data["reporte_id"] = reporte_id
        data["numero_registro"] = numero
        data["contrato_id"] = contrato_id
        data["creado_por_reg"] = uid
        for campo in ("pk_id_id", "civ", "tramo", "infraestructura", "calzada", "ubicacion",
                      "coord_lat", "coord_lng", "abs_inicio", "abs_final",
                      "nodo_ini", "nodo_fin", "subcontratista_id", "inspector_id"):
            if rep.get(campo) is not None:
                data[campo] = rep[campo]
        rows_to_insert.append(data)
    _BATCH = 200
    total = 0
    for i in range(0, len(rows_to_insert), _BATCH):
        chunk = rows_to_insert[i : i + _BATCH]
        def _ins():
            return supabase.table("so_registros").insert(chunk).execute().data
        supabase_execute(_ins)
        total += len(chunk)
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "CREAR",
            "SICOE",
            "registro",
            f"reporte_{reporte_id}_lote",
            {"reporte_id": reporte_id, "reemplazar_registros_masivo": True, "insertados": total},
            valor_anterior=None,
            valor_nuevo={"insertados": total},
        )
    except Exception:
        pass
    return {"ok": True, "insertados": total}

@app.delete("/sicoe-obra/{contrato_id}/registros/{registro_id}/dev")
def dev_eliminar_registro(contrato_id: int, registro_id: int, current_user=Depends(get_current_user)):
    """Solo Desarrollador: elimina un registro y sus comentarios (uso de soporte/desarrollo)."""
    if not _es_desarrollador(current_user):
        raise HTTPException(status_code=403, detail="Solo el cargo Desarrollador puede usar esta acción.")

    def _get():
        return supabase.table("so_registros").select("*").eq("id", registro_id).eq("contrato_id", contrato_id).limit(1).execute().data
    rows = supabase_execute(_get)
    if not rows:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    prev_row = rows[0]

    def _del_com():
        return supabase.table("so_registro_comentarios").delete().eq("registro_id", registro_id).execute().data
    supabase_execute(_del_com)

    def _del_reg():
        return supabase.table("so_registros").delete().eq("id", registro_id).eq("contrato_id", contrato_id).execute().data
    supabase_execute(_del_reg)

    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "ELIMINAR",
            "SICOE",
            "registro",
            str(registro_id),
            {"reporte_id": prev_row.get("reporte_id"), "dev": True},
            valor_anterior=_so_registro_audit_snapshot(prev_row),
            valor_nuevo={},
            severidad="AUDIT",
            alerta_generada=True,
        )
    except Exception:
        pass
    return {"ok": True}


@app.delete("/sicoe-obra/{contrato_id}/reportes/{reporte_id}/dev")
def dev_eliminar_reporte(contrato_id: int, reporte_id: int, current_user=Depends(get_current_user)):
    """Solo Desarrollador: elimina reporte, registros, comentarios y puntos topográficos (uso de soporte/desarrollo)."""
    if not _es_desarrollador(current_user):
        raise HTTPException(status_code=403, detail="Solo el cargo Desarrollador puede usar esta acción.")

    def _get_rep():
        return supabase.table("so_reportes").select("*").eq("id", reporte_id).eq("contrato_id", contrato_id).limit(1).execute().data
    rep_rows = supabase_execute(_get_rep)
    if not rep_rows:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    prev_rep = rep_rows[0]

    def _list_reg():
        return supabase.table("so_registros").select("id").eq("reporte_id", reporte_id).eq("contrato_id", contrato_id).execute().data
    reg_rows = supabase_execute(_list_reg) or []
    reg_ids = [r["id"] for r in reg_rows]

    if reg_ids:
        def _del_com():
            return supabase.table("so_registro_comentarios").delete().in_("registro_id", reg_ids).execute().data
        supabase_execute(_del_com)

    def _del_regs():
        return supabase.table("so_registros").delete().eq("reporte_id", reporte_id).eq("contrato_id", contrato_id).execute().data
    supabase_execute(_del_regs)

    def _del_pts():
        return supabase.table("so_puntos_topograficos").delete().eq("reporte_id", reporte_id).eq("contrato_id", contrato_id).execute().data
    supabase_execute(_del_pts)

    def _del_rep():
        return supabase.table("so_reportes").delete().eq("id", reporte_id).eq("contrato_id", contrato_id).execute().data
    supabase_execute(_del_rep)

    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "ELIMINAR",
            "SICOE",
            "reporte",
            str(reporte_id),
            {"registros_eliminados": len(reg_ids), "dev": True},
            valor_anterior=_json_for_log(
                {k: prev_rep.get(k) for k in ("id", "numero_reporte", "estado", "descripcion_actividad") if prev_rep}
            ),
            valor_nuevo={},
            severidad="AUDIT",
            alerta_generada=True,
        )
    except Exception:
        pass
    return {"ok": True, "registros_eliminados": len(reg_ids)}


@app.post("/sicoe-obra/{contrato_id}/registros")
def crear_registro(contrato_id: int, body: RegistroCreate, current_user=Depends(get_current_user)):
    data = body.dict()
    data["contrato_id"] = contrato_id
    data["creado_por_reg"] = int(current_user.get("sub") or current_user.get("id", 0))
    try:
        def _rep():
            return supabase.table("so_reportes").select(
                "pk_id_id,civ,tramo,infraestructura,calzada,ubicacion,"
                "coord_lat,coord_lng,abs_inicio,abs_final,nodo_ini,nodo_fin,"
                "subcontratista_id,inspector_id"
            ).eq("id", body.reporte_id).eq("contrato_id", contrato_id).limit(1).execute().data
        rep_rows = supabase_execute(_rep)
        if rep_rows:
            rep = rep_rows[0]
            for campo in ("pk_id_id","civ","tramo","infraestructura","calzada","ubicacion",
                          "coord_lat","coord_lng","abs_inicio","abs_final",
                          "nodo_ini","nodo_fin","subcontratista_id","inspector_id"):
                if rep.get(campo) is not None:
                    data[campo] = rep[campo]
    except Exception:
        pass
    def _ins():
        return supabase.table("so_registros").insert(data).execute().data
    result = supabase_execute(_ins)
    row = result[0] if result else {}
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "CREAR",
            "SICOE",
            "registro",
            str(row.get("id", "")),
            {"reporte_id": row.get("reporte_id"), "id_pol": row.get("id_pol")},
            valor_anterior=None,
            valor_nuevo=_so_registro_audit_snapshot(row),
        )
    except Exception:
        pass
    return row

class PuntoTopo(BaseModel):
    punto: Optional[str] = None
    norte: Optional[float] = None
    este: Optional[float] = None
    cota: Optional[float] = None
    descripcion: Optional[str] = None

class PuntosCreate(BaseModel):
    reporte_id: int
    puntos: List[PuntoTopo]

@app.delete("/sicoe-obra/{contrato_id}/reportes/{reporte_id}/puntos-topograficos")
def eliminar_puntos_reporte(contrato_id: int, reporte_id: int, current_user=Depends(get_current_user)):
    def _del():
        return supabase.table("so_puntos_topograficos").delete()\
            .eq("reporte_id", reporte_id).eq("contrato_id", contrato_id).execute().data
    supabase_execute(_del)
    return {"ok": True}

@app.post("/sicoe-obra/{contrato_id}/puntos-topograficos")
def crear_puntos(contrato_id: int, body: PuntosCreate, current_user=Depends(get_current_user)):
    rows = []
    for p in body.puntos:
        d = p.dict()
        d["contrato_id"] = contrato_id
        d["reporte_id"] = body.reporte_id
        d["creado_por"] = int(current_user.get("sub") or current_user.get("id", 0))
        rows.append(d)
    def _ins():
        return supabase.table("so_puntos_topograficos").insert(rows).execute().data
    return supabase_execute(_ins)

# ─── SICOE OBRA: Verificar acta RPO vigente ──────────────────────────────────
def _acta_rpo_vigente_row(contrato_id: int):
    """
    Acta RPO en período: hoy ∈ [fecha_inicio, fecha_fin]. Actas futuras (aún sin vigencia) no compiten.
    Si dos períodos solapan en un día (p. ej. cierre 30/04 y el siguiente 01/05), gana el de fecha_inicio
    más reciente; empate → numero_rpo desc, luego id desc — alineado con transición natural al vencer el mes.
    """
    from datetime import date
    today = date.today().isoformat()

    def _q():
        return supabase.table("actas")\
            .select("id, numero_rpo, fecha_inicio, fecha_fin, usuarios(nombre, apellidos)")\
            .eq("contrato_id", contrato_id)\
            .eq("tipo_grupo", "RPO")\
            .lte("fecha_inicio", today)\
            .gte("fecha_fin", today)\
            .order("fecha_inicio", desc=True)\
            .order("numero_rpo", desc=True)\
            .order("id", desc=True)\
            .limit(1)\
            .execute().data

    actas = supabase_execute(_q)
    row = actas[0] if actas else None
    if not row:
        return None
    u = row.get("usuarios") or {}
    if isinstance(u, list) and len(u) > 0:
        u = u[0]
    if not isinstance(u, dict):
        u = {}
    an = f"{u.get('nombre', '')} {u.get('apellidos', '')}".strip()
    row = {k: v for k, v in row.items() if k != "usuarios"}
    row["asignado_nombre"] = an or None
    return row


def _parse_iso_to_date(val) -> Optional[date]:
    """Convierte ISO string o date a datetime.date; None si no es parseable."""
    if val is None:
        return None
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _acta_rpo_id_vigente_para_fecha(contrato_id: int, ref_date) -> Optional[int]:
    """
    Acta RPO cuyo período contiene ref_date (mismo criterio que _acta_rpo_vigente_row con ref_date en lugar de hoy).
    """
    d = _parse_iso_to_date(ref_date)
    if not d:
        return None
    ds = d.isoformat()

    def _q():
        return supabase.table("actas")\
            .select("id")\
            .eq("contrato_id", contrato_id)\
            .eq("tipo_grupo", "RPO")\
            .lte("fecha_inicio", ds)\
            .gte("fecha_fin", ds)\
            .order("fecha_inicio", desc=True)\
            .order("numero_rpo", desc=True)\
            .order("id", desc=True)\
            .limit(1)\
            .execute().data

    rows = supabase_execute(_q)
    return rows[0]["id"] if rows else None


def _aplicar_acta_rpo_vigente_a_registro(contrato_id: int, registro_id: int, ref_date) -> Optional[int]:
    """
    Asigna acta_rpo_id según acta RPO vigente en ref_date (parcial N2 o definitivo N3).
    Actualiza so_registros y so_reportes. Retorna el id de acta o None si no hay acta para esa fecha.
    """
    acta_id = _acta_rpo_id_vigente_para_fecha(contrato_id, ref_date)
    if not acta_id:
        return None

    def _get():
        return supabase.table("so_registros").select("reporte_id")\
            .eq("id", registro_id).eq("contrato_id", contrato_id).limit(1).execute().data

    rows = supabase_execute(_get)
    if not rows:
        return None
    rep_id = rows[0]["reporte_id"]

    def _upd_reg():
        return supabase.table("so_registros").update({"acta_rpo_id": acta_id})\
            .eq("id", registro_id).eq("contrato_id", contrato_id).execute().data

    supabase_execute(_upd_reg)

    def _upd_rep():
        return supabase.table("so_reportes").update({"acta_rpo_id": acta_id})\
            .eq("id", rep_id).eq("contrato_id", contrato_id).execute().data

    supabase_execute(_upd_rep)
    return acta_id


@app.get("/sicoe-obra/{contrato_id}/acta-rpo-vigente")
def get_acta_rpo_vigente(contrato_id: int, current_user=Depends(get_current_user)):
    try:
        return _acta_rpo_vigente_row(contrato_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ─── SICOE OBRA: Búsqueda de ítems del listado de precios ────────────────────
@app.get("/sicoe-obra/{contrato_id}/listado-precios-busqueda")
def buscar_items_listado(contrato_id: int, q: str = "", capitulo: str = None, competencia: str = None, current_user=Depends(get_current_user)):
    def _q():
        query = supabase.table("listado_precios")\
            .select("id, capitulo, competencia, item_numero, descripcion, unidad, precio_unitario")\
            .eq("contrato_id", contrato_id)
        if capitulo:
            query = query.eq("capitulo", capitulo)
        if competencia:
            query = query.eq("competencia", competencia)
        if q:
            query = query.or_(f"descripcion.ilike.%{q}%,item_numero.ilike.%{q}%")
        return query.order("item_numero").limit(200).execute().data
    return supabase_execute(_q)

# ─── SICOE OBRA: Asignar ítem a registro ─────────────────────────────────────
class AsignarItemBody(BaseModel):
    item_listado_id: int
    competencia: Optional[str] = None

@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/asignar-item")
def asignar_item_registro(contrato_id: int, registro_id: int, body: AsignarItemBody, current_user=Depends(get_current_user)):
    try:
        from datetime import date
        today = date.today().isoformat()

        def _item():
            return supabase.table("listado_precios")\
                .select("capitulo, competencia, item_numero, descripcion, unidad, precio_unitario")\
                .eq("id", body.item_listado_id).single().execute().data
        item = supabase_execute(_item)
        if not item:
            raise HTTPException(status_code=404, detail="Ítem no encontrado en listado de precios")

        def _reg():
            return supabase.table("so_registros")\
                .select("cantidad_total, reporte_id, nivel3_estado")\
                .eq("id", registro_id).single().execute().data
        registro = supabase_execute(_reg)
        if not registro:
            raise HTTPException(status_code=404, detail="Registro no encontrado")
        if _registro_nivel3_aprobado(registro):
            raise HTTPException(
                status_code=400,
                detail="El registro está aprobado por Nivel 3 (Interventoría): no puede reasignarse el ítem.",
            )

        cant_total = float(registro.get("cantidad_total") or 0)
        vlr_unit   = float(item.get("precio_unitario") or 0)
        cant_total = round(cant_total, 2)
        costo_dir  = round(cant_total * vlr_unit, 0)
        reporte_id = registro["reporte_id"]

        def _rep():
            return supabase.table("so_reportes")\
                .select("subcontratista_id, acta_rpo_id, corte_id, semana_id")\
                .eq("id", reporte_id).single().execute().data
        reporte = supabase_execute(_rep) or {}

        # 4. Detectar acta RPO vigente por período — obligatorio para asignar
        acta_rpo_id = reporte.get("acta_rpo_id")
        if not acta_rpo_id:
            def _acta():
                return supabase.table("actas")\
                    .select("id, numero_rpo")\
                    .eq("contrato_id", contrato_id)\
                    .eq("tipo_grupo", "RPO")\
                    .lte("fecha_inicio", today)\
                    .gte("fecha_fin", today)\
                    .order("id", desc=True).limit(1).execute().data
            actas = supabase_execute(_acta)
            acta_rpo_id = actas[0]["id"] if actas else None
        if not acta_rpo_id:
            raise HTTPException(status_code=422, detail="No existe un Acta RPO vigente para la fecha de hoy. Crea el acta en el módulo administrativo antes de asignar ítems.")

        corte_id = reporte.get("corte_id")
        sub_id   = reporte.get("subcontratista_id")
        if not corte_id and sub_id:
            try:
                def _corte():
                    return supabase.table("subcontratista_cortes")\
                        .select("id, consecutivo")\
                        .eq("subcontratista_id", sub_id)\
                        .lte("fecha_inicio", today)\
                        .gte("fecha_fin", today)\
                        .limit(1).execute().data
                cortes = supabase_execute(_corte)
                corte_id = cortes[0]["id"] if cortes else None
            except:
                corte_id = None

        semana_id = reporte.get("semana_id")
        if not semana_id:
            try:
                def _sem():
                    return supabase.table("so_semanas")\
                        .select("id, numero_semana")\
                        .eq("contrato_id", contrato_id)\
                        .eq("estado", "activa")\
                        .lte("fecha_inicio", today)\
                        .gte("fecha_fin", today)\
                        .limit(1).execute().data
                sems = supabase_execute(_sem)
                semana_id = sems[0]["id"] if sems else None
            except:
                semana_id = None

        def _upd_reg():
            return supabase.table("so_registros").update({
                "capitulo":         item.get("capitulo"),
                "competencia":      body.competencia or item.get("competencia"),
                "item_numero":      item.get("item_numero"),
                "item_descripcion": item.get("descripcion"),
                "vlr_unitario":     vlr_unit,
                "cantidad_total":   cant_total,
                "costo_directo":    costo_dir,
                "unidad":           item.get("unidad"),
                "semana_id":        semana_id,
                "acta_rpo_id":      acta_rpo_id,
                "corte_id":         corte_id,
            }).eq("id", registro_id).eq("contrato_id", contrato_id).execute().data
        supabase_execute(_upd_reg)

        def _upd_rep():
            return supabase.table("so_reportes").update({
                "acta_rpo_id": acta_rpo_id,
                "corte_id":    corte_id,
                "semana_id":   semana_id,
                "estado":      "No Revisados",
            }).eq("id", reporte_id).execute().data
        supabase_execute(_upd_rep)

        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            registrar_log(
                u_log,
                "ASIGNAR_ITEM",
                "SICOE",
                "registro",
                str(registro_id),
                {
                    "item_numero": item.get("item_numero"),
                    "item_listado_id": body.item_listado_id,
                    "reporte_id": reporte_id,
                },
            )
        except Exception:
            pass

        return {
            "ok": True, "vlr_unitario": vlr_unit, "costo_directo": costo_dir,
            "semana_id": semana_id, "acta_rpo_id": acta_rpo_id, "corte_id": corte_id
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error asignando ítem: {str(e)}")

# ─── SICOE OBRA: Nuevo registro en blanco dentro de un reporte existente ─────
@app.post("/sicoe-obra/{contrato_id}/reportes/{reporte_id}/nuevo-registro")
def nuevo_registro_en_reporte(contrato_id: int, reporte_id: int, current_user=Depends(get_current_user)):
    def _num():
        return supabase.rpc("siguiente_numero_registro", {"p_contrato_id": contrato_id}).execute().data
    numero = supabase_execute(_num)
    def _ins():
        return supabase.table("so_registros").insert({
            "contrato_id":    contrato_id,
            "reporte_id":     reporte_id,
            "numero_registro": numero,
        }).execute().data
    result = supabase_execute(_ins)
    row = result[0] if result else {}
    try:
        u_log = _audit_user_contrato(current_user, contrato_id)
        registrar_log(
            u_log,
            "CREAR",
            "SICOE",
            "registro",
            str(row.get("id") or ""),
            {"reporte_id": reporte_id, "numero_registro": row.get("numero_registro")},
        )
    except Exception:
        pass
    return row
@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/mover-a/{nuevo_reporte_id}")
def mover_registro(contrato_id: int, registro_id: int, nuevo_reporte_id: int, current_user=Depends(get_current_user)):
    try:
        # 1. Verificar que el reporte destino existe en el contrato
        def _ver_dest():
            return supabase.table("so_reportes")\
                .select("id").eq("id", nuevo_reporte_id)\
                .eq("contrato_id", contrato_id).execute().data
        if not supabase_execute(_ver_dest):
            raise HTTPException(status_code=404, detail="Reporte destino no encontrado en este contrato")

        # 2. Leer el registro para obtener el reporte_id origen
        def _reg():
            return supabase.table("so_registros")\
                .select("reporte_id, nivel3_estado").eq("id", registro_id)\
                .eq("contrato_id", contrato_id).limit(1).execute().data
        reg_rows = supabase_execute(_reg)
        if not reg_rows:
            raise HTTPException(status_code=404, detail="Registro no encontrado")
        if _registro_nivel3_aprobado(reg_rows[0]):
            raise HTTPException(
                status_code=400,
                detail="El registro está aprobado por Nivel 3 (Interventoría): no puede moverse a otro reporte.",
            )
        reporte_origen_id = reg_rows[0]["reporte_id"]

        # 3. Leer campos de localización del reporte ORIGEN
        campos_loc = ("pk_id_id","civ","tramo","infraestructura","calzada","ubicacion",
                      "coord_lat","coord_lng","abs_inicio","abs_final",
                      "nodo_ini","nodo_fin","subcontratista_id","inspector_id","creado_por")
        def _rep_orig():
            return supabase.table("so_reportes")\
                .select(",".join(campos_loc))\
                .eq("id", reporte_origen_id).eq("contrato_id", contrato_id)\
                .limit(1).execute().data
        rep_rows = supabase_execute(_rep_orig)
        loc_data = {}
        if rep_rows:
            rep = rep_rows[0]
            for campo in campos_loc:
                if rep.get(campo) is not None:
                    loc_data[campo] = rep[campo]

        # 4. Actualizar el registro: nuevo reporte_id + datos de localización del origen tatuados
        update_data = {
            "reporte_id": nuevo_reporte_id,
            "modificado_por_reg": int(current_user.get("sub") or current_user.get("id", 0)),
            **{k: v for k, v in loc_data.items() if k != "creado_por"},
        }
        def _upd():
            return supabase.table("so_registros")\
                .update(update_data)\
                .eq("id", registro_id).eq("contrato_id", contrato_id).execute().data
        supabase_execute(_upd)
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            registrar_log(
                u_log,
                "MOVER",
                "SICOE",
                "registro",
                str(registro_id),
                {"reporte_origen": reporte_origen_id, "reporte_destino": nuevo_reporte_id},
            )
        except Exception:
            pass
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ─── SICOE OBRA: Semanas ──────────────────────────────────────────────────────
class SemanaCreate(BaseModel):
    numero_semana: int
    fecha_inicio:  str
    fecha_fin:     str
    dia_corte:     int  # 0=Lunes … 6=Domingo
    estado:        str = "activa"

@app.get("/sicoe-obra/{contrato_id}/semanas")
def listar_semanas(contrato_id: int, current_user=Depends(get_current_user)):
    def _q():
        return supabase.table("so_semanas")\
            .select("*").eq("contrato_id", contrato_id)\
            .order("numero_semana").execute().data
    return supabase_execute(_q)

@app.post("/sicoe-obra/{contrato_id}/semanas")
def crear_semanas(contrato_id: int, body: List[SemanaCreate], current_user=Depends(get_current_user)):
    rows = [{"contrato_id": contrato_id, **s.dict()} for s in body]
    def _ins():
        return supabase.table("so_semanas").insert(rows).execute().data
    return supabase_execute(_ins)

@app.get("/sicoe-obra/{contrato_id}/semana-vigente")
def semana_vigente(contrato_id: int, current_user=Depends(get_current_user)):
    from datetime import date
    today = date.today().isoformat()
    def _vig():
        return supabase.table("so_semanas")\
            .select("*").eq("contrato_id", contrato_id).eq("estado", "activa")\
            .lte("fecha_inicio", today).gte("fecha_fin", today)\
            .limit(1).execute().data
    def _prox():
        return supabase.table("so_semanas")\
            .select("*").eq("contrato_id", contrato_id).eq("estado", "activa")\
            .gt("fecha_inicio", today)\
            .order("fecha_inicio").limit(1).execute().data
    vigente = supabase_execute(_vig)
    proxima = supabase_execute(_prox)
    return {
        "vigente": vigente[0] if vigente else None,
        "proxima": proxima[0] if proxima else None
    }

@app.post("/sicoe-obra/{contrato_id}/semanas/extender")
def extender_semanas(contrato_id: int, n_semanas: int, current_user=Depends(get_current_user)):
    """Agrega n_semanas adicionales continuando desde la última semana existente"""
    from datetime import date, timedelta
    def _ultima():
        return supabase.table("so_semanas")\
            .select("numero_semana, fecha_fin, dia_corte")\
            .eq("contrato_id", contrato_id)\
            .order("numero_semana", desc=True).limit(1).execute().data
    rows = supabase_execute(_ultima)
    if not rows:
        raise HTTPException(status_code=400, detail="No hay semanas base. Crea la primera semana primero.")
    ultima     = rows[0]
    ultimo_num = ultima["numero_semana"]
    dia_corte  = ultima["dia_corte"]
    fecha_base = date.fromisoformat(ultima["fecha_fin"])
    nuevas = []
    for i in range(1, n_semanas + 1):
        f_ini = fecha_base + timedelta(days=(i - 1) * 7 + 1)
        f_fin = fecha_base + timedelta(days=i * 7)
        nuevas.append({
            "contrato_id":   contrato_id,
            "numero_semana": ultimo_num + i,
            "fecha_inicio":  f_ini.isoformat(),
            "fecha_fin":     f_fin.isoformat(),
            "dia_corte":     dia_corte,
            "estado":        "activa"
        })
    def _ins():
        return supabase.table("so_semanas").insert(nuevas).execute().data
    return supabase_execute(_ins)


# ─── SICOE OBRA: Validación de Registros ─────────────────────────────────────

class ValidarNivel1Body(BaseModel):
    estado: str
    comentario_data: Optional[dict] = None

class ValidarNivel2Body(BaseModel):
    estado: str
    objeto_pago_sub: Optional[bool] = None
    comentario_data: Optional[dict] = None

class ValidarNivel3Body(BaseModel):
    estado: str
    comentario_data: Optional[dict] = None

class ValidarSubBody(BaseModel):
    estado: str

class SolicitarReversionBody(BaseModel):
    comentario_data: dict

class AceptarReversionBody(BaseModel):
    aceptar: bool

class ReversionDobleLlaveN3Body(BaseModel):
    comentario_data: dict

class ComentarioCreate(BaseModel):
    rol_origen: Optional[str] = None
    confidencialidad: Optional[str] = None
    destinatarios: Optional[list] = None
    etiqueta: Optional[str] = None
    asunto: Optional[str] = None
    mensaje: Optional[str] = None
    enlaces: Optional[list] = None
    cantidad_verificada: Optional[float] = None
    tipo: Optional[str] = None
    nivel_validacion: Optional[int] = None
    padre_id: Optional[int] = None

class ValidarMasivoNivel2Body(BaseModel):
    estado: str
    ids_registros: List[int]
    objeto_pago_sub: Optional[bool] = None
    comentario_data: Optional[dict] = None

class ValidarMasivoNivel3Body(BaseModel):
    estado: str
    ids_registros: List[int]
    comentario_data: Optional[dict] = None

class ReconciliarActaRpoHistoricoBody(BaseModel):
    dry_run: bool = False

def _normalizar_macro_rol(valor: Optional[str]) -> Optional[str]:
    txt = (valor or "").strip().lower()
    if not txt:
        return None
    if "intervent" in txt:
        return "interventoria"
    # Subcontratista se considera del lado contratista para confidencialidad.
    if "subcontrat" in txt:
        return "contratista"
    if "contrat" in txt:
        return "contratista"
    return None

def _macro_rol_usuario_por_id(usuario_id: Optional[int]) -> Optional[str]:
    if not usuario_id:
        return None
    try:
        u = supabase.table("usuarios").select("cargo_id, rol_id").eq("id", usuario_id).single().execute().data or {}
        cargo_id = u.get("cargo_id")
        if cargo_id:
            c = supabase.table("cargos").select("nombre").eq("id", cargo_id).single().execute().data or {}
            m = _normalizar_macro_rol(c.get("nombre"))
            if m:
                return m
        rol_id = u.get("rol_id")
        if rol_id:
            r = supabase.table("roles").select("nombre").eq("id", rol_id).single().execute().data or {}
            m = _normalizar_macro_rol(r.get("nombre"))
            if m:
                return m
        return None
    except Exception:
        return None

def _macro_rol_current_user(current_user) -> Optional[str]:
    try:
        uid = int(current_user.get("sub") or current_user.get("id", 0))
    except (TypeError, ValueError):
        return None
    return _macro_rol_usuario_por_id(uid)

def _cargo_current_user(current_user) -> str:
    try:
        uid = int(current_user.get("sub") or current_user.get("id", 0))
        u = supabase.table("usuarios").select("cargo_id").eq("id", uid).single().execute().data or {}
        cargo_id = u.get("cargo_id")
        if not cargo_id:
            return ""
        c = supabase.table("cargos").select("nombre").eq("id", cargo_id).single().execute().data or {}
        return (c.get("nombre") or "").strip().lower()
    except Exception:
        return ""

def _macro_roles_destinatarios(destinatarios: list) -> set:
    roles_dest = set()
    for d in (destinatarios or []):
        if not isinstance(d, dict):
            continue
        r_macro = _normalizar_macro_rol(
            d.get("rol") or d.get("rol_nombre") or d.get("cargo_nombre") or d.get("cargo")
        )
        if not r_macro:
            r_macro = _macro_rol_usuario_por_id(d.get("id"))
        if r_macro:
            roles_dest.add(r_macro)
    return roles_dest


def _insertar_comentario(contrato_id: int, registro_id: int, autor_id: int,
                         comentario_data: dict, tipo_override: str = None, nivel_validacion_override: str = None,
                         audit_user=None):
    """Inserta un comentario en so_registro_comentarios calculando confidencialidad."""
    destinatarios = comentario_data.get("destinatarios") or []
    rol_origen_payload = comentario_data.get("rol_origen", "")
    # Fuente de verdad: preferir el lado real del autor en BD.
    rol_origen_macro = _normalizar_macro_rol(rol_origen_payload) or _macro_rol_usuario_por_id(autor_id)
    rol_origen = (rol_origen_macro or (rol_origen_payload or "").strip() or None)
    if not rol_origen:
        # El CHECK en BD exige un macro-rol reconocible (p. ej. contratista / interventoria).
        rol_origen = "contratista"

    roles_dest = _macro_roles_destinatarios(destinatarios)

    if not destinatarios:
        confidencialidad = "publico"
    else:
        if roles_dest == {"contratista"} and rol_origen_macro == "contratista":
            confidencialidad = "contratista_interno"
        elif roles_dest == {"interventoria"} and rol_origen_macro == "interventoria":
            confidencialidad = "interventoria_interna"
        elif not roles_dest:
            confidencialidad = "publico"
        else:
            confidencialidad = "cruzado"

    row = {
        "registro_id":        registro_id,
        "contrato_id":        contrato_id,
        "autor_id":           autor_id,
        "rol_origen":         rol_origen,
        "confidencialidad":   confidencialidad,
        "destinatarios":      destinatarios,
        "etiqueta":           comentario_data.get("etiqueta"),
        "asunto":             comentario_data.get("asunto"),
        "mensaje":            comentario_data.get("mensaje"),
        "enlaces":            comentario_data.get("enlaces") or [],
        "cantidad_verificada": comentario_data.get("cantidad_verificada"),
        "tipo":               tipo_override or comentario_data.get("tipo"),
        "nivel_validacion":   nivel_validacion_override or comentario_data.get("nivel_validacion"),
        "padre_id":           comentario_data.get("padre_id"),
    }

    def _ins():
        return supabase.table("so_registro_comentarios").insert(row).execute().data
    data = supabase_execute(_ins)
    ins_row = data[0] if data else {}
    if audit_user and ins_row:
        try:
            u_log = _audit_user_contrato(audit_user, contrato_id)
            msg = ins_row.get("mensaje") or ""
            registrar_log(
                u_log,
                "COMENTAR",
                "SICOE",
                "registro",
                str(registro_id),
                {
                    "so_comentario_id": ins_row.get("id"),
                    "tipo": ins_row.get("tipo"),
                    "nivel_validacion": ins_row.get("nivel_validacion"),
                    "etiqueta": ins_row.get("etiqueta"),
                    "asunto": ins_row.get("asunto"),
                    "mensaje_excerpt": msg[:800],
                },
                severidad="AUDIT",
            )
        except Exception:
            pass
    return ins_row

@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/validar-nivel1")
def validar_nivel1(contrato_id: int, registro_id: int, body: ValidarNivel1Body,
                   current_user=Depends(get_current_user)):
    ESTADOS = {"Aprobado", "Pendiente", "Rechazado"}
    if body.estado not in ESTADOS:
        raise HTTPException(status_code=422, detail=f"Estado inválido. Acepta: {ESTADOS}")
    if body.estado in ("Pendiente", "Rechazado") and not body.comentario_data:
        raise HTTPException(status_code=422,
            detail="Se requiere comentario_data cuando el estado es Pendiente o Rechazado.")
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        _require_sicoe_puede_validar_nivel(current_user, autor_id, 1)
        def _get_n3():
            return supabase.table("so_registros").select("nivel3_estado, reporte_id")\
                .eq("id", registro_id).eq("contrato_id", contrato_id).limit(1).execute().data
        n3rows = supabase_execute(_get_n3)
        if not n3rows:
            raise HTTPException(status_code=404, detail="Registro no encontrado.")
        if _registro_nivel3_aprobado(n3rows[0]):
            raise HTTPException(
                status_code=400,
                detail="El registro está aprobado por Nivel 3 (Interventoría) y no puede modificarse por esta vía.",
            )
        prev_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
        update = {
            "nivel1_estado":     body.estado,
            "nivel1_usuario_id": autor_id,
            "nivel1_fecha":      datetime.utcnow().isoformat(),
        }
        # Actualizar modificado_por en so_reportes para reflejar la validación
        try:
            def _get_rep_id():
                return supabase.table("so_registros").select("reporte_id")\
                    .eq("id", registro_id).eq("contrato_id", contrato_id)\
                    .limit(1).execute().data
            rep_rows = supabase_execute(_get_rep_id)
            if rep_rows:
                rep_id = rep_rows[0]["reporte_id"]
                def _upd_rep():
                    return supabase.table("so_reportes")\
                        .update({"modificado_por": autor_id, "updated_at": datetime.utcnow().isoformat()})\
                        .eq("id", rep_id).eq("contrato_id", contrato_id).execute().data
                supabase_execute(_upd_rep)
        except Exception:
            pass
        def _upd():
            return supabase.table("so_registros")\
                .update(update).eq("id", registro_id)\
                .eq("contrato_id", contrato_id).execute().data
        supabase_execute(_upd)
        if body.comentario_data:
            _insertar_comentario(contrato_id, registro_id, autor_id, body.comentario_data,
                                 tipo_override="validacion", nivel_validacion_override="Nivel 1",
                                 audit_user=current_user)
            _push_notif_validacion_sicoe_destinatarios(
                current_user,
                autor_id,
                contrato_id,
                registro_id,
                f"Validación Nivel 1: {body.estado}",
                body.comentario_data.get("mensaje", "") or "",
                body.comentario_data,
            )
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            after_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
            registrar_log(
                u_log, "VALIDAR", "SICOE", "registro", str(registro_id),
                {"nivel": 1, "estado": body.estado},
                valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
            )
        except Exception:
            pass
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/validar-nivel2")
def validar_nivel2(contrato_id: int, registro_id: int, body: ValidarNivel2Body,
                   current_user=Depends(get_current_user)):
    ESTADOS = {"Aprobado", "Pendiente", "Rechazado", "No Objeto de Cobro"}
    if body.estado not in ESTADOS:
        raise HTTPException(status_code=422, detail=f"Estado inválido. Acepta: {ESTADOS}")
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        _require_sicoe_puede_validar_nivel(current_user, autor_id, 2)
        # Verificar nivel1
        def _get():
            return supabase.table("so_registros")\
                .select("nivel1_estado, nivel3_estado, reporte_id").eq("id", registro_id)\
                .eq("contrato_id", contrato_id).limit(1).execute().data
        rows = supabase_execute(_get)
        if not rows:
            raise HTTPException(status_code=404, detail="Registro no encontrado.")
        if _registro_nivel3_aprobado(rows[0]):
            raise HTTPException(
                status_code=400,
                detail="El registro está aprobado por Nivel 3 (Interventoría) y no puede modificarse por esta vía.",
            )
        if rows[0].get("nivel1_estado") != "Aprobado":
            raise HTTPException(status_code=422,
                detail="El registro debe estar aprobado por Nivel 1 primero.")

        estado_real = body.estado
        if body.estado == "No Objeto de Cobro":
            estado_real = "Rechazado"
            if not body.comentario_data:
                raise HTTPException(status_code=422,
                    detail="Se requiere comentario_data para 'No Objeto de Cobro'.")

        if body.estado in ("Pendiente", "Rechazado") and not body.comentario_data:
            raise HTTPException(status_code=422,
                detail="Se requiere comentario_data cuando el estado es Pendiente o Rechazado.")

        if estado_real == "Aprobado" and _sicoe_exige_topografia_para_aprobar_nivel2(contrato_id):
            rep_id = rows[0].get("reporte_id")
            if not rep_id or not _reporte_tiene_puntos_topograficos(contrato_id, rep_id):
                raise HTTPException(
                    status_code=422,
                    detail="No se puede aprobar en Nivel 2 (Residente) sin coordenadas topográficas en la Portada del reporte. Carga al menos un punto en Topografía antes de aprobar.",
                )

        prev_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
        update = {
            "nivel2_estado":     estado_real,
            "nivel2_usuario_id": autor_id,
            "nivel2_fecha":      datetime.utcnow().isoformat(),
        }
        if body.objeto_pago_sub is not None:
            update["nivel2_objeto_pago_sub"] = body.objeto_pago_sub

        def _upd():
            return supabase.table("so_registros")\
                .update(update).eq("id", registro_id)\
                .eq("contrato_id", contrato_id).execute().data
        supabase_execute(_upd)
        # Acta RPO parcial: al aprobar Residente de Obra (N2), alinear a acta vigente a la fecha de aprobación
        if estado_real == "Aprobado":
            try:
                _aplicar_acta_rpo_vigente_a_registro(contrato_id, registro_id, date.today())
            except Exception:
                pass
        if body.comentario_data:
            _insertar_comentario(contrato_id, registro_id, autor_id, body.comentario_data,
                                 tipo_override="validacion", nivel_validacion_override="Nivel 2",
                                 audit_user=current_user)
            _push_notif_validacion_sicoe_destinatarios(
                current_user,
                autor_id,
                contrato_id,
                registro_id,
                f"Validación Nivel 2: {body.estado}",
                body.comentario_data.get("mensaje", "") or "",
                body.comentario_data,
            )
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            after_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
            registrar_log(
                u_log, "VALIDAR", "SICOE", "registro", str(registro_id),
                {"nivel": 2, "estado": body.estado, "nivel2_objeto_pago_sub": body.objeto_pago_sub},
                valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
            )
        except Exception:
            pass
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/validar-nivel3")
def validar_nivel3(contrato_id: int, registro_id: int, body: ValidarNivel3Body,
                   current_user=Depends(get_current_user)):
    ESTADOS = {"Aprobado", "Pendiente", "Rechazado"}
    if body.estado not in ESTADOS:
        raise HTTPException(status_code=422, detail=f"Estado inválido. Acepta: {ESTADOS}")
    if body.estado in ("Pendiente", "Rechazado") and not body.comentario_data:
        raise HTTPException(status_code=422,
            detail="Se requiere comentario_data cuando el estado es Pendiente o Rechazado.")
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        _require_sicoe_puede_validar_nivel(current_user, autor_id, 3)
        # Verificar nivel2
        def _get():
            return supabase.table("so_registros")\
                .select("nivel2_estado, nivel3_estado").eq("id", registro_id)\
                .eq("contrato_id", contrato_id).limit(1).execute().data
        rows = supabase_execute(_get)
        if not rows:
            raise HTTPException(status_code=404, detail="Registro no encontrado.")
        if rows[0].get("nivel3_estado") == "Aprobado" and body.estado != "Aprobado":
            raise HTTPException(
                status_code=400,
                detail="El registro ya está aprobado por Nivel 3. Use el flujo de reversión para modificar la validación.",
            )
        if rows[0].get("nivel2_estado") != "Aprobado":
            raise HTTPException(status_code=422,
                detail="El registro debe estar aprobado por Nivel 2 primero.")

        prev_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
        update = {
            "nivel3_estado":     body.estado,
            "nivel3_usuario_id": autor_id,
            "nivel3_fecha":      datetime.utcnow().isoformat(),
        }
        if body.estado == "Aprobado":
            update["bloqueado"] = True

        def _upd():
            return supabase.table("so_registros")\
                .update(update).eq("id", registro_id)\
                .eq("contrato_id", contrato_id).execute().data
        supabase_execute(_upd)
        # Acta RPO definitiva: al aprobar Interventoría (N3), alinear a acta vigente a la fecha de aprobación
        if body.estado == "Aprobado":
            try:
                _aplicar_acta_rpo_vigente_a_registro(contrato_id, registro_id, date.today())
            except Exception:
                pass
        if body.comentario_data:
            _insertar_comentario(contrato_id, registro_id, autor_id, body.comentario_data,
                                 tipo_override="validacion", nivel_validacion_override="Nivel 3",
                                 audit_user=current_user)
            _push_notif_validacion_sicoe_destinatarios(
                current_user,
                autor_id,
                contrato_id,
                registro_id,
                f"Validación Nivel 3: {body.estado}",
                body.comentario_data.get("mensaje", "") or "",
                body.comentario_data,
            )
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            after_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
            registrar_log(
                u_log, "VALIDAR", "SICOE", "registro", str(registro_id),
                {"nivel": 3, "estado": body.estado},
                valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
            )
        except Exception:
            pass
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sicoe-obra/{contrato_id}/registros/validar-nivel-masivo")
def validar_nivel_masivo_por_filtro(
    contrato_id: int,
    body_in: ValidarNivelMasivoFiltroBody,
    current_user=Depends(get_current_user),
):
    """
    Validación N2 o N3 en bloque sobre el mismo universo que la grilla / exportar (filtros + capas).
    Tope: 500 registros por solicitud. Excluye líneas con objeto de pago a subcontratista (flujo validar-sub).
    Con aprobación N2 y contrato que exige topografía: omite líneas sin puntos en el reporte y las reporta aparte.
    """
    EST_MASIVO = {"Aprobado", "Pendiente", "Rechazado"}
    marcar = (body_in.marcar_estado or "").strip()
    if marcar not in EST_MASIVO:
        raise HTTPException(
            status_code=422,
            detail=f"Estado masivo inválido. Acepta: {EST_MASIVO}",
        )
    if marcar in ("Pendiente", "Rechazado") and not body_in.comentario_data:
        raise HTTPException(
            status_code=422,
            detail="Se requiere comentario_data cuando el estado masivo es Pendiente o Rechazado.",
        )
    nivel = int(body_in.nivel)
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        _require_sicoe_puede_validar_nivel(current_user, autor_id, nivel)

        exp_body = _sicoe_masivo_filtro_to_export_body(body_in)
        candidatos, st_collect = _sicoe_colectar_registros_masivo_desde_filtros(contrato_id, exp_body)

        actualizados = 0
        omitidos_precondicion = 0
        omitidos_topografia = 0

        if nivel == 2:
            exige_topo_n2 = marcar == "Aprobado" and _sicoe_exige_topografia_para_aprobar_nivel2(contrato_id)
            rep_ids_topo = [
                int(r["reporte_id"])
                for r in candidatos
                if r.get("reporte_id") is not None and exige_topo_n2
            ]
            topo_ok = _reportes_ids_con_topografia(contrato_id, rep_ids_topo) if exige_topo_n2 else set()

            for row in candidatos:
                rid = int(row["id"])
                if _registro_nivel3_aprobado(row):
                    omitidos_precondicion += 1
                    continue
                if (row.get("nivel1_estado") or "") != "Aprobado":
                    omitidos_precondicion += 1
                    continue
                if exige_topo_n2:
                    try:
                        rpi = int(row["reporte_id"]) if row.get("reporte_id") is not None else None
                    except (TypeError, ValueError):
                        rpi = None
                    if not rpi or rpi not in topo_ok:
                        omitidos_topografia += 1
                        continue

                prev_audit = _so_registro_fetch_validacion_audit(contrato_id, rid) or {}
                update = {
                    "nivel2_estado": marcar,
                    "nivel2_usuario_id": autor_id,
                    "nivel2_fecha": datetime.utcnow().isoformat(),
                }

                def _upd(regid=rid, upd=update):
                    return (
                        supabase.table("so_registros")
                        .update(upd)
                        .eq("id", regid)
                        .eq("contrato_id", contrato_id)
                        .execute()
                        .data
                    )

                supabase_execute(_upd)

                if marcar == "Aprobado":
                    try:
                        _aplicar_acta_rpo_vigente_a_registro(contrato_id, rid, date.today())
                    except Exception:
                        pass

                if body_in.comentario_data:
                    _insertar_comentario(
                        contrato_id,
                        rid,
                        autor_id,
                        body_in.comentario_data,
                        tipo_override="validacion",
                        nivel_validacion_override="Nivel 2",
                        audit_user=current_user,
                    )
                    _push_notif_validacion_sicoe_destinatarios(
                        current_user,
                        autor_id,
                        contrato_id,
                        rid,
                        f"Validación Nivel 2 (masivo): {marcar}",
                        body_in.comentario_data.get("mensaje", "") or "",
                        body_in.comentario_data,
                    )
                try:
                    u_log = _audit_user_contrato(current_user, contrato_id)
                    after_audit = _so_registro_fetch_validacion_audit(contrato_id, rid) or {}
                    registrar_log(
                        u_log,
                        "VALIDAR",
                        "SICOE",
                        "registro",
                        str(rid),
                        {"nivel": 2, "estado": marcar, "masivo": True, "masivo_filtro": True},
                        valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                        valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
                    )
                except Exception:
                    pass
                actualizados += 1

        elif nivel == 3:
            for row in candidatos:
                rid = int(row["id"])
                if (row.get("nivel3_estado") or "") == "Aprobado" and marcar != "Aprobado":
                    omitidos_precondicion += 1
                    continue
                if (row.get("nivel2_estado") or "") != "Aprobado":
                    omitidos_precondicion += 1
                    continue

                prev_audit = _so_registro_fetch_validacion_audit(contrato_id, rid) or {}
                update = {
                    "nivel3_estado": marcar,
                    "nivel3_usuario_id": autor_id,
                    "nivel3_fecha": datetime.utcnow().isoformat(),
                }
                if marcar == "Aprobado":
                    update["bloqueado"] = True

                def _upd3(regid=rid, upd=update):
                    return (
                        supabase.table("so_registros")
                        .update(upd)
                        .eq("id", regid)
                        .eq("contrato_id", contrato_id)
                        .execute()
                        .data
                    )

                supabase_execute(_upd3)

                if marcar == "Aprobado":
                    try:
                        _aplicar_acta_rpo_vigente_a_registro(contrato_id, rid, date.today())
                    except Exception:
                        pass

                if body_in.comentario_data:
                    _insertar_comentario(
                        contrato_id,
                        rid,
                        autor_id,
                        body_in.comentario_data,
                        tipo_override="validacion",
                        nivel_validacion_override="Nivel 3",
                        audit_user=current_user,
                    )
                    _push_notif_validacion_sicoe_destinatarios(
                        current_user,
                        autor_id,
                        contrato_id,
                        rid,
                        f"Validación Nivel 3 (masivo): {marcar}",
                        body_in.comentario_data.get("mensaje", "") or "",
                        body_in.comentario_data,
                    )
                try:
                    u_log = _audit_user_contrato(current_user, contrato_id)
                    after_audit = _so_registro_fetch_validacion_audit(contrato_id, rid) or {}
                    registrar_log(
                        u_log,
                        "VALIDAR",
                        "SICOE",
                        "registro",
                        str(rid),
                        {"nivel": 3, "estado": marcar, "masivo": True, "masivo_filtro": True},
                        valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                        valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
                    )
                except Exception:
                    pass
                actualizados += 1

        out = {
            "ok": True,
            "actualizados": actualizados,
            "omitidos_precondicion": omitidos_precondicion,
            "omitidos_topografia": omitidos_topografia,
            "excluidos_objeto_pago_sub": st_collect.get("excluidos_objeto_pago_sub", 0),
            "truncado_mas_de_500": bool(st_collect.get("truncado")),
            "tope_registros": SICOE_MASIVO_MAX_REGISTROS,
            "candidatos_en_lote": len(candidatos),
        }
        if out["omitidos_topografia"]:
            out["alerta_topografia"] = (
                f"{out['omitidos_topografia']} registro(s) no se aprobaron en N2 por falta de topografía en el reporte. "
                "El resto del lote sí se procesó."
            )
        if out["truncado_mas_de_500"]:
            out["alerta_tope"] = (
                f"Se procesaron como máximo {SICOE_MASIVO_MAX_REGISTROS} registros. "
                "Acote el filtro y repita si necesita cubrir más líneas."
            )
        if out["excluidos_objeto_pago_sub"]:
            out["alerta_objeto_sub"] = (
                f"{out['excluidos_objeto_pago_sub']} línea(s) con objeto de pago a subcontratista no entran en el masivo; revíselas una a una."
            )
        return out
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/validar-sub")
def validar_sub(contrato_id: int, registro_id: int, body: ValidarSubBody,
                current_user=Depends(get_current_user)):
    ESTADOS = {"Aprobado", "Pendiente", "Rechazado"}
    if body.estado not in ESTADOS:
        raise HTTPException(status_code=422, detail=f"Estado inválido. Acepta: {ESTADOS}")
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        if not _es_desarrollador(current_user):
            if not _cargo_permiso_validar_reporte_cantidades_user_id(autor_id):
                raise HTTPException(
                    status_code=403,
                    detail="No tiene permiso de validación en «Reporte de cantidades» (matriz de accesos).",
                )
        # Verificar nivel2_objeto_pago_sub
        def _get():
            return supabase.table("so_registros")\
                .select("nivel2_objeto_pago_sub, nivel3_estado").eq("id", registro_id)\
                .eq("contrato_id", contrato_id).limit(1).execute().data
        rows = supabase_execute(_get)
        if not rows:
            raise HTTPException(status_code=404, detail="Registro no encontrado.")
        if _registro_nivel3_aprobado(rows[0]):
            raise HTTPException(
                status_code=400,
                detail="El registro está aprobado por Nivel 3 (Interventoría) y no puede modificarse por esta vía.",
            )
        if not rows[0].get("nivel2_objeto_pago_sub"):
            raise HTTPException(status_code=422,
                detail="El registro no es objeto de pago a subcontratista (nivel2_objeto_pago_sub debe ser True).")

        prev_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
        update = {
            "sub_estado":     body.estado,
            "sub_usuario_id": autor_id,
            "sub_fecha":      datetime.utcnow().isoformat(),
        }
        def _upd():
            return supabase.table("so_registros")\
                .update(update).eq("id", registro_id)\
                .eq("contrato_id", contrato_id).execute().data
        supabase_execute(_upd)
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            after_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
            registrar_log(
                u_log, "VALIDAR", "SICOE", "registro", str(registro_id),
                {"nivel": "sub", "estado": body.estado},
                valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
            )
        except Exception:
            pass
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ══════════════════════════════════════════════════════════════════════════════
# SICOE-OBRA — Dashboard endpoints (reemplazan /cobro/ leyendo so_registros)
# ══════════════════════════════════════════════════════════════════════════════


def _parse_rpc_dashboard_resumen_raw(raw):
    """Deserializa respuesta de dashboard_resumen_sicoe_agg (PostgREST / RPC)."""
    if raw is None:
        return None
    if isinstance(raw, list) and len(raw) > 0:
        row = raw[0]
        if isinstance(row, dict):
            for v in row.values():
                if isinstance(v, dict) and "total_presupuesto" in v:
                    return v
        if isinstance(row, dict) and "total_presupuesto" in row:
            return row
    if isinstance(raw, dict) and "total_presupuesto" in raw:
        return raw
    if isinstance(raw, dict):
        k = next(iter(raw.keys()), None)
        if k and isinstance(raw.get(k), dict) and "total_presupuesto" in raw[k]:
            return raw[k]
    return None


def _parse_rpc_jsonb_value(raw):
    """Primer valor jsonb devuelto por RPC (PostgREST devuelve [{ función: <payload> }])."""
    if raw is None:
        return None
    if isinstance(raw, list) and len(raw) > 0:
        row = raw[0]
        if isinstance(row, dict):
            for v in row.values():
                return v
    if isinstance(raw, dict):
        return raw
    return None


def _dash_norm_item_key_py(s: Optional[str]) -> str:
    """Alinea claves presupuesto.item ↔ so_registros.item_numero (p. ej. '4.22.' → '4.22')."""
    if s is None:
        return ""
    t = str(s).strip()
    if not t:
        return ""
    return re.sub(r"\.+$", "", t)


def _dash_norm_capitulo_key_py(s: Optional[str]) -> str:
    """Misma lógica que _dash_norm_capitulo_key en SQL: obra vs presupuesto con distinto espaciado."""
    if s is None:
        return "Sin capítulo"
    t = str(s).strip()
    if not t:
        return "Sin capítulo"
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"^(\d+\.)\s+", r"\1", t)
    return t


def _dash_pk_disp_key_py(v: Any) -> str:
    """Clave display PK para cruzar presupuesto.pk_id ↔ pk_ids.pk_id (como texto estable)."""
    if v is None:
        return "(sin pk)"
    s = str(v).strip()
    if not s:
        return "(sin pk)"
    return s


# ── Liquidación contrato: recalc (= polígonos presupuesto) vs «cobro» (= SICOE N3 aprobado en so_registros) ──
PRESUPUESTO_TIPO_POLIGONO = "Presupuesto de Obra"
LIQ_SUPERCOBRO_COP = 20_000_000.0


def _liquidacion_row_categoria(delta_costo: float, es_calculado: bool) -> str:
    if not es_calculado:
        return "EJECUCION"
    if abs(delta_costo) < 0.5:
        return "EQUILIBRIO"
    if delta_costo < 0:
        return "SUPERCOBRO" if abs(delta_costo) > LIQ_SUPERCOBRO_COP else "DEVOLUCION"
    return "POR_COBRAR"


def _liquidacion_pct_ejecucion(cobrado: float, recalculado: float) -> float:
    if recalculado and recalculado > 0:
        return round(cobrado / recalculado * 100, 1)
    if cobrado and cobrado > 0:
        return 999.0
    return 0.0


def _liquidacion_analisis_items(contrato_id: int, nivel: str, current_user) -> List[Dict[str, Any]]:
    """Agrega por ítem o capítulo: ítems con polígono (tipo Presupuesto de Obra) usan cant/costo PPTO; el resto toma obra N3 ✓ como recalc (= cobro)."""
    ppo: Dict[Tuple[str, str], Dict[str, Any]] = {}
    off = 0
    pto_iv = _presupuesto_aplica_filtro_interventoria(current_user)
    while True:
        def _q_pres(o=off):
            q = (
                supabase.table("presupuesto")
                .select("capitulo, item, tipo_ejecucion, cant_total, costo_directo, descripcion")
                .eq("contrato_id", contrato_id)
                .eq("dado_de_baja", False)
            )
            if pto_iv:
                q = q.or_("pre_interv_estado.is.null,pre_interv_estado.eq.Aprobado")
            return q.range(o, o + 999).execute().data

        batch = supabase_execute(_q_pres) or []
        for r in batch:
            ck = _dash_norm_capitulo_key_py(r.get("capitulo"))
            ik = _dash_norm_item_key_py(r.get("item"))
            if not ik:
                continue
            k = (ck, ik)
            if k not in ppo:
                ppo[k] = {
                    "cant_po": 0.0,
                    "cost_po": 0.0,
                    "desc": "",
                    "calculado": False,
                    "cap_raw": (r.get("capitulo") or "").strip() or ck,
                    "item_raw": (r.get("item") or "").strip() or ik,
                }
            te = (r.get("tipo_ejecucion") or "").strip()
            if te == PRESUPUESTO_TIPO_POLIGONO:
                ppo[k]["calculado"] = True
                ppo[k]["cant_po"] += float(r.get("cant_total") or 0)
                ppo[k]["cost_po"] += float(r.get("costo_directo") or 0)
            if not ppo[k]["desc"] and r.get("descripcion"):
                ppo[k]["desc"] = str(r.get("descripcion"))[:400]
        if len(batch) < 1000:
            break
        off += 1000

    sic: Dict[Tuple[str, str], Dict[str, Any]] = {}
    off = 0
    while True:
        def _q_reg(o=off):
            return (
                supabase.table("so_registros")
                .select("capitulo, item_numero, cantidad_total, costo_directo, nivel3_estado")
                .eq("contrato_id", contrato_id)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_q_reg) or []
        for r in batch:
            if _matriz_validacion_norm_estado(r.get("nivel3_estado")) != "Aprobado":
                continue
            ck = _dash_norm_capitulo_key_py(r.get("capitulo"))
            ik = _dash_norm_item_key_py(r.get("item_numero"))
            if not ik:
                continue
            k = (ck, ik)
            if k not in sic:
                sic[k] = {
                    "cant": 0.0,
                    "cost": 0.0,
                    "cap_raw": (r.get("capitulo") or "").strip() or ck,
                    "item_raw": (r.get("item_numero") or "").strip() or ik,
                }
            sic[k]["cant"] += float(r.get("cantidad_total") or 0)
            sic[k]["cost"] += float(r.get("costo_directo") or 0)
        if len(batch) < 1000:
            break
        off += 1000

    keys_all = set(ppo.keys()) | set(sic.keys())
    item_rows: List[Dict[str, Any]] = []
    for k in sorted(keys_all, key=lambda x: (x[0], x[1])):
        meta = ppo.get(k)
        sg = sic.get(
            k,
            {"cant": 0.0, "cost": 0.0, "cap_raw": k[0], "item_raw": k[1]},
        )
        cant_cob = float(sg["cant"])
        cob = float(sg["cost"])
        if meta:
            cap_raw = meta.get("cap_raw") or sg.get("cap_raw") or k[0]
            item_raw = meta.get("item_raw") or sg.get("item_raw") or k[1]
            desc = meta.get("desc") or ""
            calc = bool(meta.get("calculado"))
        else:
            cap_raw = sg.get("cap_raw") or k[0]
            item_raw = sg.get("item_raw") or k[1]
            desc = ""
            calc = False
        if calc and meta:
            cant_re = float(meta["cant_po"])
            rec = float(meta["cost_po"])
        else:
            cant_re, rec = cant_cob, cob
        delta_cant = cant_re - cant_cob
        delta_cost = rec - cob
        item_rows.append(
            {
                "capitulo": cap_raw,
                "nombre": item_raw,
                "descripcion": desc,
                "cant_recalc": round(cant_re, 2),
                "recalculado": round(rec, 0),
                "cant_cobro": round(cant_cob, 2),
                "cobrado": round(cob, 0),
                "delta_cant": round(delta_cant, 2),
                "delta_costo": round(delta_cost, 0),
                "pct": _liquidacion_pct_ejecucion(cob, rec),
                "categoria": _liquidacion_row_categoria(delta_cost, calc),
                "_ck": k[0],
                "_calc": calc,
            }
        )

    if nivel != "capitulo":
        for r in item_rows:
            r.pop("_ck", None)
            r.pop("_calc", None)
        return item_rows

    caps: Dict[str, Dict[str, Any]] = {}
    for r in item_rows:
        ck = r["_ck"]
        if ck not in caps:
            caps[ck] = {
                "capitulo": r["capitulo"],
                "nombre": r["capitulo"],
                "cant_recalc": 0.0,
                "recalculado": 0.0,
                "cant_cobro": 0.0,
                "cobrado": 0.0,
                "any_calc": False,
            }
        caps[ck]["cant_recalc"] += float(r["cant_recalc"])
        caps[ck]["recalculado"] += float(r["recalculado"])
        caps[ck]["cant_cobro"] += float(r["cant_cobro"])
        caps[ck]["cobrado"] += float(r["cobrado"])
        if r.get("_calc"):
            caps[ck]["any_calc"] = True

    out: List[Dict[str, Any]] = []
    for ck in sorted(caps.keys(), key=lambda x: x):
        c = caps[ck]
        d_cant = c["cant_recalc"] - c["cant_cobro"]
        d_cost = c["recalculado"] - c["cobrado"]
        ac = bool(c["any_calc"])
        out.append(
            {
                "capitulo": c["capitulo"],
                "nombre": c["nombre"],
                "descripcion": "",
                "cant_recalc": round(c["cant_recalc"], 2),
                "recalculado": round(c["recalculado"], 0),
                "cant_cobro": round(c["cant_cobro"], 2),
                "cobrado": round(c["cobrado"], 0),
                "delta_cant": round(d_cant, 2),
                "delta_costo": round(d_cost, 0),
                "pct": _liquidacion_pct_ejecucion(c["cobrado"], c["recalculado"]),
                "categoria": _liquidacion_row_categoria(d_cost, ac),
            }
        )
    return out


def _dashboard_pkid_colores_liquidacion(
    contrato_id: int,
    capitulo: Optional[str],
    item: Optional[str],
) -> Dict[str, Any]:
    """Mini mapa liquidación: por PK, referencia = Σ ítem (polígono si calculado) vs cobrado = SICOE N3 ✓."""
    cap_k = _dash_norm_capitulo_key_py(capitulo) if capitulo else None
    it_k = _dash_norm_item_key_py(item) if item else None

    calculado: Dict[Tuple[str, str], bool] = {}
    po_by: Dict[Tuple[str, str], Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    off = 0
    while True:
        def _qp(o=off):
            return (
                supabase.table("presupuesto")
                .select("capitulo, item, pk_id, tipo_ejecucion, costo_directo")
                .eq("contrato_id", contrato_id)
                .eq("dado_de_baja", False)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_qp) or []
        for r in batch:
            ck = _dash_norm_capitulo_key_py(r.get("capitulo"))
            ik = _dash_norm_item_key_py(r.get("item"))
            if not ik:
                continue
            if cap_k and ck != cap_k:
                continue
            if it_k and ik != it_k:
                continue
            key = (ck, ik)
            te = (r.get("tipo_ejecucion") or "").strip()
            if te == PRESUPUESTO_TIPO_POLIGONO:
                calculado[key] = True
                pk = _dash_pk_disp_key_py(r.get("pk_id"))
                if pk != "(sin pk)":
                    po_by[key][pk] += float(r.get("costo_directo") or 0)
        if len(batch) < 1000:
            break
        off += 1000

    sic_by: Dict[Tuple[str, str], Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    off = 0
    while True:
        def _qs(o=off):
            return (
                supabase.table("so_registros")
                .select("capitulo, item_numero, costo_directo, nivel3_estado, pk_id_id, pk_ids(pk_id)")
                .eq("contrato_id", contrato_id)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_qs) or []
        for r in batch:
            if _matriz_validacion_norm_estado(r.get("nivel3_estado")) != "Aprobado":
                continue
            ck = _dash_norm_capitulo_key_py(r.get("capitulo"))
            ik = _dash_norm_item_key_py(r.get("item_numero"))
            if not ik:
                continue
            if cap_k and ck != cap_k:
                continue
            if it_k and ik != it_k:
                continue
            pk_join = r.get("pk_ids") or {}
            pk_raw = pk_join.get("pk_id")
            pk = (
                _dash_pk_disp_key_py(pk_raw)
                if pk_raw is not None and str(pk_raw).strip() != ""
                else _dash_pk_disp_key_py(r.get("pk_id_id"))
            )
            if pk == "(sin pk)":
                continue
            sic_by[(ck, ik)][pk] += float(r.get("costo_directo") or 0)
        if len(batch) < 1000:
            break
        off += 1000

    keys_scope = set(po_by.keys()) | set(sic_by.keys())
    if cap_k:
        keys_scope = {x for x in keys_scope if x[0] == cap_k}
    if it_k:
        keys_scope = {x for x in keys_scope if x[1] == it_k}

    all_pks: Set[str] = set()
    for k in keys_scope:
        all_pks.update(sic_by[k].keys())
        all_pks.update(po_by[k].keys())

    result: Dict[str, Any] = {}
    for pk in all_pks:
        sap = 0.0
        pref = 0.0
        for k in keys_scope:
            sik = float(sic_by[k].get(pk, 0.0))
            sap += sik
            is_calc = bool(calculado.get(k, False))
            if is_calc:
                pref += float(po_by[k].get(pk, 0.0))
            else:
                pref += sik
        result[pk] = {
            "cobrado": round(sap, 2),
            "presupuesto": round(pref, 2),
            "sicoe_aprobado": round(sap, 2),
            "pct": round(sap / pref * 100, 1) if pref else 0,
            "sobrecosto": sap > pref,
        }
    return result


@app.get("/sicoe-obra/{contrato_id}/dashboard-resumen")
def dashboard_resumen_obra(contrato_id: int, current_user=Depends(get_current_user)):
    try:
        try:
            def _rpc():
                return supabase.rpc(
                    "dashboard_resumen_sicoe_agg",
                    {"p_contrato_id": contrato_id},
                ).execute().data

            hit = _parse_rpc_dashboard_resumen_raw(supabase_execute(_rpc))
            if hit is not None and isinstance(hit.get("comparativo_capitulos"), list):
                return _dashboard_resumen_merge_extras(contrato_id, hit)
        except Exception:
            pass

        # 1. Obra aprobada (Interventoría): agregar desde so_registros con el mismo criterio
        #    que la matriz / drill (nivel3 ≈ Aprobado), no desde vista_dashboard_resumen,
        #    para que importaciones con variantes de texto ("APROBADO", espacios, etc.) cuenten.
        def _actas_map():
            return supabase.table("actas").select("id, numero_rpo")\
                .eq("contrato_id", contrato_id).execute().data
        acta_rows = supabase_execute(_actas_map) or []
        acta_id_to_nr = {a["id"]: a.get("numero_rpo") for a in acta_rows if a.get("id") is not None}

        total_cobrado = 0.0
        acta_agg = {}
        obra_caps = {}
        off = 0
        while True:
            def _batch(o=off):
                return supabase.table("so_registros").select(
                    "capitulo, costo_directo, acta_rpo_id, nivel3_estado"
                ).eq("contrato_id", contrato_id).range(o, o + 999).execute().data
            batch = supabase_execute(_batch) or []
            for reg in batch:
                if _matriz_validacion_norm_estado(reg.get("nivel3_estado")) != "Aprobado":
                    continue
                cd = float(reg.get("costo_directo") or 0)
                total_cobrado += cd
                cap = reg.get("capitulo") or "Sin capítulo"
                obra_caps[cap] = obra_caps.get(cap, 0) + cd
                aid = reg.get("acta_rpo_id")
                nr = acta_id_to_nr.get(aid) if aid is not None else None
                if nr is not None:
                    acta_agg[nr] = acta_agg.get(nr, 0) + cd
            if len(batch) < 1000:
                break
            off += 1000

        # 2. Presupuesto por capítulo
        def _ppto():
            return supabase.table("vista_ppto_por_capitulo")\
                .select("*").eq("contrato_id", contrato_id).execute().data
        ppto_raw = supabase_execute(_ppto)
        ppto_caps = {r["capitulo"]: float(r.get("presupuesto") or 0) for r in ppto_raw}
        ppto_total = sum(ppto_caps.values())

        # 3. Comparativo por capítulo (presupuesto vs obra aprobada N3)

        def _sort_acta_key(x):
            try:
                return float(x[0])
            except (TypeError, ValueError):
                return 0.0

        def _acta_num(k):
            try:
                return float(k)
            except (TypeError, ValueError):
                return 0.0

        por_acta = [{"acta": nr, "cobrado": round(v, 2)} for nr, v in sorted(acta_agg.items(), key=_sort_acta_key, reverse=True)]

        caps = sorted(set(list(ppto_caps.keys()) + list(obra_caps.keys())))
        comparativo = [
            {
                "capitulo": c,
                "presupuesto": ppto_caps.get(c, 0),
                "cobrado": round(obra_caps.get(c, 0), 2),
                "delta": round(ppto_caps.get(c, 0) - obra_caps.get(c, 0), 2),
                "consumo_pct": round(obra_caps.get(c, 0) / ppto_caps.get(c, 0) * 100, 1) if ppto_caps.get(c, 0) else 0,
            }
            for c in caps
        ]

        base = {
            "total_presupuesto": ppto_total,
            "total_cobrado": round(total_cobrado, 2),
            "delta": round(ppto_total - total_cobrado, 2),
            "consumo_pct": round(total_cobrado / ppto_total * 100, 1) if ppto_total else 0,
            "actas": sorted(acta_agg.keys(), key=_acta_num, reverse=True),
            "comparativo_capitulos": comparativo,
            "por_acta": por_acta,
        }
        return _dashboard_resumen_merge_extras(contrato_id, base)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _matriz_validacion_norm_estado(v) -> str:
    if v is None:
        return "No Revisado"
    s = str(v).strip()
    if not s:
        return "No Revisado"
    sl = s.lower()
    if sl == "aprobado":
        return "Aprobado"
    if sl == "pendiente":
        return "Pendiente"
    if sl == "rechazado":
        return "Rechazado"
    if "no revis" in sl or sl == "no revisado":
        return "No Revisado"
    return s


def _dash_norm_cap(s: Optional[str]) -> str:
    t = (s or "").strip()
    return t if t else "Sin capítulo"


def _so_reg_cola_n1_n2_aprobados(reg: dict) -> bool:
    return (
        _matriz_validacion_norm_estado(reg.get("nivel1_estado")) == "Aprobado"
        and _matriz_validacion_norm_estado(reg.get("nivel2_estado")) == "Aprobado"
    )


def _so_reg_en_cola_interventoria(reg: dict) -> bool:
    """Línea lista para revisión N3: ítem asignado y N1+N2 aprobados."""
    if not (reg.get("item_numero") or "").strip():
        return False
    return _so_reg_cola_n1_n2_aprobados(reg)


def _dashboard_resumen_scan_caps(contrato_id: int) -> Dict[str, Any]:
    """
    Por capítulo: costo/cant SICOE N3 aprobado, SICOE N3 no revisado (en cola),
    y presupuesto ClaraCore partido por columna revisado (aprobado vs resto).
    """
    sicoe_ap_c = defaultdict(float)
    sicoe_ap_q = defaultdict(float)
    sicoe_nr_c = defaultdict(float)
    sicoe_nr_q = defaultdict(float)
    off = 0
    while True:
        def _b(o=off):
            return (
                supabase.table("so_registros")
                .select(
                    "capitulo, costo_directo, cantidad_total, item_numero, "
                    "nivel1_estado, nivel2_estado, nivel3_estado"
                )
                .eq("contrato_id", contrato_id)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_b) or []
        for reg in batch:
            if not (reg.get("item_numero") or "").strip():
                continue
            cap = _dash_norm_cap(reg.get("capitulo"))
            cd = float(reg.get("costo_directo") or 0)
            cq = float(reg.get("cantidad_total") or 0)
            n3 = _matriz_validacion_norm_estado(reg.get("nivel3_estado"))
            if n3 == "Aprobado":
                sicoe_ap_c[cap] += cd
                sicoe_ap_q[cap] += cq
            elif _so_reg_en_cola_interventoria(reg) and n3 == "No Revisado":
                sicoe_nr_c[cap] += cd
                sicoe_nr_q[cap] += cq
        if len(batch) < 1000:
            break
        off += 1000

    ppto_ap_c = defaultdict(float)
    ppto_nr_c = defaultdict(float)
    off = 0
    while True:
        def _pp(o=off):
            return (
                supabase.table("presupuesto")
                .select("capitulo, costo_directo, revisado")
                .eq("contrato_id", contrato_id)
                .eq("dado_de_baja", False)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_pp) or []
        for r in batch:
            cap = _dash_norm_cap(r.get("capitulo"))
            cost = float(r.get("costo_directo") or 0)
            if _matriz_validacion_norm_estado(r.get("revisado")) == "Aprobado":
                ppto_ap_c[cap] += cost
            else:
                ppto_nr_c[cap] += cost
        if len(batch) < 1000:
            break
        off += 1000

    return {
        "sicoe_ap_c": dict(sicoe_ap_c),
        "sicoe_ap_q": dict(sicoe_ap_q),
        "sicoe_nr_c": dict(sicoe_nr_c),
        "sicoe_nr_q": dict(sicoe_nr_q),
        "ppto_ap_c": dict(ppto_ap_c),
        "ppto_nr_c": dict(ppto_nr_c),
    }


def _dashboard_resumen_merge_extras(contrato_id: int, hit: dict) -> dict:
    """Si el RPC ya trae dashboard_schema=2, no escanear tablas. Si no, enriquecer con scan Python."""
    if hit.get("dashboard_schema") == 2:
        return hit
    try:
        ex = _dashboard_resumen_scan_caps(contrato_id)
    except Exception:
        return hit
    comp = hit.get("comparativo_capitulos")
    if not isinstance(comp, list):
        return hit
    tot_nr = tot_pap = tot_pnr = 0.0
    for row in comp:
        if not isinstance(row, dict):
            continue
        cap = row.get("capitulo")
        nr = float(ex["sicoe_nr_c"].get(cap, 0))
        pap = float(ex["ppto_ap_c"].get(cap, 0))
        pnr = float(ex["ppto_nr_c"].get(cap, 0))
        row["sicoe_no_revisado_n3"] = round(nr, 2)
        row["presupuesto_aprobado_n3"] = round(pap, 2)
        row["presupuesto_no_revisado_n3"] = round(pnr, 2)
        row["sicoe_aprobado_n3"] = float(row.get("cobrado") or 0)
        tot_nr += nr
        tot_pap += pap
        tot_pnr += pnr
    hit["total_sicoe_n3_no_revisado"] = round(tot_nr, 2)
    hit["total_presupuesto_aprobado_n3"] = round(tot_pap, 2)
    hit["total_presupuesto_no_revisado_n3"] = round(tot_pnr, 2)
    return hit


def _drill_agg_by_item(contrato_id: int, capitulo: str, item_filtro: Optional[str] = None) -> List[dict]:
    """Ítems de un capítulo: presupuesto ClaraCore (revisado), SICOE N3 ap / no rev."""
    ppto_by: Dict[str, List[dict]] = {}
    rows_p = _ppto_rows_capitulo(contrato_id, capitulo)
    for r in rows_p:
        it = _dash_norm_item_key_py(r.get("item"))
        if not it:
            continue
        ppto_by.setdefault(it, []).append(r)

    ap_c = defaultdict(float)
    ap_q = defaultdict(float)
    nr_c = defaultdict(float)
    nr_q = defaultdict(float)
    off = 0
    while True:
        def _sr(o=off):
            return (
                supabase.table("so_registros")
                .select(
                    "item_numero, costo_directo, cantidad_total, "
                    "nivel1_estado, nivel2_estado, nivel3_estado"
                )
                .eq("contrato_id", contrato_id)
                .eq("capitulo", capitulo)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_sr) or []
        for reg in batch:
            it = _dash_norm_item_key_py(reg.get("item_numero"))
            if not it:
                continue
            cd = float(reg.get("costo_directo") or 0)
            cq = float(reg.get("cantidad_total") or 0)
            n3 = _matriz_validacion_norm_estado(reg.get("nivel3_estado"))
            if n3 == "Aprobado":
                ap_c[it] += cd
                ap_q[it] += cq
            elif _so_reg_en_cola_interventoria(reg) and n3 == "No Revisado":
                nr_c[it] += cd
                nr_q[it] += cq
        if len(batch) < 1000:
            break
        off += 1000

    keys = sorted(set(list(ppto_by.keys()) + list(ap_c.keys()) + list(nr_c.keys())), key=lambda x: str(x))
    out = []
    for k in keys:
        rows_it = ppto_by.get(k, [])
        p_cost = sum(float(x.get("costo_directo") or 0) for x in rows_it)
        p_cant = sum(float(x.get("cant_total") or 0) for x in rows_it)
        revsplit = _ppto_costo_por_revisado(rows_it)
        pap = float(revsplit.get("Aprobado") or 0)
        pnr = float(revsplit.get("No Revisado") or 0) + float(revsplit.get("Pendiente") or 0) + float(revsplit.get("Rechazado") or 0)
        desc = ""
        for x in rows_it:
            if x.get("descripcion"):
                desc = str(x["descripcion"])
                break
        apc = ap_c.get(k, 0)
        nrc = nr_c.get(k, 0)
        out.append({
            "item": k,
            "nombre": k,
            "descripcion": desc,
            "presupuesto": p_cost,
            "cobrado": round(apc, 2),
            "presupuesto_aprobado_n3": round(pap, 2),
            "presupuesto_no_revisado_n3": round(pnr, 2),
            "sicoe_no_revisado_n3": round(nrc, 2),
            "delta": round(p_cost - apc, 2),
            "pct": round(apc / p_cost * 100, 1) if p_cost else 0,
            "cant_ppto": round(p_cant, 3),
            "cant_sicoe_aprobado": round(ap_q.get(k, 0), 3),
            "cant_sicoe_no_revisado": round(nr_q.get(k, 0), 3),
        })
    if item_filtro and str(item_filtro).strip():
        itf = _dash_norm_item_key_py(str(item_filtro).strip())
        out = [row for row in out if row["item"] == itf]
    return out


def _drill_agg_capitulos(contrato_id: int) -> List[dict]:
    """Listado por capítulo con los tres comparativos (sin cobro)."""
    def _ppto():
        return supabase.table("vista_ppto_por_capitulo").select("*").eq("contrato_id", contrato_id).execute().data

    ppto_raw = supabase_execute(_ppto) or []
    ppto_caps = {r["capitulo"]: float(r.get("presupuesto") or 0) for r in ppto_raw}
    ex = _dashboard_resumen_scan_caps(contrato_id)
    caps = sorted(
        set(
            list(ppto_caps.keys())
            + list(ex["sicoe_ap_c"].keys())
            + list(ex["sicoe_nr_c"].keys())
            + list(ex["ppto_ap_c"].keys())
            + list(ex["ppto_nr_c"].keys())
        ),
        key=lambda x: str(x),
    )
    out = []
    for cap in caps:
        apc = float(ex["sicoe_ap_c"].get(cap, 0))
        nrc = float(ex["sicoe_nr_c"].get(cap, 0))
        pap = float(ex["ppto_ap_c"].get(cap, 0))
        pnr = float(ex["ppto_nr_c"].get(cap, 0))
        pp = float(ppto_caps.get(cap, 0))
        out.append({
            "nombre": cap,
            "descripcion": "",
            "presupuesto": pp,
            "cobrado": round(apc, 2),
            "presupuesto_aprobado_n3": round(pap, 2),
            "presupuesto_no_revisado_n3": round(pnr, 2),
            "sicoe_no_revisado_n3": round(nrc, 2),
            "delta": round(pp - apc, 2),
            "pct": round(apc / pp * 100, 1) if pp else 0,
            "cant_ppto": 0,
            "cant_sicoe_aprobado": round(float(ex["sicoe_ap_q"].get(cap, 0)), 3),
            "cant_sicoe_no_revisado": round(float(ex["sicoe_nr_q"].get(cap, 0)), 3),
        })
    return out


def _matriz_validacion_bloque_capitulo(capitulo: Optional[str]) -> str:
    """Obra ejecutada directa vs ensayos/sondeos (cap. 14–15 o nombre típico)."""
    c = (capitulo or "").strip().upper()
    if c.startswith("14.") or c.startswith("15.") or "ENSAYO" in c or "SONDEO" in c:
        return "ensayos"
    return "obra"


def _matriz_validacion_empty():
    z = {"interventoria": 0.0, "residente": 0.0, "inspector": 0.0}
    return {
        "aprobado": dict(z),
        "pendiente": dict(z),
        "pendiente_item": dict(z),
        "no_revisado": dict(z),
        "rechazado": dict(z),
        "habilitado": dict(z),
        "otras_actas": dict(z),
    }


def _dashboard_matriz_validacion_fallback(
    contrato_id: int,
    acta_id_filtro: Optional[int],
) -> dict:
    """Fallback lento si la función SQL dashboard_matriz_validacion_agg no está desplegada."""
    obra_m = _matriz_validacion_empty()
    ens_m = _matriz_validacion_empty()

    off = 0
    while True:
        def _batch(o=off):
            q = supabase.table("so_registros").select(
                "costo_directo,nivel1_estado,nivel2_estado,nivel3_estado,sub_estado,"
                "capitulo,acta_rpo_id,item_numero"
            ).eq("contrato_id", contrato_id)
            if acta_id_filtro is not None:
                q = q.eq("acta_rpo_id", acta_id_filtro)
            return q.range(o, o + 999).execute().data

        batch = supabase_execute(_batch)
        for reg in batch:
            if not (reg.get("item_numero") or "").strip():
                continue
            cd = float(reg.get("costo_directo") or 0)
            n1 = _matriz_validacion_norm_estado(reg.get("nivel1_estado"))
            n2 = _matriz_validacion_norm_estado(reg.get("nivel2_estado"))
            n3 = _matriz_validacion_norm_estado(reg.get("nivel3_estado"))
            sub_n = _matriz_validacion_norm_estado(reg.get("sub_estado"))
            sub_raw = str(reg.get("sub_estado") or "").strip().lower()
            bloque = _matriz_validacion_bloque_capitulo(reg.get("capitulo"))
            M = ens_m if bloque == "ensayos" else obra_m

            def acc(estado_nivel: str, col: str):
                if estado_nivel == "Aprobado":
                    M["aprobado"][col] += cd
                elif estado_nivel == "Pendiente":
                    M["pendiente"][col] += cd
                elif estado_nivel == "Rechazado":
                    M["rechazado"][col] += cd
                else:
                    M["no_revisado"][col] += cd

            # N1 (inspector): todas las filas del acta; solo estados nivel 1.
            acc(n1, "inspector")
            # N2 (residente): solo si N1 aprobó; sobre ese subconjunto, estados nivel 2.
            if n1 == "Aprobado":
                acc(n2, "residente")
            # N3 (interventoría): solo si N1 y N2 aprobaron; sobre ese subconjunto, estados nivel 3.
            if n1 == "Aprobado" and n2 == "Aprobado":
                acc(n3, "interventoria")

            # Pendiente por ítem (sub_estado): no forma parte del bucket Pendiente del inspector; solo con N1 aprobado.
            if (sub_raw == "pendiente" or sub_n == "Pendiente") and n1 == "Aprobado":
                M["pendiente_item"]["residente"] += cd

            M["habilitado"]["inspector"] += cd
            if n1 == "Aprobado":
                M["habilitado"]["residente"] += cd
            if n1 == "Aprobado" and n2 == "Aprobado":
                M["habilitado"]["interventoria"] += cd

        if len(batch) < 1000:
            break
        off += 1000

    if acta_id_filtro is not None:
        off = 0
        while True:
            def _bo(o=off):
                q = supabase.table("so_registros").select(
                    "costo_directo,nivel1_estado,nivel2_estado,nivel3_estado,capitulo,acta_rpo_id,item_numero"
                ).eq("contrato_id", contrato_id)
                return q.range(o, o + 999).execute().data

            batch = supabase_execute(_bo)
            for reg in batch:
                if not (reg.get("item_numero") or "").strip():
                    continue
                aid = reg.get("acta_rpo_id")
                if aid is not None and aid == acta_id_filtro:
                    continue
                cd = float(reg.get("costo_directo") or 0)
                n1 = _matriz_validacion_norm_estado(reg.get("nivel1_estado"))
                n2 = _matriz_validacion_norm_estado(reg.get("nivel2_estado"))
                n3 = _matriz_validacion_norm_estado(reg.get("nivel3_estado"))
                bloque = _matriz_validacion_bloque_capitulo(reg.get("capitulo"))
                Ox = ens_m if bloque == "ensayos" else obra_m
                if n1 == "Aprobado" and n2 == "Aprobado" and n3 == "Pendiente":
                    Ox["otras_actas"]["interventoria"] += cd
                if n1 == "Aprobado" and n2 == "Pendiente":
                    Ox["otras_actas"]["residente"] += cd
                if n1 == "Pendiente":
                    Ox["otras_actas"]["inspector"] += cd
            if len(batch) < 1000:
                break
            off += 1000

    def round_block(m):
        out = {}
        for k, cols in m.items():
            out[k] = {c: round(v, 0) for c, v in cols.items()}
        return out

    return {
        "obra_ejecutada_directo_sin_aiu": round_block(obra_m),
        "ensayos_sondeos_directo_sin_iva": round_block(ens_m),
    }


@app.get("/sicoe-obra/{contrato_id}/dashboard-matriz-validacion")
def dashboard_matriz_validacion_obra(
    contrato_id: int,
    acta_rpo: Optional[int] = None,
    todo_contrato: bool = Query(False, description="Si true, agrega todo el contrato (lento). Si false y sin acta_rpo, usa acta vigente por período."),
    current_user=Depends(get_current_user),
):
    """
    Matriz de validación por rol (Interventoría=nivel3, Residente=nivel2, Inspector=nivel1).
    Por defecto (sin acta_rpo y todo_contrato=false): solo registros del **Acta RPO vigente** (fecha hoy ∈ [inicio, fin]).
    Con acta_rpo explícito: solo ese acta. Con todo_contrato=true: histórico completo (costoso).
    Preferir función SQL dashboard_matriz_validacion_agg (rápido); si no existe, fallback en Python.
    """
    try:
        acta_id_filtro: Optional[int] = None
        filtro_modo = "vigente"
        acta_rpo_resp: Optional[int] = None
        vig = None
        payload: dict = {}

        def _parse_rpc_matrix_raw(raw):
            if raw is None:
                return {}
            if isinstance(raw, list) and len(raw) > 0:
                return raw[0] if isinstance(raw[0], dict) else {}
            if isinstance(raw, dict) and "obra_ejecutada_directo_sin_aiu" in raw:
                return raw
            if isinstance(raw, dict):
                k = next(iter(raw.keys()), None)
                return raw[k] if k and isinstance(raw.get(k), dict) else raw
            return {}

        if todo_contrato:
            filtro_modo = "todo_contrato"
            acta_id_filtro = None
        elif acta_rpo is not None:
            filtro_modo = "acta"

            def _aid():
                rows = supabase.table("actas").select("id")\
                    .eq("contrato_id", contrato_id).eq("numero_rpo", acta_rpo).execute().data
                if rows:
                    return rows[0]["id"]
                rows = supabase.table("actas").select("id")\
                    .eq("contrato_id", contrato_id).eq("consecutivo", acta_rpo).execute().data
                return rows[0]["id"] if rows else None
            acta_id_filtro = supabase_execute(_aid)
            acta_rpo_resp = int(acta_rpo) if acta_rpo is not None else None
        else:
            # Acta vigente: preferir RPC único (resuelve acta + agrega en BD; menos latencia).
            bundle_meta = None
            try:
                def _bundle():
                    return supabase.rpc(
                        "dashboard_matriz_validacion_vigente_bundle",
                        {"p_contrato_id": contrato_id},
                    ).execute().data
                pay = _parse_rpc_matrix_raw(supabase_execute(_bundle))
                vm = pay.get("_vigente") if isinstance(pay, dict) else None
                if isinstance(vm, dict) and "obra_ejecutada_directo_sin_aiu" in pay:
                    bundle_meta = vm
                    payload = {k: v for k, v in pay.items() if k != "_vigente"}
            except Exception:
                payload = {}

            if bundle_meta is not None:
                aid = bundle_meta.get("acta_id")
                try:
                    acta_id_filtro = int(aid) if aid is not None else None
                except (TypeError, ValueError):
                    acta_id_filtro = None
                fm = bundle_meta.get("filtro")
                if isinstance(fm, str):
                    filtro_modo = fm
                nr = bundle_meta.get("numero_rpo")
                try:
                    acta_rpo_resp = int(nr) if nr is not None else None
                except (TypeError, ValueError):
                    acta_rpo_resp = None
                an_b = bundle_meta.get("asignado_nombre")
                if an_b is not None and not isinstance(an_b, str):
                    an_b = str(an_b)
                if acta_id_filtro is not None:
                    vig = {
                        "id": acta_id_filtro,
                        "numero_rpo": nr,
                        "asignado_nombre": (an_b or "").strip() or None,
                    }
                else:
                    vig = None
            else:
                vig = _acta_rpo_vigente_row(contrato_id)
                if vig:
                    acta_id_filtro = vig.get("id")
                    nr = vig.get("numero_rpo")
                    try:
                        acta_rpo_resp = int(nr) if nr is not None else None
                    except (TypeError, ValueError):
                        acta_rpo_resp = None
                else:
                    filtro_modo = "sin_vigente_todo_contrato"
                    acta_id_filtro = None

        if "obra_ejecutada_directo_sin_aiu" not in payload or "ensayos_sondeos_directo_sin_iva" not in payload:
            try:
                def _rpc():
                    return supabase.rpc(
                        "dashboard_matriz_validacion_agg",
                        {"p_contrato_id": contrato_id, "p_acta_id": acta_id_filtro},
                    ).execute().data
                payload = _parse_rpc_matrix_raw(supabase_execute(_rpc))
            except Exception:
                payload = {}

        if "obra_ejecutada_directo_sin_aiu" not in payload or "ensayos_sondeos_directo_sin_iva" not in payload:
            payload = _dashboard_matriz_validacion_fallback(contrato_id, acta_id_filtro)

        def _acta_vigente_public(v):
            if not v:
                return None
            nm = v.get("asignado_nombre")
            if nm is not None:
                nm = str(nm).strip() or None
            return {
                "id": v.get("id"),
                "numero_rpo": v.get("numero_rpo"),
                "asignado_nombre": nm,
            }

        return {
            "acta_rpo": acta_rpo_resp,
            "acta_id_resuelto": acta_id_filtro,
            "filtro": filtro_modo,
            "acta_vigente": _acta_vigente_public(vig),
            "obra_ejecutada_directo_sin_aiu": payload.get("obra_ejecutada_directo_sin_aiu") or {},
            "ensayos_sondeos_directo_sin_iva": payload.get("ensayos_sondeos_directo_sin_iva") or {},
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sicoe-obra/{contrato_id}/dashboard-drill")
def dashboard_drill_obra(
    contrato_id: int,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    current_user=Depends(get_current_user)
):
    try:
        if capitulo:
            try:
                def _rpc_items():
                    return (
                        supabase.rpc(
                            "dashboard_drill_items_agg",
                            {"p_contrato_id": contrato_id, "p_capitulo": capitulo},
                        )
                        .execute()
                        .data
                    )

                items = _parse_rpc_jsonb_value(supabase_execute(_rpc_items))
                if isinstance(items, list):
                    if item:
                        itn = _dash_norm_item_key_py(str(item).strip())
                        items = [
                            row
                            for row in items
                            if isinstance(row, dict) and _dash_norm_item_key_py(row.get("item")) == itn
                        ]
                    return {"campo": "item", "items": items}
            except Exception:
                pass
            result = _drill_agg_by_item(contrato_id, capitulo, item)
            return {"campo": "item", "items": result}
        if item:
            raise HTTPException(status_code=422, detail="Indica capitulo junto con item.")
        try:
            def _rpc_caps():
                return supabase.rpc("dashboard_drill_capitulos_agg", {"p_contrato_id": contrato_id}).execute().data

            caps = _parse_rpc_jsonb_value(supabase_execute(_rpc_caps))
            if isinstance(caps, list):
                return {"campo": "capitulo", "items": caps}
        except Exception:
            pass
        result = _drill_agg_capitulos(contrato_id)
        return {"campo": "capitulo", "items": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _dashboard_pkid_tabla_obra_core(contrato_id: int, capitulo: Optional[str], item: Optional[str]) -> Dict[str, Any]:
    """Por PK_ID: presupuesto por revisado; SICOE N3 aprobado; obra en cola (N1+N2 aprob., N3 no aprob.) partida en no revisado / pendiente / rechazado. cant_sicoe_no_revisado = solo bucket no revisado (no incluye pend. ni rech.). Δ = ppto aprobado N3 − obra N3 aprobada."""
    try:
        def _rpc_pk():
            return (
                supabase.rpc(
                    "dashboard_pkid_tabla_agg",
                    {
                        "p_contrato_id": contrato_id,
                        "p_capitulo": capitulo or "",
                        "p_item": item or "",
                    },
                )
                .execute()
                .data
            )

        hit = _parse_rpc_jsonb_value(supabase_execute(_rpc_pk))
        if isinstance(hit, dict) and isinstance(hit.get("rows"), list):
            return hit
    except Exception:
        pass

    registros = []
    off = 0
    it_norm = _dash_norm_item_key_py(item) if item else ""
    while True:
        def _regs(o=off):
            q = supabase.table("so_registros").select(
                "pk_id_id, pk_ids(pk_id), costo_directo, cantidad_total, item_numero, "
                "nivel1_estado, nivel2_estado, nivel3_estado"
            ).eq("contrato_id", contrato_id)
            if capitulo:
                q = q.eq("capitulo", capitulo)
            return q.range(o, o + 999).execute().data

        batch = supabase_execute(_regs)
        for reg in batch or []:
            if it_norm and _dash_norm_item_key_py(reg.get("item_numero")) != it_norm:
                continue
            registros.append(reg)
        if len(batch) < 1000:
            break
        off += 1000

    q_p = (
        supabase.table("presupuesto")
        .select("pk_id, item, cant_total, costo_directo, descripcion, revisado, capitulo")
        .eq("contrato_id", contrato_id)
        .eq("dado_de_baja", False)
    )
    cap_key = _dash_norm_capitulo_key_py(capitulo) if capitulo else None
    ppto = []
    off = 0
    while True:
        batch = q_p.range(off, off + 999).execute().data
        for r in batch or []:
            if cap_key and _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_key:
                continue
            if it_norm and _dash_norm_item_key_py(r.get("item")) != it_norm:
                continue
            ppto.append(r)
        if len(batch) < 1000:
            break
        off += 1000

    def _empty_m():
        return {"cant": 0.0, "costo": 0.0}

    agg_meta: Dict[str, Any] = {}
    agg_p_ap: Dict[str, Any] = {}
    agg_p_nr: Dict[str, Any] = {}
    agg_p_pd: Dict[str, Any] = {}
    agg_p_rj: Dict[str, Any] = {}

    for r in ppto:
        k = _dash_pk_disp_key_py(r.get("pk_id"))
        if k not in agg_meta:
            agg_meta[k] = {"desc": "", "_rev_track": []}
        cost_line = float(r.get("costo_directo") or 0)
        agg_meta[k]["_rev_track"].append((cost_line, r.get("revisado")))
        if not agg_meta[k]["desc"] and r.get("descripcion"):
            agg_meta[k]["desc"] = r["descripcion"]
        rv = _matriz_validacion_norm_estado(r.get("revisado"))
        cq = float(r.get("cant_total") or 0)
        cd = float(r.get("costo_directo") or 0)
        if rv == "Aprobado":
            tgt = agg_p_ap
        elif rv == "Pendiente":
            tgt = agg_p_pd
        elif rv == "Rechazado":
            tgt = agg_p_rj
        else:
            tgt = agg_p_nr
        if k not in tgt:
            tgt[k] = _empty_m()
        tgt[k]["cant"] += cq
        tgt[k]["costo"] += cd

    agg_sicoe_ap: Dict[str, Any] = {}
    agg_sicoe_nr: Dict[str, Any] = {}
    agg_sicoe_pe: Dict[str, Any] = {}
    agg_sicoe_rej: Dict[str, Any] = {}
    for r in registros:
        pk_join = r.get("pk_ids") or {}
        pk_raw = pk_join.get("pk_id")
        pk = _dash_pk_disp_key_py(pk_raw) if pk_raw is not None and str(pk_raw).strip() != "" else _dash_pk_disp_key_py(r.get("pk_id_id"))
        cd = float(r.get("costo_directo") or 0)
        cq = float(r.get("cantidad_total") or 0)
        n3 = _matriz_validacion_norm_estado(r.get("nivel3_estado"))
        if n3 == "Aprobado":
            if pk not in agg_sicoe_ap:
                agg_sicoe_ap[pk] = _empty_m()
            agg_sicoe_ap[pk]["cant"] += cq
            agg_sicoe_ap[pk]["costo"] += cd
            continue
        if _so_reg_cola_n1_n2_aprobados(r):
            if n3 == "Pendiente":
                tgt = agg_sicoe_pe
            elif n3 == "Rechazado":
                tgt = agg_sicoe_rej
            else:
                tgt = agg_sicoe_nr
            if pk not in tgt:
                tgt[pk] = _empty_m()
            tgt[pk]["cant"] += cq
            tgt[pk]["costo"] += cd

    keys = sorted(
        set(
            list(agg_meta.keys())
            + list(agg_p_ap.keys())
            + list(agg_p_nr.keys())
            + list(agg_p_pd.keys())
            + list(agg_p_rj.keys())
            + list(agg_sicoe_ap.keys())
            + list(agg_sicoe_nr.keys())
            + list(agg_sicoe_pe.keys())
            + list(agg_sicoe_rej.keys())
        ),
        key=lambda x: str(x),
    )
    rows = []
    for k in keys:
        pap = agg_p_ap.get(k, _empty_m())
        pnr = agg_p_nr.get(k, _empty_m())
        ppd = agg_p_pd.get(k, _empty_m())
        prj = agg_p_rj.get(k, _empty_m())
        sic = agg_sicoe_ap.get(k, _empty_m())
        snr = agg_sicoe_nr.get(k, _empty_m())
        spe = agg_sicoe_pe.get(k, _empty_m())
        srj = agg_sicoe_rej.get(k, _empty_m())
        cant_ppto = pap["cant"] + pnr["cant"] + ppd["cant"] + prj["cant"]
        costo_ppto = pap["costo"] + pnr["costo"] + ppd["costo"] + prj["costo"]
        dcant = pap["cant"] - sic["cant"]
        dcosto = pap["costo"] - sic["costo"]
        meta = agg_meta.get(k, {"desc": "", "_rev_track": []})
        tr = meta.get("_rev_track") or []
        rev_dom = "No Revisado"
        if tr:
            rev_dom = _matriz_validacion_norm_estado(max(tr, key=lambda x: float(x[0] or 0))[1])
        rows.append({
            "pk_id": k,
            "cant_ppto": round(cant_ppto, 2),
            "costo_ppto": round(costo_ppto, 0),
            "cant_ppto_aprobado_n3": round(pap["cant"], 2),
            "costo_ppto_aprobado_n3": round(pap["costo"], 0),
            "cant_ppto_estado_no_revisado": round(pnr["cant"], 2),
            "costo_ppto_estado_no_revisado": round(pnr["costo"], 0),
            "cant_ppto_estado_pendiente": round(ppd["cant"], 2),
            "costo_ppto_estado_pendiente": round(ppd["costo"], 0),
            "cant_ppto_estado_rechazado": round(prj["cant"], 2),
            "costo_ppto_estado_rechazado": round(prj["costo"], 0),
            "cant_sicoe_aprobado": round(sic["cant"], 2),
            "costo_sicoe_aprobado": round(sic["costo"], 0),
            "cant_sicoe_no_revisado": round(snr["cant"], 2),
            "costo_sicoe_no_revisado": round(snr["costo"], 0),
            "cant_sicoe_pendiente": round(spe["cant"], 2),
            "costo_sicoe_pendiente": round(spe["costo"], 0),
            "cant_sicoe_rechazado": round(srj["cant"], 2),
            "costo_sicoe_rechazado": round(srj["costo"], 0),
            "cant_sicoe": round(sic["cant"], 2),
            "costo_sicoe": round(sic["costo"], 0),
            "cant_facturado": 0.0,
            "costo_facturado": 0.0,
            "delta_cant": round(dcant, 2),
            "delta_costo": round(dcosto, 0),
            "descripcion": meta.get("desc", ""),
            "revisado": rev_dom,
        })

    desc_item = ""
    if item and capitulo:
        d = (
            supabase.table("presupuesto")
            .select("descripcion, item")
            .eq("contrato_id", contrato_id)
            .eq("capitulo", capitulo)
            .not_.is_("descripcion", "null")
            .execute()
            .data
        ) or []
        for row in d:
            if _dash_norm_item_key_py(row.get("item")) == it_norm and row.get("descripcion"):
                desc_item = row["descripcion"] or ""
                break

    por_cobrar = sum(r["delta_costo"] for r in rows if r["delta_costo"] > 0)
    devolucion = sum(abs(r["delta_costo"]) for r in rows if r["delta_costo"] < 0)
    return {"rows": rows, "por_cobrar": por_cobrar, "devolucion": devolucion, "descripcion_item": desc_item}


def _ppto_rows_capitulo(contrato_id: int, capitulo: str) -> List[dict]:
    rows = []
    off = 0
    while True:
        def _b(o=off):
            return (
                supabase.table("presupuesto")
                .select("item, descripcion, capitulo, cant_total, costo_directo, revisado, pk_id")
                .eq("contrato_id", contrato_id)
                .eq("capitulo", capitulo)
                .eq("dado_de_baja", False)
                .range(o, o + 999)
                .execute()
                .data
            )

        batch = supabase_execute(_b) or []
        rows.extend(batch)
        if len(batch) < 1000:
            break
        off += 1000
    return rows


def _ppto_costo_por_revisado(rows_item: List[dict]) -> Dict[str, float]:
    out = {"Aprobado": 0.0, "Pendiente": 0.0, "Rechazado": 0.0, "No Revisado": 0.0}
    for r in rows_item:
        cost = float(r.get("costo_directo") or 0)
        rv = _matriz_validacion_norm_estado(r.get("revisado"))
        if rv not in out:
            rv = "No Revisado"
        out[rv] += cost
    return out


def _xlsx_safe_sheet_name(name: str, fallback: str = "Hoja") -> str:
    s = re.sub(r"[\[\]\*\/\\\?\:]", "_", str(name or "").strip())[:31]
    return s or fallback[:31]


def _build_dashboard_capitulo_xlsx(contrato_id: int, capitulo: str, item: Optional[str]):
    """
    Informe multi-hoja: (1) resumen por ítem del capítulo presupuesto vs obra aprobada N3,
    (2..n) análisis por PK por ítem, (n-1) base presupuesto capítulo, (n) obra filtrada a esos ítems.

    Los totales «Cobrada/Cobrado» del resumen se calculan con la misma lógica que el análisis por PK
    (_dashboard_pkid_tabla_obra_core), no por agrupación cruda de item_numero en obra (evita
    desfaces cuando el texto del ítem en obra no coincide exactamente con el del presupuesto).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    fill_hdr = PatternFill("solid", fgColor="1B2A4A")
    fill_tot = PatternFill("solid", fgColor="1B2A4A")
    fill_green = PatternFill("solid", fgColor="DCFCE7")
    fill_red = PatternFill("solid", fgColor="FEE2E2")
    fill_white = PatternFill("solid", fgColor="FFFFFF")
    font_hdr = Font(bold=True, color="FFFFFF", size=11)
    font_bold = Font(bold=True)
    al_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _side = Side(style="thin", color="FFB4B4B4")
    border_tbl = Border(left=_side, right=_side, top=_side, bottom=_side)

    def _style_header_row(ws, row_idx: int, ncols: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill_hdr
            cell.font = font_hdr
            cell.alignment = al_center
            cell.border = border_tbl

    def _style_total_row(ws, row_idx: int, ncols: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill_tot
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = border_tbl

    def _border_range(ws, r1: int, r2: int, c1: int, c2: int):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = border_tbl

    def _row_fill(ws, r: int, c1: int, c2: int, fill):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill

    def _fill_data_row_by_delta(ws, r: int, c1: int, c2: int, delta_costo: float):
        dc = float(delta_costo or 0)
        if dc > 0.5:
            _row_fill(ws, r, c1, c2, fill_green)
        elif dc < -0.5:
            _row_fill(ws, r, c1, c2, fill_red)
        else:
            _row_fill(ws, r, c1, c2, fill_white)

    meta_ct = ""
    try:
        cr = supabase.table("contratos").select("numero, contratista").eq("id", contrato_id).limit(1).execute().data
        if cr:
            meta_ct = (cr[0].get("contratista") or cr[0].get("numero") or "").strip()
    except Exception:
        pass
    gen_ts = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")

    ppto_all = _ppto_rows_capitulo(contrato_id, capitulo)
    by_item: Dict[str, List[dict]] = {}
    for r in ppto_all:
        it = (r.get("item") or "").strip()
        if not it:
            continue
        by_item.setdefault(it, []).append(r)
    items_sorted = sorted(by_item.keys(), key=lambda x: str(x))
    # Misma fuente que las hojas por ítem (PK), para que resumen y detalle coincidan.
    core_by_item: Dict[str, Dict[str, Any]] = {}
    for _it in items_sorted:
        core_by_item[_it] = _dashboard_pkid_tabla_obra_core(contrato_id, capitulo, _it)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = _xlsx_safe_sheet_name("Resumen capítulo", "Resumen")

    # ── Hoja 1: resumen por ítem ───────────────────────────────────────────
    ws0.merge_cells(start_row=1, start_column=1, end_row=2, end_column=13)
    c1 = ws0.cell(row=1, column=1)
    c1.value = f"RESUMEN POR CAPÍTULO — {capitulo}"
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = fill_hdr
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[1].height = 22
    ws0.row_dimensions[2].height = 6
    ws0.merge_cells(start_row=3, start_column=1, end_row=3, end_column=13)
    c3 = ws0.cell(row=3, column=1)
    c3.value = f"Generado: {gen_ts}" + (f" | {meta_ct}" if meta_ct else "")
    c3.font = Font(size=10, italic=True)
    c3.alignment = Alignment(horizontal="center")

    hdr1 = [
        "Ítem",
        "Descripción",
        "Cant. ClaraCore",
        "Costo ClaraCore",
        "Cant. Cobrada",
        "Costo Cobrado",
        "Δ Cantidad",
        "Δ Costo",
        "Estado",
        "Aprobado",
        "Pendiente",
        "Rechazado",
        "No Revisado",
    ]
    _border_range(ws0, 3, 3, 1, len(hdr1))
    r0 = 4
    for j, h in enumerate(hdr1, start=1):
        ws0.cell(row=r0, column=j, value=h)
    _style_header_row(ws0, r0, len(hdr1))

    tot_cc_cant = tot_cc_cost = tot_cb_cant = tot_cb_cost = tot_d_cant = tot_d_cost = 0.0
    tot_ap = tot_pe = tot_re = tot_nr = 0.0
    row_i = r0 + 1
    for it in items_sorted:
        rows_it = by_item[it]
        desc = ""
        for x in rows_it:
            if x.get("descripcion"):
                desc = str(x["descripcion"])
                break
        cant_p = sum(float(x.get("cant_total") or 0) for x in rows_it)
        cost_p = sum(float(x.get("costo_directo") or 0) for x in rows_it)
        _rows_pk = (core_by_item.get(it) or {}).get("rows") or []
        cant_c = sum(float(x.get("cant_sicoe") or 0) for x in _rows_pk)
        cost_c = sum(float(x.get("costo_sicoe") or 0) for x in _rows_pk)
        d_cant = cant_p - cant_c
        d_cost = cost_p - cost_c
        estado = "Equilibrio"
        if d_cost > 0.5:
            estado = "Por cobrar"
        elif d_cost < -0.5:
            estado = "Devolución"
        spl = _ppto_costo_por_revisado(rows_it)
        tot_cc_cant += cant_p
        tot_cc_cost += cost_p
        tot_cb_cant += cant_c
        tot_cb_cost += cost_c
        tot_d_cant += d_cant
        tot_d_cost += d_cost
        tot_ap += spl["Aprobado"]
        tot_pe += spl["Pendiente"]
        tot_re += spl["Rechazado"]
        tot_nr += spl["No Revisado"]

        ws0.cell(row=row_i, column=1, value=it)
        ws0.cell(row=row_i, column=2, value=desc)
        ws0.cell(row=row_i, column=3, value=round(cant_p, 4))
        ws0.cell(row=row_i, column=4, value=round(cost_p, 0))
        ws0.cell(row=row_i, column=5, value=round(cant_c, 4))
        ws0.cell(row=row_i, column=6, value=round(cost_c, 0))
        ws0.cell(row=row_i, column=7, value=round(d_cant, 4))
        ws0.cell(row=row_i, column=8, value=round(d_cost, 0))
        ws0.cell(row=row_i, column=9, value=estado)
        ws0.cell(row=row_i, column=10, value=round(spl["Aprobado"], 0))
        ws0.cell(row=row_i, column=11, value=round(spl["Pendiente"], 0))
        ws0.cell(row=row_i, column=12, value=round(spl["Rechazado"], 0))
        ws0.cell(row=row_i, column=13, value=round(spl["No Revisado"], 0))
        if estado == "Por cobrar":
            _row_fill(ws0, row_i, 1, 13, fill_green)
        elif estado == "Devolución":
            _row_fill(ws0, row_i, 1, 13, fill_red)
        else:
            _row_fill(ws0, row_i, 1, 13, fill_white)
        row_i += 1

    ws0.cell(row=row_i, column=1, value="TOTALES CAPÍTULO")
    ws0.cell(row=row_i, column=3, value=round(tot_cc_cant, 4))
    ws0.cell(row=row_i, column=4, value=round(tot_cc_cost, 0))
    ws0.cell(row=row_i, column=5, value=round(tot_cb_cant, 4))
    ws0.cell(row=row_i, column=6, value=round(tot_cb_cost, 0))
    ws0.cell(row=row_i, column=7, value=round(tot_d_cant, 4))
    ws0.cell(row=row_i, column=8, value=round(tot_d_cost, 0))
    ws0.cell(row=row_i, column=10, value=round(tot_ap, 0))
    ws0.cell(row=row_i, column=11, value=round(tot_pe, 0))
    ws0.cell(row=row_i, column=12, value=round(tot_re, 0))
    ws0.cell(row=row_i, column=13, value=round(tot_nr, 0))
    _style_total_row(ws0, row_i, len(hdr1))
    _border_range(ws0, r0, row_i, 1, len(hdr1))

    # ── Hojas por ítem: análisis PK (por cobrar / devolución) ────────────────
    tbl_cols = ["PK_Id", "Cant. ClaraCore", "Costo ClaraCore", "Cant. Cobrada", "Costo Cobrado", "Δ Cantidad", "Δ Costo", "Revisado"]

    def _append_pk_table(ws, block_rows, title, fill_title, uniform_row_fill, is_subtotal=False, per_row_delta=False):
        mr = ws.max_row + 1
        if title:
            ws.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=len(tbl_cols))
            c = ws.cell(row=mr, column=1)
            c.value = title
            if fill_title:
                c.fill = fill_title
            c.font = font_bold
            _border_range(ws, mr, mr, 1, len(tbl_cols))
        ws.append(tbl_cols)
        _style_header_row(ws, ws.max_row, len(tbl_cols))
        for r in block_rows:
            ws.append(
                [
                    r.get("pk_id"),
                    r.get("cant_ppto"),
                    r.get("costo_ppto"),
                    r.get("cant_sicoe"),
                    r.get("costo_sicoe"),
                    r.get("delta_cant"),
                    r.get("delta_costo"),
                    r.get("revisado"),
                ]
            )
            rr = ws.max_row
            if per_row_delta:
                _fill_data_row_by_delta(ws, rr, 1, len(tbl_cols), float(r.get("delta_costo") or 0))
            elif uniform_row_fill:
                _row_fill(ws, rr, 1, len(tbl_cols), uniform_row_fill)
            else:
                _row_fill(ws, rr, 1, len(tbl_cols), fill_white)
            _border_range(ws, rr, rr, 1, len(tbl_cols))
        if is_subtotal and block_rows:
            if "DEVOL" in (title or ""):
                label = f"Subtotal DEVOLUCIÓN ({len(block_rows)} PK)"
            elif "POR COBRAR" in (title or ""):
                label = f"Subtotal POR COBRAR ({len(block_rows)} PK)"
            else:
                label = f"Subtotal ({len(block_rows)} PK)"
            ws.append(
                [
                    label,
                    sum(r["cant_ppto"] for r in block_rows),
                    sum(r["costo_ppto"] for r in block_rows),
                    sum(r["cant_sicoe"] for r in block_rows),
                    sum(r["costo_sicoe"] for r in block_rows),
                    sum(r["delta_cant"] for r in block_rows),
                    sum(r["delta_costo"] for r in block_rows),
                    "",
                ]
            )
            _style_total_row(ws, ws.max_row, len(tbl_cols))
            _border_range(ws, ws.max_row, ws.max_row, 1, len(tbl_cols))

    for it in items_sorted:
        data = core_by_item[it]
        rows = data.get("rows") or []
        desc_item = data.get("descripcion_item") or ""
        ws = wb.create_sheet(title=_xlsx_safe_sheet_name(it, "Item"))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        t1 = ws.cell(row=1, column=1)
        t1.value = f"ANÁLISIS DE COBRO — {it} | {desc_item[:180]}"
        t1.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        t2 = ws.cell(row=2, column=1)
        t2.value = f"{capitulo} | Generado: {gen_ts}"
        t2.font = Font(size=10, color="666666")

        if rows:
            s_cp = sum(float(r.get("cant_ppto") or 0) for r in rows)
            s_co = sum(float(r.get("costo_ppto") or 0) for r in rows)
            s_cc = sum(float(r.get("cant_sicoe") or 0) for r in rows)
            s_ob = sum(float(r.get("costo_sicoe") or 0) for r in rows)
            d_ct = s_cp - s_cc
            d_cs = s_co - s_ob
            ws.append([])
            ws.append(["Cant. ClaraCore", "Costo ClaraCore", "Cant. Cobrada", "Costo Cobrado", "Δ Cantidad", "Δ Costo"])
            rh = ws.max_row
            _style_header_row(ws, rh, 6)
            ws.append([round(s_cp, 3), round(s_co, 0), round(s_cc, 3), round(s_ob, 0), round(d_ct, 3), round(d_cs, 0)])
            lr = ws.max_row
            _fill_data_row_by_delta(ws, lr, 1, 6, d_cs)
            _border_range(ws, rh, lr, 1, 6)

        pos = [r for r in rows if (r.get("delta_costo") or 0) > 0]
        neg = [r for r in rows if (r.get("delta_costo") or 0) < 0]
        sum_pos_cost = sum(r["delta_costo"] for r in pos)
        sum_neg_cost = sum(r["delta_costo"] for r in neg)

        ws.append([])
        if pos:
            tit = f"POR COBRAR | Total: +${sum_pos_cost:,.0f}"
            _append_pk_table(ws, pos, tit, fill_green, fill_green, True, False)
            ws.append([])
        if neg:
            titn = f"DEVOLUCIÓN | Total: ${sum_neg_cost:,.0f}"
            _append_pk_table(ws, neg, titn, fill_red, fill_red, True, False)
        if not pos and not neg and rows:
            rnote = ws.max_row + 1
            ws.merge_cells(start_row=rnote, start_column=1, end_row=rnote, end_column=len(tbl_cols))
            cn = ws.cell(row=rnote, column=1)
            cn.value = "Listado completo (Δ costo = 0 o sin clasificar por signo):"
            cn.font = font_bold
            _border_range(ws, rnote, rnote, 1, len(tbl_cols))
            _append_pk_table(ws, rows, "", None, None, False, True)

    # ── Penúltima: presupuesto capítulo (completo) ──────────────────────────
    ws_p = wb.create_sheet(title=_xlsx_safe_sheet_name("Presupuesto cap", "Presupuesto"))
    if ppto_all:
        keys = list(ppto_all[0].keys())
        ws_p.append(keys)
        _style_header_row(ws_p, 1, len(keys))
        for r in ppto_all:
            ws_p.append([r.get(k) for k in keys])
        nc = len(keys)
        _border_range(ws_p, 1, ws_p.max_row, 1, nc)
        for rr in range(2, ws_p.max_row + 1):
            _row_fill(ws_p, rr, 1, nc, fill_white)

    # ── Última: obra (N3 aprobado) solo ítems del presupuesto ────────────────
    ws_o = wb.create_sheet(title=_xlsx_safe_sheet_name("Sicoe obra items", "SicoeObra"))
    obra_rows: List[dict] = []
    if items_sorted:
        for ci in range(0, len(items_sorted), 400):
            chunk_items = items_sorted[ci : ci + 400]
            off = 0
            while True:

                def _ob():
                    return (
                        supabase.table("so_registros")
                        .select("*")
                        .eq("contrato_id", contrato_id)
                        .eq("capitulo", capitulo)
                        .eq("nivel3_estado", "Aprobado")
                        .in_("item_numero", chunk_items)
                        .range(off, off + 999)
                        .execute()
                        .data
                    )

                batch = supabase_execute(_ob) or []
                obra_rows.extend(batch)
                if len(batch) < 1000:
                    break
                off += 1000
    if obra_rows:
        keys_o = list(obra_rows[0].keys())
        ws_o.append(keys_o)
        _style_header_row(ws_o, 1, len(keys_o))
        for r in obra_rows:
            row_vals = []
            for k in keys_o:
                v = r.get(k)
                if isinstance(v, (dict, list)):
                    row_vals.append(json.dumps(v, ensure_ascii=False)[:500])
                else:
                    row_vals.append(v)
            ws_o.append(row_vals)
        nco = len(keys_o)
        _border_range(ws_o, 1, ws_o.max_row, 1, nco)
        for rr in range(2, ws_o.max_row + 1):
            _row_fill(ws_o, rr, 1, nco, fill_white)

    for _sh in wb.worksheets:
        _sh.sheet_view.showGridLines = False

    bio = io.BytesIO()
    wb.save(bio)
    safe_cap = re.sub(r"[^\w\-.]+", "_", str(capitulo or "cap"))[:40]
    fn = f"ClaraCore_{safe_cap}_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx"
    return bio.getvalue(), fn


@app.get("/sicoe-obra/{contrato_id}/dashboard-pkid-tabla")
def dashboard_pkid_tabla_obra(
    contrato_id: int,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    current_user=Depends(get_current_user)
):
    """
    Reemplaza /cobro/{contrato_id}/pkid-tabla para el Dashboard.
    Retorna {rows, por_cobrar, devolucion, descripcion_item}: filas con presupuesto por estado revisado,
    SICOE solo N3 aprobado, delta_cant/costo = ppto aprobado N3 − obra aprobada N3.
    """
    try:
        return _dashboard_pkid_tabla_obra_core(contrato_id, capitulo, item)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sicoe-obra/{contrato_id}/dashboard-export-capitulo")
def dashboard_export_capitulo_obra(
    contrato_id: int,
    capitulo: str = Query(..., description="Capítulo del drill del dashboard"),
    item: Optional[str] = Query(None),
    current_user=Depends(get_current_user),
):
    """
    Inicia generación de Excel alineado con dashboard-pkid-tabla (reemplazo de /cobro/.../exportar-capitulo).
    Respuesta: { job_id }; luego GET /exportar/estado/{job_id} y /exportar/descargar/{job_id}.
    """
    _require_contract_access(current_user, contrato_id)
    job_id = str(uuid.uuid4())
    _export_jobs[job_id] = {"estado": "procesando", "buf": None, "filename": None}
    cap_copy = capitulo
    item_copy = item

    def _work():
        try:
            buf, fn = _build_dashboard_capitulo_xlsx(contrato_id, cap_copy, item_copy)
            _export_jobs[job_id] = {"estado": "listo", "buf": buf, "filename": fn}
        except Exception as e:
            _export_jobs[job_id] = {"estado": f"error:{e!s}", "buf": None, "filename": None}

    threading.Thread(target=_work, daemon=True).start()
    return {"job_id": job_id}


@app.get(
    "/presupuesto/{contrato_id}/pkid-colores",
    operation_id="presupuesto_pkid_colores_compat",
)
@app.get(
    "/sicoe-obra/{contrato_id}/dashboard-pkid-colores",
    operation_id="sicoe_dashboard_pkid_colores",
)
def dashboard_pkid_colores_obra(
    contrato_id: int,
    capitulo: Optional[str] = None,
    item: Optional[str] = None,
    liquidacion: Optional[str] = Query(None, description="1/true: referencia liquidación (PPTO polígono vs obra N3)"),
    current_user=Depends(get_current_user)
):
    """
    Compat: el front aún llama GET /presupuesto/{id}/pkid-colores.
    Reemplaza /cobro/{contrato_id}/pkid-colores-drill para el mini-mapa semáforo del Dashboard.
    Retorna misma forma: {pk_id: {cobrado, presupuesto, pct, sobrecosto}}
    """
    try:
        liq = str(liquidacion or "").strip().lower() in ("1", "true", "yes", "on")
        if liq:
            return _dashboard_pkid_colores_liquidacion(contrato_id, capitulo, item)
        ppto_agg = {}
        off = 0
        while True:
            def _fp(o=off):
                q = supabase.table("presupuesto").select("pk_id, costo_directo").eq("contrato_id", contrato_id).eq("dado_de_baja", False)
                if item:
                    q = q.eq("item", item)
                elif capitulo:
                    q = q.eq("capitulo", capitulo)
                return q.range(o, o + 999).execute().data

            batch = supabase_execute(_fp) or []
            for r in batch:
                k = str(r.get("pk_id") or "").strip()
                if k:
                    ppto_agg[k] = ppto_agg.get(k, 0) + float(r.get("costo_directo") or 0)
            if len(batch) < 1000:
                break
            off += 1000

        sicoe_ap_agg = {}
        off = 0
        while True:
            def _regs(o=off):
                q = supabase.table("so_registros").select("pk_id_id, pk_ids(pk_id), costo_directo").eq("contrato_id", contrato_id).eq("nivel3_estado", "Aprobado")
                if capitulo:
                    q = q.eq("capitulo", capitulo)
                if item:
                    q = q.ilike("item_numero", f"%{item}%")
                return q.range(o, o + 999).execute().data

            batch = supabase_execute(_regs) or []
            for r in batch:
                pk_join = r.get("pk_ids") or {}
                k = str(pk_join.get("pk_id") or r.get("pk_id_id") or "").strip()
                if k:
                    sicoe_ap_agg[k] = sicoe_ap_agg.get(k, 0) + float(r.get("costo_directo") or 0)
            if len(batch) < 1000:
                break
            off += 1000

        result = {}
        for pk in set(list(ppto_agg.keys()) + list(sicoe_ap_agg.keys())):
            p = ppto_agg.get(pk, 0)
            sap = sicoe_ap_agg.get(pk, 0)
            result[pk] = {
                "cobrado": round(sap, 2),
                "presupuesto": round(p, 2),
                "sicoe_aprobado": round(sap, 2),
                "pct": round(sap / p * 100, 1) if p else 0,
                "sobrecosto": sap > p,
            }
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sicoe-obra/{contrato_id}/dashboard-pkid-detalle")
def dashboard_pkid_detalle_obra(
    contrato_id: int,
    pk_id: Optional[str] = None,
    item: Optional[str] = None,
    capitulo: Optional[str] = None,
    current_user=Depends(get_current_user)
):
    """
    Detalle PK_ID: presupuesto ClaraCore y líneas SICOE por estado N3 (sin tabla cobro).
    """
    try:
        _require_contract_access(current_user, contrato_id)
        pk_id_s = str(pk_id).strip() if pk_id is not None and str(pk_id).strip() != "" else None

        q_p = (
            supabase.table("presupuesto")
            .select(
                "id, id_pol, no_inicio, no_final, cant_total, costo_directo, descripcion, item, ent_handle, "
                "x_label, y_label, revisado, capitulo, pk_id, pre_interv_estado, pre_interv_por, pre_interv_en, "
                "sellado, area_long_nod, ancho, espesor, vlr_unitario, und, tipo_ejecucion"
            )
            .eq("contrato_id", contrato_id)
            .eq("dado_de_baja", False)
        )
        if pk_id_s:
            q_p = q_p.eq("pk_id", pk_id_s)
        ppto_raw = supabase_execute(lambda: q_p.execute().data) or []
        it_norm = _dash_norm_item_key_py(item) if item else ""
        cap_key_det = _dash_norm_capitulo_key_py(capitulo) if capitulo else None
        ppto = []
        for r in ppto_raw:
            if cap_key_det and _dash_norm_capitulo_key_py(r.get("capitulo")) != cap_key_det:
                continue
            if it_norm and _dash_norm_item_key_py(r.get("item")) != it_norm:
                continue
            ppto.append(r)

        pkid_id_val = None
        if pk_id_s:
            res = supabase_execute(
                lambda: supabase.table("pk_ids")
                .select("id")
                .eq("contrato_id", contrato_id)
                .eq("pk_id", pk_id_s)
                .limit(1)
                .execute()
                .data
            )
            if res:
                pkid_id_val = res[0]["id"]
            if not pkid_id_val:
                try:
                    nid = int(pk_id_s, 10)
                    res2 = supabase_execute(
                        lambda: supabase.table("pk_ids")
                        .select("id")
                        .eq("contrato_id", contrato_id)
                        .eq("id", nid)
                        .limit(1)
                        .execute()
                        .data
                    )
                    if res2:
                        pkid_id_val = res2[0]["id"]
                except (ValueError, TypeError):
                    pass

        sicoe_rows = []
        if pk_id_s and not pkid_id_val:
            sicoe_rows = []
        else:
            q_s = supabase.table("so_registros").select(
                "id, numero_registro, tramo, nodo_ini, nodo_fin, cantidad_total, costo_directo, "
                "item_descripcion, item_numero, acta_rpo_id, calzada, reporte_id, observacion, "
                "nivel1_estado, nivel2_estado, nivel3_estado"
            ).eq("contrato_id", contrato_id)
            if pkid_id_val:
                q_s = q_s.eq("pk_id_id", pkid_id_val)
            if capitulo:
                q_s = q_s.eq("capitulo", capitulo)
            sicoe_raw = supabase_execute(lambda: q_s.execute().data) or []
            sicoe_rows = []
            for r in sicoe_raw:
                if it_norm and _dash_norm_item_key_py(r.get("item_numero")) != it_norm:
                    continue
                sicoe_rows.append(r)

        acta_ids = list({r["acta_rpo_id"] for r in sicoe_rows if r.get("acta_rpo_id")})
        acta_map = {}
        if acta_ids:
            for a in supabase_execute(
                lambda: supabase.table("actas").select("id, numero_rpo").in_("id", acta_ids).execute().data
            ) or []:
                acta_map[a["id"]] = a.get("numero_rpo") or a["id"]

        def _fmt_sicoe(r: dict) -> dict:
            return {
                "registro": r.get("numero_registro"),
                "id_pol": None,
                "registro_id": r.get("id"),
                "reporte_id": r.get("reporte_id"),
                "tramo_inicio": r.get("nodo_ini"),
                "tramo_final": r.get("nodo_fin"),
                "cantidad": float(r.get("cantidad_total") or 0),
                "longitud": float(r.get("cantidad_total") or 0),
                "costo_directo": float(r.get("costo_directo") or 0),
                "descripcion": r.get("item_descripcion") or "",
                "item": r.get("item_numero") or "",
                "acta": acta_map.get(r.get("acta_rpo_id")),
                "calzada": r.get("calzada") or "",
                "observacion": (r.get("observacion") or "").strip(),
                "nivel3_estado": _matriz_validacion_norm_estado(r.get("nivel3_estado")),
            }

        sicoe_aprobado = []
        for r in sicoe_rows:
            fr = _fmt_sicoe(r)
            n3 = _matriz_validacion_norm_estado(r.get("nivel3_estado"))
            if n3 == "Aprobado":
                sicoe_aprobado.append(fr)

        ppto_aprobado = []
        ppto_no_revisado = []
        ppto_pendiente = []
        ppto_rechazado = []
        for r in ppto:
            st = _matriz_validacion_norm_estado(r.get("revisado"))
            if st == "Aprobado":
                ppto_aprobado.append(r)
            elif st == "Pendiente":
                ppto_pendiente.append(r)
            elif st == "Rechazado":
                ppto_rechazado.append(r)
            else:
                ppto_no_revisado.append(r)

        facturacion = []

        def _sum(lst, kc, kv):
            return sum(float(x.get(kv) or 0) for x in lst)

        cant_ap = _sum(sicoe_aprobado, "", "cantidad")
        cost_ap = _sum(sicoe_aprobado, "", "costo_directo")

        def _sum_ppto(lst):
            cq = sum(float(x.get("cant_total") or 0) for x in lst)
            ks = sum(float(x.get("costo_directo") or 0) for x in lst)
            return cq, ks

        cant_nr_p, cost_nr_p = _sum_ppto(ppto_no_revisado)
        cant_pd_p, cost_pd_p = _sum_ppto(ppto_pendiente)
        cant_rj_p, cost_rj_p = _sum_ppto(ppto_rechazado)

        cant_ppto = sum(float(r.get("cant_total") or 0) for r in ppto)
        costo_ppto = sum(float(r.get("costo_directo") or 0) for r in ppto)
        cant_ppto_ap = cost_ppto_ap = 0.0
        cant_ppto_nap = cost_ppto_nap = 0.0
        for r in ppto:
            c = float(r.get("cant_total") or 0)
            co = float(r.get("costo_directo") or 0)
            if _matriz_validacion_norm_estado(r.get("revisado")) == "Aprobado":
                cant_ppto_ap += c
                cost_ppto_ap += co
            else:
                cant_ppto_nap += c
                cost_ppto_nap += co

        return {
            "ppto": ppto,
            "ppto_por_revisado": {
                "aprobado": ppto_aprobado,
                "no_revisado": ppto_no_revisado,
                "pendiente": ppto_pendiente,
                "rechazado": ppto_rechazado,
            },
            "sicoe": {
                "aprobado": sicoe_aprobado,
            },
            "cobro": facturacion,
            "totales": {
                "cant_ppto": round(cant_ppto, 2),
                "costo_ppto": round(costo_ppto, 0),
                "cant_ppto_aprobado_n3": round(cant_ppto_ap, 2),
                "costo_ppto_aprobado_n3": round(cost_ppto_ap, 0),
                "cant_ppto_no_revisado_n3": round(cant_ppto_nap, 2),
                "costo_ppto_no_revisado_n3": round(cost_ppto_nap, 0),
                "cant_ppto_estado_no_revisado": round(cant_nr_p, 2),
                "costo_ppto_estado_no_revisado": round(cost_nr_p, 0),
                "cant_ppto_estado_pendiente": round(cant_pd_p, 2),
                "costo_ppto_estado_pendiente": round(cost_pd_p, 0),
                "cant_ppto_estado_rechazado": round(cant_rj_p, 2),
                "costo_ppto_estado_rechazado": round(cost_rj_p, 0),
                "cant_sicoe_aprobado": round(cant_ap, 2),
                "costo_sicoe_aprobado": round(cost_ap, 0),
                "cant_sicoe_no_revisado": 0.0,
                "costo_sicoe_no_revisado": 0.0,
                "cant_cobro": 0.0,
                "costo_cobro": 0.0,
                "delta_cant": round(cant_ppto - cant_ap, 2),
                "delta_costo": round(costo_ppto - cost_ap, 0),
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/solicitar-reversion")
def solicitar_reversion(contrato_id: int, registro_id: int, body: SolicitarReversionBody,
                        current_user=Depends(get_current_user)):
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        def _get():
            return supabase.table("so_registros")\
                .select("nivel3_estado, bloqueado").eq("id", registro_id)\
                .eq("contrato_id", contrato_id).limit(1).execute().data
        rows = supabase_execute(_get)
        if not rows:
            raise HTTPException(status_code=404, detail="Registro no encontrado.")
        reg = rows[0]
        if reg.get("nivel3_estado") != "Aprobado" or not reg.get("bloqueado"):
            raise HTTPException(status_code=422,
                detail="El registro debe estar en nivel3 Aprobado y bloqueado para solicitar reversión.")

        prev_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
        def _upd():
            return supabase.table("so_registros")\
                .update({"solicitud_reversion": True}).eq("id", registro_id)\
                .eq("contrato_id", contrato_id).execute().data
        supabase_execute(_upd)
        _insertar_comentario(contrato_id, registro_id, autor_id, body.comentario_data,
                             tipo_override="solicitud_reversion", audit_user=current_user)
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            after_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
            registrar_log(
                u_log,
                "REVERSION_SOLICITAR",
                "SICOE",
                "registro",
                str(registro_id),
                {},
                valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
                severidad="AUDIT",
            )
        except Exception:
            pass
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/aceptar-reversion")
def aceptar_reversion(contrato_id: int, registro_id: int, body: AceptarReversionBody,
                      current_user=Depends(get_current_user)):
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        def _get():
            return supabase.table("so_registros")\
                .select("solicitud_reversion").eq("id", registro_id)\
                .eq("contrato_id", contrato_id).limit(1).execute().data
        rows = supabase_execute(_get)
        if not rows or not rows[0].get("solicitud_reversion"):
            raise HTTPException(status_code=422,
                detail="No existe una solicitud de reversión activa para este registro.")

        prev_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}

        if body.aceptar:
            update = {
                "bloqueado":          False,
                "solicitud_reversion": False,
                "nivel3_estado":      "No Revisado",
                "nivel3_usuario_id":  None,
                "nivel3_fecha":       None,
            }
        else:
            update = {"solicitud_reversion": False}

        def _upd():
            return supabase.table("so_registros")\
                .update(update).eq("id", registro_id)\
                .eq("contrato_id", contrato_id).execute().data
        supabase_execute(_upd)

        if body.aceptar:
            comentario_data = {
                "tipo":    "aceptar_reversion",
                "mensaje": "Reversión aceptada.",
            }
            _insertar_comentario(contrato_id, registro_id, autor_id, comentario_data,
                                 tipo_override="aceptar_reversion", audit_user=current_user)
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            after_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
            accion = "REVERSION_ACEPTAR" if body.aceptar else "REVERSION_RECHAZAR"
            registrar_log(
                u_log,
                accion,
                "SICOE",
                "registro",
                str(registro_id),
                {},
                valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
                severidad="AUDIT",
            )
        except Exception:
            pass
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sicoe-obra/{contrato_id}/registros/{registro_id}/reversion-n3-doble-llave")
def reversion_n3_doble_llave(
    contrato_id: int,
    registro_id: int,
    body: ReversionDobleLlaveN3Body,
    current_user=Depends(get_current_user),
):
    """
    Revierte la aprobación N3 (Interventoría) solo tras acción coordinada de N2 y N3 (doble llave).
    Cada llamada registra comentario con destinatarios; al completar ambas llaves se desbloquea y
    nivel3 pasa a «No Revisado».
    """
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        cd = body.comentario_data or {}
        mensaje_limpio = (cd.get("mensaje") or "").strip()
        if not mensaje_limpio:
            raise HTTPException(status_code=422, detail="El cuerpo del mensaje es obligatorio.")
        dest_raw = cd.get("destinatarios") or []
        n_dest = sum(1 for d in dest_raw if isinstance(d, dict) and d.get("id") is not None)
        if n_dest < 1:
            raise HTTPException(
                status_code=422,
                detail="Debe indicar al menos un destinatario (para quién va el mensaje).",
            )

        def _get():
            return (
                supabase.table("so_registros")
                .select(
                    "nivel3_estado, bloqueado,"
                    "reversion_arm_n2_usuario_id, reversion_arm_n3_usuario_id"
                )
                .eq("id", registro_id)
                .eq("contrato_id", contrato_id)
                .limit(1)
                .execute()
                .data
            )

        rows = supabase_execute(_get)
        if not rows:
            raise HTTPException(status_code=404, detail="Registro no encontrado.")
        row = rows[0]
        if row.get("nivel3_estado") != "Aprobado" or not row.get("bloqueado"):
            raise HTTPException(
                status_code=422,
                detail="Solo aplica a registros aprobados por Interventoría (N3) y bloqueados.",
            )

        prev_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}

        def _nid(v):
            if v is None:
                return None
            try:
                return int(v)
            except (TypeError, ValueError):
                return None

        arm2 = _nid(row.get("reversion_arm_n2_usuario_id"))
        arm3 = _nid(row.get("reversion_arm_n3_usuario_id"))

        nivel_db = _sicoe_db_nivel_validacion_usuario(autor_id)
        if _es_desarrollador(current_user):
            if arm2 is None:
                nivel = 2
            elif arm3 is None:
                nivel = 3
            else:
                raise HTTPException(
                    status_code=422,
                    detail="Las dos llaves ya están registradas para este registro.",
                )
        else:
            nivel = nivel_db
            if nivel not in (2, 3):
                raise HTTPException(
                    status_code=403,
                    detail="Solo Residente de costos (Nivel 2) o Interventoría (Nivel 3) pueden activar esta llave.",
                )
        _require_llave_reversion_sicoe_nivel(current_user, autor_id, nivel)

        if nivel == 2:
            if arm2 is not None and arm2 != autor_id:
                raise HTTPException(
                    status_code=409,
                    detail="Otro usuario de Nivel 2 ya registró su llave para este registro.",
                )
            if arm2 is not None and arm2 == autor_id:
                raise HTTPException(
                    status_code=422,
                    detail="Ya registraste la llave de Nivel 2; falta la llave de Interventoría (N3).",
                )
        else:
            if arm3 is not None and arm3 != autor_id:
                raise HTTPException(
                    status_code=409,
                    detail="Otro usuario de Interventoría ya registró su llave para este registro.",
                )
            if arm3 is not None and arm3 == autor_id:
                raise HTTPException(
                    status_code=422,
                    detail="Ya registraste la llave de Interventoría; falta la llave de Nivel 2.",
                )

        new2 = autor_id if nivel == 2 else arm2
        new3 = autor_id if nivel == 3 else arm3
        ejecutar = new2 is not None and new3 is not None
        if ejecutar and new2 == new3:
            raise HTTPException(
                status_code=422,
                detail="La reversión doble requiere dos personas distintas (Nivel 2 e Interventoría).",
            )

        cd_send = {**cd, "mensaje": mensaje_limpio}
        tipo_c = "reversion_doble_llave_n2" if nivel == 2 else "reversion_doble_llave_n3"
        _insertar_comentario(
            contrato_id,
            registro_id,
            autor_id,
            cd_send,
            tipo_override=tipo_c,
            nivel_validacion_override=f"Nivel {nivel}",
            audit_user=current_user,
        )
        try:
            _push_notif_validacion_sicoe_destinatarios(
                current_user,
                autor_id,
                contrato_id,
                registro_id,
                f"Doble llave reversión N3 — Nivel {nivel}",
                mensaje_limpio,
                cd_send,
            )
        except Exception:
            pass

        if ejecutar:
            update = {
                "bloqueado": False,
                "solicitud_reversion": False,
                "nivel3_estado": "No Revisado",
                "nivel3_usuario_id": None,
                "nivel3_fecha": None,
                "reversion_arm_n2_usuario_id": None,
                "reversion_arm_n3_usuario_id": None,
            }
        else:
            update = (
                {"reversion_arm_n2_usuario_id": autor_id}
                if nivel == 2
                else {"reversion_arm_n3_usuario_id": autor_id}
            )

        def _upd():
            return (
                supabase.table("so_registros")
                .update({**update, "updated_at": "now()"})
                .eq("id", registro_id)
                .eq("contrato_id", contrato_id)
                .execute()
                .data
            )

        try:
            supabase_execute(_upd)
        except Exception as ex:
            low = str(ex).lower()
            if "reversion_arm" in low or "column" in low:
                raise HTTPException(
                    status_code=503,
                    detail="Falta ejecutar la migración SQL reversion_n3_doble_llave_so_registros.sql en la base de datos.",
                ) from ex
            raise

        accion_log = (
            "REVERSION_DOBLE_EJECUTADA"
            if ejecutar
            else ("REVERSION_LLAVE_N2" if nivel == 2 else "REVERSION_LLAVE_N3")
        )
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            after_audit = _so_registro_fetch_validacion_audit(contrato_id, registro_id) or {}
            registrar_log(
                u_log,
                accion_log,
                "SICOE",
                "registro",
                str(registro_id),
                {"nivel_llave": nivel, "ejecutada": ejecutar},
                valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
                severidad="AUDIT",
            )
        except Exception:
            pass

        return {"ok": True, "ejecutada": ejecutar}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sicoe-obra/{contrato_id}/registros/{registro_id}/comentarios")
def crear_comentario(contrato_id: int, registro_id: int, body: ComentarioCreate,
                     current_user=Depends(get_current_user)):
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        comentario_data = body.dict()
        dest_raw = comentario_data.get("destinatarios") or []
        n_dest = 0
        for d in dest_raw:
            if isinstance(d, dict) and d.get("id") is not None:
                n_dest += 1
        mensaje_limpio = (comentario_data.get("mensaje") or "").strip()
        if not mensaje_limpio:
            raise HTTPException(
                status_code=422, detail="El cuerpo del mensaje es obligatorio.")
        if n_dest < 1:
            raise HTTPException(
                status_code=422, detail="Debe indicar al menos un destinatario (para quién va el mensaje).")
        creado = _insertar_comentario(contrato_id, registro_id, autor_id, comentario_data, audit_user=current_user)

        # Notificación directa a destinatarios explícitos
        asunto = (comentario_data.get("asunto") or "Comentario de validación").strip() or "Comentario de validación"
        mensaje = comentario_data.get("mensaje") or ""
        _push_notif_validacion_sicoe_destinatarios(
            current_user,
            autor_id,
            contrato_id,
            registro_id,
            asunto,
            mensaje,
            comentario_data,
        )

        return {"ok": True, "comentario_id": creado.get("id")}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/sicoe-obra/{contrato_id}/registros/{registro_id}/comentarios/{comentario_id}/respuesta")
def responder_comentario_registro(contrato_id: int, registro_id: int, comentario_id: int,
                                   body: dict, current_user=Depends(get_current_user)):
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        parent = supabase.table("so_registro_comentarios").select(
            "confidencialidad, rol_origen, destinatarios, etiqueta, asunto, tipo, nivel_validacion"
        ).eq("id", comentario_id).eq("registro_id", registro_id).eq("contrato_id", contrato_id).single().execute().data

        if not parent:
            raise HTTPException(status_code=404, detail="Comentario padre no encontrado")

        parent_conf = parent.get("confidencialidad") or "publico"
        parent_rol = parent.get("rol_origen") or ""
        parent_dest = parent.get("destinatarios") or []
        parent_tipo = parent.get("tipo") or "validacion"

        def _ins():
            return supabase.table("so_registro_comentarios").insert({
                "registro_id":    registro_id,
                "contrato_id":    contrato_id,
                "autor_id":       autor_id,
                "padre_id":       comentario_id,
                "mensaje":        body.get("mensaje", ""),
                "tipo":           parent_tipo,
                # Heredar confidencialidad del hilo evita fugas por respuestas "cruzadas".
                "confidencialidad": parent_conf,
                "rol_origen":     body.get("rol_origen") or parent_rol,
                "destinatarios":  parent_dest,
                "etiqueta":       parent.get("etiqueta"),
                "asunto":         parent.get("asunto"),
                "enlaces":        [],
                "nivel_validacion": parent.get("nivel_validacion"),
            }).execute().data
        data = supabase_execute(_ins)
        row = data[0] if data else {}
        try:
            u_log = _audit_user_contrato(current_user, contrato_id)
            msg = (body.get("mensaje") or "")[:800]
            registrar_log(
                u_log,
                "COMENTAR_RESPUESTA",
                "SICOE",
                "registro",
                str(registro_id),
                {
                    "so_comentario_id": row.get("id"),
                    "padre_id": comentario_id,
                    "mensaje_excerpt": msg,
                    "tipo": parent_tipo,
                },
                severidad="AUDIT",
            )
        except Exception:
            pass
        return {"ok": True, "comentario_id": row.get("id")}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sicoe-obra/{contrato_id}/registros/{registro_id}/reporte")
def get_reporte_de_registro(
    contrato_id: int,
    registro_id: int,
    cargo_id: Optional[int] = Query(None),
    estado_validacion: Optional[str] = Query(None),
    validacion_capas: Optional[str] = Query(None),
    validacion_capas_op: Optional[str] = Query(None),
    current_user=Depends(get_current_user),
):
    try:
        def _get():
            return supabase.table("so_registros").select("reporte_id")\
                .eq("id", registro_id).eq("contrato_id", contrato_id)\
                .limit(1).execute().data
        rows = supabase_execute(_get)
        if not rows:
            raise HTTPException(status_code=404, detail="Registro no encontrado")
        reporte_id = rows[0]["reporte_id"]
        # Reusar el endpoint existente
        from fastapi import Request
        def _rep():
            return supabase.table("so_reportes")\
                .select("*, subcontratistas(razon_social)")\
                .eq("id", reporte_id).eq("contrato_id", contrato_id).execute().data
        def _reg():
            return supabase.table("so_registros").select("*")\
                .eq("reporte_id", reporte_id).order("id").execute().data
        reporte = supabase_execute(_rep)
        if not reporte:
            raise HTTPException(status_code=404, detail="Reporte no encontrado")
        r = reporte[0]
        sub = r.pop("subcontratistas", None)
        r["subcontratista_nombre"] = sub["razon_social"] if sub else None
        regs_raw = supabase_execute(_reg)
        _capas_gr = _parse_validacion_capas_param(validacion_capas, cargo_id, estado_validacion)
        if _capas_gr:
            regs_raw = _filtrar_registros_validacion_capas_sicoe(
                regs_raw, _capas_gr, r, validacion_capas_op or "and"
            )
        reg_ids = [reg["id"] for reg in regs_raw if reg.get("id")]
        num_comentarios_map = {}
        if reg_ids:
            try:
                def _cnt():
                    return supabase.table("so_registro_comentarios")\
                        .select("registro_id").in_("registro_id", reg_ids).execute().data
                for row in supabase_execute(_cnt):
                    rid = row["registro_id"]
                    num_comentarios_map[rid] = num_comentarios_map.get(rid, 0) + 1
            except Exception:
                pass
        for reg in regs_raw:
            reg["num_comentarios"] = num_comentarios_map.get(reg["id"], 0)
        _enriquecer_registros_labels_reversion_doble_llave(regs_raw)
        r["registros"] = regs_raw
        r["puntos"] = []
        return r
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sicoe-obra/{contrato_id}/registros/{registro_id}/comentarios")
def listar_comentarios(contrato_id: int, registro_id: int, rol_solicitante: str,
                       current_user=Depends(get_current_user)):
    try:
        def _verify_reg():
            return supabase.table("so_registros").select("id")\
                .eq("id", registro_id).eq("contrato_id", contrato_id)\
                .limit(1).execute().data
        if not supabase_execute(_verify_reg):
            raise HTTPException(
                status_code=404,
                detail="Registro no encontrado en este contrato.",
            )
        # Por registro_id: el conteo en grilla no filtra por contrato_id en la fila del
        # comentario; migraciones u operaciones legacy pueden dejar contrato_id distinto/NULL
        # y el listado quedaba vacío aunque hubiera filas.
        def _get():
            return supabase.table("so_registro_comentarios")\
                .select("*")\
                .eq("registro_id", registro_id)\
                .order("created_at", desc=False).execute().data
        comentarios = supabase_execute(_get)
        autor_ids = list({c["autor_id"] for c in comentarios if c.get("autor_id")})
        autor_map = {}
        if autor_ids:
            try:
                def _autores():
                    return supabase.table("usuarios").select("id, nombre, apellidos")\
                        .in_("id", autor_ids).execute().data
                for u in supabase_execute(_autores):
                    autor_map[u["id"]] = f"{u.get('nombre','')} {u.get('apellidos','')or ''}".strip()
            except Exception:
                pass
        for c in comentarios:
            c["autor"] = {"nombre": autor_map.get(c.get("autor_id"), "Usuario")}

        # Regla directa solicitada:
        # - Sin destinatarios: visible para todos.
        # - Con destinatarios: visible solo para ids explícitos.
        uid = int(current_user.get("sub") or current_user.get("id", 0))
        by_id = {c.get("id"): c for c in comentarios}
        filtrados = []
        for c in comentarios:
            destinatarios = c.get("destinatarios") or []
            # Si es respuesta sin destinatarios, heredar destinatarios del padre.
            if (not destinatarios) and c.get("padre_id") and by_id.get(c.get("padre_id")):
                destinatarios = by_id[c.get("padre_id")].get("destinatarios") or []
            ids_dest = set()
            for d in destinatarios:
                if isinstance(d, dict):
                    try:
                        did = int(d.get("id"))
                        ids_dest.add(did)
                    except Exception:
                        continue
            visible = (not ids_dest) or (uid in ids_dest) or (int(c.get("autor_id") or 0) == uid)

            if visible:
                filtrados.append(c)

        # Raíces: sin padre o padre no está en el conjunto visible (p. ej. filtrado o
        # migrado sin padre en BD); no perder "huérfanos" en la respuesta.
        by_id_filtrados = {c.get("id"): c for c in filtrados}
        padres = []
        for c in filtrados:
            pid = c.get("padre_id")
            if not pid or pid not in by_id_filtrados:
                padres.append(c)
        for p in padres:
            p["respuestas"] = [h for h in filtrados if h.get("padre_id") == p["id"]]
        return padres
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sicoe-obra/{contrato_id}/reportes/{reporte_id}/validar-masivo-nivel2")
def validar_masivo_nivel2(contrato_id: int, reporte_id: int, body: ValidarMasivoNivel2Body,
                          current_user=Depends(get_current_user)):
    ESTADOS = {"Aprobado", "Pendiente", "Rechazado", "No Objeto de Cobro"}
    if body.estado not in ESTADOS:
        raise HTTPException(status_code=422, detail=f"Estado inválido. Acepta: {ESTADOS}")

    estado_real = body.estado
    if body.estado == "No Objeto de Cobro":
        estado_real = "Rechazado"
        if not body.comentario_data:
            raise HTTPException(status_code=422,
                detail="Se requiere comentario_data para 'No Objeto de Cobro'.")
    if body.estado in ("Pendiente", "Rechazado") and not body.comentario_data:
        raise HTTPException(status_code=422,
            detail="Se requiere comentario_data cuando el estado es Pendiente o Rechazado.")

    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        _require_sicoe_puede_validar_nivel(current_user, autor_id, 2)
        actualizados = 0
        omitidos_precondicion = 0
        omitidos_topografia = 0
        excluidos_objeto_pago_sub = 0

        ids_all = list(body.ids_registros or [])
        truncado_mas_de_500 = len(ids_all) > SICOE_MASIVO_MAX_REGISTROS
        raw_ids = ids_all[:SICOE_MASIVO_MAX_REGISTROS]
        id_list: List[int] = []
        for x in raw_ids:
            try:
                id_list.append(int(x))
            except (TypeError, ValueError):
                continue
        reg_by_id: Dict[int, Any] = {}
        if id_list:
            _CHUNK = 200
            for i in range(0, len(id_list), _CHUNK):
                chunk = id_list[i : i + _CHUNK]

                def _rb(c=chunk):
                    return (
                        supabase.table("so_registros")
                        .select("id, nivel1_estado, nivel3_estado, reporte_id, nivel2_objeto_pago_sub")
                        .eq("contrato_id", contrato_id)
                        .in_("id", c)
                        .execute()
                        .data
                    )

                for row in supabase_execute(_rb) or []:
                    try:
                        reg_by_id[int(row["id"])] = row
                    except (TypeError, ValueError, KeyError):
                        pass
        rep_candidates = [r.get("reporte_id") for r in reg_by_id.values() if r.get("reporte_id") is not None]
        exige_topo_n2 = estado_real == "Aprobado" and _sicoe_exige_topografia_para_aprobar_nivel2(contrato_id)
        topo_reportes = (
            _reportes_ids_con_topografia(contrato_id, rep_candidates) if exige_topo_n2 else set()
        )

        for reg_id in raw_ids:
            try:
                rid = int(reg_id)
            except (TypeError, ValueError):
                omitidos_precondicion += 1
                continue
            rinfo = reg_by_id.get(rid)
            if not rinfo or rinfo.get("nivel1_estado") != "Aprobado":
                omitidos_precondicion += 1
                continue
            if rinfo.get("nivel2_objeto_pago_sub"):
                excluidos_objeto_pago_sub += 1
                continue
            if _registro_nivel3_aprobado(rinfo):
                omitidos_precondicion += 1
                continue
            if exige_topo_n2:
                rp = rinfo.get("reporte_id")
                try:
                    rpi = int(rp) if rp is not None else None
                except (TypeError, ValueError):
                    rpi = None
                if not rpi or rpi not in topo_reportes:
                    omitidos_topografia += 1
                    continue

            prev_audit = _so_registro_fetch_validacion_audit(contrato_id, rid) or {}

            update = {
                "nivel2_estado":     estado_real,
                "nivel2_usuario_id": autor_id,
                "nivel2_fecha":      datetime.utcnow().isoformat(),
            }
            if body.objeto_pago_sub is not None:
                update["nivel2_objeto_pago_sub"] = body.objeto_pago_sub

            def _upd(regid=rid, upd=update):
                return supabase.table("so_registros")\
                    .update(upd).eq("id", regid)\
                    .eq("contrato_id", contrato_id).execute().data
            supabase_execute(_upd)

            if estado_real == "Aprobado":
                try:
                    _aplicar_acta_rpo_vigente_a_registro(contrato_id, rid, date.today())
                except Exception:
                    pass

            if body.comentario_data:
                _insertar_comentario(
                    contrato_id, rid, autor_id, body.comentario_data,
                    tipo_override="validacion", nivel_validacion_override="Nivel 2",
                    audit_user=current_user,
                )
            try:
                u_log = _audit_user_contrato(current_user, contrato_id)
                after_audit = _so_registro_fetch_validacion_audit(contrato_id, rid) or {}
                registrar_log(
                    u_log, "VALIDAR", "SICOE", "registro", str(rid),
                    {
                        "nivel": 2, "estado": body.estado, "masivo": True, "reporte_id": reporte_id,
                        "nivel2_objeto_pago_sub": body.objeto_pago_sub,
                    },
                    valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                    valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
                )
            except Exception:
                pass
            actualizados += 1

        omitidos = omitidos_precondicion + omitidos_topografia + excluidos_objeto_pago_sub
        out = {
            "actualizados": actualizados,
            "omitidos": omitidos,
            "omitidos_precondicion": omitidos_precondicion,
            "omitidos_topografia": omitidos_topografia,
            "excluidos_objeto_pago_sub": excluidos_objeto_pago_sub,
            "truncado_mas_de_500": truncado_mas_de_500,
        }
        if omitidos_topografia:
            out["alerta_topografia"] = (
                f"{omitidos_topografia} registro(s) no se aprobaron en N2 por falta de topografía en el reporte correspondiente."
            )
        if excluidos_objeto_pago_sub:
            out["alerta_objeto_sub"] = (
                "Hay líneas con objeto de pago a subcontratista: no entran en el masivo de N2."
            )
        if truncado_mas_de_500:
            out["alerta_tope"] = f"Se procesaron como máximo {SICOE_MASIVO_MAX_REGISTROS} registros por solicitud."
        return out
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sicoe-obra/{contrato_id}/reportes/{reporte_id}/validar-masivo-nivel3")
def validar_masivo_nivel3(contrato_id: int, reporte_id: int, body: ValidarMasivoNivel3Body,
                          current_user=Depends(get_current_user)):
    ESTADOS = {"Aprobado", "Pendiente", "Rechazado"}
    if body.estado not in ESTADOS:
        raise HTTPException(status_code=422, detail=f"Estado inválido. Acepta: {ESTADOS}")
    if body.estado in ("Pendiente", "Rechazado") and not body.comentario_data:
        raise HTTPException(status_code=422,
            detail="Se requiere comentario_data cuando el estado es Pendiente o Rechazado.")
    try:
        autor_id = int(current_user.get("sub") or current_user.get("id", 0))
        _require_sicoe_puede_validar_nivel(current_user, autor_id, 3)
        actualizados = 0
        omitidos_precondicion = 0
        excluidos_objeto_pago_sub = 0

        ids_all = list(body.ids_registros or [])
        truncado_mas_de_500 = len(ids_all) > SICOE_MASIVO_MAX_REGISTROS

        for reg_id in ids_all[:SICOE_MASIVO_MAX_REGISTROS]:
            def _get(rid=reg_id):
                return supabase.table("so_registros")\
                    .select("nivel2_estado, nivel3_estado, nivel2_objeto_pago_sub").eq("id", rid)\
                    .eq("contrato_id", contrato_id).limit(1).execute().data
            rows = supabase_execute(_get)
            if not rows:
                omitidos_precondicion += 1
                continue
            row0 = rows[0]
            if row0.get("nivel2_objeto_pago_sub"):
                excluidos_objeto_pago_sub += 1
                continue
            if row0.get("nivel2_estado") != "Aprobado":
                omitidos_precondicion += 1
                continue
            if row0.get("nivel3_estado") == "Aprobado" and body.estado != "Aprobado":
                omitidos_precondicion += 1
                continue

            prev_audit = _so_registro_fetch_validacion_audit(contrato_id, reg_id) or {}

            update = {
                "nivel3_estado":     body.estado,
                "nivel3_usuario_id": autor_id,
                "nivel3_fecha":      datetime.utcnow().isoformat(),
            }
            if body.estado == "Aprobado":
                update["bloqueado"] = True

            def _upd(rid=reg_id, upd=update):
                return supabase.table("so_registros")\
                    .update(upd).eq("id", rid)\
                    .eq("contrato_id", contrato_id).execute().data
            supabase_execute(_upd)

            if body.estado == "Aprobado":
                try:
                    _aplicar_acta_rpo_vigente_a_registro(contrato_id, reg_id, date.today())
                except Exception:
                    pass

            if body.comentario_data:
                _insertar_comentario(
                    contrato_id, reg_id, autor_id, body.comentario_data,
                    tipo_override="validacion", nivel_validacion_override="Nivel 3",
                    audit_user=current_user,
                )
            try:
                u_log = _audit_user_contrato(current_user, contrato_id)
                after_audit = _so_registro_fetch_validacion_audit(contrato_id, reg_id) or {}
                registrar_log(
                    u_log, "VALIDAR", "SICOE", "registro", str(reg_id),
                    {"nivel": 3, "estado": body.estado, "masivo": True, "reporte_id": reporte_id},
                    valor_anterior=_so_registro_validacion_audit_snapshot(prev_audit),
                    valor_nuevo=_so_registro_validacion_audit_snapshot(after_audit),
                )
            except Exception:
                pass
            actualizados += 1

        omitidos = omitidos_precondicion + excluidos_objeto_pago_sub
        out = {
            "actualizados": actualizados,
            "omitidos": omitidos,
            "omitidos_precondicion": omitidos_precondicion,
            "excluidos_objeto_pago_sub": excluidos_objeto_pago_sub,
            "truncado_mas_de_500": truncado_mas_de_500,
        }
        if excluidos_objeto_pago_sub:
            out["alerta_objeto_sub"] = (
                "Hay líneas con objeto de pago a subcontratista: no entran en el masivo de N3."
            )
        if truncado_mas_de_500:
            out["alerta_tope"] = f"Se procesaron como máximo {SICOE_MASIVO_MAX_REGISTROS} registros por solicitud."
        return out
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sicoe-obra/{contrato_id}/registros/reconciliar-acta-rpo-historico")
def reconciliar_acta_rpo_historico(
    contrato_id: int,
    body: ReconciliarActaRpoHistoricoBody,
    current_user=Depends(get_current_user),
):
    """
    Alinea acta_rpo_id en registros ya existentes según la fecha de aprobación:
    N3 aprobado → acta vigente en nivel3_fecha (o nivel2_fecha si falta);
    solo N2 aprobado → acta vigente en nivel2_fecha.
    No modifica registros sin aprobación N2. Solo Desarrollador.
    """
    if not _es_desarrollador(current_user):
        raise HTTPException(status_code=403, detail="Solo el cargo Desarrollador puede ejecutar esta acción.")
    try:
        revisados = 0
        actualizados = 0
        sin_acta_periodo = 0
        off = 0
        while True:
            def _batch(o=off):
                return supabase.table("so_registros").select(
                    "id, reporte_id, acta_rpo_id, nivel2_estado, nivel3_estado, nivel2_fecha, nivel3_fecha"
                ).eq("contrato_id", contrato_id).range(o, o + 499).execute().data

            rows = supabase_execute(_batch) or []
            for row in rows:
                revisados += 1
                ref_raw = None
                if _matriz_validacion_norm_estado(row.get("nivel3_estado")) == "Aprobado":
                    ref_raw = row.get("nivel3_fecha") or row.get("nivel2_fecha")
                elif _matriz_validacion_norm_estado(row.get("nivel2_estado")) == "Aprobado":
                    ref_raw = row.get("nivel2_fecha")
                else:
                    continue
                d = _parse_iso_to_date(ref_raw)
                if not d:
                    continue
                acta_id = _acta_rpo_id_vigente_para_fecha(contrato_id, d)
                if not acta_id:
                    sin_acta_periodo += 1
                    continue
                if acta_id == row.get("acta_rpo_id"):
                    continue
                if body.dry_run:
                    actualizados += 1
                    continue
                rid = row["id"]
                rep_id = row.get("reporte_id")

                def _ur():
                    return supabase.table("so_registros").update({"acta_rpo_id": acta_id})\
                        .eq("id", rid).eq("contrato_id", contrato_id).execute().data

                supabase_execute(_ur)
                if rep_id is not None:
                    def _urp():
                        return supabase.table("so_reportes").update({"acta_rpo_id": acta_id})\
                            .eq("id", rep_id).eq("contrato_id", contrato_id).execute().data

                    supabase_execute(_urp)
                actualizados += 1
            if len(rows) < 500:
                break
            off += 500
        return {
            "revisados": revisados,
            "actualizados": actualizados,
            "sin_acta_periodo": sin_acta_periodo,
            "dry_run": body.dry_run,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

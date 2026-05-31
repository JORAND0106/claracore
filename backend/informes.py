import io
import json
import logging
import html
import hashlib
import base64
import difflib
import math
import os as _os
import re
import threading
import time
from datetime import datetime, date
import pytz
from concurrent.futures import ThreadPoolExecutor, as_completed
from zoneinfo import ZoneInfo
from typing import Any, Dict, List, Literal, Optional, Tuple

_log = logging.getLogger("uvicorn.error")

_GERENCIA_MATRIZ_CACHE: Dict[tuple, Dict[str, Any]] = {}
_GERENCIA_MATRIZ_CACHE_LOCK = threading.Lock()
_GERENCIA_MATRIZ_CACHE_TTL_SEC = 600
_GERENCIA_MATRIZ_CACHE_STALE_SEC = 1800
_GERENCIA_PDF_CACHE: Dict[tuple, Dict[str, Any]] = {}
_GERENCIA_PDF_CACHE_LOCK = threading.Lock()
_GERENCIA_PDF_CACHE_TTL_SEC = 600


def _gerencia_matriz_cache_key(contrato_id: int, acta_presente_override: Optional[int]) -> tuple:
    ap = int(acta_presente_override) if acta_presente_override is not None else None
    # v5: col.2/3 = nivel máximo activo del contrato (no solo N3 fijo)
    return ("gerencia_matriz", 5, int(contrato_id), ap)


def _gerencia_matriz_cache_get(key: tuple) -> Optional[Dict[str, Any]]:
    now = time.time()
    with _GERENCIA_MATRIZ_CACHE_LOCK:
        entry = _GERENCIA_MATRIZ_CACHE.get(key)
        if not entry:
            return None
        age = now - entry["ts"]
        if age <= _GERENCIA_MATRIZ_CACHE_STALE_SEC:
            return entry["data"]
        _GERENCIA_MATRIZ_CACHE.pop(key, None)
    return None


def _gerencia_matriz_cache_set(key: tuple, data: Dict[str, Any]) -> None:
    with _GERENCIA_MATRIZ_CACHE_LOCK:
        _GERENCIA_MATRIZ_CACHE[key] = {"data": data, "ts": time.time()}
        if len(_GERENCIA_MATRIZ_CACHE) > 32:
            oldest = min(_GERENCIA_MATRIZ_CACHE, key=lambda k: _GERENCIA_MATRIZ_CACHE[k]["ts"])
            _GERENCIA_MATRIZ_CACHE.pop(oldest, None)


def _gerencia_pdf_cache_key(contrato_id: int, acta_presente: Optional[int], modo: str, con_sello: bool) -> tuple:
    ap = int(acta_presente) if acta_presente is not None else -1
    return ("gerencia_pdf", 4, int(contrato_id), ap, (modo or "matriz").lower().strip(), bool(con_sello))


def _gerencia_pdf_cache_get(key: tuple) -> Optional[Tuple[bytes, str]]:
    now = time.time()
    with _GERENCIA_PDF_CACHE_LOCK:
        entry = _GERENCIA_PDF_CACHE.get(key)
        if not entry:
            return None
        if now - entry["ts"] <= _GERENCIA_PDF_CACHE_TTL_SEC:
            return entry["data"], str(entry.get("nr") or "ger")
        _GERENCIA_PDF_CACHE.pop(key, None)
    return None


def _gerencia_pdf_cache_set(key: tuple, data: bytes, nr: str = "ger") -> None:
    with _GERENCIA_PDF_CACHE_LOCK:
        _GERENCIA_PDF_CACHE[key] = {"data": data, "ts": time.time(), "nr": nr}
        if len(_GERENCIA_PDF_CACHE) > 16:
            oldest = min(_GERENCIA_PDF_CACHE, key=lambda k: _GERENCIA_PDF_CACHE[k]["ts"])
            _GERENCIA_PDF_CACHE.pop(oldest, None)


def _gerencia_caches_clear() -> None:
    with _GERENCIA_MATRIZ_CACHE_LOCK:
        _GERENCIA_MATRIZ_CACHE.clear()
    with _GERENCIA_PDF_CACHE_LOCK:
        _GERENCIA_PDF_CACHE.clear()


from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from xhtml2pdf import pisa
from main import get_current_user as _get_user
from main import get_current_user_optional as _get_user_optional
from mail_smtp import try_send_text_email
from supabase import create_client as _create_client
from ccd_conciliacion import (
    _bloque_capitulo_matriz,
    _nivel_norm_matriz as _norm_estado_n3,
    _orden_titulo_capitulo_obra,
    _registro_aprobado_matriz_panel,
    aggregate_items_conciliacion,
    fetch_registros_acta_todas_sico_obra,
    fetch_registros_conciliacion,
    fetch_registros_informe_cc_mes_por_acta,
    fetch_registros_memoria_cc_mes_alineado_acta,
    fetch_registros_memoria_conciliacion,
    informe_gerencia_matriz_maps_por_rpc,
    _norm_cap_informe_gerencia,
    rpo_conciliacion_por_contrato,
    rpo_conciliacion_un_acta_rpc,
    rpo_resumen_actas_rpc,
    registro_tiene_pendiente_matriz,
    suma_por_capitulo_desde_registros,
    suma_por_capitulo_solo_cdirecto_almacenado,
    suma_por_capitulo_interventoria_sellada,
    _fetch_cascade_interventoria_actas_rpo,
    matriz_params_contrato,
)

# ClaraCore Documentación (CCD): código único por tipo de formato (gestión documental).
CODIGO_FORMATO_CCD_CC_SUB_001 = "CC-SUB-001"
CODIGO_FORMATO_CCD_CC_SUB_002 = "CC-SUB-002"
CODIGO_FORMATO_CCD_CC_SEM_001 = "CC-SEM-001"
CODIGO_FORMATO_CCD_CC_SEM_002 = "CC-SEM-002"
CODIGO_FORMATO_CCD_CC_MES_001 = "CC-MES-001"
CODIGO_FORMATO_CCD_CC_MES_002 = "CC-MES-002"
# Formato contractual entidad contratante (IDU — gestión documental ClaraCore).
CODIGO_FORMATO_IDU_FO_EO_04_V2 = "FO-IDU-EO-04-V2"
# Informe gerencia: comparativo de costo directo (cascada N1–N3) entre dos actas RPO; firmas CCD = acta «presente».
CODIGO_FORMATO_CCD_CC_GER_001 = "CC-GER-001"
_sb = _create_client(
    _os.getenv("SUPABASE_URL", ""),
    _os.getenv("SUPABASE_KEY", "")
)

router = APIRouter(tags=["informes"])


def _cargar_permisos_cargo_por_sub(uid: int) -> List[dict]:
    """
    El JWT de get_current_user no incluye la matriz de permisos (solo se envía en el body del login al cliente).
    Replica la carga de /auth/log-in para decidir acceso a Informes.
    """
    if not uid:
        return []
    try:
        urows = (
            _sb.table("usuarios")
            .select("cargo_id, subcontratista_id")
            .eq("id", uid)
            .limit(1)
            .execute()
            .data
            or []
        )
    except Exception as e0:
        _log.debug("informes perms: usuario: %s", e0)
        return []
    if not urows:
        return []
    urow = urows[0]
    cargo_id = urow.get("cargo_id")
    if not cargo_id:
        return []
    cargo_nombre = ""
    try:
        c = _sb.table("cargos").select("nombre").eq("id", int(cargo_id)).limit(1).execute().data
        if c:
            cargo_nombre = (c[0].get("nombre") or "").strip().lower()
    except Exception:
        pass
    if cargo_nombre == "subcontratista" and not urow.get("subcontratista_id"):
        return []
    try:
        permisos_raw = _sb.table("permisos").select("*").eq("cargo_id", cargo_id).execute().data or []
        funciones_rows = _sb.table("funciones").select("id, nombre").execute().data or []
    except Exception as e1:
        _log.debug("informes perms: matriz: %s", e1)
        return []
    fmap = {f["id"]: f["nombre"] for f in funciones_rows}
    out = [{**p, "funcion_nombre": fmap.get(p.get("funcion_id"), "")} for p in permisos_raw]
    if cargo_nombre == "desarrollador":
        out = [{**p, "exportar": True, "ver": True} for p in (out or [])]
    return out or []


def _perm_informes_ccd(user: Any, necesita: Literal["ver", "editar", "validar", "exportar"]) -> None:
    """
    Módulo Informes CCD / funciones: matriz «informes ccd» (ver, editar, validar, exportar).
    Desarrollador y administrador: acceso completo.
    """
    if user is None:
        raise HTTPException(401, "No autenticado")
    try:
        u = user if isinstance(user, dict) else dict(user)
    except Exception:
        u = {}
    # Ver nota en _cargar_permisos_cargo_por_sub: el token JWT no trae "permisos".
    pl = u.get("permisos")
    if not (isinstance(pl, list) and len(pl) > 0):
        try:
            uid = int(str(u.get("sub") or "0").strip() or 0)
        except (TypeError, ValueError):
            uid = 0
        if uid:
            u = {**u, "permisos": _cargar_permisos_cargo_por_sub(uid)}
    cn = (u.get("cargo_nombre") or "").strip().lower()
    if cn in ("desarrollador", "administrador"):
        return
    for p in (u.get("permisos") or []):
        if (p.get("funcion_nombre") or "").strip().lower() != "informes ccd":
            continue
        if necesita == "ver" and p.get("ver"):
            return
        if necesita == "editar" and p.get("editar"):
            return
        if necesita == "validar" and p.get("validar"):
            return
        if necesita == "exportar" and p.get("exportar"):
            return
    raise HTTPException(
        403,
        f"Sin permiso para esta acción en Informes (se requiere permiso «{necesita}» en la función Informes CCD).",
    )


def _perm_informes_ccd_ver_o_validar(user: Any) -> None:
    """Estado de firmas registradas: «ver» o «validar» (quien firma suele tener solo validar en la matriz)."""
    try:
        _perm_informes_ccd(user, "ver")
        return
    except HTTPException as ex:
        if ex.status_code != 403:
            raise
    _perm_informes_ccd(user, "validar")


# ── Biblioteca CCD (gestión documental): metadatos por código de formato ─────
# Más adelante el contrato podrá elegir qué códigos aplican; las plantillas siguen en código.
FORMATOS_CCD: Dict[str, Dict[str, Any]] = {
    CODIGO_FORMATO_CCD_CC_SUB_001: {
        "titulo": "Informe corte de subcontratista",
        "descripcion": "Preacta / corte de cantidades aprobadas por subcontratista",
        "plantilla_html": "cc_sub_001_v1_plain",
        "motor_pdf": "xhtml2pdf",
        "layout": {
            "encabezado_institucional_solo_primera_hoja": True,
            "tabla_items_continua_en_siguientes_hojas": True,
            "firmas_solo_ultima_hoja": True,
            "firmas_bloque_inferior": True,
        },
        # Misma estructura de slots para futuros formatos (Elaboró / Revisó configurables; Aprobó según reglas del formato).
        "slots_firma": [
            {"id": "elaboro", "label": "Elaboró", "origen": "configuracion"},
            {"id": "reviso", "label": "Revisó", "origen": "configuracion"},
            {"id": "aprobo", "label": "Aprobó", "origen": "subcontratista"},
        ],
        # Si False, perfiles de interventoría pueden ocultar el formato en la biblioteca (UI).
        "acceso_interventoria": False,
    },
    CODIGO_FORMATO_CCD_CC_SUB_002: {
        "titulo": "Memoria por ítem (corte subcontratista)",
        "descripcion": "Detalle de cantidades aprobadas y registro fotográfico por ítem",
        "plantilla_html": "memoria_item",
        "motor_pdf": "xhtml2pdf",
        "layout": {
            "orientacion": "landscape",
            "encabezado_institucional_solo_primera_hoja": True,
            "firmas_solo_ultima_hoja": True,
            "firmas_bloque_inferior": True,
        },
        "slots_firma": [
            {"id": "elaboro", "label": "Elaboró", "origen": "configuracion"},
            {"id": "reviso", "label": "Revisó", "origen": "configuracion"},
            {"id": "aprobo", "label": "Aprobó", "origen": "subcontratista"},
        ],
        "acceso_interventoria": False,
    },
    CODIGO_FORMATO_CCD_CC_SEM_001: {
        "titulo": "Informe ejecución semanal (conciliación interventoría–contratista)",
        "descripcion": "Cantidades nivel 3 aprobadas y bloqueadas, filtradas por semana de aprobación (so_semanas).",
        "plantilla_html": "cc_conc_sem_001_v1",
        "motor_pdf": "xhtml2pdf",
        "layout": {
            "encabezado_institucional_solo_primera_hoja": True,
            "tabla_items_continua_en_siguientes_hojas": True,
            "firmas_solo_ultima_hoja": True,
            "firmas_bloque_inferior": True,
        },
        "slots_firma": [
            {"id": "elaboro", "label": "Elaboró", "origen": "configuracion"},
            {"id": "reviso", "label": "Revisó", "origen": "configuracion"},
            {"id": "aprobo", "label": "Aprobó", "origen": "configuracion"},
        ],
        "acceso_interventoria": True,
    },
    CODIGO_FORMATO_CCD_CC_SEM_002: {
        "titulo": "Memoria por ítem — semanal (conciliación)",
        "descripcion": "Detalle y anexo fotográfico por ítem; mismo criterio de filtro que CC-SEM-001.",
        "plantilla_html": "memoria_item_conc_sem",
        "motor_pdf": "xhtml2pdf",
        "layout": {
            "orientacion": "landscape",
            "encabezado_institucional_solo_primera_hoja": True,
            "firmas_solo_ultima_hoja": True,
            "firmas_bloque_inferior": True,
        },
        "slots_firma": [
            {"id": "elaboro", "label": "Elaboró", "origen": "configuracion"},
            {"id": "reviso", "label": "Revisó", "origen": "configuracion"},
            {"id": "aprobo", "label": "Aprobó", "origen": "configuracion"},
        ],
        "acceso_interventoria": True,
    },
    CODIGO_FORMATO_CCD_CC_MES_001: {
        "titulo": "Informe ejecución mensual (conciliación interventoría–contratista)",
        "descripcion": "Cantidades nivel 3 aprobadas y bloqueadas, filtradas por acta RPO.",
        "plantilla_html": "cc_conc_mes_001_v1",
        "motor_pdf": "xhtml2pdf",
        "layout": {
            "encabezado_institucional_solo_primera_hoja": True,
            "tabla_items_continua_en_siguientes_hojas": True,
            "firmas_solo_ultima_hoja": True,
            "firmas_bloque_inferior": True,
        },
        "slots_firma": [
            {"id": "elaboro", "label": "Elaboró", "origen": "configuracion"},
            {"id": "reviso", "label": "Revisó", "origen": "configuracion"},
            {"id": "aprobo", "label": "Aprobó", "origen": "configuracion"},
        ],
        "acceso_interventoria": True,
    },
    CODIGO_FORMATO_CCD_CC_MES_002: {
        "titulo": "Memoria por ítem — mensual (conciliación)",
        "descripcion": "Detalle y anexo fotográfico por ítem; mismo criterio de filtro que CC-MES-001.",
        "plantilla_html": "memoria_item_conc_mes",
        "motor_pdf": "xhtml2pdf",
        "layout": {
            "orientacion": "landscape",
            "encabezado_institucional_solo_primera_hoja": True,
            "firmas_solo_ultima_hoja": True,
            "firmas_bloque_inferior": True,
        },
        "slots_firma": [
            {"id": "elaboro", "label": "Elaboró", "origen": "configuracion"},
            {"id": "reviso", "label": "Revisó", "origen": "configuracion"},
            {"id": "aprobo", "label": "Aprobó", "origen": "configuracion"},
        ],
        "acceso_interventoria": True,
    },
    CODIGO_FORMATO_CCD_CC_GER_001: {
        "titulo": "Informe de gerencia (avance ejecución de obra — comparativo por acta RPO)",
        "descripcion": (
            "Comparación de costo directo aprobado en cascada (SICOE Obra, N1·N2·N3) por capítulo, "
            "acta presente frente a acta de referencia. Firmas CCD vinculadas al acta presente."
        ),
        "plantilla_html": "informe_gerencia_v1",
        "motor_pdf": "xhtml2pdf",
        "layout": {
            "encabezado_institucional_solo_primera_hoja": True,
            "tabla_items_continua_en_siguientes_hojas": True,
            "firmas_solo_ultima_hoja": True,
            "firmas_bloque_inferior": True,
        },
        "slots_firma": [
            {"id": "elaboro", "label": "Elaboró", "origen": "configuracion"},
            {"id": "reviso", "label": "Revisó", "origen": "configuracion"},
            {"id": "aprobo", "label": "Aprobó", "origen": "configuracion"},
        ],
        "acceso_interventoria": True,
    },
    CODIGO_FORMATO_IDU_FO_EO_04_V2: {
        "titulo": "Memorias IDU FO-EO-04 V2.0",
        "descripcion": "Memoria de cálculo de cantidades de obra — Instituto de Desarrollo Urbano (IDU).",
        "plantilla_html": "idu_memoria_fo_eo_04_v2",
        "motor_pdf": "xhtml2pdf",
        "grupo_ccd": "entidades_externas",
        "layout": {
            "orientacion": "portrait",
            "encabezado_institucional_solo_primera_hoja": True,
            "firmas_solo_ultima_hoja": True,
            "firmas_bloque_inferior": True,
        },
        "slots_firma": [
            {"id": "elaboro", "label": "Elaboró", "origen": "configuracion"},
            {"id": "reviso", "label": "Revisó", "origen": "configuracion"},
        ],
        "acceso_interventoria": True,
    },
}


def _safe_filename_part(s: object) -> str:
    """Nombre de archivo seguro para Content-Disposition: solo ASCII (HTTP/latin-1).
    Antes se usaba \\w (Unicode); un acento en razón social rompía el envío de cabeceras → 500 genérico."""
    raw = str(s if s is not None else "").strip()
    t = re.sub(r"[^A-Za-z0-9._\-]+", "_", raw)
    return (t or "x")[:80]


def _natural_sort_key_cadena(s: object) -> Tuple[Any, ...]:
    """Orden natural por trozos numéricos vs texto: 1.10.1 < 1.10.2; NP-004 < NP-205; 4.01. < 4.26."""
    t = str(s if s is not None else "").strip()
    if not t:
        return (2, ())
    parts: List[Tuple[int, Any]] = []
    for part in re.split(r"(\d+)", t):
        if not part:
            continue
        if part.isdigit():
            parts.append((0, int(part)))
        else:
            parts.append((1, part.lower()))
    return (0, tuple(parts))


def _capitulo_sort_key_asc(s: object) -> Tuple[Any, ...]:
    """Primer entero al inicio del texto de capítulo (1.PRELIMIN… → 1; 16. ZONA… → 16). Sin número: antes que vacío."""
    t = str(s if s is not None else "").strip()
    if not t:
        return (2, 0, "")
    m = re.match(r"^\s*(\d+)", t)
    if m:
        try:
            return (0, int(m.group(1)), t.lower())
        except ValueError:
            pass
    return (1, t.lower())


def _capitulos_por_item_numero_desde_items(items: List[dict]) -> Dict[str, str]:
    """Un capítulo por código de ítem (prioriza el primero no vacío al fusionar)."""
    m: Dict[str, str] = {}
    for it in items or []:
        k = (it.get("item_numero") or "").strip()
        if not k:
            continue
        cap = str(it.get("capitulo") or "").strip()
        if k not in m:
            m[k] = cap
        elif cap and not m[k]:
            m[k] = cap
    return m


def _sort_identificadores_item_asc(
    identificadores: List[str],
    capitulos_por_item: Optional[Dict[str, str]] = None,
) -> List[str]:
    if capitulos_por_item:
        return sorted(
            identificadores,
            key=lambda k: (
                _capitulo_sort_key_asc((capitulos_por_item or {}).get(k, "")),
                _natural_sort_key_cadena(k),
            ),
        )
    return sorted(identificadores, key=_natural_sort_key_cadena)


def _sort_items_corte_por_item_numero_asc(items: List[dict]) -> None:
    """Orden in-place: primero por capítulo (número inicial), luego por código de ítem (orden natural)."""
    items.sort(
        key=lambda it: (
            _capitulo_sort_key_asc(it.get("capitulo")),
            _natural_sort_key_cadena(it.get("item_numero")),
        )
    )


def _nombre_archivo_cc_sub_001(corte: dict, sub: dict, corte_id: int) -> str:
    """Patrón CCD: {CODIGO}_Corte_{n}_{Subcontratista}.pdf (ASCII seguro)."""
    co = corte.get("consecutivo")
    if co is None or co == "":
        co = corte.get("id") if corte.get("id") is not None else corte_id
    try:
        num = f"{int(float(str(co))):02d}"
    except Exception:
        num = _safe_filename_part(str(co))[:12] or "00"
    sub_nom = _safe_filename_part((sub.get("razon_social") or "sub")[:56])
    base = f"{CODIGO_FORMATO_CCD_CC_SUB_001}_Corte_{num}_{sub_nom}.pdf"
    return _safe_filename_part(base.replace(".pdf", "")) + ".pdf"


def _parse_numero_contrato(numero: str) -> tuple:
    """Extrae (numero_corto, año) de strings como 'IDU-1551-2017' o 'ICCU-CTO-1614-2025'.

    Regla: el último segmento separado por '-' es el año y el penúltimo es el número.
    Ejemplos:
      'IDU-1551-2017'      → ('1551', '2017')
      'ICCU-CTO-1614-2025' → ('1614', '2025')
    Si no hay suficientes segmentos devuelve (numero_completo, '').
    """
    partes = (numero or "").strip().split("-")
    if len(partes) >= 2:
        return partes[-2].strip(), partes[-1].strip()
    return (numero or "").strip(), ""


def _logo_url_pdf_safe(url: str) -> str:
    """Convierte un data-URL de GIF a PNG (xhtml2pdf no renderiza GIF).

    Para URLs externas (http/https) o data-URLs PNG/JPEG devuelve el valor tal cual.
    Si la conversión falla, devuelve el URL original y se deja que xhtml2pdf lo intente.
    """
    if not url or not url.startswith("data:image/gif;base64,"):
        return url
    try:
        import base64, io
        from PIL import Image  # Pillow es dependencia de xhtml2pdf

        b64_data = url.split(",", 1)[1]
        img_bytes = base64.b64decode(b64_data)
        img = Image.open(io.BytesIO(img_bytes))
        img = img.convert("RGBA")  # preserva transparencia del GIF
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        png_b64 = base64.b64encode(buf.getvalue()).decode()
        return f"data:image/png;base64,{png_b64}"
    except Exception:
        return url  # fallback silencioso


def _html_logo_entidad(logo_url: Optional[str]) -> str:
    """Recuadro del logo de la entidad para el encabezado del FO-EO-04.

    Si no hay URL muestra un placeholder con borde punteado.
    Convierte GIF a PNG antes de embeber (limitación de xhtml2pdf).
    """
    if not logo_url or not str(logo_url).strip():
        return (
            '<div style="border:1px dashed #cbd5e1;min-height:50px;'
            'text-align:center;padding:6px 4px;font-size:5pt;color:#94a3b8;">'
            'Logo<br/>entidad</div>'
        )
    u = html.escape(_logo_url_pdf_safe(str(logo_url).strip()), quote=True)
    return (
        f'<img src="{u}" alt="" '
        'style="max-width:100%;max-height:44px;display:block;margin:0 auto;object-fit:contain;" />'
    )


def _html_logo_contratista(contrato: dict, compact: bool = False, compact_box_height: str = "") -> str:
    """Logo del contratista (URL en BD). xhtml2pdf intenta cargar la URL; si falla, queda placeholder.
    compact=True: CC-SUB-002; si compact_box_height (p.ej. \"1.28cm\") el logo llena ese rectángulo (object-fit: contain)."""
    url = contrato.get("logo_contratista")
    box = (compact_box_height or "").strip()
    if url is None or str(url).strip() == "":
        h = box if (compact and box) else ("40px" if compact else "84px")
        fs = "5.5pt" if compact else "6.5pt"
        # display:table + table-cell: xhtml2pdf maneja mejor que flex el centrado del placeholder.
        return (
            f'<div style="width:100%;height:{h};display:table;box-sizing:border-box;border:1px dashed #cbd5e1;">'
            f'<div style="display:table-cell;width:100%;height:{h};vertical-align:middle;text-align:center;'
            f'font-size:{fs};color:#94a3b8;">LOGO</div></div>'
        )
    u = html.escape(_logo_url_pdf_safe(str(url).strip()), quote=True)
    if compact and box:
        return (
            f'<div style="width:100%;height:{box};min-height:{box};box-sizing:border-box;padding:0;margin:0;'
            'display:table;table-layout:fixed;overflow:hidden;">'
            f'<div style="display:table-cell;width:100%;height:{box};vertical-align:middle;text-align:center;padding:0;">'
            f'<img src="{u}" alt="" style="max-width:100%;max-height:100%;width:auto;height:auto;display:inline-block;'
            'object-fit:contain;vertical-align:middle;margin:0;" />'
            "</div></div>"
        )
    if compact:
        return (
            f'<img src="{u}" alt="" style="max-width:118px;max-height:40px;display:block;margin:0 auto;'
            'object-fit:contain;" />'
        )
    return (
        f'<img src="{u}" alt="" style="max-width:132px;max-height:92px;display:block;margin:0 auto;'
        'object-fit:contain;" />'
    )


def _row(table: str, select: str, **eq: Any) -> Optional[Dict[str, Any]]:
    """Una fila o None; evita excepciones de PostgREST por `.single()` sin resultados."""
    q = _sb.table(table).select(select)
    for k, v in eq.items():
        q = q.eq(k, v)
    rows = q.limit(1).execute().data or []
    return rows[0] if rows else None


class CcdEstiloPdfBody(BaseModel):
    """Colores PDF por formato (hex #RRGGBB). Vacío = defaults del backend."""

    section_bar_bg: Optional[str] = None
    section_bar_text: Optional[str] = None
    thead_bg: Optional[str] = None
    row_even_bg: Optional[str] = None
    row_odd_bg: Optional[str] = None
    subtotal_bg: Optional[str] = None
    capitulo_subtotal_bg: Optional[str] = None
    # CC-GER-001 (informe de gerencia, matriz 4 columnas): filas resumidas
    ger_titulo_bloque_bg: Optional[str] = None
    ger_subtotal_obra_con_aiu_bg: Optional[str] = None
    ger_fila_tasa_aiu_bg: Optional[str] = None
    ger_cdirecto_mas_aiu_bg: Optional[str] = None
    ger_filas_post_cdu_bg: Optional[str] = None
    ger_vtot_obra_ajustes_bg: Optional[str] = None
    ger_subtotal_obra_con_iva_bg: Optional[str] = None
    ger_fila_tasa_iva_bg: Optional[str] = None
    ger_cdirecto_mas_iva_bg: Optional[str] = None
    ger_vtot_obra_iva_bg: Optional[str] = None
    ger_valor_total_acta_bg: Optional[str] = None


class CcdFirmaConfigBody(BaseModel):
    elaboro_nombre: Optional[str] = None
    elaboro_cargo: Optional[str] = None
    elaboro2_nombre: Optional[str] = None
    elaboro2_cargo: Optional[str] = None
    reviso_nombre: Optional[str] = None
    reviso_cargo: Optional[str] = None
    reviso2_nombre: Optional[str] = None
    reviso2_cargo: Optional[str] = None
    aprobo_nombre: Optional[str] = None
    aprobo_cargo: Optional[str] = None
    elaboro_usuario_id: Optional[int] = None
    elaboro2_usuario_id: Optional[int] = None
    reviso_usuario_id: Optional[int] = None
    reviso2_usuario_id: Optional[int] = None
    aprobo_usuario_id: Optional[int] = None
    estilo_pdf: Optional[CcdEstiloPdfBody] = None


def _sanitize_ccd_hex_color(val: object, default: str) -> str:
    """Acepta #RGB / #RRGGBB; si es inválido devuelve default."""
    d = (default or "#000000").strip()
    if val is None:
        return d
    s = str(val).strip()
    if not s:
        return d
    if not s.startswith("#"):
        s = "#" + s
    if re.match(r"^#[0-9A-Fa-f]{3}$", s):
        r, g, b = s[1], s[2], s[3]
        s = f"#{r}{r}{g}{g}{b}{b}"
    if re.match(r"^#[0-9A-Fa-f]{6}$", s):
        return s.lower()
    return d


def _default_estilo_pdf(formato_codigo: str) -> Dict[str, str]:
    """Valores por defecto alineados a las plantillas actuales (CC-SUB-001 / CC-SUB-002)."""
    if formato_codigo in (
        CODIGO_FORMATO_CCD_CC_SUB_001,
        CODIGO_FORMATO_CCD_CC_SEM_001,
        CODIGO_FORMATO_CCD_CC_MES_001,
    ):
        d: Dict[str, str] = {
            "section_bar_bg": "#e5e7eb",
            "section_bar_text": "#111827",
            "thead_bg": "#e8e8e8",
            "row_even_bg": "#ffffff",
            "row_odd_bg": "#f9fafb",
            "subtotal_bg": "#dbeafe",
        }
        if formato_codigo in (CODIGO_FORMATO_CCD_CC_SEM_001, CODIGO_FORMATO_CCD_CC_MES_001):
            d["capitulo_subtotal_bg"] = "#93c5fd"
        return d
    if formato_codigo == CODIGO_FORMATO_CCD_CC_GER_001:
        d: Dict[str, str] = {
            "section_bar_bg": "#e5e7eb",
            "section_bar_text": "#111827",
            "thead_bg": "#e8e8e8",
            "row_even_bg": "#ffffff",
            "row_odd_bg": "#f9fafb",
            "subtotal_bg": "#dbeafe",
            "capitulo_subtotal_bg": "#93c5fd",
            "ger_titulo_bloque_bg": "#bfdbfe",
            "ger_subtotal_obra_con_aiu_bg": "#e0f2fe",
            "ger_fila_tasa_aiu_bg": "#dbeafe",
            "ger_cdirecto_mas_aiu_bg": "#c7d8f0",
            "ger_filas_post_cdu_bg": "#e8edf5",
            "ger_vtot_obra_ajustes_bg": "#a8bfdb",
            "ger_subtotal_obra_con_iva_bg": "#e0f2fe",
            "ger_fila_tasa_iva_bg": "#e8eeff",
            "ger_cdirecto_mas_iva_bg": "#d4dcf5",
            "ger_vtot_obra_iva_bg": "#c3d0f0",
            "ger_valor_total_acta_bg": "#93c5fd",
        }
        return d
    if formato_codigo in (
        CODIGO_FORMATO_CCD_CC_SUB_002,
        CODIGO_FORMATO_CCD_CC_SEM_002,
        CODIGO_FORMATO_CCD_CC_MES_002,
    ):
        return {
            "section_bar_bg": "#e5e7eb",
            "section_bar_text": "#111827",
            "thead_bg": "#f3f4f6",
            "row_even_bg": "#f8fafc",
            "row_odd_bg": "#ffffff",
            "subtotal_bg": "#e5e7eb",
        }
    return {
        "section_bar_bg": "#e5e7eb",
        "section_bar_text": "#111827",
        "thead_bg": "#f3f4f6",
        "row_even_bg": "#f8fafc",
        "row_odd_bg": "#ffffff",
        "subtotal_bg": "#e5e7eb",
    }


def _merge_estilo_pdf(raw: Any, formato_codigo: str) -> Dict[str, str]:
    base = _default_estilo_pdf(formato_codigo)
    if raw is None:
        return base
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            return base
    if not isinstance(raw, dict):
        return base
    out = dict(base)
    for k in base:
        if k in raw and raw[k] is not None and str(raw[k]).strip() != "":
            out[k] = _sanitize_ccd_hex_color(raw[k], base[k])
    return out


def _estilo_pdf_from_body(body: Optional[CcdEstiloPdfBody], formato_codigo: str) -> Dict[str, str]:
    if body is None:
        return _default_estilo_pdf(formato_codigo)
    d = body.model_dump(exclude_none=True)
    return _merge_estilo_pdf(d, formato_codigo)


def _resolve_estilo_for_upsert(
    body: CcdFirmaConfigBody, contrato_id: int, formato_codigo: str
) -> Dict[str, str]:
    """Si el cliente no envía estilo_pdf, conserva el guardado (no resetea a defaults)."""
    if body.estilo_pdf is not None:
        return _estilo_pdf_from_body(body.estilo_pdf, formato_codigo)
    prev = _get_ccd_firma_config(contrato_id, formato_codigo)
    raw = prev.get("estilo_pdf")
    if isinstance(raw, dict) and raw:
        return _merge_estilo_pdf(raw, formato_codigo)
    return _merge_estilo_pdf(None, formato_codigo)


def _is_ccd_formato_firma_table_missing(exc: BaseException) -> bool:
    s = str(exc).lower()
    return "pgrst205" in s or ("could not find the table" in s and "ccd_formato_firma" in s)


def _opt_usuario_id(val: Any) -> Optional[int]:
    if val is None:
        return None
    try:
        i = int(val)
        return i if i > 0 else None
    except (TypeError, ValueError):
        return None


def _ccd_config_from_row(r: Dict[str, Any], formato_codigo: str) -> Dict[str, Any]:
    return {
        "elaboro_nombre": str(r.get("elaboro_nombre") or "").strip(),
        "elaboro_cargo": str(r.get("elaboro_cargo") or "").strip(),
        "elaboro2_nombre": str(r.get("elaboro2_nombre") or "").strip(),
        "elaboro2_cargo": str(r.get("elaboro2_cargo") or "").strip(),
        "reviso_nombre": str(r.get("reviso_nombre") or "").strip(),
        "reviso_cargo": str(r.get("reviso_cargo") or "").strip(),
        "reviso2_nombre": str(r.get("reviso2_nombre") or "").strip(),
        "reviso2_cargo": str(r.get("reviso2_cargo") or "").strip(),
        "aprobo_nombre": str(r.get("aprobo_nombre") or "").strip(),
        "aprobo_cargo": str(r.get("aprobo_cargo") or "").strip(),
        "elaboro_usuario_id": _opt_usuario_id(r.get("elaboro_usuario_id")),
        "elaboro2_usuario_id": _opt_usuario_id(r.get("elaboro2_usuario_id")),
        "reviso_usuario_id": _opt_usuario_id(r.get("reviso_usuario_id")),
        "reviso2_usuario_id": _opt_usuario_id(r.get("reviso2_usuario_id")),
        "aprobo_usuario_id": _opt_usuario_id(r.get("aprobo_usuario_id")),
        "estilo_pdf": _merge_estilo_pdf(r.get("estilo_pdf"), formato_codigo),
    }


def _ccd_formato_firma_row_get(contrato_id: int, formato_codigo: str) -> Optional[Dict[str, Any]]:
    """Lee fila; reintenta con menos columnas si el esquema aún no tiene aprobo_* o estilo_pdf."""
    selects = (
        # Completo: incluye elaboro2/reviso2 y todos los _usuario_id
        (
            "elaboro_nombre, elaboro_cargo, elaboro2_nombre, elaboro2_cargo,"
            " reviso_nombre, reviso_cargo, reviso2_nombre, reviso2_cargo,"
            " aprobo_nombre, aprobo_cargo, estilo_pdf,"
            " elaboro_usuario_id, elaboro2_usuario_id, reviso_usuario_id, reviso2_usuario_id, aprobo_usuario_id"
        ),
        # Sin elaboro2/reviso2
        (
            "elaboro_nombre, elaboro_cargo, reviso_nombre, reviso_cargo,"
            " aprobo_nombre, aprobo_cargo, estilo_pdf,"
            " elaboro_usuario_id, reviso_usuario_id, aprobo_usuario_id"
        ),
        # Sin aprobo
        "elaboro_nombre, elaboro_cargo, reviso_nombre, reviso_cargo, estilo_pdf, elaboro_usuario_id, reviso_usuario_id, aprobo_usuario_id",
        # Sin _usuario_id de aprobo
        "elaboro_nombre, elaboro_cargo, reviso_nombre, reviso_cargo, estilo_pdf, elaboro_usuario_id, reviso_usuario_id",
        # Solo nombres básicos + estilo
        "elaboro_nombre, elaboro_cargo, reviso_nombre, reviso_cargo, estilo_pdf",
        # Mínimo absoluto
        "elaboro_nombre, elaboro_cargo, reviso_nombre, reviso_cargo",
    )
    last_exc: Optional[BaseException] = None
    for sel in selects:
        try:
            rows = (
                _sb.table("ccd_formato_firma")
                .select(sel)
                .eq("contrato_id", contrato_id)
                .eq("formato_codigo", formato_codigo)
                .limit(1)
                .execute()
                .data
                or []
            )
            if not rows:
                return None
            r = dict(rows[0])
            if "estilo_pdf" not in r:
                r["estilo_pdf"] = None
            return r
        except Exception as e:
            last_exc = e
            continue
    if last_exc:
        _log.warning("ccd_formato_firma_row_get: %s", last_exc)
    return None


def _get_ccd_firma_from_contrato_json(contrato_id: int, formato_codigo: str) -> Dict[str, Any]:
    """Fallback: contratos.ccd_firma_config[formato_codigo] (firmas + estilo_pdf opcional)."""
    try:
        rows = (
            _sb.table("contratos")
            .select("ccd_firma_config")
            .eq("id", contrato_id)
            .limit(1)
            .execute()
            .data
            or []
        )
        _blank = {
            "elaboro_nombre": "", "elaboro_cargo": "",
            "elaboro2_nombre": "", "elaboro2_cargo": "",
            "reviso_nombre": "", "reviso_cargo": "",
            "reviso2_nombre": "", "reviso2_cargo": "",
            "aprobo_nombre": "", "aprobo_cargo": "",
            "elaboro_usuario_id": None, "elaboro2_usuario_id": None,
            "reviso_usuario_id": None, "reviso2_usuario_id": None,
                "aprobo_usuario_id": None,
                "estilo_pdf": _merge_estilo_pdf(None, formato_codigo),
            }
        if not rows:
            return dict(_blank)
        raw = rows[0].get("ccd_firma_config")
        if raw is None:
            return dict(_blank)
        if isinstance(raw, str):
            raw = json.loads(raw)
        if not isinstance(raw, dict):
            return dict(_blank)
        block = raw.get(formato_codigo)
        if not isinstance(block, dict):
            return dict(_blank)
        return {
            "elaboro_nombre": str(block.get("elaboro_nombre") or "").strip(),
            "elaboro_cargo": str(block.get("elaboro_cargo") or "").strip(),
            "elaboro2_nombre": str(block.get("elaboro2_nombre") or "").strip(),
            "elaboro2_cargo": str(block.get("elaboro2_cargo") or "").strip(),
            "reviso_nombre": str(block.get("reviso_nombre") or "").strip(),
            "reviso_cargo": str(block.get("reviso_cargo") or "").strip(),
            "reviso2_nombre": str(block.get("reviso2_nombre") or "").strip(),
            "reviso2_cargo": str(block.get("reviso2_cargo") or "").strip(),
            "aprobo_nombre": str(block.get("aprobo_nombre") or "").strip(),
            "aprobo_cargo": str(block.get("aprobo_cargo") or "").strip(),
            "elaboro_usuario_id": _opt_usuario_id(block.get("elaboro_usuario_id")),
            "elaboro2_usuario_id": _opt_usuario_id(block.get("elaboro2_usuario_id")),
            "reviso_usuario_id": _opt_usuario_id(block.get("reviso_usuario_id")),
            "reviso2_usuario_id": _opt_usuario_id(block.get("reviso2_usuario_id")),
            "aprobo_usuario_id": _opt_usuario_id(block.get("aprobo_usuario_id")),
            "estilo_pdf": _merge_estilo_pdf(block.get("estilo_pdf"), formato_codigo),
        }
    except Exception as e:
        _log.warning("ccd_firma_config en contratos no disponible: %s", e)
        return {
            "elaboro_nombre": "", "elaboro_cargo": "",
            "elaboro2_nombre": "", "elaboro2_cargo": "",
            "reviso_nombre": "", "reviso_cargo": "",
            "reviso2_nombre": "", "reviso2_cargo": "",
            "aprobo_nombre": "", "aprobo_cargo": "",
            "elaboro_usuario_id": None, "elaboro2_usuario_id": None,
            "reviso_usuario_id": None, "reviso2_usuario_id": None,
            "aprobo_usuario_id": None,
            "estilo_pdf": _merge_estilo_pdf(None, formato_codigo),
        }


def _get_ccd_firma_config(contrato_id: int, formato_codigo: str) -> Dict[str, Any]:
    """Lee Elaboró/Revisó y estilo_pdf desde ccd_formato_firma o contratos.ccd_firma_config."""
    try:
        row = _ccd_formato_firma_row_get(contrato_id, formato_codigo)
    except Exception as e:
        if _is_ccd_formato_firma_table_missing(e):
            _log.info("ccd_formato_firma ausente; usando contratos.ccd_firma_config si existe")
            return _get_ccd_firma_from_contrato_json(contrato_id, formato_codigo)
        _log.warning("ccd_formato_firma: %s", e)
        return _get_ccd_firma_from_contrato_json(contrato_id, formato_codigo)
    if row:
        cfg = _ccd_config_from_row(row, formato_codigo)
        jcfg = _get_ccd_firma_from_contrato_json(contrato_id, formato_codigo)
        # El JSON backup se actualiza en cada guardado (fuente de verdad para
        # nombre/cargo/id). La tabla puede tener datos desactualizados si el
        # UPDATE falló silenciosamente (columnas faltantes, RLS, etc.).
        # Regla: JSON gana sobre tabla para todos los campos de firmantes.
        _CAMPOS_FIRMANTE = (
            "elaboro_nombre",  "elaboro_cargo",  "elaboro_usuario_id",
            "elaboro2_nombre", "elaboro2_cargo", "elaboro2_usuario_id",
            "reviso_nombre",   "reviso_cargo",   "reviso_usuario_id",
            "reviso2_nombre",  "reviso2_cargo",  "reviso2_usuario_id",
            "aprobo_nombre",   "aprobo_cargo",   "aprobo_usuario_id",
        )
        for _k in _CAMPOS_FIRMANTE:
            jv = jcfg.get(_k)
            # JSON tiene dato válido → usarlo (puede ser None/vacío = "borrado")
            if _k in jcfg:
                cfg[_k] = jv
        # estilo_pdf: tabla gana (columna JSONB dedicada), JSON como fallback
        ep = row.get("estilo_pdf")
        if ep in (None, "", {}):
            cfg["estilo_pdf"] = jcfg.get("estilo_pdf") or cfg["estilo_pdf"]
        return cfg
    return _get_ccd_firma_from_contrato_json(contrato_id, formato_codigo)


def _list_firmantes_candidatos_contrato(contrato_id: int) -> List[Dict[str, Any]]:
    """Usuarios activos del contrato (principal + usuario_contratos) con cargo para elegir Elaboró/Revisó."""
    try:
        # IDs desde la tabla de relación many-to-many
        uc = (
            _sb.table("usuario_contratos")
            .select("usuario_id")
            .eq("contrato_id", contrato_id)
            .execute()
            .data
            or []
        )
        ids_uc = [r["usuario_id"] for r in uc]
        # IDs de usuarios cuyo contrato principal es este
        pr = (
            _sb.table("usuarios")
            .select("id")
            .eq("contrato_id", contrato_id)
            .execute()
            .data
            or []
        )
        ids_principal = [u["id"] for u in pr]
        todos_ids = list(dict.fromkeys(ids_uc + ids_principal))
        if not todos_ids:
            return []
        # Recuperar datos; campo activo = True/False (no "estado")
        rows = (
            _sb.table("usuarios")
            .select("id, nombre, apellidos, cargo_id, activo")
            .in_("id", todos_ids)
            .execute()
            .data
            or []
        )
        # Preferir solo activos; si ninguno lo es, mostrar todos
        rows_activos = [r for r in rows if r.get("activo") is True]
        rows = rows_activos if rows_activos else rows
        # Mapa cargo_id → nombre de cargo
        carg_rows = _sb.table("cargos").select("id, nombre").execute().data or []
        cmap = {c["id"]: (c.get("nombre") or "").strip() for c in carg_rows}
        out: List[Dict[str, Any]] = []
        for r in rows:
            nom = f"{r.get('nombre') or ''} {r.get('apellidos') or ''}".strip()
            cid = r.get("cargo_id")
            out.append(
                {
                    "id": r.get("id"),
                    "nombre_completo": nom or "—",
                    "cargo": cmap.get(cid, "—"),
                }
            )
        out.sort(key=lambda x: (x.get("nombre_completo") or "").lower())
        return out
    except Exception as e:
        _log.warning("firmantes candidatos: %s", e)
        return []


def _upsert_ccd_firma_in_contrato_json(contrato_id: int, formato_codigo: str, body: CcdFirmaConfigBody) -> None:
    est = _resolve_estilo_for_upsert(body, contrato_id, formato_codigo)
    patch = {
        "elaboro_nombre": (body.elaboro_nombre or "").strip() or None,
        "elaboro_cargo": (body.elaboro_cargo or "").strip() or None,
        "elaboro2_nombre": (body.elaboro2_nombre or "").strip() or None,
        "elaboro2_cargo": (body.elaboro2_cargo or "").strip() or None,
        "reviso_nombre": (body.reviso_nombre or "").strip() or None,
        "reviso_cargo": (body.reviso_cargo or "").strip() or None,
        "reviso2_nombre": (body.reviso2_nombre or "").strip() or None,
        "reviso2_cargo": (body.reviso2_cargo or "").strip() or None,
        "aprobo_nombre": (body.aprobo_nombre or "").strip() or None,
        "aprobo_cargo": (body.aprobo_cargo or "").strip() or None,
        "elaboro_usuario_id": body.elaboro_usuario_id,
        "elaboro2_usuario_id": body.elaboro2_usuario_id,
        "reviso_usuario_id": body.reviso_usuario_id,
        "reviso2_usuario_id": body.reviso2_usuario_id,
        "aprobo_usuario_id": body.aprobo_usuario_id,
        "estilo_pdf": est,
    }
    rows = (
        _sb.table("contratos").select("ccd_firma_config").eq("id", contrato_id).limit(1).execute().data or []
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Contrato no encontrado")
    raw = rows[0].get("ccd_firma_config")
    if raw is None:
        cfg: Dict[str, Any] = {}
    elif isinstance(raw, str):
        cfg = json.loads(raw) if raw.strip() else {}
    elif isinstance(raw, dict):
        cfg = dict(raw)
    else:
        cfg = {}
    cfg[formato_codigo] = patch
    _sb.table("contratos").update({"ccd_firma_config": cfg}).eq("id", contrato_id).execute()


def _upsert_ccd_firma_config(contrato_id: int, formato_codigo: str, body: CcdFirmaConfigBody) -> Dict[str, Any]:
    est = _resolve_estilo_for_upsert(body, contrato_id, formato_codigo)
    row = {
        "contrato_id": contrato_id,
        "formato_codigo": formato_codigo,
        "elaboro_nombre": (body.elaboro_nombre or "").strip() or None,
        "elaboro_cargo": (body.elaboro_cargo or "").strip() or None,
        "elaboro2_nombre": (body.elaboro2_nombre or "").strip() or None,
        "elaboro2_cargo": (body.elaboro2_cargo or "").strip() or None,
        "reviso_nombre": (body.reviso_nombre or "").strip() or None,
        "reviso_cargo": (body.reviso_cargo or "").strip() or None,
        "reviso2_nombre": (body.reviso2_nombre or "").strip() or None,
        "reviso2_cargo": (body.reviso2_cargo or "").strip() or None,
        "aprobo_nombre": (body.aprobo_nombre or "").strip() or None,
        "aprobo_cargo": (body.aprobo_cargo or "").strip() or None,
        "elaboro_usuario_id": body.elaboro_usuario_id,
        "elaboro2_usuario_id": body.elaboro2_usuario_id,
        "reviso_usuario_id": body.reviso_usuario_id,
        "reviso2_usuario_id": body.reviso2_usuario_id,
        "aprobo_usuario_id": body.aprobo_usuario_id,
        "estilo_pdf": est,
    }
    try:
        ex = (
            _sb.table("ccd_formato_firma")
            .select("id")
            .eq("contrato_id", contrato_id)
            .eq("formato_codigo", formato_codigo)
            .limit(1)
            .execute()
            .data
            or []
        )
        row_write = dict(row)

        def _err_txt(exc: BaseException) -> str:
            return f"{exc!s} {repr(exc)}".lower()

        def _do_write(rw: Dict[str, Any]) -> None:
            if ex:
                upd = {k: v for k, v in rw.items() if k not in ("contrato_id", "formato_codigo")}
                res = _sb.table("ccd_formato_firma").update(upd).eq("id", ex[0]["id"]).execute()
                if hasattr(res, "error") and res.error:
                    raise Exception(f"Supabase update error: {res.error}")
                # Nota: res.data == [] puede ocurrir en supabase-py v2 incluso en UPDATEs exitosos
                # (depende del header Prefer/RETURNING). No tratarlo como fallo.
            else:
                res = _sb.table("ccd_formato_firma").insert(rw).execute()
                if hasattr(res, "error") and res.error:
                    raise Exception(f"Supabase insert error: {res.error}")

        def _missing_schema_err(exc: BaseException) -> bool:
            et = _err_txt(exc)
            return any(
                x in et
                for x in (
                    "pgrst204",
                    "schema cache",
                    "could not find",
                    "column",
                    "42703",
                    "does not exist",
                )
            )

        _RM_APROBO = ("aprobo_nombre", "aprobo_cargo", "aprobo_usuario_id")
        _RM_UIDS = ("elaboro_usuario_id", "reviso_usuario_id", "aprobo_usuario_id",
                    "elaboro2_usuario_id", "reviso2_usuario_id")
        _RM_ELABORO2_REVISO2 = (
            "elaboro2_nombre", "elaboro2_cargo", "elaboro2_usuario_id",
            "reviso2_nombre", "reviso2_cargo", "reviso2_usuario_id",
        )
        _RM_ESTILO = ("estilo_pdf",)

        try:
            _do_write(row_write)
            # Siempre sincronizar el JSON backup (fuente de verdad para lectura).
            _upsert_ccd_firma_in_contrato_json(contrato_id, formato_codigo, body)
        except Exception as e2:
            if not _missing_schema_err(e2):
                raise
            last_strip_exc: Optional[BaseException] = e2
            wrote = False
            for rm in (
                _RM_ELABORO2_REVISO2,
                _RM_APROBO,
                tuple(set(_RM_ELABORO2_REVISO2) | set(_RM_APROBO)),
                _RM_UIDS,
                tuple(set(_RM_APROBO) | set(_RM_UIDS)),
                tuple(set(_RM_ELABORO2_REVISO2) | set(_RM_APROBO) | set(_RM_UIDS)),
                tuple(set(_RM_APROBO) | set(_RM_UIDS) | set(_RM_ESTILO)),
                tuple(set(_RM_ELABORO2_REVISO2) | set(_RM_APROBO) | set(_RM_UIDS) | set(_RM_ESTILO)),
            ):
                rm_set = set(rm)
                rw = {k: v for k, v in row_write.items() if k not in rm_set}
                try:
                    _do_write(rw)
                    wrote = True
                    if rm_set:
                        _log.info(
                            "ccd_formato_firma: columnas omitidas en tabla %s; copia completa en contratos.ccd_firma_config",
                            sorted(rm_set),
                        )
                    _upsert_ccd_firma_in_contrato_json(contrato_id, formato_codigo, body)
                    break
                except Exception as ex:
                    last_strip_exc = ex
                    if not _missing_schema_err(ex):
                        raise
                    continue
            if not wrote:
                _upsert_ccd_firma_in_contrato_json(contrato_id, formato_codigo, body)
                _log.warning(
                    "ccd_formato_firma: solo contratos.ccd_firma_config (tabla no aceptó el registro). Último error: %s",
                    last_strip_exc,
                )
    except Exception as e:
        if _is_ccd_formato_firma_table_missing(e):
            try:
                _upsert_ccd_firma_in_contrato_json(contrato_id, formato_codigo, body)
            except HTTPException:
                raise
            except Exception as e2:
                _log.exception("upsert ccd_firma_config en contratos")
                raise HTTPException(
                    status_code=503,
                    detail=(
                        "No se pudo guardar la configuración de firmas en la base de datos. "
                        "Esto no tiene que ver con el subcontratista vacío: solo se guardan Elaboró y Revisó; "
                        "Aprobó se toma del subcontratista al generar el PDF. "
                        "En Supabase → SQL Editor, ejecuta el archivo backend/sql/ccd_formato_firma.sql "
                        "(incluye ALTER TABLE … ccd_firma_config y CREATE TABLE ccd_formato_firma). "
                        f"Detalle: {e2!s}"
                    ),
                ) from e2
            return _get_ccd_firma_config(contrato_id, formato_codigo)
        try:
            _upsert_ccd_firma_in_contrato_json(contrato_id, formato_codigo, body)
            _log.warning(
                "ccd_formato_firma: guardado degradado en contratos.ccd_firma_config tras error en tabla: %s",
                e,
            )
            return _get_ccd_firma_config(contrato_id, formato_codigo)
        except HTTPException:
            raise
        except Exception as e3:
            _log.exception("upsert ccd_formato_firma; JSON contratos también falló")
            raise HTTPException(
                status_code=503,
                detail="No se pudo guardar la configuración de firmas. ¿Existe la tabla ccd_formato_firma? "
                f"Ver backend/sql/ccd_formato_firma.sql (incluye columnas aprobo_*). Detalle tabla: {e!s}. JSON: {e3!s}",
            ) from e3
    return _get_ccd_firma_config(contrato_id, formato_codigo)


# ── REGISTRO DE FORMATOS ────────────────────────────────────────────────────────
# CC-SUB-001 : Corte Subcontratista
# CC-SUB-002 : Memorias Corte Subcontratista (por ítem)

# ── Selectores ────────────────────────────────────────────────────────────────

@router.get("/{contrato_id}/subcontratistas")
def inf_subcontratistas(contrato_id: int, current_user=Depends(_get_user)):
    _perm_informes_ccd(current_user, "ver")
    rows = _sb.table("subcontratistas")\
        .select("id, razon_social, nit, nombre_contacto, telefono")\
        .order("razon_social").execute().data
    return rows or []

@router.get("/{contrato_id}/cortes/{sub_id}")
def inf_cortes(contrato_id: int, sub_id: int, current_user=Depends(_get_user)):
    _perm_informes_ccd(current_user, "ver")
    rows = _sb.table("subcontratista_cortes").select("*")\
        .eq("subcontratista_id", sub_id)\
        .order("consecutivo").execute().data
    return rows or []

@router.get("/{contrato_id}/items-corte/{corte_id}")
def inf_items_corte(contrato_id: int, corte_id: int, current_user=Depends(_get_user)):
    """Ítems únicos aprobados por el sub en un corte dado."""
    _perm_informes_ccd(current_user, "ver")
    try:
        rows = _sb.table("so_registros")\
            .select("item_numero, item_descripcion, unidad, capitulo")\
            .eq("contrato_id", contrato_id)\
            .eq("corte_id", corte_id)\
            .eq("sub_estado", "Aprobado")\
            .execute().data or []
    except Exception as e0:
        err = str(e0).lower()
        if "capitulo" in err or "column" in err or "schema cache" in err:
            _log.warning("so_registros sin columna capitulo (items-corte); reintento sin ella: %s", e0)
            rows = _sb.table("so_registros")\
                .select("item_numero, item_descripcion, unidad")\
                .eq("contrato_id", contrato_id)\
                .eq("corte_id", corte_id)\
                .eq("sub_estado", "Aprobado")\
                .execute().data or []
        else:
            raise
    seen = {}
    for r in rows:
        k = r.get("item_numero")
        if k and k not in seen:
            seen[k] = r
    out = list(seen.values())
    _sort_items_corte_por_item_numero_asc(out)
    return out


def _contexto_corte_sub(contrato_id: int, corte_id: int, current_user: dict) -> Dict[str, Any]:
    """Datos compartidos por vista previa (JSON) y PDF."""
    if not isinstance(current_user, dict):
        try:
            current_user = dict(current_user)
        except Exception:
            current_user = {}

    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")

    corte = _row("subcontratista_cortes", "*", id=corte_id)
    if not corte:
        raise HTTPException(404, "Corte no encontrado")

    sub_id = corte.get("subcontratista_id")
    if sub_id is None:
        raise HTTPException(400, "Corte sin subcontratista asociado (subcontratista_id nulo)")

    sub = _row(
        "subcontratistas",
        "razon_social, nit, nombre_contacto, objeto_contrato",
        id=sub_id,
    ) or {}

    try:
        try:
            registros = _sb.table("so_registros")\
                .select("item_numero, item_descripcion, unidad, cantidad_total, vlr_unitario_subcontratista, capitulo")\
                .eq("contrato_id", contrato_id)\
                .eq("corte_id", corte_id)\
                .eq("sub_estado", "Aprobado")\
                .execute().data or []
        except Exception as e0:
            err = str(e0).lower()
            if "capitulo" in err or "column" in err or "schema cache" in err:
                _log.warning("so_registros sin columna capitulo; reintento sin ella: %s", e0)
                registros = _sb.table("so_registros")\
                    .select("item_numero, item_descripcion, unidad, cantidad_total, vlr_unitario_subcontratista")\
                    .eq("contrato_id", contrato_id)\
                    .eq("corte_id", corte_id)\
                    .eq("sub_estado", "Aprobado")\
                    .execute().data or []
            else:
                raise
    except Exception as e:
        _log.exception("so_registros corte sub")
        raise HTTPException(503, f"No se pudieron leer las cantidades aprobadas: {e!s}") from e

    items_map = {}
    for r in registros:
        k = r.get("item_numero") or "SIN_ITEM"
        if k not in items_map:
            items_map[k] = {
                "item_numero":      r.get("item_numero", ""),
                "item_descripcion": r.get("item_descripcion", ""),
                "unidad":           r.get("unidad", ""),
                "cantidad":         0.0,
                "vlr_unitario_sub": 0.0,
                "costo_directo":    0.0,
                "capitulo":         "",
            }
        cap = str(r.get("capitulo") or "").strip()
        if cap and not items_map[k].get("capitulo"):
            items_map[k]["capitulo"] = cap
        items_map[k]["cantidad"] += _sf(r.get("cantidad_total"), 0.0)
        vu = _sf(r.get("vlr_unitario_subcontratista"), 0.0)
        if items_map[k]["vlr_unitario_sub"] == 0.0 and vu != 0.0:
            items_map[k]["vlr_unitario_sub"] = vu

    for _k, it in items_map.items():
        cd = _sf(it.get("cantidad"), 0.0) * _sf(it.get("vlr_unitario_sub"), 0.0)
        if not math.isfinite(cd):
            cd = 0.0
        it["costo_directo"] = cd

    items = list(items_map.values())
    _sort_items_corte_por_item_numero_asc(items)
    total_costo = sum(_sf(i.get("costo_directo"), 0.0) for i in items)
    if not math.isfinite(total_costo):
        total_costo = 0.0

    usuario_nombre = f"{current_user.get('nombre','')} {current_user.get('apellidos','')}".strip() or "—"
    usuario_cargo = current_user.get("cargo_nombre", "—") or "—"

    return {
        "contrato": contrato,
        "sub": sub,
        "corte": corte,
        "items": items,
        "total_costo": total_costo,
        "usuario_nombre": usuario_nombre,
        "usuario_cargo": usuario_cargo,
    }


def _contexto_memoria_item(
    contrato_id: int,
    corte_id: int,
    item_numero: str,
    current_user: dict,
    *,
    item_exacto: bool = False,
) -> Dict[str, Any]:
    if not isinstance(current_user, dict):
        try:
            current_user = dict(current_user)
        except Exception:
            current_user = {}

    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")

    corte = _row("subcontratista_cortes", "*", id=corte_id)
    if not corte:
        raise HTTPException(404, "Corte no encontrado")

    mem_sub_id = corte.get("subcontratista_id")
    if mem_sub_id is None:
        raise HTTPException(400, "Corte sin subcontratista asociado (subcontratista_id nulo)")

    sub = _row(
        "subcontratistas",
        "razon_social, nit, nombre_contacto, objeto_contrato",
        id=mem_sub_id,
    ) or {}

    q = (
        _sb.table("so_registros")
        .select(
            "numero_registro, abs_inicio, abs_final, pk_id_id, pk_ids(pk_id), calzada, longitud, ancho, espesor, cantidad, cantidad_total, observacion, foto_url, foto_numero, item_numero, item_descripcion, unidad"
        )
        .eq("contrato_id", contrato_id)
        .eq("corte_id", corte_id)
        .eq("sub_estado", "Aprobado")
    )
    if item_exacto:
        q = q.eq("item_numero", (item_numero or "").strip())
    else:
        q = q.ilike("item_numero", f"%{item_numero}%")
    registros = q.order("numero_registro").execute().data or []

    if not registros:
        raise HTTPException(404, "No hay registros aprobados para este ítem en el corte")

    item_info = {
        "item_numero":      registros[0].get("item_numero", item_numero),
        "item_descripcion": registros[0].get("item_descripcion", ""),
        "unidad":           registros[0].get("unidad", ""),
    }

    usuario_nombre = f"{current_user.get('nombre','')} {current_user.get('apellidos','')}".strip() or "—"
    usuario_cargo = current_user.get("cargo_nombre", "—") or "—"

    return {
        "contrato": contrato,
        "sub": sub,
        "corte": corte,
        "item_info": item_info,
        "registros": registros,
        "usuario_nombre": usuario_nombre,
        "usuario_cargo": usuario_cargo,
    }


def _list_item_numeros_memoria_corte(contrato_id: int, corte_id: int) -> List[str]:
    """Ítems distintos con al menos un registro aprobado en el corte.

    Orden: capítulo (número inicial) ascendente, luego código de ítem (orden natural), alineado a CC-SUB-001.
    """
    try:
        rows = (
            _sb.table("so_registros")
            .select("item_numero, capitulo")
            .eq("contrato_id", contrato_id)
            .eq("corte_id", corte_id)
            .eq("sub_estado", "Aprobado")
            .order("numero_registro")
            .execute()
            .data
            or []
        )
    except Exception as e0:
        err = str(e0).lower()
        if "capitulo" in err or "column" in err or "schema cache" in err:
            _log.warning("so_registros sin columna capitulo (memoria ítems); reintento sin ella: %s", e0)
            rows = (
                _sb.table("so_registros")
                .select("item_numero")
                .eq("contrato_id", contrato_id)
                .eq("corte_id", corte_id)
                .eq("sub_estado", "Aprobado")
                .order("numero_registro")
                .execute()
                .data
                or []
            )
        else:
            raise
    seen: set[str] = set()
    out: List[str] = []
    caps: Dict[str, str] = {}
    for r in rows:
        k = (r.get("item_numero") or "").strip()
        if k and k not in seen:
            seen.add(k)
            out.append(k)
        if k:
            cap = str(r.get("capitulo") or "").strip()
            if k not in caps:
                caps[k] = cap
            elif cap and not caps[k]:
                caps[k] = cap
    return _sort_identificadores_item_asc(out, caps if caps else None)


# ── CC-SUB-001 : Corte Subcontratista ─────────────────────────────────────────

def _respuesta_json_corte(contrato_id: int, corte_id: int, current_user: dict) -> Dict[str, Any]:
    ctx = _contexto_corte_sub(contrato_id, corte_id, current_user)
    return {"formato": "CC-SUB-001", **ctx}


def _respuesta_json_memoria(contrato_id: int, corte_id: int, item_numero: str, current_user: dict) -> Dict[str, Any]:
    ctx = _contexto_memoria_item(contrato_id, corte_id, item_numero, current_user)
    return {"formato": "CC-SUB-002", **ctx}


# Rutas JSON de vista previa: definidas en main.py (evita duplicar y asegura registro en la app).

@router.get("/test-sin-auth")
def test_sin_auth():
    return {"ok": True}


@router.get("/formatos-ccd")
def listar_formatos_ccd(current_user=Depends(_get_user)):
    """Biblioteca de formatos ClaraCore Documentación (CCD); convive con asignación futura por contrato."""
    _perm_informes_ccd(current_user, "ver")
    return [{"codigo": k, **v} for k, v in FORMATOS_CCD.items()]


@router.get("/{contrato_id}/ccd/biblioteca")
def ccd_biblioteca_contrato(contrato_id: int, current_user=Depends(_get_user)):
    """Formatos CCD con slots de firma y configuración guardada (Elaboró/Revisó) para este contrato."""
    _perm_informes_ccd(current_user, "ver")
    out: List[Dict[str, Any]] = []
    for codigo, meta in FORMATOS_CCD.items():
        cfg = _get_ccd_firma_config(contrato_id, codigo)
        out.append({"codigo": codigo, **meta, "config_firma": cfg})
    return out


@router.get("/{contrato_id}/ccd/firmantes-candidatos")
def ccd_firmantes_candidatos(contrato_id: int, current_user=Depends(_get_user)):
    """Usuarios del contrato con cargo — para asignar Elaboró y Revisó en la biblioteca CCD."""
    _perm_informes_ccd(current_user, "editar")
    return _list_firmantes_candidatos_contrato(contrato_id)


@router.get("/{contrato_id}/ccd/config-firma/{formato_codigo}")
def ccd_get_config_firma(contrato_id: int, formato_codigo: str, current_user=Depends(_get_user)):
    _perm_informes_ccd(current_user, "ver")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    return _get_ccd_firma_config(contrato_id, formato_codigo)


@router.put("/{contrato_id}/ccd/config-firma/{formato_codigo}")
def ccd_put_config_firma(
    contrato_id: int,
    formato_codigo: str,
    body: CcdFirmaConfigBody,
    current_user=Depends(_get_user),
):
    _perm_informes_ccd(current_user, "editar")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    _upsert_ccd_firma_config(contrato_id, formato_codigo, body)
    # Devuelve el config tal como quedó en BD (permite al frontend detectar fallos silenciosos)
    return _get_ccd_firma_config(contrato_id, formato_codigo)


def _corte_pertenece_contrato(contrato_id: int, corte_id: int) -> bool:
    c = _row("subcontratista_cortes", "id, subcontratista_id", id=corte_id)
    if not c:
        return False
    sid = c.get("subcontratista_id")
    if sid is None:
        return False
    s = _row("subcontratistas", "id, contrato_id", id=sid)
    return bool(s and int(s.get("contrato_id") or 0) == int(contrato_id))


_SLOT_FIRMA_ES = {"elaboro": "Elaboró", "reviso": "Revisó", "aprobo": "Aprobó"}


def _ccd_enviar_correo_confirmacion_firma(
    current_user: dict,
    contrato_id: int,
    formato_codigo: str,
    slot: str,
    *,
    corte_id: Optional[int] = None,
    contexto_ref: Optional[str] = None,
) -> None:
    """Correo al usuario que firmó (misma cuenta). SMTP opcional vía mail_smtp; fallos no bloquean."""
    email = (current_user.get("email") or "").strip()
    if not email:
        return
    slot_es = _SLOT_FIRMA_ES.get(slot, slot)
    num_contrato = ""
    try:
        cr = _row("contratos", "numero", id=contrato_id)
        num_contrato = str((cr or {}).get("numero") or "").strip() or f"id {contrato_id}"
    except Exception:
        num_contrato = f"id {contrato_id}"
    ref_doc = ""
    if contexto_ref:
        ref_doc = str(contexto_ref).strip()
    elif corte_id is not None:
        try:
            cor = _row("subcontratista_cortes", "consecutivo", id=corte_id)
            ref_doc = f"Corte N°: {str((cor or {}).get('consecutivo') or '').strip() or '—'}"
        except Exception:
            ref_doc = "Corte N°: —"
    ahora = datetime.now(ZoneInfo("America/Bogota")).strftime("%d/%m/%Y %H:%M %Z")
    ref_line = (f"{ref_doc}\n" if ref_doc else "")
    body = (
        f"ClaraCore — Confirmación de firma registrada\n\n"
        f"Se registró tu firma digital en el documento CCD asociado a tu cuenta.\n\n"
        f"Rol en el PDF: {slot_es}\n"
        f"Formato: {formato_codigo}\n"
        f"Contrato: {num_contrato}\n"
        f"{ref_line}"
        f"Fecha y hora (registro): {ahora}\n\n"
        f"Si no fuiste tú quien firmó, cambia tu contraseña y avisa al administrador del contrato.\n"
    )
    subj = f"ClaraCore: firma registrada ({slot_es} · {formato_codigo})"
    r = try_send_text_email(email, subj, body)
    if r is None:
        _log.debug("CCD firma: correo no enviado (SMTP no configurado)")
    elif r is False:
        _log.warning("CCD firma: no se pudo enviar correo de confirmación a %s", email)


# ── Snapshot inmutable de firmantes por documento ─────────────────────────────

def _snapshot_key_corte(contrato_id: int, formato_codigo: str, corte_id: int) -> Optional[Dict[str, Any]]:
    """Lee el snapshot guardado para un corte, o None si no existe."""
    try:
        rows = (
            _sb.table("ccd_documento_firma_snapshot")
            .select("*")
            .eq("contrato_id", contrato_id)
            .eq("formato_codigo", formato_codigo)
            .eq("corte_id", corte_id)
            .limit(1)
            .execute()
            .data
            or []
        )
        return rows[0] if rows else None
    except Exception as e:
        _log.debug("ccd_documento_firma_snapshot (corte) get: %s", e)
        return None


def _snapshot_key_contexto(
    contrato_id: int, formato_codigo: str, contexto_tipo: str, contexto_id: int
) -> Optional[Dict[str, Any]]:
    """Lee el snapshot guardado para un documento de contexto (semana/acta_rpo), o None."""
    try:
        rows = (
            _sb.table("ccd_documento_firma_snapshot")
            .select("*")
            .eq("contrato_id", contrato_id)
            .eq("formato_codigo", formato_codigo)
            .eq("contexto_tipo", contexto_tipo)
            .eq("contexto_id", contexto_id)
            .limit(1)
            .execute()
            .data
            or []
        )
        return rows[0] if rows else None
    except Exception as e:
        _log.debug("ccd_documento_firma_snapshot (contexto) get: %s", e)
        return None


def _snapshot_to_firma_cfg(snap: Dict[str, Any]) -> Dict[str, Any]:
    """Convierte un snapshot de tabla en el mismo formato que _get_ccd_firma_config."""
    return {
        "elaboro_nombre":      str(snap.get("elaboro_nombre") or "").strip(),
        "elaboro_cargo":       str(snap.get("elaboro_cargo") or "").strip(),
        "elaboro_usuario_id":  _opt_usuario_id(snap.get("elaboro_usuario_id")),
        "elaboro2_nombre":     str(snap.get("elaboro2_nombre") or "").strip(),
        "elaboro2_cargo":      str(snap.get("elaboro2_cargo") or "").strip(),
        "elaboro2_usuario_id": _opt_usuario_id(snap.get("elaboro2_usuario_id")),
        "reviso_nombre":       str(snap.get("reviso_nombre") or "").strip(),
        "reviso_cargo":        str(snap.get("reviso_cargo") or "").strip(),
        "reviso_usuario_id":   _opt_usuario_id(snap.get("reviso_usuario_id")),
        "reviso2_nombre":      str(snap.get("reviso2_nombre") or "").strip(),
        "reviso2_cargo":       str(snap.get("reviso2_cargo") or "").strip(),
        "reviso2_usuario_id":  _opt_usuario_id(snap.get("reviso2_usuario_id")),
        "aprobo_nombre":       str(snap.get("aprobo_nombre") or "").strip(),
        "aprobo_cargo":        str(snap.get("aprobo_cargo") or "").strip(),
        "aprobo_usuario_id":   _opt_usuario_id(snap.get("aprobo_usuario_id")),
    }


def _guardar_snapshot_si_no_existe(
    contrato_id: int,
    formato_codigo: str,
    fc: Dict[str, Any],
    *,
    corte_id: Optional[int] = None,
    contexto_tipo: Optional[str] = None,
    contexto_id: Optional[int] = None,
) -> None:
    """
    Guarda un snapshot del config de firmantes la primera vez que se firma un documento.
    Si ya existe un snapshot, no hace nada (inmutabilidad garantizada).
    Silencia cualquier error para no bloquear el flujo principal.
    """
    try:
        # Verificar si ya existe
        if corte_id is not None:
            if _snapshot_key_corte(contrato_id, formato_codigo, corte_id):
                return  # ya hay snapshot → no sobreescribir
            row = {"corte_id": corte_id}
        elif contexto_tipo and contexto_id is not None:
            if _snapshot_key_contexto(contrato_id, formato_codigo, contexto_tipo, contexto_id):
                return  # ya hay snapshot → no sobreescribir
            row = {"contexto_tipo": contexto_tipo, "contexto_id": contexto_id}
        else:
            return

        row.update({
            "contrato_id":         contrato_id,
            "formato_codigo":      formato_codigo,
            "elaboro_nombre":      fc.get("elaboro_nombre") or None,
            "elaboro_cargo":       fc.get("elaboro_cargo") or None,
            "elaboro_usuario_id":  fc.get("elaboro_usuario_id") or None,
            "elaboro2_nombre":     fc.get("elaboro2_nombre") or None,
            "elaboro2_cargo":      fc.get("elaboro2_cargo") or None,
            "elaboro2_usuario_id": fc.get("elaboro2_usuario_id") or None,
            "reviso_nombre":       fc.get("reviso_nombre") or None,
            "reviso_cargo":        fc.get("reviso_cargo") or None,
            "reviso_usuario_id":   fc.get("reviso_usuario_id") or None,
            "reviso2_nombre":      fc.get("reviso2_nombre") or None,
            "reviso2_cargo":       fc.get("reviso2_cargo") or None,
            "reviso2_usuario_id":  fc.get("reviso2_usuario_id") or None,
            "aprobo_nombre":       fc.get("aprobo_nombre") or None,
            "aprobo_cargo":        fc.get("aprobo_cargo") or None,
            "aprobo_usuario_id":   fc.get("aprobo_usuario_id") or None,
        })
        _sb.table("ccd_documento_firma_snapshot").insert(row).execute()
        _log.info(
            "ccd_documento_firma_snapshot: snapshot guardado — formato=%s contrato=%s %s",
            formato_codigo, contrato_id,
            f"corte={corte_id}" if corte_id else f"{contexto_tipo}={contexto_id}",
        )
    except Exception as e:
        _log.warning("ccd_documento_firma_snapshot: no se pudo guardar snapshot: %s", e)


def _get_firma_cfg_para_documento(
    contrato_id: int,
    formato_codigo: str,
    *,
    corte_id: Optional[int] = None,
    contexto_tipo: Optional[str] = None,
    contexto_id: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Devuelve la configuración de firmantes para un documento:
    - Si existe snapshot inmutable → lo usa (prioridad absoluta).
    - Si no existe → usa la biblioteca actual (_get_ccd_firma_config).
    Así los documentos firmados conservan sus firmantes originales.

    FO-IDU-EO-04-V2: no usa snapshot por acta. La memoria de cálculo sigue la biblioteca
    mientras se ajustan firmantes; el registro de imagen por slot queda en ccd_firma_registro.
    """
    if formato_codigo == CODIGO_FORMATO_IDU_FO_EO_04_V2:
        return _get_ccd_firma_config(contrato_id, formato_codigo)
    snap: Optional[Dict[str, Any]] = None
    if corte_id is not None:
        snap = _snapshot_key_corte(contrato_id, formato_codigo, corte_id)
    elif contexto_tipo and contexto_id is not None:
        snap = _snapshot_key_contexto(contrato_id, formato_codigo, contexto_tipo, contexto_id)

    if snap:
        cfg_snap = _snapshot_to_firma_cfg(snap)
        # Complementar con estilo_pdf y otros campos no-snapshot desde la biblioteca
        cfg_lib = _get_ccd_firma_config(contrato_id, formato_codigo)
        cfg_snap["estilo_pdf"] = cfg_lib.get("estilo_pdf", {})
        return cfg_snap

    return _get_ccd_firma_config(contrato_id, formato_codigo)


@router.post("/{contrato_id}/ccd/corte/{corte_id}/registrar-firma/{formato_codigo}")
def ccd_registrar_firma_corte(
    contrato_id: int,
    corte_id: int,
    formato_codigo: str,
    current_user: dict = Depends(_get_user),
):
    """Guarda la URL de firma del perfil para Elaboró o Revisó según asignación en biblioteca CCD."""
    _perm_informes_ccd(current_user, "validar")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    uid = _uid_session(current_user)
    if not uid:
        raise HTTPException(status_code=401, detail="Sesión requerida")
    if not _corte_pertenece_contrato(contrato_id, corte_id):
        raise HTTPException(status_code=404, detail="Corte no encontrado en este contrato")
    fc = _get_ccd_firma_config(contrato_id, formato_codigo)
    slot = _slot_firma_usuario_en_config(fc, uid, current_user)
    if not slot:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tu usuario no coincide con Elaboró ni Revisó en la biblioteca CCD para este formato. "
                "Elige al usuario en los desplegables (no solo el texto) y guarda, o revisa que el nombre "
                "coincida con tu perfil."
            ),
        )
    if formato_codigo in (CODIGO_FORMATO_CCD_CC_SUB_001, CODIGO_FORMATO_CCD_CC_SUB_002) and slot == "aprobo":
        raise HTTPException(
            status_code=400,
            detail="En CC-SUB el Aprobó corresponde al subcontratista en el PDF; aquí solo se registra firma para Elaboró o Revisó.",
        )
    url = _firma_imagen_url_usuario(uid)
    if not url:
        raise HTTPException(status_code=400, detail="Configura la imagen de firma en tu perfil antes de registrar.")
    row_ins = {
        "contrato_id": contrato_id,
        "corte_id": corte_id,
        "formato_codigo": formato_codigo,
        "slot": slot,
        "usuario_id": uid,
        "firma_imagen_url": url,
    }
    try:
        _sb.table("ccd_corte_firma_registro").upsert(row_ins, on_conflict="corte_id,formato_codigo,slot").execute()
    except Exception as e:
        _log.exception("ccd_corte_firma_registro upsert")
        raise HTTPException(
            status_code=503,
            detail=(
                "No se pudo registrar la firma. Si la tabla no existe, ejecuta en Supabase el archivo "
                "backend/sql/ccd_corte_firma_registro.sql. "
                f"Detalle: {e!s}"
            ),
        ) from e
    # Snapshot inmutable: guardar config de firmantes al momento de la primera firma.
    _guardar_snapshot_si_no_existe(contrato_id, formato_codigo, fc, corte_id=corte_id)
    _ccd_enviar_correo_confirmacion_firma(
        current_user, contrato_id, formato_codigo, slot, corte_id=corte_id
    )
    return {"ok": True, "slot": slot, "corte_id": corte_id, "formato_codigo": formato_codigo}


@router.get("/{contrato_id}/ccd/corte/{corte_id}/firmas-registradas/{formato_codigo}")
def ccd_firmas_registradas_corte(
    contrato_id: int,
    corte_id: int,
    formato_codigo: str,
    current_user: dict = Depends(_get_user),
):
    """Quién ya registró firma en este corte (Elaboró / Revisó)."""
    _perm_informes_ccd(current_user, "ver")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    if not _corte_pertenece_contrato(contrato_id, corte_id):
        raise HTTPException(status_code=404, detail="Corte no encontrado en este contrato")
    try:
        rows = (
            _sb.table("ccd_corte_firma_registro")
            .select("slot, usuario_id, firma_imagen_url, created_at")
            .eq("corte_id", corte_id)
            .eq("formato_codigo", formato_codigo)
            .execute()
            .data
            or []
        )
    except Exception as e:
        _log.debug("firmas registradas: %s", e)
        return {"elaboro": None, "reviso": None, "tabla_disponible": False}
    out_e = None
    out_r = None
    for r in rows:
        sl = (r.get("slot") or "").strip()
        item = {
            "usuario_id": r.get("usuario_id"),
            "created_at": r.get("created_at"),
            "registrada": True,
        }
        if sl == "elaboro":
            out_e = item
        elif sl == "reviso":
            out_r = item
    return {"elaboro": out_e, "reviso": out_r, "tabla_disponible": True}


def _semana_pertenece_contrato(contrato_id: int, semana_id: int) -> bool:
    r = _row("so_semanas", "id, contrato_id", id=semana_id)
    return bool(r and int(r.get("contrato_id") or 0) == int(contrato_id))


def _acta_pertenece_contrato(contrato_id: int, acta_id: int) -> bool:
    r = _row("actas", "id, contrato_id", id=acta_id)
    return bool(r and int(r.get("contrato_id") or 0) == int(contrato_id))


def _resolver_acta_id_en_contrato(contrato_id: int, acta_ref: Optional[int]) -> Optional[int]:
    """Resuelve una referencia de acta al id real dentro del contrato.

    Acepta: id, numero_rpo o consecutivo. Devuelve id o None si no coincide.
    """
    if acta_ref is None:
        return None
    sref = str(acta_ref).strip()
    if not sref:
        return None
    # 1) Intentar como id directo
    try:
        aid = int(sref)
    except (TypeError, ValueError):
        aid = None
    if aid is not None:
        r = _row("actas", "id, contrato_id", id=aid)
        if r and int(r.get("contrato_id") or 0) == int(contrato_id):
            return int(r.get("id"))
    # 2) Intentar como numero_rpo (texto visible)
    try:
        rows = (
            _sb.table("actas")
            .select("id")
            .eq("contrato_id", int(contrato_id))
            .eq("numero_rpo", sref)
            .order("id", desc=True)
            .limit(1)
            .execute()
            .data
            or []
        )
        if rows and rows[0].get("id") is not None:
            return int(rows[0]["id"])
    except Exception:
        pass
    # 3) Intentar como consecutivo (numérico)
    if aid is not None:
        try:
            rows = (
                _sb.table("actas")
                .select("id")
                .eq("contrato_id", int(contrato_id))
                .eq("consecutivo", aid)
                .order("id", desc=True)
                .limit(1)
                .execute()
                .data
                or []
            )
            if rows and rows[0].get("id") is not None:
                return int(rows[0]["id"])
        except Exception:
            pass
    return None


def _distinct_semana_ids_nivel3_rpc(contrato_id: int) -> Optional[set[int]]:
    """Una sola llamada Postgres (distinct). Requiere backend/sql/ccd_distinct_semanas_nivel3_rpc.sql."""
    try:
        res = _sb.rpc(
            "ccd_distinct_semanas_nivel3_aprobado",
            {"p_contrato_id": int(contrato_id)},
        ).execute()
        raw = getattr(res, "data", None)
        if raw is None:
            return set()
        out: set[int] = set()
        if isinstance(raw, list):
            for x in raw:
                if x is None:
                    continue
                try:
                    out.add(int(x))
                except (TypeError, ValueError):
                    continue
            return out
    except Exception as e:
        _log.info("ccd_distinct_semanas_nivel3_aprobado RPC: %s", e)
    return None


def _distinct_semana_ids_nivel3_fallback_paginado(contrato_id: int, *, max_pages: int = 20) -> set[int]:
    """Reserva si no existe la RPC: pagina so_registros (lento; acotado en páginas)."""
    out: set[int] = set()
    chunk = 1000
    start = 0
    pages = 0
    while pages < max_pages:
        try:
            page = (
                _sb.table("so_registros")
                .select("semana_id")
                .eq("contrato_id", contrato_id)
                .eq("nivel3_estado", "Aprobado")
                .range(start, start + chunk - 1)
                .execute()
                .data
                or []
            )
        except Exception as e:
            _log.warning("ccd semanas: paginación registros nivel3: %s", e)
            break
        if not page:
            break
        for r in page:
            sid = r.get("semana_id")
            if sid is None:
                continue
            try:
                out.add(int(sid))
            except (TypeError, ValueError):
                continue
        pages += 1
        if len(page) < chunk:
            break
        start += chunk
    if pages >= max_pages:
        _log.warning(
            "ccd semanas: límite de paginación (%s páginas); despliega la RPC ccd_distinct_semanas_nivel3_aprobado en Supabase.",
            max_pages,
        )
    return out


def _distinct_semana_ids_nivel3_aprobado_interventoria(contrato_id: int) -> set[int]:
    """IDs de so_semanas con al menos un registro nivel 3 «Aprobado» (interventoría)."""
    s = _distinct_semana_ids_nivel3_rpc(contrato_id)
    if s is not None:
        return s
    return _distinct_semana_ids_nivel3_fallback_paginado(contrato_id)


def _fetch_semanas_rows_por_ids(contrato_id: int, sem_ids: List[int]) -> List[Dict[str, Any]]:
    """PostgREST limita el tamaño del filtro `in`; partimos en lotes."""
    if not sem_ids:
        return []
    batch_size = 120
    rows: List[Dict[str, Any]] = []
    for i in range(0, len(sem_ids), batch_size):
        part = sem_ids[i : i + batch_size]
        try:
            chunk = (
                _sb.table("so_semanas")
                .select("id, numero_semana, fecha_inicio, fecha_fin, estado")
                .eq("contrato_id", contrato_id)
                .in_("id", part)
                .execute()
                .data
                or []
            )
            rows.extend(chunk)
        except Exception as e:
            _log.warning("ccd semanas: fetch so_semanas batch: %s", e)
    return rows


@router.get("/{contrato_id}/ccd/semanas")
def ccd_listar_semanas(contrato_id: int, current_user=Depends(_get_user)):
    _perm_informes_ccd(current_user, "ver")
    sem_ids = _distinct_semana_ids_nivel3_aprobado_interventoria(contrato_id)
    if not sem_ids:
        return []
    rows = _fetch_semanas_rows_por_ids(contrato_id, list(sem_ids))
    # Más reciente primero: fecha_fin descendente, luego número de semana descendente
    def _sort_key(r: Dict[str, Any]) -> tuple:
        ff = str(r.get("fecha_fin") or "")
        fi = str(r.get("fecha_inicio") or "")
        try:
            nsem = int(r.get("numero_semana") or 0)
        except (TypeError, ValueError):
            nsem = 0
        return (ff, fi, nsem)

    rows.sort(key=_sort_key, reverse=True)
    return rows


def _fo_eo_04_list_imagenes_acta(contrato_id: int, acta_id: int) -> List[Dict[str, Any]]:
    """
    Fotos/gráficos de ítems del acta con sellado interventoría (nivel máximo del contrato).
    Una entrada por ítem del PDF, no por cada línea suelta de so_registros.
    """
    items = _fetch_items_n3_acta(int(acta_id), int(contrato_id))
    out: List[Dict[str, Any]] = []
    seen_f: set = set()
    seen_g: set = set()
    for it in items:
        itn = (it.get("item_numero") or "").strip()
        cap = (it.get("capitulo") or "").strip()
        desc = (it.get("item_descripcion") or "").strip() or f"{itn} · {cap}".strip(" ·")
        fn = it.get("foto_numero")
        fu = (it.get("foto_url") or "").strip()
        if fn is not None and fu:
            try:
                n = int(fn)
                if n not in seen_f:
                    seen_f.add(n)
                    out.append({
                        "item_numero": itn,
                        "capitulo": cap,
                        "item_descripcion": desc,
                        "foto_url": fu,
                        "foto_numero": n,
                        "grafico_url": None,
                        "grafico_numero": None,
                    })
            except (TypeError, ValueError):
                pass
        gn = it.get("grafico_numero")
        gu = (it.get("grafico_url") or "").strip()
        if gn is not None and gu:
            try:
                n = int(gn)
                if n not in seen_g:
                    seen_g.add(n)
                    out.append({
                        "item_numero": itn,
                        "capitulo": cap,
                        "item_descripcion": desc,
                        "foto_url": None,
                        "foto_numero": None,
                        "grafico_url": gu,
                        "grafico_numero": n,
                    })
            except (TypeError, ValueError):
                pass
    _log.info(
        "fo_eo_04 list_imagenes_acta: acta=%s items_pdf=%s entradas_img=%s",
        acta_id,
        len(items),
        len(out),
    )
    return out


@router.get("/{contrato_id}/ccd/fo-eo-04/fotos-acta")
def ccd_fo_eo_04_fotos_acta(
    contrato_id: int,
    acta_id: int = Query(..., description="Id del acta RPO"),
    current_user=Depends(_get_user),
):
    """Fotos/gráficos de ítems sellados por interventoría (mismos que el PDF FO-EO-04)."""
    _perm_informes_ccd(current_user, "ver")
    acta_norm = _resolver_acta_id_en_contrato(contrato_id, acta_id)
    if not acta_norm:
        raise HTTPException(404, "Acta no encontrada en este contrato")
    return _fo_eo_04_list_imagenes_acta(int(contrato_id), int(acta_norm))


@router.get("/{contrato_id}/ccd/fo-eo-04/diagnostico")
def ccd_fo_eo_04_diagnostico(
    contrato_id: int,
    acta_id: int = Query(..., description="Id del acta RPO"),
    current_user=Depends(_get_user),
):
    """Diagnóstico: por qué el FO-EO-04 puede salir sin cantidades (misma regla que panel Actas)."""
    _perm_informes_ccd(current_user, "ver")
    acta_norm = _resolver_acta_id_en_contrato(contrato_id, acta_id)
    if not acta_norm:
        raise HTTPException(404, "Acta no encontrada en este contrato")
    matriz = matriz_params_contrato(_sb, int(contrato_id))
    campo_mx, niveles_act = matriz
    cascade = _fetch_cascade_interventoria_actas_rpo(
        _sb,
        int(contrato_id),
        [int(acta_norm)],
        campo_nivel_max=campo_mx,
        niveles_activos=niveles_act,
    )
    raw = _fo_eo_04_fetch_registros_acta(int(contrato_id), int(acta_norm))
    sellados = _fo_eo_04_registros_sellados_acta(
        int(contrato_id), int(acta_norm), matriz=matriz
    )
    items = _fetch_items_n3_acta(int(acta_norm), int(contrato_id))
    prev_ids = _fo_eo_04_actas_anteriores_ids(int(contrato_id), int(acta_norm))
    raw_prev = (
        _fo_eo_04_fetch_registros_actas_ids(int(contrato_id), prev_ids) if prev_ids else []
    )
    tot_batch, tot_meta = (
        _fetch_totales_batch(int(contrato_id), int(acta_norm), items) if items else ({}, {})
    )
    muestra_ant = None
    if items:
        it0 = items[0]
        k0 = _fo_eo_04_norm_item_key(it0.get("item_numero") or "", it0.get("capitulo") or "")
        muestra_ant = tot_batch.get(k0, 0.0)
    return {
        "acta_id": acta_norm,
        "campo_nivel_maximo": campo_mx,
        "niveles_activos": niveles_act,
        "registros_raw_acta": len(raw),
        "registros_cascade_panel_actas": len(cascade or []),
        "registros_sellados_memoria": len(sellados),
        "items_pdf": len(items),
        "actas_anteriores_ids": len(prev_ids),
        "registros_actas_anteriores_raw": len(raw_prev),
        "muestra_total_anteriores_item0": muestra_ant,
        "totales_meta": tot_meta,
        "mensaje": (
            "OK: hay ítems para el PDF."
            if items
            else (
                "Sin ítems: no hay líneas con el nivel máximo del contrato en Aprobado para este acta. "
                "Revise matriz SICOE y acta_rpo_id en registros/reportes."
            )
        ),
    }


@router.get("/{contrato_id}/ccd/actas-rpo")
def ccd_listar_actas_rpo(contrato_id: int, current_user=Depends(_get_user)):
    """Actas de cobro RPO (excluye administrativas u otros grupos)."""
    _perm_informes_ccd(current_user, "ver")
    rows = (
        _sb.table("actas")
        .select("id, numero_rpo, consecutivo, fecha_inicio, fecha_fin")
        .eq("contrato_id", contrato_id)
        .eq("tipo_grupo", "RPO")
        .order("consecutivo", desc=True)
        .execute()
        .data
        or []
    )
    return rows


@router.get("/{contrato_id}/ccd/conciliacion/semana/{semana_id}/items")
def ccd_items_conciliacion_semana(contrato_id: int, semana_id: int, current_user=Depends(_get_user)):
    _perm_informes_ccd(current_user, "ver")
    if not _semana_pertenece_contrato(contrato_id, semana_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    reg = fetch_registros_conciliacion(_sb, contrato_id, semana_id=semana_id)
    items, total = aggregate_items_conciliacion(reg)
    _sort_items_corte_por_item_numero_asc(items)
    caps = _capitulos_por_item_numero_desde_items(items)
    nums = _sort_identificadores_item_asc(
        list({(it.get("item_numero") or "").strip() for it in items if (it.get("item_numero") or "").strip()}),
        caps,
    )
    return {"items": items, "total_costo": total, "item_numeros": nums, "registros_raw": len(reg)}


@router.get("/{contrato_id}/ccd/conciliacion/acta/{acta_id}/items")
def ccd_items_conciliacion_acta(contrato_id: int, acta_id: int, current_user=Depends(_get_user)):
    _perm_informes_ccd(current_user, "ver")
    if not _acta_pertenece_contrato(contrato_id, acta_id):
        raise HTTPException(404, "Acta no encontrada en este contrato")
    reg = fetch_registros_informe_cc_mes_por_acta(_sb, contrato_id, acta_id)
    items, total = aggregate_items_conciliacion(reg)
    _sort_items_corte_por_item_numero_asc(items)
    caps = _capitulos_por_item_numero_desde_items(items)
    nums = _sort_identificadores_item_asc(
        list({(it.get("item_numero") or "").strip() for it in items if (it.get("item_numero") or "").strip()}),
        caps,
    )
    return {"items": items, "total_costo": total, "item_numeros": nums, "registros_raw": len(reg)}


@router.post("/{contrato_id}/ccd/contexto/{contexto_tipo}/{contexto_id}/registrar-firma/{formato_codigo}")
def ccd_registrar_firma_contexto(
    contrato_id: int,
    contexto_tipo: str,
    contexto_id: int,
    formato_codigo: str,
    current_user: dict = Depends(_get_user),
):
    _perm_informes_ccd(current_user, "validar")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    if contexto_tipo not in ("semana", "acta_rpo"):
        raise HTTPException(400, "contexto_tipo debe ser semana o acta_rpo")
    uid = _uid_session(current_user)
    if not uid:
        raise HTTPException(status_code=401, detail="Sesión requerida")
    if contexto_tipo == "semana" and not _semana_pertenece_contrato(contrato_id, contexto_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    if contexto_tipo == "acta_rpo" and not _acta_pertenece_contrato(contrato_id, contexto_id):
        raise HTTPException(404, "Acta no encontrada en este contrato")
    fc = _get_ccd_firma_config(contrato_id, formato_codigo)
    slot = _slot_firma_usuario_en_config(fc, uid, current_user)
    if not slot:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tu usuario no coincide con Elaboró, Revisó ni Aprobó configurados en la biblioteca CCD para este formato. "
                "Vuelve a guardar la biblioteca (asigna tu usuario en el desplegable o revisa nombres). "
                "Si al guardar falla el servidor, ejecuta en Supabase los ALTER de aprobo_* en backend/sql/ccd_formato_firma.sql."
            ),
        )
    url = _firma_imagen_url_usuario(uid)
    if not url:
        raise HTTPException(status_code=400, detail="Configura la imagen de firma en tu perfil antes de registrar.")
    row_ins = {
        "contrato_id": contrato_id,
        "formato_codigo": formato_codigo,
        "contexto_tipo": contexto_tipo,
        "contexto_id": contexto_id,
        "slot": slot,
        "usuario_id": uid,
        "firma_imagen_url": url,
    }
    try:
        _sb.table("ccd_firma_registro").upsert(
            row_ins,
            on_conflict="contrato_id,formato_codigo,contexto_tipo,contexto_id,slot",
        ).execute()
    except Exception as e:
        _log.exception("ccd_firma_registro upsert")
        raise HTTPException(
            status_code=503,
            detail=(
                "No se pudo registrar la firma. Si el error menciona check constraint en slot, ejecuta en Supabase "
                "backend/sql/ccd_firma_registro_slot_elaboro2_reviso2.sql (slots elaboro2/reviso2 para FO-EO-04). "
                "Si falta la tabla, ejecuta backend/sql/ccd_firma_registro_contexto.sql. "
                f"Detalle: {e!s}"
            ),
        ) from e
    # Snapshot inmutable (no aplica a FO-IDU-EO-04-V2: siempre biblioteca viva).
    if formato_codigo != CODIGO_FORMATO_IDU_FO_EO_04_V2:
        _guardar_snapshot_si_no_existe(
            contrato_id, formato_codigo, fc,
            contexto_tipo=contexto_tipo, contexto_id=contexto_id,
        )
    ref = ""
    if contexto_tipo == "semana":
        sm = _row("so_semanas", "numero_semana", id=contexto_id)
        ref = f"Semana N°: {str((sm or {}).get('numero_semana') or contexto_id)}"
    elif contexto_tipo == "acta_rpo":
        ac = _row("actas", "numero_rpo, consecutivo", id=contexto_id)
        ref = f"Acta RPO: {str((ac or {}).get('numero_rpo') or (ac or {}).get('consecutivo') or contexto_id)}"
    _ccd_enviar_correo_confirmacion_firma(
        current_user, contrato_id, formato_codigo, slot, contexto_ref=ref
    )
    return {"ok": True, "slot": slot, "contexto_tipo": contexto_tipo, "contexto_id": contexto_id, "formato_codigo": formato_codigo}


@router.get("/{contrato_id}/ccd/contexto/{contexto_tipo}/{contexto_id}/firmas-registradas/{formato_codigo}")
def ccd_firmas_registradas_contexto(
    contrato_id: int,
    contexto_tipo: str,
    contexto_id: int,
    formato_codigo: str,
    current_user: dict = Depends(_get_user),
):
    _perm_informes_ccd_ver_o_validar(current_user)
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    if contexto_tipo not in ("semana", "acta_rpo"):
        raise HTTPException(400, "contexto_tipo debe ser semana o acta_rpo")
    if contexto_tipo == "semana" and not _semana_pertenece_contrato(contrato_id, contexto_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    if contexto_tipo == "acta_rpo" and not _acta_pertenece_contrato(contrato_id, contexto_id):
        raise HTTPException(404, "Acta no encontrada en este contrato")
    try:
        rows = (
            _sb.table("ccd_firma_registro")
            .select("slot, usuario_id, firma_imagen_url, created_at")
            .eq("contrato_id", contrato_id)
            .eq("contexto_tipo", contexto_tipo)
            .eq("contexto_id", contexto_id)
            .eq("formato_codigo", formato_codigo)
            .execute()
            .data
            or []
        )
    except Exception as e:
        _log.debug("firmas contexto: %s", e)
        return {"elaboro": None, "elaboro2": None, "reviso": None, "reviso2": None, "aprobo": None, "tabla_disponible": False}
    out: Dict[str, Any] = {"elaboro": None, "elaboro2": None, "reviso": None, "reviso2": None, "aprobo": None, "tabla_disponible": True}
    for r in rows:
        sl = (r.get("slot") or "").strip()
        item = {"usuario_id": r.get("usuario_id"), "created_at": r.get("created_at"), "registrada": True}
        if sl in out:
            out[sl] = item
    return out


def _pdf_bytes_conciliacion_informe_v1(
    contrato_id: int,
    current_user: dict,
    *,
    formato_codigo: str,
    contexto_tipo: str,
    contexto_id: int,
    titulo_documento: str,
    c3_label: str,
    c3_value: str,
    c4_label: str,
    c4_value: str,
    pie_contexto: str,
    items: List[dict],
    total_costo: float,
) -> bytes:
    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    u = current_user if isinstance(current_user, dict) else dict(current_user)
    usuario_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
    usuario_cargo = u.get("cargo_nombre", "—") or "—"
    firma_cfg = _get_firma_cfg_para_documento(
        contrato_id, formato_codigo, contexto_tipo=contexto_tipo, contexto_id=contexto_id
    )
    fc = firma_cfg or {}
    e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
    r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
    a_uid = _opt_usuario_id(fc.get("aprobo_usuario_id"))
    enom = str(fc.get("elaboro_nombre") or "").strip()
    rnom = str(fc.get("reviso_nombre") or "").strip()
    anom = str(fc.get("aprobo_nombre") or "").strip()
    elaboro_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, contexto_tipo, contexto_id, formato_codigo, "elaboro", e_uid, enom, current_user
    )
    reviso_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, contexto_tipo, contexto_id, formato_codigo, "reviso", r_uid, rnom, current_user
    )
    aprobo_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, contexto_tipo, contexto_id, formato_codigo, "aprobo", a_uid, anom, current_user
    )
    html = _html_cc_conciliacion_informe_v1(
        contrato,
        items,
        total_costo,
        usuario_nombre,
        usuario_cargo,
        codigo_ccd=formato_codigo,
        titulo_documento=titulo_documento,
        c3_label=c3_label,
        c3_value=c3_value,
        c4_label=c4_label,
        c4_value=c4_value,
        pie_contexto=pie_contexto,
        firma_cfg=firma_cfg,
        elaboro_firma_data_uri=elaboro_uri,
        reviso_firma_data_uri=reviso_uri,
        aprobo_firma_data_uri=aprobo_uri,
        estilo_formato_codigo=formato_codigo,
    )
    return _to_pdf(html)


def _rpo_gerencia_vacio() -> Dict[str, Any]:
    return {
        "costo_directo_total": 0.0,
        "registros_cascade_interventoria": 0,
        "registros_n3_aprobado": 0,
        "por_capitulo": [],
        "secciones": {},
    }


def _map_por_capitulo_a_montos(por: Optional[List[Dict[str, Any]]]) -> Dict[str, float]:
    m: Dict[str, float] = {}
    for r in por or []:
        c = (str(r.get("capitulo") or "—").strip()) or "—"
        m[c] = float(r.get("costo_directo") or 0)
    return m


def _merge_filas_gerencia_dos_actas(
    p_list: Optional[List[Dict[str, Any]]],
    a_list: Optional[List[Dict[str, Any]]],
) -> List[Tuple[str, float, float, float]]:
    m_p = _map_por_capitulo_a_montos(p_list)
    m_a = _map_por_capitulo_a_montos(a_list)
    keys = set(m_p) | set(m_a)
    caps = sorted(keys, key=lambda c: _orden_titulo_capitulo_obra(c))
    out: List[Tuple[str, float, float, float]] = []
    for c in caps:
        vp, va = m_p.get(c, 0.0), m_a.get(c, 0.0)
        out.append((c, vp, va, vp - va))
    return out


def _cobro_item_numeros_contrato(contrato_id: int) -> set:
    try:
        rows = _sb.table("cobro").select("item").eq("contrato_id", contrato_id).execute().data or []
    except Exception:
        return set()
    return {str(r.get("item") or "").strip() for r in rows if str(r.get("item") or "").strip()}


def _actas_rpo_consecutivo_ascendente(contrato_id: int) -> List[Dict[str, Any]]:
    rows = (
        _sb.table("actas")
        .select("id, consecutivo, numero_rpo, fecha_inicio, fecha_fin, tipo_grupo, pct_proyectado_ajustes")
        .eq("contrato_id", contrato_id)
        .eq("tipo_grupo", "RPO")
        .order("consecutivo", desc=False)
        .execute()
        .data
        or []
    )
    return [r for r in rows if r.get("id") is not None]


def _parse_fecha_acta_informe(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()[:10]
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _acta_rpo_vigente_desde_lista_actas(actas: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Misma semántica que main._acta_rpo_vigente_row (orden SQL: fecha_inicio desc, numero_rpo desc, id desc)."""
    today = date.today()
    cand: List[Dict[str, Any]] = []
    for a in actas:
        fi = _parse_fecha_acta_informe(a.get("fecha_inicio"))
        ff = _parse_fecha_acta_informe(a.get("fecha_fin"))
        if fi and ff and fi <= today <= ff:
            cand.append(a)
    if not cand:
        return None

    def _sort_key(a: Dict[str, Any]):
        fi = _parse_fecha_acta_informe(a.get("fecha_inicio")) or date.min
        try:
            nr = int(a.get("numero_rpo") or 0)
        except (TypeError, ValueError):
            nr = 0
        try:
            aid = int(a.get("id") or 0)
        except (TypeError, ValueError):
            aid = 0
        return (-fi.toordinal(), -nr, -aid)

    return sorted(cand, key=_sort_key)[0]


_ECON_OH_KEYS = (
    "valor_comp_ambiental",
    "valor_comp_social",
    "valor_comp_pmt",
    "valor_cobrado_adicional",
    "ajuste_iccp",
    "ajuste_icociv",
    "ajuste_ipc",
)


def _suma_historico_economia_rpo_antes_presente(
    contrato_id: int, acta_presente_id: int
) -> Dict[str, float]:
    """
    Suma en COP (solo actas RPO estrictamente anteriores al acta presente del informe).
    El aporte del acta vigente (col.1: factor VR o líneas) se agrega en _embalaje, no en BD si no está diligenciado.
    """
    z = {k: 0.0 for k in _ECON_OH_KEYS}
    if contrato_id <= 0 or acta_presente_id <= 0:
        return z
    actas = _actas_rpo_consecutivo_ascendente(contrato_id)
    idx = next(
        (i for i, a in enumerate(actas) if int(a.get("id") or 0) == int(acta_presente_id)),
        None,
    )
    if idx is None or idx < 1:
        return z
    ids: List[int] = []
    for a in actas[:idx]:
        aid = a.get("id")
        if aid is not None:
            try:
                ids.append(int(aid))
            except (TypeError, ValueError):
                continue
    if not ids:
        return z
    cols = ", ".join(_ECON_OH_KEYS)
    for off in range(0, len(ids), 200):
        chunk = ids[off : off + 200]
        try:
            rws = _sb.table("actas").select(cols).in_("id", chunk).execute().data or []
        except Exception:
            rws = []
        for row in rws or []:
            if not isinstance(row, dict):
                continue
            for k in _ECON_OH_KEYS:
                try:
                    p = float(row.get(k) or 0.0)
                except (TypeError, ValueError):
                    p = 0.0
                if not math.isfinite(p):
                    p = 0.0
                z[k] += p
    for k in z:
        z[k] = float(round(z[k], 0))
    return z


def _resolver_acta_gerencia_presente_y_anterior(
    contrato_id: int, acta_presente_id: Optional[int] = None
) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Acta «presente» (columna 1) por defecto: **RPO en período** — hoy ∈ [fecha_inicio, fecha_fin],
    alineado con la matriz SICOE / acta vigente. Si no hay ninguna (p. ej. sin actas o hueco entre períodos),
    se usa la de **mayor consecutivo** (legado). Con `acta_presente_id` se fuerza un acta concreta.
    Anterior = RPO inmediata previa por consecutivo respecto al presente.
    Devuelve también la lista de actas RPO ya cargada (orden ascendente por consecutivo) para evitar otra consulta.
    """
    actas = _actas_rpo_consecutivo_ascendente(contrato_id)
    if not actas:
        return None, None, []
    if acta_presente_id is not None and _acta_pertenece_contrato(contrato_id, int(acta_presente_id)):
        for i, a in enumerate(actas):
            if int(a.get("id") or 0) == int(acta_presente_id):
                return a, (actas[i - 1] if i > 0 else None), actas
    vig = _acta_rpo_vigente_desde_lista_actas(actas)
    if vig and vig.get("id"):
        vid = int(vig["id"])
        for i, a in enumerate(actas):
            if int(a.get("id") or 0) == vid:
                return a, (actas[i - 1] if i > 0 else None), actas
        c_v = int(vig.get("consecutivo") or 0)
        anterior = None
        best_c = -1
        for a in actas:
            try:
                ac = int(a.get("consecutivo") or 0)
            except (TypeError, ValueError):
                continue
            if ac < c_v and ac > best_c:
                best_c = ac
                anterior = a
        return vig, anterior, actas
    presente = actas[-1]
    anterior = actas[-2] if len(actas) >= 2 else None
    return presente, anterior, actas


def _construir_datos_informe_gerencia_matriz_fallback_por_capitulo(
    contrato_id: int,
    pres: dict,
    ant: Optional[dict],
    ap_id: int,
    c_pres: int,
    a_ant: Optional[int],
    ids_cascade_hasta: List[int],
    t2: float,
    t_acc: float,
    aiu_c: Optional[float] = None,
    iva_c: Optional[float] = None,
    vr_contr: Optional[dict] = None,
) -> Dict[str, Any]:
    """Lento: paginación Python; se usa si los RPC de informe_gerencia no existen aún en Supabase."""
    todas = fetch_registros_acta_todas_sico_obra(_sb, contrato_id, ap_id)
    c1r = list(todas)
    m1 = suma_por_capitulo_solo_cdirecto_almacenado(c1r)
    m4 = suma_por_capitulo_desde_registros([r for r in todas if registro_tiene_pendiente_matriz(r)])
    rmap = rpo_conciliacion_por_contrato(_sb, contrato_id, [ap_id, a_ant] if a_ant else [ap_id])
    d_p = rmap.get(ap_id) or _rpo_gerencia_vacio()
    m2: Dict[str, float] = {}
    if a_ant is not None:
        d_a = rmap.get(int(a_ant)) or _rpo_gerencia_vacio()
        for row in d_a.get("por_capitulo") or []:
            c2 = (str(row.get("capitulo") or "—").strip()) or "—"
            m2[c2] = m2.get(c2, 0.0) + float(row.get("costo_directo") or 0.0)
    campo_max, niveles_act = matriz_params_contrato(_sb, contrato_id)
    sellados = _fetch_cascade_interventoria_actas_rpo(
        _sb,
        contrato_id,
        [int(x) for x in ids_cascade_hasta],
        campo_nivel_max=campo_max,
        niveles_activos=niveles_act,
    )
    m3 = suma_por_capitulo_interventoria_sellada(
        sellados,
        niveles_activos=niveles_act,
        campo_nivel_max=campo_max,
    )
    caps = set(m1) | set(m2) | set(m3) | set(m4)
    caps_orden = sorted(caps, key=lambda c: _orden_titulo_capitulo_obra(c))
    c1b: Dict[Tuple[str, str], float] = {}
    c2b: Dict[Tuple[str, str], float] = {}
    c3b: Dict[Tuple[str, str], float] = {}
    c4b: Dict[Tuple[str, str], float] = {}
    for c in m1:
        nk = _norm_cap_informe_gerencia(c)
        bl = _bloque_capitulo_matriz(c)
        c1b[(nk, bl)] = c1b.get((nk, bl), 0.0) + m1.get(c, 0.0)
    for c in m2:
        nk = _norm_cap_informe_gerencia(c)
        c2b[(nk, _bloque_capitulo_matriz(c))] = c2b.get((nk, _bloque_capitulo_matriz(c)), 0.0) + m2.get(c, 0.0)
    for c in m3:
        nk = _norm_cap_informe_gerencia(c)
        c3b[(nk, _bloque_capitulo_matriz(c))] = c3b.get((nk, _bloque_capitulo_matriz(c)), 0.0) + m3.get(c, 0.0)
    for c in m4:
        nk = _norm_cap_informe_gerencia(c)
        c4b[(nk, _bloque_capitulo_matriz(c))] = c4b.get((nk, _bloque_capitulo_matriz(c)), 0.0) + m4.get(c, 0.0)
    rpv = rpo_resumen_actas_rpc(_sb, contrato_id, [ap_id]) or {}
    t_pl = float((rpv.get(int(ap_id)) or {}).get("costo_directo_total", 0) or 0)
    vr = vr_contr
    if vr is None:
        vr = (
            _row(
                "contratos",
                "aiu, iva, costo_directo_contrato, valor_componente_ambiental, "
                "valor_componente_social, valor_componente_pmt, costos_adicionales_lista",
                id=contrato_id,
            )
            or {}
        )
    return _embalaje_informe_gerencia_bloques(
        pres,
        ant,
        c_pres,
        ids_cascade_hasta,
        t2,
        t_acc,
        t_pl,
        c1b,
        c2b,
        c3b,
        c4b,
        d_p,
        a_ant,
        aiu_c,
        iva_c,
        vr_contr=vr,
        contrato_id=contrato_id,
    )  # t2, t_acc, t_pl desde rpo_resumen (listado actas) para coherencia


def _tasa_fraccion_desde_valor_contrato(tasa_raw: Optional[Any]) -> float:
    """
    Panel admin (AIU, IVA): lo habitual es 0,19 = 19 %. Si el valor es >1,
    asumir porcentaje (p. ej. 19) y normalizar a fracción.
    """
    if tasa_raw is None:
        return 0.0
    try:
        f = float(tasa_raw)
    except (TypeError, ValueError):
        return 0.0
    if not math.isfinite(f) or f < 0.0:
        return 0.0
    if f > 1.0:
        f = f / 100.0
    if f > 1.0:
        f = 1.0
    return f


def _embalaje_informe_gerencia_bloques(
    pres: dict,
    ant: Optional[dict],
    c_pres: int,
    ids_cascade_hasta: List[int],
    t2: float,
    t_acc: float,
    t_presente_lista: float,
    c1: Dict[Tuple[str, str], float],
    c2: Dict[Tuple[str, str], float],
    c3: Dict[Tuple[str, str], float],
    c4: Dict[Tuple[str, str], float],
    d_p: dict,
    a_ant: Optional[int],
    aiu_pct: Optional[float] = None,
    iva_pct: Optional[float] = None,
    vr_contr: Optional[dict] = None,
    contrato_id: Optional[int] = None,
) -> Dict[str, Any]:
    all_keys: set = set()
    for m in (c1, c2, c3, c4):
        all_keys |= set(m.keys())
    aiu = _tasa_fraccion_desde_valor_contrato(aiu_pct)
    iva = _tasa_fraccion_desde_valor_contrato(iva_pct)

    def _fila_bloque(bid: str, tit: str) -> dict:
        keys = sorted([k for k in all_keys if k[1] == bid], key=lambda t: _orden_titulo_capitulo_obra(t[0]))
        fils = [
            {
                "capitulo": k[0],
                "bloque": k[1],
                "c1": float(c1.get(k, 0.0)),
                "c2": float(c2.get(k, 0.0)) if a_ant is not None else 0.0,
                "c3": float(c3.get(k, 0.0)),
                "c4": float(c4.get(k, 0.0)),
            }
            for k in keys
        ]
        st = {
            "c1": sum(f["c1"] for f in fils) if fils else 0.0,
            "c2": sum(f["c2"] for f in fils) if fils else 0.0,
            "c3": sum(f["c3"] for f in fils) if fils else 0.0,
            "c4": sum(f["c4"] for f in fils) if fils else 0.0,
        }
        out: Dict[str, Any] = {
            "id": bid,
            "titulo": tit,
            "filas": fils,
            "subtotal": st,
        }
        if bid == "obra" and aiu > 0.0 and (st.get("c1", 0) or st.get("c2", 0) or st.get("c3", 0) or st.get("c4", 0)):
            out["fila_aiu"] = {
                "c1": st["c1"] * aiu,
                "c2": st["c2"] * aiu,
                "c3": st["c3"] * aiu,
                "c4": st["c4"] * aiu,
            }
            out["fila_costo_directo_mas_aiu"] = {
                "c1": st["c1"] + out["fila_aiu"]["c1"],
                "c2": st["c2"] + out["fila_aiu"]["c2"],
                "c3": st["c3"] + out["fila_aiu"]["c3"],
                "c4": st["c4"] + out["fila_aiu"]["c4"],
            }
        if bid == "obra" and isinstance(vr_contr, dict):
            f_cdu = out.get("fila_costo_directo_mas_aiu")
            n_acta = 0.0
            if isinstance(f_cdu, dict) and f_cdu:
                n_acta = float(f_cdu.get("c1") or 0.0)
            else:
                n_acta = float(st.get("c1") or 0.0)
            n_acta_r = round(n_acta, 0)
            try:
                cd0 = float(vr_contr.get("costo_directo_contrato") or 0.0)
            except (TypeError, ValueError):
                cd0 = 0.0
            n_con = round(cd0 * (1.0 + aiu), 0)
            factor = 0.0
            if n_con > 0.0 and math.isfinite(n_con) and n_con == n_con:
                factor = float(n_acta_r) / float(n_con)
            if not math.isfinite(factor):
                factor = 0.0

            n_acta4 = 0.0
            if isinstance(f_cdu, dict) and f_cdu:
                n_acta4 = float(f_cdu.get("c4") or 0.0)
            else:
                n_acta4 = float(st.get("c4") or 0.0)
            n_acta4_r = round(n_acta4, 0)
            factor4 = 0.0
            if n_con > 0.0 and math.isfinite(n_con) and n_con == n_con:
                factor4 = float(n_acta4_r) / float(n_con)
            if not math.isfinite(factor4):
                factor4 = 0.0

            def _v1(vv) -> float:
                try:
                    w = float(vv)
                except (TypeError, ValueError):
                    w = 0.0
                return float(round(factor * w, 0))

            def _v4(vv) -> float:
                """Escala a columna 4 (pendiente) con la misma base de contrato que c1, pero con CD+AIU de pendientes."""
                try:
                    w = float(vv)
                except (TypeError, ValueError):
                    w = 0.0
                return float(round(factor4 * w, 0))

            # Col.2 (acta de referencia anterior a la vigente): importes reales de esa acta (no factores c1).
            ant_vals: Optional[dict] = None
            if a_ant is not None:
                try:
                    ant_vals = _row(
                        "actas",
                        "valor_comp_ambiental, valor_comp_social, valor_comp_pmt, "
                        "valor_cobrado_adicional, ajuste_iccp, ajuste_icociv, ajuste_ipc",
                        id=int(a_ant),
                    )
                except (TypeError, ValueError):
                    ant_vals = None
            if not isinstance(ant_vals, dict):
                ant_vals = None

            def _a2(k: str) -> float:
                if not ant_vals:
                    return 0.0
                try:
                    v = float(ant_vals.get(k) or 0.0)
                except (TypeError, ValueError):
                    return 0.0
                if not math.isfinite(v):
                    return 0.0
                return float(round(v, 0))

            def _a2r(k: str) -> float:
                if not ant_vals:
                    return 0.0
                try:
                    v = float(ant_vals.get(k) or 0.0)
                except (TypeError, ValueError):
                    return 0.0
                if not math.isfinite(v):
                    return 0.0
                return v

            c2_cobrado_adic = _a2("valor_cobrado_adicional")
            ap0 = int((pres or {}).get("id") or 0)
            cid0 = int(contrato_id or 0)
            if cid0 and ap0:
                econ_antes = _suma_historico_economia_rpo_antes_presente(cid0, ap0)
            else:
                econ_antes = {k: 0.0 for k in _ECON_OH_KEYS}
            c1_camb = _v1(vr_contr.get("valor_componente_ambiental"))
            c1_csoc = _v1(vr_contr.get("valor_componente_social"))
            c1_cpmt = _v1(vr_contr.get("valor_componente_pmt"))
            pres_econ: Optional[dict] = None
            if ap0:
                try:
                    pe_cols = (
                        "valor_comp_ambiental, valor_comp_social, valor_comp_pmt, "
                        "valor_cobrado_adicional, ajuste_iccp, ajuste_icociv, ajuste_ipc"
                    )
                    pres_econ = _row("actas", pe_cols, id=int(ap0)) or None
                except (TypeError, ValueError):
                    pres_econ = None
            if not isinstance(pres_econ, dict):
                pres_econ = None

            def _pv(k: str) -> float:
                if not pres_econ:
                    return 0.0
                try:
                    w = float(pres_econ.get(k) or 0.0)
                except (TypeError, ValueError):
                    w = 0.0
                if not math.isfinite(w):
                    w = 0.0
                return w

            def _vig_mas_c1(db_v: float, c1b: float) -> float:
                try:
                    w = float(db_v or 0.0)
                except (TypeError, ValueError):
                    w = 0.0
                if not math.isfinite(w):
                    w = 0.0
                return w if w else float(c1b)

            c3_amb = float(econ_antes.get("valor_comp_ambiental") or 0.0) + _vig_mas_c1(
                _pv("valor_comp_ambiental"), c1_camb
            )
            c3_soc = float(econ_antes.get("valor_comp_social") or 0.0) + _vig_mas_c1(
                _pv("valor_comp_social"), c1_csoc
            )
            c3_pmt = float(econ_antes.get("valor_comp_pmt") or 0.0) + _vig_mas_c1(
                _pv("valor_comp_pmt"), c1_cpmt
            )
            c3_cobrado_antes = float(econ_antes.get("valor_cobrado_adicional") or 0.0)
            try:
                pres_aj = float(
                    round(
                        _pv("ajuste_iccp")
                        + _pv("ajuste_icociv")
                        + _pv("ajuste_ipc"),
                        0,
                    )
                )
            except (TypeError, ValueError):
                pres_aj = 0.0
            if not math.isfinite(pres_aj):
                pres_aj = 0.0
            c3_ajus_hist = float(
                round(
                    float(econ_antes.get("ajuste_iccp") or 0.0)
                    + float(econ_antes.get("ajuste_icociv") or 0.0)
                    + float(econ_antes.get("ajuste_ipc") or 0.0)
                    + pres_aj,
                    0,
                )
            )
            filas_t = [
                {
                    "id": "comp_amb",
                    "label": "Componente Ambiental",
                    "c1": c1_camb,
                    "c2": _a2("valor_comp_ambiental"),
                    "c3": float(round(c3_amb, 0)),
                    "c4": _v4(vr_contr.get("valor_componente_ambiental")),
                },
                {
                    "id": "comp_soc",
                    "label": "Componente Social",
                    "c1": c1_csoc,
                    "c2": _a2("valor_comp_social"),
                    "c3": float(round(c3_soc, 0)),
                    "c4": _v4(vr_contr.get("valor_componente_social")),
                },
                {
                    "id": "comp_pmt",
                    "label": "Componente PMT",
                    "c1": c1_cpmt,
                    "c2": _a2("valor_comp_pmt"),
                    "c3": float(round(c3_pmt, 0)),
                    "c4": _v4(vr_contr.get("valor_componente_pmt")),
                },
            ]
            raw_ads: Any = vr_contr.get("costos_adicionales_lista")
            if isinstance(raw_ads, str) and str(raw_ads).strip():
                try:
                    raw_ads = json.loads(raw_ads)
                except (TypeError, ValueError, json.JSONDecodeError):
                    raw_ads = []
            if not isinstance(raw_ads, list):
                raw_ads = []
            suma_c1_adic = 0.0
            suma_c4_adic = 0.0
            adic_c2_puesto = False
            adic_c3_puesto = False
            for idx, it in enumerate(raw_ads):
                if not isinstance(it, dict):
                    continue
                cap = (str(it.get("concepto_contractual") or "")).strip()
                if not cap:
                    continue
                vmm: float
                try:
                    vmm = float(it.get("valor_mensual") or 0.0)
                except (TypeError, ValueError):
                    vmm = 0.0
                if vmm == 0.0 and (it.get("valor_mensual") in (None, "", "null")) and it.get("valor") is not None:
                    try:
                        vtot_leg = float(it.get("valor") or 0.0)
                    except (TypeError, ValueError):
                        vtot_leg = 0.0
                    try:
                        tm_leg = float(it.get("tiempo_meses") or 0.0) or 0.0
                    except (TypeError, ValueError):
                        tm_leg = 0.0
                    dtm = max(tm_leg, 1.0) if vtot_leg or tm_leg else 1.0
                    vmm = vtot_leg / dtm
                # Col.1: mensual completo (COP) entero. Col.4: misma tasa de obra que componentes = factor4 × mensual
                c1_ads = float(round(float(vmm), 0))
                c4_ads = float(round(factor4 * c1_ads, 0))
                suma_c1_adic += c1_ads
                suma_c4_adic += c4_ads
                c2_line = 0.0
                if not adic_c2_puesto:
                    c2_line = c2_cobrado_adic
                    adic_c2_puesto = True
                c3_line = 0.0
                if not adic_c3_puesto:
                    c3_line = float(
                        round(
                            c3_cobrado_antes
                            + _vig_mas_c1(_pv("valor_cobrado_adicional"), float(c1_ads)),
                            0,
                        )
                    )
                    adic_c3_puesto = True
                filas_t.append(
                    {
                        "id": f"adic_{idx}",
                        "label": cap,
                        "c1": c1_ads,
                        "c2": c2_line,
                        "c3": c3_line,
                        "c4": c4_ads,
                    }
                )
            c1_cd_aiu_aj = 0.0
            if isinstance(f_cdu, dict):
                c1_cd_aiu_aj = float(f_cdu.get("c1") or 0.0)
            else:
                c1_cd_aiu_aj = float(st.get("c1") or 0.0)
            c4_camb = _v4(vr_contr.get("valor_componente_ambiental"))
            c4_csoc = _v4(vr_contr.get("valor_componente_social"))
            c4_cpmt = _v4(vr_contr.get("valor_componente_pmt"))
            c4_cd_aiu_aj = 0.0
            if isinstance(f_cdu, dict):
                c4_cd_aiu_aj = float(f_cdu.get("c4") or 0.0)
            else:
                c4_cd_aiu_aj = float(st.get("c4") or 0.0)
            base_aj = float(c1_cd_aiu_aj) + float(c1_camb) + float(c1_csoc) + float(c1_cpmt) + float(suma_c1_adic)
            base4_aj = float(c4_cd_aiu_aj) + float(c4_camb) + float(c4_csoc) + float(c4_cpmt) + float(suma_c4_adic)
            # col.1/4 subtotal obra = costo directo (sin fila AIU)
            c1_directo_sin_aiu = float(st.get("c1") or 0.0)
            c4_directo_sin_aiu = float(st.get("c4") or 0.0)
            p_aj = _tasa_fraccion_desde_valor_contrato(
                (pres or {}).get("pct_proyectado_ajustes")
                if isinstance(pres, dict)
                else None
            )
            if int(contrato_id or 0) == 2:
                c1_ajustes = float(round(c1_directo_sin_aiu * float(p_aj), 0))
                c4_ajustes = float(round(c4_directo_sin_aiu * float(p_aj), 0))
            else:
                c1_ajustes = float(round(float(base_aj) * float(p_aj), 0))
                c4_ajustes = float(round(float(base4_aj) * float(p_aj), 0))
            c1_vtot_obr = float(round(float(base_aj) + c1_ajustes, 0))
            c4_vtot_obr = float(round(float(base4_aj) + c4_ajustes, 0))
            c2_cd_aiu_aj = 0.0
            if isinstance(f_cdu, dict):
                c2_cd_aiu_aj = float(f_cdu.get("c2") or 0.0)
            else:
                c2_cd_aiu_aj = float(st.get("c2") or 0.0)
            c2_ajustes = float(
                round(_a2r("ajuste_iccp") + _a2r("ajuste_icociv") + _a2r("ajuste_ipc"), 0)
            )
            c2_camb = _a2("valor_comp_ambiental")
            c2_csoc = _a2("valor_comp_social")
            c2_cpmt = _a2("valor_comp_pmt")
            base2_aj = (
                float(c2_cd_aiu_aj)
                + float(c2_camb)
                + float(c2_csoc)
                + float(c2_cpmt)
                + float(c2_cobrado_adic)
            )
            c2_vtot_obr = float(round(float(base2_aj) + float(c2_ajustes), 0))
            c3_cd_aiu_aj = 0.0
            if isinstance(f_cdu, dict):
                c3_cd_aiu_aj = float(f_cdu.get("c3") or 0.0)
            else:
                c3_cd_aiu_aj = float(st.get("c3") or 0.0)
            c3_cobrado_h = 0.0
            for _f in filas_t:
                if not isinstance(_f, dict) or not str(_f.get("id") or "").startswith("adic_"):
                    continue
                try:
                    c3_cobrado_h = float(_f.get("c3") or 0.0)
                except (TypeError, ValueError):
                    c3_cobrado_h = 0.0
                break
            if not c3_cobrado_h and c3_cobrado_antes:
                c3_cobrado_h = c3_cobrado_antes
            base3_aj = (
                float(c3_cd_aiu_aj)
                + float(c3_amb)
                + float(c3_soc)
                + float(c3_pmt)
                + float(c3_cobrado_h)
            )
            c3_vtot_obr = float(round(float(base3_aj) + float(c3_ajus_hist), 0))
            filas_t += [
                {
                    "id": "ajustes",
                    "label": "Ajustes",
                    "c1": c1_ajustes,
                    "c2": c2_ajustes,
                    "c3": c3_ajus_hist,
                    "c4": c4_ajustes,
                },
                {
                    "id": "vtot_aj",
                    "label": "VALOR TOTAL OBRA CON AIU Y AJUSTES",
                    "c1": c1_vtot_obr,
                    "c2": c2_vtot_obr,
                    "c3": c3_vtot_obr,
                    "c4": c4_vtot_obr,
                },
            ]
            out["filas_tras_costo_directo_mas_aiu"] = filas_t
            out["_calc_componentes_meta"] = {
                "n_acta_cd_aiu_c1": n_acta_r,
                "n_contrato_cd_aiu": n_con,
                "factor_sobre_vr_componente": round(factor, 12) if factor else 0.0,
            }
        if bid == "ensayos" and iva > 0.0 and (st.get("c1", 0) or st.get("c2", 0) or st.get("c3", 0) or st.get("c4", 0)):
            out["fila_iva"] = {
                "c1": st["c1"] * iva,
                "c2": st["c2"] * iva,
                "c3": st["c3"] * iva,
                "c4": st["c4"] * iva,
            }
            out["fila_directo_mas_iva"] = {
                "c1": st["c1"] + out["fila_iva"]["c1"],
                "c2": st["c2"] + out["fila_iva"]["c2"],
                "c3": st["c3"] + out["fila_iva"]["c3"],
                "c4": st["c4"] + out["fila_iva"]["c4"],
            }
            fd0 = out["fila_directo_mas_iva"]
            out["fila_valor_total_obra_iva"] = {
                "c1": float(fd0.get("c1") or 0.0),
                "c2": float(fd0.get("c2") or 0.0),
                "c3": float(fd0.get("c3") or 0.0),
                "c4": float(fd0.get("c4") or 0.0),
            }
        return out

    blq = [
        _fila_bloque("obra", "Obra ejecutada (directo con AIU — sección SICOE / matriz)"),
        _fila_bloque("ensayos", "Ensayos, sondeos y 14/15 (IVA — sección SICOE / matriz)"),
    ]
    c_vtot_ob: Dict[str, float] = {"c1": 0.0, "c2": 0.0, "c3": 0.0, "c4": 0.0}
    c_ens_iva: Dict[str, float] = {"c1": 0.0, "c2": 0.0, "c3": 0.0, "c4": 0.0}
    for _b in blq:
        if not isinstance(_b, dict):
            continue
        if _b.get("id") == "obra":
            for _ft in _b.get("filas_tras_costo_directo_mas_aiu") or []:
                if not isinstance(_ft, dict) or str(_ft.get("id") or "") != "vtot_aj":
                    continue
                c_vtot_ob = {
                    "c1": float(_ft.get("c1") or 0.0),
                    "c2": float(_ft.get("c2") or 0.0),
                    "c3": float(_ft.get("c3") or 0.0),
                    "c4": float(_ft.get("c4") or 0.0),
                }
        if _b.get("id") == "ensayos":
            fdiv = _b.get("fila_valor_total_obra_iva") or _b.get("fila_directo_mas_iva")
            if isinstance(fdiv, dict):
                c_ens_iva = {
                    "c1": float(fdiv.get("c1") or 0.0),
                    "c2": float(fdiv.get("c2") or 0.0),
                    "c3": float(fdiv.get("c3") or 0.0),
                    "c4": float(fdiv.get("c4") or 0.0),
                }
    totales_doc = {
        "c1": float(c_vtot_ob.get("c1", 0.0) + c_ens_iva.get("c1", 0.0)),
        "c2": float(c_vtot_ob.get("c2", 0.0) + c_ens_iva.get("c2", 0.0)),
        "c3": float(c_vtot_ob.get("c3", 0.0) + c_ens_iva.get("c3", 0.0)),
        "c4": float(c_vtot_ob.get("c4", 0.0) + c_ens_iva.get("c4", 0.0)),
    }
    filas_unidas: List[Dict[str, Any]] = []
    for b in blq:
        for f in b["filas"]:
            filas_unidas.append(f)
    return {
        "acta_presente": pres,
        "acta_anterior": ant,
        "criterio": (
            "C1: acta **presente** = RPO en período (hoy ∈ [inicio, fin]), alineada a matriz SICOE; matriz «HABILITADO» — costo directo con ítem. "
            "C2: acta anterior en consecutivo, cascada N1·N2·N3. "
            "C3: total aprobado solo interventoría (nivel 3), acumulado hasta el acta presente. "
            "C4: pendiente en al menos un nivel. Tras subtotal bloque obra: fila AIU = subtotal·tasa AIU (contrato). "
            "Tras subtotal bloque ensayos/14-15: fila IVA = subtotal·tasa IVA (contrato, p. ej. 0,19=19 %)."
        ),
        "aiu_pactado": aiu,
        "iva_pactado": iva,
        "resolucion": {
            "consecutivo_presente": c_pres,
            "ids_cascade_acumulado": ids_cascade_hasta,
        },
        "validacion_misma_que_lista_actas": {
            "cascade_acta_presente_cdirecto_lista": float(t_presente_lista),
            "cascade_acta_anterior_cdirecto": float(t2) if a_ant is not None else None,
            "cascade_acumulado_todas_las_rpo_cdirecto": float(t_acc),
        },
        "filas_por_bloque": blq,
        "totales": {
            "c1": totales_doc["c1"],
            "c2": totales_doc["c2"],
            "c3": totales_doc["c3"],
            "c4": totales_doc["c4"],
        },
        "valor_total_acta": {
            "c1": totales_doc["c1"],
            "c2": totales_doc["c2"],
            "c3": totales_doc["c3"],
            "c4": totales_doc["c4"],
        },
        "filas_orden_capitulo": filas_unidas,
        "secciones_presente": d_p.get("secciones") or {},
    }


def _construir_datos_informe_gerencia_matriz(
    contrato_id: int,
    acta_presente_override: Optional[int] = None,
    *,
    skip_cache: bool = False,
) -> Dict[str, Any]:
    cache_key = _gerencia_matriz_cache_key(contrato_id, acta_presente_override)
    if not skip_cache:
        cached = _gerencia_matriz_cache_get(cache_key)
        if cached is not None:
            return cached

    pres, ant, actas_asc = _resolver_acta_gerencia_presente_y_anterior(
        int(contrato_id), acta_presente_override
    )
    if not pres or not pres.get("id"):
        raise HTTPException(404, "No hay actas RPO (cobro) en el contrato para el informe de gerencia")
    ap_id = int(pres["id"])
    c_pres = int(pres.get("consecutivo") or 0) or 0
    ids_cascade_hasta: List[int] = [
        int(a["id"])
        for a in actas_asc
        if a.get("id") is not None and int(a.get("consecutivo") or 0) <= c_pres
    ]
    a_ant = int(ant["id"]) if (ant and ant.get("id")) else None
    _seen_rsum: set = set()
    ids_rsum: List[int] = []
    for x in [ap_id] + ([int(a_ant)] if a_ant is not None else []) + list(ids_cascade_hasta):
        try:
            xi = int(x)
        except (TypeError, ValueError):
            continue
        if xi not in _seen_rsum:
            _seen_rsum.add(xi)
            ids_rsum.append(xi)

    def _rpc_resumen() -> Dict[Any, Any]:
        return rpo_resumen_actas_rpc(_sb, contrato_id, ids_rsum) or {}

    def _row_contrato() -> Dict[str, Any]:
        return (
            _row(
                "contratos",
                "aiu, iva, costo_directo_contrato, valor_componente_ambiental, "
                "valor_componente_social, valor_componente_pmt, costos_adicionales_lista",
                id=contrato_id,
            )
            or {}
        )

    def _rpc_maps() -> Optional[Dict[str, Dict[Tuple[str, str], float]]]:
        return informe_gerencia_matriz_maps_por_rpc(
            _sb, contrato_id, ap_id, a_ant, ids_cascade_hasta, None
        )

    def _rpc_conc_p() -> Optional[Dict[str, Any]]:
        return rpo_conciliacion_un_acta_rpc(_sb, contrato_id, ap_id)

    with ThreadPoolExecutor(max_workers=4) as _pool_ger:
        _f_res = _pool_ger.submit(_rpc_resumen)
        _f_ctr = _pool_ger.submit(_row_contrato)
        _f_map = _pool_ger.submit(_rpc_maps)
        _f_con = _pool_ger.submit(_rpc_conc_p)
        r_all = _f_res.result()
        ctr_r = _f_ctr.result()
        maps = _f_map.result()
        d_p = _f_con.result()

    if not isinstance(r_all, dict):
        r_all = {}

    t_presente_lista = float((r_all.get(int(ap_id)) or {}).get("costo_directo_total", 0) or 0)
    t2 = 0.0
    if a_ant is not None:
        t2 = float((r_all.get(int(a_ant)) or {}).get("costo_directo_total", 0) or 0)
    t_acc = 0.0
    for _oid in ids_cascade_hasta:
        t_acc += float((r_all.get(int(_oid)) or {}).get("costo_directo_total", 0) or 0)
    aiu_c: Optional[float] = None
    iva_c: Optional[float] = None
    if ctr_r.get("aiu") is not None:
        try:
            aiu_c = float(ctr_r["aiu"])
        except (TypeError, ValueError):
            aiu_c = None
    if ctr_r.get("iva") is not None:
        try:
            iva_c = float(ctr_r["iva"])
        except (TypeError, ValueError):
            iva_c = None

    if d_p is None:
        d_p = _rpo_gerencia_vacio()
        _log.warning(
            "CC-GER-001: rpo_panel_acta_por_capitulo_bloque no disponible; secciones AIU en cero."
        )
    if maps is None:
        _log.warning(
            "CC-GER-001: RPC informe gerencia incompleto; usando fallback Python (lento, datos completos). "
            "Verifique funciones rpo_ger_* en Supabase (backend/sql/rpo_informe_gerencia.sql)."
        )
        result = _construir_datos_informe_gerencia_matriz_fallback_por_capitulo(
            contrato_id, pres, ant, ap_id, c_pres, a_ant, ids_cascade_hasta, t2, t_acc, aiu_c, iva_c, ctr_r
        )
        _gerencia_matriz_cache_set(cache_key, result)
        return result

    c1, c2, c3, c4 = maps["c1"], maps["c2"], maps["c3"], maps["c4"]
    result = _embalaje_informe_gerencia_bloques(
        pres,
        ant,
        c_pres,
        ids_cascade_hasta,
        t2,
        t_acc,
        t_presente_lista,
        c1,
        c2,
        c3,
        c4,
        d_p,
        a_ant,
        aiu_c,
        iva_c,
        vr_contr=ctr_r,
        contrato_id=contrato_id,
    )
    _gerencia_matriz_cache_set(cache_key, result)
    return result


def _acta_cab_gerencia(acta_id: int) -> Dict[str, Any]:
    r = _row("actas", "id, numero_rpo, consecutivo, fecha_inicio, fecha_fin, tipo_grupo", id=acta_id) or {}
    if not r:
        return r
    return r


def _html_informe_gerencia_cc_ger_001(
    contrato: dict,
    acta_p: dict,
    acta_a: Optional[dict],
    d_p: dict,
    d_a: Optional[dict],
    *,
    usuario_nombre: str,
    usuario_cargo: str,
    formato_ccd: str,
    contexto_tipo: str,
    contexto_id: int,
    firma_cfg: Optional[Dict[str, Any]],
    elaboro_firma_data_uri: Optional[str],
    reviso_firma_data_uri: Optional[str],
    aprobo_firma_data_uri: Optional[str],
) -> str:
    """PDF CC-GER-001: matriz SICOE (N1 N2 N3) por capítulo; dos actas; firmas terciarias."""
    bd = "border:1px solid #9ca3af"
    bd_blk = "border:1px solid #1f2937"
    fmt = formato_ccd
    est = _merge_estilo_pdf((firma_cfg or {}).get("estilo_pdf"), fmt)
    fc = firma_cfg or {}
    logo_html = _html_logo_contratista(contrato)
    fecha_gen = _fmt_informe_fecha_generacion()
    contratista_nom = _h(str(contrato.get("contratista") or ""))
    interv = _h(str(contrato.get("interventoria") or ""))
    nit_raw = str(contrato.get("nit") or "").strip()
    nit_en_valor = f' <span style="font-size:6.5pt;color:#444;">(NIT: {_h(nit_raw)})</span>' if nit_raw else ""
    lbl = "font-size:6pt;font-weight:bold;color:#111;text-transform:uppercase;letter-spacing:0.2px;"
    und = "border-bottom:1px solid #1f2937;font-size:7pt;padding:1px 0 2px 0;margin-top:1px;"
    objeto = str(contrato.get("objeto") or "—")
    mrg = _merge_filas_gerencia_dos_actas(d_p.get("por_capitulo"), d_a.get("por_capitulo") if d_a else None)
    t_p = float(d_p.get("costo_directo_total") or 0)
    t_a = float(d_a.get("costo_directo_total") or 0) if d_a else 0.0

    def _lbl_acta(ra: dict) -> str:
        n = str(ra.get("numero_rpo") or ra.get("consecutivo") or "—")
        fi = str((ra.get("fecha_inicio") or "") or "")[:10] or "—"
        ff = str((ra.get("fecha_fin") or "") or "")[:10] or "—"
        return f"RPO {n} — {fi} → {ff} (consecutivo: {str(ra.get('consecutivo') or '—')})"

    c3l, c3v = "ACTA (PRESENTE — CCD / FIRMA)", _lbl_acta(acta_p)
    if acta_a:
        c4l, c4v = "ACTA (REFERENCIA / COMPARACIÓN)", _lbl_acta(acta_a)
    else:
        c4l, c4v = "ACTA (REFERENCIA / COMPARACIÓN)", "— (sin acta de referencia; solo se lista el acta presente)"

    elaboro_n = _h(str(fc.get("elaboro_nombre") or "").strip() or "—")
    elaboro_c = _h(str(fc.get("elaboro_cargo") or "").strip() or "—")
    reviso_n = _h(str(fc.get("reviso_nombre") or "").strip() or "—")
    reviso_c = _h(str(fc.get("reviso_cargo") or "").strip() or "—")
    aprobo_n = _h(str(fc.get("aprobo_nombre") or "").strip() or "—")
    aprobo_c = _h(str(fc.get("aprobo_cargo") or "").strip() or "—")

    thead_bg = _sanitize_ccd_hex_color(est.get("thead_bg"), "#1e3a5f")
    reven = _sanitize_ccd_hex_color(est.get("row_even_bg"), "#f8fafc")
    rodd = _sanitize_ccd_hex_color(est.get("row_odd_bg"), "#e5e7eb")

    filas = []
    for i, (cap, vp, va, dlt) in enumerate(mrg):
        rbg = reven if i % 2 == 0 else rodd
        cshow = (cap if len(cap) <= 100 else cap[:97] + "…")
        if acta_a:
            filas.append(
                f'<tr style="background:{rbg}">'
                f'<td style="{bd};padding:3px 4px;font-size:7pt;text-align:left;word-wrap:break-word">{_h(cshow)}</td>'
                f'<td style="{bd};padding:3px 4px;font-size:7pt;text-align:right">{_fm(vp)}</td>'
                f'<td style="{bd};padding:3px 4px;font-size:7pt;text-align:right">{_fm(va)}</td>'
                f'<td style="{bd};padding:3px 4px;font-size:7pt;text-align:right;font-weight:600">{_fm(dlt)}</td>'
                f"</tr>"
            )
        else:
            filas.append(
                f'<tr style="background:{rbg}">'
                f'<td style="{bd};padding:3px 4px;font-size:7pt;text-align:left;word-wrap:break-word">{_h(cshow)}</td>'
                f'<td style="{bd};padding:3px 4px;font-size:7pt;text-align:right">{_fm(vp)}</td>'
                f"</tr>"
            )
    tbody = "".join(filas) if filas else (
        f'<tr><td colspan="{4 if acta_a else 2}" style="{bd};padding:6px;font-size:7pt;color:#6b7280">'
        "Sin capítulos con registro aprobado en matriz (revisa sincronización SICOE / acta RPO).</td></tr>"
    )
    tot_row = (
        f'<tr style="background:{_sanitize_ccd_hex_color(est.get("subtotal_bg"), "#dbeafe")}">'
        f'<td style="{bd};padding:4px 6px;font-size:7.5pt;font-weight:800;text-align:right">'
        f'TOTAL costo directo aprob. (N1·N2·N3 en cascada)</td>'
        f'<td style="{bd};padding:4px 6px;font-size:7.5pt;font-weight:800;text-align:right">{_fm(t_p)}</td>'
    )
    if acta_a:
        tot_row += (
            f'<td style="{bd};padding:4px 6px;font-size:7.5pt;font-weight:800;text-align:right">{_fm(t_a)}</td>'
            f'<td style="{bd};padding:4px 6px;font-size:7.5pt;font-weight:800;text-align:right">{_fm(t_p - t_a)}</td>'
        )
    tot_row += "</tr>"

    pie = (
        f"Informe gerencia · criterio matriz: mismos N1, N2 y N3 aprobado en SICOE Obra que en conciliación. "
        f"Contexto CCD: {contexto_tipo} {contexto_id}."
    )

    th_act = (
        f'<th style="{bd};background:{thead_bg};color:#fff;padding:3px 4px;font-size:6.5pt">CAPÍTULO (obra según título SICOE)</th>'
        f'<th style="{bd};background:{thead_bg};color:#fff;padding:3px 4px;font-size:6.5pt">ACTA PRESENTE · COP$</th>'
    )
    if acta_a:
        th_act += (
            f'<th style="{bd};background:{thead_bg};color:#fff;padding:3px 4px;font-size:6.5pt">'
            f"ACTA REFERENCIA · COP$</th>"
            f'<th style="{bd};background:{thead_bg};color:#fff;padding:3px 4px;font-size:6.5pt">DIF. · COP$</th>'
        )

    elaboro_td = _html_cc_sub_td_firma_columna(bd, "Elaboró:", elaboro_n, elaboro_c, elaboro_firma_data_uri)
    reviso_td = _html_cc_sub_td_firma_columna(bd, "Revisó:", reviso_n, reviso_c, reviso_firma_data_uri)
    aprobo_td = _html_cc_sub_td_firma_columna(bd, "Aprobó:", aprobo_n, aprobo_c, aprobo_firma_data_uri)

    secc = d_p.get("secciones") or {}
    sec_txt = []
    for _k, blo in (secc.items() if isinstance(secc, dict) else []):
        if isinstance(blo, dict) and blo.get("titulo") and (blo.get("subtotal") is not None):
            sec_txt.append(f"{_h(str(blo.get('titulo') or ''))} · {_h(_fm(blo.get('subtotal')))}")
    sum_sec = f"<p style=\"font-size:6.5pt;color:#374151;margin:6px 0 4px 0\">{' &nbsp;|&nbsp; '.join(sec_txt)}</p>" if sec_txt else ""

    return f"""<!DOCTYPE html>
<html xmlns="http://www.w3.org/1999/xhtml" xmlns:pdf="http://www.xhtml2pdf.org/pdf">
<head><meta charset="UTF-8"/><title>{_h(formato_ccd)}</title>
<style type="text/css">
@page {{ size: letter; margin: 8mm 10mm; }}
</style></head>
<body style="margin:0;padding:4px;font-family:Arial,Helvetica,sans-serif;font-size:7.5pt;color:#111;">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd_blk};table-layout:fixed;">
<tr>
<td style="width:20%;{bd_blk};vertical-align:middle;padding:2px;text-align:center;background:#fff">
{logo_html}
</td>
<td style="width:48%;{bd_blk};vertical-align:middle;text-align:center;font-weight:bold;font-size:7.5pt;padding:3px 5px;line-height:1.1;">
INFORME DE AVANCE (GERENCIAL) — EJECUCIÓN DE OBRA
</td>
<td style="width:32%;{bd_blk};vertical-align:middle;text-align:center;padding:3px 5px;">
<div style="color:#1e3a8a;font-weight:bold;font-size:12pt;letter-spacing:0.3px;">{_h(formato_ccd)}</div>
<div style="font-size:8.5pt;color:#1e3a8a;font-weight:bold;">CCD · ClaraCore</div>
</td>
</tr>
<tr><td colspan="3" style="padding:0;border:none;">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd_blk};background:#fff">
<tr><td style="padding:2px 6px;border:none;">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:none;">
<tr>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">CONTRATO</div>
<div style="{und}">{_h(contrato.get("numero", ""))}</div>
</td>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">FECHA DE EMISIÓN</div>
<div style="{und}">{_h(fecha_gen)}</div>
</td>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">{_h(c3l)}</div>
<div style="{und}">{_h(c3v)}</div>
</td>
<td style="width:25%;padding:0 0 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">{_h(c4l)}</div>
<div style="{und}">{_h(c4v)}</div>
</td>
</tr>
<tr>
<td colspan="2" style="padding:3px 5px 1px 0;border:none;vertical-align:top;">
<div style="{lbl}">CONTRATISTA</div>
<div style="{und}">{contratista_nom}{nit_en_valor}</div>
</td>
<td colspan="2" style="padding:3px 0 1px 0;border:none;vertical-align:top;">
<div style="{lbl}">INTERVENTORÍA</div>
<div style="{und}">{interv}</div>
</td>
</tr>
<tr>
<td colspan="4" style="padding:4px 5px 0 0;border:none;vertical-align:top;">
<div style="{lbl}">OBJETO</div>
<div style="border-bottom:1px solid #1f2937;font-size:7pt;padding:1px 0 2px 0;word-wrap:break-word">{_h(objeto)}</div>
</td>
</tr>
</table>
</td></tr>
</table>
</td></tr>
</table>
{sum_sec}
<div style="font-size:6.3pt;color:#4b5563;margin:4px 0 2px 0">Datos: sincronizados con SICOE Obra; costo = línea directa aprobada en interventoría, sin Excel intermedio.</div>
<table width="100%" cellspacing="0" cellpadding="0" class="cc001-tabla-ger" style="border-collapse:collapse;">
<thead><tr>{th_act}</tr></thead>
<tbody>
{tbody}
{tot_row}
</tbody>
</table>
<div class="ccd-cc001-firmas-wrap" style="margin-top:4mm;page-break-inside:avoid;">
<table class="ccd-cc001-firmas-tbl" style="width:100%;border-collapse:collapse" cellspacing="0" cellpadding="0">
<tr>
{elaboro_td}
{reviso_td}
{aprobo_td}
</tr>
</table>
</div>
<p style="font-size:6pt;color:#64748b;margin-top:6px;text-align:center;">{_h(pie)} · Generado ClaraCore · {_h(usuario_cargo)} · {_h(usuario_nombre)}</p>
</body></html>"""


def _html_informe_gerencia_cc_ger_001_matriz(
    contrato: dict,
    datos: Dict[str, Any],
    *,
    usuario_nombre: str,
    usuario_cargo: str,
    formato_ccd: str,
    ap_id: int,
    firma_cfg: Optional[Dict[str, Any]],
    elaboro_firma_data_uri: Optional[str],
    reviso_firma_data_uri: Optional[str],
    aprobo_firma_data_uri: Optional[str],
) -> str:
    """CC-GER-001: 4 columnas, horizontal, compacto; Obra/AIU y Ensayos/IVA con subtotales independientes."""
    bd = "border:0.4px solid #b8c5d0"
    bd_blk = "border:0.4px solid #9ca3af"
    fmt = formato_ccd
    est = _merge_estilo_pdf((firma_cfg or {}).get("estilo_pdf"), fmt)
    fc = firma_cfg or {}
    logo_html = _html_logo_contratista(contrato, compact=True, compact_box_height="0.9cm")
    fecha_gen = _fmt_informe_fecha_generacion()
    contratista_nom = _h(str(contrato.get("contratista") or ""))
    interv = _h(str(contrato.get("interventoria") or ""))
    nit_raw = str(contrato.get("nit") or "").strip()
    nit_en_valor = f' <span style="font-size:4.2pt;color:#444;">(NIT: {_h(nit_raw)})</span>' if nit_raw else ""
    # Encabezado compacto (una sola hoja): no reducir contenido, solo padding/tipos.
    lbl = "font-size:4.5pt;font-weight:bold;color:#111;text-transform:uppercase;letter-spacing:0.1px;padding-top:0;line-height:1.05;"
    und = "border-bottom:0.4px solid #4b5563;font-size:4.65pt;padding:0.5px 0 0.5px 0;margin-top:0;line-height:1.08;word-wrap:break-word;"
    # Bloque metadatos (CONTRATO…OBJETO, recuadro bajo título, no fila de logos): +20% altura / aire.
    _meta_h = 1.2
    _pye = 0.5 * _meta_h
    lbl_m = (
        f"font-size:4.5pt;font-weight:bold;color:#111;text-transform:uppercase;letter-spacing:0.1px;"
        f"padding-top:0;line-height:{1.05 * _meta_h:.2f};"
    )
    und_m = (
        f"border-bottom:0.4px solid #4b5563;font-size:4.65pt;padding:{_pye:.2f}px 0 {_pye:.2f}px 0;margin-top:0;"
        f"line-height:{1.08 * _meta_h:.2f};word-wrap:break-word;"
    )
    pad_m4 = f"padding:{_pye:.2f}px 2px {_pye:.2f}px 0"
    pad_m4l = f"padding:{_pye:.2f}px 0 {_pye:.2f}px 0"
    pad_m2l = f"padding:{_pye:.2f}px 3px {_pye:.2f}px 0"
    pad_mct1 = f"padding:{_pye:.2f}px 2px {_pye:.2f}px 0"
    pad_mct2 = f"padding:{_pye:.2f}px 0 {_pye:.2f}px 0"
    pad_mob = f"padding:{_pye:.2f}px 0 {_pye:.2f}px 3px"
    objeto = str(contrato.get("objeto") or "—")
    ap = (datos.get("acta_presente") or {}) or {}
    aa = (datos.get("acta_anterior") or {}) or {}
    val = (datos.get("validacion_misma_que_lista_actas") or {}) or {}
    tit_acta = f"RPO {ap.get('numero_rpo') or ap.get('consecutivo') or '—'}"
    if aa and aa.get("id"):
        subt_ref = f"RPO ant. {aa.get('numero_rpo') or aa.get('consecutivo') or '—'}"
    else:
        subt_ref = "RPO ant.: —"
    t = (datos.get("valor_total_acta") or datos.get("totales") or {}) or {}
    aiu_pac = float(datos.get("aiu_pactado") or 0) or 0.0
    iva_pac = float(datos.get("iva_pactado") or 0) or 0.0
    aiu_lbl = f"AIU ({(aiu_pac * 100):g} % contrato)"
    iva_lbl = f"IVA ({(iva_pac * 100):g} % contrato)"
    thead_bg = _sanitize_ccd_hex_color(est.get("thead_bg"), "#1e3a5f")
    reven = _sanitize_ccd_hex_color(est.get("row_even_bg"), "#f8fafc")
    rodd = _sanitize_ccd_hex_color(est.get("row_odd_bg"), "#eef2f7")
    sub_bg = _sanitize_ccd_hex_color(est.get("subtotal_bg"), "#bfdbfe")
    g_tit_blo = _sanitize_ccd_hex_color(est.get("ger_titulo_bloque_bg"), sub_bg)
    g_su_aiu = _sanitize_ccd_hex_color(est.get("ger_subtotal_obra_con_aiu_bg"), "#e0f2fe")
    g_ta_aiu = _sanitize_ccd_hex_color(est.get("ger_fila_tasa_aiu_bg"), "#dbeafe")
    g_cdu = _sanitize_ccd_hex_color(est.get("ger_cdirecto_mas_aiu_bg"), "#c7d8f0")
    g_post = _sanitize_ccd_hex_color(est.get("ger_filas_post_cdu_bg"), "#e8edf5")
    g_vo_aj = _sanitize_ccd_hex_color(est.get("ger_vtot_obra_ajustes_bg"), "#a8bfdb")
    g_su_iva = _sanitize_ccd_hex_color(est.get("ger_subtotal_obra_con_iva_bg"), "#e0f2fe")
    g_ta_iva = _sanitize_ccd_hex_color(est.get("ger_fila_tasa_iva_bg"), "#e8eeff")
    g_cdiva = _sanitize_ccd_hex_color(est.get("ger_cdirecto_mas_iva_bg"), "#d4dcf5")
    g_v_iva = _sanitize_ccd_hex_color(est.get("ger_vtot_obra_iva_bg"), "#c3d0f0")
    g_vacta = _sanitize_ccd_hex_color(est.get("ger_valor_total_acta_bg"), "#93c5fd")
    sec_tit = "#0c4a6e"
    # Grilla: +1 pt respecto a la base compacta original (antes se probó +2; una hoja: +1).
    _dpt = 1.0
    fs = f"{4.9 + _dpt}pt"
    fs_th = f"{5.5 + _dpt}pt"
    fs_th_sub = f"{4.0 + _dpt}pt"
    pd_cell = "0.6px 1.6px"

    def _th_encab_cabecera_rpo(ra: Optional[dict], *, sin_referencia: str) -> str:
        if not ra or not ra.get("id"):
            return f'<div style="line-height:1.1;font-size:{fs_th}">{_h(sin_referencia)}</div>'
        n = str(ra.get("numero_rpo") or ra.get("id") or "—")
        c = str(ra.get("consecutivo") or "").strip() or "—"
        fi0 = (ra.get("fecha_inicio") or "") or ""
        ff0 = (ra.get("fecha_fin") or "") or ""
        fi = (str(fi0)[:10] if fi0 else "—") or "—"
        ff = (str(ff0)[:10] if ff0 else "—") or "—"
        l1 = f"RPO {n} · cons. {c}"
        l2 = f"{fi} → {ff}"
        return (
            f'<div style="line-height:1.0;font-size:{fs_th}">{_h(l1)}</div>'
            f'<div style="line-height:1.0;font-size:{fs_th_sub};font-weight:600;opacity:0.95;padding-top:0">{_h(l2)}</div>'
        )

    # 10,2% c/u +15% de ancho → 11,73% c/u; capítulo el resto (xhtml2pdf: celdas + CSS nth-child).
    w_num = "11.73%"
    w_cap = "53.08%"  # 100 − 4×11,73%
    st_w_cap = f"width:{w_cap};max-width:{w_cap};min-width:0;box-sizing:border-box;"
    st_w_num = f"width:{w_num};max-width:{w_num};min-width:0;box-sizing:border-box;word-wrap:break-word;"
    colgrp = (
        f'<colgroup>'
        f'<col style="{st_w_cap}" />'
        f'<col style="{st_w_num}" />'
        f'<col style="{st_w_num}" />'
        f'<col style="{st_w_num}" />'
        f'<col style="{st_w_num}" />'
        f"</colgroup>"
    )
    th_c1 = _th_encab_cabecera_rpo(ap if ap.get("id") else None, sin_referencia="— (sin acta presente)")
    th_c2 = _th_encab_cabecera_rpo(aa if (aa and aa.get("id")) else None, sin_referencia="— (sin RPO previa)")
    th = (
        f'<th style="{st_w_cap}{bd};background:{thead_bg};color:#fff;padding:0.5px 1px;font-size:{fs_th};line-height:1.05">'
        f"Capítulo SICOE</th>"
        f'<th style="{st_w_num}{bd};background:{thead_bg};color:#fff;padding:0.5px 0;font-size:{fs_th};line-height:1.05;text-align:center;vertical-align:middle">'
        f"{th_c1}</th>"
        f'<th style="{st_w_num}{bd};background:{thead_bg};color:#fff;padding:0.5px 0;font-size:{fs_th};line-height:1.05;text-align:center;vertical-align:middle">'
        f"{th_c2}</th>"
        f'<th style="{st_w_num}{bd};background:{thead_bg};color:#fff;padding:0.5px 0;font-size:{fs_th};line-height:1.05">'
        f"Total aprobados</th>"
        f'<th style="{st_w_num}{bd};background:{thead_bg};color:#fff;padding:0.5px 0;font-size:{fs_th};line-height:1.05">Total pendientes</th>'
    )
    # Sin max-width:0 (provocaba que la 1.ª col se colapsara y las numéricas comían el ancho en el PDF).
    td_cap = f"{st_w_cap}word-wrap:break-word;overflow-wrap:break-word;white-space:normal;{bd}"
    td_num = f"{st_w_num}{bd}"

    def _row_dat(r: Dict[str, Any], i: int) -> str:
        rbg = reven if i % 2 == 0 else rodd
        cap = str(r.get("capitulo") or "—")
        cshow = cap
        return (
            f'<tr style="background:{rbg}">'
            f'<td style="{td_cap};padding:{pd_cell};font-size:{fs};line-height:1.1">{_h(cshow)}</td>'
            f'<td style="{td_num};padding:{pd_cell};font-size:{fs};line-height:1.05;text-align:right">{_fm(float(r.get("c1") or 0))}</td>'
            f'<td style="{td_num};padding:{pd_cell};font-size:{fs};line-height:1.05;text-align:right">{_fm(float(r.get("c2") or 0))}</td>'
            f'<td style="{td_num};padding:{pd_cell};font-size:{fs};line-height:1.05;text-align:right">{_fm(float(r.get("c3") or 0))}</td>'
            f'<td style="{td_num};padding:{pd_cell};font-size:{fs};line-height:1.05;text-align:right">{_fm(float(r.get("c4") or 0))}</td>'
            f"</tr>"
        )

    bdy: list = []
    bloques = datos.get("filas_por_bloque")
    if isinstance(bloques, list) and bloques:
        for j, bsec in enumerate(bloques):
            if not isinstance(bsec, dict):
                continue
            titb = str(bsec.get("titulo") or "")
            bdy.append(
                f'<tr><td colspan="5" style="{bd};background:{g_tit_blo};color:{sec_tit};font-size:{4.7 + _dpt}pt;'
                f'font-weight:800;padding:0.5px 1px;">{_h(titb)}</td></tr>'
            )
            for i, frow in enumerate(bsec.get("filas") or []):
                if not isinstance(frow, dict):
                    continue
                bdy.append(_row_dat(frow, (j * 200) + i))
            st = bsec.get("subtotal") or {}
            st_lbl = "Subtotal"
            st_row_bg = g_su_aiu
            if bsec.get("id") == "obra":
                st_lbl = "Subtotal Obra con AIU"
                st_row_bg = g_su_aiu
            elif bsec.get("id") == "ensayos":
                st_lbl = "Subtotal Obra con IVA"
                st_row_bg = g_su_iva
            bdy.append(
                f'<tr style="background:{st_row_bg}">'
                f'<td style="{td_cap};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:left;padding:{pd_cell}">{_h(st_lbl)}</td>'
                f'<td style="{td_num};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">{_fm(float(st.get("c1") or 0))}</td>'
                f'<td style="{td_num};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">{_fm(float(st.get("c2") or 0))}</td>'
                f'<td style="{td_num};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">{_fm(float(st.get("c3") or 0))}</td>'
                f'<td style="{td_num};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">{_fm(float(st.get("c4") or 0))}</td>'
                f"</tr>"
            )
            f_aiu = bsec.get("fila_aiu")
            f_cd_aiu = bsec.get("fila_costo_directo_mas_aiu")
            if bsec.get("id") == "obra" and isinstance(f_aiu, dict):
                bdy.append(
                    f'<tr style="background:{g_ta_aiu}">'
                    f'<td style="{td_cap};font-size:{4.6 + _dpt}pt;font-weight:700;text-align:left;padding:{pd_cell}">'
                    f"{_h(aiu_lbl)}</td>"
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:700;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_aiu.get("c1") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:700;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_aiu.get("c2") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:700;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_aiu.get("c3") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:700;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_aiu.get("c4") or 0))}</td>'
                    f"</tr>"
                )
            if bsec.get("id") == "obra" and isinstance(f_cd_aiu, dict):
                bdy.append(
                    f'<tr style="background:{g_cdu}">'
                    f'<td style="{td_cap};font-size:{4.6 + _dpt}pt;font-weight:800;text-align:left;padding:{pd_cell}">'
                    f"Costo Directo + AIU</td>"
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_cd_aiu.get("c1") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_cd_aiu.get("c2") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_cd_aiu.get("c3") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_cd_aiu.get("c4") or 0))}</td>'
                    f"</tr>"
                )
            for f_aux in bsec.get("filas_tras_costo_directo_mas_aiu") or []:
                if not isinstance(f_aux, dict) or not f_aux.get("label"):
                    continue
                lab2 = str(f_aux.get("label") or "—")
                fx_id = str(f_aux.get("id") or "")
                fxb = g_vo_aj if fx_id == "vtot_aj" else g_post
                c1s = "—" if f_aux.get("c1") is None else _fm(float(f_aux.get("c1") or 0.0))
                c2s = "—" if f_aux.get("c2") is None else _fm(float(f_aux.get("c2") or 0.0))
                c3s = "—" if f_aux.get("c3") is None else _fm(float(f_aux.get("c3") or 0.0))
                c4s = "—" if f_aux.get("c4") is None else _fm(float(f_aux.get("c4") or 0.0))
                wfx = "800" if fx_id == "vtot_aj" else "700"
                bdy.append(
                    f'<tr style="background:{fxb}">'
                    f'<td style="{td_cap};font-size:{4.4 + _dpt}pt;font-weight:{wfx};text-align:left;padding:{pd_cell}">'
                    f"{_h(lab2)}</td>"
                    f'<td style="{td_num};font-size:{4.4 + _dpt}pt;font-weight:{wfx};text-align:right;padding:{pd_cell}">'
                    f"{c1s}</td>"
                    f'<td style="{td_num};font-size:{4.4 + _dpt}pt;font-weight:{wfx};text-align:right;padding:{pd_cell}">'
                    f"{c2s}</td>"
                    f'<td style="{td_num};font-size:{4.4 + _dpt}pt;font-weight:{wfx};text-align:right;padding:{pd_cell}">'
                    f"{c3s}</td>"
                    f'<td style="{td_num};font-size:{4.4 + _dpt}pt;font-weight:{wfx};text-align:right;padding:{pd_cell}">'
                    f"{c4s}</td>"
                    f"</tr>"
                )
            f_iva = bsec.get("fila_iva")
            f_vt_iva = bsec.get("fila_valor_total_obra_iva")
            if bsec.get("id") == "ensayos" and isinstance(f_iva, dict):
                bdy.append(
                    f'<tr style="background:{g_ta_iva}">'
                    f'<td style="{td_cap};font-size:{4.6 + _dpt}pt;font-weight:700;text-align:left;padding:{pd_cell}">'
                    f"{_h(iva_lbl)}</td>"
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:700;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_iva.get("c1") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:700;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_iva.get("c2") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:700;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_iva.get("c3") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:700;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_iva.get("c4") or 0))}</td>'
                    f"</tr>"
                )
            if bsec.get("id") == "ensayos" and isinstance(f_vt_iva, dict):
                bdy.append(
                    f'<tr style="background:{g_v_iva}">'
                    f'<td style="{td_cap};font-size:{4.6 + _dpt}pt;font-weight:800;text-align:left;padding:{pd_cell}">'
                    f"VALOR TOTAL OBRA CON IVA</td>"
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_vt_iva.get("c1") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_vt_iva.get("c2") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_vt_iva.get("c3") or 0))}</td>'
                    f'<td style="{td_num};font-size:{4.7 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">'
                    f'{_fm(float(f_vt_iva.get("c4") or 0))}</td>'
                    f"</tr>"
                )
    else:
        filas = datos.get("filas_orden_capitulo") or []
        for i, row in enumerate(filas):
            if not isinstance(row, dict):
                continue
            bdy.append(_row_dat(row, i))
    tbody = "".join(bdy) if bdy else (
        f'<tr><td colspan="5" style="{bd};padding:4px;font-size:{6 + _dpt}pt;color:#6b7280">Sin filas SICOE.</td></tr>'
    )
    tot_tr = (
        f'<tr style="background:{g_vacta}">'
        f'<td style="{td_cap};padding:{pd_cell};font-size:{4.8 + _dpt}pt;font-weight:800">VALOR TOTAL ACTA</td>'
        f'<td style="{td_num};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">{_fm(float(t.get("c1") or 0))}</td>'
        f'<td style="{td_num};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">{_fm(float(t.get("c2") or 0))}</td>'
        f'<td style="{td_num};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">{_fm(float(t.get("c3") or 0))}</td>'
        f'<td style="{td_num};font-size:{4.8 + _dpt}pt;font-weight:800;text-align:right;padding:{pd_cell}">{_fm(float(t.get("c4") or 0))}</td>'
        f"</tr>"
    )
    v_a = val.get("cascade_acta_anterior_cdirecto")
    s_ant = _h(_fm(float(v_a))) if v_a is not None else "—"
    s_acc = _h(_fm(float(val.get("cascade_acumulado_todas_las_rpo_cdirecto") or 0)))
    s_pr = _h(_fm(float(val.get("cascade_acta_presente_cdirecto_lista") or 0)))
    vline = f"Presente (lista actas) CD: {s_pr} · RPO ant. CD: {s_ant} · Acum. CD: {s_acc}"
    elaboro_n = _h(str(fc.get("elaboro_nombre") or "").strip() or "—")
    elaboro_c = _h(str(fc.get("elaboro_cargo") or "").strip() or "—")
    reviso_n = _h(str(fc.get("reviso_nombre") or "").strip() or "—")
    reviso_c = _h(str(fc.get("reviso_cargo") or "").strip() or "—")
    aprobo_n = _h(str(fc.get("aprobo_nombre") or "").strip() or "—")
    aprobo_c = _h(str(fc.get("aprobo_cargo") or "").strip() or "—")
    pie2 = f"Firmas CCD · {tit_acta} · {vline}"
    elaboro_td = _html_cc_sub_td_firma_columna(
        bd, "Elaboró:", elaboro_n, elaboro_c, elaboro_firma_data_uri, memoria_compact=True
    )
    reviso_td = _html_cc_sub_td_firma_columna(
        bd, "Revisó:", reviso_n, reviso_c, reviso_firma_data_uri, memoria_compact=True
    )
    aprobo_td = _html_cc_sub_td_firma_columna(
        bd, "Aprobó:", aprobo_n, aprobo_c, aprobo_firma_data_uri, memoria_compact=True
    )
    return f"""<!DOCTYPE html>
<html xmlns="http://www.w3.org/1999/xhtml" xmlns:pdf="http://www.xhtml2pdf.org/pdf">
<head><meta charset="UTF-8"/><title>{_h(formato_ccd)}</title>
<style type="text/css">
@page {{ size: letter landscape; margin: 2.8mm 3.2mm; }}
th, td {{ word-wrap: break-word; }}
table.cc-ger-mat {{ table-layout: fixed; width: 100%; border-collapse: collapse; page-break-inside: auto; }}
table.cc-ger-mat th, table.cc-ger-mat td {{ box-sizing: border-box; }}
/* xhtml2pdf a veces ignora colgroup: reforzar anchos (mismos % que celdas). */
table.cc-ger-mat th:nth-child(1), table.cc-ger-mat td:nth-child(1) {{ width: {w_cap} !important; max-width: {w_cap} !important; }}
table.cc-ger-mat th:nth-child(2), table.cc-ger-mat td:nth-child(2),
table.cc-ger-mat th:nth-child(3), table.cc-ger-mat td:nth-child(3),
table.cc-ger-mat th:nth-child(4), table.cc-ger-mat td:nth-child(4),
table.cc-ger-mat th:nth-child(5), table.cc-ger-mat td:nth-child(5) {{ width: {w_num} !important; max-width: {w_num} !important; }}
</style></head>
<body style="margin:0;padding:0;font-family:Arial,Helvetica,sans-serif;font-size:4.8pt;color:#0f172a;">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd_blk};table-layout:fixed">
<tr>
<td style="width:13%;{bd_blk};vertical-align:middle;padding:0;text-align:center;background:#fff">
{logo_html}
</td>
<td style="width:55%;{bd_blk};vertical-align:middle;text-align:center;font-weight:bold;font-size:5.05pt;padding:0 2px;line-height:1.04;">
INFORME DE AVANCE (GERENCIAL) — EJECUCIÓN DE OBRA
</td>
<td style="width:32%;{bd_blk};vertical-align:middle;text-align:center;padding:0 1px;">
<div style="color:#1d4ed8;font-weight:800;font-size:4.4pt;letter-spacing:0.1px;line-height:1.02;">{_h(formato_ccd)}</div>
<div style="font-size:3.1pt;color:#1d4ed8;font-weight:700;line-height:1.0">CCD · ClaraCore</div>
</td>
</tr>
</table>
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd_blk};background:#fff">
<tr>
<td style="padding:0;border:none">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:none">
<tr>
<td style="width:25%;{pad_m4};border:none;vertical-align:top">
<div style="{lbl_m}">CONTRATO</div>
<div style="{und_m}">{_h(str(contrato.get("numero") or ""))}</div>
</td>
<td style="width:25%;{pad_m4};border:none;vertical-align:top">
<div style="{lbl_m}">FECHA DE EMISIÓN</div>
<div style="{und_m}">{_h(fecha_gen)}</div>
</td>
<td style="width:25%;{pad_m4};border:none;vertical-align:top">
<div style="{lbl_m}">ACTA VIGENTE</div>
<div style="{und_m}">{_h(tit_acta)}</div>
</td>
<td style="width:25%;{pad_m4l};border:none;vertical-align:top">
<div style="{lbl_m}">ACTA ANTERIOR</div>
<div style="{und_m}">{_h(subt_ref)}</div>
</td>
</tr>
</table>
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:none;margin:0;table-layout:fixed">
<tr>
<td style="width:50%;border:none;vertical-align:top;{pad_m2l}">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:none;table-layout:fixed">
<tr>
<td style="width:50%;{pad_mct1};border:none;vertical-align:top">
<div style="{lbl_m}">CONTRATISTA</div>
<div style="{und_m}">{contratista_nom}{nit_en_valor}</div>
</td>
<td style="width:50%;{pad_mct2};border:none;vertical-align:top">
<div style="{lbl_m}">INTERVENTORÍA</div>
<div style="{und_m}">{interv}</div>
</td>
</tr>
</table>
</td>
<td style="width:50%;border:none;border-left:0.4px solid #b8c5d0;vertical-align:top;{pad_mob}">
<div style="{lbl_m}">OBJETO</div>
<div style="{und_m}">{_h(objeto)}</div>
</td>
</tr>
</table>
</td>
</tr>
</table>
<div style="font-size:3.25pt;color:#334155;padding:0 1px 0.5px 1px;line-height:1.1">{_h(pie2)}</div>
<table width="100%" style="border-collapse:collapse" cellspacing="0" cellpadding="0" class="cc-ger-mat">
{colgrp}
<thead><tr>{th}</tr></thead>
<tbody>{tbody}{tot_tr}</tbody>
</table>
<div style="margin-top:0;page-break-inside:avoid">
<table style="width:100%;border-collapse:collapse;font-size:3.45pt" cellspacing="0" cellpadding="0">
<tr>
{elaboro_td}
{reviso_td}
{aprobo_td}
</tr>
</table>
</div>
<div style="font-size:3.0pt;color:#64748b;padding-top:0">ClaraCore · {_h(usuario_cargo)} · {_h(usuario_nombre)}</div>
</body></html>"""


def _pdf_bytes_informe_gerencia_pareja(
    contrato_id: int,
    acta_presente_id: int,
    acta_anterior_id: Optional[int],
    current_user: dict,
) -> bytes:
    if not _acta_pertenece_contrato(contrato_id, acta_presente_id):
        raise HTTPException(404, "Acta (presente) no encontrada en este contrato")
    a_ant = int(acta_anterior_id) if acta_anterior_id is not None else None
    if a_ant is not None and (
        not _acta_pertenece_contrato(contrato_id, a_ant) or a_ant == int(acta_presente_id)
    ):
        raise HTTPException(400, "Acta de referencia no pertenece al contrato o es la misma que el acta presente")
    ap = int(acta_presente_id)
    contrato = _row(
        "contratos", "id, numero, objeto, contratista, nit, interventoria, logo_contratista", id=contrato_id
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    acta_p = _acta_cab_gerencia(ap)
    if not acta_p.get("id"):
        raise HTTPException(404, "Acta (presente) no encontrada")
    acta_ref = _acta_cab_gerencia(a_ant) if a_ant is not None else None
    ids: List[int] = [ap]
    if a_ant is not None:
        ids.append(a_ant)
    rmap = rpo_conciliacion_por_contrato(_sb, contrato_id, ids)
    d_p = rmap.get(ap) or _rpo_gerencia_vacio()
    d_a: Optional[dict] = None
    if a_ant is not None:
        d_a = rmap.get(int(a_ant)) or _rpo_gerencia_vacio()
    u = current_user if isinstance(current_user, dict) else dict(current_user)
    un = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
    uc = u.get("cargo_nombre", "—") or "—"
    fmtc = CODIGO_FORMATO_CCD_CC_GER_001
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmtc, contexto_tipo="acta_rpo", contexto_id=ap)
    fc2 = firma_cfg or {}
    e_uid2 = _opt_usuario_id(fc2.get("elaboro_usuario_id"))
    r_uid2 = _opt_usuario_id(fc2.get("reviso_usuario_id"))
    a_uid2 = _opt_usuario_id(fc2.get("aprobo_usuario_id"))
    enom2 = str(fc2.get("elaboro_nombre") or "").strip()
    rnom2 = str(fc2.get("reviso_nombre") or "").strip()
    anom2 = str(fc2.get("aprobo_nombre") or "").strip()
    el_ur = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", ap, fmtc, "elaboro", e_uid2, enom2, current_user
    )
    re_ur = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", ap, fmtc, "reviso", r_uid2, rnom2, current_user
    )
    ap_ur = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", ap, fmtc, "aprobo", a_uid2, anom2, current_user
    )
    html = _html_informe_gerencia_cc_ger_001(
        contrato,
        acta_p,
        acta_ref,
        d_p,
        d_a,
        usuario_nombre=un,
        usuario_cargo=uc,
        formato_ccd=fmtc,
        contexto_tipo="acta_rpo",
        contexto_id=ap,
        firma_cfg=firma_cfg,
        elaboro_firma_data_uri=el_ur,
        reviso_firma_data_uri=re_ur,
        aprobo_firma_data_uri=ap_ur,
    )
    return _to_pdf(html)


def _pdf_bytes_informe_gerencia_matriz(
    contrato_id: int,
    current_user: dict,
    acta_presente_override: Optional[int] = None,
    datos: Optional[Dict[str, Any]] = None,
) -> bytes:
    if datos is None:
        datos = _construir_datos_informe_gerencia_matriz(contrato_id, acta_presente_override)
    ap = datos.get("acta_presente") or {}
    ap_id = int(ap.get("id") or 0)
    if not ap_id:
        raise HTTPException(404, "No se pudo determinar el acta presente")
    contrato = _row(
        "contratos", "id, numero, objeto, contratista, nit, interventoria, logo_contratista", id=contrato_id
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    u = current_user if isinstance(current_user, dict) else dict(current_user)
    un = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
    uc = u.get("cargo_nombre", "—") or "—"
    fmtc = CODIGO_FORMATO_CCD_CC_GER_001
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmtc, contexto_tipo="acta_rpo", contexto_id=ap_id)
    fc2 = firma_cfg or {}
    e_uid2 = _opt_usuario_id(fc2.get("elaboro_usuario_id"))
    r_uid2 = _opt_usuario_id(fc2.get("reviso_usuario_id"))
    a_uid2 = _opt_usuario_id(fc2.get("aprobo_usuario_id"))
    enom2 = str(fc2.get("elaboro_nombre") or "").strip()
    rnom2 = str(fc2.get("reviso_nombre") or "").strip()
    anom2 = str(fc2.get("aprobo_nombre") or "").strip()
    el_ur = re_ur = ap_ur = None
    with ThreadPoolExecutor(max_workers=3) as _pool:
        fut_el = _pool.submit(
            _firma_data_uri_para_slot_contexto,
            contrato_id, "acta_rpo", ap_id, fmtc, "elaboro", e_uid2, enom2, current_user,
        )
        fut_re = _pool.submit(
            _firma_data_uri_para_slot_contexto,
            contrato_id, "acta_rpo", ap_id, fmtc, "reviso", r_uid2, rnom2, current_user,
        )
        fut_ap = _pool.submit(
            _firma_data_uri_para_slot_contexto,
            contrato_id, "acta_rpo", ap_id, fmtc, "aprobo", a_uid2, anom2, current_user,
        )
        el_ur = fut_el.result()
        re_ur = fut_re.result()
        ap_ur = fut_ap.result()
    html = _html_informe_gerencia_cc_ger_001_matriz(
        contrato,
        datos,
        usuario_nombre=un,
        usuario_cargo=uc,
        formato_ccd=fmtc,
        ap_id=ap_id,
        firma_cfg=firma_cfg,
        elaboro_firma_data_uri=el_ur,
        reviso_firma_data_uri=re_ur,
        aprobo_firma_data_uri=ap_ur,
    )
    return _to_pdf(html)


def _generar_pdf_bytes_corte_sub_desde_ctx(
    ctx: Dict[str, Any],
    contrato_id: int,
    corte_id: int,
    current_user: Optional[dict] = None,
) -> bytes:
    """Mismo PDF que la vista previa CC-SUB-001. Firmas en Elaboró/Revisó según usuarios configurados y registros por corte."""
    contrato = ctx["contrato"]
    sub = ctx["sub"]
    corte = ctx["corte"]
    items = ctx["items"]
    total_costo = ctx["total_costo"]
    usuario_nombre = ctx["usuario_nombre"]
    usuario_cargo = ctx["usuario_cargo"]
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, CODIGO_FORMATO_CCD_CC_SUB_001, corte_id=corte_id)
    fmt001 = CODIGO_FORMATO_CCD_CC_SUB_001
    e_uid = _opt_usuario_id(firma_cfg.get("elaboro_usuario_id"))
    r_uid = _opt_usuario_id(firma_cfg.get("reviso_usuario_id"))
    enom = str(firma_cfg.get("elaboro_nombre") or "").strip()
    rnom = str(firma_cfg.get("reviso_nombre") or "").strip()
    elaboro_firma_uri = _firma_data_uri_para_slot_corte(
        contrato_id, corte_id, fmt001, "elaboro", e_uid, enom, current_user
    )
    reviso_firma_uri = _firma_data_uri_para_slot_corte(
        contrato_id, corte_id, fmt001, "reviso", r_uid, rnom, current_user
    )

    errs: list[str] = []
    pdf_bytes: bytes | None = None
    for name, fn in (
        (
            "v1_plana",
            lambda: _html_cc_sub_v1_plain(
                contrato,
                sub,
                corte,
                items,
                total_costo,
                usuario_nombre,
                usuario_cargo,
                firma_cfg=firma_cfg,
                elaboro_firma_data_uri=elaboro_firma_uri,
                reviso_firma_data_uri=reviso_firma_uri,
            ),
        ),
        ("modo_seguro", lambda: _html_corte_sub_fallback(contrato, sub, corte, items, total_costo, usuario_nombre, usuario_cargo)),
        ("minima", lambda: _html_corte_sub_minima(
            contrato, sub, corte, items, total_costo, usuario_nombre, usuario_cargo, "prev", "prev",
        )),
    ):
        try:
            pdf_bytes = _to_pdf(fn())
            break
        except Exception as e:
            errs.append(f"{name}:{e!s}"[:220])
            _log.warning("pdf_corte_sub falló %s: %s", name, e)

    if pdf_bytes is None:
        try:
            html_last = _html_cc_sub_v1_plain(
                contrato,
                sub,
                corte,
                items,
                total_costo,
                usuario_nombre,
                usuario_cargo,
                firma_cfg=firma_cfg,
                elaboro_firma_data_uri=elaboro_firma_uri,
                reviso_firma_data_uri=reviso_firma_uri,
            )
            pdf_bytes = _to_pdf_corte_garantizado(html_last)
        except Exception as e:
            _log.exception("pdf_corte_sub: sin PDF tras cadena de respaldo")
            raise HTTPException(
                status_code=500,
                detail="PDF corte: " + " | ".join(errs)[:900] + f" | garantizado: {e!s}"[:1200],
            ) from e
    return pdf_bytes


def _merge_pdf_bytes(parte_principal: bytes, parte_anexa: bytes) -> bytes:
    return _merge_pdf_bytes_list([parte_principal, parte_anexa])


def _merge_pdf_bytes_list(partes: List[bytes]) -> bytes:
    from pypdf import PdfReader, PdfWriter
    w = PdfWriter()
    for parte in partes:
        if not parte:
            continue
        for page in PdfReader(io.BytesIO(parte)).pages:
            w.add_page(page)
    out = io.BytesIO()
    w.write(out)
    return out.getvalue()


def _merge_pdf_bytes_tree(partes: List[bytes]) -> bytes:
    """Fusiona muchos PDFs en árbol (menos picos de RAM que un solo merge gigante)."""
    chunk = [p for p in partes if p]
    if not chunk:
        return b""
    if len(chunk) == 1:
        return chunk[0]
    while len(chunk) > 1:
        next_lvl: List[bytes] = []
        for i in range(0, len(chunk), 12):
            next_lvl.append(_merge_pdf_bytes_list(chunk[i : i + 12]))
        chunk = next_lvl
    return chunk[0]


def _attachment_pdf_con_pagina_sello_usuario(
    pdf_main: bytes,
    current_user: dict,
    *,
    titulo_doc: str,
    formato_ccd: str,
    contrato_numero: str,
    nombre_archivo_pdf: str,
) -> Response:
    """
    Anexa la página de sello ClaraCore (firma de perfil, fecha, SHA-256 del PDF previo).
    Misma lógica que corte CC-SUB-001 /con-sello-firma.
    """
    uid = int(current_user.get("sub") or 0)
    if not uid:
        raise HTTPException(status_code=401, detail="Debes iniciar sesión para descargar el PDF firmado.")

    digest = hashlib.sha256(pdf_main).hexdigest()
    ufir = _usuario_datos_firma_sello(uid)
    if not ufir.get("firma_imagen_url"):
        raise HTTPException(
            status_code=400,
            detail="Configura tu imagen de firma en Perfil antes de descargar el PDF con sello.",
        )
    try:
        img_uri = _fetch_img_data_uri(ufir["firma_imagen_url"])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo cargar la imagen de firma: {e!s}") from e

    tz = ZoneInfo("America/Bogota")
    ahora = datetime.now(tz)
    cuando = ahora.strftime("%Y-%m-%d %H:%M:%S %Z")

    html_sello = _html_pagina_sello_firma_claracore(
        titulo_doc=str(titulo_doc or "")[:280],
        digest_sha256_hex=digest,
        cuando_bogota=cuando,
        contrato_numero=str(contrato_numero or ""),
        formato_ccd=formato_ccd,
        nombre_firmante=ufir["nombre_completo"],
        cargo_firmante=ufir["cargo"],
        img_data_uri=img_uri,
    )
    pdf_sello = _to_pdf(html_sello)
    out = _merge_pdf_bytes(pdf_main, pdf_sello)
    base = nombre_archivo_pdf.strip() or "documento.pdf"
    fname = (base[:-4] + "_firmado.pdf") if base.lower().endswith(".pdf") else (base + "_firmado.pdf")
    return Response(
        content=out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _usuario_datos_firma_sello(uid: int) -> Dict[str, Any]:
    rows = _sb.table("usuarios").select("nombre, apellidos, firma_imagen_url, cargo_id").eq("id", uid).limit(1).execute().data or []
    if not rows:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    row = rows[0]
    cargo_nom = "—"
    cid = row.get("cargo_id")
    if cid:
        cr = _sb.table("cargos").select("nombre").eq("id", cid).limit(1).execute().data or []
        if cr:
            cargo_nom = str(cr[0].get("nombre") or "—").strip() or "—"
    nombre = f"{row.get('nombre') or ''} {row.get('apellidos') or ''}".strip() or "—"
    return {
        "nombre_completo": nombre,
        "cargo": cargo_nom,
        "firma_imagen_url": (row.get("firma_imagen_url") or "").strip(),
    }


def _fetch_img_data_uri(url: str) -> str:
    """Convierte una URL de imagen a data-URI. Acepta http(s) o data: (p. ej. logos guardados desde el admin)."""
    u = (url or "").strip()
    if u.startswith("data:"):
        return _logo_url_pdf_safe(u)
    import requests
    r = requests.get(u, timeout=25)
    r.raise_for_status()
    ct = (r.headers.get("content-type") or "image/png").split(";")[0].strip()
    if not ct.startswith("image/"):
        ct = "image/png"
    b64 = base64.b64encode(r.content).decode("ascii")
    return f"data:{ct};base64,{b64}"


def _firma_imagen_url_usuario(uid: int) -> Optional[str]:
    rows = _sb.table("usuarios").select("firma_imagen_url").eq("id", uid).limit(1).execute().data or []
    if not rows:
        return None
    u = (rows[0].get("firma_imagen_url") or "").strip()
    return u or None


def _uid_session(current_user: Optional[dict]) -> Optional[int]:
    if not current_user or not current_user.get("sub"):
        return None
    try:
        return int(current_user["sub"])
    except (TypeError, ValueError):
        return None


def _usuario_firma_img_data_uri_opcional(current_user: Optional[dict]) -> Optional[str]:
    """Data URI de la imagen de firma del usuario en sesión (perfil)."""
    uid = _uid_session(current_user)
    if not uid:
        return None
    url = _firma_imagen_url_usuario(uid)
    if not url:
        return None
    try:
        return _fetch_img_data_uri(url)
    except Exception as e:
        _log.warning("Firma de perfil omitida: %s", e)
        return None


def _ccd_firma_registro_get(corte_id: int, formato_codigo: str, slot: str) -> Optional[Dict[str, Any]]:
    try:
        rows = (
            _sb.table("ccd_corte_firma_registro")
            .select("firma_imagen_url, usuario_id, created_at")
            .eq("corte_id", corte_id)
            .eq("formato_codigo", formato_codigo)
            .eq("slot", slot)
            .order("created_at", desc=True)
            .limit(1)
            .execute()
            .data
            or []
        )
        return rows[0] if rows else None
    except Exception as e:
        _log.debug("ccd_corte_firma_registro: %s", e)
        return None


def _ccd_firma_registro_contexto_get(
    contrato_id: int,
    contexto_tipo: str,
    contexto_id: int,
    formato_codigo: str,
    slot: str,
) -> Optional[Dict[str, Any]]:
    try:
        rows = (
            _sb.table("ccd_firma_registro")
            .select("firma_imagen_url, usuario_id, created_at")
            .eq("contrato_id", contrato_id)
            .eq("contexto_tipo", contexto_tipo)
            .eq("contexto_id", contexto_id)
            .eq("formato_codigo", formato_codigo)
            .eq("slot", slot)
            .order("created_at", desc=True)
            .limit(1)
            .execute()
            .data
            or []
        )
        return rows[0] if rows else None
    except Exception as e:
        _log.debug("ccd_firma_registro: %s", e)
        return None


def _fecha_hora_marca_bo_desde_created_at(created_at: Optional[object]) -> Tuple[str, str]:
    """(fecha dd/mm/aaaa, hora HH:MM) en America/Bogota desde `ccd_firma_registro.created_at`."""
    if created_at is None:
        return "", ""
    try:
        tz_bo = ZoneInfo("America/Bogota")
        if isinstance(created_at, datetime):
            dt = created_at
        elif isinstance(created_at, date):
            dt = datetime.combine(created_at, datetime.min.time(), tzinfo=tz_bo)
        else:
            s = str(created_at).strip()
            if not s:
                return "", ""
            if " " in s and "T" not in s.upper():
                s = s.replace(" ", "T", 1)
            if s.endswith("Z"):
                s = s[:-1] + "+00:00"
            dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        dt_bo = dt.astimezone(tz_bo)
        return dt_bo.strftime("%d/%m/%Y"), dt_bo.strftime("%H:%M")
    except Exception as e:
        _log.debug("marca fecha/hora firma: %s", e)
        return "", ""


def _ccd_usuario_nombre_completo_por_id(usuario_id: Optional[int]) -> str:
    """Nombre y apellidos desde `usuarios.id` (firma registrada)."""
    uid = _opt_usuario_id(usuario_id)
    if not uid:
        return ""
    try:
        rows = (
            _sb.table("usuarios")
            .select("nombre, apellidos")
            .eq("id", int(uid))
            .limit(1)
            .execute()
            .data
            or []
        )
        if not rows:
            return ""
        r = rows[0]
        return f"{r.get('nombre') or ''} {r.get('apellidos') or ''}".strip()
    except Exception as e:
        _log.debug("nombre usuario id=%s: %s", uid, e)
        return ""


def _firma_data_uri_para_slot_contexto(
    contrato_id: int,
    contexto_tipo: str,
    contexto_id: int,
    formato_codigo: str,
    slot: str,
    config_usuario_id: Optional[int],
    nombre_configurado: str,
    current_user: Optional[dict],
) -> Optional[str]:
    """Firma del slot SOLO desde registro explícito (no autocompleta desde perfil)."""
    reg = _ccd_firma_registro_contexto_get(contrato_id, contexto_tipo, contexto_id, formato_codigo, slot)
    if reg and reg.get("firma_imagen_url"):
        try:
            return _fetch_img_data_uri(reg["firma_imagen_url"])
        except Exception as e:
            _log.warning("Firma registrada contexto slot %s: %s", slot, e)
    return None


def _normalizar_nombre_firma_txt(s: str) -> str:
    """Quita signos, colapsa espacios — útil cuando el config tiene «Jaime!» o el usuario «Jaimes»."""
    t = (s or "").lower().strip()
    t = re.sub(r"[!?.¡¿,;:]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _nombre_apellidos_usuario_sesion(current_user: Optional[dict]) -> str:
    """El JWT suele traer solo `sub`; completa nombre desde tabla usuarios si hace falta."""
    if not current_user:
        return ""
    n = f"{current_user.get('nombre', '')} {current_user.get('apellidos', '')}".strip()
    if n:
        return n
    uid = _uid_session(current_user)
    if not uid:
        return ""
    try:
        rows = _sb.table("usuarios").select("nombre, apellidos").eq("id", uid).limit(1).execute().data or []
        if rows:
            return f"{rows[0].get('nombre') or ''} {rows[0].get('apellidos') or ''}".strip()
    except Exception as e:
        _log.debug("nombre usuario sesión: %s", e)
    return ""


def _nombre_coincide_firma_cfg(nombre_en_config: str, current_user: Optional[dict]) -> bool:
    """Compatibilidad sin usuario_id: compara nombre del config con nombre/apellidos del usuario."""
    if not current_user:
        return False
    nc = _normalizar_nombre_firma_txt(nombre_en_config)
    if not nc or nc in ("—", "ej. elaboró", "ej. revisó"):
        return False
    nom = _normalizar_nombre_firma_txt(_nombre_apellidos_usuario_sesion(current_user))
    if not nom:
        return False
    if nom == nc or nom in nc or nc in nom:
        return True
    # Tipos «Jaime» vs «Jaimes», espacios o puntuación
    if difflib.SequenceMatcher(None, nom, nc).ratio() >= 0.82:
        return True
    return False


def _slot_firma_usuario_en_config(fc: Dict[str, Any], uid: int, current_user: dict) -> Optional[str]:
    """Elaboró / Elaboró2 / Revisó / Revisó2 / Aprobó según ids guardados o, si faltan, por nombre."""
    uid_i = int(uid)
    slots_id = [
        ("elaboro",  _opt_usuario_id(fc.get("elaboro_usuario_id"))),
        ("elaboro2", _opt_usuario_id(fc.get("elaboro2_usuario_id"))),
        ("reviso",   _opt_usuario_id(fc.get("reviso_usuario_id"))),
        ("reviso2",  _opt_usuario_id(fc.get("reviso2_usuario_id"))),
        ("aprobo",   _opt_usuario_id(fc.get("aprobo_usuario_id"))),
    ]
    for slot_name, sid in slots_id:
        if sid and int(sid) == uid_i:
            return slot_name
    # Fallback por nombre cuando no hay usuario_id registrado
    slots_nombre = [
        ("elaboro",  str(fc.get("elaboro_nombre") or "")),
        ("elaboro2", str(fc.get("elaboro2_nombre") or "")),
        ("reviso",   str(fc.get("reviso_nombre") or "")),
        ("reviso2",  str(fc.get("reviso2_nombre") or "")),
        ("aprobo",   str(fc.get("aprobo_nombre") or "")),
    ]
    for slot_name, nombre in slots_nombre:
        if _nombre_coincide_firma_cfg(nombre, current_user):
            return slot_name
    return None


def _firma_data_uri_para_slot_corte(
    contrato_id: int,
    corte_id: int,
    formato_codigo: str,
    slot: str,
    config_usuario_id: Optional[int],
    nombre_configurado: str,
    current_user: Optional[dict],
) -> Optional[str]:
    """
    Imagen para Elaboró o Revisó: primero firma registrada en BD; si no hay,
    el usuario asignado a ese slot ve su imagen de perfil al generar/descargar.
    Si no hay usuario_id en config (datos viejos), se intenta coincidencia por nombre.
    """
    _ = contrato_id  # reservado p. ej. validaciones futuras
    reg = _ccd_firma_registro_get(corte_id, formato_codigo, slot)
    if reg and reg.get("firma_imagen_url"):
        try:
            return _fetch_img_data_uri(reg["firma_imagen_url"])
        except Exception as e:
            _log.warning("Firma registrada slot %s: %s", slot, e)
    cu = _opt_usuario_id(config_usuario_id)
    sess = _uid_session(current_user)
    if cu and sess and int(cu) == int(sess):
        return _usuario_firma_img_data_uri_opcional(current_user)
    if (not cu) and _nombre_coincide_firma_cfg(nombre_configurado, current_user):
        return _usuario_firma_img_data_uri_opcional(current_user)
    return None


def _html_pagina_sello_firma_claracore(
    *,
    titulo_doc: str,
    digest_sha256_hex: str,
    cuando_bogota: str,
    contrato_numero: str,
    formato_ccd: str,
    nombre_firmante: str,
    cargo_firmante: str,
    img_data_uri: str,
) -> str:
    esc = lambda s: html.escape(str(s if s is not None else ""), quote=True)
    td0 = str(titulo_doc or "")
    doc_corto = td0[:120] + ("…" if len(td0) > 120 else "")
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/><style>
    @page {{ size: A4; margin: 14mm 16mm; }}
    body {{ font-family: DejaVu Sans, Arial, Helvetica, sans-serif; font-size: 6.5pt; color: #0f172a; margin: 0; padding: 0; }}
    .box {{ border: 1px solid #0e7490; border-radius: 3px; padding: 5px 7px; background: #f0fdfa; page-break-inside: avoid; }}
    .t1 {{ font-size: 7pt; font-weight: bold; color: #0f766e; text-transform: uppercase; letter-spacing: 0.3px; margin: 0; padding: 0 0 3px 0; border-bottom: 1px solid #99f6e4; }}
    .muted {{ font-size: 5.5pt; color: #64748b; margin-top: 2px; line-height: 1.2; }}
    .hash {{ font-family: DejaVu Sans Mono, Consolas, monospace; font-size: 5pt; word-break: break-all; color: #334155; line-height: 1.15; margin-top: 1px; }}
    .sigwrap {{ height: 2.4cm; max-height: 2.4cm; overflow: hidden; text-align: center; line-height: 0; }}
    .sigwrap img {{ max-width: 100%; max-height: 2.2cm; width: auto; height: auto; display: block; margin: 0 auto; }}
    .sigcap {{ font-size: 5pt; color: #94a3b8; margin-top: 2px; }}
    </style></head><body>
    <div class="box">
      <div class="t1">Revisado y aprobado por</div>
      <table style="width:100%;margin-top:4px;border-collapse:collapse;">
        <tr>
          <td style="width:58%;vertical-align:top;padding-right:6px;">
            <div style="font-size:6.5pt;line-height:1.15;"><strong>{esc(cargo_firmante)}</strong></div>
            <div style="font-size:7pt;margin-top:1px;font-weight:bold;">{esc(nombre_firmante)}</div>
            <div class="muted">Doc.: {esc(doc_corto)} · {esc(formato_ccd)}</div>
            <div class="muted">Contrato: {esc(contrato_numero)}</div>
            <div class="muted">Firma (America/Bogotá): {esc(cuando_bogota)}</div>
            <div class="muted">SHA-256 (PDF previo al sello):</div>
            <div class="hash">{esc(digest_sha256_hex)}</div>
            <div class="muted" style="margin-top:3px;">
              Protocolo ClaraCore: huella de integridad del binario previo; no es firma electrónica avanzada (p. ej. Adobe Sign).
            </div>
          </td>
          <td style="width:42%;vertical-align:top;text-align:center;padding-left:4px;">
            <div class="sigwrap"><img src="{img_data_uri}" alt="Firma"/></div>
            <div class="sigcap">Imagen de firma del perfil</div>
          </td>
        </tr>
      </table>
    </div>
    </body></html>"""


@router.get("/{contrato_id}/pdf/corte-subcontratista/{corte_id}", dependencies=[])
def pdf_corte_sub(contrato_id: int, corte_id: int, current_user: dict = Depends(_get_user)):
    _perm_informes_ccd(current_user, "ver")
    try:
        try:
            ctx = _contexto_corte_sub(contrato_id, corte_id, current_user)
        except HTTPException:
            raise
        except Exception as e:
            _log.exception("pdf_corte_sub: contexto")
            raise HTTPException(503, f"No se pudieron cargar los datos del informe: {e!s}") from e

        pdf_bytes = _generar_pdf_bytes_corte_sub_desde_ctx(ctx, contrato_id, corte_id, current_user)
        corte = ctx["corte"]
        sub = ctx["sub"]
        fname = _nombre_archivo_cc_sub_001(corte, sub, corte_id)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print("ERROR pdf_corte_sub:", tb, flush=True)
        raise HTTPException(status_code=500, detail=f"PDF corte: {type(e).__name__}: {e!s} | {tb[-500:]}") from e


@router.get("/{contrato_id}/pdf/corte-subcontratista/{corte_id}/con-sello-firma")
def pdf_corte_sub_con_sello_firma(
    contrato_id: int,
    corte_id: int,
    current_user: dict = Depends(_get_user),
):
    """
    PDF CC-SUB-001 + página final de sello: imagen de firma del usuario, cargo, nombre,
    fecha/hora (America/Bogotá), contrato, y SHA-256 del PDF del informe *antes* de añadir el sello.
    Requiere imagen de firma en perfil (firma_imagen_url).
    """
    _perm_informes_ccd(current_user, "ver")
    try:
        ctx = _contexto_corte_sub(contrato_id, corte_id, current_user)
        pdf_main = _generar_pdf_bytes_corte_sub_desde_ctx(ctx, contrato_id, corte_id, current_user)
        contrato = ctx["contrato"]
        corte = ctx["corte"]
        sub = ctx["sub"]
        titulo = f"Informe corte subcontratista — Corte {_corte_consecutivo_fmt(corte)} — {sub.get('razon_social') or ''}"
        base = _nombre_archivo_cc_sub_001(corte, sub, corte_id)
        return _attachment_pdf_con_pagina_sello_usuario(
            pdf_main,
            current_user,
            titulo_doc=titulo,
            formato_ccd=CODIGO_FORMATO_CCD_CC_SUB_001,
            contrato_numero=str(contrato.get("numero") or ""),
            nombre_archivo_pdf=base,
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("pdf_corte_sub_con_sello_firma")
        raise HTTPException(status_code=500, detail=f"PDF firmado: {e!s}") from e


# ── CC-SUB-002 : Memorias Corte Subcontratista ─────────────────────────────────

@router.get("/{contrato_id}/pdf/memoria-item/{corte_id}")
def pdf_memoria_item(
    contrato_id: int,
    corte_id:    int,
    item_numero: str = Query(...),
    current_user=Depends(_get_user)
):
    _perm_informes_ccd(current_user, "ver")
    try:
        ctx = _contexto_memoria_item(contrato_id, corte_id, item_numero, current_user)
        contrato = ctx["contrato"]
        sub = ctx["sub"]
        corte = ctx["corte"]
        item_info = ctx["item_info"]
        registros = ctx["registros"]
        usuario_nombre = ctx["usuario_nombre"]
        usuario_cargo = ctx["usuario_cargo"]
        firma_cfg = _get_firma_cfg_para_documento(contrato_id, CODIGO_FORMATO_CCD_CC_SUB_002, corte_id=corte_id)
        fc = firma_cfg or {}
        fmt002 = CODIGO_FORMATO_CCD_CC_SUB_002
        e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
        r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
        enom = str(fc.get("elaboro_nombre") or "").strip()
        rnom = str(fc.get("reviso_nombre") or "").strip()
        elaboro_firma_uri = _firma_data_uri_para_slot_corte(
            contrato_id, corte_id, fmt002, "elaboro", e_uid, enom, current_user
        )
        reviso_firma_uri = _firma_data_uri_para_slot_corte(
            contrato_id, corte_id, fmt002, "reviso", r_uid, rnom, current_user
        )

        try:
            html = _html_memoria_item(
                contrato,
                sub,
                corte,
                item_info,
                registros,
                usuario_nombre,
                usuario_cargo,
                firma_cfg=firma_cfg,
                elaboro_firma_data_uri=elaboro_firma_uri,
                reviso_firma_data_uri=reviso_firma_uri,
            )
            pdf_bytes = _to_pdf(html)
        except Exception as e:
            _log.exception("pdf_memoria_item: fallo HTML/PDF, intento plantilla mínima")
            try:
                html_min = _html_memoria_minima(contrato, sub, corte, item_info, registros, usuario_nombre, usuario_cargo, str(e))
                pdf_bytes = _to_pdf(html_min)
            except Exception as e2:
                raise HTTPException(500, f"Error generando PDF memoria: {e!s} | mínima: {e2!s}") from e2

        item_safe = _safe_filename_part(item_numero.replace("/", "-").replace(" ", ""))
        fname = _safe_filename_part(f"CC-SUB-002_Corte{corte.get('consecutivo', '')}_{item_safe}.pdf")
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("pdf_memoria_item: error no controlado")
        raise HTTPException(500, f"Error interno memoria PDF: {repr(e)}")


@router.get("/{contrato_id}/pdf/memoria-corte-completo/{corte_id}")
def pdf_memoria_corte_completo(
    contrato_id: int,
    corte_id: int,
    current_user=Depends(_get_user),
):
    """Un solo PDF CC-SUB-002: todos los ítems del corte (mismo formato que por ítem, separados con salto de página)."""
    _perm_informes_ccd(current_user, "ver")
    try:
        numeros = _list_item_numeros_memoria_corte(contrato_id, corte_id)
        if not numeros:
            raise HTTPException(
                status_code=404,
                detail="No hay registros aprobados en este corte para generar memorias",
            )

        firma_cfg = _get_firma_cfg_para_documento(contrato_id, CODIGO_FORMATO_CCD_CC_SUB_002, corte_id=corte_id)
        fc = firma_cfg or {}
        fmt002 = CODIGO_FORMATO_CCD_CC_SUB_002
        e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
        r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
        enom = str(fc.get("elaboro_nombre") or "").strip()
        rnom = str(fc.get("reviso_nombre") or "").strip()
        elaboro_firma_uri = _firma_data_uri_para_slot_corte(
            contrato_id, corte_id, fmt002, "elaboro", e_uid, enom, current_user
        )
        reviso_firma_uri = _firma_data_uri_para_slot_corte(
            contrato_id, corte_id, fmt002, "reviso", r_uid, rnom, current_user
        )
        est = _merge_estilo_pdf(fc.get("estilo_pdf"), CODIGO_FORMATO_CCD_CC_SUB_002)
        estilo_css = _memoria_pdf_estilo_css(est)

        parts: list[str] = []
        corte_out: dict = {}
        for i, item_numero in enumerate(numeros):
            ctx = _contexto_memoria_item(
                contrato_id,
                corte_id,
                item_numero,
                current_user,
                item_exacto=True,
            )
            contrato = ctx["contrato"]
            sub = ctx["sub"]
            corte = ctx["corte"]
            corte_out = corte
            item_info = ctx["item_info"]
            registros = ctx["registros"]
            usuario_nombre = ctx["usuario_nombre"]
            usuario_cargo = ctx["usuario_cargo"]
            inner = _html_memoria_item_body(
                contrato,
                sub,
                corte,
                item_info,
                registros,
                usuario_nombre,
                usuario_cargo,
                firma_cfg=firma_cfg,
                elaboro_firma_data_uri=elaboro_firma_uri,
                reviso_firma_data_uri=reviso_firma_uri,
            )
            if i > 0:
                parts.append("<pdf:nextpage />")
            parts.append(inner)

        html = _wrap_memoria_item_html("".join(parts), estilo_css)
        try:
            pdf_bytes = _to_pdf(html)
        except Exception as e:
            _log.exception("pdf_memoria_corte_completo: fallo PDF")
            raise HTTPException(500, f"Error generando PDF memoria completa: {e!s}") from e

        cons = corte_out.get("consecutivo", "")
        fname = _safe_filename_part(f"CC-SUB-002_Corte{cons}_todos-items.pdf")
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("pdf_memoria_corte_completo: error no controlado")
        raise HTTPException(500, f"Error interno memoria PDF completa: {repr(e)}")


# ── CC-SEM / CC-MES : Conciliación interventoría–contratista ───────────────────


def _sub_corte_dummy_memoria() -> tuple:
    return ({"razon_social": "—", "nombre_contacto": "—"}, {"consecutivo": "—", "fecha_inicio": None, "fecha_fin": None})


@router.get("/{contrato_id}/pdf/cc-sem-001/semana/{semana_id}")
def pdf_cc_sem_001_semana(
    contrato_id: int,
    semana_id: int,
    current_user=Depends(_get_user),
):
    _perm_informes_ccd(current_user, "ver")
    if not _semana_pertenece_contrato(contrato_id, semana_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    reg = fetch_registros_conciliacion(_sb, contrato_id, semana_id=semana_id)
    items, total = aggregate_items_conciliacion(reg)
    _sort_items_corte_por_item_numero_asc(items)
    sm = _row("so_semanas", "id, numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
    nsem = sm.get("numero_semana")
    fi = str(sm.get("fecha_inicio") or "—")
    ff = str(sm.get("fecha_fin") or "—")
    pdf_bytes = _pdf_bytes_conciliacion_informe_v1(
        contrato_id,
        current_user,
        formato_codigo=CODIGO_FORMATO_CCD_CC_SEM_001,
        contexto_tipo="semana",
        contexto_id=semana_id,
        titulo_documento="INFORME EJECUCIÓN SEMANAL (CONCILIACIÓN INTERVENTORÍA–CONTRATISTA)",
        c3_label="SEMANA",
        c3_value=f"N° {nsem}",
        c4_label="VIGENCIA",
        c4_value=f"{fi} — {ff}",
        pie_contexto=f"Semana N° {nsem} · {fi} — {ff} · Registros nivel 3 aprobados y bloqueados",
        items=items,
        total_costo=total,
    )
    fname = _safe_filename_part(f"CC-SEM-001_semana_{nsem}.pdf")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{contrato_id}/pdf/cc-mes-001/acta/{acta_id}")
def pdf_cc_mes_001_acta(
    contrato_id: int,
    acta_id: int,
    current_user=Depends(_get_user),
):
    _perm_informes_ccd(current_user, "ver")
    if not _acta_pertenece_contrato(contrato_id, acta_id):
        raise HTTPException(404, "Acta no encontrada en este contrato")
    reg = fetch_registros_informe_cc_mes_por_acta(_sb, contrato_id, acta_id)
    items, total = aggregate_items_conciliacion(reg)
    _sort_items_corte_por_item_numero_asc(items)
    ac = _row("actas", "id, numero_rpo, consecutivo", id=acta_id) or {}
    nrpo = str(ac.get("numero_rpo") or ac.get("consecutivo") or acta_id)
    cons = str(ac.get("consecutivo") or "—")
    fa = "—"
    pdf_bytes = _pdf_bytes_conciliacion_informe_v1(
        contrato_id,
        current_user,
        formato_codigo=CODIGO_FORMATO_CCD_CC_MES_001,
        contexto_tipo="acta_rpo",
        contexto_id=acta_id,
        titulo_documento="INFORME EJECUCIÓN MENSUAL (CONCILIACIÓN INTERVENTORÍA–CONTRATISTA)",
        c3_label="ACTA RPO",
        c3_value=nrpo,
        c4_label="FECHA ACTA",
        c4_value=fa,
        pie_contexto=f"Acta RPO {nrpo} · consecutivo {cons} · Misma lógica que módulo Actas (N1·N2·N3 aprob. en cascada, costo directo por línea)",
        items=items,
        total_costo=total,
    )
    fname = _safe_filename_part(f"CC-MES-001_acta_{nrpo}.pdf")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{contrato_id}/pdf/cc-sem-002/semana/{semana_id}")
def pdf_cc_sem_002_semana(
    contrato_id: int,
    semana_id: int,
    item_numero: str = Query(...),
    current_user=Depends(_get_user),
):
    _perm_informes_ccd(current_user, "ver")
    if not _semana_pertenece_contrato(contrato_id, semana_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    registros = fetch_registros_memoria_conciliacion(
        _sb, contrato_id, item_numero, semana_id=semana_id, item_exacto=True
    )
    if not registros:
        raise HTTPException(404, "No hay registros para este ítem y semana")
    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    u = current_user if isinstance(current_user, dict) else dict(current_user)
    usuario_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
    usuario_cargo = u.get("cargo_nombre", "—") or "—"
    item_info = {
        "item_numero": registros[0].get("item_numero", item_numero),
        "item_descripcion": registros[0].get("item_descripcion", ""),
        "unidad": registros[0].get("unidad", ""),
    }
    sm = _row("so_semanas", "numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
    nsem = sm.get("numero_semana")
    fi = str(sm.get("fecha_inicio") or "—")
    ff = str(sm.get("fecha_fin") or "—")
    conc_meta = {
        "titulo": "RESUMEN ACTIVIDADES — CONCILIACIÓN SEMANAL (INTERVENTORÍA–CONTRATISTA)",
        "codigo": CODIGO_FORMATO_CCD_CC_SEM_002,
        "cells": [
            ("CONTRATO", str(contrato.get("numero") or "")),
            ("SEMANA", f"N° {nsem}"),
            ("VIGENCIA", f"{fi} — {ff}"),
            ("REFERENCIA", "Cantidades ejecutadas — conciliación"),
        ],
    }
    fmt = CODIGO_FORMATO_CCD_CC_SEM_002
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmt, contexto_tipo="semana", contexto_id=semana_id)
    fc = firma_cfg or {}
    e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
    r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
    a_uid = _opt_usuario_id(fc.get("aprobo_usuario_id"))
    enom = str(fc.get("elaboro_nombre") or "").strip()
    rnom = str(fc.get("reviso_nombre") or "").strip()
    anom = str(fc.get("aprobo_nombre") or "").strip()
    elaboro_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "semana", semana_id, fmt, "elaboro", e_uid, enom, current_user
    )
    reviso_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "semana", semana_id, fmt, "reviso", r_uid, rnom, current_user
    )
    aprobo_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "semana", semana_id, fmt, "aprobo", a_uid, anom, current_user
    )
    sub, corte = _sub_corte_dummy_memoria()
    html = _html_memoria_item(
        contrato,
        sub,
        corte,
        item_info,
        registros,
        usuario_nombre,
        usuario_cargo,
        firma_cfg=firma_cfg,
        elaboro_firma_data_uri=elaboro_uri,
        reviso_firma_data_uri=reviso_uri,
        conc_meta=conc_meta,
        aprobo_firma_data_uri=aprobo_uri,
        aprobo_interventoria_desde_config=True,
        pie_fotos_contexto=f"Semana N° {nsem} · {fi} — {ff}",
        estilo_formato_codigo=fmt,
    )
    pdf_bytes = _to_pdf(html)
    item_safe = _safe_filename_part(item_numero.replace("/", "-").replace(" ", ""))
    fname = _safe_filename_part(f"CC-SEM-002_semana{nsem}_{item_safe}.pdf")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{contrato_id}/pdf/cc-sem-002/semana/{semana_id}/completo")
def pdf_cc_sem_002_semana_completo(
    contrato_id: int,
    semana_id: int,
    current_user=Depends(_get_user),
):
    """Un solo PDF CC-SEM-002: todos los ítems de la semana (mismo formato que por ítem, salto entre ítems)."""
    _perm_informes_ccd(current_user, "ver")
    try:
        if not _semana_pertenece_contrato(contrato_id, semana_id):
            raise HTTPException(404, "Semana no encontrada en este contrato")
        reg_agg = fetch_registros_conciliacion(_sb, contrato_id, semana_id=semana_id)
        items_agg, _total = aggregate_items_conciliacion(reg_agg)
        _sort_items_corte_por_item_numero_asc(items_agg)
        numeros = [
            (it.get("item_numero") or "").strip()
            for it in items_agg
            if (it.get("item_numero") or "").strip()
        ]
        if not numeros:
            raise HTTPException(
                status_code=404,
                detail="No hay registros en esta semana para generar memorias",
            )

        contrato = _row(
            "contratos",
            "numero, objeto, contratista, nit, interventoria, logo_contratista",
            id=contrato_id,
        )
        if not contrato:
            raise HTTPException(404, "Contrato no encontrado")
        u = current_user if isinstance(current_user, dict) else dict(current_user)
        usuario_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
        usuario_cargo = u.get("cargo_nombre", "—") or "—"
        sm = _row("so_semanas", "numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
        nsem = sm.get("numero_semana")
        fi = str(sm.get("fecha_inicio") or "—")
        ff = str(sm.get("fecha_fin") or "—")
        conc_meta = {
            "titulo": "RESUMEN ACTIVIDADES — CONCILIACIÓN SEMANAL (INTERVENTORÍA–CONTRATISTA)",
            "codigo": CODIGO_FORMATO_CCD_CC_SEM_002,
            "cells": [
                ("CONTRATO", str(contrato.get("numero") or "")),
                ("SEMANA", f"N° {nsem}"),
                ("VIGENCIA", f"{fi} — {ff}"),
                ("REFERENCIA", "Cantidades ejecutadas — conciliación"),
            ],
        }
        fmt = CODIGO_FORMATO_CCD_CC_SEM_002
        firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmt, contexto_tipo="semana", contexto_id=semana_id)
        fc = firma_cfg or {}
        e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
        r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
        a_uid = _opt_usuario_id(fc.get("aprobo_usuario_id"))
        enom = str(fc.get("elaboro_nombre") or "").strip()
        rnom = str(fc.get("reviso_nombre") or "").strip()
        anom = str(fc.get("aprobo_nombre") or "").strip()
        elaboro_uri = _firma_data_uri_para_slot_contexto(
            contrato_id, "semana", semana_id, fmt, "elaboro", e_uid, enom, current_user
        )
        reviso_uri = _firma_data_uri_para_slot_contexto(
            contrato_id, "semana", semana_id, fmt, "reviso", r_uid, rnom, current_user
        )
        aprobo_uri = _firma_data_uri_para_slot_contexto(
            contrato_id, "semana", semana_id, fmt, "aprobo", a_uid, anom, current_user
        )
        sub, corte = _sub_corte_dummy_memoria()
        est = _merge_estilo_pdf(fc.get("estilo_pdf"), fmt)
        estilo_css = _memoria_pdf_estilo_css(est)

        parts: list[str] = []
        for item_numero in numeros:
            registros = fetch_registros_memoria_conciliacion(
                _sb, contrato_id, item_numero, semana_id=semana_id, item_exacto=True
            )
            if not registros:
                continue
            item_info = {
                "item_numero": registros[0].get("item_numero", item_numero),
                "item_descripcion": registros[0].get("item_descripcion", ""),
                "unidad": registros[0].get("unidad", ""),
            }
            inner = _html_memoria_item_body(
                contrato,
                sub,
                corte,
                item_info,
                registros,
                usuario_nombre,
                usuario_cargo,
                firma_cfg=firma_cfg,
                elaboro_firma_data_uri=elaboro_uri,
                reviso_firma_data_uri=reviso_uri,
                conc_meta=conc_meta,
                aprobo_firma_data_uri=aprobo_uri,
                aprobo_interventoria_desde_config=True,
                pie_fotos_contexto=f"Semana N° {nsem} · {fi} — {ff}",
            )
            if parts:
                parts.append("<pdf:nextpage />")
            parts.append(inner)

        if not parts:
            raise HTTPException(
                status_code=404,
                detail="No hay registros aprobados por ítem en esta semana",
            )

        html = _wrap_memoria_item_html("".join(parts), estilo_css)
        try:
            pdf_bytes = _to_pdf(html)
        except Exception as e:
            _log.exception("pdf_cc_sem_002_semana_completo: fallo PDF")
            raise HTTPException(500, f"Error generando PDF memoria semanal completa: {e!s}") from e

        fname = _safe_filename_part(f"CC-SEM-002_semana{nsem}_todos-items.pdf")
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("pdf_cc_sem_002_semana_completo: error no controlado")
        raise HTTPException(500, f"Error interno memoria PDF semanal completa: {repr(e)}")


@router.get("/{contrato_id}/pdf/cc-mes-002/acta/{acta_id}")
def pdf_cc_mes_002_acta(
    contrato_id: int,
    acta_id: int,
    item_numero: str = Query(...),
    current_user=Depends(_get_user),
):
    _perm_informes_ccd(current_user, "ver")
    if not _acta_pertenece_contrato(contrato_id, acta_id):
        raise HTTPException(404, "Acta no encontrada en este contrato")
    registros = fetch_registros_memoria_cc_mes_alineado_acta(
        _sb, contrato_id, item_numero, acta_rpo_id=acta_id, item_exacto=True
    )
    if not registros:
        raise HTTPException(404, "No hay registros para este ítem y acta")
    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    u = current_user if isinstance(current_user, dict) else dict(current_user)
    usuario_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
    usuario_cargo = u.get("cargo_nombre", "—") or "—"
    item_info = {
        "item_numero": registros[0].get("item_numero", item_numero),
        "item_descripcion": registros[0].get("item_descripcion", ""),
        "unidad": registros[0].get("unidad", ""),
    }
    ac = _row("actas", "numero_rpo, consecutivo", id=acta_id) or {}
    nrpo = str(ac.get("numero_rpo") or ac.get("consecutivo") or acta_id)
    cons = str(ac.get("consecutivo") or "—")
    fa = "—"
    conc_meta = {
        "titulo": "RESUMEN ACTIVIDADES — CONCILIACIÓN MENSUAL (INTERVENTORÍA–CONTRATISTA)",
        "codigo": CODIGO_FORMATO_CCD_CC_MES_002,
        "cells": [
            ("CONTRATO", str(contrato.get("numero") or "")),
            ("ACTA RPO", nrpo),
            ("CONSECUTIVO", cons),
            ("FECHA ACTA", fa),
        ],
    }
    fmt = CODIGO_FORMATO_CCD_CC_MES_002
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmt, contexto_tipo="acta_rpo", contexto_id=acta_id)
    fc = firma_cfg or {}
    e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
    r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
    a_uid = _opt_usuario_id(fc.get("aprobo_usuario_id"))
    enom = str(fc.get("elaboro_nombre") or "").strip()
    rnom = str(fc.get("reviso_nombre") or "").strip()
    anom = str(fc.get("aprobo_nombre") or "").strip()
    elaboro_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", acta_id, fmt, "elaboro", e_uid, enom, current_user
    )
    reviso_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", acta_id, fmt, "reviso", r_uid, rnom, current_user
    )
    aprobo_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", acta_id, fmt, "aprobo", a_uid, anom, current_user
    )
    sub, corte = _sub_corte_dummy_memoria()
    html = _html_memoria_item(
        contrato,
        sub,
        corte,
        item_info,
        registros,
        usuario_nombre,
        usuario_cargo,
        firma_cfg=firma_cfg,
        elaboro_firma_data_uri=elaboro_uri,
        reviso_firma_data_uri=reviso_uri,
        conc_meta=conc_meta,
        aprobo_firma_data_uri=aprobo_uri,
        aprobo_interventoria_desde_config=True,
        pie_fotos_contexto=f"Acta RPO {nrpo} · cons. {cons}",
        estilo_formato_codigo=fmt,
    )
    pdf_bytes = _to_pdf(html)
    item_safe = _safe_filename_part(item_numero.replace("/", "-").replace(" ", ""))
    fname = _safe_filename_part(f"CC-MES-002_acta_{nrpo}_{item_safe}.pdf")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{contrato_id}/pdf/cc-sem-001/semana/{semana_id}/con-sello-firma")
def pdf_cc_sem_001_semana_con_sello_firma(
    contrato_id: int,
    semana_id: int,
    current_user=Depends(_get_user),
):
    """Mismo PDF que CC-SEM-001 + página final de sello (firma de perfil, SHA-256, fecha)."""
    _perm_informes_ccd(current_user, "ver")
    if not _semana_pertenece_contrato(contrato_id, semana_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    reg = fetch_registros_conciliacion(_sb, contrato_id, semana_id=semana_id)
    items, total = aggregate_items_conciliacion(reg)
    _sort_items_corte_por_item_numero_asc(items)
    sm = _row("so_semanas", "id, numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
    nsem = sm.get("numero_semana")
    fi = str(sm.get("fecha_inicio") or "—")
    ff = str(sm.get("fecha_fin") or "—")
    pdf_bytes = _pdf_bytes_conciliacion_informe_v1(
        contrato_id,
        current_user,
        formato_codigo=CODIGO_FORMATO_CCD_CC_SEM_001,
        contexto_tipo="semana",
        contexto_id=semana_id,
        titulo_documento="INFORME EJECUCIÓN SEMANAL (CONCILIACIÓN INTERVENTORÍA–CONTRATISTA)",
        c3_label="SEMANA",
        c3_value=f"N° {nsem}",
        c4_label="VIGENCIA",
        c4_value=f"{fi} — {ff}",
        pie_contexto=f"Semana N° {nsem} · {fi} — {ff} · Registros nivel 3 aprobados y bloqueados",
        items=items,
        total_costo=total,
    )
    fname = _safe_filename_part(f"CC-SEM-001_semana_{nsem}.pdf")
    ctr = _row("contratos", "numero", id=contrato_id) or {}
    return _attachment_pdf_con_pagina_sello_usuario(
        pdf_bytes,
        current_user,
        titulo_doc=f"Informe ejecución semanal CC-SEM-001 — Semana N° {nsem} — {fi} — {ff}",
        formato_ccd=CODIGO_FORMATO_CCD_CC_SEM_001,
        contrato_numero=str(ctr.get("numero") or ""),
        nombre_archivo_pdf=fname,
    )


@router.get("/{contrato_id}/pdf/cc-mes-001/acta/{acta_id}/con-sello-firma")
def pdf_cc_mes_001_acta_con_sello_firma(
    contrato_id: int,
    acta_id: int,
    current_user=Depends(_get_user),
):
    """Mismo PDF que CC-MES-001 + página final de sello."""
    _perm_informes_ccd(current_user, "ver")
    if not _acta_pertenece_contrato(contrato_id, acta_id):
        raise HTTPException(404, "Acta no encontrada en este contrato")
    reg = fetch_registros_informe_cc_mes_por_acta(_sb, contrato_id, acta_id)
    items, total = aggregate_items_conciliacion(reg)
    _sort_items_corte_por_item_numero_asc(items)
    ac = _row("actas", "id, numero_rpo, consecutivo", id=acta_id) or {}
    nrpo = str(ac.get("numero_rpo") or ac.get("consecutivo") or acta_id)
    cons = str(ac.get("consecutivo") or "—")
    fa = "—"
    pdf_bytes = _pdf_bytes_conciliacion_informe_v1(
        contrato_id,
        current_user,
        formato_codigo=CODIGO_FORMATO_CCD_CC_MES_001,
        contexto_tipo="acta_rpo",
        contexto_id=acta_id,
        titulo_documento="INFORME EJECUCIÓN MENSUAL (CONCILIACIÓN INTERVENTORÍA–CONTRATISTA)",
        c3_label="ACTA RPO",
        c3_value=nrpo,
        c4_label="FECHA ACTA",
        c4_value=fa,
        pie_contexto=f"Acta RPO {nrpo} · consecutivo {cons} · Misma lógica que módulo Actas (N1·N2·N3 aprob. en cascada, costo directo por línea)",
        items=items,
        total_costo=total,
    )
    fname = _safe_filename_part(f"CC-MES-001_acta_{nrpo}.pdf")
    ctr = _row("contratos", "numero", id=contrato_id) or {}
    return _attachment_pdf_con_pagina_sello_usuario(
        pdf_bytes,
        current_user,
        titulo_doc=f"Informe ejecución mensual CC-MES-001 — Acta RPO {nrpo}",
        formato_ccd=CODIGO_FORMATO_CCD_CC_MES_001,
        contrato_numero=str(ctr.get("numero") or ""),
        nombre_archivo_pdf=fname,
    )


@router.get("/{contrato_id}/json/informe-gerencia")
def json_informe_gerencia_pareja(
    contrato_id: int,
    acta_presente: int = Query(..., description="Acta RPO cuyo contexto CCD aplica a firmas"),
    acta_referencia: Optional[int] = Query(
        None, description="Otro acta RPO para columnas de comparación (opcional)"
    ),
    current_user=Depends(_get_user),
):
    """JSON (modo clásico) dos actas. Preferir /json/informe-gerencia-matriz para 4 columnas v2."""
    _perm_informes_ccd(current_user, "ver")
    if not _acta_pertenece_contrato(contrato_id, acta_presente):
        raise HTTPException(404, "Acta (presente) no encontrada")
    a_ref = int(acta_referencia) if acta_referencia is not None else None
    if a_ref is not None and (not _acta_pertenece_contrato(contrato_id, a_ref) or a_ref == int(acta_presente)):
        raise HTTPException(400, "Acta de referencia no válida")
    ids: List[int] = [int(acta_presente)]
    if a_ref is not None:
        ids.append(a_ref)
    rmap = rpo_conciliacion_por_contrato(_sb, contrato_id, ids)
    d_p = rmap.get(int(acta_presente)) or _rpo_gerencia_vacio()
    d_a = rmap.get(int(a_ref)) if a_ref is not None else None
    if d_a is None and a_ref is not None:
        d_a = _rpo_gerencia_vacio()
    filas = _merge_filas_gerencia_dos_actas(d_p.get("por_capitulo"), d_a.get("por_capitulo") if d_a else None)
    return {
        "formato_ccd": CODIGO_FORMATO_CCD_CC_GER_001,
        "criterio": "Costo directo aprobado en interventoría: N1, N2 y N3 aprobado en cascada (SICOE Obra).",
        "acta_presente_id": int(acta_presente),
        "acta_referencia_id": a_ref,
        "contexto_firma_ccd": {"tipo": "acta_rpo", "id": int(acta_presente)},
        "totales": {
            "acta_presente_cdirecto": d_p.get("costo_directo_total"),
            "acta_referencia_cdirecto": (d_a or {}).get("costo_directo_total") if a_ref is not None else None,
        },
        "por_capitulo_merged": [
            {"capitulo": a, "costo_presente": b, "costo_referencia": c, "diferencia": d} for a, b, c, d in filas
        ],
        "secciones_presente": d_p.get("secciones") or {},
    }


@router.get("/{contrato_id}/json/informe-gerencia-matriz")
def json_informe_gerencia_matriz(
    contrato_id: int,
    acta_presente: Optional[int] = Query(
        None,
        description="Opcional. Si se omite: acta RPO **en período** (hoy ∈ fechas del acta), "
        "igual que la matriz SICOE; si no aplica ninguna, la de mayor consecutivo.",
    ),
    refresh: bool = Query(
        False,
        description="true = ignora caché en memoria y recalcula (usar tras correcciones o si el PDF salió mal).",
    ),
    current_user=Depends(_get_user),
):
    """4 columnas (habil. cobro, acta ant. N3, acum. N3, pendiente) por capítulo, costo directo."""
    _perm_informes_ccd(current_user, "ver")
    if refresh:
        _gerencia_caches_clear()
    d = _construir_datos_informe_gerencia_matriz(
        contrato_id, acta_presente, skip_cache=refresh
    )
    ap = d.get("acta_presente") or {}
    return {
        "formato_ccd": CODIGO_FORMATO_CCD_CC_GER_001,
        "v": 2,
        "contexto_firma_ccd": {"tipo": "acta_rpo", "id": int(ap.get("id") or 0)},
        **d,
    }


def _pdf_cc_ger_001_bytes_cached(
    contrato_id: int,
    current_user,
    acta_presente: Optional[int],
    acta_referencia: Optional[int],
    modo: str,
    *,
    skip_cache: bool = False,
) -> Tuple[bytes, str]:
    m = (modo or "").lower().strip()
    cache_key = _gerencia_pdf_cache_key(contrato_id, acta_presente, m, False)
    if not skip_cache:
        cached = _gerencia_pdf_cache_get(cache_key)
        if cached is not None:
            return cached

    if m in ("pareja", "2", "dos", "clásico", "clasico"):
        raw = _pdf_bytes_informe_gerencia_pareja(
            contrato_id, int(acta_presente), acta_referencia, current_user
        )
        ac = _row("actas", "numero_rpo, consecutivo", id=int(acta_presente)) or {}
        nr = str(ac.get("numero_rpo") or ac.get("consecutivo") or acta_presente)
    else:
        datosg = _construir_datos_informe_gerencia_matriz(
            contrato_id, acta_presente, skip_cache=skip_cache
        )
        raw = _pdf_bytes_informe_gerencia_matriz(
            contrato_id, current_user, acta_presente, datos=datosg
        )
        apd = (datosg.get("acta_presente") or {}) or {}
        nr = str(apd.get("numero_rpo") or apd.get("consecutivo") or apd.get("id") or "ger")
    if not skip_cache:
        _gerencia_pdf_cache_set(cache_key, raw, nr)
    return raw, nr


@router.get("/{contrato_id}/pdf/cc-ger-001")
def pdf_cc_ger_001_pareja(
    contrato_id: int,
    acta_presente: Optional[int] = Query(
        None,
        description="Opcional. Si se omite: acta RPO en período (matriz SICOE); si no hay, mayor consecutivo.",
    ),
    acta_referencia: Optional[int] = Query(
        None, description="(Deprecada en v2) se ignora; usaba comparación 2 actas en modo clásico."
    ),
    modo: str = Query(
        "matriz", description="matriz = 4 col. (defecto). pareja = dos actas requiere acta_presente (Query obligatorio manual)."
    ),
    refresh: bool = Query(False, description="true = recalcula sin caché PDF/matríz."),
    current_user=Depends(_get_user),
):
    """
    PDF CC-GER-001. Por defecto: 4 columnas (horizontal). Acta presente = RPO en período (misma lógica que matriz SICOE), salvo `acta_presente`.
    modo=pareja: comparación 2 actas; acta_presente requerida; acta_referencia opcional.
    """
    _perm_informes_ccd(current_user, "ver")
    m = (modo or "").lower().strip()
    if m in ("pareja", "2", "dos", "clásico", "clasico") and acta_presente is None:
        raise HTTPException(400, "Con modo=pareja debe indicar acta_presente")
    if refresh:
        _gerencia_caches_clear()
    raw, nr = _pdf_cc_ger_001_bytes_cached(
        contrato_id, current_user, acta_presente, acta_referencia, m, skip_cache=refresh
    )
    fname = _safe_filename_part(f"CC-GER-001_RPO{nr}.pdf")
    return Response(
        content=raw,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{contrato_id}/pdf/cc-ger-001/con-sello-firma")
def pdf_cc_ger_001_pareja_con_sello_firma(
    contrato_id: int,
    acta_presente: Optional[int] = Query(
        None,
        description="Opcional en matriz: RPO en período por defecto; forzar id de acta si se requiere.",
    ),
    acta_referencia: Optional[int] = Query(
        None, description="(Deprecada) solo modo pareja; ignorada en matriz v2"
    ),
    modo: str = Query("matriz", description="matriz | pareja"),
    refresh: bool = Query(False, description="true = recalcula sin caché PDF/matríz."),
    current_user=Depends(_get_user),
):
    """Mismo PDF CC-GER-001 (matriz v2 o pareja) + página de sello (huella, fecha)."""
    _perm_informes_ccd(current_user, "ver")
    m = (modo or "").lower().strip()
    if m in ("pareja", "2", "dos", "clásico", "clasico") and acta_presente is None:
        raise HTTPException(400, "Con modo=pareja debe indicar acta_presente")
    if refresh:
        _gerencia_caches_clear()
    pdf_bytes, nr = _pdf_cc_ger_001_bytes_cached(
        contrato_id, current_user, acta_presente, acta_referencia, m, skip_cache=refresh
    )
    fname = _safe_filename_part(f"CC-GER-001_RPO{nr}.pdf")
    ctr = _row("contratos", "numero", id=contrato_id) or {}
    return _attachment_pdf_con_pagina_sello_usuario(
        pdf_bytes,
        current_user,
        titulo_doc=f"Informe gerencia CC-GER-001 — Acta RPO {nr}",
        formato_ccd=CODIGO_FORMATO_CCD_CC_GER_001,
        contrato_numero=str(ctr.get("numero") or ""),
        nombre_archivo_pdf=fname,
    )


@router.get("/{contrato_id}/pdf/cc-sem-002/semana/{semana_id}/con-sello-firma")
def pdf_cc_sem_002_semana_con_sello_firma(
    contrato_id: int,
    semana_id: int,
    item_numero: str = Query(...),
    current_user=Depends(_get_user),
):
    """Misma memoria CC-SEM-002 que por ítem + página de sello."""
    _perm_informes_ccd(current_user, "ver")
    if not _semana_pertenece_contrato(contrato_id, semana_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    registros = fetch_registros_memoria_conciliacion(
        _sb, contrato_id, item_numero, semana_id=semana_id, item_exacto=True
    )
    if not registros:
        raise HTTPException(404, "No hay registros para este ítem y semana")
    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    u = current_user if isinstance(current_user, dict) else dict(current_user)
    usuario_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
    usuario_cargo = u.get("cargo_nombre", "—") or "—"
    item_info = {
        "item_numero": registros[0].get("item_numero", item_numero),
        "item_descripcion": registros[0].get("item_descripcion", ""),
        "unidad": registros[0].get("unidad", ""),
    }
    sm = _row("so_semanas", "numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
    nsem = sm.get("numero_semana")
    fi = str(sm.get("fecha_inicio") or "—")
    ff = str(sm.get("fecha_fin") or "—")
    conc_meta = {
        "titulo": "RESUMEN ACTIVIDADES — CONCILIACIÓN SEMANAL (INTERVENTORÍA–CONTRATISTA)",
        "codigo": CODIGO_FORMATO_CCD_CC_SEM_002,
        "cells": [
            ("CONTRATO", str(contrato.get("numero") or "")),
            ("SEMANA", f"N° {nsem}"),
            ("VIGENCIA", f"{fi} — {ff}"),
            ("REFERENCIA", "Cantidades ejecutadas — conciliación"),
        ],
    }
    fmt = CODIGO_FORMATO_CCD_CC_SEM_002
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmt, contexto_tipo="semana", contexto_id=semana_id)
    fc = firma_cfg or {}
    e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
    r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
    a_uid = _opt_usuario_id(fc.get("aprobo_usuario_id"))
    enom = str(fc.get("elaboro_nombre") or "").strip()
    rnom = str(fc.get("reviso_nombre") or "").strip()
    anom = str(fc.get("aprobo_nombre") or "").strip()
    elaboro_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "semana", semana_id, fmt, "elaboro", e_uid, enom, current_user
    )
    reviso_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "semana", semana_id, fmt, "reviso", r_uid, rnom, current_user
    )
    aprobo_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "semana", semana_id, fmt, "aprobo", a_uid, anom, current_user
    )
    sub, corte = _sub_corte_dummy_memoria()
    html = _html_memoria_item(
        contrato,
        sub,
        corte,
        item_info,
        registros,
        usuario_nombre,
        usuario_cargo,
        firma_cfg=firma_cfg,
        elaboro_firma_data_uri=elaboro_uri,
        reviso_firma_data_uri=reviso_uri,
        conc_meta=conc_meta,
        aprobo_firma_data_uri=aprobo_uri,
        aprobo_interventoria_desde_config=True,
        pie_fotos_contexto=f"Semana N° {nsem} · {fi} — {ff}",
        estilo_formato_codigo=fmt,
    )
    pdf_bytes = _to_pdf(html)
    item_safe = _safe_filename_part(item_numero.replace("/", "-").replace(" ", ""))
    fname = _safe_filename_part(f"CC-SEM-002_semana{nsem}_{item_safe}.pdf")
    return _attachment_pdf_con_pagina_sello_usuario(
        pdf_bytes,
        current_user,
        titulo_doc=f"Memoria CC-SEM-002 — Ítem {item_numero} — Semana N° {nsem}",
        formato_ccd=CODIGO_FORMATO_CCD_CC_SEM_002,
        contrato_numero=str(contrato.get("numero") or ""),
        nombre_archivo_pdf=fname,
    )


@router.get("/{contrato_id}/pdf/cc-sem-002/semana/{semana_id}/completo/con-sello-firma")
def pdf_cc_sem_002_semana_completo_con_sello_firma(
    contrato_id: int,
    semana_id: int,
    current_user=Depends(_get_user),
):
    """Mismo PDF «todos los ítems» CC-SEM-002 + página de sello."""
    _perm_informes_ccd(current_user, "ver")
    try:
        if not _semana_pertenece_contrato(contrato_id, semana_id):
            raise HTTPException(404, "Semana no encontrada en este contrato")
        reg_agg = fetch_registros_conciliacion(_sb, contrato_id, semana_id=semana_id)
        items_agg, _total = aggregate_items_conciliacion(reg_agg)
        _sort_items_corte_por_item_numero_asc(items_agg)
        numeros = [
            (it.get("item_numero") or "").strip()
            for it in items_agg
            if (it.get("item_numero") or "").strip()
        ]
        if not numeros:
            raise HTTPException(
                status_code=404,
                detail="No hay registros en esta semana para generar memorias",
            )

        contrato = _row(
            "contratos",
            "numero, objeto, contratista, nit, interventoria, logo_contratista",
            id=contrato_id,
        )
        if not contrato:
            raise HTTPException(404, "Contrato no encontrado")
        u = current_user if isinstance(current_user, dict) else dict(current_user)
        usuario_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
        usuario_cargo = u.get("cargo_nombre", "—") or "—"
        sm = _row("so_semanas", "numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
        nsem = sm.get("numero_semana")
        fi = str(sm.get("fecha_inicio") or "—")
        ff = str(sm.get("fecha_fin") or "—")
        conc_meta = {
            "titulo": "RESUMEN ACTIVIDADES — CONCILIACIÓN SEMANAL (INTERVENTORÍA–CONTRATISTA)",
            "codigo": CODIGO_FORMATO_CCD_CC_SEM_002,
            "cells": [
                ("CONTRATO", str(contrato.get("numero") or "")),
                ("SEMANA", f"N° {nsem}"),
                ("VIGENCIA", f"{fi} — {ff}"),
                ("REFERENCIA", "Cantidades ejecutadas — conciliación"),
            ],
        }
        fmt = CODIGO_FORMATO_CCD_CC_SEM_002
        firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmt, contexto_tipo="semana", contexto_id=semana_id)
        fc = firma_cfg or {}
        e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
        r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
        a_uid = _opt_usuario_id(fc.get("aprobo_usuario_id"))
        enom = str(fc.get("elaboro_nombre") or "").strip()
        rnom = str(fc.get("reviso_nombre") or "").strip()
        anom = str(fc.get("aprobo_nombre") or "").strip()
        elaboro_uri = _firma_data_uri_para_slot_contexto(
            contrato_id, "semana", semana_id, fmt, "elaboro", e_uid, enom, current_user
        )
        reviso_uri = _firma_data_uri_para_slot_contexto(
            contrato_id, "semana", semana_id, fmt, "reviso", r_uid, rnom, current_user
        )
        aprobo_uri = _firma_data_uri_para_slot_contexto(
            contrato_id, "semana", semana_id, fmt, "aprobo", a_uid, anom, current_user
        )
        sub, corte = _sub_corte_dummy_memoria()
        est = _merge_estilo_pdf(fc.get("estilo_pdf"), fmt)
        estilo_css = _memoria_pdf_estilo_css(est)

        parts: list[str] = []
        for item_numero in numeros:
            registros = fetch_registros_memoria_conciliacion(
                _sb, contrato_id, item_numero, semana_id=semana_id, item_exacto=True
            )
            if not registros:
                continue
            item_info = {
                "item_numero": registros[0].get("item_numero", item_numero),
                "item_descripcion": registros[0].get("item_descripcion", ""),
                "unidad": registros[0].get("unidad", ""),
            }
            inner = _html_memoria_item_body(
                contrato,
                sub,
                corte,
                item_info,
                registros,
                usuario_nombre,
                usuario_cargo,
                firma_cfg=firma_cfg,
                elaboro_firma_data_uri=elaboro_uri,
                reviso_firma_data_uri=reviso_uri,
                conc_meta=conc_meta,
                aprobo_firma_data_uri=aprobo_uri,
                aprobo_interventoria_desde_config=True,
                pie_fotos_contexto=f"Semana N° {nsem} · {fi} — {ff}",
            )
            if parts:
                parts.append("<pdf:nextpage />")
            parts.append(inner)

        if not parts:
            raise HTTPException(
                status_code=404,
                detail="No hay registros aprobados por ítem en esta semana",
            )

        html = _wrap_memoria_item_html("".join(parts), estilo_css)
        try:
            pdf_bytes = _to_pdf(html)
        except Exception as e:
            _log.exception("pdf_cc_sem_002_semana_completo_con_sello_firma: fallo PDF")
            raise HTTPException(500, f"Error generando PDF memoria semanal completa: {e!s}") from e

        fname = _safe_filename_part(f"CC-SEM-002_semana{nsem}_todos-items.pdf")
        return _attachment_pdf_con_pagina_sello_usuario(
            pdf_bytes,
            current_user,
            titulo_doc=f"Memorias CC-SEM-002 — Semana N° {nsem} — todos los ítems",
            formato_ccd=CODIGO_FORMATO_CCD_CC_SEM_002,
            contrato_numero=str(contrato.get("numero") or ""),
            nombre_archivo_pdf=fname,
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("pdf_cc_sem_002_semana_completo_con_sello_firma: error no controlado")
        raise HTTPException(500, f"Error interno memoria PDF semanal completa: {repr(e)}")


@router.get("/{contrato_id}/pdf/cc-mes-002/acta/{acta_id}/con-sello-firma")
def pdf_cc_mes_002_acta_con_sello_firma(
    contrato_id: int,
    acta_id: int,
    item_numero: str = Query(...),
    current_user=Depends(_get_user),
):
    """Misma memoria CC-MES-002 + página de sello."""
    _perm_informes_ccd(current_user, "ver")
    if not _acta_pertenece_contrato(contrato_id, acta_id):
        raise HTTPException(404, "Acta no encontrada en este contrato")
    registros = fetch_registros_memoria_cc_mes_alineado_acta(
        _sb, contrato_id, item_numero, acta_rpo_id=acta_id, item_exacto=True
    )
    if not registros:
        raise HTTPException(404, "No hay registros para este ítem y acta")
    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    u = current_user if isinstance(current_user, dict) else dict(current_user)
    usuario_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
    usuario_cargo = u.get("cargo_nombre", "—") or "—"
    item_info = {
        "item_numero": registros[0].get("item_numero", item_numero),
        "item_descripcion": registros[0].get("item_descripcion", ""),
        "unidad": registros[0].get("unidad", ""),
    }
    ac = _row("actas", "numero_rpo, consecutivo", id=acta_id) or {}
    nrpo = str(ac.get("numero_rpo") or ac.get("consecutivo") or acta_id)
    cons = str(ac.get("consecutivo") or "—")
    fa = "—"
    conc_meta = {
        "titulo": "RESUMEN ACTIVIDADES — CONCILIACIÓN MENSUAL (INTERVENTORÍA–CONTRATISTA)",
        "codigo": CODIGO_FORMATO_CCD_CC_MES_002,
        "cells": [
            ("CONTRATO", str(contrato.get("numero") or "")),
            ("ACTA RPO", nrpo),
            ("CONSECUTIVO", cons),
            ("FECHA ACTA", fa),
        ],
    }
    fmt = CODIGO_FORMATO_CCD_CC_MES_002
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmt, contexto_tipo="acta_rpo", contexto_id=acta_id)
    fc = firma_cfg or {}
    e_uid = _opt_usuario_id(fc.get("elaboro_usuario_id"))
    r_uid = _opt_usuario_id(fc.get("reviso_usuario_id"))
    a_uid = _opt_usuario_id(fc.get("aprobo_usuario_id"))
    enom = str(fc.get("elaboro_nombre") or "").strip()
    rnom = str(fc.get("reviso_nombre") or "").strip()
    anom = str(fc.get("aprobo_nombre") or "").strip()
    elaboro_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", acta_id, fmt, "elaboro", e_uid, enom, current_user
    )
    reviso_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", acta_id, fmt, "reviso", r_uid, rnom, current_user
    )
    aprobo_uri = _firma_data_uri_para_slot_contexto(
        contrato_id, "acta_rpo", acta_id, fmt, "aprobo", a_uid, anom, current_user
    )
    sub, corte = _sub_corte_dummy_memoria()
    html = _html_memoria_item(
        contrato,
        sub,
        corte,
        item_info,
        registros,
        usuario_nombre,
        usuario_cargo,
        firma_cfg=firma_cfg,
        elaboro_firma_data_uri=elaboro_uri,
        reviso_firma_data_uri=reviso_uri,
        conc_meta=conc_meta,
        aprobo_firma_data_uri=aprobo_uri,
        aprobo_interventoria_desde_config=True,
        pie_fotos_contexto=f"Acta RPO {nrpo} · cons. {cons}",
        estilo_formato_codigo=fmt,
    )
    pdf_bytes = _to_pdf(html)
    item_safe = _safe_filename_part(item_numero.replace("/", "-").replace(" ", ""))
    fname = _safe_filename_part(f"CC-MES-002_acta_{nrpo}_{item_safe}.pdf")
    return _attachment_pdf_con_pagina_sello_usuario(
        pdf_bytes,
        current_user,
        titulo_doc=f"Memoria CC-MES-002 — Ítem {item_numero} — Acta RPO {nrpo}",
        formato_ccd=CODIGO_FORMATO_CCD_CC_MES_002,
        contrato_numero=str(contrato.get("numero") or ""),
        nombre_archivo_pdf=fname,
    )


_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.get("/{contrato_id}/excel/memoria-item/{corte_id}")
def excel_memoria_item(
    contrato_id: int,
    corte_id: int,
    item_numero: str = Query(...),
    current_user=Depends(_get_user),
):
    _perm_informes_ccd(current_user, "exportar")
    try:
        ctx = _contexto_memoria_item(contrato_id, corte_id, item_numero, current_user)
        firma_cfg = _get_firma_cfg_para_documento(contrato_id, CODIGO_FORMATO_CCD_CC_SUB_002, corte_id=corte_id)
        xbytes = _memoria_item_excel_bytes(
            ctx["contrato"],
            ctx["sub"],
            ctx["corte"],
            ctx["item_info"],
            ctx["registros"],
            firma_cfg,
        )
        item_safe = _safe_filename_part(item_numero.replace("/", "-").replace(" ", ""))
        cons = (ctx["corte"] or {}).get("consecutivo", "")
        fname = _safe_filename_part(f"CC-SUB-002_Corte{cons}_{item_safe}.xlsx")
        return Response(
            content=xbytes,
            media_type=_XLSX_MEDIA,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("excel_memoria_item")
        raise HTTPException(500, f"Error generando Excel memoria: {repr(e)}") from e


@router.get("/{contrato_id}/excel/memoria-corte-completo/{corte_id}")
def excel_memoria_corte_completo(
    contrato_id: int,
    corte_id: int,
    current_user=Depends(_get_user),
):
    """Un libro con una hoja por ítem (orden ascendente por código, igual que el PDF completo)."""
    _perm_informes_ccd(current_user, "exportar")
    try:
        numeros = _list_item_numeros_memoria_corte(contrato_id, corte_id)
        if not numeros:
            raise HTTPException(
                status_code=404,
                detail="No hay registros aprobados en este corte para generar memorias",
            )
        firma_cfg = _get_firma_cfg_para_documento(contrato_id, CODIGO_FORMATO_CCD_CC_SUB_002, corte_id=corte_id)
        xbytes = _memoria_corte_completo_excel_bytes(
            contrato_id, corte_id, numeros, current_user, firma_cfg
        )
        corte_row = _row("subcontratista_cortes", "consecutivo", id=corte_id) or {}
        cons = corte_row.get("consecutivo", "")
        fname = _safe_filename_part(f"CC-SUB-002_Corte{cons}_todos-items.xlsx")
        return Response(
            content=xbytes,
            media_type=_XLSX_MEDIA,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("excel_memoria_corte_completo")
        raise HTTPException(500, f"Error generando Excel memoria completa: {repr(e)}") from e


@router.get("/{contrato_id}/excel/corte-subcontratista/{corte_id}")
def excel_corte_subcontratista(
    contrato_id: int,
    corte_id: int,
    current_user=Depends(_get_user),
):
    """Excel CC-SUB-001: mismo contenido lógico que el informe de corte (PDF)."""
    _perm_informes_ccd(current_user, "exportar")
    try:
        ctx = _contexto_corte_sub(contrato_id, corte_id, current_user)
        firma_cfg = _get_firma_cfg_para_documento(contrato_id, CODIGO_FORMATO_CCD_CC_SUB_001, corte_id=corte_id)
        xbytes = _corte_sub_001_excel_bytes(
            ctx["contrato"],
            ctx["sub"],
            ctx["corte"],
            ctx["items"],
            float(ctx["total_costo"] or 0.0),
            ctx["usuario_nombre"],
            ctx["usuario_cargo"],
            firma_cfg,
        )
        fname = _nombre_archivo_cc_sub_001(ctx["corte"], ctx["sub"], corte_id).replace(".pdf", ".xlsx")
        return Response(
            content=xbytes,
            media_type=_XLSX_MEDIA,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("excel_corte_subcontratista")
        raise HTTPException(500, f"Error generando Excel corte subcontratista: {repr(e)}") from e


@router.get("/{contrato_id}/excel/cc-sem-001/semana/{semana_id}")
def excel_cc_sem_001_semana(
    contrato_id: int,
    semana_id: int,
    current_user=Depends(_get_user),
):
    """Excel CC-SEM-001: mismo contenido lógico que el informe semanal (PDF)."""
    _perm_informes_ccd(current_user, "exportar")
    try:
        xbytes = _cc_sem_001_excel_bytes(contrato_id, semana_id, current_user)
        sm = _row("so_semanas", "numero_semana", id=semana_id) or {}
        nsem = sm.get("numero_semana")
        fname = _safe_filename_part(f"CC-SEM-001_semana_{nsem}.xlsx")
        return Response(
            content=xbytes,
            media_type=_XLSX_MEDIA,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("excel_cc_sem_001_semana")
        raise HTTPException(500, f"Error generando Excel CC-SEM-001: {repr(e)}") from e


@router.get("/{contrato_id}/excel/cc-sem-002/semana/{semana_id}")
def excel_cc_sem_002_semana(
    contrato_id: int,
    semana_id: int,
    item_numero: str = Query(...),
    current_user=Depends(_get_user),
):
    """Excel CC-SEM-002 por ítem: misma estructura que la memoria PDF semanal."""
    _perm_informes_ccd(current_user, "exportar")
    try:
        xbytes = _cc_sem_002_item_excel_bytes(contrato_id, semana_id, item_numero, current_user)
        sm = _row("so_semanas", "numero_semana", id=semana_id) or {}
        nsem = sm.get("numero_semana")
        item_safe = _safe_filename_part(item_numero.replace("/", "-").replace(" ", ""))
        fname = _safe_filename_part(f"CC-SEM-002_semana{nsem}_{item_safe}.xlsx")
        return Response(
            content=xbytes,
            media_type=_XLSX_MEDIA,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("excel_cc_sem_002_semana")
        raise HTTPException(500, f"Error generando Excel CC-SEM-002: {repr(e)}") from e


@router.get("/{contrato_id}/excel/cc-sem-002/semana/{semana_id}/completo")
def excel_cc_sem_002_semana_completo(
    contrato_id: int,
    semana_id: int,
    current_user=Depends(_get_user),
):
    """Excel CC-SEM-002: una hoja por ítem (todos los ítems de la semana)."""
    _perm_informes_ccd(current_user, "exportar")
    try:
        xbytes = _cc_sem_002_semana_completo_excel_bytes(contrato_id, semana_id, current_user)
        sm = _row("so_semanas", "numero_semana", id=semana_id) or {}
        nsem = sm.get("numero_semana")
        fname = _safe_filename_part(f"CC-SEM-002_semana{nsem}_todos-items.xlsx")
        return Response(
            content=xbytes,
            media_type=_XLSX_MEDIA,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("excel_cc_sem_002_semana_completo")
        raise HTTPException(500, f"Error generando Excel CC-SEM-002 completo: {repr(e)}") from e


# ── Preview / desarrollo ───────────────────────────────────────────────────────

def _orden_item_numero(num: str) -> tuple:
    """Sort key natural para números de ítem: NP-001 < NP-010, 1.01 < 1.10 < 2.01."""
    parts = re.split(r"(\d+)", (num or "").strip().upper())
    return tuple(int(p) if p.isdigit() else p for p in parts)


def _fetch_foto_grafico_item(
    contrato_id: int, acta_id: int, item_numero: str, capitulo: str = ""
) -> dict:
    """
    Devuelve dict con {foto_url, foto_numero, grafico_url, grafico_numero} del primer
    registro que coincida con contrato + acta + ítem + capítulo.
    """
    empty = {"foto_url": None, "foto_numero": None, "grafico_url": None, "grafico_numero": None}
    if not item_numero:
        return empty

    def _extraer(rows: list) -> dict:
        foto_url = next(
            (r.get("foto_url") for r in rows if str(r.get("foto_url") or "").strip()), None
        )
        foto_num = next(
            (r.get("foto_numero") for r in rows if str(r.get("foto_url") or "").strip()), None
        )
        graf_url = next(
            (r.get("grafico_url") for r in rows if str(r.get("grafico_url") or "").strip()), None
        )
        graf_num = next(
            (r.get("grafico_numero") for r in rows if str(r.get("grafico_url") or "").strip()), None
        )
        return {"foto_url": foto_url, "foto_numero": foto_num,
                "grafico_url": graf_url, "grafico_numero": graf_num}

    def _q_registros(q):
        """Aplica filtros comunes a una query de so_registros."""
        q = q.eq("item_numero", item_numero)
        if (capitulo or "").strip():
            q = q.eq("capitulo", capitulo.strip())
        return q

    try:
        sel = "foto_url, foto_numero, grafico_url, grafico_numero"

        # 1. Vía so_reportes.acta_rpo_id → reporte_ids → so_registros (semántica dashboard)
        if acta_id:
            rp_batch = (
                _sb.table("so_reportes")
                .select("id")
                .eq("contrato_id", int(contrato_id))
                .eq("acta_rpo_id", int(acta_id))
                .limit(1000)
                .execute()
                .data or []
            )
            reporte_ids = [r["id"] for r in rp_batch if r.get("id")]
            if reporte_ids:
                CHUNK = 200
                rows: list = []
                for i in range(0, len(reporte_ids), CHUNK):
                    q = (
                        _sb.table("so_registros")
                        .select(sel)
                        .eq("contrato_id", int(contrato_id))
                        .in_("reporte_id", reporte_ids[i : i + CHUNK])
                    )
                    part = _q_registros(q).limit(200).execute().data or []
                    rows.extend(part)
                result = _extraer(rows)
                if result["foto_url"] or result["grafico_url"]:
                    _log.info("fetch_foto_grafico via reportes: item=%s cap=%s acta=%s foto=#%s grf=#%s",
                              item_numero, capitulo, acta_id, result["foto_numero"], result["grafico_numero"])
                    return result

        # 2. Fallback directo por so_registros.acta_rpo_id
        if acta_id:
            q = (
                _sb.table("so_registros")
                .select(sel)
                .eq("contrato_id", int(contrato_id))
                .eq("acta_rpo_id", int(acta_id))
            )
            rows_d = _q_registros(q).limit(100).execute().data or []
            result = _extraer(rows_d)
            if result["foto_url"] or result["grafico_url"]:
                _log.info("fetch_foto_grafico via acta_rpo_id: item=%s cap=%s foto=#%s grf=#%s",
                          item_numero, capitulo, result["foto_numero"], result["grafico_numero"])
                return result

        _log.info("fetch_foto_grafico: sin imagen para item=%s cap=%s acta=%s", item_numero, capitulo, acta_id)
        return empty

    except Exception as exc:
        _log.warning("fetch_foto_grafico_item ERROR item=%s: %s", item_numero, exc)
        return empty


def _url_a_data_url_pdf(url: Optional[str]) -> Optional[str]:
    """
    Descarga una imagen desde una URL y la convierte a data-URL base64 (PDF-safe).
    Convierte GIF → PNG automáticamente (xhtml2pdf no renderiza GIF).
    Devuelve None si falla o si la URL está vacía.
    """
    if not url or not str(url).strip().startswith("http"):
        return None
    try:
        import base64, io
        resp = requests.get(url.strip(), timeout=_FO_EO04_IMG_FETCH_TIMEOUT_SEC)
        resp.raise_for_status()
        ct = resp.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()
        data = resp.content
        from PIL import Image
        img = Image.open(io.BytesIO(data))
        if ct == "image/gif" or url.lower().endswith(".gif"):
            img = img.convert("RGBA")
            ct = "image/png"
        elif img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
            ct = "image/jpeg"
        elif img.mode != "RGB":
            img = img.convert("RGB")
            ct = "image/jpeg"
        w, h = img.size
        if max(w, h) > _FO_EO04_IMG_MAX_PX:
            scale = _FO_EO04_IMG_MAX_PX / float(max(w, h))
            img = img.resize((max(1, int(w * scale)), max(1, int(h * scale))), Image.Resampling.LANCZOS)
        buf = io.BytesIO()
        if ct == "image/png":
            img.save(buf, format="PNG", optimize=True)
        else:
            img.save(buf, format="JPEG", quality=75, optimize=True)
            ct = "image/jpeg"
        data = buf.getvalue()
        b64 = base64.b64encode(data).decode()
        return f"data:{ct};base64,{b64}"
    except Exception as exc:
        _log.warning("_url_a_data_url_pdf ERROR url=%s: %s", (url or "")[:80], exc)
        return None


_FO_EO_04_SEL_REGISTROS = (
    "id, acta_rpo_id, reporte_id, capitulo, item_numero, item_descripcion, unidad, "
    "nivel1_estado, nivel2_estado, nivel3_estado, nivel4_estado, nivel5_estado, nivel6_estado, "
    "cantidad_total, semana_id, foto_url, foto_numero, grafico_url, grafico_numero"
)
_FO_EO_04_SEL_TOTALES = (
    "id, acta_rpo_id, reporte_id, capitulo, item_numero, "
    "nivel1_estado, nivel2_estado, nivel3_estado, nivel4_estado, nivel5_estado, nivel6_estado, "
    "cantidad_total"
)


def _fo_eo_04_paginar_so_registros(q_builder) -> list:
    """Pagina una query de so_registros (máx. 200k filas)."""
    _PAGE = 1000
    rows: list = []
    off = 0
    for _ in range(200):
        part = q_builder().range(off, off + _PAGE - 1).execute().data or []
        rows.extend(part)
        if len(part) < _PAGE:
            break
        off += _PAGE
    return rows


def _fo_eo_04_fetch_registros_acta(contrato_id: int, acta_id: int) -> List[Dict[str, Any]]:
    """
    Líneas del acta: acta_rpo_id en el registro y, además, registros de reportes con
    so_reportes.acta_rpo_id = acta (muchos contratos no replican acta_rpo_id en cada línea).
    Excluye líneas con acta_rpo_id distinto al acta seleccionado.
    """
    cid, aid = int(contrato_id), int(acta_id)

    def _base_q():
        return (
            _sb.table("so_registros")
            .select(_FO_EO_04_SEL_REGISTROS)
            .eq("contrato_id", cid)
            .not_.is_("item_numero", "null")
            .neq("item_numero", "")
        )

    rows_direct = _fo_eo_04_paginar_so_registros(lambda: _base_q().eq("acta_rpo_id", aid))

    reporte_ids: List[int] = []
    try:
        rp = (
            _sb.table("so_reportes")
            .select("id")
            .eq("contrato_id", cid)
            .eq("acta_rpo_id", aid)
            .execute()
            .data
            or []
        )
        reporte_ids = [int(x["id"]) for x in rp if x.get("id") is not None]
    except Exception as exc:
        _log.warning("fo_eo_04 reportes acta: %s", exc)

    rows_rep: list = []
    for i in range(0, len(reporte_ids), 80):
        chunk = reporte_ids[i : i + 80]

        def _q_rep(ids=chunk):
            # Solo líneas sin acta en registro (las de acta_rpo_id=acta ya están en rows_direct)
            return _base_q().in_("reporte_id", ids).is_("acta_rpo_id", "null")

        rows_rep.extend(_fo_eo_04_paginar_so_registros(_q_rep))

    by_id: Dict[Any, Dict[str, Any]] = {}
    for r in rows_direct:
        rid = r.get("id")
        if rid is not None:
            by_id[rid] = r
    for r in rows_rep:
        rid = r.get("id")
        if rid is None:
            continue
        ara = r.get("acta_rpo_id")
        if ara is not None:
            try:
                if int(ara) != aid:
                    continue
            except (TypeError, ValueError):
                continue
        by_id[rid] = r

    rows = list(by_id.values())
    _log.info(
        "fo_eo_04 registros_acta: acta=%s direct=%s via_reporte=%s uniq=%s reportes=%s",
        aid,
        len(rows_direct),
        len(rows_rep),
        len(rows),
        len(reporte_ids),
    )
    return rows


def _fo_eo_04_norm_item_key(item_numero: str, capitulo: str = "") -> tuple:
    """Clave ítem+capítulo alineada con UI (1.02 vs 1.02.)."""
    it = (item_numero or "").strip()
    while it.endswith("."):
        it = it[:-1].strip()
    return (it, (capitulo or "").strip())


def _fo_eo_04_item_keys_match(a_item: str, a_cap: str, b_item: str, b_cap: str) -> bool:
    return _fo_eo_04_norm_item_key(a_item, a_cap) == _fo_eo_04_norm_item_key(b_item, b_cap)


def _fo_eo_04_es_acta_rpo(acta_row: Dict[str, Any]) -> bool:
    return (str(acta_row.get("tipo_grupo") or "").strip().upper()) == "RPO"


def _fo_eo_04_parse_entero_acta(val: object) -> Optional[int]:
    """numero_rpo / consecutivo: int, float 3.0, texto '3' (prod y local pueden diferir)."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        pass
    try:
        f = float(val)
        if f == int(f):
            return int(f)
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    if not s:
        return None
    import re

    m = re.search(r"\d+", s)
    if m:
        try:
            return int(m.group(0))
        except ValueError:
            pass
    return None


def _fo_eo_04_list_actas_contrato(contrato_id: int) -> List[Dict[str, Any]]:
    """Todas las actas del contrato (paginado; en prod Supabase limita ~1000 filas por request)."""
    cid = int(contrato_id)
    _PAGE = 1000
    rows: List[Dict[str, Any]] = []
    off = 0
    for _ in range(50):
        part = (
            _sb.table("actas")
            .select("id, numero_rpo, consecutivo, tipo_grupo")
            .eq("contrato_id", cid)
            .range(off, off + _PAGE - 1)
            .execute()
            .data
            or []
        )
        rows.extend(part)
        if len(part) < _PAGE:
            break
        off += _PAGE
    return rows


def _fo_eo_04_actas_anteriores_ids(contrato_id: int, acta_id: int) -> List[int]:
    """
    Actas RPO anteriores al acta actual.
    Orden: numero_rpo (como recibo parcial); si falta en actas viejas, consecutivo; último fallback id.
    """
    try:
        acta_row = (
            _sb.table("actas")
            .select("id, consecutivo, numero_rpo, tipo_grupo")
            .eq("id", int(acta_id))
            .single()
            .execute()
            .data
            or {}
        )
    except Exception:
        return []
    n_cur = _fo_eo_04_parse_entero_acta(acta_row.get("numero_rpo"))
    c_cur = _fo_eo_04_parse_entero_acta(acta_row.get("consecutivo"))
    try:
        todas = _fo_eo_04_list_actas_contrato(int(contrato_id))
    except Exception:
        return []
    rpo_actas = [a for a in todas if _fo_eo_04_es_acta_rpo(a)]
    prev: List[int] = []
    seen: set = set()

    def _add(aid: int) -> None:
        if aid not in seen:
            seen.add(aid)
            prev.append(aid)

    if n_cur is not None:
        for a in rpo_actas:
            aid = a.get("id")
            if aid is None:
                continue
            nr = _fo_eo_04_parse_entero_acta(a.get("numero_rpo"))
            if nr is not None and nr < n_cur:
                _add(int(aid))
            elif nr is None and c_cur is not None:
                ac = _fo_eo_04_parse_entero_acta(a.get("consecutivo"))
                if ac is not None and ac < c_cur:
                    _add(int(aid))
    if not prev and c_cur is not None:
        for a in rpo_actas:
            aid = a.get("id")
            if aid is None:
                continue
            ac = _fo_eo_04_parse_entero_acta(a.get("consecutivo"))
            if ac is not None and ac < c_cur:
                _add(int(aid))
    if not prev:
        try:
            aid_cur = int(acta_id)
            for a in rpo_actas:
                aid = a.get("id")
                if aid is not None and int(aid) < aid_cur:
                    _add(int(aid))
        except (TypeError, ValueError):
            pass
    _log.info(
        "fo_eo_04 actas_anteriores_ids: acta=%s num_rpo=%s consec=%s prev=%s (rpo_en_contrato=%s)",
        acta_id,
        acta_row.get("numero_rpo"),
        acta_row.get("consecutivo"),
        len(prev),
        len(rpo_actas),
    )
    if not prev and (n_cur is not None and n_cur > 1 or c_cur is not None and c_cur > 1):
        _log.warning(
            "fo_eo_04 actas_anteriores VACÍO con acta aparentemente no primera: acta=%s num_rpo=%s consec=%s rpo=%s",
            acta_id,
            acta_row.get("numero_rpo"),
            acta_row.get("consecutivo"),
            len(rpo_actas),
        )
    return prev


def _fo_eo_04_fetch_registros_actas_ids(
    contrato_id: int,
    acta_ids: List[int],
    *,
    select_cols: str = _FO_EO_04_SEL_REGISTROS,
    item_numeros: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """
    Líneas de varias actas: acta_rpo_id en el conjunto y registros vía so_reportes.acta_rpo_id
    (misma regla que _fo_eo_04_fetch_registros_acta; necesario para totales acumulados).
    item_numeros: si se pasa, solo trae esos ítems (mucho más rápido en prod).
    """
    if not acta_ids:
        return []
    cid = int(contrato_id)
    ids_u = sorted({int(x) for x in acta_ids})
    by_id: Dict[Any, Dict[str, Any]] = {}
    item_chunks: List[Optional[List[str]]] = [None]
    if item_numeros:
        inums = sorted({(x or "").strip() for x in item_numeros if (x or "").strip()})
        if inums:
            item_chunks = [inums[i : i + 40] for i in range(0, len(inums), 40)]

    def _base_q():
        return (
            _sb.table("so_registros")
            .select(select_cols)
            .eq("contrato_id", cid)
            .not_.is_("item_numero", "null")
            .neq("item_numero", "")
        )

    for i in range(0, len(ids_u), 80):
        chunk = ids_u[i : i + 80]
        chunk_set = set(chunk)

        for ic in item_chunks:
            rows_direct = _fo_eo_04_paginar_so_registros(
                lambda c=chunk, ic=ic: (
                    _base_q().in_("acta_rpo_id", c).in_("item_numero", ic)
                    if ic is not None
                    else _base_q().in_("acta_rpo_id", c)
                )
            )

            reporte_ids: List[int] = []
            try:
                rp = (
                    _sb.table("so_reportes")
                    .select("id")
                    .eq("contrato_id", cid)
                    .in_("acta_rpo_id", chunk)
                    .execute()
                    .data
                    or []
                )
                reporte_ids = [int(x["id"]) for x in rp if x.get("id") is not None]
            except Exception as exc:
                _log.warning("fo_eo_04 reportes actas_prev: %s", exc)

            rows_rep: list = []
            for j in range(0, len(reporte_ids), 80):
                rep_chunk = reporte_ids[j : j + 80]

                def _q_rep(ids=rep_chunk, ic=ic):
                    q = _base_q().in_("reporte_id", ids).is_("acta_rpo_id", "null")
                    if ic is not None:
                        q = q.in_("item_numero", ic)
                    return q

                rows_rep.extend(_fo_eo_04_paginar_so_registros(_q_rep))

            for r in rows_direct:
                rid = r.get("id")
                if rid is not None:
                    by_id[rid] = r
            for r in rows_rep:
                rid = r.get("id")
                if rid is None:
                    continue
                ara = r.get("acta_rpo_id")
                if ara is not None:
                    try:
                        if int(ara) not in chunk_set:
                            continue
                    except (TypeError, ValueError):
                        continue
                by_id[rid] = r

    rows = list(by_id.values())
    _log.info(
        "fo_eo_04 registros_actas_ids: actas=%s items_filtro=%s registros=%s",
        len(ids_u),
        len(item_numeros or []),
        len(rows),
    )
    return rows


def _fo_eo_04_registro_aprobado_nivel_max(
    reg: Dict[str, Any],
    contrato_id: int,
    *,
    matriz: Optional[Tuple[str, List[int]]] = None,
) -> bool:
    """
    FO-EO-04 / acumulados: solo «Aprobado» en el nivel máximo activo del contrato.
    No exige cascada N1·N2·… (eso es sellado / panel matriz; aquí no aplica).
    Misma regla que SICOE KPI dashboard (_registro_nivel_max_aprobado en main).
    """
    if not (str(reg.get("item_numero") or "").strip()):
        return False
    if matriz is None:
        matriz = matriz_params_contrato(_sb, int(contrato_id))
    campo_mx, _niveles_act = matriz
    return _norm_estado_n3(reg.get(campo_mx)) == "Aprobado"


def _fo_eo_04_registro_aprobado_interventoria(
    reg: Dict[str, Any],
    contrato_id: int,
    *,
    matriz: Optional[Tuple[str, List[int]]] = None,
) -> bool:
    """Alias histórico → aprobado en nivel máximo (no sellado en cascada)."""
    return _fo_eo_04_registro_aprobado_nivel_max(reg, contrato_id, matriz=matriz)


def _fo_eo_04_registros_sellados_acta(
    contrato_id: int,
    acta_id: int,
    *,
    matriz: Optional[Tuple[str, List[int]]] = None,
) -> List[Dict[str, Any]]:
    """Líneas del acta con nivel máximo en Aprobado (vínculo acta/reporte)."""
    if matriz is None:
        matriz = matriz_params_contrato(_sb, int(contrato_id))
    campo_mx, niveles_act = matriz
    raw = _fo_eo_04_fetch_registros_acta(int(contrato_id), int(acta_id))
    aprobados = [
        r
        for r in raw
        if _fo_eo_04_registro_aprobado_nivel_max(r, int(contrato_id), matriz=matriz)
    ]
    _log.info(
        "fo_eo_04 aprobados_nivel_max_acta: acta=%s campo_max=%s niveles=%s raw=%s aprobados=%s",
        acta_id,
        campo_mx,
        niveles_act,
        len(raw),
        len(aprobados),
    )
    return aprobados


def _fetch_total_actas_anteriores(
    contrato_id: int, acta_id: int, item_numero: str, capitulo: str = ""
) -> float:
    """
    Opción A: suma de cantidad_total N3-Aprobado para el capítulo+ítem en todas
    las actas del contrato con consecutivo < al del acta actual.
    Fallback: si consecutivo es NULL, compara por id < acta_id.
    """
    if not acta_id or not item_numero:
        return 0.0
    try:
        prev_ids = _fo_eo_04_actas_anteriores_ids(int(contrato_id), int(acta_id))
        if not prev_ids:
            return 0.0

        matriz = matriz_params_contrato(_sb, int(contrato_id))
        total = 0.0
        cap_strip = (capitulo or "").strip()
        itn = (item_numero or "").strip()

        key_tgt = _fo_eo_04_norm_item_key(itn, cap_strip)
        raw_prev = _fo_eo_04_fetch_registros_actas_ids(
            int(contrato_id), prev_ids, select_cols=_FO_EO_04_SEL_TOTALES
        )
        for r in raw_prev:
            if not _fo_eo_04_registro_aprobado_nivel_max(
                r, int(contrato_id), matriz=matriz
            ):
                continue
            if _fo_eo_04_norm_item_key(
                r.get("item_numero") or "", r.get("capitulo") or ""
            ) != key_tgt:
                continue
            total += float(r.get("cantidad_total") or 0)

        _log.info(
            "fo_eo_04 actas_anteriores: contrato=%s acta=%s item=%s cap=%s prev_actas=%s total=%.3f",
            contrato_id, acta_id, item_numero, cap_strip, len(prev_ids), total,
        )
        return total
    except Exception as exc:
        _log.warning("fetch_total_actas_anteriores ERROR item=%s: %s", item_numero, exc)
        return 0.0


def _fetch_totales_batch(contrato_id: int, acta_id: int, items: list) -> Tuple[dict, Dict[str, Any]]:
    """
    Totales de actas anteriores: cantidad_total con nivel máximo del contrato en Aprobado
    (sin exigir cascada / sellado en niveles inferiores).
    Returns: (dict {(item_numero, capitulo): total_float}, meta diagnóstico).
    """
    from collections import defaultdict
    matriz0 = matriz_params_contrato(_sb, int(contrato_id)) if acta_id and items else None
    campo_mx0 = matriz0[0] if matriz0 else ""
    meta: Dict[str, Any] = {
        "prev_actas_count": 0,
        "registros_prev_raw": 0,
        "registros_prev_nivel_max": 0,
        "items_con_acumulado": 0,
        "criterio_aprobacion": "nivel_max_aprobado",
        "campo_nivel_maximo": campo_mx0,
    }
    if not items or not acta_id:
        return {}, meta
    default_result: dict = {
        _fo_eo_04_norm_item_key(
            item.get("item_numero") or "", item.get("capitulo") or ""
        ): 0.0
        for item in items
    }
    try:
        prev_ids = _fo_eo_04_actas_anteriores_ids(int(contrato_id), int(acta_id))
        meta["prev_actas_count"] = len(prev_ids)
        if not prev_ids:
            _log.warning(
                "fo_eo_04 totales_batch sin actas previas: contrato=%s acta=%s items=%s",
                contrato_id,
                acta_id,
                len(items),
            )
            return default_result, meta

        matriz = matriz_params_contrato(_sb, int(contrato_id))
        items_set = {
            _fo_eo_04_norm_item_key(
                item.get("item_numero") or "", item.get("capitulo") or ""
            )
            for item in items
        }
        totales: dict = defaultdict(float)

        nums_pdf = [
            (item.get("item_numero") or "").strip()
            for item in items
            if (item.get("item_numero") or "").strip()
        ]
        raw_prev = _fo_eo_04_fetch_registros_actas_ids(
            int(contrato_id),
            prev_ids,
            select_cols=_FO_EO_04_SEL_TOTALES,
            item_numeros=nums_pdf,
        )
        meta["registros_prev_raw"] = len(raw_prev)
        n_nivel_max = 0
        for r in raw_prev:
            if not _fo_eo_04_registro_aprobado_nivel_max(
                r, int(contrato_id), matriz=matriz
            ):
                continue
            n_nivel_max += 1
            k = _fo_eo_04_norm_item_key(
                r.get("item_numero") or "", r.get("capitulo") or ""
            )
            if k in items_set:
                totales[k] += float(r.get("cantidad_total") or 0)
        meta["registros_prev_nivel_max"] = n_nivel_max

        meta["items_con_acumulado"] = sum(1 for v in totales.values() if v)
        _log.info(
            "fo_eo_04 totales_batch: acta=%s prev_actas=%s raw_prev=%s nivel_max=%s items=%s con_total=%s campo=%s",
            acta_id,
            len(prev_ids),
            len(raw_prev),
            n_nivel_max,
            len(items_set),
            meta["items_con_acumulado"],
            meta.get("campo_nivel_maximo"),
        )
        # Combinar: ítems con datos + ítems sin registros anteriores (→ 0.0)
        result = {**default_result, **dict(totales)}
        return result, meta
    except Exception as exc:
        _log.exception(
            "fetch_totales_batch ERROR contrato=%s acta=%s (prod suele ser timeout o límite Supabase): %s",
            contrato_id,
            acta_id,
            exc,
        )
        return default_result, meta


def _fetch_items_n3_acta(acta_id: int, contrato_id: int) -> list:
    """
    Ítems del acta con nivel máximo del contrato en Aprobado (FO-EO-04; no cascada sellado).
    """
    from collections import defaultdict
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if not acta_id:
        return []
    try:
        matriz = matriz_params_contrato(_sb, int(contrato_id))
        aprobados = _fo_eo_04_registros_sellados_acta(
            int(contrato_id), int(acta_id), matriz=matriz
        )
        if not aprobados:
            return []

        # 3. Agrupar por (item_numero, capitulo) × semana_id
        #    En algunas entidades externas puede repetirse item_numero en capítulos distintos.
        items_sem: dict = defaultdict(lambda: defaultdict(float))
        items_meta: dict = {}
        items_foto: dict = {}

        for r in aprobados:
            num = (r.get("item_numero") or "").strip()
            cap = (r.get("capitulo") or "").strip()
            if not num:
                continue
            key = (num, cap)
            sem_id = r.get("semana_id")
            items_sem[key][sem_id] += float(r.get("cantidad_total") or 0)
            if key not in items_meta:
                items_meta[key] = {
                    "item_numero": num,
                    "unidad": (r.get("unidad") or "").strip(),
                    "capitulo": cap,
                    "item_descripcion": (r.get("item_descripcion") or "").strip(),
                }
            if key not in items_foto:
                items_foto[key] = {"foto_url": None, "foto_numero": None,
                                   "grafico_url": None, "grafico_numero": None}
            fi = items_foto[key]
            if not fi["foto_url"] and str(r.get("foto_url") or "").strip():
                fi["foto_url"] = r["foto_url"]
                fi["foto_numero"] = r.get("foto_numero")
            if not fi["grafico_url"] and str(r.get("grafico_url") or "").strip():
                fi["grafico_url"] = r["grafico_url"]
                fi["grafico_numero"] = r.get("grafico_numero")
        _log.info("fo_eo_04 fetch_items_n3: items_unicos=%s", len(items_meta))

        # 4+5. so_semanas y listado_precios EN PARALELO (son independientes)
        all_sem_ids = list({sid for sems in items_sem.values() for sid in sems if sid})
        sem_info: dict = {}
        esp_map: dict = {}

        def _fetch_semanas():
            if not all_sem_ids:
                return {}
            try:
                s_rows = (
                    _sb.table("so_semanas")
                    .select("id, numero_semana, fecha_inicio, fecha_fin")
                    .in_("id", all_sem_ids)
                    .execute()
                    .data or []
                )
                return {s["id"]: s for s in s_rows}
            except Exception as exc:
                _log.warning("fetch_items_n3: semanas: %s", exc)
                return {}

        def _fetch_esp():
            nums = list({k[0] for k in items_meta if k[0]})
            if not nums:
                return {}
            esp: dict = {}
            try:
                for i in range(0, len(nums), 100):
                    part = nums[i : i + 100]
                    lp = (
                        _sb.table("listado_precios")
                        .select("item_numero, especificacion_tecnica")
                        .eq("contrato_id", int(contrato_id))
                        .in_("item_numero", part)
                        .execute()
                        .data
                        or []
                    )
                    for r in lp:
                        n = (r.get("item_numero") or "").strip()
                        if n and n not in esp:
                            esp[n] = (r.get("especificacion_tecnica") or "")
            except Exception as exc:
                _log.warning("fetch_items_n3: listado_precios: %s", exc)
            return esp

        with ThreadPoolExecutor(max_workers=2) as pool:
            f_sem = pool.submit(_fetch_semanas)
            f_esp = pool.submit(_fetch_esp)
            try:
                sem_info = f_sem.result()
            except Exception as exc:
                _log.warning("fetch_items_n3: semanas pool: %s", exc)
                sem_info = {}
            try:
                esp_map = f_esp.result()
            except Exception as exc:
                _log.warning("fetch_items_n3: esp pool: %s", exc)
                esp_map = {}

        # 6. Construir lista final con semanas ordenadas por numero_semana
        result: list = []
        for key, meta in items_meta.items():
            num = meta.get("item_numero") or ""
            sem_entries: list = []
            total_acta = 0.0
            for sem_id, cantidad in items_sem[key].items():
                info = sem_info.get(sem_id, {}) if sem_id else {}
                sem_entries.append({
                    "numero_semana": info.get("numero_semana"),
                    "fecha_inicio": (info.get("fecha_inicio") or "")[:10],
                    "fecha_fin": (info.get("fecha_fin") or "")[:10],
                    "cantidad": cantidad,
                })
                total_acta += cantidad
            sem_entries.sort(key=lambda s: (s.get("numero_semana") or 0))
            meta["semanas"] = sem_entries
            meta["total_acta"] = total_acta
            meta["especificacion_tecnica"] = esp_map.get(num, "")
            # Fotos de registros sellados en el nivel máximo del contrato
            fi = items_foto.get(key, {})
            meta["foto_url"]       = fi.get("foto_url")
            meta["foto_numero"]    = fi.get("foto_numero")
            meta["grafico_url"]    = fi.get("grafico_url")
            meta["grafico_numero"] = fi.get("grafico_numero")
            _log.info("fo_eo_04 item=%s foto=#%s grf=#%s", num, meta["foto_numero"], meta["grafico_numero"])
            result.append(meta)

        return sorted(
            result,
            key=lambda r: (
                _orden_titulo_capitulo_obra(r.get("capitulo") or ""),
                _orden_item_numero(r.get("item_numero") or ""),
            ),
        )
    except Exception as exc:
        _log.warning("fetch_items_n3_acta ERROR: %s", exc)
        return []


def _combine_html_pages(htmls: list) -> str:
    """Une múltiples páginas HTML en un solo documento con saltos de página entre ellas."""
    if not htmls:
        return ""
    if len(htmls) == 1:
        return htmls[0]
    # Quita el cierre del primer documento
    base = htmls[0].rstrip()
    if base.endswith("</body></html>"):
        base = base[: -len("</body></html>")].rstrip()
    parts = [base]
    for page_html in htmls[1:]:
        # Extrae solo el contenido del body
        body_start = page_html.find("<body>")
        body_end = page_html.rfind("</body>")
        if body_start != -1 and body_end != -1:
            body_content = page_html[body_start + len("<body>") : body_end].strip()
        else:
            body_content = page_html
        parts.append(
            '\n<div style="page-break-before:always;height:0;margin:0;padding:0;"></div>\n'
            + body_content
        )
    parts.append("\n</body></html>")
    return "".join(parts)


def _render_items_n3_html(
    items: list,
    navy: str,
    navy_hdr: str,
    bg_hdr: str,
    bd: str,
    z: str,
) -> str:
    """
    Genera HTML para la sección de ítems N3 aprobados del acta.

    Sin ítems → plantilla estática (ÍTEM No / UNIDAD / ESPECIFICACIÓN / CAPÍTULO / DESCRIPCIÓN)
                con 13 filas vacías.
    Con ítems → bloque por ítem (encabezado CAPÍTULO + info ítem + tabla de medidas de 5 filas),
                ordenados por capítulo y número de ítem.
    """
    bd_top_none = f"border:1px solid {navy};border-top:none"
    col_labels = ["UBICACI\u00d3N", "LARGO", "ANCHO", "ALTO", "CANTIDAD", "TOTAL"]
    col_widths = ["40%", "13%", "13%", "13%", "12%", "9%"]
    tot_labels = [
        "TOTAL EJECUTADO PRESENTE ACTA",
        "TOTAL EJECUTADO ACTAS ANTERIORES",
        "TOTAL EJECUTADO ACUMULADO",
        "TOTAL POR EJECUTAR",
    ]

    def _meas_header() -> str:
        cells = "".join(
            f'<td style="{bd};border-top:none;padding:3px 4px;font-size:6.5pt;'
            f'font-weight:bold;text-align:center;color:{navy_hdr};width:{w};">{lbl}</td>'
            for lbl, w in zip(col_labels, col_widths)
        )
        return (
            f'<table width="100%" cellspacing="0" cellpadding="0" '
            f'style="border-collapse:collapse;{bd_top_none};">'
            f'<tr style="background-color:{bg_hdr};">{cells}</tr>'
        )

    def _meas_row_empty() -> str:
        return (
            f'<tr style="height:15px;">'
            f'<td style="{bd};padding:2px 4px;font-size:6pt;color:{navy};">{z}</td>'
            f'<td style="{bd};padding:2px 3px;font-size:6pt;text-align:center;color:{navy};">{z}</td>'
            f'<td style="{bd};padding:2px 3px;font-size:6pt;text-align:center;color:{navy};">{z}</td>'
            f'<td style="{bd};padding:2px 3px;font-size:6pt;text-align:center;color:{navy};">{z}</td>'
            f'<td style="{bd};padding:2px 3px;font-size:6pt;text-align:center;color:{navy};">{z}</td>'
            f'<td style="{bd};padding:2px 3px;font-size:6pt;text-align:right;color:{navy};">{z}</td>'
            f'</tr>'
        )

    def _tot_rows() -> str:
        return "".join(
            f'<tr><td colspan="4" style="{bd};padding:3px 5px;font-size:6.5pt;text-align:right;'
            f'font-weight:bold;color:{navy_hdr};background-color:{bg_hdr};">{_h(lbl)}</td>'
            f'<td colspan="2" style="{bd};padding:3px 5px;font-size:6.5pt;text-align:right;color:{navy};">{z}</td></tr>'
            for lbl in tot_labels
        )

    def _meas_table(n_filas: int) -> str:
        return _meas_header() + _meas_row_empty() * n_filas + _tot_rows() + "</table>"

    # ── Plantilla vacía (sin ítems) ──────────────────────────────────────────
    if not items:
        return (
            f'<table width="100%" cellspacing="0" cellpadding="0" '
            f'style="border-collapse:collapse;{bd_top_none};">'
            # Fila 1: ÍTEM No | UNIDAD | ESPECIFICACIÓN TÉCNICA
            f'<tr>'
            f'<td style="border-right:1px solid {navy};width:20%;padding:4px 7px;vertical-align:top;">'
            f'<div style="font-weight:bold;font-size:6pt;color:{navy_hdr};">&#205;TEM No</div>'
            f'<div style="min-height:14px;border-bottom:1px solid {navy};">&nbsp;</div>'
            f'</td>'
            f'<td style="border-right:1px solid {navy};width:12%;padding:4px 7px;vertical-align:top;">'
            f'<div style="font-weight:bold;font-size:6pt;color:{navy_hdr};">UNIDAD</div>'
            f'<div style="min-height:14px;border-bottom:1px solid {navy};">&nbsp;</div>'
            f'</td>'
            f'<td style="width:68%;padding:4px 7px;vertical-align:top;">'
            f'<div style="font-weight:bold;font-size:6pt;color:{navy_hdr};">ESPECIFICACI&#211;N T&#201;CNICA</div>'
            f'<div style="min-height:14px;border-bottom:1px solid {navy};">&nbsp;</div>'
            f'</td>'
            f'</tr>'
            # Fila 2: CAPÍTULO (debajo de ESPECIFICACIÓN TÉCNICA, sin línea separadora)
            f'<tr>'
            f'<td colspan="3" style="border-top:1px solid {navy};padding:4px 7px;">'
            f'<div style="font-weight:bold;font-size:6pt;color:{navy_hdr};">CAP&#205;TULO</div>'
            f'<div style="min-height:14px;border-bottom:1px solid {navy};">&nbsp;</div>'
            f'</td>'
            f'</tr>'
            # Fila 3: DESCRIPCIÓN (sin línea separadora desde CAPÍTULO)
            f'<tr>'
            f'<td colspan="3" style="border-top:1px solid {navy};padding:4px 7px;">'
            f'<div style="font-weight:bold;font-size:6pt;color:{navy_hdr};">DESCRIPCI&#211;N DEL &#205;TEM</div>'
            f'<div style="min-height:15px;">&nbsp;</div>'
            f'</td>'
            f'</tr>'
            f'</table>'
            + _meas_table(13)
        )

    # ── Ítems reales: un bloque por ítem, agrupados por capítulo ─────────────
    parts: list = []
    current_cap: Optional[str] = None

    for item in items:
        cap = (item.get("capitulo") or "Sin cap\u00edtulo").strip()
        item_num = _h(item.get("item_numero") or "")
        unidad = _h(item.get("unidad") or "")
        descripcion = _h(item.get("item_descripcion") or "")

        # Encabezado de capítulo cuando cambia
        if cap != current_cap:
            current_cap = cap
            parts.append(
                f'<table width="100%" cellspacing="0" cellpadding="0" '
                f'style="border-collapse:collapse;{bd_top_none};">'
                f'<tr style="background-color:{bg_hdr};">'
                f'<td style="padding:3px 8px;font-weight:bold;font-size:7pt;color:{navy_hdr};">'
                f'CAP&#205;TULO: {_h(cap)}'
                f'</td></tr></table>'
            )

        # Info del ítem: ÍTEM No | UNIDAD | DESCRIPCIÓN
        parts.append(
            f'<table width="100%" cellspacing="0" cellpadding="0" '
            f'style="border-collapse:collapse;{bd_top_none};">'
            f'<tr>'
            f'<td style="border-right:1px solid {navy};width:20%;padding:4px 7px;vertical-align:top;">'
            f'<div style="font-weight:bold;font-size:6pt;color:{navy_hdr};">&#205;TEM No</div>'
            f'<div style="font-size:8pt;font-weight:bold;color:{navy_hdr};">{item_num}</div>'
            f'</td>'
            f'<td style="border-right:1px solid {navy};width:12%;padding:4px 7px;vertical-align:top;">'
            f'<div style="font-weight:bold;font-size:6pt;color:{navy_hdr};">UNIDAD</div>'
            f'<div style="font-size:8pt;color:{navy_hdr};">{unidad}</div>'
            f'</td>'
            f'<td style="width:68%;padding:4px 7px;vertical-align:top;">'
            f'<div style="font-weight:bold;font-size:6pt;color:{navy_hdr};">DESCRIPCI&#211;N DEL &#205;TEM</div>'
            f'<div style="font-size:7pt;color:{navy_hdr};">{descripcion}</div>'
            f'</td>'
            f'</tr>'
            f'</table>'
        )

        # Tabla de medidas (5 filas vacías por ítem)
        parts.append(_meas_table(5))

    return "".join(parts)


def _fo_eo_04_firmas_data_uris(
    contrato_id: int,
    formato_codigo: str,
    acta_id: Optional[int],
    firma_cfg: Dict[str, Any],
    current_user: Optional[dict],
) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    """Imágenes de firma (ccd_firma_registro + acta_rpo) para los 4 slots del FO-EO-04."""
    if not acta_id:
        return None, None, None, None
    ct, cid = "acta_rpo", int(acta_id)
    fc = firma_cfg or {}

    def _slot(slot: str, uid_key: str, nombre_key: str) -> Optional[str]:
        return _firma_data_uri_para_slot_contexto(
            contrato_id,
            ct,
            cid,
            formato_codigo,
            slot,
            _opt_usuario_id(fc.get(uid_key)),
            str(fc.get(nombre_key) or "").strip(),
            current_user,
        )

    return (
        _slot("elaboro", "elaboro_usuario_id", "elaboro_nombre"),
        _slot("elaboro2", "elaboro2_usuario_id", "elaboro2_nombre"),
        _slot("reviso", "reviso_usuario_id", "reviso_nombre"),
        _slot("reviso2", "reviso2_usuario_id", "reviso2_nombre"),
    )


def _fo_eo_04_marca_detalle_por_slot(
    contrato_id: int,
    formato_codigo: str,
    acta_id: Optional[int],
    slot: str,
) -> Tuple[str, str, str]:
    """Nombre del usuario que registró la firma + fecha y hora (Bogotá). Vacíos si no hay registro."""
    if not acta_id:
        return "", "", ""
    reg = _ccd_firma_registro_contexto_get(
        contrato_id, "acta_rpo", int(acta_id), formato_codigo, slot
    )
    if not reg:
        return "", "", ""
    nom = _ccd_usuario_nombre_completo_por_id(reg.get("usuario_id"))
    mf, mh = _fecha_hora_marca_bo_desde_created_at(reg.get("created_at"))
    return nom, mf, mh


def _fo_eo_04_firmas_marcas_registro(
    contrato_id: int,
    formato_codigo: str,
    acta_id: Optional[int],
) -> Tuple[
    Tuple[str, str, str],
    Tuple[str, str, str],
    Tuple[str, str, str],
    Tuple[str, str, str],
]:
    """Cuatro ternas (nombre firmante, fecha, hora) por slot — solo si existe fila en ccd_firma_registro."""
    if not acta_id:
        return ("", "", ""), ("", "", ""), ("", "", ""), ("", "", "")
    return (
        _fo_eo_04_marca_detalle_por_slot(contrato_id, formato_codigo, acta_id, "elaboro"),
        _fo_eo_04_marca_detalle_por_slot(contrato_id, formato_codigo, acta_id, "elaboro2"),
        _fo_eo_04_marca_detalle_por_slot(contrato_id, formato_codigo, acta_id, "reviso"),
        _fo_eo_04_marca_detalle_por_slot(contrato_id, formato_codigo, acta_id, "reviso2"),
    )


# FO-IDU-EO-04-V2: xhtml2pdf ignora a menudo overflow en <div>; misma técnica que CC-SUB-001 (_html_cc_sub_td_firma_columna).
# Alturas fijas: caja firma + celda verificación alineada; reducción acumulada vs 63/53 + compactación tipográfica en marca.
_FO_EO04_FIRMA_BOX = "45px"
_FO_EO04_FIRMA_IMG = "38px"
# Descarga paralela de foto + gráfico por ítem (alineado con informe gerencia matriz).
_FO_EO04_ITEM_IMG_WORKERS = 16
_FO_EO04_IMG_MAX_PX = 720
_FO_EO04_MAX_DATA_URI_LEN = 400_000
_FO_EO04_PDF_PARALLEL_MIN_PAGES = 2
_FO_EO04_PDF_PAGES_PER_TASK = max(
    1,
    int(_os.environ.get("FO_EO04_PDF_PAGES_PER_TASK", "2") or "2"),
)
_FO_EO04_PDF_TASK_TIMEOUT_SEC = max(
    45,
    int(_os.environ.get("FO_EO04_PDF_TASK_TIMEOUT_SEC", "90") or "90"),
)
_FO_EO04_IMG_FETCH_TIMEOUT_SEC = max(
    5,
    int(_os.environ.get("FO_EO04_IMG_FETCH_TIMEOUT_SEC", "12") or "12"),
)
# Auto-test del pool al crearlo: si el hijo no responde en este tiempo, se asume roto
# (p. ej. /dev/shm minúsculo o sin permiso de fork en el host) y se cae a secuencial.
_FO_EO04_POOL_PING_TIMEOUT_SEC = max(
    8,
    int(_os.environ.get("FO_EO04_PDF_POOL_PING_TIMEOUT_SEC", "25") or "25"),
)


def _fo_eo_04_pdf_page_workers() -> int:
    """Nº de procesos para renderizar páginas en paralelo (meta ≈ 2 mem/s).

    Configurable con FO_EO04_PDF_PAGE_WORKERS. IMPORTANTE: en Azure App Service
    `os.cpu_count()` reporta TODOS los núcleos del HOST (no los del plan), así que
    dimensionar el pool con cpu_count creaba demasiados procesos → OOM → cuelgue
    silencioso. Por eso el default en Azure es conservador (3) y NO depende de cpu_count.
    """
    raw = (_os.environ.get("FO_EO04_PDF_PAGE_WORKERS") or "").strip()
    if raw:
        try:
            n = int(raw)
            if n >= 1:
                return min(n, 12)
        except ValueError:
            pass
    if _os.environ.get("WEBSITE_SITE_NAME") or _os.environ.get("WEBSITE_INSTANCE_ID"):
        return 3
    return max(3, min(8, (_os.cpu_count() or 4)))


def _fo_eo_04_pdf_use_processes() -> bool:
    """
    Renderizar el PDF en procesos hijos mantiene vivo el event loop del worker:
    el trabajo CPU-bound de xhtml2pdf no bloquea uvicorn (el hilo espera el futuro y
    suelta el GIL) ⇒ gunicorn --timeout no mata al worker, y además paraleliza.

    Antes se desactivaba en Azure por cuelgues, pero esos cuelgues eran por exceso de
    procesos (RAM, ver _fo_eo_04_pdf_page_workers). Ahora el nº de procesos es
    conservador/configurable y hay auto-test con fallback. FO_EO04_PDF_USE_PROCESSES=0
    fuerza el modo secuencial (hilo principal) si hiciera falta.
    """
    env = (_os.environ.get("FO_EO04_PDF_USE_PROCESSES") or "").strip().lower()
    if env in ("0", "false", "no"):
        return False
    return True
_FO_EO04_PROCESS_POOL = None
_FO_EO04_POOL_LOCK = threading.Lock()
_FO_EO04_POOL_BROKEN = False
_FO_EO04_MARCA_LBL_ELABORO = "Elaborado y firmado por:"
_FO_EO04_MARCA_LBL_REVISO = "Revisado y Aprobado por:"


def _html_fo_eo_04_firmas_4(
    navy: str,
    navy_hdr: str,
    bd: str,
    elaboro_nombre: str,
    elaboro_cargo: str,
    elaboro2_nombre: str,
    elaboro2_cargo: str,
    reviso_nombre: str,
    reviso_cargo: str,
    reviso2_nombre: str,
    reviso2_cargo: str,
    elaboro_firma_data_uri: Optional[str] = None,
    elaboro2_firma_data_uri: Optional[str] = None,
    reviso_firma_data_uri: Optional[str] = None,
    reviso2_firma_data_uri: Optional[str] = None,
    elaboro_marca_fecha: str = "",
    elaboro_marca_hora: str = "",
    elaboro2_marca_fecha: str = "",
    elaboro2_marca_hora: str = "",
    reviso_marca_fecha: str = "",
    reviso_marca_hora: str = "",
    reviso2_marca_fecha: str = "",
    reviso2_marca_hora: str = "",
    elaboro_marca_usuario: str = "",
    elaboro2_marca_usuario: str = "",
    reviso_marca_usuario: str = "",
    reviso2_marca_usuario: str = "",
) -> str:
    """Bloque de firmas: col izquierda = Elaboró (1 arriba, 2 abajo); col derecha = Revisó (1 arriba, 2 abajo)."""
    import html as _html_mod

    def _fila_firma(
        nombre: str,
        cargo: str,
        borde_top: bool = False,
        firma_uri: Optional[str] = None,
        marca_fecha: str = "",
        marca_hora: str = "",
        marca_usuario: str = "",
        marca_es_elaboro: bool = True,
    ) -> str:
        nombre_h = _html_mod.escape(nombre) if nombre else "&nbsp;"
        cargo_h  = _html_mod.escape(cargo) if cargo else "&nbsp;"
        bt = f"border-top:1px solid {navy};" if borde_top else ""
        bp = _FO_EO04_FIRMA_BOX
        ip = _FO_EO04_FIRMA_IMG
        if firma_uri:
            src_h = _html_mod.escape(firma_uri, quote=True)
            sig_td = (
                f'<img src="{src_h}" alt="" style="display:block;margin:0 auto;max-width:100%;width:auto;'
                f'height:{ip};max-height:{ip};border:0;padding:0;"/>'
            )
        else:
            sig_td = "&nbsp;"
        fd_h = _html_mod.escape(marca_fecha.strip()) if (marca_fecha or "").strip() else ""
        hr_h = _html_mod.escape(marca_hora.strip()) if (marca_hora or "").strip() else ""
        usr_h = _html_mod.escape(marca_usuario.strip()) if (marca_usuario or "").strip() else ""
        if fd_h and hr_h:
            fh_line = f"{fd_h} &#8211; {hr_h}"
        elif fd_h:
            fh_line = fd_h
        elif hr_h:
            fh_line = hr_h
        else:
            fh_line = ""

        tiene_marca = bool(usr_h or fh_line)
        lbl = _FO_EO04_MARCA_LBL_ELABORO if marca_es_elaboro else _FO_EO04_MARCA_LBL_REVISO
        if tiene_marca:
            marca_block = (
                f'<div style="font-size:6pt;font-weight:bold;color:{navy_hdr};'
                f'line-height:1.05;margin:0;padding:0;text-align:center;">{_html_mod.escape(lbl)}</div>'
                f'<div style="font-size:6.5pt;font-weight:bold;color:{navy};line-height:1.05;'
                f'margin:0;padding:0;text-align:center;">{usr_h if usr_h else "&#8212;"}</div>'
                f'<div style="font-size:6.5pt;color:#64748b;line-height:1.05;text-align:center;margin:0;padding:0;">'
                f'{fh_line if fh_line else "&#8212;"}</div>'
            )
        else:
            marca_block = (
                '<div style="font-size:6pt;color:#cbd5e1;line-height:1.05;margin:0;padding:0;">&nbsp;</div>'
                '<div style="font-size:6pt;color:#cbd5e1;line-height:1.05;margin:0;padding:0;">&nbsp;</div>'
                '<div style="font-size:6pt;color:#cbd5e1;line-height:1.05;margin:0;padding:0;">&nbsp;</div>'
            )
        marca_cell_inner = marca_block
        firma_row = f"""<table cellspacing="0" cellpadding="0" width="100%" style="border-collapse:collapse;table-layout:fixed;margin:0 0 3px 0;">
<tr>
<td style="width:76%;vertical-align:middle;padding:0;border:none;">
<table cellspacing="0" cellpadding="0" width="100%" style="border-collapse:collapse;table-layout:fixed;">
<tr><td style="height:{bp};max-height:{bp};min-height:{bp};overflow:hidden;vertical-align:middle;text-align:center;line-height:0;font-size:0;padding:0;border:none;border-bottom:1px solid {navy};">
{sig_td}
</td></tr>
</table>
</td>
<td style="width:24%;height:{bp};max-height:{bp};min-height:{bp};overflow:hidden;vertical-align:middle;text-align:center;border-bottom:1px solid {navy};padding:1px 2px;line-height:1;">
{marca_cell_inner}
</td>
</tr>
</table>"""
        return f"""<tr>
<td style="vertical-align:top;{bt}padding:5px 8px;">
  {firma_row}
  <div style="font-size:8.5pt;font-weight:bold;color:{navy};text-align:center;border-bottom:1px solid {navy};padding:1px 0;margin-bottom:1px;line-height:1.15;">{nombre_h}</div>
  <div style="font-size:7.5pt;color:#475569;text-align:center;line-height:1.15;">{cargo_h}</div>
</td>
</tr>"""

    def _columna(
        titulo: str,
        n1: str,
        c1: str,
        n2: str,
        c2: str,
        uri1: Optional[str] = None,
        uri2: Optional[str] = None,
        m1_f: str = "",
        m1_h: str = "",
        m1_u: str = "",
        m2_f: str = "",
        m2_h: str = "",
        m2_u: str = "",
        columna_es_elaboro: bool = True,
    ) -> str:
        return f"""<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
<tr>
<td style="font-weight:bold;text-align:center;font-size:8.5pt;padding:3px 5px;
           background-color:#dce8f0;color:{navy_hdr};
           border-bottom:1px solid {navy};">{titulo}</td>
</tr>
{_fila_firma(n1, c1, borde_top=False, firma_uri=uri1, marca_fecha=m1_f, marca_hora=m1_h, marca_usuario=m1_u, marca_es_elaboro=columna_es_elaboro)}
{_fila_firma(n2, c2, borde_top=True, firma_uri=uri2, marca_fecha=m2_f, marca_hora=m2_h, marca_usuario=m2_u, marca_es_elaboro=columna_es_elaboro)}
</table>"""

    elab_col  = _columna(
        "ELABOR\u00d3",
        elaboro_nombre,
        elaboro_cargo,
        elaboro2_nombre,
        elaboro2_cargo,
        elaboro_firma_data_uri,
        elaboro2_firma_data_uri,
        elaboro_marca_fecha,
        elaboro_marca_hora,
        elaboro_marca_usuario,
        elaboro2_marca_fecha,
        elaboro2_marca_hora,
        elaboro2_marca_usuario,
        columna_es_elaboro=True,
    )
    revis_col = _columna(
        "REVIS\u00d3",
        reviso_nombre,
        reviso_cargo,
        reviso2_nombre,
        reviso2_cargo,
        reviso_firma_data_uri,
        reviso2_firma_data_uri,
        reviso_marca_fecha,
        reviso_marca_hora,
        reviso_marca_usuario,
        reviso2_marca_fecha,
        reviso2_marca_hora,
        reviso2_marca_usuario,
        columna_es_elaboro=False,
    )

    return f"""<table width="100%" cellspacing="0" cellpadding="0"
       style="border-collapse:collapse;margin-top:5px;{bd};">
<tr>
<td width="50%" style="vertical-align:top;{bd};padding:0;">
{elab_col}
</td>
<td width="50%" style="vertical-align:top;{bd};padding:0;border-left:none;">
{revis_col}
</td>
</tr>
</table>"""


_FO_EO_04_PREVIEW_HEAD_EXTRA = """
<style>
.cc-rotate-btn { font-family: system-ui, sans-serif; }
.cc-rotate-btn:hover { background: #eef2ff !important; }
.cc-rotate-btn:disabled { opacity: 0.55; cursor: wait; }
@media print { .cc-rotate-btn { display: none !important; } }
</style>
"""


def _fo_eo_04_wrap_preview_html(html: str) -> str:
    if not html:
        return html
    if "<head>" in html:
        return html.replace("<head>", "<head>" + _FO_EO_04_PREVIEW_HEAD_EXTRA, 1)
    return _FO_EO_04_PREVIEW_HEAD_EXTRA + html


def _fo_eo_04_html_embed_fragment(html: str) -> str:
    """Fragmento sin documento HTML anidado (mejor para embeber en React)."""
    if not html:
        return html
    bs = html.lower().find("<body")
    be = html.lower().rfind("</body>")
    if bs != -1 and be != -1:
        start = html.find(">", bs)
        if start != -1:
            return html[start + 1 : be].strip()
    return html


def _fo_eo_04_img_block_html(
    url: Optional[str],
    numero: Optional[int],
    tipo: str,
    *,
    preview_ui: bool,
    empty_label: str,
    max_height: str = "228px",
    mostrar_rotar: bool = True,
) -> str:
    if not url or not str(url).strip():
        return f'<div style="font-size:6pt;color:#94a3b8;padding-top:100px;">{empty_label}</div>'
    src = html.escape(str(url).strip(), quote=True)
    # En preview HTML (navegador) cargamos las imágenes en diferido para no saturar
    # la red con 100+ fotos a la vez; xhtml2pdf ignora estos atributos en el PDF.
    extra_attr = ' loading="lazy" decoding="async"' if preview_ui else ""
    img = (
        f'<img data-cc-img="1" src="{src}" alt=""{extra_attr} '
        f'style="max-width:98%;max-height:{max_height};display:block;margin:0 auto;object-fit:contain;" />'
    )
    if not preview_ui or numero is None or not mostrar_rotar:
        return img
    num = int(numero)
    tipo_esc = html.escape(tipo, quote=True)
    btn = (
        f'<button type="button" class="cc-rotate-btn" data-cc-rotate="1" '
        f'data-cc-tipo="{tipo_esc}" data-cc-num="{num}" '
        f'style="position:absolute;top:6px;right:6px;z-index:5;padding:6px 10px;'
        f'font-size:11px;font-weight:700;border:1px solid #6366f1;border-radius:8px;'
        f'background:rgba(255,255,255,0.92);cursor:pointer;box-shadow:0 1px 4px rgba(0,0,0,.15);">'
        f'&#8635; Girar</button>'
    )
    return (
        f'<div class="cc-img-slot" data-cc-slot="1" data-cc-tipo="{tipo_esc}" data-cc-num="{num}" '
        f'style="position:relative;height:100%;min-height:200px;display:flex;align-items:center;'
        f'justify-content:center;">{img}{btn}</div>'
    )


def _html_idu_fo_eo_04_v2_plantilla_vacia(
    logo_entidad: Optional[str] = None,
    logo_entidad_html: Optional[str] = None,
    entidad: str = "IDU",
    subsistema: str = "vial",
    num_contrato: str = "",
    año_contrato: str = "",
    objeto_contrato: str = "",
    num_acta: str = "",
    fecha_desde: str = "",
    fecha_hasta: str = "",
    # Partes firmantes (del contrato / admin)
    contratista: str = "",
    interventoria: str = "",
    supervisor: str = "",
    # Elaboró / Revisó (Biblioteca CCD – 2 por sección)
    elaboro_nombre: str = "",
    elaboro_cargo: str = "",
    elaboro2_nombre: str = "",
    elaboro2_cargo: str = "",
    reviso_nombre: str = "",
    reviso_cargo: str = "",
    reviso2_nombre: str = "",
    reviso2_cargo: str = "",
    # Firmas registradas (ccd_firma_registro) o perfil — data URI para el PDF
    elaboro_firma_data_uri: Optional[str] = None,
    elaboro2_firma_data_uri: Optional[str] = None,
    reviso_firma_data_uri: Optional[str] = None,
    reviso2_firma_data_uri: Optional[str] = None,
    # Marcas fecha/hora (registro en ccd_firma_registro) a la derecha de cada firma
    elaboro_marca_fecha: str = "",
    elaboro_marca_hora: str = "",
    elaboro2_marca_fecha: str = "",
    elaboro2_marca_hora: str = "",
    reviso_marca_fecha: str = "",
    reviso_marca_hora: str = "",
    reviso2_marca_fecha: str = "",
    reviso2_marca_hora: str = "",
    elaboro_marca_usuario: str = "",
    elaboro2_marca_usuario: str = "",
    reviso_marca_usuario: str = "",
    reviso2_marca_usuario: str = "",
    # Datos específicos del ítem (vacíos = plantilla en blanco)
    item_numero: str = "",
    item_unidad: str = "",
    item_especificacion: str = "",
    item_capitulo: str = "",
    item_descripcion: str = "",
    # Filas semanales: lista de {numero_semana, fecha_inicio, fecha_fin, cantidad}
    semanas_item: Optional[list] = None,
    # Totales de actas anteriores (Opción A: actas con consecutivo < actual)
    total_actas_anteriores: Optional[float] = None,
    foto_url: Optional[str] = None,
    foto_numero: Optional[int] = None,
    grafico_url: Optional[str] = None,
    grafico_numero: Optional[int] = None,
    preview_ui: bool = False,
    mostrar_rotar: bool = True,
) -> str:
    """Plantilla FO-EO-04 V2.0: vista previa sin datos de obra.

    Estructura y proporciones extraídas del SVG original (A3 portrait, 841.96×1201.02 pt).
    Colores: navy=#102a42 (texto), navy_hdr=#192739 (encabezados), bg_hdr=#dce8f0 (fondo filas cabecera).
    Columnas: UBICACIÓN 40% · LARGO 13% · ANCHO 13% · ALTO 13% · CANTIDAD 12% · TOTAL 9%.
    Filas de datos: 13 (coincide con el formulario original IDU).

    entidad:   'IDU' | 'ICCU' | 'ENEL' | 'EAB' | cualquier nombre — determina la sección institucional.
    subsistema: 'vial' | 'transporte' — solo aplica cuando entidad == 'IDU'.
    """
    # Paleta de colores extraída del SVG original
    navy = "#102a42"        # texto principal de filas
    navy_hdr = "#192739"    # texto de encabezados y etiquetas clave
    bg_hdr = "#dce8f0"      # fondo suave azul-marino para filas de cabecera (≈ rgba(28,55,90,0.16) sobre blanco)
    bd = f"border:1px solid {navy}"
    z = "\u2014"  # em dash como valor vacío

    obs = (
        "Las dimensiones de la presente cantidad se soportan en los archivos que surgen de las respectivas "
        "conciliaciones semanales, realizadas entre contratista e interventoría. Toda esta información se adjunta "
        "en la entrega mensual para la verificación final del acta de recibo parcial de obra."
    )

    # Logo precargado en data-URI (evita N descargas HTTP al renderizar cada memoria)
    if not logo_entidad_html:
        logo_entidad_html = _html_logo_entidad(logo_entidad)

    # ── Sección institucional debajo del encabezado ──────────────────────────
    _lbl_inst = (
        f'font-size:6.5pt;font-weight:bold;color:{navy_hdr};text-transform:uppercase;'
    )
    entidad_upper = (entidad or "IDU").strip().upper()

    if entidad_upper == "IDU":
        _sub_txt = (
            "DE EJECUCI&#211;N DEL SUBSISTEMA VIAL"
            if subsistema == "vial"
            else "DE EJECUCI&#211;N DEL SUBSISTEMA DE TRANSPORTE"
        )
        seccion_institucional = (
            # TD exterior: borde izq/der/top pero SIN borde inferior (el siguiente bloque lo provee)
            f'<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">'
            f'<tr><td style="border:1px solid {navy};border-top:none;border-bottom:none;padding:0;">'
            # Fila 1: Subdirección General (div, sin borde)
            f'<div style="text-align:center;padding:3px 8px;{_lbl_inst}">'
            f'SUBDIRECCI&#211;N GENERAL DE INFRAESTRUCTURA'
            f'</div>'
            # Fila 2: Dirección + Subdirección (tabla interna sin bordes)
            f'<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">'
            f'<tr>'
            f'<td style="width:38%;padding:3px 7px;{_lbl_inst}">DIRECCI&#211;N T&#201;CNICA DE CONSTRUCCIONES</td>'
            f'<td style="width:62%;padding:3px 7px;{_lbl_inst}">SUBDIRECCI&#211;N T&#201;CNICA &nbsp; {_sub_txt}</td>'
            f'</tr>'
            f'</table>'
            f'</td></tr>'
            f'</table>'
        )
    elif entidad_upper == "ICCU":
        seccion_institucional = (
            f'<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">'
            f'<tr><td style="border:1px solid {navy};border-top:none;border-bottom:none;text-align:center;padding:3px 8px;{_lbl_inst}">'
            f'INSTITUTO DE CAMINOS Y CONSTRUCCIONES DE CUNDINAMARCA'
            f'</td></tr></table>'
        )
    else:
        seccion_institucional = (
            f'<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">'
            f'<tr><td style="border:1px solid {navy};border-top:none;border-bottom:none;text-align:center;padding:3px 8px;{_lbl_inst}">'
            f'{_h(entidad)}'
            f'</td></tr></table>'
        )

    def _fmt_fecha(iso: str) -> str:
        """'2024-03-11' → '11/03/2024'"""
        if not iso or len(iso) < 10:
            return iso or ""
        try:
            y, m, d = iso[:10].split("-")
            return f"{d}/{m}/{y}"
        except Exception:
            return iso

    def _fmt_cant(v) -> str:
        """Formato numérico con 3 decimales (punto miles, coma decimal)."""
        try:
            n = float(v)
            formatted = f"{n:,.3f}"
            return formatted.replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return z

    # ── Estilos de la grilla: líneas muy sutiles punteadas ──────────────────
    # Color muy suave para que las líneas sean casi imperceptibles
    bd_row = "border:1px dotted #aec6d4"
    # Tamaño de fuente de datos: +2pt respecto al diseño base (6pt → 8pt)
    fs_data = "8pt"
    fs_hdr_grid = "8.5pt"
    # Siempre 10 filas visibles en la grilla
    TOTAL_FILAS = 10

    # ── Filas de datos: una por semana N3 + vacías hasta completar 10 ───────
    if semanas_item:
        data_rows: list = []
        for s in semanas_item:
            num_sem = s.get("numero_semana")
            fi = _fmt_fecha(s.get("fecha_inicio", ""))
            ff = _fmt_fecha(s.get("fecha_fin", ""))
            periodo = f"{fi} - {ff}" if (fi or ff) else z
            ubicacion = f"Semana {num_sem} \u00b7 {periodo}" if num_sem else f"Semana {z} \u00b7 {periodo}"
            cant_fmt = _fmt_cant(s.get("cantidad"))
            data_rows.append((ubicacion, cant_fmt))
        total_acta_fmt = _fmt_cant(sum(s.get("cantidad", 0) for s in semanas_item))
    else:
        data_rows = []
        total_acta_fmt = z

    _vm = "vertical-align:middle;"
    filas_det = ""
    for i in range(TOTAL_FILAS):
        if i < len(data_rows):
            ubicacion, cant_fmt = data_rows[i]
            filas_det += (
                f'<tr style="height:22px;">'
                f'<td style="{bd_row};padding:3px 5px;font-size:{fs_data};color:{navy};{_vm}">{_h(ubicacion)}</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">{z}</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">{z}</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">{z}</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">{z}</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;font-weight:bold;color:{navy};{_vm}">{cant_fmt}</td>'
                f'</tr>'
            )
        else:
            filas_det += (
                f'<tr style="height:22px;">'
                f'<td style="{bd_row};padding:3px 4px;font-size:{fs_data};color:{navy};{_vm}">&nbsp;</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">&nbsp;</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">&nbsp;</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">&nbsp;</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">&nbsp;</td>'
                f'<td style="{bd_row};padding:3px 3px;font-size:{fs_data};text-align:center;color:{navy};{_vm}">&nbsp;</td>'
                f'</tr>'
            )

    # Calcular totales
    _t_presente = sum(s.get("cantidad", 0) for s in semanas_item) if semanas_item else 0.0
    _t_anteriores = total_actas_anteriores if total_actas_anteriores is not None else None
    _t_acumulado = (_t_presente + _t_anteriores) if _t_anteriores is not None else None

    tot_rows = [
        ("TOTAL EJECUTADO PRESENTE ACTA",    total_acta_fmt),
        ("TOTAL EJECUTADO ACTAS ANTERIORES", _fmt_cant(_t_anteriores) if _t_anteriores is not None else z),
        ("TOTAL EJECUTADO ACUMULADO",        _fmt_cant(_t_acumulado)  if _t_acumulado  is not None else z),
        ("TOTAL POR EJECUTAR",               z),
    ]
    tot_block = "".join(
        f'<tr><td colspan="4" style="{bd};padding:3px 5px;font-size:{fs_hdr_grid};text-align:right;'
        f'font-weight:bold;color:{navy_hdr};background-color:{bg_hdr};">'
        f"{_h(lbl)}</td>"
        f'<td colspan="2" style="{bd};padding:3px 5px;font-size:{fs_hdr_grid};text-align:center;'
        f'font-weight:{"bold" if val != z else "normal"};color:{navy};">{val}</td></tr>'
        for lbl, val in tot_rows
    )

    return f"""<!DOCTYPE html>
<html xmlns="http://www.w3.org/1999/xhtml">
<head><meta charset="UTF-8"/><title>FO-EO-04 V2.0 \u2014 IDU</title>
<style type="text/css">
@page {{ size: A3 portrait; margin: 10mm 12mm; }}
body {{ margin:0;padding:0;font-family:Arial,Helvetica,sans-serif;font-size:7pt;color:{navy};line-height:1.2; }}
.lbl {{ font-size:5.5pt;font-weight:bold;text-transform:uppercase;color:{navy_hdr}; }}
.hint {{ font-size:5pt;color:#4a5568;margin-top:1px; }}
</style></head><body>

<!-- ═══ ENCABEZADO INSTITUCIONAL ═══ -->
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd}">

  <!-- FILA 1: Título (colspan 3) + Logo (rowspan 2) -->
  <tr>
    <td colspan="3" style="text-align:center;padding:3px 14px;
                           border-bottom:1px solid {navy};border-right:1px solid {navy};">
      <div style="font-size:6pt;color:{navy};">FORMATO</div>
      <div style="font-size:8.5pt;font-weight:bold;color:{navy_hdr};margin-top:1px;">
        MEMORIA DE C&#193;LCULO DE CANTIDADES DE OBRA
      </div>
</td>
    <td rowspan="2" style="width:14%;vertical-align:middle;text-align:center;padding:2px 6px;">
      {logo_entidad_html}
</td>
  </tr>

  <!-- FILA 2: CÓDIGO | PROCESO | VERSIÓN -->
  <tr>
    <td style="width:17%;border-right:1px solid {navy};padding:2px 6px;
               text-align:center;vertical-align:middle;">
      <div class="lbl">C&#243;digo</div>
      <div style="font-size:7pt;font-weight:bold;color:{navy_hdr};">FO-EO-04</div>
</td>
    <td style="width:55%;border-right:1px solid {navy};padding:2px 6px;
               text-align:center;vertical-align:middle;">
      <div class="lbl">Proceso</div>
      <div style="font-size:7pt;font-weight:bold;color:{navy_hdr};">CONSTRUCCI&#211;N DE PROYECTOS</div>
    </td>
    <td style="width:14%;padding:2px 6px;text-align:center;vertical-align:middle;">
      <div class="lbl">Versi&#243;n</div>
      <div style="font-size:7pt;font-weight:bold;color:{navy_hdr};">2.0</div>
</td>
</tr>

</table>

<!-- ═══ SECCIÓN INSTITUCIONAL (condicional por entidad) ═══ -->
{seccion_institucional}

<!-- ═══ DATOS GENERALES DEL CONTRATO ═══ -->
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd};border-top:none;">
<tr>
  <td style="{bd};padding:4px 7px;border-top:none;border-bottom:none;">
    <table width="50%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
    <tr>
      <td style="padding-right:5px;font-weight:bold;color:{navy_hdr};font-size:7pt;white-space:nowrap;vertical-align:bottom;">
        CONTRATO No
      </td>
      <td style="padding:0 6px 0 0;vertical-align:bottom;">
        <div style="border-bottom:1px solid {navy};min-width:55px;text-align:center;
                    font-size:8pt;font-weight:bold;color:{navy_hdr};padding:0 4px;">
          {_h(num_contrato)}
        </div>
        <div style="font-size:4.5pt;color:#6b7280;text-align:center;margin-top:1px;">
          (N&#250;mero de Contrato)
        </div>
      </td>
      <td style="padding-right:5px;font-weight:bold;color:{navy_hdr};font-size:7pt;white-space:nowrap;vertical-align:bottom;">
        DE
      </td>
      <td style="padding:0;vertical-align:bottom;">
        <div style="border-bottom:1px solid {navy};min-width:40px;text-align:center;
                    font-size:8pt;font-weight:bold;color:{navy_hdr};padding:0 4px;">
          {_h(año_contrato)}
        </div>
        <div style="font-size:4.5pt;color:#6b7280;text-align:center;margin-top:1px;">
          (A&#241;o de suscripci&#243;n)
        </div>
</td>
</tr>
    </table>
    <div style="margin-top:5px;font-weight:bold;color:{navy_hdr};">OBJETO DEL CONTRATO</div>
    <div style="border:1px solid {navy};padding:4px 6px;margin-top:2px;
                font-size:7pt;color:{navy_hdr};min-height:22px;line-height:1.4;">
      {_h(objeto_contrato) if objeto_contrato else "&nbsp;"}
    </div>
</td>
</tr>
<tr>
  <td style="{bd};padding:4px 7px;border-top:none;">
    <!-- RECIBO PARCIAL: 6 celdas alternando label / valor -->
    <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;margin-bottom:4px;"><tr>
    <td style="width:16%;font-weight:bold;color:{navy_hdr};vertical-align:bottom;white-space:nowrap;">RECIBO PARCIAL No.</td>
    <td style="width:10%;border-bottom:1px solid {navy};text-align:center;font-size:8pt;font-weight:bold;color:{navy_hdr};vertical-align:bottom;">{_h(num_acta)}</td>
    <td style="width:36%;font-weight:bold;color:{navy_hdr};vertical-align:bottom;padding-left:8px;white-space:nowrap;">CORRESPONDIENTE AL PERIODO DESDE</td>
    <td style="width:14%;border-bottom:1px solid {navy};text-align:center;font-size:8pt;font-weight:bold;color:{navy_hdr};vertical-align:bottom;">{_h(fecha_desde)}</td>
    <td style="width:6%;font-weight:bold;color:{navy_hdr};vertical-align:bottom;padding-left:8px;white-space:nowrap;">HASTA</td>
    <td style="width:18%;border-bottom:1px solid {navy};text-align:center;font-size:8pt;font-weight:bold;color:{navy_hdr};vertical-align:bottom;">{_h(fecha_hasta)}</td>
    </tr></table>
    <!-- ITEM No | UNIDAD | ESPECIFICACIÓN TÉCNICA | CAPÍTULO (una sola línea) -->
    <table width="100%" cellspacing="0" cellpadding="0" style="margin-top:3px;border-collapse:collapse;">
    <tr>
      <td style="width:15%;white-space:nowrap;vertical-align:bottom;padding-bottom:2px;">
        <span style="font-weight:bold;color:{navy_hdr};">ITEM No</span>&nbsp;
        <span style="border-bottom:1px solid {navy};display:inline-block;min-width:42px;font-size:8pt;font-weight:bold;color:{navy_hdr};">{_h(item_numero) if item_numero else "&nbsp;"}</span>
</td>
      <td style="width:10%;white-space:nowrap;vertical-align:bottom;padding-bottom:2px;">
        <span style="font-weight:bold;color:{navy_hdr};">UNIDAD</span>&nbsp;
        <span style="border-bottom:1px solid {navy};display:inline-block;min-width:30px;font-size:8pt;color:{navy_hdr};">{_h(item_unidad) if item_unidad else "&nbsp;"}</span>
</td>
      <td style="width:45%;white-space:nowrap;vertical-align:bottom;padding-bottom:2px;">
        <span style="font-weight:bold;color:{navy_hdr};">ESPECIFICACI&#211;N T&#201;CNICA</span>&nbsp;
        <span style="border-bottom:1px solid {navy};display:inline-block;min-width:80px;font-size:8pt;color:{navy_hdr};">{_h(item_especificacion) if item_especificacion else "&nbsp;"}</span>
</td>
      <td style="width:30%;vertical-align:bottom;padding-bottom:2px;">
        <span style="border-bottom:1px solid {navy};display:inline-block;width:100%;font-size:8pt;font-weight:bold;color:{navy_hdr};text-align:center;">{_h(item_capitulo) if item_capitulo else "&nbsp;"}</span>
</td>
</tr>
    </table>
    <!-- DESCRIPCIÓN DEL ÍTEM: en recuadro (como OBJETO DEL CONTRATO) -->
    <div style="margin-top:3px;font-weight:bold;font-size:6pt;color:{navy_hdr};">DESCRIPCI&#211;N DEL &#205;TEM</div>
    <div style="border:1px solid {navy};padding:3px 6px;margin-top:1px;font-size:7pt;
                color:{navy_hdr};min-height:18px;line-height:1.4;">
      {_h(item_descripcion) if item_descripcion else "&nbsp;"}
    </div>
</td>
</tr>
<tr>
  <td style="{bd};padding:3px 7px;border-top:none;">
    <!-- CONTRATISTA -->
    <table width="100%" cellspacing="0" cellpadding="0" style="margin-bottom:2px;">
    <tr>
      <td style="width:20%;font-weight:bold;font-size:6.5pt;color:{navy_hdr};white-space:nowrap;vertical-align:bottom;">CONTRATISTA</td>
      <td style="border-bottom:1px solid {navy};font-size:7.5pt;font-weight:bold;color:{navy_hdr};vertical-align:bottom;padding-left:4px;">{_h(contratista) if contratista else "&nbsp;"}</td>
</tr>
    </table>
    <!-- INTERVENTORÍA -->
    <table width="100%" cellspacing="0" cellpadding="0" style="margin-bottom:2px;">
    <tr>
      <td style="width:20%;font-weight:bold;font-size:6.5pt;color:{navy_hdr};white-space:nowrap;vertical-align:bottom;">INTERVENTOR&#205;A</td>
      <td style="border-bottom:1px solid {navy};font-size:7.5pt;font-weight:bold;color:{navy_hdr};vertical-align:bottom;padding-left:4px;">{_h(interventoria) if interventoria else "&nbsp;"}</td>
</tr>
    </table>
    <!-- SUPERVISOR -->
    <table width="100%" cellspacing="0" cellpadding="0">
    <tr>
      <td style="width:20%;font-weight:bold;font-size:6.5pt;color:{navy_hdr};white-space:nowrap;vertical-align:bottom;">SUPERVISOR(A)</td>
      <td style="border-bottom:1px solid {navy};font-size:7.5pt;font-weight:bold;color:{navy_hdr};vertical-align:bottom;padding-left:4px;">{_h(supervisor) if supervisor else "&nbsp;"}</td>
    </tr>
    </table>
</td>
</tr>
</table>

<!-- ═══ TABLA DE CANTIDADES ═══
     Proporciones del SVG original: UBICACIÓN 40% · LARGO/ANCHO/ALTO 13% c/u · CANTIDAD 12% · TOTAL 9% -->
<table width="100%" cellspacing="0" cellpadding="0"
       style="border-collapse:collapse;{bd};border-top:none;margin-top:0;">
<tr style="background-color:{bg_hdr};">
  <td style="{bd};padding:3px 5px;font-size:{fs_hdr_grid};font-weight:bold;text-align:center;
            width:40%;color:{navy_hdr};">UBICACI&#211;N</td>
  <td style="{bd};padding:3px 4px;font-size:{fs_hdr_grid};font-weight:bold;text-align:center;
            width:13%;color:{navy_hdr};">LARGO</td>
  <td style="{bd};padding:3px 4px;font-size:{fs_hdr_grid};font-weight:bold;text-align:center;
            width:13%;color:{navy_hdr};">ANCHO</td>
  <td style="{bd};padding:3px 4px;font-size:{fs_hdr_grid};font-weight:bold;text-align:center;
            width:13%;color:{navy_hdr};">ALTO</td>
  <td style="{bd};padding:3px 4px;font-size:{fs_hdr_grid};font-weight:bold;text-align:center;
            width:12%;color:{navy_hdr};">CANTIDAD</td>
  <td style="{bd};padding:3px 4px;font-size:{fs_hdr_grid};font-weight:bold;text-align:center;
            width:9%;color:{navy_hdr};">TOTAL</td>
</tr>
{filas_det}
{tot_block}
</table>

<!-- ═══ PLANO / FOTOGRAFÍA ═══ -->
<table width="100%" cellspacing="0" cellpadding="0"
       style="border-collapse:collapse;margin-top:6px;{bd};border-top:none;">
<tr>
<td width="50%" style="vertical-align:top;{bd};padding:0;border-top:none;">
  <div style="font-weight:bold;font-size:6.5pt;padding:3px 5px;
              border-bottom:1px solid {navy};text-align:center;
              background-color:{bg_hdr};color:{navy_hdr};">PLANO/ESQUEMA</div>
  <div style="height:240px;min-height:240px;max-height:240px;padding:6px;text-align:center;overflow:hidden;">
    {_fo_eo_04_img_block_html(grafico_url, grafico_numero, "grafico", preview_ui=preview_ui, empty_label="Sin gr&aacute;fico", mostrar_rotar=mostrar_rotar)}
  </div>
  <div style="height:16px;min-height:16px;font-size:5.5pt;text-align:center;padding:2px;border-top:1px solid {navy};color:{navy_hdr};">
    {"Gr&aacute;fico #" + str(grafico_numero) if grafico_numero else "&nbsp;"}
  </div>
</td>
<td width="50%" style="vertical-align:top;{bd};padding:0;border-top:none;border-left:none;">
  <div style="font-weight:bold;font-size:6.5pt;padding:3px 5px;
              border-bottom:1px solid {navy};text-align:center;
              background-color:{bg_hdr};color:{navy_hdr};">FOTOGRAF&#205;A</div>
  <div style="height:240px;min-height:240px;max-height:240px;padding:6px;text-align:center;overflow:hidden;">
    {_fo_eo_04_img_block_html(foto_url, foto_numero, "foto", preview_ui=preview_ui, empty_label="Sin fotograf&iacute;a", mostrar_rotar=mostrar_rotar)}
  </div>
  <div style="height:16px;min-height:16px;font-size:5.5pt;text-align:center;padding:2px;border-top:1px solid {navy};color:{navy_hdr};">
    {"Foto #" + str(foto_numero) if foto_numero else "&nbsp;"}
  </div>
</td>
</tr>
</table>

<!-- ═══ OBSERVACIONES ═══ -->
<div style="margin-top:6px;{bd};padding:5px 7px;">
  <div style="font-weight:bold;margin-bottom:3px;color:{navy_hdr};">OBSERVACIONES</div>
  <div style="font-size:7pt;text-align:justify;line-height:1.4;color:{navy};">{_h(obs)}</div>
</div>

<!-- ═══ FIRMAS: ELABORÓ / REVISÓ (2 × 2) ═══ -->
{_html_fo_eo_04_firmas_4(navy, navy_hdr, bd, elaboro_nombre, elaboro_cargo, elaboro2_nombre, elaboro2_cargo, reviso_nombre, reviso_cargo, reviso2_nombre, reviso2_cargo,
    elaboro_firma_data_uri=elaboro_firma_data_uri, elaboro2_firma_data_uri=elaboro2_firma_data_uri, reviso_firma_data_uri=reviso_firma_data_uri, reviso2_firma_data_uri=reviso2_firma_data_uri,
    elaboro_marca_fecha=elaboro_marca_fecha, elaboro_marca_hora=elaboro_marca_hora, elaboro2_marca_fecha=elaboro2_marca_fecha, elaboro2_marca_hora=elaboro2_marca_hora,
    reviso_marca_fecha=reviso_marca_fecha, reviso_marca_hora=reviso_marca_hora, reviso2_marca_fecha=reviso2_marca_fecha, reviso2_marca_hora=reviso2_marca_hora,
    elaboro_marca_usuario=elaboro_marca_usuario, elaboro2_marca_usuario=elaboro2_marca_usuario,
    reviso_marca_usuario=reviso_marca_usuario, reviso2_marca_usuario=reviso2_marca_usuario)}

<!-- ═══ DISTRIBUCIÓN ═══ -->
<div style="font-size:5.5pt;text-align:center;margin-top:10px;padding:4px;color:#334155;">
ORIGINAL: INTERVENTOR&#205;A &nbsp;|&nbsp; 1ra COPIA: CONTRATISTA &nbsp;|&nbsp; 2da COPIA: DEPENDENCIA RESPONSABLE DEL CONTRATO
</div>
<div style="font-size:5pt;text-align:center;margin-top:3px;color:#64748b;">
Vista previa ClaraCore &middot; sin datos de obra &middot; {CODIGO_FORMATO_IDU_FO_EO_04_V2}
</div>
</body></html>"""


def _fo_eo_04_img_src_for_pdf(http_url: str, cache: Dict[str, str]) -> Optional[str]:
    """
    src para xhtml2pdf: data-URI embebida si cabe; si falla o es enorme, URL http
    (el worker usa link_callback para descargarla al renderizar).
    """
    if not http_url or not str(http_url).strip():
        return None
    u = str(http_url).strip()
    if u.startswith("data:"):
        return u
    if u in cache:
        return cache[u] or u
    return u


def _fo_eo_04_prefetch_item_imgs_data_uri(
    items_n3: list,
    on_progress=None,
) -> List[Tuple[Optional[str], Optional[str]]]:
    """Data URI por ítem; cada URL única se descarga una sola vez (mismo foto_numero en muchos ítems)."""
    n = len(items_n3)
    if not n:
        return []
    unique_urls: set = set()
    for item in items_n3:
        for key in ("foto_url", "grafico_url"):
            u = (item.get(key) or "").strip()
            if u.startswith("http"):
                unique_urls.add(u)
    url_to_uri: Dict[str, str] = {}
    if unique_urls:
        n_u = len(unique_urls)
        done_u = 0
        with ThreadPoolExecutor(max_workers=min(_FO_EO04_ITEM_IMG_WORKERS, n_u)) as pool:
            futs = {pool.submit(_url_a_data_url_pdf, u): u for u in unique_urls}
            for fut in as_completed(futs):
                u = futs[fut]
                done_u += 1
                if on_progress:
                    on_progress({
                        "pct": 50 + int(4 * done_u / max(n_u, 1)),
                        "msg": f"Descargando fotos/gráficos ({done_u}/{n_u})…",
                    })
                try:
                    uri = fut.result(timeout=_FO_EO04_IMG_FETCH_TIMEOUT_SEC + 5)
                    if uri and len(uri) <= _FO_EO04_MAX_DATA_URI_LEN:
                        url_to_uri[u] = uri
                    else:
                        url_to_uri[u] = u
                except Exception as exc:
                    _log.warning("fo_eo_04 prefetch url=%s: %s", u[:80], exc)
                    url_to_uri[u] = u
        ok_embed = sum(1 for v in url_to_uri.values() if v.startswith("data:"))
        _log.info(
            "fo_eo_04 prefetch: %s urls, %s embebidas, %s http para %s ítems",
            len(unique_urls),
            ok_embed,
            len(unique_urls) - ok_embed,
            n,
        )
    out: List[Tuple[Optional[str], Optional[str]]] = []
    for item in items_n3:
        fu = (item.get("foto_url") or "").strip()
        gu = (item.get("grafico_url") or "").strip()
        out.append((
            _fo_eo_04_img_src_for_pdf(fu, url_to_uri) if fu else None,
            _fo_eo_04_img_src_for_pdf(gu, url_to_uri) if gu else None,
        ))
    return out


def _fo_eo_04_init_pdf_pool() -> None:
    """Crea el pool en el hilo de la petición HTTP (Windows no permite hijos desde hilos daemon).

    Hace un ping con timeout para detectar hosts donde el ProcessPool no arranca
    (p. ej. /dev/shm minúsculo en App Service): en ese caso marca el pool como roto
    y el render cae a modo secuencial en vez de quedarse colgado indefinidamente.
    """
    global _FO_EO04_PROCESS_POOL, _FO_EO04_POOL_BROKEN
    if _FO_EO04_POOL_BROKEN or _FO_EO04_PROCESS_POOL is not None:
        return
    with _FO_EO04_POOL_LOCK:
        if _FO_EO04_POOL_BROKEN or _FO_EO04_PROCESS_POOL is not None:
            return
        import multiprocessing
        from concurrent.futures import ProcessPoolExecutor

        workers = _fo_eo_04_pdf_page_workers()
        ctx = multiprocessing.get_context("spawn")
        pool = ProcessPoolExecutor(max_workers=workers, mp_context=ctx)
        try:
            from fo_eo_04_pdf_worker import ping as _pool_ping

            pool.submit(_pool_ping).result(timeout=_FO_EO04_POOL_PING_TIMEOUT_SEC)
        except Exception as exc:
            _FO_EO04_POOL_BROKEN = True
            _log.warning(
                "fo_eo_04 PDF pool no operativo (%s); se usará modo secuencial. "
                "Ajusta FO_EO04_PDF_PAGE_WORKERS o FO_EO04_PDF_USE_PROCESSES=0 si persiste.",
                exc,
            )
            try:
                pool.shutdown(wait=False, cancel_futures=True)
            except Exception:
                pass
            return
        _FO_EO04_PROCESS_POOL = pool
        _log.info("fo_eo_04 PDF pool: %s workers (ping OK)", workers)


def _fo_eo_04_pdf_from_pages(
    pages_html: List[str],
    on_progress=None,
) -> bytes:
    """
    Una memoria → un PDF pequeño → fusión. Procesos en lotes de 2 páginas (~meta 2 mem/s).
    """
    n = len(pages_html)
    if n == 0:
        return _to_pdf("<!DOCTYPE html><html><head><meta charset=\"UTF-8\"/></head><body></body></html>")
    if n == 1:
        if on_progress:
            on_progress({"pct": 88, "msg": "Renderizando PDF…", "current_item": 1, "total_items": 1})
        return _to_pdf(pages_html[0])

    t0 = time.time()
    done_pages = 0

    def _report(done: int) -> None:
        if not on_progress:
            return
        elapsed = max(0.001, time.time() - t0)
        rate = done / elapsed
        on_progress({
            "pct": 76 + int(22 * done / n),
            "msg": f"Renderizando PDF… ({done}/{n} páginas · ~{rate:.1f} mem/s)",
            "current_item": done,
            "total_items": n,
        })

    use_processes = (
        _fo_eo_04_pdf_use_processes()
        and n >= _FO_EO04_PDF_PARALLEL_MIN_PAGES
        and _FO_EO04_PROCESS_POOL is not None
    )
    per_task = max(1, _FO_EO04_PDF_PAGES_PER_TASK)
    page_pdfs: List[Optional[bytes]] = [None] * n

    if use_processes:
        from fo_eo_04_pdf_worker import render_html_batch

        pool = _FO_EO04_PROCESS_POOL
        batches: List[Tuple[int, List[str]]] = []
        for start in range(0, n, per_task):
            batches.append((start, pages_html[start : start + per_task]))
        futs = {pool.submit(render_html_batch, batch_html): start for start, batch_html in batches}
        for fut in as_completed(futs):
            start = futs[fut]
            batch_html = pages_html[start : start + per_task]
            try:
                chunk_pdfs = fut.result(timeout=_FO_EO04_PDF_TASK_TIMEOUT_SEC)
            except Exception as exc:
                _log.warning(
                    "fo_eo_04 lote PDF start=%s timeout/err (%s); secuencial",
                    start,
                    exc,
                )
                chunk_pdfs = [_to_pdf(h) for h in batch_html]
            for j, pdf in enumerate(chunk_pdfs):
                if start + j < n:
                    page_pdfs[start + j] = pdf
            done_pages = sum(1 for p in page_pdfs if p is not None)
            _report(done_pages)
    else:
        _log.info("fo_eo_04 PDF secuencial (sin ProcessPool): %s páginas", n)
        for i, h in enumerate(pages_html):
            if on_progress:
                on_progress({
                    "pct": 76 + int(22 * (i + 1) / max(n, 1)),
                    "msg": f"Renderizando PDF página {i + 1}/{n}…",
                    "current_item": i + 1,
                    "total_items": n,
                })
            page_pdfs[i] = _to_pdf_unlocked(h)
            _report(i + 1)

    parts = [p for p in page_pdfs if p]
    if not parts:
        raise ValueError("No se generó ninguna página PDF")
    merged = _merge_pdf_bytes_tree(parts)
    elapsed = time.time() - t0
    _log.info(
        "fo_eo_04 PDF listo: %s páginas en %.1fs (%.2f mem/s) procesos=%s",
        n,
        elapsed,
        n / max(0.001, elapsed),
        use_processes,
    )
    return merged


def _build_fo_eo_04_html(
    contrato_id: int,
    formato_codigo: str,
    subsistema: str,
    acta_id: Optional[int],
    supervisor: str,
    current_user: Optional[dict] = None,
    *,
    preview_ui: bool = False,
    on_progress=None,
    standalone: bool = False,
    mostrar_rotar: bool = True,
) -> tuple:
    """
    Arma el HTML del FO-IDU-EO-04-V2 (todas las páginas del acta).
    preview_ui=True: URLs directas (solo vista previa en pantalla).
    standalone=True: devuelve un documento HTML completo (para iframe), no un fragmento.
    mostrar_rotar=False: omite los botones «Girar» (no funcionan fuera de React/iframe).
    Devuelve (html_combinado, fname, contrato_numero, lista_html_por_página).
    """

    def _prog(
        pct: int,
        msg: str,
        current_item: Optional[int] = None,
        total_items: Optional[int] = None,
        **extra,
    ):
        if on_progress:
            payload = {
                "pct": pct,
                "msg": msg,
                "current_item": current_item,
                "total_items": total_items,
            }
            payload.update(extra)
            on_progress(payload)

    def _fmt_fecha(f: object) -> str:
        if not f:
            return ""
        try:
            from datetime import date as _date
            d = _date.fromisoformat(str(f)[:10])
            return f"{d.day:02d}/{d.month:02d}/{d.year}"
        except Exception:
            return str(f)[:10]

    acta_id_norm = _resolver_acta_id_en_contrato(contrato_id, acta_id)

    _prog(5, "Leyendo datos del contrato…")
    logo_entidad: Optional[str] = None
    entidad: str = "IDU"
    num_contrato: str = ""
    año_contrato: str = ""
    objeto_contrato: str = ""
    contratista_nombre: str = ""
    interventoria_nombre: str = ""
    contrato_numero_raw: str = ""
    try:
        contrato_row = _row(
            "contratos",
            "logo_entidad, entidad, entidad_otra, numero, objeto, contratista, interventoria",
            id=contrato_id,
        )
        if contrato_row:
            logo_entidad = contrato_row.get("logo_entidad") or None
            ent = (contrato_row.get("entidad") or "").strip().upper()
            if ent == "OTRA":
                ent = (contrato_row.get("entidad_otra") or "OTRA").strip().upper()
            entidad = ent or "IDU"
            num_contrato, año_contrato = _parse_numero_contrato(contrato_row.get("numero") or "")
            objeto_contrato = (contrato_row.get("objeto") or "").strip()
            contratista_nombre = (contrato_row.get("contratista") or "").strip()
            interventoria_nombre = (contrato_row.get("interventoria") or "").strip()
            contrato_numero_raw = (contrato_row.get("numero") or "").strip()
    except Exception:
        pass

    logo_entidad_html = _html_logo_entidad(None)
    if logo_entidad:
        try:
            logo_du = _url_a_data_url_pdf(logo_entidad)
            logo_entidad_html = _html_logo_entidad(logo_du or logo_entidad)
        except Exception as exc:
            _log.warning("fo_eo_04 logo entidad: %s", exc)
            logo_entidad_html = _html_logo_entidad(logo_entidad)

    _prog(15, "Consultando datos del acta…")
    num_acta: str = ""
    fecha_desde: str = ""
    fecha_hasta: str = ""
    if acta_id_norm:
        try:
            acta_row = _row("actas", "id, numero_rpo, consecutivo, fecha_inicio, fecha_fin", id=acta_id_norm)
            if acta_row:
                num_acta = str(acta_row.get("numero_rpo") or acta_row.get("consecutivo") or "")
                fecha_desde = _fmt_fecha(acta_row.get("fecha_inicio"))
                fecha_hasta = _fmt_fecha(acta_row.get("fecha_fin"))
        except Exception:
            pass

    _prog(22, "Consultando configuración de firmas…")
    subsistema_norm = subsistema.strip().lower()
    if subsistema_norm not in ("vial", "transporte"):
        subsistema_norm = "vial"

    firma_cfg: Dict[str, Any] = {}
    try:
        firma_cfg = _get_firma_cfg_para_documento(
            contrato_id,
            formato_codigo,
            contexto_tipo="acta_rpo" if acta_id_norm else None,
            contexto_id=acta_id_norm if acta_id_norm else None,
        )
    except Exception:
        pass

    _base_kwargs = dict(
        logo_entidad=logo_entidad,
        logo_entidad_html=logo_entidad_html,
        entidad=entidad,
        subsistema=subsistema_norm,
        num_contrato=num_contrato,
        año_contrato=año_contrato,
        objeto_contrato=objeto_contrato,
        num_acta=num_acta,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        contratista=contratista_nombre,
        interventoria=interventoria_nombre,
        supervisor=supervisor.strip(),
        elaboro_nombre=str(firma_cfg.get("elaboro_nombre") or "").strip(),
        elaboro_cargo=str(firma_cfg.get("elaboro_cargo") or "").strip(),
        elaboro2_nombre=str(firma_cfg.get("elaboro2_nombre") or "").strip(),
        elaboro2_cargo=str(firma_cfg.get("elaboro2_cargo") or "").strip(),
        reviso_nombre=str(firma_cfg.get("reviso_nombre") or "").strip(),
        reviso_cargo=str(firma_cfg.get("reviso_cargo") or "").strip(),
        reviso2_nombre=str(firma_cfg.get("reviso2_nombre") or "").strip(),
        reviso2_cargo=str(firma_cfg.get("reviso2_cargo") or "").strip(),
    )
    e_uri, e2_uri, r_uri, r2_uri = _fo_eo_04_firmas_data_uris(
        contrato_id, formato_codigo, acta_id_norm, firma_cfg, current_user
    )
    (em_u, em_f, em_h), (e2m_u, e2m_f, e2m_h), (rm_u, rm_f, rm_h), (r2m_u, r2m_f, r2m_h) = _fo_eo_04_firmas_marcas_registro(
        contrato_id, formato_codigo, acta_id_norm
    )
    _base_kwargs.update(
        elaboro_firma_data_uri=e_uri,
        elaboro2_firma_data_uri=e2_uri,
        reviso_firma_data_uri=r_uri,
        reviso2_firma_data_uri=r2_uri,
        elaboro_marca_usuario=em_u,
        elaboro_marca_fecha=em_f,
        elaboro_marca_hora=em_h,
        elaboro2_marca_usuario=e2m_u,
        elaboro2_marca_fecha=e2m_f,
        elaboro2_marca_hora=e2m_h,
        reviso_marca_usuario=rm_u,
        reviso_marca_fecha=rm_f,
        reviso_marca_hora=rm_h,
        reviso2_marca_usuario=r2m_u,
        reviso2_marca_fecha=r2m_f,
        reviso2_marca_hora=r2m_h,
    )

    _prog(30, "Obteniendo ítems (aprobados en nivel máximo del contrato)…")
    items_n3 = _fetch_items_n3_acta(acta_id_norm, contrato_id) if acta_id_norm else []
    n_items = len(items_n3)
    if acta_id_norm and not items_n3:
        campo_mx, niveles_act = matriz_params_contrato(_sb, int(contrato_id))
        raise ValueError(
            f"No hay líneas aprobadas en {campo_mx} (niveles activos {niveles_act}) "
            f"para el acta {acta_id_norm}."
        )

    if items_n3:
        _prog(
            48,
            f"Calculando totales de actas anteriores ({n_items} ítem{'s' if n_items != 1 else ''})…",
            total_items=n_items,
        )
        totales_batch, fo_totales_meta = (
            _fetch_totales_batch(contrato_id, acta_id_norm, items_n3)
            if acta_id_norm
            else ({}, {})
        )
        _prog(
            49,
            f"Totales actas anteriores: {fo_totales_meta.get('items_con_acumulado', 0)} ítem(s) con acumulado",
            total_items=n_items,
            fo_totales_meta=fo_totales_meta,
        )
        if not preview_ui:
            _prog(
                50,
                f"Descargando fotos y gráficos ({n_items} ítem{'s' if n_items != 1 else ''})…",
                total_items=n_items,
            )
            img_pairs = _fo_eo_04_prefetch_item_imgs_data_uri(
                items_n3,
                on_progress=lambda d: _prog(
                    d.get("pct", 52),
                    d.get("msg", "Descargando imágenes…"),
                    total_items=n_items,
                )
                if on_progress
                else None,
            )
        else:
            img_pairs = None

        pages = []
        for i, item in enumerate(items_n3):
            item_num = (item.get("item_numero") or "").strip()
            item_cap = (item.get("capitulo") or "").strip()
            desc_corta = (item.get("item_descripcion") or item_num or "")[:40]
            if on_progress:
                pct_page = 55 + int(20 * (i + 1) / max(n_items, 1))
                _prog(pct_page, f"Página {i + 1}/{n_items}: {desc_corta}", current_item=i + 1, total_items=n_items)
            t_anteriores = totales_batch.get(_fo_eo_04_norm_item_key(item_num, item_cap), 0.0) if acta_id_norm else None
            if preview_ui:
                foto_src = (item.get("foto_url") or "").strip() or None
                graf_src = (item.get("grafico_url") or "").strip() or None
            else:
                foto_src, graf_src = img_pairs[i]
            pages.append(
                _html_idu_fo_eo_04_v2_plantilla_vacia(
                    **_base_kwargs,
                    item_numero=item_num,
                    item_unidad=item.get("unidad") or "",
                    item_especificacion=item.get("especificacion_tecnica") or "",
                    item_capitulo=item_cap,
                    item_descripcion=item.get("item_descripcion") or "",
                    semanas_item=item.get("semanas") or [],
                    total_actas_anteriores=t_anteriores,
                    foto_url=foto_src,
                    foto_numero=item.get("foto_numero"),
                    grafico_url=graf_src,
                    grafico_numero=item.get("grafico_numero"),
                    preview_ui=preview_ui,
                    mostrar_rotar=mostrar_rotar,
                )
            )
        _prog(76, f"Documento listo ({n_items} página{'s' if n_items != 1 else ''})…", total_items=n_items)
        html = _combine_html_pages(pages)
        pages_out = pages
    else:
        _prog(76, "Generando plantilla vacía…")
        pages_out = [_html_idu_fo_eo_04_v2_plantilla_vacia(**_base_kwargs, preview_ui=preview_ui, mostrar_rotar=mostrar_rotar)]
        html = pages_out[0]

    if preview_ui:
        if standalone:
            # Documento HTML completo (con estilos del preview) para mostrar en un iframe.
            html = _fo_eo_04_wrap_preview_html(html)
        else:
            html = _fo_eo_04_wrap_preview_html(_fo_eo_04_html_embed_fragment(html))

    suffix = f"_acta{num_acta}" if num_acta else ""
    fname = _safe_filename_part(f"{formato_codigo}{suffix}.pdf")
    return html, fname, contrato_numero_raw, pages_out


def _build_fo_eo_04_pdf_bytes(
    contrato_id: int,
    formato_codigo: str,
    subsistema: str,
    acta_id: Optional[int],
    supervisor: str,
    current_user: Optional[dict] = None,
) -> tuple:
    """Genera los bytes del PDF FO-IDU-EO-04-V2. Devuelve (pdf_bytes, fname, contrato_numero)."""
    cu = current_user if isinstance(current_user, dict) else dict(current_user or {})
    _html, fname, contrato_numero_raw, pages = _build_fo_eo_04_html(
        contrato_id,
        formato_codigo,
        subsistema,
        acta_id,
        supervisor,
        current_user=cu,
        preview_ui=False,
    )
    pdf_bytes = _fo_eo_04_pdf_from_pages(pages)
    return pdf_bytes, fname, contrato_numero_raw


# ── Sistema de jobs en background para generación progresiva de PDF ──────────

import uuid as _uuid_mod
import threading as _threading_mod
import time as _time_mod
import tempfile as _tempfile_mod
import secrets as _secrets_mod
import shutil as _shutil_mod

_pdf_jobs: dict = {}
_pdf_jobs_lock = _threading_mod.Lock()
_PDF_JOBS_ROOT = _os.environ.get(
    "FO_EO04_PDF_JOBS_DIR",
    _os.path.join(_tempfile_mod.gettempdir(), "claracore_pdf_jobs"),
)
_PDF_JOB_DISK_TTL_SEC = 1800
_PDF_JOB_DL_TOKEN_TTL_SEC = 900


def _pdf_job_disk_dir(job_id: str) -> str:
    root = _PDF_JOBS_ROOT
    _os.makedirs(root, exist_ok=True)
    d = _os.path.join(root, str(job_id))
    _os.makedirs(d, exist_ok=True)
    return d


def _pdf_job_disk_meta_path(job_id: str) -> str:
    return _os.path.join(_pdf_job_disk_dir(job_id), "meta.json")


def _pdf_job_disk_pdf_path(job_id: str) -> str:
    return _os.path.join(_pdf_job_disk_dir(job_id), "memoria.pdf")


def _pdf_job_disk_save(job_id: str, patch: dict) -> None:
    """Persiste estado en disco para que cualquier worker Gunicorn vea el job."""
    try:
        path = _pdf_job_disk_meta_path(job_id)
        cur: dict = {}
        if _os.path.isfile(path):
            with open(path, encoding="utf-8") as f:
                cur = json.load(f)
        cur.update(patch)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(cur, f)
        _os.replace(tmp, path)
    except Exception as exc:
        _log.warning("pdf_job disk save %s: %s", job_id[:8], exc)


def _pdf_job_disk_load(job_id: str) -> Optional[dict]:
    try:
        path = _pdf_job_disk_meta_path(job_id)
        if not _os.path.isfile(path):
            return None
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else None
    except Exception as exc:
        _log.warning("pdf_job disk load %s: %s", job_id[:8], exc)
        return None


def _pdf_job_resolve(job_id: str) -> Optional[dict]:
    with _pdf_jobs_lock:
        mem = dict(_pdf_jobs.get(job_id) or {})
    disk = _pdf_job_disk_load(job_id) or {}
    if not mem and not disk:
        return None
    merged = {**disk, **mem}
    pdf_disk = _pdf_job_disk_pdf_path(job_id)
    if _os.path.isfile(pdf_disk):
        merged["pdf_path"] = pdf_disk
    return merged


def _pdf_job_issue_download_token(job_id: str) -> str:
    tok = _secrets_mod.token_urlsafe(24)
    exp = _time_mod.time() + _PDF_JOB_DL_TOKEN_TTL_SEC
    _pdf_job_disk_save(job_id, {"download_token": tok, "download_token_exp": exp})
    with _pdf_jobs_lock:
        if job_id in _pdf_jobs:
            _pdf_jobs[job_id]["download_token"] = tok
            _pdf_jobs[job_id]["download_token_exp"] = exp
    return tok


def _pdf_job_validate_download_token(contrato_id: int, job_id: str, token: str) -> dict:
    job = _pdf_job_resolve(job_id)
    if not job or int(job.get("contrato_id") or 0) != int(contrato_id):
        raise HTTPException(status_code=404, detail="Job no encontrado")
    if job.get("status") != "listo":
        raise HTTPException(status_code=409, detail="PDF aún no está listo")
    exp = job.get("download_token_exp")
    try:
        exp_f = float(exp)
    except (TypeError, ValueError):
        exp_f = 0.0
    if not token or token != str(job.get("download_token") or ""):
        raise HTTPException(status_code=403, detail="Token de descarga inválido")
    if _time_mod.time() > exp_f:
        raise HTTPException(status_code=410, detail="Enlace de descarga expirado; genere de nuevo.")
    return job


def _pdf_job_unlink_path(path: Optional[str]) -> None:
    if not path:
        return
    try:
        _os.unlink(path)
    except OSError:
        pass


def _pdf_job_remove_dir(job_id: str) -> None:
    try:
        _shutil_mod.rmtree(_pdf_job_disk_dir(job_id), ignore_errors=True)
    except Exception:
        pass


def _pdf_job_store_pdf(job_id: str, pdf_bytes: bytes) -> str:
    """Guarda en carpeta del job (compartida entre workers en el mismo servidor)."""
    path = _pdf_job_disk_pdf_path(job_id)
    tmp = path + ".part"
    with open(tmp, "wb") as f:
        f.write(pdf_bytes)
    _os.replace(tmp, path)
    return path


def _pdf_job_read_bytes(job: dict) -> bytes:
    path = job.get("pdf_path")
    if path and _os.path.isfile(path):
        with open(path, "rb") as f:
            return f.read()
    raw = job.get("bytes")
    if raw:
        return raw
    raise HTTPException(status_code=410, detail="PDF del job ya no está disponible; vuelva a generar.")


def _pdf_jobs_cleanup():
    """Elimina jobs en memoria y carpetas en disco con más de 30 minutos."""
    ahora = _time_mod.time()
    with _pdf_jobs_lock:
        caducados: List[str] = []
        for k, v in list(_pdf_jobs.items()):
            ca = v.get("created_at", 0)
            try:
                ca_f = float(ca)
            except (TypeError, ValueError):
                caducados.append(k)
                continue
            if ahora - ca_f > _PDF_JOB_DISK_TTL_SEC:
                caducados.append(k)
        for k in caducados:
            _pdf_jobs.pop(k, None)
    try:
        if _os.path.isdir(_PDF_JOBS_ROOT):
            for name in _os.listdir(_PDF_JOBS_ROOT):
                d = _os.path.join(_PDF_JOBS_ROOT, name)
                if not _os.path.isdir(d):
                    continue
                meta_p = _os.path.join(d, "meta.json")
                ca_f = ahora
                if _os.path.isfile(meta_p):
                    try:
                        with open(meta_p, encoding="utf-8") as f:
                            meta = json.load(f)
                        ca_f = float(meta.get("created_at") or 0)
                    except Exception:
                        ca_f = 0.0
                if ahora - ca_f > _PDF_JOB_DISK_TTL_SEC:
                    _pdf_job_remove_dir(name)
    except Exception as exc:
        _log.warning("pdf_jobs_cleanup disk: %s", exc)


def _build_fo_eo_04_pdf_bytes_prog(
    contrato_id: int,
    formato_codigo: str,
    subsistema: str,
    acta_id: Optional[int],
    supervisor: str,
    on_progress=None,
    current_user: Optional[dict] = None,
) -> tuple:
    """Igual que _build_fo_eo_04_pdf_bytes con reporte de progreso."""
    cu = current_user if isinstance(current_user, dict) else dict(current_user or {})
    _html, fname, contrato_numero_raw, pages = _build_fo_eo_04_html(
        contrato_id,
        formato_codigo,
        subsistema,
        acta_id,
        supervisor,
        current_user=cu,
        preview_ui=False,
        on_progress=on_progress,
    )
    pdf_bytes = _fo_eo_04_pdf_from_pages(pages, on_progress=on_progress)
    if on_progress:
        on_progress({"pct": 100, "msg": "¡Listo!"})
    return pdf_bytes, fname, contrato_numero_raw


@router.post("/{contrato_id}/ccd/pdf-job/iniciar")
def ccd_pdf_job_iniciar(
    contrato_id: int,
    formato_codigo: str,
    subsistema: str = "vial",
    acta_id: Optional[int] = None,
    supervisor: str = "",
    current_user=Depends(_get_user),
):
    """Inicia la generación del PDF en background. Devuelve {job_id} para hacer polling."""
    _perm_informes_ccd(current_user, "ver")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    meta = FORMATOS_CCD[formato_codigo]
    if meta.get("plantilla_html") != "idu_memoria_fo_eo_04_v2":
        raise HTTPException(status_code=404, detail="Job de PDF no disponible para este formato")
    if acta_id is not None and _resolver_acta_id_en_contrato(contrato_id, acta_id) is None:
        raise HTTPException(status_code=404, detail="Acta no encontrada en este contrato")

    if _fo_eo_04_pdf_use_processes():
        try:
            _fo_eo_04_init_pdf_pool()
        except Exception as pool_exc:
            _log.warning("fo_eo_04 pool al iniciar job (se usará modo secuencial): %s", pool_exc)
    else:
        _log.info("fo_eo_04 PDF job: modo secuencial (Azure o FO_EO04_PDF_USE_PROCESSES=0)")
    _pdf_jobs_cleanup()
    job_id = str(_uuid_mod.uuid4())
    created_at = _time_mod.time()
    job_base = {
        "status": "pendiente",
        "pct": 0,
        "msg": "Iniciando…",
        "current_item": None,
        "total_items": None,
        "bytes": None,
        "fname": None,
        "error": None,
        "contrato_id": contrato_id,
        "created_at": created_at,
    }
    with _pdf_jobs_lock:
        _pdf_jobs[job_id] = dict(job_base)
    _pdf_job_disk_save(job_id, job_base)

    cu_pdf = current_user if isinstance(current_user, dict) else dict(current_user)

    def _run():
        if _fo_eo_04_pdf_use_processes():
            try:
                _fo_eo_04_init_pdf_pool()
            except Exception as pool_exc:
                _log.warning("fo_eo_04 pool en worker: %s", pool_exc)

        def on_progress(d: dict):
            patch = {
                "status": "progresando",
                "pct": d.get("pct", 0),
                "msg": d.get("msg", ""),
                "current_item": d.get("current_item"),
                "total_items": d.get("total_items"),
            }
            if d.get("fo_totales_meta"):
                patch["fo_totales_meta"] = d["fo_totales_meta"]
            with _pdf_jobs_lock:
                if job_id in _pdf_jobs:
                    _pdf_jobs[job_id].update(patch)
            _pdf_job_disk_save(job_id, patch)

        try:
            pdf_bytes, fname, contrato_numero = _build_fo_eo_04_pdf_bytes_prog(
                contrato_id, formato_codigo, subsistema, acta_id, supervisor,
                on_progress=on_progress,
                current_user=cu_pdf,
            )
            pdf_path = _pdf_job_store_pdf(job_id, pdf_bytes)
            dl_tok = _pdf_job_issue_download_token(job_id)
            job_snap = _pdf_job_resolve(job_id) or {}
            done_patch = {
                "status": "listo",
                "pct": 100,
                "msg": "¡Listo!",
                "bytes": None,
                "pdf_path": pdf_path,
                "fname": fname,
                "contrato_numero": contrato_numero,
                "formato_codigo": formato_codigo,
                "acta_id": acta_id,
                "download_token": dl_tok,
                "fo_totales_meta": job_snap.get("fo_totales_meta"),
            }
            with _pdf_jobs_lock:
                if job_id in _pdf_jobs:
                    _pdf_jobs[job_id].update(done_patch)
            _pdf_job_disk_save(job_id, done_patch)
            _log.info(
                "ccd_pdf_job listo job_id=%s bytes=%s",
                job_id[:8],
                len(pdf_bytes),
            )
        except Exception as exc:
            _log.exception("ccd_pdf_job error job_id=%s", job_id)
            err_patch = {"status": "error", "error": str(exc)[:2000]}
            with _pdf_jobs_lock:
                if job_id in _pdf_jobs:
                    _pdf_jobs[job_id].update(err_patch)
            _pdf_job_disk_save(job_id, err_patch)

    t = _threading_mod.Thread(target=_run, daemon=False, name=f"pdf-job-{job_id[:8]}")
    t.start()
    return {"job_id": job_id}


@router.get("/{contrato_id}/ccd/pdf-job/{job_id}/estado")
def ccd_pdf_job_estado(
    contrato_id: int,
    job_id: str,
    current_user=Depends(_get_user),
):
    """Retorna el estado actual del job (para polling desde el frontend)."""
    _perm_informes_ccd(current_user, "ver")
    job = _pdf_job_resolve(job_id)
    if not job or job.get("contrato_id") != contrato_id:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    out = {
        "status": job.get("status"),
        "pct": job.get("pct", 0),
        "msg": job.get("msg", ""),
        "current_item": job.get("current_item"),
        "total_items": job.get("total_items"),
        "error": job.get("error"),
    }
    if job.get("fo_totales_meta"):
        out["fo_totales_meta"] = job.get("fo_totales_meta")
    if job.get("status") == "listo" and job.get("download_token"):
        out["download_token"] = job.get("download_token")
        pdf_p = job.get("pdf_path") or _pdf_job_disk_pdf_path(job_id)
        try:
            if pdf_p and _os.path.isfile(pdf_p):
                out["pdf_size_bytes"] = _os.path.getsize(pdf_p)
        except OSError:
            pass
    return out


def _ccd_pdf_job_file_response(job: dict):
    path = job.get("pdf_path") or ""
    fname = job.get("fname") or "memoria.pdf"
    if path and _os.path.isfile(path):
        from fastapi.responses import FileResponse

        return FileResponse(
            path,
            media_type="application/pdf",
            filename=fname,
            headers={
                "Content-Disposition": f'inline; filename="{fname}"',
                "Cache-Control": "private, max-age=300",
            },
        )
    if job.get("bytes"):
        return Response(
            content=job["bytes"],
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="{fname}"'},
        )
    raise HTTPException(status_code=410, detail="PDF no disponible; genere de nuevo.")


@router.get("/{contrato_id}/ccd/pdf-job/{job_id}/pdf")
def ccd_pdf_job_resultado(
    contrato_id: int,
    job_id: str,
    token: Optional[str] = Query(None, description="Token de descarga (iframe sin Authorization)"),
    current_user=Depends(_get_user_optional),
):
    """PDF del job: FileResponse en streaming. ?token= (visor) o JWT Bearer."""
    dl_tok = (token or "").strip()
    if dl_tok:
        job = _pdf_job_validate_download_token(contrato_id, job_id, dl_tok)
    elif current_user:
        _perm_informes_ccd(current_user, "ver")
        job = _pdf_job_resolve(job_id)
        if not job or job.get("contrato_id") != contrato_id:
            raise HTTPException(status_code=404, detail="Job no encontrado")
        if job.get("status") != "listo":
            raise HTTPException(status_code=409, detail="PDF aún no está listo")
    else:
        raise HTTPException(
            status_code=401,
            detail="Indique ?token= de descarga (estado del job) o Authorization: Bearer.",
        )
    return _ccd_pdf_job_file_response(job)


@router.get("/{contrato_id}/ccd/pdf-job/{job_id}/con-sello-firma")
def ccd_pdf_job_con_sello_firma(
    contrato_id: int,
    job_id: str,
    current_user=Depends(_get_user),
):
    """Anexa sello al PDF ya generado por el job (sin volver a armar las memorias)."""
    _perm_informes_ccd(current_user, "ver")
    job = _pdf_job_resolve(job_id) or {}
    if not job or job.get("contrato_id") != contrato_id:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    if job.get("status") != "listo":
        raise HTTPException(
            status_code=409,
            detail="Primero genere el PDF del acta; luego podrá descargar la versión con sello.",
        )
    fmt = str(job.get("formato_codigo") or "FO-IDU-EO-04-V2")
    acta_ref = job.get("acta_id")
    titulo = f"Memoria de cálculo {fmt}" + (f" — Acta {acta_ref}" if acta_ref else "")
    return _attachment_pdf_con_pagina_sello_usuario(
        _pdf_job_read_bytes(job),
        current_user,
        titulo_doc=titulo,
        formato_ccd=fmt,
        contrato_numero=str(job.get("contrato_numero") or ""),
        nombre_archivo_pdf=str(job.get("fname") or "memoria.pdf"),
    )


@router.get("/{contrato_id}/ccd/fo-eo-04/preview-html")
def ccd_fo_eo_04_preview_html(
    contrato_id: int,
    subsistema: str = "vial",
    acta_id: Optional[int] = None,
    supervisor: str = "",
    formato_codigo: str = "FO-IDU-EO-04-V2",
    standalone: bool = False,
    current_user=Depends(_get_user),
):
    """Vista previa HTML del acta (todas las páginas), renderizada por el navegador.

    Es el camino rápido del botón «lupa»: NO usa xhtml2pdf (sin bloqueo de CPU en el
    worker), así que 120 memorias se muestran en segundos en cualquier plan de Azure.
    standalone=True devuelve un documento HTML completo (para cargar en un iframe) y
    omite los botones «Girar» (la orientación se hace en su propio modal).
    """
    _perm_informes_ccd(current_user, "ver")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    meta = FORMATOS_CCD[formato_codigo]
    if meta.get("plantilla_html") != "idu_memoria_fo_eo_04_v2":
        raise HTTPException(status_code=404, detail="Vista previa HTML no disponible para este formato")
    if acta_id is not None and _resolver_acta_id_en_contrato(contrato_id, acta_id) is None:
        raise HTTPException(status_code=404, detail="Acta no encontrada en este contrato")
    cu = current_user if isinstance(current_user, dict) else dict(current_user)
    acta_norm = _resolver_acta_id_en_contrato(contrato_id, acta_id) if acta_id is not None else None
    try:
        html, _, _, _pages = _build_fo_eo_04_html(
            contrato_id,
            formato_codigo,
            subsistema,
            acta_id,
            supervisor,
            current_user=cu,
            preview_ui=True,
            standalone=standalone,
            mostrar_rotar=not standalone,
        )
    except ValueError as ex:
        raise HTTPException(status_code=404, detail=str(ex)) from ex
    n_items = 0
    if acta_norm:
        n_items = len(_fetch_items_n3_acta(int(acta_norm), int(contrato_id)))
    content = html if standalone else _fo_eo_04_html_embed_fragment(html)
    return HTMLResponse(
        content=content,
        media_type="text/html; charset=utf-8",
        headers={"X-CC-Fo-Eo-04-Items": str(n_items)},
    )


@router.get("/{contrato_id}/ccd/preview-plantilla-vacia/{formato_codigo}")
def ccd_preview_plantilla_vacia_pdf(
    contrato_id: int,
    formato_codigo: str,
    subsistema: str = "vial",
    acta_id: Optional[int] = None,
    supervisor: str = "",
    current_user=Depends(_get_user),
):
    """PDF de plantilla (vista previa inline) — FO-IDU-EO-04-V2.

    Query params:
      subsistema: 'vial' (default) | 'transporte'  — solo aplica cuando la entidad es IDU.
    """
    _perm_informes_ccd(current_user, "ver")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    meta = FORMATOS_CCD[formato_codigo]
    if meta.get("plantilla_html") != "idu_memoria_fo_eo_04_v2":
        raise HTTPException(status_code=404, detail="Vista previa vacía no disponible para este formato")
    if acta_id is not None and _resolver_acta_id_en_contrato(contrato_id, acta_id) is None:
        raise HTTPException(status_code=404, detail="Acta no encontrada en este contrato")
    cu_pdf = current_user if isinstance(current_user, dict) else dict(current_user)
    pdf_bytes, fname, _ = _build_fo_eo_04_pdf_bytes(
        contrato_id, formato_codigo, subsistema, acta_id, supervisor,
        current_user=cu_pdf,
    )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'},
    )


@router.get("/{contrato_id}/ccd/preview-plantilla-vacia/{formato_codigo}/con-sello-firma")
def ccd_preview_plantilla_vacia_con_sello(
    contrato_id: int,
    formato_codigo: str,
    subsistema: str = "vial",
    acta_id: Optional[int] = None,
    supervisor: str = "",
    current_user=Depends(_get_user),
):
    """PDF FO-IDU-EO-04-V2 + página de sello ClaraCore (firma de perfil, SHA-256, fecha)."""
    _perm_informes_ccd(current_user, "ver")
    if formato_codigo not in FORMATOS_CCD:
        raise HTTPException(status_code=404, detail="Código de formato CCD desconocido")
    meta = FORMATOS_CCD[formato_codigo]
    if meta.get("plantilla_html") != "idu_memoria_fo_eo_04_v2":
        raise HTTPException(status_code=404, detail="Vista previa vacía no disponible para este formato")
    if acta_id is not None and _resolver_acta_id_en_contrato(contrato_id, acta_id) is None:
        raise HTTPException(status_code=404, detail="Acta no encontrada en este contrato")
    cu_pdf = current_user if isinstance(current_user, dict) else dict(current_user)
    pdf_bytes, fname, contrato_numero = _build_fo_eo_04_pdf_bytes(
        contrato_id, formato_codigo, subsistema, acta_id, supervisor,
        current_user=cu_pdf,
    )
    return _attachment_pdf_con_pagina_sello_usuario(
        pdf_bytes,
        current_user,
        titulo_doc=f"Memoria de cálculo {formato_codigo}" + (f" — Acta {acta_id}" if acta_id else ""),
        formato_ccd=formato_codigo,
        contrato_numero=contrato_numero,
        nombre_archivo_pdf=fname,
    )


@router.get("/preview/corte-sub-001")
def preview_corte_sub():
    contrato = {
        "numero": "IDU-1551-2017",
        "objeto": "Mantenimiento y rehabilitación de la malla vial local",
        "contratista": "CONSORCIO CONCRESCOL INZIGNIA",
        "nit": "901.234.567-8",
        "interventoria": "SETEC INGENIEROS CONSULTORES",
        "logo_contratista": None
    }
    sub = {
        "razon_social": "SOILING SAS",
        "nit": "900.512.882-1",
        "nombre_contacto": "Carlos Soiling",
        "objeto_contrato": "Excavaciones y perforaciones"
    }
    corte = {
        "consecutivo": 5,
        "fecha_inicio": "2026-03-18",
        "fecha_fin": "2026-04-01",
        "tipo_periodo": "quincenal"
    }
    items = [
        {"capitulo": "IV", "item_numero": "NP-491", "item_descripcion": "PERFORACIÓN PREBARRENADO EN DIAMETRO 3\" HASTA 12M", "unidad": "ML", "cantidad": 144.0, "vlr_unitario_sub": 144716, "costo_directo": 20839104},
        {"capitulo": "IV", "item_numero": "NP-492", "item_descripcion": "EXCAVACIÓN MANUAL EN MATERIAL COMÚN", "unidad": "M3", "cantidad": 38.5, "vlr_unitario_sub": 85000, "costo_directo": 3272500},
        {"capitulo": "IV", "item_numero": "NP-493", "item_descripcion": "RELLENO COMPACTADO CON MATERIAL SELECCIONADO", "unidad": "M3", "cantidad": 22.0, "vlr_unitario_sub": 120000, "costo_directo": 2640000},
        {"capitulo": "IV", "item_numero": "NP-494", "item_descripcion": "SUMINISTRO E INSTALACIÓN TUBERÍA PVC D=8\"", "unidad": "ML", "cantidad": 65.0, "vlr_unitario_sub": 195000, "costo_directo": 12675000},
        {"capitulo": "IV", "item_numero": "NP-495", "item_descripcion": "CONCRETO DE LIMPIEZA f'c=140 kg/cm2", "unidad": "M3", "cantidad": 8.2, "vlr_unitario_sub": 380000, "costo_directo": 3116000},
    ]
    total_costo = sum(i["costo_directo"] for i in items)
    html = _html_cc_sub_v1_plain(
        contrato,
        sub,
        corte,
        items,
        total_costo,
        "Jorge Andrés Jaimes",
        "Desarrollador / Controlador de Obra",
        firma_cfg={
            "elaboro_nombre": "Ej. Elaboró",
            "elaboro_cargo": "Cargo",
            "reviso_nombre": "Ej. Revisó",
            "reviso_cargo": "Cargo",
        },
    )
    pdf_bytes = _to_pdf(html)
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": "inline; filename=preview_CC-SUB-001.pdf"})

# ── Utilidades ─────────────────────────────────────────────────────────────────

_TO_PDF_LOCK = threading.Lock()


def _to_pdf(html: str) -> bytes:
    """Genera PDF. xhtml2pdf a veces marca `err` por advertencias aun con salida válida."""
    with _TO_PDF_LOCK:
        return _to_pdf_unlocked(html)


def _to_pdf_unlocked(html: str) -> bytes:
    from fo_eo_04_pdf_worker import _pdf_link_callback

    buf = io.BytesIO()
    # StringIO + caracteres raros en Windows puede fallar; UTF-8 explícito reduce errores 500.
    src = io.BytesIO(html.encode("utf-8", errors="replace"))
    result = pisa.CreatePDF(
        src,
        dest=buf,
        encoding="utf-8",
        link_callback=_pdf_link_callback,
    )
    buf.seek(0)
    out = buf.read()
    if not out:
        raise ValueError("xhtml2pdf no produjo bytes (PDF vacío)")
    if getattr(result, "err", 0):
        _log.warning("xhtml2pdf reportó advertencias (err=%s); se devuelve PDF de %s bytes", result.err, len(out))
    return out


def _to_pdf_corte_garantizado(html: str) -> bytes:
    """Intenta PDF; si falla, HTML mínimo de una línea (siempre debe producir bytes)."""
    try:
        return _to_pdf(html)
    except Exception as e:
        _log.warning("pisa falló primer HTML (%s); intento mínimo", e)
        try:
            return _to_pdf(
                "<!DOCTYPE html><html><head><meta charset=\"UTF-8\"/></head><body>"
                "<p style=\"font-family:Arial\">ClaraCore: el informe no pudo renderizarse con la plantilla. "
                "Contacte soporte o reintente.</p></body></html>"
            )
        except Exception as e2:
            _log.exception("pisa falló incluso HTML mínimo: %s", e2)
            raise RuntimeError(f"pisa: {e!s} | minimo: {e2!s}") from e2

def _fd(d):
    """Formatea fecha ISO → dd/mm/yyyy."""
    if not d: return "—"
    try:    return datetime.fromisoformat(str(d)).strftime("%d/%m/%Y")
    except: return str(d)

def _fn(n, dec=2):
    """Formatea número con decimales (inf/nan no rompen el PDF: format ',.' falla con inf)."""
    if n is None: return "—"
    try:
        x = float(n)
        if math.isnan(x):
            return "—"
        if math.isinf(x):
            return "> max" if x > 0 else "< min"
        return f"{x:,.{dec}f}"
    except Exception:
        return str(n)

def _sf(n, default=0.0):
    """Convierte a float sin romper el endpoint (strings, comas, vacíos)."""
    if n is None or n == "":
        return float(default)
    try:
        return float(n)
    except Exception:
        try:
            return float(str(n).replace(",", "").replace(" ", "").strip())
        except Exception:
            return float(default)

def _fm(n):
    """Formatea como moneda colombiana (evita ValueError con inf en f'{x:,.0f}')."""
    if n is None: return "—"
    try:
        x = float(n)
        if math.isnan(x):
            return "—"
        if math.isinf(x):
            return "$ (valor fuera de rango)"
        return f"$ {x:,.0f}"
    except Exception:
        return str(n)

def _h(v):
    """Escape de caracteres especiales para HTML/PDF."""
    return html.escape("" if v is None else str(v), quote=True)


def _descripcion_memoria_compacta(s: object) -> str:
    """Descripción u observación en CC-SUB-002: minúsculas para PDF más compacto (la BD suele venir en mayúsculas)."""
    t = "" if s is None else str(s).strip()
    return t.lower()

# ── CSS base (compatible xhtml2pdf) ────────────────────────────────────────────
# CC-SUB-002 NO debe concatenar PAGE_CSS_PORTRAIT: xhtml2pdf aplica el primer @page a
# todas las hojas; un segundo @page no cambia orientación en páginas siguientes → queda
# vertical rota / columnas aplastadas. Memoria usa solo BASE_CSS_SHARED + MEMORIA002_CSS.

PAGE_CSS_LETTER_PORTRAIT = """
@page {
    size: letter;
    margin: 1.2cm 1.2cm 1.4cm 1.2cm;
}
"""

BASE_CSS_SHARED = """
body { font-family: Arial, sans-serif; font-size: 8pt; color: #1a1a2e; }
table { border-collapse: collapse; }
.w100 { width: 100%; }
.hdr-outer { width: 100%; border: 1px solid #9ca3af; margin-bottom: 6px; }
.hdr-logo  { width: 150px; border-right: 1px solid #9ca3af; padding: 6px; text-align: center; vertical-align: middle; }
.hdr-logo img { max-width: 138px; max-height: 48px; }
.hdr-main  { padding: 5px 8px; vertical-align: middle; text-align: center; }
.doc-title { font-size: 10.5pt; font-weight: bold; color: #111827; text-transform: uppercase; }
.doc-code  { font-size: 10pt; font-weight: bold; color: #374151; margin: 2px 0 4px 0; text-align: center; }
.lbl { color: #555; font-weight: normal; }
.val { font-weight: bold; color: #1a1a2e; }
.section-bar {
    background: #e5e7eb; color: #111827;
    font-size: 8pt; font-weight: bold;
    border: 1px solid #9ca3af;
    padding: 3px 8px; margin: 6px 0 3px 0;
}
.data-th {
    background: #f3f4f6; color: #111827;
    font-weight: bold; padding: 4px 5px;
    border: 1px solid #9ca3af; text-align: center;
    font-size: 7.5pt;
}
.data-td {
    padding: 3px 5px;
    border: 1px solid #e2e8f0;
    font-size: 7.5pt;
    vertical-align: middle;
}
.even { background: #f8fafc; }
.total-td {
    background: #e5e7eb; color: #111827;
    font-weight: bold; padding: 4px 5px;
    border: 1px solid #9ca3af; font-size: 7.5pt;
}
.firma-linea { border-top: 1px solid #333; margin-top: 55px; margin-bottom: 4px; }
.firma-nombre { font-weight: bold; font-size: 8pt; }
.firma-cargo  { font-size: 7pt; color: #555; }
.firma-dato   { font-size: 7pt; color: #333; }
.foto-caption { font-size: 7pt; color: #555; margin-top: 3px; text-align: center; }
.sep { border: none; border-top: 1px solid #e2e8f0; margin: 4px 0; }
.doc-meta td { border: 1px solid #9ca3af; padding: 3px 5px; font-size: 7.4pt; vertical-align: top; }
.doc-footer {
    margin-top: 8px;
    border-top: 1px solid #9ca3af;
    padding-top: 4px;
    font-size: 6.7pt;
    color: #4b5563;
    text-align: center;
}
"""

BASE_CSS = PAGE_CSS_LETTER_PORTRAIT + BASE_CSS_SHARED

# CC-SUB-002: un único @page (landscape) + estilos; ir junto a BASE_CSS_SHARED, sin PAGE_CSS_PORTRAIT.
MEMORIA002_CSS = """
@page {
  size: letter landscape;
  margin: 0.85cm 0.8cm 0.95cm 0.8cm;
}
body.mem002-doc {
  font-family: Arial, sans-serif;
  font-size: 6pt;
  color: #1a1a2e;
}
body.mem002-doc .mem002-head-wrap {
  width: 100%;
  page-break-inside: avoid;
  page-break-after: avoid;
}
body.mem002-doc .mem002-section {
  page-break-after: avoid;
  margin-top: 1px;
  margin-bottom: 2px;
  font-size: 7pt;
  padding: 2px 6px;
}
body.mem002-doc .mem002-detail {
  width: 100%;
  border-collapse: collapse;
  /* auto: xhtml2pdf + colgroup/table-layout:fixed aplastaba OBSERVACIÓN; anchos en <th> */
  table-layout: auto;
}
body.mem002-doc .mem002-detail th.data-th {
  padding: 2px 2px;
  font-size: 6pt;
  line-height: 1.1;
}
body.mem002-doc .mem002-detail td.data-td {
  padding: 1px 2px;
  font-size: 6pt;
  line-height: 1.12;
  min-height: 0.26cm;
  vertical-align: middle;
}
body.mem002-doc .mem002-detail td.mem002-obs {
  vertical-align: top;
  word-wrap: break-word;
  overflow-wrap: break-word;
}
body.mem002-doc .mem002-detail .total-td {
  font-size: 6pt;
  padding: 2px 3px;
}
/* Totales fuera de .mem002-detail: evita colspan que en xhtml2pdf deforma la última hoja del detalle */
body.mem002-doc .mem002-total-wrap {
  width: 100%;
  border-collapse: collapse;
  table-layout: fixed;
}
body.mem002-doc .mem002-total-wrap td.total-td {
  font-size: 6pt;
  padding: 2px 3px;
}
body.mem002-doc .doc-footer {
  margin-top: 4px;
  padding-top: 2px;
  font-size: 6pt;
}
body.mem002-doc .mem002-foto-grid td { vertical-align: top; }
/* Bloque firmas (misma lógica que CC-SUB-001: Elaboró / Revisó / Aprobó subcontratista) */
body.mem002-doc .mem002-firmas-wrap {
  margin-top: 2mm;
  page-break-inside: avoid;
}
body.mem002-doc .mem002-firmas-tbl { width: 100%; border-collapse: collapse; table-layout: fixed; }
body.mem002-doc .mem002-firmas-tbl td {
  vertical-align: top !important;
  padding: 1px 3px !important;
}
body.mem002-doc .mem002-firma-slot-hdr {
  font-weight: bold;
  font-size: 6pt;
  margin: 0;
  padding: 0;
  color: #111;
  line-height: 1;
}
body.mem002-doc .mem002-firma-slot-body {
  height: 0.45cm;
  min-height: 0.45cm;
  max-height: 0.45cm;
  overflow: hidden;
  box-sizing: border-box;
  padding: 0;
  margin: 0;
}
body.mem002-doc .mem002-firma-line {
  border: none;
  border-top: 1px solid #111;
  width: 100%;
  height: 0;
  line-height: 0;
  font-size: 0;
  margin: 0 0 1px 0;
  padding: 0;
  overflow: hidden;
}
body.mem002-doc .mem002-firma-nombre {
  font-weight: bold;
  font-size: 6.5pt;
  line-height: 1.05;
  margin: 0;
  padding: 0;
  word-wrap: break-word;
}
body.mem002-doc .mem002-firma-cargo {
  font-size: 6pt;
  color: #444;
  line-height: 1.05;
  margin: 0;
  padding: 0;
  word-wrap: break-word;
}
body.mem002-doc .mem002-firma-rep {
  font-size: 6pt;
  color: #333;
  line-height: 1.05;
  margin: 0;
  padding: 0;
  word-wrap: break-word;
}
/* _html_cc_sub_td_firma_columna usa clases ccd-firma-* (no mem002-*). CC-SUB-001 incrusta estas
   reglas en su <style>; sin ellas xhtml2pdf estira .ccd-firma-slot-body / la línea y toda la fila
   de firmas queda desproporcionada (hueco entre imagen y nombre en CC-SEM/CC-MES memorias). */
body.mem002-doc .mem002-firmas-wrap .ccd-firma-slot-hdr {
  font-weight: bold;
  font-size: 6pt;
  margin: 0;
  padding: 0;
  color: #111;
  line-height: 1;
}
body.mem002-doc .mem002-firmas-wrap .ccd-firma-slot-body {
  /* Columnas sin imagen: más bajo que CC-SUB-001 para no estirar la fila frente a la columna con firma */
  height: 0.45cm;
  min-height: 0.45cm;
  max-height: 0.45cm;
  overflow: hidden;
  box-sizing: border-box;
  padding: 0;
  margin: 0;
}
body.mem002-doc .mem002-firmas-wrap .ccd-firma-line {
  border: none;
  border-top: 1px solid #111;
  width: 100%;
  height: 0;
  line-height: 0;
  font-size: 0;
  margin: 0 0 1px 0;
  padding: 0;
  overflow: hidden;
}
/* Memorias: jaula más baja que CC-SUB-001 (~22pt); !important gana estilos inline del HTML. */
body.mem002-doc .mem002-firmas-wrap table.ccd-firma-img-cage {
  height: 22pt !important;
  max-height: 22pt !important;
  margin: 0 !important;
  overflow: hidden !important;
}
body.mem002-doc .mem002-firmas-wrap .ccd-firma-img-cage td {
  overflow: hidden !important;
  padding: 0 !important;
  height: 22pt !important;
  max-height: 22pt !important;
  line-height: 0 !important;
  vertical-align: middle !important;
}
body.mem002-doc .mem002-firmas-wrap .ccd-firma-img-cage img {
  max-width: 100% !important;
  height: 20pt !important;
  max-height: 20pt !important;
  width: auto !important;
  display: block !important;
  margin: 0 auto !important;
}
body.mem002-doc .mem002-firmas-wrap .ccd-firma-nombre {
  word-wrap: break-word;
  font-weight: bold;
  font-size: 6.5pt;
  line-height: 1.05;
  margin: 0 !important;
  padding: 0;
}
body.mem002-doc .mem002-firmas-wrap .ccd-firma-cargo {
  word-wrap: break-word;
  font-size: 6pt;
  color: #444;
  line-height: 1.05;
  margin: 0 !important;
  padding: 0;
}
"""

# ── Template CC-SUB-001 (réplica maqueta institucional; sin <pdf:nextpage/>; tablas + estilos inline) ──

def _fmt_informe_fecha_generacion() -> str:
    """Ej.: 15 Abr 26, 02:04 pm (alineado a la maqueta)."""
    n = datetime.now(pytz.timezone("America/Bogota"))
    meses = ("Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic")
    h12 = n.strftime("%I").lstrip("0") or "12"
    mi = n.strftime("%M")
    ap = n.strftime("%p").lower()
    return f"{n.day:02d} {meses[n.month - 1]} {n.year % 100:02d}, {h12}:{mi} {ap}"


def _sello_verificado_por(usuario_nombre: str) -> str:
    nom = (usuario_nombre or "—").strip().upper()
    n = datetime.now(pytz.timezone("America/Bogota"))
    return f"Verificado y aprobado por: {nom} {n.strftime('%Y.%m.%d')} - {n.strftime('%H:%M:%S')}"


def _corte_consecutivo_fmt(corte: dict) -> str:
    co = corte.get("consecutivo")
    if co is None or co == "":
        return "—"
    try:
        return f"{int(float(str(co))):02d}"
    except Exception:
        return str(co)


def _excel_sheet_name(item_numero: object) -> str:
    """Nombre de hoja Excel (máx. 31, sin caracteres prohibidos)."""
    raw = str(item_numero if item_numero is not None else "Item").strip() or "Item"
    t = re.sub(r"[\[\]\*\?\/\\:]", "_", raw)[:31]
    return t or "Item"


def _excel_unique_sheet_name(wb: Workbook, base: str) -> str:
    b = _excel_sheet_name(base)
    if b not in wb.sheetnames:
        return b
    for i in range(2, 500):
        suf = f"_{i}"
        t = (b[: 31 - len(suf)] + suf)[:31]
        if t not in wb.sheetnames:
            return t
    return f"H{len(wb.sheetnames)}"[:31]


def _fill_memoria_excel_ws(
    ws,
    contrato: dict,
    sub: dict,
    corte: dict,
    item_info: dict,
    registros: List[dict],
    firma_cfg: Optional[Dict[str, Any]],
    *,
    conc_meta: Optional[Dict[str, Any]] = None,
    pie_fotos_contexto: Optional[str] = None,
    aprobo_interventoria_desde_config: bool = False,
) -> None:
    """Hoja CC-SUB-002 / CC-SEM-002 / CC-MES-002: encabezado, detalle, total, fotos, firmas."""
    fc = firma_cfg or {}
    elaboro_n = str(fc.get("elaboro_nombre") or "").strip() or "—"
    elaboro_c = str(fc.get("elaboro_cargo") or "").strip() or "—"
    reviso_n = str(fc.get("reviso_nombre") or "").strip() or "—"
    reviso_c = str(fc.get("reviso_cargo") or "").strip() or "—"
    if aprobo_interventoria_desde_config:
        aprobo_emp = str(fc.get("aprobo_nombre") or "").strip() or "—"
        aprobo_rep = str(fc.get("aprobo_cargo") or "").strip() or "—"
    else:
        aprobo_emp = str(sub.get("razon_social") or "").strip() or "—"
        aprobo_rep = str(sub.get("nombre_contacto") or "").strip() or "—"

    thin = Side(style="thin", color="9CA3AF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_bar = PatternFill("solid", fgColor="E5E7EB")
    fill_th = PatternFill("solid", fgColor="F3F4F6")
    fill_tot = PatternFill("solid", fgColor="E5E7EB")

    for i, w in enumerate([5, 9, 9, 8, 10, 8, 8, 8, 9, 10, 36], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_cant = sum(_sf(r.get("cantidad_total"), 0.0) for r in registros)
    corte_lbl = _corte_consecutivo_fmt(corte)
    periodo = f"{_fd(corte.get('fecha_inicio'))} — {_fd(corte.get('fecha_fin'))}"

    titulo_h = "RESUMEN ACTIVIDADES CONCILIACIÓN CORTE SUBCONTRATISTA"
    codigo_h = "CC-SUB-002"
    if conc_meta:
        titulo_h = str(conc_meta.get("titulo") or titulo_h)
        codigo_h = str(conc_meta.get("codigo") or codigo_h)

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = titulo_h
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A2:K2")
    ws["A2"] = f"{codigo_h} · CCD · Memoria por ítem"
    ws["A2"].font = Font(bold=True, size=10, color="1E40AF")
    ws["A2"].alignment = Alignment(horizontal="center")

    if conc_meta:
        cells = list(conc_meta.get("cells") or [])
        while len(cells) < 4:
            cells.append(("", ""))
        ws["A3"] = str(cells[0][0] or "")
        ws["A3"].font = Font(size=8, color="64748B")
        ws.merge_cells("B3:E3")
        ws["B3"] = str(cells[0][1] or "—")
        ws["F3"] = str(cells[1][0] or "")
        ws["F3"].font = Font(size=8, color="64748B")
        ws.merge_cells("G3:K3")
        ws["G3"] = str(cells[1][1] or "—")
        ws["A4"] = str(cells[2][0] or "")
        ws["A4"].font = Font(size=8, color="64748B")
        ws.merge_cells("B4:E4")
        ws["B4"] = str(cells[2][1] or "—")
        ws["F4"] = str(cells[3][0] or "")
        ws["F4"].font = Font(size=8, color="64748B")
        ws.merge_cells("G4:K4")
        ws["G4"] = str(cells[3][1] or "—")
        for rr in (3, 4):
            for cc in (1, 2, 6, 7):
                ws.cell(row=rr, column=cc).border = bd
        ws["A5"], ws["B5"] = "ÍTEM", str(item_info.get("item_numero") or "—")
        ws.merge_cells("B5:C5")
        ws["D5"] = "DESCRIPCIÓN"
        ws["D5"].font = Font(size=8, color="64748B")
        ws.merge_cells("E5:I5")
        ws["E5"] = _descripcion_memoria_compacta(item_info.get("item_descripcion"))
        ws["E5"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["J5"] = "UNIDAD"
        ws["J5"].font = Font(size=8, color="64748B")
        ws["K5"] = str(item_info.get("unidad") or "—")
        for cc in (1, 2, 4, 5, 10, 11):
            ws.cell(row=5, column=cc).border = bd
        r0 = 7
    else:
        ws["A3"], ws["B3"] = "CONTRATO", str(contrato.get("numero") or "—")
        ws.merge_cells("B3:E3")
        ws["F3"], ws["G3"] = "SUB CONTRATISTA", str(sub.get("razon_social") or "—")
        ws.merge_cells("G3:K3")
        for x in ("A3", "F3"):
            ws[x].font = Font(size=8, color="64748B")
            ws[x].alignment = Alignment(vertical="top")

        ws["A4"], ws["B4"] = "CORTE N°", corte_lbl
        ws.merge_cells("B4:E4")
        ws["F4"], ws["G4"] = "PERÍODO", periodo
        ws.merge_cells("G4:K4")
        for x in ("A4", "F4"):
            ws[x].font = Font(size=8, color="64748B")

        ws["A5"], ws["B5"] = "ÍTEM", str(item_info.get("item_numero") or "—")
        ws.merge_cells("B5:C5")
        ws["D5"] = "DESCRIPCIÓN"
        ws["D5"].font = Font(size=8, color="64748B")
        ws.merge_cells("E5:I5")
        ws["E5"] = _descripcion_memoria_compacta(item_info.get("item_descripcion"))
        ws["E5"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["J5"] = "UNIDAD"
        ws["J5"].font = Font(size=8, color="64748B")
        ws["K5"] = str(item_info.get("unidad") or "—")
        for rr in (3, 4):
            for cc in (1, 2, 6, 7):
                ws.cell(row=rr, column=cc).border = bd
        for cc in (1, 2, 4, 5, 10, 11):
            ws.cell(row=5, column=cc).border = bd
        r0 = 7
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=11)
    bar = ws.cell(row=r0, column=1, value="DETALLE DE CANTIDADES APROBADAS")
    bar.fill = fill_bar
    bar.font = Font(bold=True, size=9)
    bar.alignment = Alignment(horizontal="center", vertical="center")
    bar.border = bd

    hdr = [
        "N°",
        "ABS INI",
        "ABS FIN",
        "PK ID",
        "COSTADO",
        "LONG",
        "ANCHO",
        "ESP",
        "CANT",
        "CANT TOT",
        "OBSERVACIÓN",
    ]
    hr = r0 + 1
    for col, h in enumerate(hdr, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.fill = fill_th
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bd

    data_row = hr + 1
    for i, r in enumerate(registros):
        obs = r.get("observacion") or ""
        fn = r.get("foto_numero")
        if fn:
            obs = f"{obs} [Foto {fn}]".strip()
        obs = _descripcion_memoria_compacta(obs)
        pkv = (r.get("pk_ids") or {}).get("pk_id")
        row = data_row + i
        vals = [
            r.get("numero_registro"),
            r.get("abs_inicio") or "—",
            r.get("abs_final") or "—",
            pkv if pkv is not None else "—",
            r.get("calzada") or "—",
            _fn(r.get("longitud")),
            _fn(r.get("ancho")),
            _fn(r.get("espesor")),
            _fn(r.get("cantidad")),
            _fn(r.get("cantidad_total")),
            (obs or "")[:500],
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = bd
            cell.font = Font(size=8)
            if col in (6, 7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col == 11:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if i % 2 == 0:
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F8FAFC")

    tot_r = data_row + len(registros)
    ws.merge_cells(start_row=tot_r, start_column=1, end_row=tot_r, end_column=9)
    ws.cell(row=tot_r, column=1, value="CANTIDAD TOTAL DEL ÍTEM").alignment = Alignment(
        horizontal="right", vertical="center"
    )
    ws.cell(row=tot_r, column=1).font = Font(bold=True, size=9)
    c_tot = ws.cell(row=tot_r, column=10, value=_fn(total_cant))
    c_tot.font = Font(bold=True, size=9)
    c_tot.alignment = Alignment(horizontal="right")
    c_tot.border = bd
    c_tot.fill = fill_tot
    c_lab = ws.cell(row=tot_r, column=1)
    c_lab.border = bd
    c_lab.fill = fill_tot
    c_e = ws.cell(row=tot_r, column=11)
    c_e.border = bd
    c_e.fill = fill_tot

    fotos = [r for r in registros if (r.get("foto_url") or "").strip()]
    fr = tot_r + 2
    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=11)
    cap_foto = f"REGISTRO FOTOGRÁFICO — ÍTEM {item_info.get('item_numero', '')} | Corte N° {corte.get('consecutivo', '')}"
    if conc_meta and (pie_fotos_contexto or "").strip():
        cap_foto = f"REGISTRO FOTOGRÁFICO — ÍTEM {item_info.get('item_numero', '')} | {pie_fotos_contexto}"
    fb = ws.cell(row=fr, column=1, value=cap_foto)
    fb.fill = fill_bar
    fb.font = Font(bold=True, size=9)
    fb.alignment = Alignment(horizontal="center", vertical="center")
    fb.border = bd

    fhdr = fr + 1
    ws.cell(row=fhdr, column=1, value="Reg.").fill = fill_th
    ws.cell(row=fhdr, column=1).font = Font(bold=True, size=8)
    ws.cell(row=fhdr, column=1).border = bd
    ws.cell(row=fhdr, column=2, value="Foto N°").fill = fill_th
    ws.cell(row=fhdr, column=2).font = Font(bold=True, size=8)
    ws.cell(row=fhdr, column=2).border = bd
    ws.merge_cells(start_row=fhdr, start_column=3, end_row=fhdr, end_column=8)
    c3 = ws.cell(row=fhdr, column=3, value="Enlace / URL")
    c3.fill = fill_th
    c3.font = Font(bold=True, size=8)
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.border = bd
    ws.merge_cells(start_row=fhdr, start_column=9, end_row=fhdr, end_column=11)
    c9 = ws.cell(row=fhdr, column=9, value="Observación")
    c9.fill = fill_th
    c9.font = Font(bold=True, size=8)
    c9.alignment = Alignment(horizontal="center", vertical="center")
    c9.border = bd

    pr = fhdr + 1
    if not fotos:
        ws.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=11)
        sf = ws.cell(row=pr, column=1, value="Sin fotos en este ítem.")
        sf.font = Font(size=8, italic=True, color="64748B")
        sf.alignment = Alignment(horizontal="left")
        sf.border = bd
        pr += 1
    else:
        for r in fotos:
            fu = (r.get("foto_url") or "").strip()
            obs_f = _descripcion_memoria_compacta((r.get("observacion") or "")[:300])
            ws.cell(row=pr, column=1, value=r.get("numero_registro")).border = bd
            ws.cell(row=pr, column=2, value=r.get("foto_numero")).border = bd
            ws.merge_cells(start_row=pr, start_column=3, end_row=pr, end_column=8)
            fu_cell = fu if len(fu) < 8000 else fu[:7990] + "…"
            lc = ws.cell(row=pr, column=3, value=fu_cell)
            lc.font = Font(size=8, color="0563C1", underline="single")
            if fu.startswith(("http://", "https://")):
                lc.hyperlink = fu
            lc.alignment = Alignment(wrap_text=True, vertical="top")
            lc.border = bd
            ws.merge_cells(start_row=pr, start_column=9, end_row=pr, end_column=11)
            oc = ws.cell(row=pr, column=9, value=obs_f)
            oc.font = Font(size=8)
            oc.alignment = Alignment(wrap_text=True, vertical="top")
            oc.border = bd
            pr += 1

    sr = pr + 1
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=11)
    ws.cell(row=sr, column=1, value="FIRMAS").fill = fill_bar
    ws.cell(row=sr, column=1).font = Font(bold=True, size=9)
    ws.cell(row=sr, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=sr, column=1).border = bd

    fr0 = sr + 1
    ws.cell(row=fr0, column=1, value="Elaboró:").font = Font(size=8, bold=True)
    ws.merge_cells(start_row=fr0, start_column=2, end_row=fr0, end_column=4)
    ws.cell(row=fr0, column=2, value=f"{elaboro_n}\n{elaboro_c}").alignment = Alignment(wrap_text=True)
    ws.cell(row=fr0, column=2).font = Font(size=8)
    ws.cell(row=fr0, column=5, value="Revisó:").font = Font(size=8, bold=True)
    ws.merge_cells(start_row=fr0, start_column=6, end_row=fr0, end_column=8)
    ws.cell(row=fr0, column=6, value=f"{reviso_n}\n{reviso_c}").alignment = Alignment(wrap_text=True)
    ws.cell(row=fr0, column=6).font = Font(size=8)
    lbl_apro = "Aprobó (interventoría):" if aprobo_interventoria_desde_config else "Aprobó (subcontratista):"
    txt_apro = f"{aprobo_emp}\n{aprobo_rep}" if aprobo_interventoria_desde_config else f"{aprobo_emp}\nRepresentante: {aprobo_rep}"
    ws.cell(row=fr0, column=9, value=lbl_apro).font = Font(size=8, bold=True)
    ws.merge_cells(start_row=fr0, start_column=10, end_row=fr0, end_column=11)
    ws.cell(row=fr0, column=10, value=txt_apro).alignment = Alignment(wrap_text=True)
    ws.cell(row=fr0, column=10).font = Font(size=8)
    for cc in (1, 2, 5, 6, 9, 10):
        ws.cell(row=fr0, column=cc).border = bd

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.print_options.horizontalCentered = True


def _memoria_item_excel_bytes(
    contrato: dict,
    sub: dict,
    corte: dict,
    item_info: dict,
    registros: List[dict],
    firma_cfg: Optional[Dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _excel_sheet_name(item_info.get("item_numero"))
    _fill_memoria_excel_ws(ws, contrato, sub, corte, item_info, registros, firma_cfg)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _memoria_corte_completo_excel_bytes(
    contrato_id: int,
    corte_id: int,
    numeros: List[str],
    current_user: dict,
    firma_cfg: Optional[Dict[str, Any]],
) -> bytes:
    wb = Workbook()
    first = True
    for item_numero in numeros:
        ctx = _contexto_memoria_item(
            contrato_id, corte_id, item_numero, current_user, item_exacto=True
        )
        contrato = ctx["contrato"]
        sub = ctx["sub"]
        corte = ctx["corte"]
        item_info = ctx["item_info"]
        registros = ctx["registros"]
        if first:
            ws = wb.active
            assert ws is not None
            ws.title = _excel_unique_sheet_name(wb, item_info.get("item_numero"))
            _fill_memoria_excel_ws(ws, contrato, sub, corte, item_info, registros, firma_cfg)
            first = False
        else:
            ws = wb.create_sheet(title=_excel_unique_sheet_name(wb, item_info.get("item_numero")))
            _fill_memoria_excel_ws(ws, contrato, sub, corte, item_info, registros, firma_cfg)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cc_sem_002_item_excel_bytes(
    contrato_id: int,
    semana_id: int,
    item_numero: str,
    current_user: dict,
) -> bytes:
    """Excel CC-SEM-002 por ítem: mismo contenido lógico que el PDF semanal."""
    if not _semana_pertenece_contrato(contrato_id, semana_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    registros = fetch_registros_memoria_conciliacion(
        _sb, contrato_id, item_numero, semana_id=semana_id, item_exacto=True
    )
    if not registros:
        raise HTTPException(404, "No hay registros para este ítem y semana")
    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    item_info = {
        "item_numero": registros[0].get("item_numero", item_numero),
        "item_descripcion": registros[0].get("item_descripcion", ""),
        "unidad": registros[0].get("unidad", ""),
    }
    sm = _row("so_semanas", "numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
    nsem = sm.get("numero_semana")
    fi = str(sm.get("fecha_inicio") or "—")
    ff = str(sm.get("fecha_fin") or "—")
    conc_meta = {
        "titulo": "RESUMEN ACTIVIDADES — CONCILIACIÓN SEMANAL (INTERVENTORÍA–CONTRATISTA)",
        "codigo": CODIGO_FORMATO_CCD_CC_SEM_002,
        "cells": [
            ("CONTRATO", str(contrato.get("numero") or "")),
            ("SEMANA", f"N° {nsem}"),
            ("VIGENCIA", f"{fi} — {ff}"),
            ("REFERENCIA", "Cantidades ejecutadas — conciliación"),
        ],
    }
    fmt = CODIGO_FORMATO_CCD_CC_SEM_002
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmt, contexto_tipo="semana", contexto_id=semana_id)
    sub, corte = _sub_corte_dummy_memoria()
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _excel_sheet_name(item_info.get("item_numero"))
    _fill_memoria_excel_ws(
        ws,
        contrato,
        sub,
        corte,
        item_info,
        registros,
        firma_cfg,
        conc_meta=conc_meta,
        pie_fotos_contexto=f"Semana N° {nsem} · {fi} — {ff}",
        aprobo_interventoria_desde_config=True,
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cc_sem_002_semana_completo_excel_bytes(contrato_id: int, semana_id: int, current_user: dict) -> bytes:
    """Excel CC-SEM-002: una hoja por ítem (misma lógica que PDF completo)."""
    if not _semana_pertenece_contrato(contrato_id, semana_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    reg_agg = fetch_registros_conciliacion(_sb, contrato_id, semana_id=semana_id)
    items_agg, _total = aggregate_items_conciliacion(reg_agg)
    _sort_items_corte_por_item_numero_asc(items_agg)
    numeros = [
        (it.get("item_numero") or "").strip()
        for it in items_agg
        if (it.get("item_numero") or "").strip()
    ]
    if not numeros:
        raise HTTPException(404, "No hay registros en esta semana para generar memorias")

    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    sm = _row("so_semanas", "numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
    nsem = sm.get("numero_semana")
    fi = str(sm.get("fecha_inicio") or "—")
    ff = str(sm.get("fecha_fin") or "—")
    conc_meta = {
        "titulo": "RESUMEN ACTIVIDADES — CONCILIACIÓN SEMANAL (INTERVENTORÍA–CONTRATISTA)",
        "codigo": CODIGO_FORMATO_CCD_CC_SEM_002,
        "cells": [
            ("CONTRATO", str(contrato.get("numero") or "")),
            ("SEMANA", f"N° {nsem}"),
            ("VIGENCIA", f"{fi} — {ff}"),
            ("REFERENCIA", "Cantidades ejecutadas — conciliación"),
        ],
    }
    fmt = CODIGO_FORMATO_CCD_CC_SEM_002
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, fmt, contexto_tipo="semana", contexto_id=semana_id)
    sub, corte = _sub_corte_dummy_memoria()
    pie = f"Semana N° {nsem} · {fi} — {ff}"

    wb = Workbook()
    first = True
    for inum in numeros:
        registros = fetch_registros_memoria_conciliacion(
            _sb, contrato_id, inum, semana_id=semana_id, item_exacto=True
        )
        if not registros:
            continue
        item_info = {
            "item_numero": registros[0].get("item_numero", inum),
            "item_descripcion": registros[0].get("item_descripcion", ""),
            "unidad": registros[0].get("unidad", ""),
        }
        if first:
            ws = wb.active
            assert ws is not None
            ws.title = _excel_unique_sheet_name(wb, item_info.get("item_numero"))
            _fill_memoria_excel_ws(
                ws,
                contrato,
                sub,
                corte,
                item_info,
                registros,
                firma_cfg,
                conc_meta=conc_meta,
                pie_fotos_contexto=pie,
                aprobo_interventoria_desde_config=True,
            )
            first = False
        else:
            ws = wb.create_sheet(title=_excel_unique_sheet_name(wb, item_info.get("item_numero")))
            _fill_memoria_excel_ws(
                ws,
                contrato,
                sub,
                corte,
                item_info,
                registros,
                firma_cfg,
                conc_meta=conc_meta,
                pie_fotos_contexto=pie,
                aprobo_interventoria_desde_config=True,
            )
    if first:
        raise HTTPException(404, "No hay registros aprobados por ítem en esta semana")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ccd_hex_to_excel_rgb(h: str, default: str = "E8E8E8") -> str:
    t = _sanitize_ccd_hex_color(h, "#" + default).lstrip("#")
    return t if len(t) == 6 else default


def _fill_corte_sub_001_excel_ws(
    ws,
    contrato: dict,
    sub: dict,
    corte: dict,
    items: List[dict],
    total_costo: float,
    usuario_nombre: str,
    usuario_cargo: str,
    firma_cfg: Optional[Dict[str, Any]],
) -> None:
    """Hoja CC-SUB-001: encabezado, tabla de ítems (como PDF), subtotal, firmas, pie."""
    fc = firma_cfg or {}
    est = _merge_estilo_pdf(fc.get("estilo_pdf"), CODIGO_FORMATO_CCD_CC_SUB_001)
    thead_bg = _ccd_hex_to_excel_rgb(est.get("thead_bg"), "E8E8E8")
    row_even = _ccd_hex_to_excel_rgb(est.get("row_even_bg"), "F8FAFC")
    row_odd = _ccd_hex_to_excel_rgb(est.get("row_odd_bg"), "FFFFFF")
    subtotal_bg = _ccd_hex_to_excel_rgb(est.get("subtotal_bg"), "DBEAFE")

    elaboro_n = str(fc.get("elaboro_nombre") or "").strip() or "—"
    elaboro_c = str(fc.get("elaboro_cargo") or "").strip() or "—"
    reviso_n = str(fc.get("reviso_nombre") or "").strip() or "—"
    reviso_c = str(fc.get("reviso_cargo") or "").strip() or "—"
    aprobo_emp = str(sub.get("razon_social") or "").strip() or "—"
    aprobo_rep = str(sub.get("nombre_contacto") or "").strip() or "—"

    thin = Side(style="thin", color="9CA3AF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_bar = PatternFill("solid", fgColor="E5E7EB")
    fill_th = PatternFill("solid", fgColor=thead_bg)

    for i, w in enumerate([10, 9, 36, 8, 12, 11, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def meta_row(r: int, label: str, value: object) -> None:
        ws.cell(row=r, column=1, value=label).font = Font(size=8, bold=True, color="374151")
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        v = ws.cell(row=r, column=2, value=value if value is not None else "—")
        v.font = Font(size=8)
        v.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=r, column=1).border = bd
        v.border = bd

    fecha_gen = _fmt_informe_fecha_generacion()
    corte_lbl = _corte_consecutivo_fmt(corte)
    nit_raw = str(contrato.get("nit") or "").strip()
    contratista_val = str(contrato.get("contratista") or "—")
    if nit_raw:
        contratista_val = f"{contratista_val} (NIT: {nit_raw})"
    periodo = f"{_fd(corte.get('fecha_inicio'))} — {_fd(corte.get('fecha_fin'))}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    t1 = ws.cell(row=1, column=1, value="INFORME CORTE DE SUB CONTRATISTA")
    t1.font = Font(bold=True, size=12)
    t1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    t1.border = bd

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    t2 = ws.cell(row=2, column=1, value=f"{CODIGO_FORMATO_CCD_CC_SUB_001} · CCD · ClaraCore")
    t2.font = Font(bold=True, size=10, color="1E40AF")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    t2.border = bd

    meta_row(3, "CONTRATO", str(contrato.get("numero") or "—"))
    meta_row(4, "FECHA", fecha_gen)
    meta_row(5, "SUB CONTRATISTA", str(sub.get("razon_social") or "—"))
    meta_row(6, "CORTE N°", corte_lbl)
    meta_row(7, "CONTRATISTA", contratista_val)
    meta_row(8, "INTERVENTORÍA", str(contrato.get("interventoria") or "—"))
    meta_row(9, "PERÍODO DEL CORTE", periodo)

    r0 = 11
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=7)
    br = ws.cell(row=r0, column=1, value="CANTIDADES APROBADAS POR ÍTEM")
    br.fill = fill_bar
    br.font = Font(bold=True, size=9)
    br.alignment = Alignment(horizontal="center", vertical="center")
    br.border = bd

    hdr = [
        "CAPÍTULO",
        "ÍTEM",
        "DESCRIPCIÓN",
        "UNIDAD",
        "VALOR UNIT.",
        "CANTIDAD",
        "COSTO DIRECTO",
    ]
    hr = r0 + 1
    for col, h in enumerate(hdr, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.fill = fill_th
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bd

    data0 = hr + 1
    item_list = list(items or [])
    for idx, it in enumerate(item_list):
        row = data0 + idx
        cap = (it.get("capitulo") or "").strip() or "—"
        desc = str(it.get("item_descripcion", "") or "").lower()
        bg = row_even if idx % 2 == 0 else row_odd
        vals = [
            cap,
            it.get("item_numero", ""),
            desc,
            it.get("unidad", ""),
            _fm(it.get("vlr_unitario_sub")),
            _fn(it.get("cantidad")),
            _fm(it.get("costo_directo")),
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = bd
            cell.font = Font(size=8)
            cell.fill = PatternFill("solid", fgColor=bg)
            if col == 3:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col in (5, 6, 7):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    tot_r = data0 + len(item_list)
    if not item_list:
        ws.merge_cells(start_row=tot_r, start_column=1, end_row=tot_r, end_column=7)
        emp = ws.cell(row=tot_r, column=1, value="Sin ítems con estado Aprobado en este corte.")
        emp.font = Font(size=8, italic=True, color="64748B")
        emp.alignment = Alignment(horizontal="left")
        emp.border = bd
        tot_r += 1

    st_r = tot_r
    ws.merge_cells(start_row=st_r, start_column=1, end_row=st_r, end_column=5)
    s1 = ws.cell(row=st_r, column=1, value="SUB TOTAL:")
    s1.font = Font(bold=True, size=9)
    s1.alignment = Alignment(horizontal="right", vertical="center")
    s1.fill = PatternFill("solid", fgColor=subtotal_bg)
    s1.border = bd
    ws.merge_cells(start_row=st_r, start_column=6, end_row=st_r, end_column=7)
    s2 = ws.cell(row=st_r, column=6, value=_fm(total_costo))
    s2.font = Font(bold=True, size=9)
    s2.alignment = Alignment(horizontal="right", vertical="center")
    s2.fill = PatternFill("solid", fgColor=subtotal_bg)
    s2.border = bd

    sr = st_r + 2
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=7)
    ws.cell(row=sr, column=1, value="FIRMAS").fill = fill_bar
    ws.cell(row=sr, column=1).font = Font(bold=True, size=9)
    ws.cell(row=sr, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=sr, column=1).border = bd

    def firma_block(r: int, lab: str, txt: str) -> None:
        ws.cell(row=r, column=1, value=lab).font = Font(size=8, bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=txt)
        c.font = Font(size=8)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).border = bd
        c.border = bd

    firma_block(sr + 1, "Elaboró:", f"{elaboro_n}\n{elaboro_c}")
    firma_block(sr + 2, "Revisó:", f"{reviso_n}\n{reviso_c}")
    firma_block(sr + 3, "Aprobó (subcontratista):", f"{aprobo_emp}\nRepresentante: {aprobo_rep}")

    foot_r = sr + 5
    ws.merge_cells(start_row=foot_r, start_column=1, end_row=foot_r, end_column=7)
    pie = ws.cell(
        row=foot_r,
        column=1,
        value=(
            f"Período del corte: {periodo} · Generado ClaraCore · {usuario_cargo} · {usuario_nombre}"
        ),
    )
    pie.font = Font(size=7, color="64748B")
    pie.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pie.border = bd

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 1
    ws.print_options.horizontalCentered = True


def _corte_sub_001_excel_bytes(
    contrato: dict,
    sub: dict,
    corte: dict,
    items: List[dict],
    total_costo: float,
    usuario_nombre: str,
    usuario_cargo: str,
    firma_cfg: Optional[Dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "CC-SUB-001"
    _fill_corte_sub_001_excel_ws(
        ws,
        contrato,
        sub,
        corte,
        items,
        total_costo,
        usuario_nombre,
        usuario_cargo,
        firma_cfg,
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_cc_conc_001_excel_ws(
    ws,
    contrato: dict,
    items: List[dict],
    total_costo: float,
    usuario_nombre: str,
    usuario_cargo: str,
    firma_cfg: Optional[Dict[str, Any]],
    *,
    titulo_documento: str,
    codigo_ccd: str,
    c3_label: str,
    c3_value: str,
    c4_label: str,
    c4_value: str,
    pie_contexto: str,
) -> None:
    """Hoja tipo CC-SEM-001 / CC-MES-001: encabezado conciliación, tabla de ítems, subtotal, firmas interventoría."""
    fc = firma_cfg or {}
    est = _merge_estilo_pdf(fc.get("estilo_pdf"), codigo_ccd)
    thead_bg = _ccd_hex_to_excel_rgb(est.get("thead_bg"), "E8E8E8")
    row_even = _ccd_hex_to_excel_rgb(est.get("row_even_bg"), "F8FAFC")
    row_odd = _ccd_hex_to_excel_rgb(est.get("row_odd_bg"), "FFFFFF")
    subtotal_bg = _ccd_hex_to_excel_rgb(est.get("subtotal_bg"), "DBEAFE")
    cap_sub_bg = _ccd_hex_to_excel_rgb(est.get("capitulo_subtotal_bg"), "93C5FD")

    elaboro_n = str(fc.get("elaboro_nombre") or "").strip() or "—"
    elaboro_c = str(fc.get("elaboro_cargo") or "").strip() or "—"
    reviso_n = str(fc.get("reviso_nombre") or "").strip() or "—"
    reviso_c = str(fc.get("reviso_cargo") or "").strip() or "—"
    aprobo_n = str(fc.get("aprobo_nombre") or "").strip() or "—"
    aprobo_c = str(fc.get("aprobo_cargo") or "").strip() or "—"

    thin = Side(style="thin", color="9CA3AF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_bar = PatternFill("solid", fgColor="E5E7EB")
    fill_th = PatternFill("solid", fgColor=thead_bg)

    for i, w in enumerate([10, 9, 36, 8, 12, 11, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def meta_row(r: int, label: str, value: object) -> None:
        ws.cell(row=r, column=1, value=label).font = Font(size=8, bold=True, color="374151")
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        v = ws.cell(row=r, column=2, value=value if value is not None else "—")
        v.font = Font(size=8)
        v.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=r, column=1).border = bd
        v.border = bd

    fecha_gen = _fmt_informe_fecha_generacion()
    nit_raw = str(contrato.get("nit") or "").strip()
    contratista_val = str(contrato.get("contratista") or "—")
    if nit_raw:
        contratista_val = f"{contratista_val} (NIT: {nit_raw})"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    t1 = ws.cell(row=1, column=1, value=str(titulo_documento or "").strip() or "—")
    t1.font = Font(bold=True, size=11)
    t1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    t1.border = bd

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    t2 = ws.cell(row=2, column=1, value=f"{codigo_ccd} · CCD · ClaraCore")
    t2.font = Font(bold=True, size=10, color="1E40AF")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    t2.border = bd

    meta_row(3, "CONTRATO", str(contrato.get("numero") or "—"))
    meta_row(4, "FECHA", fecha_gen)
    meta_row(5, str(c3_label or "").strip() or "—", str(c3_value or "").strip() or "—")
    meta_row(6, str(c4_label or "").strip() or "—", str(c4_value or "").strip() or "—")
    meta_row(7, "CONTRATISTA", contratista_val)
    meta_row(8, "INTERVENTORÍA", str(contrato.get("interventoria") or "—"))

    r0 = 10
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=7)
    br = ws.cell(row=r0, column=1, value="CANTIDADES APROBADAS POR ÍTEM")
    br.fill = fill_bar
    br.font = Font(bold=True, size=9)
    br.alignment = Alignment(horizontal="center", vertical="center")
    br.border = bd

    hdr = [
        "CAPÍTULO",
        "ÍTEM",
        "DESCRIPCIÓN",
        "UNIDAD",
        "VALOR UNIT.",
        "CANTIDAD",
        "COSTO DIRECTO",
    ]
    hr = r0 + 1
    for col, h in enumerate(hdr, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.fill = fill_th
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bd

    data0 = hr + 1
    item_list = list(items or [])
    row = data0
    data_idx = 0
    if not item_list:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        emp = ws.cell(row=row, column=1, value="Sin registros nivel 3 aprobados para este filtro.")
        emp.font = Font(size=8, italic=True, color="64748B")
        emp.alignment = Alignment(horizontal="left")
        emp.border = bd
        row += 1
    else:
        i = 0
        while i < len(item_list):
            cap = _capitulo_norm_conc(item_list[i])
            j = i
            sub_sum = 0.0
            while j < len(item_list) and _capitulo_norm_conc(item_list[j]) == cap:
                it = item_list[j]
                cap_cell = (it.get("capitulo") or "").strip() or "—"
                desc = str(it.get("item_descripcion", "") or "").lower()
                vu = it.get("vlr_unitario")
                if vu is None:
                    vu = it.get("vlr_unitario_sub")
                bg = row_even if data_idx % 2 == 0 else row_odd
                vals = [
                    cap_cell,
                    it.get("item_numero", ""),
                    desc,
                    it.get("unidad", ""),
                    _fm(vu),
                    _fn(it.get("cantidad")),
                    _fm(it.get("costo_directo")),
                ]
                for col, v in enumerate(vals, start=1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = bd
                    cell.font = Font(size=8)
                    cell.fill = PatternFill("solid", fgColor=bg)
                    if col == 3:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    elif col in (5, 6, 7):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                sub_sum += _sf(it.get("costo_directo"))
                data_idx += 1
                row += 1
                j += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            cs1 = ws.cell(row=row, column=1, value=f"Subtotal capítulo — {cap}")
            cs1.font = Font(bold=True, size=8)
            cs1.alignment = Alignment(horizontal="right", vertical="center")
            cs1.fill = PatternFill("solid", fgColor=cap_sub_bg)
            cs1.border = bd
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
            cs2 = ws.cell(row=row, column=6, value=_fm(sub_sum))
            cs2.font = Font(bold=True, size=9)
            cs2.alignment = Alignment(horizontal="right", vertical="center")
            cs2.fill = PatternFill("solid", fgColor=cap_sub_bg)
            cs2.border = bd
            row += 1
            i = j

    tot_r = row
    st_r = tot_r
    ws.merge_cells(start_row=st_r, start_column=1, end_row=st_r, end_column=5)
    s1 = ws.cell(row=st_r, column=1, value="SUB TOTAL:")
    s1.font = Font(bold=True, size=9)
    s1.alignment = Alignment(horizontal="right", vertical="center")
    s1.fill = PatternFill("solid", fgColor=subtotal_bg)
    s1.border = bd
    ws.merge_cells(start_row=st_r, start_column=6, end_row=st_r, end_column=7)
    s2 = ws.cell(row=st_r, column=6, value=_fm(total_costo))
    s2.font = Font(bold=True, size=9)
    s2.alignment = Alignment(horizontal="right", vertical="center")
    s2.fill = PatternFill("solid", fgColor=subtotal_bg)
    s2.border = bd

    sr = st_r + 2
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=7)
    ws.cell(row=sr, column=1, value="FIRMAS").fill = fill_bar
    ws.cell(row=sr, column=1).font = Font(bold=True, size=9)
    ws.cell(row=sr, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=sr, column=1).border = bd

    def firma_block(r: int, lab: str, txt: str) -> None:
        ws.cell(row=r, column=1, value=lab).font = Font(size=8, bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=txt)
        c.font = Font(size=8)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).border = bd
        c.border = bd

    firma_block(sr + 1, "Elaboró:", f"{elaboro_n}\n{elaboro_c}")
    firma_block(sr + 2, "Revisó:", f"{reviso_n}\n{reviso_c}")
    firma_block(sr + 3, "Aprobó (interventoría):", f"{aprobo_n}\n{aprobo_c}")

    foot_r = sr + 5
    ws.merge_cells(start_row=foot_r, start_column=1, end_row=foot_r, end_column=7)
    pie = ws.cell(
        row=foot_r,
        column=1,
        value=f"{pie_contexto} · Generado ClaraCore · {usuario_cargo} · {usuario_nombre}",
    )
    pie.font = Font(size=7, color="64748B")
    pie.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pie.border = bd

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 1
    ws.print_options.horizontalCentered = True


def _cc_sem_001_excel_bytes(contrato_id: int, semana_id: int, current_user: dict) -> bytes:
    if not _semana_pertenece_contrato(contrato_id, semana_id):
        raise HTTPException(404, "Semana no encontrada en este contrato")
    reg = fetch_registros_conciliacion(_sb, contrato_id, semana_id=semana_id)
    items, total = aggregate_items_conciliacion(reg)
    _sort_items_corte_por_item_numero_asc(items)
    sm = _row("so_semanas", "id, numero_semana, fecha_inicio, fecha_fin", id=semana_id) or {}
    nsem = sm.get("numero_semana")
    fi = str(sm.get("fecha_inicio") or "—")
    ff = str(sm.get("fecha_fin") or "—")
    contrato = _row(
        "contratos",
        "numero, objeto, contratista, nit, interventoria, logo_contratista",
        id=contrato_id,
    )
    if not contrato:
        raise HTTPException(404, "Contrato no encontrado")
    u = current_user if isinstance(current_user, dict) else dict(current_user)
    usuario_nombre = f"{u.get('nombre','')} {u.get('apellidos','')}".strip() or "—"
    usuario_cargo = u.get("cargo_nombre", "—") or "—"
    firma_cfg = _get_firma_cfg_para_documento(contrato_id, CODIGO_FORMATO_CCD_CC_SEM_001, contexto_tipo="semana", contexto_id=semana_id)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "CC-SEM-001"
    _fill_cc_conc_001_excel_ws(
        ws,
        contrato,
        items,
        float(total or 0.0),
        usuario_nombre,
        usuario_cargo,
        firma_cfg,
        titulo_documento="INFORME EJECUCIÓN SEMANAL (CONCILIACIÓN INTERVENTORÍA–CONTRATISTA)",
        codigo_ccd=CODIGO_FORMATO_CCD_CC_SEM_001,
        c3_label="SEMANA",
        c3_value=f"N° {nsem}",
        c4_label="VIGENCIA",
        c4_value=f"{fi} — {ff}",
        pie_contexto=f"Semana N° {nsem} · {fi} — {ff} · Registros nivel 3 aprobados y bloqueados",
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Objetivo: ~26 ítems en la 1ª hoja (con encabezado + subtotal + firmas sin página casi vacía).
_CC_SUB_001_ROWS_PAGINA_1 = 26
_CC_SUB_001_ROWS_PAGINA_SIG = 32


def _cc_sub_001_chunk_items(items: List[dict]) -> List[List[dict]]:
    if not items:
        return [[]]
    out: List[List[dict]] = []
    i = 0
    out.append(items[i : i + _CC_SUB_001_ROWS_PAGINA_1])
    i += _CC_SUB_001_ROWS_PAGINA_1
    while i < len(items):
        out.append(items[i : i + _CC_SUB_001_ROWS_PAGINA_SIG])
        i += _CC_SUB_001_ROWS_PAGINA_SIG
    return out


def _html_cc_sub_001_tr_item(item: dict, bd: str, row_bg: str = "") -> str:
    cap = (item.get("capitulo") or "").strip() or "—"
    desc = str(item.get("item_descripcion", "") or "").lower()
    trs = f"background:{row_bg};" if row_bg else ""
    return (
        f"<tr style=\"{trs}\">"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;vertical-align:top\">{_h(cap)}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;vertical-align:top\">{_h(item.get('item_numero', ''))}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:left\">{_h(desc)}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:center\">{_h(item.get('unidad', ''))}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:right\">{_fm(item.get('vlr_unitario_sub'))}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:right\">{_fn(item.get('cantidad'))}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:right\">{_fm(item.get('costo_directo'))}</td>"
        "</tr>"
    )


def _html_cc_sub_001_thead_items(bd: str, thead_bg: str) -> str:
    th = _sanitize_ccd_hex_color(thead_bg, "#e8e8e8")
    return f"""<thead>
<tr style="background:{th};">
<th style="{bd};padding:3px 2px;font-size:6.5pt;font-weight:bold;text-align:center;width:16%;">CAPITULO</th>
<th style="{bd};padding:3px 2px;font-size:6.5pt;font-weight:bold;text-align:center;width:8%;">ITEM</th>
<th style="{bd};padding:3px 2px;font-size:6.5pt;font-weight:bold;text-align:left;width:40%;">DESCRIPCIÓN</th>
<th style="{bd};padding:3px 2px;font-size:6.5pt;font-weight:bold;text-align:center;width:6%;">UNIDAD</th>
<th style="{bd};padding:3px 2px;font-size:6.5pt;font-weight:bold;text-align:center;width:10%;">VALOR UNIT.</th>
<th style="{bd};padding:3px 2px;font-size:6.5pt;font-weight:bold;text-align:center;width:9%;">CANTIDAD</th>
<th style="{bd};padding:3px 2px;font-size:6.5pt;font-weight:bold;text-align:center;width:11%;">COSTO DIR</th>
</tr>
</thead>"""

# Área de imagen de firma en PDFs CCD (~1,5 cm). xhtml2pdf respeta mejor pt que cm; altura fija en <td>
# evita que una columna estire TODA la fila de las tres firmas.
_CCD_FIRMA_BOX_PT = "28pt"  # ≈ 1 cm (más compacto que 42pt)
_CCD_FIRMA_IMG_INNER_PT = "26pt"
# Memorias CC-SUB-002 / CC-SEM / CC-MES (landscape): fila de firmas más baja que CC-SUB-001.
_CCD_FIRMA_MEM002_BOX_PT = "22pt"
_CCD_FIRMA_MEM002_IMG_INNER_PT = "20pt"


def _html_cc_sub_td_firma_columna(
    bd: str,
    etiqueta_hdr: str,
    nombre_h: str,
    cargo_h: str,
    firma_data_uri: Optional[str],
    *,
    memoria_compact: bool = False,
) -> str:
    """Una columna Elaboró o Revisó; si hay data URI, tabla con altura fija + img acotada (evita fila de firmas gigante)."""
    td_pad = "padding:1px 3px" if memoria_compact else "padding:3px 5px"
    fs = "6pt" if memoria_compact else "6.5pt"
    img_marg = "margin:0" if memoria_compact else "margin:1px 0 2px 0"
    if firma_data_uri:
        bp = _CCD_FIRMA_MEM002_BOX_PT if memoria_compact else _CCD_FIRMA_BOX_PT
        ip = _CCD_FIRMA_MEM002_IMG_INNER_PT if memoria_compact else _CCD_FIRMA_IMG_INNER_PT
        return f"""<td style="width:33.33%;{bd};{td_pad};font-size:{fs};vertical-align:top;">
<div class="ccd-firma-slot-hdr">{etiqueta_hdr}</div>
<div class="ccd-firma-line"></div>
<table class="ccd-firma-img-cage" cellspacing="0" cellpadding="0" width="100%" style="border-collapse:collapse;table-layout:fixed;{img_marg};">
<tr><td style="height:{bp};max-height:{bp};min-height:{bp};overflow:hidden;vertical-align:middle;text-align:center;line-height:0;font-size:0;padding:0;border:none;">
<img src="{firma_data_uri}" alt="" style="display:block;margin:0 auto;max-width:100%;width:auto;height:{ip};max-height:{ip};border:0;padding:0;"/>
</td></tr>
</table>
<div class="ccd-firma-nombre">{nombre_h}</div>
<div class="ccd-firma-cargo">{cargo_h}</div>
</td>"""
    return f"""<td style="width:33.33%;{bd};{td_pad};font-size:{fs};vertical-align:top;">
<div class="ccd-firma-slot-hdr">{etiqueta_hdr}</div>
<div class="ccd-firma-slot-body">
<div class="ccd-firma-line"></div>
<div class="ccd-firma-nombre">{nombre_h}</div>
<div class="ccd-firma-cargo">{cargo_h}</div>
</div>
</td>"""


def _html_cc_sub_v1_plain(
    contrato,
    sub,
    corte,
    items,
    total_costo,
    usuario_nombre,
    usuario_cargo,
    firma_cfg: Optional[Dict[str, Any]] = None,
    elaboro_firma_data_uri: Optional[str] = None,
    reviso_firma_data_uri: Optional[str] = None,
) -> str:
    """CC-SUB-001: encabezado solo 1ª hoja; ítems paginados; firmas al cierre (Elaboró/Revisó configurables; Aprobó desde subcontratista)."""
    bd = "border:1px solid #9ca3af"
    bd_blk = "border:1px solid #1f2937"
    codigo_ccd = CODIGO_FORMATO_CCD_CC_SUB_001
    logo_html = _html_logo_contratista(contrato)
    fecha_gen = _fmt_informe_fecha_generacion()
    corte_lbl = _corte_consecutivo_fmt(corte)
    contratista_nom = _h(str(contrato.get("contratista") or ""))
    interv = _h(str(contrato.get("interventoria") or ""))
    nit_raw = str(contrato.get("nit") or "").strip()
    nit_en_valor = f' <span style="font-size:6.5pt;color:#444;">(NIT: {_h(nit_raw)})</span>' if nit_raw else ""
    fc = firma_cfg or {}
    est = _merge_estilo_pdf(fc.get("estilo_pdf"), CODIGO_FORMATO_CCD_CC_SUB_001)
    elaboro_n = _h(str(fc.get("elaboro_nombre") or "").strip() or "—")
    elaboro_c = _h(str(fc.get("elaboro_cargo") or "").strip() or "—")
    reviso_n = _h(str(fc.get("reviso_nombre") or "").strip() or "—")
    reviso_c = _h(str(fc.get("reviso_cargo") or "").strip() or "—")
    aprobo_empresa = _h(str(sub.get("razon_social") or "").strip() or "—")
    aprobo_rep = _h(str(sub.get("nombre_contacto") or "").strip() or "—")
    lbl = "font-size:6pt;font-weight:bold;color:#111;text-transform:uppercase;letter-spacing:0.2px;"
    und = "border-bottom:1px solid #1f2937;font-size:7pt;padding:1px 0 2px 0;margin-top:1px;"

    chunks = _cc_sub_001_chunk_items(list(items or []))
    nchunks = len(chunks)

    parts: list[str] = []
    parts.append(f"""<!DOCTYPE html>
<html xmlns="http://www.w3.org/1999/xhtml" xmlns:pdf="http://www.xhtml2pdf.org/pdf">
<head><meta charset="UTF-8"/><title>{_h(codigo_ccd)}</title>
<style type="text/css">
@page {{ size: letter; margin: 8mm 10mm; }}
.cc001-tabla-items {{ width:100%; border-collapse:collapse; table-layout:fixed; }}
.cc001-tabla-items thead {{ display: table-header-group; }}
.ccd-cc001-firmas-wrap {{ margin-top: 5mm; page-break-inside: avoid; }}
.ccd-cc001-firmas-tbl {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
.ccd-cc001-firmas-tbl td {{ vertical-align: top; }}
/* Altura fija por columna (0.75cm, mitad de 1.5cm); overflow recorta texto muy largo. */
.ccd-firma-slot-hdr {{ font-weight: bold; font-size: 6.5pt; margin: 0 0 1px 0; padding: 0; color: #111; line-height: 1.1; }}
.ccd-firma-slot-body {{
  height: 0.75cm;
  min-height: 0.75cm;
  max-height: 0.75cm;
  overflow: hidden;
  box-sizing: border-box;
  padding: 0;
  margin: 0;
}}
/* xhtml2pdf: un div vacío con solo border-top a veces ocupa toda la altura del padre; forzar altura 0. */
.ccd-firma-line {{
  border: none;
  border-top: 1px solid #111;
  width: 100%;
  height: 0;
  line-height: 0;
  font-size: 0;
  margin: 0 0 2px 0;
  padding: 0;
  overflow: hidden;
}}
.ccd-firma-nombre {{ font-weight: bold; font-size: 7pt; line-height: 1.1; margin: 1px 0 0 0; padding: 0; word-wrap: break-word; }}
.ccd-firma-cargo {{ font-size: 6.5pt; color: #444; line-height: 1.08; margin: 0; padding: 0; word-wrap: break-word; }}
.ccd-firma-rep {{ font-size: 6.5pt; color: #333; line-height: 1.1; margin: 1px 0 0 0; padding: 0; word-wrap: break-word; }}
/* Refuerzo: tabla .ccd-firma-img-cage (altura en línea en el HTML). */
.ccd-firma-img-cage td {{
  overflow: hidden !important;
  padding: 0 !important;
}}
.ccd-firma-img-cage img {{
  max-width: 100% !important;
  height: {_CCD_FIRMA_IMG_INNER_PT} !important;
  max-height: {_CCD_FIRMA_IMG_INNER_PT} !important;
  width: auto !important;
  display: block !important;
  margin: 0 auto !important;
}}
</style></head>
<body style="margin:0;padding:4px;font-family:Arial,Helvetica,sans-serif;font-size:7.5pt;color:#111;">
""")

    # ── Bloque encabezado (solo antes del primer corte de tabla de ítems = primera hoja) ──
    parts.append(f"""
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd_blk};table-layout:fixed;">
<tr>
<td style="width:20%;{bd_blk};vertical-align:middle;padding:2px;text-align:center;background:#fff">
{logo_html}
</td>
<td style="width:48%;{bd_blk};vertical-align:middle;text-align:center;font-weight:bold;font-size:8.2pt;padding:3px 5px;line-height:1.1;height:32px;">
INFORME CORTE DE SUB CONTRATISTA
</td>
<td style="width:32%;{bd_blk};vertical-align:middle;text-align:center;padding:3px 5px;">
<div style="color:#1e3a8a;font-weight:bold;font-size:12.5pt;letter-spacing:0.5px;line-height:1;">{_h(codigo_ccd)}</div>
<div style="font-size:8.5pt;color:#1e3a8a;font-weight:bold;margin-top:2px;">CCD · ClaraCore</div>
</td>
</tr>
<tr><td colspan="3" style="padding:0;border:none;">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd_blk};background:#fff">
<tr><td style="padding:2px 6px;border:none;">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:none;">
<tr>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">CONTRATO</div>
<div style="{und}">{_h(contrato.get('numero', ''))}</div>
</td>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">FECHA</div>
<div style="{und}">{_h(fecha_gen)}</div>
</td>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">SUB CONTRATISTA</div>
<div style="{und}">{_h(sub.get('razon_social', ''))}</div>
</td>
<td style="width:25%;padding:0 0 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">CORTE</div>
<div style="{und}">{_h(corte_lbl)}</div>
</td>
</tr>
<tr>
<td colspan="2" style="padding:3px 5px 1px 0;border:none;vertical-align:top;">
<div style="{lbl}">CONTRATISTA</div>
<div style="{und}">{contratista_nom}{nit_en_valor}</div>
</td>
<td colspan="2" style="padding:3px 0 1px 0;border:none;vertical-align:top;">
<div style="{lbl}">INTERVENTORÍA</div>
<div style="{und}">{interv}</div>
</td>
</tr>
</table>
</td></tr>
</table>
</td></tr>
</table>
""")

    for ci, chunk in enumerate(chunks):
        if ci > 0:
            parts.append('<pdf:nextpage />')
        parts.append(f'<table class="cc001-tabla-items" cellspacing="0" cellpadding="0">')
        parts.append(_html_cc_sub_001_thead_items(bd, est["thead_bg"]))
        parts.append("<tbody>")
        if not chunk and not items:
            parts.append(
                f"<tr><td colspan=\"7\" style=\"{bd};padding:5px;font-size:7pt;color:#6b7280\">"
                "Sin ítems con estado Aprobado en este corte.</td></tr>"
            )
        else:
            for idx, it in enumerate(chunk):
                row_bg = est["row_even_bg"] if idx % 2 == 0 else est["row_odd_bg"]
                parts.append(_html_cc_sub_001_tr_item(it, bd, row_bg))
        if ci == nchunks - 1:
            st = _sanitize_ccd_hex_color(est.get("subtotal_bg"), "#dbeafe")
            parts.append(
                f"""<tr style="background:{st};">
<td colspan="5" style="{bd};text-align:right;padding:4px 6px;font-weight:bold;font-size:7.5pt;">SUB TOTAL:</td>
<td colspan="2" style="{bd};text-align:right;padding:4px 6px;font-weight:bold;font-size:7.5pt;">{_fm(total_costo)}</td>
</tr>"""
            )
        parts.append("</tbody></table>")

    elaboro_td = _html_cc_sub_td_firma_columna(bd, "Elaboró:", elaboro_n, elaboro_c, elaboro_firma_data_uri)
    reviso_td = _html_cc_sub_td_firma_columna(bd, "Revisó:", reviso_n, reviso_c, reviso_firma_data_uri)

    # Firmas: solo tras cerrar la última tabla de ítems (última hoja). Izq=Elaboró, Centro=Revisó, Der=Aprobó (datos del sub en módulo administrativo).
    parts.append(f"""
<div class="ccd-cc001-firmas-wrap">
<table class="ccd-cc001-firmas-tbl" cellspacing="0" cellpadding="0">
<tr>
{elaboro_td}
{reviso_td}
<td style="width:33.33%;{bd};padding:3px 5px;font-size:6.5pt;vertical-align:top;">
<div class="ccd-firma-slot-hdr">Aprobó:</div>
<div class="ccd-firma-slot-body">
<div class="ccd-firma-line"></div>
<div class="ccd-firma-nombre">{aprobo_empresa}</div>
<div class="ccd-firma-rep">Representante: {aprobo_rep}</div>
</div>
</td>
</tr>
</table>
</div>
<p style="font-size:6pt;color:#64748b;margin-top:6px;text-align:center;">
Período del corte: {_h(_fd(corte.get('fecha_inicio')))} — {_h(_fd(corte.get('fecha_fin')))} · Generado ClaraCore · {_h(usuario_cargo)} · Sesión: {_h(usuario_nombre)}
</p>
</body></html>""")
    return "".join(parts)


def _capitulo_norm_conc(it: dict) -> str:
    return (it.get("capitulo") or "").strip() or "—"


def _cc_conc_plan_filas_con_subtotal_capitulo(items: List[dict]) -> List[tuple]:
    """Secuencia de filas: ('item', item, idx_paridad) o ('subcap', etiqueta_corta, total_costo_directo)."""
    items = list(items or [])
    if not items:
        return []
    out: List[tuple] = []
    idx = 0
    i = 0
    while i < len(items):
        cap = _capitulo_norm_conc(items[i])
        j = i
        tot = 0.0
        while j < len(items) and _capitulo_norm_conc(items[j]) == cap:
            tot += _sf(items[j].get("costo_directo"))
            out.append(("item", items[j], idx))
            idx += 1
            j += 1
        short = cap if len(cap) <= 56 else (cap[:53] + "…")
        out.append(("subcap", short, tot))
        i = j
    return out


def _html_cc_conc_001_tr_subtotal_capitulo(bd: str, cap_etiqueta: str, sum_cd: float, bg_hex: str) -> str:
    st = _sanitize_ccd_hex_color(bg_hex, "#93c5fd")
    return (
        f'<tr class="cc001-cap-sub" style="background:{st};">'
        f'<td colspan="5" style="{bd};padding:3px 6px;font-size:7pt;font-weight:bold;text-align:right;vertical-align:middle">'
        f"Subtotal capítulo — {_h(cap_etiqueta)}</td>"
        f'<td colspan="2" style="{bd};padding:3px 6px;font-size:7.5pt;font-weight:bold;text-align:right;vertical-align:middle">'
        f"{_fm(sum_cd)}</td>"
        "</tr>"
    )


def _html_cc_conc_001_tr_item(item: dict, bd: str, row_bg: str = "") -> str:
    cap = (item.get("capitulo") or "").strip() or "—"
    desc = str(item.get("item_descripcion", "") or "").lower()
    trs = f"background:{row_bg};" if row_bg else ""
    vu = item.get("vlr_unitario")
    if vu is None:
        vu = item.get("vlr_unitario_sub")
    return (
        f"<tr style=\"{trs}\">"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;vertical-align:top\">{_h(cap)}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;vertical-align:top\">{_h(item.get('item_numero', ''))}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:left\">{_h(desc)}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:center\">{_h(item.get('unidad', ''))}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:right\">{_fm(vu)}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:right\">{_fn(item.get('cantidad'))}</td>"
        f"<td style=\"{bd};padding:2px 3px;font-size:7pt;text-align:right\">{_fm(item.get('costo_directo'))}</td>"
        "</tr>"
    )


def _html_cc_conciliacion_informe_v1(
    contrato,
    items,
    total_costo,
    usuario_nombre,
    usuario_cargo,
    *,
    codigo_ccd: str,
    titulo_documento: str,
    c3_label: str,
    c3_value: str,
    c4_label: str,
    c4_value: str,
    pie_contexto: str,
    firma_cfg: Optional[Dict[str, Any]],
    elaboro_firma_data_uri: Optional[str],
    reviso_firma_data_uri: Optional[str],
    aprobo_firma_data_uri: Optional[str],
    estilo_formato_codigo: str,
) -> str:
    """Informe tipo CC-SEM-001 / CC-MES-001: sin subcontratista; Aprobó interventoría con firma opcional."""
    bd = "border:1px solid #9ca3af"
    bd_blk = "border:1px solid #1f2937"
    logo_html = _html_logo_contratista(contrato)
    fecha_gen = _fmt_informe_fecha_generacion()
    contratista_nom = _h(str(contrato.get("contratista") or ""))
    interv = _h(str(contrato.get("interventoria") or ""))
    nit_raw = str(contrato.get("nit") or "").strip()
    nit_en_valor = f' <span style="font-size:6.5pt;color:#444;">(NIT: {_h(nit_raw)})</span>' if nit_raw else ""
    fc = firma_cfg or {}
    est = _merge_estilo_pdf(fc.get("estilo_pdf"), estilo_formato_codigo)
    elaboro_n = _h(str(fc.get("elaboro_nombre") or "").strip() or "—")
    elaboro_c = _h(str(fc.get("elaboro_cargo") or "").strip() or "—")
    reviso_n = _h(str(fc.get("reviso_nombre") or "").strip() or "—")
    reviso_c = _h(str(fc.get("reviso_cargo") or "").strip() or "—")
    aprobo_n = _h(str(fc.get("aprobo_nombre") or "").strip() or "—")
    aprobo_c = _h(str(fc.get("aprobo_cargo") or "").strip() or "—")
    lbl = "font-size:6pt;font-weight:bold;color:#111;text-transform:uppercase;letter-spacing:0.2px;"
    und = "border-bottom:1px solid #1f2937;font-size:7pt;padding:1px 0 2px 0;margin-top:1px;"
    plan = _cc_conc_plan_filas_con_subtotal_capitulo(list(items or []))
    cap_st_bg = est.get("capitulo_subtotal_bg") or "#93c5fd"
    parts: list[str] = []
    parts.append(f"""<!DOCTYPE html>
<html xmlns="http://www.w3.org/1999/xhtml" xmlns:pdf="http://www.xhtml2pdf.org/pdf">
<head><meta charset="UTF-8"/><title>{_h(codigo_ccd)}</title>
<style type="text/css">
@page {{ size: letter; margin: 8mm 10mm; }}
.cc001-tabla-items {{ width:100%; border-collapse:collapse; table-layout:fixed; }}
.cc001-tabla-items thead {{ display: table-header-group; }}
.cc001-tabla-items tr.cc001-cap-sub {{ page-break-inside: avoid; }}
.cc001-tabla-items tr.cc001-grand-sub {{ page-break-inside: avoid; }}
.ccd-cc001-firmas-wrap {{ margin-top: 5mm; page-break-inside: avoid; }}
.ccd-cc001-firmas-tbl {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
.ccd-cc001-firmas-tbl td {{ vertical-align: top; }}
.ccd-firma-slot-hdr {{ font-weight: bold; font-size: 6.5pt; margin: 0 0 1px 0; padding: 0; color: #111; line-height: 1.1; }}
.ccd-firma-slot-body {{
  height: 0.75cm;
  min-height: 0.75cm;
  max-height: 0.75cm;
  overflow: hidden;
  box-sizing: border-box;
  padding: 0;
  margin: 0;
}}
.ccd-firma-line {{
  border: none;
  border-top: 1px solid #111;
  width: 100%;
  height: 0;
  line-height: 0;
  font-size: 0;
  margin: 0 0 2px 0;
  padding: 0;
  overflow: hidden;
}}
.ccd-firma-nombre {{ font-weight: bold; font-size: 7pt; line-height: 1.1; margin: 1px 0 0 0; padding: 0; word-wrap: break-word; }}
.ccd-firma-cargo {{ font-size: 6.5pt; color: #444; line-height: 1.08; margin: 0; padding: 0; word-wrap: break-word; }}
.ccd-firma-img-cage td {{
  overflow: hidden !important;
  padding: 0 !important;
}}
.ccd-firma-img-cage img {{
  max-width: 100% !important;
  height: {_CCD_FIRMA_IMG_INNER_PT} !important;
  max-height: {_CCD_FIRMA_IMG_INNER_PT} !important;
  width: auto !important;
  display: block !important;
  margin: 0 auto !important;
}}
</style></head>
<body style="margin:0;padding:4px;font-family:Arial,Helvetica,sans-serif;font-size:7.5pt;color:#111;">
""")
    parts.append(f"""
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd_blk};table-layout:fixed;">
<tr>
<td style="width:20%;{bd_blk};vertical-align:middle;padding:2px;text-align:center;background:#fff">
{logo_html}
</td>
<td style="width:48%;{bd_blk};vertical-align:middle;text-align:center;font-weight:bold;font-size:8.2pt;padding:3px 5px;line-height:1.1;height:32px;">
{_h(titulo_documento)}
</td>
<td style="width:32%;{bd_blk};vertical-align:middle;text-align:center;padding:3px 5px;">
<div style="color:#1e3a8a;font-weight:bold;font-size:12.5pt;letter-spacing:0.5px;line-height:1;">{_h(codigo_ccd)}</div>
<div style="font-size:8.5pt;color:#1e3a8a;font-weight:bold;margin-top:2px;">CCD · ClaraCore</div>
</td>
</tr>
<tr><td colspan="3" style="padding:0;border:none;">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;{bd_blk};background:#fff">
<tr><td style="padding:2px 6px;border:none;">
<table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:none;">
<tr>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">CONTRATO</div>
<div style="{und}">{_h(contrato.get('numero', ''))}</div>
</td>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">FECHA</div>
<div style="{und}">{_h(fecha_gen)}</div>
</td>
<td style="width:25%;padding:0 5px 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">{_h(c3_label)}</div>
<div style="{und}">{_h(c3_value)}</div>
</td>
<td style="width:25%;padding:0 0 2px 0;border:none;vertical-align:top;">
<div style="{lbl}">{_h(c4_label)}</div>
<div style="{und}">{_h(c4_value)}</div>
</td>
</tr>
<tr>
<td colspan="2" style="padding:3px 5px 1px 0;border:none;vertical-align:top;">
<div style="{lbl}">CONTRATISTA</div>
<div style="{und}">{contratista_nom}{nit_en_valor}</div>
</td>
<td colspan="2" style="padding:3px 0 1px 0;border:none;vertical-align:top;">
<div style="{lbl}">INTERVENTORÍA</div>
<div style="{und}">{interv}</div>
</td>
</tr>
</table>
</td></tr>
</table>
</td></tr>
</table>
""")
    parts.append('<table class="cc001-tabla-items" cellspacing="0" cellpadding="0">')
    parts.append(_html_cc_sub_001_thead_items(bd, est["thead_bg"]))
    parts.append("<tbody>")
    if not items:
        parts.append(
            f"<tr><td colspan=\"7\" style=\"{bd};padding:5px;font-size:7pt;color:#6b7280\">"
            "Sin registros nivel 3 aprobados (y bloqueados) para este filtro.</td></tr>"
        )
    else:
        for row in plan:
            if row[0] == "item":
                _, it, idx = row
                row_bg = est["row_even_bg"] if idx % 2 == 0 else est["row_odd_bg"]
                parts.append(_html_cc_conc_001_tr_item(it, bd, row_bg))
            else:
                _, cap_lab, sub_amt = row
                parts.append(_html_cc_conc_001_tr_subtotal_capitulo(bd, cap_lab, sub_amt, cap_st_bg))
        st = _sanitize_ccd_hex_color(est.get("subtotal_bg"), "#dbeafe")
        parts.append(
            f"""<tr class="cc001-grand-sub" style="background:{st};">
<td colspan="5" style="{bd};text-align:right;padding:4px 6px;font-weight:bold;font-size:7.5pt;">SUB TOTAL:</td>
<td colspan="2" style="{bd};text-align:right;padding:4px 6px;font-weight:bold;font-size:7.5pt;">{_fm(total_costo)}</td>
</tr>"""
        )
    parts.append("</tbody></table>")
    elaboro_td = _html_cc_sub_td_firma_columna(bd, "Elaboró:", elaboro_n, elaboro_c, elaboro_firma_data_uri)
    reviso_td = _html_cc_sub_td_firma_columna(bd, "Revisó:", reviso_n, reviso_c, reviso_firma_data_uri)
    aprobo_td = _html_cc_sub_td_firma_columna(bd, "Aprobó:", aprobo_n, aprobo_c, aprobo_firma_data_uri)
    parts.append(f"""
<div class="ccd-cc001-firmas-wrap">
<table class="ccd-cc001-firmas-tbl" cellspacing="0" cellpadding="0">
<tr>
{elaboro_td}
{reviso_td}
{aprobo_td}
</tr>
</table>
</div>
<p style="font-size:6pt;color:#64748b;margin-top:6px;text-align:center;">
{_h(pie_contexto)} · Generado ClaraCore · {_h(usuario_cargo)} · Sesión: {_h(usuario_nombre)}
</p>
</body></html>""")
    return "".join(parts)


def _html_corte_sub(contrato, sub, corte, items, total_costo, usuario_nombre, usuario_cargo):
    now = datetime.now(pytz.timezone("America/Bogota")).strftime("%d %b %y, %I:%M %p")
    # Sin logo en PDF: evita descargas y <img> (xhtml2pdf es frágil con imágenes/data URI largas).
    logo_td = "<span style='font-size:7pt;color:#6b7280'>LOGO CONTRATISTA</span>"

    ROWS_PER_PAGE = 40

    def bloque_encabezado():
        return f"""
<table class="w100 hdr-outer">
  <tr>
    <td class="hdr-logo">{logo_td}</td>
    <td class="hdr-main" style="border-right:1px solid #9ca3af"><div class="doc-title">INFORME CORTE DE SUB CONTRATISTA</div></td>
    <td class="hdr-main" style="width:190px;border-right:1px solid #9ca3af"><div class="doc-code">UNION TEMPORAL MURCON</div></td>
    <td class="hdr-main" style="width:95px"><div class="doc-code">SICOE</div></td>
  </tr>
</table>
<table class="w100 doc-meta">
  <tr>
    <td style="width:25%"><span class="lbl">CONTRATO</span><br/><span class="val">{_h(contrato.get('numero',''))}</span></td>
    <td style="width:20%"><span class="lbl">FECHA:</span><br/><span class="val">{_h(now)}</span></td>
    <td style="width:40%"><span class="lbl">SUB CONTRATISTA:</span><br/><span class="val">{_h((sub.get('razon_social','') or '').upper())}</span></td>
    <td style="width:15%"><span class="lbl">CORTE</span><br/><span class="val">{_h(corte.get('consecutivo',''))}</span></td>
  </tr>
  <tr>
    <td colspan="2"><span class="lbl">CONTRATISTA:</span><br/><span class="val">{_h(contrato.get('contratista',''))}</span></td>
    <td colspan="2"><span class="lbl">INTERVENTORÍA:</span><br/><span class="val">{_h(contrato.get('interventoria',''))}</span></td>
  </tr>
  <tr>
    <td colspan="4"><span class="lbl">NIT SUB:</span><br/><span class="val">{_h(sub.get('nit',''))}</span></td>
  </tr>
</table>
"""

    chunks = [items[i:i + ROWS_PER_PAGE] for i in range(0, len(items), ROWS_PER_PAGE)] or [[]]
    tablas = ""
    for ci, chunk in enumerate(chunks):
        if ci > 0:
            tablas += "<pdf:nextpage />"
        tablas += bloque_encabezado()
        tablas += """
<div class="section-bar">CANTIDADES APROBADAS POR SUBCONTRATISTA</div>
<table class="w100">
  <tr>
    <th class="data-th" style="width:10%">ITEM</th>
    <th class="data-th" style="width:8%">UNIDAD</th>
    <th class="data-th" style="width:12%">CANTIDAD</th>
    <th class="data-th" style="width:16%">VALOR UNIT.</th>
    <th class="data-th" style="width:18%">COSTO DIR.</th>
    <th class="data-th" style="width:36%">DESCRIPCIÓN</th>
  </tr>
"""
        for i, item in enumerate(chunk):
            cls = "even" if i % 2 == 0 else ""
            tablas += f"""<tr class="{cls}">
            <td class="data-td" style="text-align:center">{_h(item.get('item_numero',''))}</td>
            <td class="data-td" style="text-align:center">{_h(item.get('unidad',''))}</td>
            <td class="data-td" style="text-align:right">{_fn(item.get('cantidad'))}</td>
            <td class="data-td" style="text-align:right">{_fm(item.get('vlr_unitario_sub'))}</td>
            <td class="data-td" style="text-align:right">{_fm(item.get('costo_directo'))}</td>
            <td class="data-td">{_h(item.get('item_descripcion',''))}</td>
        </tr>"""
        if ci == len(chunks) - 1:
            tablas += f"""
  <tr>
    <td class="total-td" colspan="4" style="text-align:right;padding-right:10px">SUB TOTAL:</td>
    <td class="total-td" style="text-align:right">{_fm(total_costo)}</td>
    <td class="total-td"></td>
  </tr>
"""
        tablas += "</table>"

    # Sin bloque de firmas (xhtml2pdf es inestable con saltos extra + tablas de firmas). Se reintroducirá en una versión posterior.
    pie_periodo = (
        f"Período del corte: {_fd(corte.get('fecha_inicio'))} — {_fd(corte.get('fecha_fin'))}. "
        f"Generado en sesión: {_h(usuario_nombre)} ({_h(usuario_cargo)})."
    )

    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/>
<style>{BASE_CSS}
</style></head><body>
{tablas}

<p style="font-size:7.5pt;color:#555;margin:10px 0 6px 0">{pie_periodo}</p>
<div class="doc-footer">
  Documento institucional de control interno. Prohibida su reproduccion parcial o total sin autorizacion escrita.
</div>

</body></html>"""

def _html_corte_sub_fallback(contrato, sub, corte, items, total_costo, usuario_nombre, usuario_cargo):
    """Plantilla simple y estable (sin logo ni layout complejo) para no bloquear operación."""
    filas = ""
    for item in items:
        filas += f"""<tr>
          <td style="border:1px solid #999;padding:4px">{_h(item.get('item_numero',''))}</td>
          <td style="border:1px solid #999;padding:4px">{_h(item.get('item_descripcion',''))}</td>
          <td style="border:1px solid #999;padding:4px;text-align:center">{_h(item.get('unidad',''))}</td>
          <td style="border:1px solid #999;padding:4px;text-align:right">{_fn(item.get('cantidad',0))}</td>
          <td style="border:1px solid #999;padding:4px;text-align:right">{_fm(item.get('vlr_unitario_sub',0))}</td>
          <td style="border:1px solid #999;padding:4px;text-align:right">{_fm(item.get('costo_directo',0))}</td>
        </tr>"""

    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/></head>
<body style="font-family:Arial,sans-serif;font-size:9pt;color:#111">
  <h2 style="margin:0 0 6px 0">INFORME CORTE DE SUB CONTRATISTA (MODO SEGURO)</h2>
  <div style="margin-bottom:8px">
    <b>Contrato:</b> {_h(contrato.get('numero',''))} &nbsp;|&nbsp;
    <b>Subcontratista:</b> {_h(sub.get('razon_social',''))} &nbsp;|&nbsp;
    <b>Corte:</b> {_h(corte.get('consecutivo',''))}
  </div>
  <table style="width:100%;border-collapse:collapse">
    <tr style="background:#eee">
      <th style="border:1px solid #999;padding:4px">Item</th>
      <th style="border:1px solid #999;padding:4px">Descripcion</th>
      <th style="border:1px solid #999;padding:4px">Und</th>
      <th style="border:1px solid #999;padding:4px">Cantidad</th>
      <th style="border:1px solid #999;padding:4px">Vlr Unit.</th>
      <th style="border:1px solid #999;padding:4px">Costo Dir.</th>
    </tr>
    {filas}
    <tr>
      <td colspan="5" style="border:1px solid #999;padding:4px;text-align:right"><b>SUB TOTAL</b></td>
      <td style="border:1px solid #999;padding:4px;text-align:right"><b>{_fm(total_costo)}</b></td>
    </tr>
  </table>
  <div style="margin-top:14px">
    <b>Generado por:</b> {_h(usuario_nombre)} - {_h(usuario_cargo)}
  </div>
  <div style="margin-top:10px;font-size:7pt;color:#555">
    Documento institucional de control interno. Prohibida su reproduccion parcial o total sin autorizacion escrita.
  </div>
</body></html>"""


def _html_corte_sub_minima(contrato, sub, corte, items, total_costo, usuario_nombre, usuario_cargo, err1: str, err2: str) -> str:
    """Último recurso CC-SUB-001: tabla simple sin estilos complejos ni firmas multipágina."""
    filas = ""
    for item in items:
        filas += f"""<tr>
          <td>{_h(item.get("item_numero", ""))}</td>
          <td>{_h(item.get("unidad", ""))}</td>
          <td>{_fn(item.get("cantidad"))}</td>
          <td>{_fm(item.get("vlr_unitario_sub"))}</td>
          <td>{_fm(item.get("costo_directo"))}</td>
          <td>{_h(item.get("item_descripcion", ""))}</td>
        </tr>"""
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/></head>
<body style="font-family:Arial,sans-serif;font-size:9pt">
  <p style="color:#b45309;font-size:8pt">Vista simplificada CC-SUB-001. Motivo: {_h((err1 + " | " + err2)[:500])}</p>
  <p><b>Contrato</b> {_h(contrato.get("numero", ""))} · <b>Sub</b> {_h(sub.get("razon_social", ""))} · <b>Corte</b> {_h(corte.get("consecutivo", ""))}</p>
  <table border="1" cellpadding="4" style="border-collapse:collapse;width:100%">
    <tr><th>Ítem</th><th>Und</th><th>Cant</th><th>Vlr u.</th><th>Costo</th><th>Descripción</th></tr>
    {filas}
    <tr><td colspan="4" align="right"><b>Subtotal</b></td><td colspan="2"><b>{_fm(total_costo)}</b></td></tr>
  </table>
  <p>{_h(usuario_nombre)} — {_h(usuario_cargo)}</p>
</body></html>"""


def _html_memoria_minima(contrato, sub, corte, item_info, registros, usuario_nombre, usuario_cargo, err_note: str) -> str:
    """Si la plantilla completa falla (datos raros / xhtml2pdf), al menos un PDF legible."""
    filas = ""
    for r in registros:
        filas += f"""<tr>
          <td style="border:1px solid #999;padding:4px">{_h(r.get("numero_registro"))}</td>
          <td style="border:1px solid #999;padding:4px">{_h(r.get("abs_inicio"))}</td>
          <td style="border:1px solid #999;padding:4px">{_h(r.get("abs_final"))}</td>
          <td style="border:1px solid #999;padding:4px">{_fn(r.get("cantidad_total"))}</td>
          <td style="border:1px solid #999;padding:4px">{_h((r.get("observacion") or "")[:300])}</td>
        </tr>"""
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/></head>
<body style="font-family:Arial,sans-serif;font-size:9pt;color:#111">
  <p style="color:#b45309;font-size:8pt">Vista simplificada (hubo un problema al armar el formato completo): {_h(err_note[:400])}</p>
  <h2 style="margin:0 0 8px 0">CC-SUB-002 · {_h(item_info.get("item_numero", ""))}</h2>
  <p><b>Contrato:</b> {_h(contrato.get("numero", ""))} &nbsp; <b>Sub:</b> {_h(sub.get("razon_social", ""))} &nbsp; <b>Corte:</b> {_h(corte.get("consecutivo", ""))}</p>
  <p><b>Usuario:</b> {_h(usuario_nombre)} — {_h(usuario_cargo)}</p>
  <table style="width:100%;border-collapse:collapse">{filas}</table>
</body></html>"""


def _html_mem002_bloque_firmas(
    sub: dict,
    firma_cfg: Optional[Dict[str, Any]],
    elaboro_firma_data_uri: Optional[str] = None,
    reviso_firma_data_uri: Optional[str] = None,
    *,
    aprobo_firma_data_uri: Optional[str] = None,
    aprobo_interventoria_desde_config: bool = False,
) -> str:
    """Tres columnas Elaboró / Revisó / Aprobó: subcontratista (CC-SUB) o interventoría desde biblioteca (CC-SEM/MES)."""
    fc = firma_cfg or {}
    elaboro_n = _h(str(fc.get("elaboro_nombre") or "").strip() or "—")
    elaboro_c = _h(str(fc.get("elaboro_cargo") or "").strip() or "—")
    reviso_n = _h(str(fc.get("reviso_nombre") or "").strip() or "—")
    reviso_c = _h(str(fc.get("reviso_cargo") or "").strip() or "—")
    bd = "border:1px solid #9ca3af"
    elaboro_td = _html_cc_sub_td_firma_columna(
        bd, "Elaboró:", elaboro_n, elaboro_c, elaboro_firma_data_uri, memoria_compact=True
    )
    reviso_td = _html_cc_sub_td_firma_columna(
        bd, "Revisó:", reviso_n, reviso_c, reviso_firma_data_uri, memoria_compact=True
    )
    if aprobo_interventoria_desde_config:
        aprobo_n = _h(str(fc.get("aprobo_nombre") or "").strip() or "—")
        aprobo_c = _h(str(fc.get("aprobo_cargo") or "").strip() or "—")
        aprobo_td = _html_cc_sub_td_firma_columna(
            bd, "Aprobó:", aprobo_n, aprobo_c, aprobo_firma_data_uri, memoria_compact=True
        )
    else:
        aprobo_empresa = _h(str(sub.get("razon_social") or "").strip() or "—")
        aprobo_rep = _h(str(sub.get("nombre_contacto") or "").strip() or "—")
        aprobo_td = f"""<td style="width:33.33%;{bd};padding:1px 3px;font-size:6pt;vertical-align:top;">
<div class="mem002-firma-slot-hdr">Aprobó:</div>
<div class="mem002-firma-slot-body">
<div class="mem002-firma-line"></div>
<div class="mem002-firma-nombre">{aprobo_empresa}</div>
<div class="mem002-firma-rep">Representante: {aprobo_rep}</div>
</div>
</td>"""
    return f"""<div class="mem002-firmas-wrap">
<table class="mem002-firmas-tbl" cellspacing="0" cellpadding="0">
<tr>
{elaboro_td}
{reviso_td}
{aprobo_td}
</tr>
</table>
</div>"""


def _chunks_memoria_detalle(registros: List[Any], rows_primera_hoja: int, rows_hoja_siguiente: int) -> List[List[Any]]:
    """Trocea filas del detalle CC-SUB-002: la 1ª hoja lleva encabezado institucional (menos filas); el resto objetivo mayor.
    No fuerza “cuota” rígida en la última hoja: el último bloque puede ser más corto."""
    reg = list(registros or [])
    if not reg:
        return [[]]
    rp = max(1, rows_primera_hoja)
    rs = max(1, rows_hoja_siguiente)
    out: List[List[Any]] = []
    i = 0
    take = min(rp, len(reg) - i)
    out.append(reg[i : i + take])
    i += take
    while i < len(reg):
        take = min(rs, len(reg) - i)
        out.append(reg[i : i + take])
        i += take
    return out


def _memoria_pdf_estilo_css(est: Dict[str, str]) -> str:
    """CSS extra CC-SUB-002 según estilo guardado en biblioteca (colores ya validados)."""
    sb = _sanitize_ccd_hex_color(est.get("section_bar_bg"), "#e5e7eb")
    st = _sanitize_ccd_hex_color(est.get("section_bar_text"), "#111827")
    th = _sanitize_ccd_hex_color(est.get("thead_bg"), "#f3f4f6")
    ev = _sanitize_ccd_hex_color(est.get("row_even_bg"), "#f8fafc")
    od = _sanitize_ccd_hex_color(est.get("row_odd_bg"), "#ffffff")
    su = _sanitize_ccd_hex_color(est.get("subtotal_bg"), "#e5e7eb")
    return (
        f"body.mem002-doc .section-bar.mem002-section{{background:{sb}!important;color:{st}!important;}}"
        f"body.mem002-doc .mem002-detail th.data-th{{background:{th}!important;}}"
        f"body.mem002-doc .mem002-detail tr.even td{{background:{ev}!important;}}"
        f"body.mem002-doc .mem002-detail tr.odd td{{background:{od}!important;}}"
        f"body.mem002-doc .mem002-total-wrap td.total-td{{background:{su}!important;}}"
    )


def _wrap_memoria_item_html(body_html: str, estilo_css: str) -> str:
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/>
<style>{BASE_CSS_SHARED}{MEMORIA002_CSS}{estilo_css}</style></head><body class="mem002-doc">
{body_html}
</body></html>"""


# ── Template CC-SUB-002 ────────────────────────────────────────────────────────

def _html_memoria_item_body(
    contrato,
    sub,
    corte,
    item_info,
    registros,
    usuario_nombre,
    usuario_cargo,
    firma_cfg: Optional[Dict[str, Any]] = None,
    elaboro_firma_data_uri: Optional[str] = None,
    reviso_firma_data_uri: Optional[str] = None,
    *,
    conc_meta: Optional[Dict[str, Any]] = None,
    aprobo_firma_data_uri: Optional[str] = None,
    aprobo_interventoria_desde_config: bool = False,
    pie_fotos_contexto: Optional[str] = None,
):
    total_cant = sum(_sf(r.get("cantidad_total"), 0.0) for r in registros)

    # 1ª hoja: encabezado + barra consumen altura — objetivo ~26 filas para evitar 1 sola fila en página siguiente.
    ROWS_MEMORIA_PRIMERA_HOJA = 26
    ROWS_MEMORIA_SIGUIENTES = 30
    FOTOS_PER_PAGE = 6

    fotos = [r for r in registros if r.get("foto_url")]

    def encabezado():
        # CC-SUB-002: encabezado compacto en 3 filas (logo | título | código → contrato | sub | corte | período → ítem | descripción | und).
        # conc_meta: CC-SEM-002 / CC-MES-002 (sin sub/corte en metadatos).
        bd = "border:1px solid #9ca3af"
        titulo = "RESUMEN ACTIVIDADES CONCILIACIÓN CORTE SUBCONTRATISTA"
        codigo_h = _h("CC-SUB-002")
        num_contrato = _h(str(contrato.get("numero") or ""))
        sub_nom = _h(str(sub.get("razon_social") or ""))
        corte_lbl = _h(_corte_consecutivo_fmt(corte))
        periodo = f"{_h(_fd(corte.get('fecha_inicio')))} — {_h(_fd(corte.get('fecha_fin')))}"
        if conc_meta:
            titulo = str(conc_meta.get("titulo") or titulo)
            codigo_h = _h(str(conc_meta.get("codigo") or ""))
        it_num = _h(str(item_info.get("item_numero") or ""))
        it_desc = _h(_descripcion_memoria_compacta(item_info.get("item_descripcion")))
        it_und = _h(str(item_info.get("unidad") or ""))
        lbl = "font-size:6pt;color:#555;font-weight:normal;"
        val = "font-size:7pt;font-weight:bold;color:#1a1a2e;line-height:1.15;"
        # Descripción: +1 pt respecto a 5.5 pt anterior → 6.5 pt; interlineado ajustado para fila más baja.
        val_desc = (
            "font-size:6.5pt;font-weight:normal;color:#1a1a2e;line-height:1.04;"
            "text-transform:none;word-wrap:break-word;margin:1px 0 0 0;padding:0;display:block;"
        )
        fila1_h = "1.14cm"
        # Área útil del logo = altura de fila menos padding mínimo de celda.
        logo_box_h = "1.08cm"
        logo_html = _html_logo_contratista(contrato, compact=True, compact_box_height=logo_box_h)
        lbl_blk = "font-size:6pt;color:#555;font-weight:normal;display:block;margin:0;padding:0;line-height:1.05;"
        if conc_meta:
            cells = list(conc_meta.get("cells") or [])
            while len(cells) < 4:
                cells.append(("", ""))
            c4 = cells[:4]
            row2 = ""
            for lab, celval in c4:
                row2 += f"""<td style="width:25%;{bd};padding:3px 5px;vertical-align:top;">
<span style="{lbl}">{_h(str(lab))}</span><br/><span style="{val}">{_h(str(celval))}</span>
</td>"""
            return f"""<div class="mem002-head-wrap"><table class="w100" cellspacing="0" cellpadding="0" style="border-collapse:collapse;table-layout:fixed;width:100%;margin-bottom:3px;border:1px solid #9ca3af;">
<tr>
<td style="width:17%;{bd};padding:1px 2px;vertical-align:middle;text-align:center;height:{fila1_h};min-height:{fila1_h};">
{logo_html}
</td>
<td style="width:65%;{bd};padding:2px 6px;vertical-align:middle;text-align:center;height:{fila1_h};min-height:{fila1_h};">
<div style="font-size:6.8pt;font-weight:bold;color:#111827;text-transform:uppercase;line-height:1.08;">{_h(titulo)}</div>
</td>
<td style="width:18%;{bd};padding:2px;vertical-align:middle;text-align:center;height:{fila1_h};min-height:{fila1_h};">
<div style="font-size:10pt;font-weight:bold;color:#1e40af;line-height:1;">{codigo_h}</div>
<div style="font-size:6pt;color:#64748b;margin-top:1px;">CCD</div>
</td>
</tr>
<tr>
<td colspan="3" style="padding:0;">
<table class="w100" cellspacing="0" cellpadding="0" style="border-collapse:collapse;table-layout:fixed;width:100%;">
<tr>
{row2}
</tr>
</table>
</td>
</tr>
<tr>
<td colspan="3" style="padding:0;">
<table class="w100" cellspacing="0" cellpadding="0" style="border-collapse:collapse;table-layout:fixed;width:100%;">
<tr>
<td style="width:13%;{bd};padding:1px 4px;vertical-align:top;">
<span style="{lbl_blk}">ÍTEM</span><span style="{val};display:block;margin:1px 0 0 0;padding:0;">{it_num}</span>
</td>
<td style="width:74%;{bd};padding:1px 4px;vertical-align:top;">
<span style="{lbl_blk}">DESCRIPCIÓN</span><span style="{val_desc}">{it_desc}</span>
</td>
<td style="width:13%;{bd};padding:1px 4px;vertical-align:top;">
<span style="{lbl_blk}">UNIDAD</span><span style="{val};display:block;margin:1px 0 0 0;padding:0;">{it_und}</span>
</td>
</tr>
</table>
</td>
</tr>
</table></div>"""
        return f"""<div class="mem002-head-wrap"><table class="w100" cellspacing="0" cellpadding="0" style="border-collapse:collapse;table-layout:fixed;width:100%;margin-bottom:3px;border:1px solid #9ca3af;">
<tr>
<td style="width:17%;{bd};padding:1px 2px;vertical-align:middle;text-align:center;height:{fila1_h};min-height:{fila1_h};">
{logo_html}
</td>
<td style="width:65%;{bd};padding:2px 6px;vertical-align:middle;text-align:center;height:{fila1_h};min-height:{fila1_h};">
<div style="font-size:6.8pt;font-weight:bold;color:#111827;text-transform:uppercase;line-height:1.08;">{_h(titulo)}</div>
</td>
<td style="width:18%;{bd};padding:2px;vertical-align:middle;text-align:center;height:{fila1_h};min-height:{fila1_h};">
<div style="font-size:10pt;font-weight:bold;color:#1e40af;line-height:1;">{codigo_h}</div>
<div style="font-size:6pt;color:#64748b;margin-top:1px;">CCD</div>
</td>
</tr>
<tr>
<td colspan="3" style="padding:0;">
<table class="w100" cellspacing="0" cellpadding="0" style="border-collapse:collapse;table-layout:fixed;width:100%;">
<tr>
<td style="width:25%;{bd};padding:3px 5px;vertical-align:top;">
<span style="{lbl}">CONTRATO</span><br/><span style="{val}">{num_contrato}</span>
</td>
<td style="width:25%;{bd};padding:3px 5px;vertical-align:top;">
<span style="{lbl}">SUB CONTRATISTA</span><br/><span style="{val}">{sub_nom}</span>
</td>
<td style="width:25%;{bd};padding:3px 5px;vertical-align:top;">
<span style="{lbl}">CORTE N°</span><br/><span style="{val}">{corte_lbl}</span>
</td>
<td style="width:25%;{bd};padding:3px 5px;vertical-align:top;">
<span style="{lbl}">PERÍODO</span><br/><span style="{val}">{periodo}</span>
</td>
</tr>
</table>
</td>
</tr>
<tr>
<td colspan="3" style="padding:0;">
<table class="w100" cellspacing="0" cellpadding="0" style="border-collapse:collapse;table-layout:fixed;width:100%;">
<tr>
<td style="width:13%;{bd};padding:1px 4px;vertical-align:top;">
<span style="{lbl_blk}">ÍTEM</span><span style="{val};display:block;margin:1px 0 0 0;padding:0;">{it_num}</span>
</td>
<td style="width:74%;{bd};padding:1px 4px;vertical-align:top;">
<span style="{lbl_blk}">DESCRIPCIÓN</span><span style="{val_desc}">{it_desc}</span>
</td>
<td style="width:13%;{bd};padding:1px 4px;vertical-align:top;">
<span style="{lbl_blk}">UNIDAD</span><span style="{val};display:block;margin:1px 0 0 0;padding:0;">{it_und}</span>
</td>
</tr>
</table>
</td>
</tr>
</table></div>"""

    chunks_foto = [fotos[i : i + FOTOS_PER_PAGE] for i in range(0, len(fotos), FOTOS_PER_PAGE)]
    chunks_reg = _chunks_memoria_detalle(registros, ROWS_MEMORIA_PRIMERA_HOJA, ROWS_MEMORIA_SIGUIENTES)

    thead_detalle = """<tr>
            <th class="data-th" style="width:4%">N°</th>
            <th class="data-th" style="width:5.5%">ABS INI</th>
            <th class="data-th" style="width:5.5%">ABS FIN</th>
            <th class="data-th" style="width:7%">PK ID</th>
            <th class="data-th" style="width:8%">COSTADO</th>
            <th class="data-th" style="width:7%">LONG</th>
            <th class="data-th" style="width:7%">ANCHO</th>
            <th class="data-th" style="width:7%">ESP</th>
            <th class="data-th" style="width:9%">CANT</th>
            <th class="data-th" style="width:9%">CANT TOT</th>
            <th class="data-th" style="width:38%">OBSERVACIÓN</th>
        </tr>"""

    body = ""
    row_offset = 0
    for ci, chunk in enumerate(chunks_reg):
        if ci > 0:
            body += "<pdf:nextpage />"
        if ci == 0:
            body += encabezado()
        body += '<div class="section-bar mem002-section">DETALLE DE CANTIDADES APROBADAS</div>'
        body += f"""<table class="w100 mem002-detail" cellspacing="0" cellpadding="0">{thead_detalle}"""

        for j, r in enumerate(chunk):
            i = row_offset + j
            cls = "even" if i % 2 == 0 else "odd"
            obs = r.get("observacion") or ""
            fn = r.get("foto_numero")
            if fn:
                obs = f"{obs} [Foto {fn}]".strip()
            obs = _descripcion_memoria_compacta(obs)
            pkv = (r.get("pk_ids") or {}).get("pk_id")
            body += f"""<tr class="{cls}">
                <td class="data-td" style="text-align:center">{_h(r.get('numero_registro',''))}</td>
                <td class="data-td" style="text-align:center">{_h(r.get('abs_inicio') or '—')}</td>
                <td class="data-td" style="text-align:center">{_h(r.get('abs_final') or '—')}</td>
                <td class="data-td" style="text-align:center">{_h(pkv if pkv is not None else '—')}</td>
                <td class="data-td" style="text-align:center">{_h(r.get('calzada') or '—')}</td>
                <td class="data-td" style="text-align:right">{_fn(r.get('longitud'))}</td>
                <td class="data-td" style="text-align:right">{_fn(r.get('ancho'))}</td>
                <td class="data-td" style="text-align:right">{_fn(r.get('espesor'))}</td>
                <td class="data-td" style="text-align:right">{_fn(r.get('cantidad'))}</td>
                <td class="data-td" style="text-align:right;font-weight:bold">{_fn(r.get('cantidad_total'))}</td>
                <td class="data-td mem002-obs">{_h((obs or '')[:500])}</td>
            </tr>"""

        row_offset += len(chunk)
        body += "</table>"

    # Fila de total en tabla aparte (sin colspan): xhtml2pdf suele aplastar OBS/CANT TOT en la última página del detalle.
    body += f"""<table class="w100 mem002-total-wrap" cellspacing="0" cellpadding="0">
        <tr>
            <td class="total-td" style="width:60%;text-align:right;padding-right:8px">CANTIDAD TOTAL DEL ÍTEM</td>
            <td class="total-td" style="width:9%;text-align:right">{_fn(total_cant)}</td>
            <td class="total-td" style="width:31%">&nbsp;</td>
        </tr>
    </table>"""

    body += '<div class="doc-footer">Documento institucional de control interno. Prohibida su reproduccion parcial o total sin autorizacion escrita.</div>'

    # Firmas justo después del detalle (antes de fotos): visibles en vista previa sin pasar el anexo gráfico.
    body += _html_mem002_bloque_firmas(
        sub,
        firma_cfg,
        elaboro_firma_data_uri,
        reviso_firma_data_uri,
        aprobo_firma_data_uri=aprobo_firma_data_uri,
        aprobo_interventoria_desde_config=aprobo_interventoria_desde_config,
    )

    # Páginas de fotos después del bloque de firmas (sin encabezado repetido).
    if pie_fotos_contexto:
        pie_foto = _h(pie_fotos_contexto)
    else:
        pie_foto = f'Corte N° {_h(corte.get("consecutivo",""))}'
    for foto_chunk in chunks_foto:
        body += '<pdf:nextpage />'
        body += f'<div class="section-bar mem002-section">REGISTRO FOTOGRÁFICO — ÍTEM {_h(item_info.get("item_numero",""))} | {pie_foto}</div>'
        body += '<table class="w100 mem002-foto-grid">'
        for row_start in range(0, len(foto_chunk), 3):
            body += "<tr>"
            fila = foto_chunk[row_start:row_start+3]
            for r in fila:
                obs_f = _descripcion_memoria_compacta((r.get("observacion") or "")[:120])
                fu = (r.get("foto_url") or "").strip()
                if fu:
                    body += f"""<td style="width:33%;text-align:center;padding:8px;vertical-align:top">
                    <img src="{_h(fu)}" style="max-width:155px;max-height:115px;border:1px solid #dee2e6"/>
                    <div class="foto-caption">Foto {_h(r.get('foto_numero',''))} — Reg. {_h(r.get('numero_registro',''))}</div>
                    <div style="font-size:6pt;color:#666;margin-top:2px">{_h(obs_f)}</div>
                </td>"""
                else:
                    body += """<td style="width:33%;text-align:center;padding:8px;vertical-align:top;color:#888">Sin foto</td>"""
            # Completar fila si tiene menos de 3
            for _ in range(3 - len(fila)):
                body += '<td style="width:33%"></td>'
            body += "</tr>"
        body += "</table>"
        body += '<div class="doc-footer">Documento institucional de control interno. Prohibida su reproduccion parcial o total sin autorizacion escrita.</div>'

    return body


def _html_memoria_item(
    contrato,
    sub,
    corte,
    item_info,
    registros,
    usuario_nombre,
    usuario_cargo,
    firma_cfg: Optional[Dict[str, Any]] = None,
    elaboro_firma_data_uri: Optional[str] = None,
    reviso_firma_data_uri: Optional[str] = None,
    *,
    conc_meta: Optional[Dict[str, Any]] = None,
    aprobo_firma_data_uri: Optional[str] = None,
    aprobo_interventoria_desde_config: bool = False,
    pie_fotos_contexto: Optional[str] = None,
    estilo_formato_codigo: str = CODIGO_FORMATO_CCD_CC_SUB_002,
):
    fc = firma_cfg or {}
    est = _merge_estilo_pdf(fc.get("estilo_pdf"), estilo_formato_codigo)
    estilo_css = _memoria_pdf_estilo_css(est)
    inner = _html_memoria_item_body(
        contrato,
        sub,
        corte,
        item_info,
        registros,
        usuario_nombre,
        usuario_cargo,
        firma_cfg,
        elaboro_firma_data_uri=elaboro_firma_data_uri,
        reviso_firma_data_uri=reviso_firma_data_uri,
        conc_meta=conc_meta,
        aprobo_firma_data_uri=aprobo_firma_data_uri,
        aprobo_interventoria_desde_config=aprobo_interventoria_desde_config,
        pie_fotos_contexto=pie_fotos_contexto,
    )
    return _wrap_memoria_item_html(inner, estilo_css)
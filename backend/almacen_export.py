"""
Exportación Excel — inventario Almacén.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import List
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from almacen_service import list_inventario

_FILL_HDR = PatternFill("solid", fgColor="0077B6")
_FONT_HDR = Font(bold=True, color="FFFFFF", size=10)
_FONT_TITLE = Font(bold=True, size=14, color="0077B6")
_SIDE = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
_AL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_AL_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_AL_RIGHT = Alignment(horizontal="right", vertical="center")

_SEMAFORO_LABEL = {"verde": "Dentro de presupuesto", "amarillo": "Cerca del límite", "rojo": "Superado"}
_BOGOTA = ZoneInfo("America/Bogota")


def build_inventario_xlsx(contrato_id: int, contrato_numero: str = "") -> bytes:
    rows = list_inventario(contrato_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    titulo = f"Inventario de Almacén — {contrato_numero or f'Contrato {contrato_id}'}"
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = _FONT_TITLE
    generado = datetime.now(_BOGOTA).strftime("%d/%m/%Y %H:%M")
    ws.cell(row=2, column=1, value=f"Generado: {generado}")

    headers = [
        "Capítulo", "Ítem", "Material", "Unidad", "Stock disponible",
        "Presupuestado", "Ingresado acum.", "Semáforo",
    ]
    hr = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.fill = _FILL_HDR
        cell.font = _FONT_HDR
        cell.alignment = _AL_CENTER
        cell.border = _BORDER

    for i, r in enumerate(rows, hr + 1):
        ppto_id = r.get("presupuesto_id")
        vals = [
            r.get("capitulo") or "",
            r.get("item") or str(ppto_id or ""),
            r.get("material_descripcion") or "",
            r.get("unidad") or "",
            float(r.get("stock_disponible") or 0),
            float(r.get("cant_presupuestada") or 0),
            float(r.get("ingresado_acumulado") or 0),
            _SEMAFORO_LABEL.get(r.get("semaforo") or "", r.get("semaforo") or ""),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = _BORDER
            cell.alignment = _AL_RIGHT if col >= 5 and col <= 7 else _AL_LEFT

    widths = [12, 10, 36, 8, 14, 14, 14, 22]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

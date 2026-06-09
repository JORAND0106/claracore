"""
Fase 4 — Curva S: inversión acumulada baseline / vigente / ejecutado.
"""
from __future__ import annotations

import base64
import html
import io
import re
from calendar import monthrange
from collections import defaultdict
from datetime import date, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

from prog_obra_compare import fetch_compare_nodes, supplement_nodes_missing_presupuesto
from prog_obra_costos_presupuesto import (
    apply_ppto_cost_overlay,
    build_cost_overlay_maps,
    fetch_ppto_baseline_version_id,
    fetch_ppto_borrador_version_id,
    ppto_scope_direct_total,
    _fetch_agrupadores_meta,
)
from prog_obra_pk_filter import filter_nodes_by_pk, parse_pk_ids_param, parse_tramos_param
from prog_obra_service import fetch_baseline_version_id, fetch_borrador_activo, fetch_vigente_meta


def _parse_d(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, date):
        return val
    s = str(val).strip()[:10]
    if not s:
        return None
    try:
        y, m, d = s.split("-")
        return date(int(y), int(m), int(d))
    except (ValueError, TypeError):
        return None


def _month_key(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def _month_label(key: str) -> str:
    y, m = key.split("-")
    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    return f"{meses[int(m) - 1]} {y}"


def _linear_monthly_distribution(fi: date, ff: date, costo: float) -> Dict[str, float]:
    """Distribuye costo linealmente por mes entre fi y ff (inclusive)."""
    if costo <= 0 or not fi or not ff or ff < fi:
        return {}
    total_days = (ff - fi).days + 1
    if total_days <= 0:
        return {}
    out: Dict[str, float] = defaultdict(float)
    cur = fi
    while cur <= ff:
        mk = _month_key(cur)
        month_end = date(cur.year, cur.month, monthrange(cur.year, cur.month)[1])
        seg_end = min(ff, month_end)
        days = (seg_end - cur).days + 1
        out[mk] += costo * (days / total_days)
        cur = seg_end + timedelta(days=1)
    return dict(out)


def _scale_monthly_to_target(monthly: Dict[str, float], target_total: float) -> Tuple[Dict[str, float], float]:
    """Escala flujo mensual para alcanzar un total contractual (p. ej. V0 completo por tramo)."""
    cur = sum(float(v) for v in monthly.values())
    target = float(target_total)
    if cur <= 0 or target <= 0:
        return dict(monthly), cur
    if abs(target - cur) <= 0.01:
        return dict(monthly), cur
    factor = target / cur
    scaled: Dict[str, float] = {k: round(float(v) * factor, 2) for k, v in monthly.items()}
    diff = round(target - sum(scaled.values()), 2)
    if diff and scaled:
        last = sorted(scaled.keys())[-1]
        scaled[last] = round(scaled[last] + diff, 2)
    return scaled, target


def _nodes_with_ppto_costs(
    sb,
    version_id: str,
    contrato_id: int,
    version_ppto_id: Optional[str],
    pk_ids: Optional[Set[str]] = None,
    tramos: Optional[List[str]] = None,
    *,
    schedule_mode: str = "cpm",
) -> Dict[str, dict]:
    nodes = fetch_compare_nodes(sb, version_id, contrato_id, schedule_mode=schedule_mode)
    nodes = filter_nodes_by_pk(nodes, pk_ids)
    if not version_ppto_id:
        return nodes
    tramo = tramos[0] if tramos and len(tramos) == 1 else None
    ag_costs, item_costs = build_cost_overlay_maps(
        sb,
        contrato_id,
        str(version_ppto_id),
        tramo=tramo,
        tramos=tramos,
        pk_ids=pk_ids,
    )
    if schedule_mode == "programada":
        ag_meta = _fetch_agrupadores_meta(sb, contrato_id)
        nodes = supplement_nodes_missing_presupuesto(nodes, ag_costs, ag_meta)
    return apply_ppto_cost_overlay(nodes, ag_costs, item_costs)


def _aggregate_version_monthly(
    sb,
    version_id: str,
    contrato_id: int,
    version_ppto_id: Optional[str] = None,
    pk_ids: Optional[Set[str]] = None,
    tramos: Optional[List[str]] = None,
    *,
    schedule_mode: str = "cpm",
) -> Tuple[Dict[str, float], float]:
    nodes = _nodes_with_ppto_costs(
        sb, version_id, contrato_id, version_ppto_id, pk_ids=pk_ids, tramos=tramos,
        schedule_mode=schedule_mode,
    )
    monthly: Dict[str, float] = defaultdict(float)
    total = 0.0
    for n in nodes.values():
        costo = float(n.get("costo_programado") or 0)
        if costo <= 0:
            continue
        fi = n.get("fecha_inicio")
        ff = n.get("fecha_fin")
        if not fi or not ff:
            continue
        total += costo
        for mk, part in _linear_monthly_distribution(fi, ff, costo).items():
            monthly[mk] += part
    return dict(monthly), total


def _registro_aprobado(row: dict) -> bool:
    for n in (6, 5, 4, 3, 2, 1):
        st = (row.get(f"nivel{n}_estado") or "").strip().lower()
        if st in ("aprobado", "validación aprobada", "validacion aprobada"):
            return True
    return False


def _fecha_aprobacion(row: dict) -> Optional[date]:
    for n in (6, 5, 4, 3, 2, 1):
        st = (row.get(f"nivel{n}_estado") or "").strip().lower()
        if st in ("aprobado", "validación aprobada", "validacion aprobada"):
            return _parse_d(row.get(f"nivel{n}_fecha"))
    return None


def _fetch_ejecutado_mensual(sb, contrato_id: int) -> Tuple[Dict[str, float], float]:
    """Ejecutado real desde so_registros aprobados, agrupado por mes de aprobación."""
    monthly: Dict[str, float] = defaultdict(float)
    total = 0.0
    off = 0
    cols = (
        "costo_directo,cantidad_total,vlr_unitario,"
        "nivel1_estado,nivel1_fecha,nivel2_estado,nivel2_fecha,"
        "nivel3_estado,nivel3_fecha,nivel4_estado,nivel4_fecha,"
        "nivel5_estado,nivel5_fecha,nivel6_estado,nivel6_fecha"
    )
    while True:
        batch = (
            sb.table("so_registros")
            .select(cols)
            .eq("contrato_id", int(contrato_id))
            .range(off, off + 999)
            .execute()
            .data
            or []
        )
        for r in batch:
            if not _registro_aprobado(r):
                continue
            costo = float(r.get("costo_directo") or 0)
            if costo <= 0:
                cant = float(r.get("cantidad_total") or 0)
                vlr = float(r.get("vlr_unitario") or 0)
                costo = cant * vlr
            if costo <= 0:
                continue
            fd = _fecha_aprobacion(r)
            if not fd:
                continue
            mk = _month_key(fd)
            monthly[mk] += costo
            total += costo
        if len(batch) < 1000:
            break
        off += 1000
    return dict(monthly), total


def _detalle_pk_mensual(
    sb,
    version_id: str,
    contrato_id: int,
    version_ppto_id: Optional[str] = None,
    pk_ids: Optional[Set[str]] = None,
    tramos: Optional[List[str]] = None,
    *,
    schedule_mode: str = "cpm",
) -> List[dict]:
    from prog_obra_costos_presupuesto import _fetch_agrupadores_meta

    nodes = _nodes_with_ppto_costs(
        sb, version_id, contrato_id, version_ppto_id, pk_ids=pk_ids, tramos=tramos,
        schedule_mode=schedule_mode,
    )
    ag_meta = _fetch_agrupadores_meta(sb, contrato_id)
    out: List[dict] = []
    for n in nodes.values():
        costo = float(n.get("costo_programado") or 0)
        fi = n.get("fecha_inicio")
        ff = n.get("fecha_fin")
        if costo <= 0 or not fi or not ff:
            continue
        ag_raw = n.get("agrupador_id")
        meta = {}
        if ag_raw is not None:
            try:
                meta = ag_meta.get(int(ag_raw)) or {}
            except (TypeError, ValueError):
                meta = {}
        wbs = str(n.get("codigo_wbs") or meta.get("codigo_wbs") or n.get("label") or "").strip()
        nombre = str(meta.get("nombre") or wbs or "").strip()
        dist = _linear_monthly_distribution(fi, ff, costo)
        out.append(
            {
                "pk_id": n.get("pk_id"),
                "capitulo": n.get("capitulo"),
                "agrupador_id": n.get("agrupador_id"),
                "label": n.get("label"),
                "codigo_wbs": wbs,
                "agrupador_nombre": nombre,
                "fecha_inicio": fi,
                "fecha_fin": ff,
                "costo_total": round(costo, 2),
                "distribucion_mensual": {k: round(v, 2) for k, v in dist.items()},
            }
        )
    return out


def _resolve_version_ppto_id(sb, contrato_id: int, version_ppto_id: Optional[str]) -> Optional[str]:
    vid = (version_ppto_id or "").strip()
    if vid:
        return vid
    return fetch_ppto_borrador_version_id(sb, contrato_id)


def _resolve_ppto_vigente_curva_id(sb, contrato_id: int) -> Optional[str]:
    """Presupuesto vigente en edición (borrador); la línea Vigente de la curva S siempre usa esto."""
    return fetch_ppto_borrador_version_id(sb, contrato_id)


def _apply_ppto_scope_total_scale(
    sb,
    contrato_id: int,
    monthly: Dict[str, float],
    total: float,
    version_ppto_id: Optional[str],
    pk_ids: Optional[Set[str]],
    tramos_list: Optional[List[str]],
) -> Tuple[Dict[str, float], float]:
    """Ajusta flujo mensual al total contractual del alcance (comparación de presupuesto)."""
    ppto_id = (version_ppto_id or "").strip()
    if not ppto_id or total <= 0:
        return monthly, total
    tramo_one = tramos_list[0] if tramos_list and len(tramos_list) == 1 else None
    scope_total = ppto_scope_direct_total(
        sb,
        contrato_id,
        ppto_id,
        tramo=tramo_one,
        tramos=tramos_list,
        pk_ids=pk_ids,
    )
    if scope_total > 0 and abs(scope_total - total) > 0.01:
        return _scale_monthly_to_target(monthly, scope_total)
    return monthly, total


def build_curva_s(
    sb,
    contrato_id: int,
    baseline_id: Optional[str] = None,
    target_id: Optional[str] = None,
    version_ppto_id: Optional[str] = None,
    pk_ids: Optional[str] = None,
    tramos: Optional[str] = None,
) -> dict:
    pk_set = parse_pk_ids_param(pk_ids)
    tramos_list = parse_tramos_param(tramos)
    bid = (baseline_id or fetch_baseline_version_id(sb, contrato_id) or "").strip() or None
    tid = (target_id or "").strip()
    if not tid:
        borrador = fetch_borrador_activo(sb, contrato_id)
        if borrador and borrador.get("id"):
            tid = str(borrador["id"])
        else:
            vid, _ = fetch_vigente_meta(sb, contrato_id)
            tid = str(vid) if vid else ""
    tid = tid.strip() or None
    if not tid and not bid:
        raise ValueError("No hay versión de programación para la curva S")

    ppto_vigente_id = _resolve_version_ppto_id(sb, contrato_id, version_ppto_id)
    ppto_baseline_id = fetch_ppto_baseline_version_id(sb, contrato_id) or ppto_vigente_id

    if bid:
        base_m, base_total = _aggregate_version_monthly(
            sb, str(bid), contrato_id, ppto_baseline_id, pk_ids=pk_set, tramos=tramos_list,
            schedule_mode="programada",
        )
        base_m, base_total = _apply_ppto_scope_total_scale(
            sb, contrato_id, base_m, base_total, ppto_baseline_id, pk_set, tramos_list,
        )
    else:
        base_m, base_total = {}, 0.0

    if tid:
        tgt_m, tgt_total = _aggregate_version_monthly(
            sb, str(tid), contrato_id, ppto_vigente_id, pk_ids=pk_set, tramos=tramos_list,
            schedule_mode="cpm",
        )
    elif bid:
        tgt_m, tgt_total = base_m, base_total
        tid = bid
    else:
        tgt_m, tgt_total = {}, 0.0

    ej_m, ej_total = _fetch_ejecutado_mensual(sb, contrato_id)

    all_months = sorted(set(base_m.keys()) | set(tgt_m.keys()) | set(ej_m.keys()))
    if not all_months:
        all_months = [_month_key(date.today())]

    rows: List[dict] = []
    acc_b = acc_t = acc_e = 0.0
    hoy = date.today()
    hoy_key = _month_key(hoy)

    for mk in all_months:
        acc_b += float(base_m.get(mk, 0))
        acc_t += float(tgt_m.get(mk, 0))
        acc_e += float(ej_m.get(mk, 0))
        delta_v = ((acc_t - acc_b) / acc_b * 100) if acc_b > 0 else 0.0
        delta_e = ((acc_e - acc_b) / acc_b * 100) if acc_b > 0 else 0.0
        rows.append(
            {
                "mes": mk,
                "mes_label": _month_label(mk),
                "baseline_mes": round(float(base_m.get(mk, 0)), 2),
                "vigente_mes": round(float(tgt_m.get(mk, 0)), 2),
                "ejecutado_mes": round(float(ej_m.get(mk, 0)), 2),
                "baseline_acum": round(acc_b, 2),
                "vigente_acum": round(acc_t, 2),
                "ejecutado_acum": round(acc_e, 2),
                "delta_vigente_pct": round(delta_v, 1),
                "delta_ejecutado_pct": round(delta_e, 1),
            }
        )

    prog_a_fecha = acc_t if hoy_key in base_m or hoy_key in tgt_m else acc_t
    ej_a_fecha = acc_e
    presupuesto_total = max(base_total, tgt_total)
    pct_prog = (prog_a_fecha / presupuesto_total * 100) if presupuesto_total > 0 else 0.0
    pct_ej = (ej_a_fecha / presupuesto_total * 100) if presupuesto_total > 0 else 0.0
    desv = ej_a_fecha - prog_a_fecha
    desv_pct = (desv / prog_a_fecha * 100) if prog_a_fecha > 0 else 0.0

    detalle_baseline = (
        _detalle_pk_mensual(
            sb, str(bid), contrato_id, ppto_baseline_id, pk_ids=pk_set, tramos=tramos_list,
            schedule_mode="programada",
        )
        if bid
        else []
    )
    detalle_vigente = (
        _detalle_pk_mensual(
            sb, str(tid), contrato_id, ppto_vigente_id, pk_ids=pk_set, tramos=tramos_list,
            schedule_mode="cpm",
        )
        if tid
        else []
    )

    return {
        "baseline_id": str(bid) if bid else None,
        "target_id": str(tid) if tid else None,
        "version_ppto_id": ppto_vigente_id,
        "version_ppto_baseline_id": ppto_baseline_id,
        "indicadores": {
            "presupuesto_total": round(presupuesto_total, 2),
            "programado_a_fecha": round(prog_a_fecha, 2),
            "programado_pct": round(pct_prog, 1),
            "ejecutado_a_fecha": round(ej_a_fecha, 2),
            "ejecutado_pct": round(pct_ej, 1),
            "desviacion_valor": round(desv, 2),
            "desviacion_pct": round(desv_pct, 1),
        },
        "meses": rows,
        "detalle_pk": {
            "baseline": detalle_baseline,
            "vigente": detalle_vigente,
        },
    }


MAX_ESCENARIOS_CURVA_S = 5


def build_curva_s_escenarios(
    sb,
    contrato_id: int,
    version_prog_id: str,
    version_ppto_ids: List[str],
) -> dict:
    """Comparación de flujo de inversión con el mismo cronograma y distintos presupuestos."""
    if not version_ppto_ids:
        raise ValueError("version_ppto_ids requerido")
    if len(version_ppto_ids) > MAX_ESCENARIOS_CURVA_S:
        raise ValueError(f"Máximo {MAX_ESCENARIOS_CURVA_S} versiones de presupuesto por comparación")

    from prog_obra_costos_presupuesto import assert_ppto_version_contrato

    series: List[dict] = []
    all_months: set = set()
    for ppto_id in version_ppto_ids:
        vrow = assert_ppto_version_contrato(sb, contrato_id, str(ppto_id))
        monthly, total = _aggregate_version_monthly(
            sb, str(version_prog_id), contrato_id, str(ppto_id)
        )
        all_months.update(monthly.keys())
        series.append(
            {
                "version_ppto_id": str(ppto_id),
                "numero_version": int(vrow.get("numero_version") or 0),
                "etiqueta": (vrow.get("etiqueta") or "").strip(),
                "es_vigente_aprobada": bool(vrow.get("es_vigente_aprobada")),
                "es_vigente": bool(vrow.get("es_vigente")),
                "costo_total": round(total, 2),
                "distribucion_mensual": {k: round(v, 2) for k, v in monthly.items()},
            }
        )

    if not all_months:
        all_months = {_month_key(date.today())}

    meses_sorted = sorted(all_months)
    acc_by_ppto: Dict[str, float] = {s["version_ppto_id"]: 0.0 for s in series}
    filas: List[dict] = []
    for mk in meses_sorted:
        row: dict = {"mes": mk, "mes_label": _month_label(mk), "series": {}}
        for s in series:
            pid = s["version_ppto_id"]
            acc_by_ppto[pid] += float(s["distribucion_mensual"].get(mk, 0))
            row["series"][pid] = round(acc_by_ppto[pid], 2)
        filas.append(row)

    return {
        "version_prog_id": str(version_prog_id),
        "escenarios": series,
        "meses": filas,
    }


def _cap_sort_key(c: str) -> tuple:
    s = (c or "").strip()
    m = re.match(r"^(\d+)", s)
    if m:
        return (0, int(m.group(1)), s)
    return (1, 0, s)


def _h(val: Any) -> str:
    return html.escape(str(val if val is not None else ""))


def _fmt_duracion(val: Any) -> str:
    if val is None or val == "":
        return "—"
    try:
        return str(int(val))
    except (TypeError, ValueError):
        return "—"


def fetch_cronograma_pdf_tree(
    sb,
    version_id: str,
    contrato_id: int,
    pk_ids: Optional[Set[str]] = None,
) -> List[dict]:
    """Árbol PK → capítulo → agrupador WBS → ítems para PDF cronograma."""
    from prog_obra_costos_presupuesto import _fetch_agrupadores_meta
    from prog_obra_service import _listado_agrupador_por_item

    vid = (version_id or "").strip()
    if not vid:
        return []

    act_rows = (
        sb.table("prog_actividades")
        .select(
            "pk_id,capitulo,item,fecha_inicio,fecha_fin_calculada,duracion_dias_habiles,"
            "agrupador_id,codigo_wbs"
        )
        .eq("version_id", vid)
        .eq("contrato_id", int(contrato_id))
        .execute()
        .data
        or []
    )

    ag_meta = _fetch_agrupadores_meta(sb, contrato_id)
    _, desc_lp = _listado_agrupador_por_item(sb, contrato_id)

    headers: Dict[Tuple[str, str, int], dict] = {}
    items_map: Dict[Tuple[str, str, int], List[dict]] = defaultdict(list)

    for r in act_rows:
        pk = str(r.get("pk_id") or "").strip()
        if pk_ids and pk not in pk_ids:
            continue
        cap = str(r.get("capitulo") or "").strip()
        ag_raw = r.get("agrupador_id")
        if not pk or not cap or ag_raw is None:
            continue
        fi = r.get("fecha_inicio")
        if not fi:
            continue
        ag_id = int(ag_raw)
        key = (pk, cap, ag_id)
        meta = ag_meta.get(ag_id) or {}
        wbs = str(r.get("codigo_wbs") or meta.get("codigo_wbs") or f"AG{ag_id}").strip()
        item = str(r.get("item") or "").strip()
        row_base = {
            "item": item,
            "descripcion": (desc_lp.get((cap, item)) or "").strip(),
            "fecha_inicio": str(fi)[:10],
            "fecha_fin": str(r.get("fecha_fin_calculada") or "")[:10] or "—",
            "duracion_dias_habiles": r.get("duracion_dias_habiles"),
        }
        if item == wbs or item.replace(" ", "") == wbs.replace(" ", ""):
            headers[key] = {
                **row_base,
                "codigo_wbs": wbs,
                "nombre": (meta.get("nombre") or wbs).strip(),
                "agrupador_id": ag_id,
                "orden": meta.get("orden") or 0,
            }
        else:
            items_map[key].append(row_base)

    if not headers:
        return []

    by_pk: Dict[str, Dict[str, List[dict]]] = defaultdict(lambda: defaultdict(list))
    for (pk, cap, ag_id), hdr in headers.items():
        hijos = sorted(
            items_map.get((pk, cap, ag_id), []),
            key=lambda x: (_cap_sort_key(x["item"]), x["item"]),
        )
        by_pk[pk][cap].append({**hdr, "items": hijos})

    tree: List[dict] = []
    for pk in sorted(by_pk.keys(), key=lambda x: (len(x), x)):
        cap_list = []
        for cap in sorted(by_pk[pk].keys(), key=_cap_sort_key):
            ags = sorted(
                by_pk[pk][cap],
                key=lambda a: (a.get("orden") or 0, a.get("codigo_wbs") or ""),
            )
            cap_list.append({"capitulo": cap, "agrupadores": ags})
        tree.append({"pk_id": pk, "capitulos": cap_list})
    return tree


CRONO_ROWS_PER_PAGE = 45
MESES_ROWS_PER_PDF_PAGE = 40
DETALLE_PK_MONTH_COLS = 14
CPM_ROWS_PER_PK_PAGE = 40

# Paleta PDF alineada al sistema (grises, azules y celestes pastel)
PDF_CLR = {
    "title": "#1e40af",
    "text": "#334155",
    "muted": "#64748b",
    "border": "#cbd5e1",
    "border_light": "#e2e8f0",
    "bg_page": "#f8fafc",
    "bg_kpi": "#eff6ff",
    "bg_header": "#dbeafe",
    "header_text": "#1e3a8a",
    "row_pk": "#e2e8f0",
    "row_cap": "#f1f5f9",
    "row_ag": "#e0f2fe",
    "row_crit": "#fee2e2",
    "row_final": "#dbeafe",
    "accent": "#2563eb",
    "accent_soft": "#93c5fd",
    "negative": "#b91c1c",
    "line_baseline": "#1e40af",
    "line_vigente": "#38bdf8",
    "line_ejecutado": "#64748b",
}


def _pdf_th_style(extra: str = "") -> str:
    return (
        f"background:{PDF_CLR['bg_header']};color:{PDF_CLR['header_text']};"
        f"padding:2px 4px;font-size:6pt;font-weight:bold;"
        f"border-bottom:0.25px solid {PDF_CLR['border']};{extra}"
    )


def _pdf_td(extra: str = "") -> str:
    return (
        f"padding:1px 3px;font-size:6pt;line-height:1.15;"
        f"border:0.25px solid {PDF_CLR['border']};{extra}"
    )


def _pdf_page_header_html(
    contrato: dict,
    title: str,
    *,
    subtitle: str = "",
    prog_meta: Optional[dict] = None,
    fecha_gen: str = "",
    pag: str = "",
) -> str:
    from informes import _html_logo_contratista

    c = PDF_CLR
    logo_html = _html_logo_contratista(contrato, compact=True, compact_box_height="0.85cm")
    prog_lbl = _prog_version_label(prog_meta)
    pag_html = f' &nbsp;|&nbsp; <span style="color:{c["muted"]};">{_h(pag)}</span>' if pag else ""
    sub_html = (
        f'<div style="font-size:7pt;color:{c["muted"]};margin-top:2px;">{_h(subtitle)}</div>'
        if subtitle
        else ""
    )
    return f"""
<table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;margin-bottom:4px;">
  <tr>
    <td style="width:14%;vertical-align:middle;text-align:center;border:0.25px solid {c['border']};padding:1px;">{logo_html}</td>
    <td style="width:62%;vertical-align:middle;text-align:center;border:0.25px solid {c['border']};padding:3px 5px;">
      <div style="font-size:10pt;font-weight:bold;color:{c['text']};text-transform:uppercase;line-height:1.15;">{_h(title)}</div>
      {sub_html}
    </td>
    <td style="width:24%;vertical-align:middle;text-align:center;border:0.25px solid {c['border']};padding:3px;">
      <div style="font-size:5.5pt;color:{c['muted']};text-transform:uppercase;font-weight:bold;">Versión programación</div>
      <div style="font-size:8pt;font-weight:bold;color:{c['accent']};margin-top:1px;">{_h(prog_lbl)}</div>
    </td>
  </tr>
</table>
<div style="font-size:6.5pt;color:{c['text']};margin-bottom:4px;padding:0 2px;">
  <strong>Generado:</strong> {_h(fecha_gen)}{pag_html}
</div>
"""


def _cpm_holgura_by_agrupador(cpm_export: Optional[dict]) -> Dict[Tuple[str, str, int], int]:
    out: Dict[Tuple[str, str, int], int] = {}
    for r in (cpm_export or {}).get("resultados") or []:
        pk = str(r.get("pk_id") or "").strip()
        cap = str(r.get("capitulo") or "").strip()
        ag_raw = r.get("agrupador_id")
        if not pk or not cap or ag_raw is None:
            continue
        try:
            ag_id = int(ag_raw)
        except (TypeError, ValueError):
            continue
        out[(pk, cap, ag_id)] = int(r.get("holgura_total") or 0)
    return out


def _fmt_chart_money(v: float) -> str:
    return f"${v:,.0f}"


def _chart_left_margin(max_y: float, font_size: float = 7.5) -> int:
    labels = [_fmt_chart_money(max_y * t / 5) for t in range(6)]
    max_chars = max(len(lb) for lb in labels)
    return max(58, int(max_chars * font_size * 0.62) + 14)


def _monotone_path_d(xs: List[float], ys: List[float]) -> str:
    """Path SVG con curva monótona (equivalente a Recharts type=\"monotone\")."""
    n = len(xs)
    if n == 0:
        return ""
    if n == 1:
        return f"M {xs[0]:.2f},{ys[0]:.2f}"
    if n == 2:
        return f"M {xs[0]:.2f},{ys[0]:.2f} L {xs[1]:.2f},{ys[1]:.2f}"

    dx = [xs[i + 1] - xs[i] for i in range(n - 1)]
    dy = [ys[i + 1] - ys[i] for i in range(n - 1)]
    slopes = [dy[i] / dx[i] if abs(dx[i]) > 1e-12 else 0.0 for i in range(n - 1)]

    tangents = [0.0] * n
    tangents[0] = slopes[0]
    tangents[-1] = slopes[-1]
    for i in range(1, n - 1):
        if slopes[i - 1] * slopes[i] <= 0:
            tangents[i] = 0.0
        else:
            tangents[i] = (slopes[i - 1] + slopes[i]) / 2

    for i in range(n - 1):
        if abs(slopes[i]) < 1e-12:
            tangents[i] = 0.0
            tangents[i + 1] = 0.0
        else:
            alpha = tangents[i] / slopes[i]
            beta = tangents[i + 1] / slopes[i]
            s = alpha * alpha + beta * beta
            if s > 9:
                t = 3 / (s**0.5)
                tangents[i] = t * alpha * slopes[i]
                tangents[i + 1] = t * beta * slopes[i]

    parts = [f"M {xs[0]:.2f},{ys[0]:.2f}"]
    for i in range(n - 1):
        seg_dx = dx[i]
        c1x = xs[i] + seg_dx / 3
        c1y = ys[i] + tangents[i] * seg_dx / 3
        c2x = xs[i + 1] - seg_dx / 3
        c2y = ys[i + 1] - tangents[i + 1] * seg_dx / 3
        parts.append(f"C {c1x:.2f},{c1y:.2f} {c2x:.2f},{c2y:.2f} {xs[i + 1]:.2f},{ys[i + 1]:.2f}")
    return " ".join(parts)


CHART_RENDER_DPI = 180


def _build_curva_s_chart_svg(meses: List[dict], width: int = 920, height: int = 240) -> str:
    """Curva S en SVG generado en código a partir de meses (baseline/vigente/ejecutado acum.)."""
    c = PDF_CLR
    if not meses:
        return (
            f'<div style="height:100px;text-align:center;color:{c["muted"]};font-size:8pt;padding-top:40px;">'
            "Sin datos para el gráfico"
            "</div>"
        )

    ml, mr, mt, mb = 58, 20, 22, 36
    series = [
        ("Baseline", "baseline_acum", c["line_baseline"], 2.5),
        ("Vigente", "vigente_acum", c["line_vigente"], 2.0),
        ("Ejecutado", "ejecutado_acum", c["line_ejecutado"], 2.0),
    ]

    vals: List[float] = []
    for r in meses:
        for _, key, _, _ in series:
            try:
                vals.append(float(r.get(key) or 0))
            except (TypeError, ValueError):
                vals.append(0.0)
    max_y = max(vals) if vals else 1.0
    if max_y <= 0:
        max_y = 1.0
    max_y *= 1.08

    n = len(meses)
    ml = _chart_left_margin(max_y)
    pw = width - ml - mr
    ph = height - mt - mb

    def x_at(i: int) -> float:
        if n <= 1:
            return ml + pw / 2
        return ml + (i / (n - 1)) * pw

    def y_at(v: float) -> float:
        return mt + ph - (float(v) / max_y) * ph

    parts: List[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}">'
    ]

    # Leyenda
    lx = ml
    for label, _, color, _ in series:
        parts.append(f'<line x1="{lx}" y1="10" x2="{lx + 16}" y2="10" stroke="{color}" stroke-width="2.5"/>')
        parts.append(
            f'<text x="{lx + 20}" y="13" font-size="8" fill="{c["text"]}" font-family="Helvetica,Arial,sans-serif">'
            f"{_h(label)}</text>"
        )
        lx += 110

    # Cuadrícula Y y eje
    parts.append(
        f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt + ph:.1f}" stroke="{c["border"]}" stroke-width="0.8"/>'
    )
    parts.append(
        f'<line x1="{ml}" y1="{mt + ph:.1f}" x2="{ml + pw:.1f}" y2="{mt + ph:.1f}" '
        f'stroke="{c["border"]}" stroke-width="0.8"/>'
    )

    for t in range(6):
        y = mt + ph - (t / 5) * ph
        val = max_y * t / 5
        parts.append(
            f'<line x1="{ml}" y1="{y:.1f}" x2="{ml + pw:.1f}" y2="{y:.1f}" '
            f'stroke="{c["border_light"]}" stroke-width="0.8" stroke-dasharray="4,3"/>'
        )
        parts.append(
            f'<text x="{ml - 5}" y="{y + 3:.1f}" text-anchor="end" font-size="7.5" '
            f'fill="{c["muted"]}" font-family="Helvetica,Arial,sans-serif">'
            f"{_h(_fmt_chart_money(val))}</text>"
        )

    parts.append(
        f'<rect x="{ml}" y="{mt}" width="{pw:.1f}" height="{ph:.1f}" fill="{c["bg_page"]}" '
        f'stroke="none"/>'
    )

    for i, r in enumerate(meses):
        x = x_at(i)
        lbl = str(r.get("mes_label") or "")[:10]
        parts.append(
            f'<text x="{x:.1f}" y="{height - 8}" text-anchor="middle" font-size="7.5" '
            f'fill="{c["muted"]}" font-family="Helvetica,Arial,sans-serif">'
            f"{_h(lbl)}</text>"
        )

    for _, key, color, sw in series:
        xs: List[float] = []
        ys: List[float] = []
        for i, r in enumerate(meses):
            try:
                v = float(r.get(key) or 0)
            except (TypeError, ValueError):
                v = 0.0
            xs.append(x_at(i))
            ys.append(y_at(v))
        path_d = _monotone_path_d(xs, ys)
        if path_d:
            parts.append(
                f'<path d="{path_d}" fill="none" stroke="{color}" stroke-width="{sw}" '
                f'stroke-linejoin="round" stroke-linecap="round"/>'
            )

    parts.append("</svg>")
    return "".join(parts)


def _svg_to_png_b64(svg: str) -> str:
    """Rasteriza SVG a PNG base64 para incrustar en PDF (xhtml2pdf no dibuja SVG inline)."""
    from reportlab.graphics import renderPM
    from svglib.svglib import svg2rlg

    drawing = svg2rlg(io.BytesIO(svg.encode("utf-8")))
    if drawing is None:
        raise ValueError("No se pudo interpretar el SVG del gráfico")
    png = renderPM.drawToString(drawing, fmt="PNG", dpi=CHART_RENDER_DPI)
    return base64.b64encode(png).decode("ascii")


def _build_curva_s_chart_block(meses: List[dict]) -> str:
    """Contenedor del gráfico en página 1 del PDF."""
    c = PDF_CLR
    width, height = 920, 240
    if not meses:
        return (
            f'<div style="width:{width}px;height:{height}px;margin:4px 0 8px;text-align:center;'
            f'color:{c["muted"]};font-size:8pt;padding-top:80px;'
            f'background:{c["bg_page"]};border:1px solid {c["border_light"]};">'
            "Sin datos para el gráfico"
            "</div>"
        )
    svg = _build_curva_s_chart_svg(meses, width=width, height=height)
    b64 = _svg_to_png_b64(svg)
    return (
        f'<div style="width:100%;max-width:{width}px;height:{height}px;margin:4px 0 8px;overflow:hidden;'
        f'background:{c["bg_page"]};border:1px solid {c["border_light"]};">'
        f'<img src="data:image/png;base64,{b64}" width="100%" height="{height}" '
        f'style="display:block;max-width:100%;" />'
        f"</div>"
    )


def _flatten_cronograma_rows(tree: List[dict]) -> List[dict]:
    flat: List[dict] = []
    for pk_node in tree:
        pk = pk_node["pk_id"]
        flat.append({"kind": "pk", "pk_id": pk})
        for cap_node in pk_node.get("capitulos") or []:
            cap = cap_node["capitulo"]
            flat.append({"kind": "cap", "capitulo": cap, "pk_id": pk})
            for ag in cap_node.get("agrupadores") or []:
                flat.append({"kind": "ag", "pk_id": pk, "capitulo": cap, **ag})
                for it in ag.get("items") or []:
                    flat.append({"kind": "item", "pk_id": pk, "capitulo": cap, **it})
    return flat


def _chunk_cronograma_rows(flat: List[dict], per_page: int = CRONO_ROWS_PER_PAGE) -> List[List[dict]]:
    """Pagina cronograma: cada PK inicia en página nueva; el overflow del mismo PK continúa."""
    if not flat:
        return [[]]

    chunks: List[List[dict]] = []
    i = 0
    while i < len(flat):
        if flat[i].get("kind") != "pk":
            i += 1
            continue
        pk_block: List[dict] = []
        while i < len(flat):
            pk_block.append(flat[i])
            i += 1
            if i < len(flat) and flat[i].get("kind") == "pk":
                break
        start = 0
        while start < len(pk_block):
            chunks.append(pk_block[start : start + per_page])
            start += per_page
    return chunks if chunks else [[]]


def _html_footer_logo_contratista(contrato: dict) -> str:
    from informes import _html_logo_contratista

    return _html_logo_contratista(contrato, compact=True, compact_box_height="0.48cm")


def _ppto_version_label(ppto_meta: Optional[dict]) -> str:
    if not ppto_meta:
        return "—"
    n = ppto_meta.get("numero_version")
    et = (ppto_meta.get("etiqueta") or "").strip()
    if n is not None and et:
        return f"v{int(n)} · {et}"
    if et:
        return et
    if n is not None:
        return f"Versión {int(n)}"
    return "—"


def _prog_version_label(prog_meta: Optional[dict]) -> str:
    if not prog_meta:
        return "—"
    n = prog_meta.get("numero_version")
    tipo = (prog_meta.get("tipo") or "").strip()
    if n is not None and tipo:
        return f"v{int(n)} ({tipo})"
    if n is not None:
        return f"v{int(n)}"
    return tipo or "—"


def _resolve_export_pk_ids(
    sb,
    contrato_id: int,
    pk_ids: Optional[str],
    tramos: Optional[str],
) -> Optional[Set[str]]:
    pk_set = parse_pk_ids_param(pk_ids)
    if pk_set:
        return pk_set
    tramos_list = parse_tramos_param(tramos)
    if not tramos_list:
        return None
    from prog_obra_service import fetch_tramos_contrato

    names = {t.strip() for t in tramos_list if t.strip()}
    out: Set[str] = set()
    for tr in fetch_tramos_contrato(sb, contrato_id):
        if (tr.get("tramo") or "").strip() not in names:
            continue
        for p in tr.get("pk_ids") or []:
            ps = str(p).strip()
            if ps:
                out.add(ps)
    return out or None


def _cpm_pdf_row_bg(estado: str) -> str:
    c = PDF_CLR
    if estado == "Ruta crítica":
        return c["row_crit"]
    if estado == "Actividad final tramo":
        return c["row_final"]
    return "transparent"


def _chunk_list(items: List, size: int) -> List[List]:
    if not items:
        return [[]]
    return [items[i : i + size] for i in range(0, len(items), size)]


def _group_cpm_resultados_by_pk(resultados: List[dict]) -> Dict[str, List[dict]]:
    by_pk: Dict[str, List[dict]] = defaultdict(list)
    for r in resultados or []:
        pk = str(r.get("pk_id") or "").strip() or "—"
        by_pk[pk].append(r)
    for pk in by_pk:
        by_pk[pk].sort(
            key=lambda x: (
                _cap_sort_key(str(x.get("capitulo") or "")),
                str(x.get("agrupador_label") or ""),
            )
        )
    return by_pk


def _pk_cronograma_resumen(pk_node: dict) -> dict:
    fechas_ini: List[str] = []
    fechas_fin: List[str] = []
    n_ag = 0
    for cap_node in pk_node.get("capitulos") or []:
        for ag in cap_node.get("agrupadores") or []:
            n_ag += 1
            fi = str(ag.get("fecha_inicio") or "").strip()[:10]
            ff = str(ag.get("fecha_fin") or "").strip()[:10]
            if fi and fi != "—":
                fechas_ini.append(fi)
            if ff and ff != "—":
                fechas_fin.append(ff)
    return {
        "agrupadores": n_ag,
        "fecha_inicio": min(fechas_ini) if fechas_ini else "—",
        "fecha_fin": max(fechas_fin) if fechas_fin else "—",
    }


def _render_cpm_pk_pdf_page(
    contrato: dict,
    pk_id: str,
    resultados_chunk: List[dict],
    prog_meta: Optional[dict],
    fecha_gen: str,
    *,
    page_num: int,
    total_pages: int,
    pk_page_idx: int,
    pk_page_total: int,
) -> str:
    c = PDF_CLR
    th = _pdf_th_style
    td = _pdf_td
    pag = f"Página {page_num} de {total_pages}" if total_pages > 1 else ""
    pk_pag = f" · hoja {pk_page_idx} de {pk_page_total}" if pk_page_total > 1 else ""
    body_rows = ""
    if not resultados_chunk:
        body_rows = (
            f'<tr><td colspan="6" style="{td("text-align:center;color:" + c["muted"] + ";")}">'
            "Sin resultados CPM para este PK.</td></tr>"
        )
    for r in resultados_chunk:
        bg = _cpm_pdf_row_bg(str(r.get("estado_cpm") or ""))
        wbs = str(r.get("agrupador_label") or "").strip()
        body_rows += (
            f'<tr style="background:{bg};">'
            f'<td style="{td()}">{_h(r.get("capitulo"))}</td>'
            f'<td style="{td()}">{_h(wbs)}</td>'
            f'<td style="{td("text-align:center;")}">{_h(r.get("fecha_inicio_temprana"))}</td>'
            f'<td style="{td("text-align:center;")}">{_h(r.get("fecha_fin_temprana"))}</td>'
            f'<td style="{td("text-align:center;")}">{_h(r.get("holgura_total"))}</td>'
            f'<td style="{td("font-weight:600;")}">{_h(r.get("estado_cpm"))}</td>'
            f"</tr>"
        )
    header = _pdf_page_header_html(
        contrato,
        "Resultados CPM",
        subtitle=f"PK {_h(pk_id)}{pk_pag}",
        prog_meta=prog_meta,
        fecha_gen=fecha_gen,
        pag=pag,
    )
    return f"""
<div class="page-cronograma">
  <div class="frame-double-outer">
    <div class="frame-double-inner">
      {header}
      <table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;">
        <thead><tr>
          <th style="{th('text-align:left;width:8%;')}">Cap.</th>
          <th style="{th('text-align:left;width:34%;')}">Actividad WBS</th>
          <th style="{th('text-align:center;width:14%;')}">Inicio</th>
          <th style="{th('text-align:center;width:14%;')}">Fin</th>
          <th style="{th('text-align:center;width:10%;')}">Holgura</th>
          <th style="{th('text-align:center;width:20%;')}">Estado</th>
        </tr></thead>
        <tbody>{body_rows}</tbody>
      </table>
    </div>
  </div>
</div>
"""


def _render_dependencias_pdf_pages(
    contrato: dict,
    dependencias: List[dict],
    prog_meta: Optional[dict],
    fecha_gen: str,
    *,
    start_page: int,
) -> Tuple[str, int]:
    dep_chunks = _chunk_list(dependencias, 35) if dependencias else [[]]
    if not dep_chunks:
        dep_chunks = [[]]
    out = ""
    total = len(dep_chunks)
    for i, chunk in enumerate(dep_chunks, start=1):
        page_num = start_page + i - 1
        pag = f"Página {page_num}" if total > 1 else ""
        th = _pdf_th_style
        td = _pdf_td
        body_rows = ""
        if not chunk:
            body_rows = (
                f'<tr><td colspan="4" style="{td("text-align:center;color:" + PDF_CLR["muted"] + ";")}">'
                "Sin dependencias definidas.</td></tr>"
            )
        for d in chunk:
            lag = int(d.get("lag_dias") or 0)
            lag_lbl = f"{lag} día{'s' if lag != 1 else ''}" if lag else "0 días"
            body_rows += (
                f"<tr>"
                f'<td style="{td()}">{_h(d.get("origen_label"))}</td>'
                f'<td style="{td("text-align:center;font-weight:700;")}">{_h(d.get("tipo"))}</td>'
                f'<td style="{td("text-align:center;")}">{_h(lag_lbl)}</td>'
                f'<td style="{td()}">{_h(d.get("destino_label"))}</td>'
                f"</tr>"
            )
        header = _pdf_page_header_html(
            contrato,
            "Dependencias CPM",
            subtitle="Listado general",
            prog_meta=prog_meta,
            fecha_gen=fecha_gen,
            pag=f"{pag} ({i} de {total})" if total > 1 else pag,
        )
        out += f"""
<div class="page-cronograma">
  <div class="frame-double-outer">
    <div class="frame-double-inner">
      {header}
      <table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;">
        <thead><tr>
          <th style="{th('text-align:left;width:38%;')}">Origen</th>
          <th style="{th('text-align:center;width:8%;')}">Tipo</th>
          <th style="{th('text-align:center;width:10%;')}">Lag</th>
          <th style="{th('text-align:left;width:38%;')}">Destino</th>
        </tr></thead>
        <tbody>{body_rows}</tbody>
      </table>
    </div>
  </div>
</div>
"""
    return out, start_page + total - 1


def _render_cpm_export_pdf_html(
    contrato: dict,
    cpm_export: Optional[dict],
    prog_meta: Optional[dict],
    fecha_gen: str,
) -> str:
    export = cpm_export or {}
    resultados = export.get("resultados") or []
    dependencias = export.get("dependencias_generales") or export.get("dependencias") or []
    by_pk = _group_cpm_resultados_by_pk(resultados)
    pk_ids = sorted(by_pk.keys(), key=lambda x: (len(x), x))

    cpm_page_specs: List[Tuple[str, List[dict], int, int]] = []
    for pk in pk_ids:
        pk_rows = by_pk.get(pk) or []
        chunks = _chunk_list(pk_rows, CPM_ROWS_PER_PK_PAGE) or [[]]
        for idx, chunk in enumerate(chunks, start=1):
            cpm_page_specs.append((pk, chunk, idx, len(chunks)))

    dep_page_count = len(_chunk_list(dependencias, 35)) if dependencias else 1
    if not cpm_page_specs:
        cpm_page_specs = [("—", [], 1, 1)]
    grand_total = len(cpm_page_specs) + dep_page_count

    out = ""
    for page_num, (pk, chunk, pk_idx, pk_total) in enumerate(cpm_page_specs, start=1):
        out += _render_cpm_pk_pdf_page(
            contrato,
            pk,
            chunk,
            prog_meta,
            fecha_gen,
            page_num=page_num,
            total_pages=grand_total,
            pk_page_idx=pk_idx,
            pk_page_total=pk_total,
        )
    deps_html, _ = _render_dependencias_pdf_pages(
        contrato,
        dependencias,
        prog_meta,
        fecha_gen,
        start_page=len(cpm_page_specs) + 1,
    )
    return out + deps_html


def _write_cpm_excel_sheet(ws, resultados: List[dict]) -> None:
    from openpyxl.styles import Alignment, Font

    headers = [
        "PK",
        "Capítulo",
        "Agrupador",
        "Inicio temprano",
        "Fin temprano",
        "Holgura total",
        "Holgura libre",
        "Estado CPM",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="1E3A8A", size=9)
        cell.fill = _xl_solid_fill("DBEAFE")
        cell.border = _xl_thin_border()
    for ri, r in enumerate(resultados, start=2):
        vals = [
            r.get("pk_id"),
            r.get("capitulo"),
            r.get("agrupador_label"),
            r.get("fecha_inicio_temprana"),
            r.get("fecha_fin_temprana"),
            r.get("holgura_total"),
            r.get("holgura_libre"),
            r.get("estado_cpm"),
        ]
        estado = str(r.get("estado_cpm") or "")
        fill = None
        if estado == "Ruta crítica":
            fill = _xl_solid_fill("FEE2E2")
        elif estado == "Actividad final tramo":
            fill = _xl_solid_fill("DBEAFE")
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.border = _xl_thin_border()
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill
            if col >= 6:
                cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 22
    ws.freeze_panes = "A2"


def _write_dependencias_excel_sheet(ws, dependencias: List[dict]) -> None:
    from openpyxl.styles import Alignment, Font

    headers = ["Origen", "Tipo", "Lag (días hábiles)", "Destino"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="1E3A8A", size=9)
        cell.fill = _xl_solid_fill("DBEAFE")
        cell.border = _xl_thin_border()
    for ri, d in enumerate(dependencias, start=2):
        vals = [
            d.get("origen_label"),
            d.get("tipo"),
            d.get("lag_dias"),
            d.get("destino_label"),
        ]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.border = _xl_thin_border()
            cell.font = Font(size=9)
            if col == 2:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=9, bold=True)
            if col == 3:
                cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 42
    ws.freeze_panes = "A2"


def _render_cronograma_body_rows(rows: List[dict]) -> str:
    if not rows:
        return (
            '<tr><td colspan="5" style="padding:12px;text-align:center;color:#64748b;font-size:8pt;">'
            "Sin actividades programadas en esta versión."
            "</td></tr>"
        )
    out: List[str] = []
    c = PDF_CLR
    for r in rows:
        kind = r.get("kind")
        if kind == "pk":
            out.append(
                f'<tr style="background:{c["row_pk"]};">'
                f'<td colspan="5" style="padding:4px 6px;font-size:8pt;font-weight:bold;color:{c["text"]};">'
                f"PK {_h(r.get('pk_id'))}</td></tr>"
            )
        elif kind == "cap":
            out.append(
                f'<tr style="background:{c["row_cap"]};">'
                f'<td colspan="5" style="padding:3px 6px;font-size:7.5pt;font-weight:bold;color:{c["text"]};">'
                f"Capítulo {_h(r.get('capitulo'))}</td></tr>"
            )
        elif kind == "ag":
            nombre = f"{r.get('codigo_wbs') or ''} · {r.get('nombre') or ''}".strip(" ·")
            out.append(
                f"<tr style=\"background:{c['row_ag']};\">"
                f"<td style=\"padding:3px 6px;font-size:7.5pt;font-weight:bold;color:{c['text']};\">{_h(nombre)}</td>"
                f"<td style=\"padding:3px 6px;font-size:7.5pt;\">{_h(r.get('fecha_inicio'))}</td>"
                f"<td style=\"padding:3px 6px;font-size:7.5pt;\">{_h(r.get('fecha_fin'))}</td>"
                f"<td style=\"padding:3px 6px;font-size:7.5pt;text-align:center;\">{_h(_fmt_duracion(r.get('duracion_dias_habiles')))}</td>"
                f"<td style=\"padding:3px 6px;font-size:7pt;color:{c['muted']};\">Agrupador WBS</td>"
                f"</tr>"
            )
        elif kind == "item":
            desc = (r.get("descripcion") or "").strip()
            label = f"{r.get('item') or ''}"
            if desc:
                label = f"{label} · {desc}"
            out.append(
                f"<tr>"
                f"<td style=\"padding:3px 6px 3px 18px;font-size:7pt;\">{_h(label)}</td>"
                f"<td style=\"padding:3px 6px;font-size:7pt;\">{_h(r.get('fecha_inicio'))}</td>"
                f"<td style=\"padding:3px 6px;font-size:7pt;\">{_h(r.get('fecha_fin'))}</td>"
                f"<td style=\"padding:3px 6px;font-size:7pt;text-align:center;\">{_h(_fmt_duracion(r.get('duracion_dias_habiles')))}</td>"
                f"<td style=\"padding:3px 6px;font-size:7pt;color:#94a3b8;\">Ítem</td>"
                f"</tr>"
            )
    return "".join(out)


def _prog_estado_display(estado: Any) -> str:
    key = (estado or "").strip()
    labels = {
        "borrador": "Borrador",
        "en_validacion": "En validación",
        "sellada": "Sellada",
    }
    return labels.get(key, key or "—")


def _aggregate_cronograma_por_capitulo(cronograma: List[dict]) -> Dict[str, dict]:
    """Agrupa agrupadores WBS del cronograma PDF por capítulo (todos los PK)."""
    by_cap: Dict[str, dict] = {}
    for pk_node in cronograma or []:
        for cap_node in pk_node.get("capitulos") or []:
            cap = str(cap_node.get("capitulo") or "").strip()
            if not cap:
                continue
            rec = by_cap.setdefault(
                cap,
                {"agrupadores": 0, "fechas_ini": [], "fechas_fin": []},
            )
            for ag in cap_node.get("agrupadores") or []:
                rec["agrupadores"] += 1
                fi = str(ag.get("fecha_inicio") or "").strip()[:10]
                ff = str(ag.get("fecha_fin") or "").strip()[:10]
                if fi and fi != "—":
                    rec["fechas_ini"].append(fi)
                if ff and ff != "—":
                    rec["fechas_fin"].append(ff)
    return by_cap


def _dias_exceso_tras_horizonte(
    fin_act: date,
    fin_version: date,
    contrato_id: int,
    cache,
) -> int:
    if fin_act <= fin_version:
        return 0
    from prog_obra_calendar import count_dias_habiles_entre

    d_start = fin_version + timedelta(days=1)
    return count_dias_habiles_entre(contrato_id, d_start, fin_act, cache)


def _fetch_distinct_capitulos_presupuesto(
    sb,
    contrato_id: int,
    version_ppto_id: Optional[str],
    pk_ids: Optional[Set[str]],
    cronograma: List[dict],
) -> List[str]:
    caps: Set[str] = set()
    ppto_id = (version_ppto_id or "").strip()
    if ppto_id:
        from prog_obra_costos_presupuesto import fetch_ppto_items_version

        for r in fetch_ppto_items_version(sb, contrato_id, ppto_id, pk_ids=pk_ids):
            cap = str(r.get("capitulo") or "").strip()
            if cap:
                caps.add(cap)
    if not caps:
        for pk_node in cronograma or []:
            for cap_node in pk_node.get("capitulos") or []:
                cap = str(cap_node.get("capitulo") or "").strip()
                if cap:
                    caps.add(cap)
    return sorted(caps, key=_cap_sort_key)


def build_resumen_ejecutivo_data(
    sb,
    contrato_id: int,
    *,
    version_ppto_id: Optional[str] = None,
    pk_ids: Optional[Set[str]] = None,
    cronograma: Optional[List[dict]] = None,
    cpm_export: Optional[dict] = None,
    prog_meta: Optional[dict] = None,
) -> dict:
    """Datos para la página de resumen ejecutivo del PDF de programación."""
    from prog_obra_calendar import CalendarioNoHabilesCache
    from prog_obra_service import make_prog_calendar_loader

    crono = cronograma or []
    meta = prog_meta or {}
    export = cpm_export or {}
    resultados = export.get("resultados") or []
    tiene_cpm = len(resultados) > 0

    cache = CalendarioNoHabilesCache(make_prog_calendar_loader(sb))
    caps = _fetch_distinct_capitulos_presupuesto(
        sb, contrato_id, version_ppto_id, pk_ids, crono
    )
    agg = _aggregate_cronograma_por_capitulo(crono)

    capitulos_rows: List[dict] = []
    for cap in caps:
        rec = agg.get(cap) or {}
        n_ag = int(rec.get("agrupadores") or 0)
        if n_ag <= 0:
            capitulos_rows.append(
                {
                    "capitulo": cap,
                    "num_agrupadores": 0,
                    "fecha_inicio": None,
                    "fecha_fin": None,
                    "dias_habiles": None,
                    "sin_programar": True,
                }
            )
            continue
        fi = min(rec["fechas_ini"]) if rec.get("fechas_ini") else None
        ff = max(rec["fechas_fin"]) if rec.get("fechas_fin") else None
        d_hab = None
        d0, d1 = _parse_d(fi), _parse_d(ff)
        if d0 and d1:
            from prog_obra_calendar import count_dias_habiles_entre

            d_hab = count_dias_habiles_entre(contrato_id, d0, d1, cache)
        capitulos_rows.append(
            {
                "capitulo": cap,
                "num_agrupadores": n_ag,
                "fecha_inicio": fi,
                "fecha_fin": ff,
                "dias_habiles": d_hab,
                "sin_programar": False,
            }
        )

    criticos: List[dict] = []
    desfases: List[dict] = []
    if tiene_cpm:
        for r in resultados:
            if not r.get("es_ruta_critica"):
                continue
            criticos.append(
                {
                    "nombre": r.get("agrupador_label") or "—",
                    "fecha_inicio": str(r.get("fecha_inicio_temprana") or "")[:10] or "—",
                    "fecha_fin": str(r.get("fecha_fin_temprana") or "")[:10] or "—",
                    "holgura": int(r.get("holgura_total") or 0),
                }
            )
        fin_ver = _parse_d(meta.get("fecha_fin"))
        if fin_ver:
            for r in resultados:
                ft = _parse_d(r.get("fecha_fin_temprana"))
                if not ft or ft <= fin_ver:
                    continue
                dias = _dias_exceso_tras_horizonte(ft, fin_ver, contrato_id, cache)
                if dias <= 0:
                    continue
                desfases.append(
                    {
                        "nombre": r.get("agrupador_label") or "—",
                        "fecha_fin": str(r.get("fecha_fin_temprana") or "")[:10] or "—",
                        "dias_exceso": dias,
                    }
                )
        desfases.sort(key=lambda x: (-x["dias_exceso"], x.get("nombre") or ""))

    cpm_estado = None
    if tiene_cpm:
        cpm_estado = "Desactualizado" if meta.get("cpm_dirty") else "Vigente"

    return {
        "capitulos": capitulos_rows,
        "tiene_cpm": tiene_cpm,
        "criticos": criticos,
        "desfases": desfases,
        "version": {
            "tiene_version": bool(meta),
            "nombre": _prog_version_label(meta) if meta else "—",
            "estado": _prog_estado_display(meta.get("estado")) if meta else "—",
            "fecha_inicio": str(meta.get("fecha_inicio") or "")[:10] or "—",
            "fecha_fin": str(meta.get("fecha_fin") or "")[:10] or "—",
            "cpm_estado": cpm_estado,
        },
    }


def _render_resumen_ejecutivo_pdf_html(
    contrato: dict,
    resumen: dict,
    prog_meta: Optional[dict],
    fecha_gen: str,
) -> str:
    c = PDF_CLR
    td = _pdf_td
    note = f"font-size:6.5pt;color:{c['muted']};font-style:italic;padding:2px 0;"
    sec = (
        f"font-size:7.5pt;font-weight:bold;color:{c['text']};"
        f"margin:6px 0 3px;text-transform:uppercase;"
    )
    tiene_cpm = bool(resumen.get("tiene_cpm"))
    ver = resumen.get("version") or {}

    if not tiene_cpm:
        ver_block = f"""
<div style="{note}">No hay resultados CPM calculados para esta versión.</div>
<table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;margin-top:2px;">
  <tbody>
    <tr>
      <td style="{td('width:22%;font-weight:bold;')}">Versión</td>
      <td style="{td()}">{_h(ver.get("nombre"))}</td>
      <td style="{td('width:18%;font-weight:bold;')}">Estado</td>
      <td style="{td()}">{_h(ver.get("estado"))}</td>
    </tr>
    <tr>
      <td style="{td('font-weight:bold;')}">Inicio cronograma</td>
      <td style="{td()}">{_h(ver.get("fecha_inicio"))}</td>
      <td style="{td('font-weight:bold;')}">Fin cronograma</td>
      <td style="{td()}">{_h(ver.get("fecha_fin"))}</td>
    </tr>
    <tr>
      <td style="{td('font-weight:bold;')}">Estado CPM</td>
      <td colspan="3" style="{td('color:' + c['muted'] + ';')}">Sin calcular</td>
    </tr>
  </tbody>
</table>
"""
    else:
        cpm_lbl = ver.get("cpm_estado") or "—"
        cpm_color = c["negative"] if cpm_lbl == "Desactualizado" else c["accent"]
        ver_block = f"""
<table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;">
  <tbody>
    <tr>
      <td style="{td('width:22%;font-weight:bold;')}">Versión</td>
      <td style="{td()}">{_h(ver.get("nombre"))}</td>
      <td style="{td('width:18%;font-weight:bold;')}">Estado</td>
      <td style="{td()}">{_h(ver.get("estado"))}</td>
    </tr>
    <tr>
      <td style="{td('font-weight:bold;')}">Inicio cronograma</td>
      <td style="{td()}">{_h(ver.get("fecha_inicio"))}</td>
      <td style="{td('font-weight:bold;')}">Fin cronograma</td>
      <td style="{td()}">{_h(ver.get("fecha_fin"))}</td>
    </tr>
    <tr>
      <td style="{td('font-weight:bold;')}">Estado CPM</td>
      <td colspan="3" style="{td('font-weight:bold;color:' + cpm_color + ';')}">{_h(cpm_lbl)}</td>
    </tr>
  </tbody>
</table>
"""

    header = _pdf_page_header_html(
        contrato,
        "Estado de la versión",
        subtitle="Programación de obra",
        prog_meta=prog_meta,
        fecha_gen=fecha_gen,
    )
    return f"""
<div class="page-cronograma">
  <div class="frame-double-outer">
    <div class="frame-double-inner">
      {header}
      <div style="{sec}">Versión activa</div>
      {ver_block}
    </div>
  </div>
</div>
"""


def _render_cronograma_pk_pdf_pages(
    contrato: dict,
    cronograma: List[dict],
    cpm_export: Optional[dict],
    prog_meta: Optional[dict],
    fecha_gen: str,
) -> str:
    holgura_map = _cpm_holgura_by_agrupador(cpm_export)
    page_specs: List[Tuple[dict, List[dict], int, int]] = []
    for pk_node in cronograma or []:
        pk = str(pk_node.get("pk_id") or "").strip()
        ag_rows: List[dict] = []
        for cap_node in pk_node.get("capitulos") or []:
            cap = str(cap_node.get("capitulo") or "").strip()
            for ag in cap_node.get("agrupadores") or []:
                ag_rows.append({"capitulo": cap, **ag})
        chunks = _chunk_list(ag_rows, CRONO_ROWS_PER_PAGE) or [[]]
        for idx, chunk in enumerate(chunks, start=1):
            page_specs.append((pk_node, chunk, idx, len(chunks)))

    if not page_specs:
        page_specs = [({"pk_id": "—", "capitulos": []}, [], 1, 1)]

    total = len(page_specs)
    out = ""
    c = PDF_CLR
    th = _pdf_th_style
    td = _pdf_td
    for page_num, (pk_node, ag_chunk, pk_idx, pk_total) in enumerate(page_specs, start=1):
        pk = str(pk_node.get("pk_id") or "—")
        resumen = _pk_cronograma_resumen(pk_node)
        pag = f"Página {page_num} de {total}" if total > 1 else ""
        pk_pag = f" · hoja {pk_idx} de {pk_total}" if pk_total > 1 else ""
        header = _pdf_page_header_html(
            contrato,
            "Cronograma de programación",
            subtitle=f"PK {_h(pk)}{pk_pag}",
            prog_meta=prog_meta,
            fecha_gen=fecha_gen,
            pag=pag,
        )
        if pk_idx == 1:
            resumen_html = f"""
<table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;margin-bottom:4px;background:{c['bg_kpi']};">
  <tr>
    <td style="{td('border:0.25px solid ' + c['border_light'] + ';font-weight:bold;width:20%;')}">Resumen PK</td>
    <td style="{td('border:0.25px solid ' + c['border_light'] + ';')}">Agrupadores: <strong>{resumen['agrupadores']}</strong></td>
    <td style="{td('border:0.25px solid ' + c['border_light'] + ';')}">Inicio: <strong>{_h(resumen['fecha_inicio'])}</strong></td>
    <td style="{td('border:0.25px solid ' + c['border_light'] + ';')}">Fin: <strong>{_h(resumen['fecha_fin'])}</strong></td>
  </tr>
</table>
"""
        else:
            resumen_html = ""

        if ag_chunk:
            body = ""
            for ag in ag_chunk:
                cap = str(ag.get("capitulo") or "").strip()
                ag_id = ag.get("agrupador_id")
                wbs = str(ag.get("codigo_wbs") or "").strip()
                nombre = str(ag.get("nombre") or wbs).strip()
                act_lbl = f"{wbs} · {nombre}".strip(" ·") if wbs else nombre
                try:
                    hol = holgura_map.get((pk, cap, int(ag_id)), "—")
                except (TypeError, ValueError):
                    hol = "—"
                body += (
                    f"<tr style=\"background:{c['row_ag']};\">"
                    f"<td style=\"{td('font-weight:bold;')}\">"
                    f"<div>{_h(cap)}</div>"
                    f"<div style=\"font-size:5.5pt;color:{c['muted']};font-weight:normal;\">{_h(act_lbl)}</div>"
                    f"</td>"
                    f"<td style=\"{td('text-align:center;')}\">{_h(ag.get('fecha_inicio'))}</td>"
                    f"<td style=\"{td('text-align:center;')}\">{_h(ag.get('fecha_fin'))}</td>"
                    f"<td style=\"{td('text-align:center;')}\">{_h(_fmt_duracion(ag.get('duracion_dias_habiles')))}</td>"
                    f"<td style=\"{td('text-align:center;')}\">{_h(hol)}</td>"
                    f"</tr>"
                )
        else:
            body = (
                f'<tr><td colspan="5" style="{td("text-align:center;color:" + c["muted"] + ";")}">'
                "Sin actividades programadas en este PK.</td></tr>"
            )

        out += f"""
<div class="page-cronograma">
  <div class="frame-double-outer">
    <div class="frame-double-inner">
      {header}
      {resumen_html}
      <table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;">
        <thead><tr>
          <th style="{th('text-align:left;width:40%;')}">Capítulo / Actividad WBS</th>
          <th style="{th('text-align:center;width:15%;')}">Inicio</th>
          <th style="{th('text-align:center;width:15%;')}">Fin</th>
          <th style="{th('text-align:center;width:15%;')}">Días háb.</th>
          <th style="{th('text-align:center;width:15%;')}">Holgura</th>
        </tr></thead>
        <tbody>{body}</tbody>
      </table>
    </div>
  </div>
</div>
"""
    return out


def _meses_table_rows_html(meses_chunk: List[dict]) -> str:
    td = _pdf_td
    rows_html = ""
    for r in meses_chunk:
        rows_html += (
            f"<tr>"
            f"<td style=\"{td()}\">{_h(r.get('mes_label'))}</td>"
            f"<td style=\"{td('text-align:right;')}\">${float(r.get('baseline_mes') or 0):,.0f}</td>"
            f"<td style=\"{td('text-align:right;')}\">${float(r.get('baseline_acum') or 0):,.0f}</td>"
            f"<td style=\"{td('text-align:right;')}\">${float(r.get('vigente_mes') or 0):,.0f}</td>"
            f"<td style=\"{td('text-align:right;')}\">${float(r.get('vigente_acum') or 0):,.0f}</td>"
            f"<td style=\"{td('text-align:right;')}\">${float(r.get('ejecutado_mes') or 0):,.0f}</td>"
            f"<td style=\"{td('text-align:right;')}\">${float(r.get('ejecutado_acum') or 0):,.0f}</td>"
            f"<td style=\"{td('text-align:right;')}\">{_h(r.get('delta_vigente_pct'))}%</td>"
            f"<td style=\"{td('text-align:right;')}\">{_h(r.get('delta_ejecutado_pct'))}%</td>"
            f"</tr>"
        )
    return rows_html


def _render_curva_s_pdf_header_block(
    contrato: dict,
    ind: dict,
    *,
    fecha_gen: str,
    ppto_meta: Optional[dict],
    prog_meta: Optional[dict],
    data: dict,
) -> str:
    c = PDF_CLR
    titulo = f"Curva S — Contrato {_h(contrato.get('numero') or contrato.get('id') or '')}"
    contratista = _h(contrato.get("contratista") or "—")
    interventoria = _h(contrato.get("interventoria") or "—")
    objeto_cto = _h(contrato.get("objeto") or "—")
    ppto_lbl = _h(_ppto_version_label(ppto_meta))
    prog_lbl = _h(_prog_version_label(prog_meta))
    bid = _h(data.get("baseline_id") or "—")
    tid = _h(data.get("target_id") or "—")
    desv_val = float(ind.get("desviacion_valor") or 0)
    desv_color = c["negative"] if desv_val < 0 else c["accent"]
    lbl = (
        f"font-size:5.5pt;font-weight:bold;color:{c['muted']};"
        f"text-transform:uppercase;letter-spacing:0.2px;"
    )
    val = f"font-size:7pt;color:{c['text']};margin-top:1px;line-height:1.15;"
    logo_contratista = _html_footer_logo_contratista(contrato)
    return f"""
<div class="title">{titulo}</div>
<div class="sub">{objeto_cto} · {contratista} · Interventoría: {interventoria}</div>
<div style="font-size:7pt;color:{c['text']};margin-bottom:6px;line-height:1.35;">
  <strong>Contrato:</strong> {_h(contrato.get('numero') or contrato.get('id') or '')}
  · <strong>Contratista:</strong> {contratista}
  · <strong>Interventoría:</strong> {interventoria}
  · <strong>Generado:</strong> {_h(fecha_gen)}
</div>
<table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;background:{c['bg_kpi']};margin-bottom:8px;border:1px solid {c['border_light']};">
  <tr>
    <td style="padding:5px 8px;font-size:8pt;white-space:nowrap;color:{c['text']};">
      Presupuesto total: <strong style="color:{c['accent']};">${float(ind.get('presupuesto_total') or 0):,.0f}</strong>
      &nbsp;&nbsp;&nbsp;Programado a la fecha: <strong style="color:{c['accent']};">${float(ind.get('programado_a_fecha') or 0):,.0f} ({ind.get('programado_pct')}%)</strong>
      &nbsp;&nbsp;&nbsp;Ejecutado a la fecha: <strong style="color:{c['accent']};">${float(ind.get('ejecutado_a_fecha') or 0):,.0f} ({ind.get('ejecutado_pct')}%)</strong>
      &nbsp;&nbsp;&nbsp;Desviación: <strong style="color:{desv_color};">${desv_val:,.0f} ({ind.get('desviacion_pct')}%)</strong>
    </td>
  </tr>
</table>
<table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;margin-bottom:8px;border-top:1px solid {c['border_light']};">
  <tr>
    <td style="width:18%;vertical-align:middle;padding:3px 4px 2px 0;border-right:1px solid {c['border_light']};">
      <div style="{lbl}">Logo contratista</div>
      <div style="{val}">{logo_contratista}</div>
    </td>
    <td style="width:32%;vertical-align:top;padding:3px 4px 2px 6px;border-right:1px solid {c['border_light']};">
      <div style="{lbl}">Contratista · Interventoría</div>
      <div style="{val}"><strong>{contratista}</strong><br/>Interventoría: {interventoria}</div>
    </td>
    <td style="width:50%;vertical-align:top;padding:3px 0 2px 6px;">
      <div style="{lbl}">Versiones programación</div>
      <div style="{val}"><strong>Baseline:</strong> {bid}<br/><strong>Vigente:</strong> {tid} · {prog_lbl}</div>
    </td>
  </tr>
  <tr>
    <td colspan="2" style="vertical-align:top;padding:3px 4px 2px 0;border-right:1px solid {c['border_light']};border-top:1px solid {c['border_light']};">
      <div style="{lbl}">Versión presupuesto activa</div>
      <div style="{val}">{ppto_lbl}</div>
    </td>
    <td style="vertical-align:top;padding:3px 0 2px 6px;border-top:1px solid {c['border_light']};">
      <div style="{lbl}">Fecha y hora de generación</div>
      <div style="{val}">{_h(fecha_gen)}</div>
    </td>
  </tr>
</table>
"""


def _render_curva_s_meses_pdf_page(
    contrato: dict,
    meses_chunk: List[dict],
    *,
    page_num: int,
    total_pages: int,
    fecha_gen: str,
    prog_meta: Optional[dict],
) -> str:
    c = PDF_CLR
    th = _pdf_th_style
    td = _pdf_td
    pag = f"Página {page_num} de {total_pages}" if total_pages > 1 else ""
    tbody = _meses_table_rows_html(meses_chunk)
    if not tbody:
        tbody = (
            f'<tr><td colspan="9" style="{td("text-align:center;color:" + c["muted"] + ";")}">'
            "Sin meses con datos de inversión programada.</td></tr>"
        )
    header = _pdf_page_header_html(
        contrato,
        "Curva S — Distribución mensual de inversión",
        prog_meta=prog_meta,
        fecha_gen=fecha_gen,
        pag=pag,
    )
    return f"""
<div class="page-cronograma">
  <div class="frame-meses-outer">
    <div class="frame-meses-inner">
      {header}
      <table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;">
        <thead><tr>
          <th style="{th('text-align:left;width:8%;')}">Mes</th>
          <th style="{th('text-align:right;width:10%;')}">Base. mes</th>
          <th style="{th('text-align:right;width:10%;')}">Base. acum.</th>
          <th style="{th('text-align:right;width:10%;')}">Vig. mes</th>
          <th style="{th('text-align:right;width:10%;')}">Vig. acum.</th>
          <th style="{th('text-align:right;width:10%;')}">Ejec. mes</th>
          <th style="{th('text-align:right;width:10%;')}">Ejec. acum.</th>
          <th style="{th('text-align:right;width:8%;')}">Δ Vig.</th>
          <th style="{th('text-align:right;width:8%;')}">Δ Ejec.</th>
        </tr></thead>
        <tbody>{tbody}</tbody>
      </table>
    </div>
  </div>
</div>
"""


def _group_detalle_by_pk(detalle: List[dict]) -> Dict[str, List[dict]]:
    by_pk: Dict[str, List[dict]] = defaultdict(list)
    for row in detalle or []:
        pk = str(row.get("pk_id") or "").strip() or "—"
        by_pk[pk].append(row)
    for pk in by_pk:
        by_pk[pk].sort(
            key=lambda r: (
                _cap_sort_key(str(r.get("capitulo") or "")),
                str(r.get("label") or ""),
            )
        )
    return by_pk


def _render_detalle_pk_pdf_pages(
    contrato: dict,
    data: dict,
    meses: List[dict],
    *,
    prog_meta: Optional[dict],
    fecha_gen: str,
) -> str:
    det = data.get("detalle_pk") or {}
    baseline_rows = det.get("baseline") or []
    vigente_rows = det.get("vigente") or []
    month_keys = [str(m.get("mes") or "") for m in meses if m.get("mes")]
    if not month_keys:
        month_keys = sorted(
            {
                mk
                for row in (baseline_rows + vigente_rows)
                for mk in (row.get("distribucion_mensual") or {}).keys()
            }
        )
    month_labels = {str(m.get("mes") or ""): str(m.get("mes_label") or "") for m in meses}
    same_version = str(data.get("baseline_id") or "") == str(data.get("target_id") or "")
    sections: List[Tuple[str, List[dict]]] = []
    if baseline_rows:
        sections.append(("Baseline", baseline_rows))
    if vigente_rows and not same_version:
        sections.append(("Vigente", vigente_rows))
    elif vigente_rows and not baseline_rows:
        sections.append(("Vigente", vigente_rows))

    c = PDF_CLR
    th = _pdf_th_style
    td = _pdf_td

    if not sections:
        header = _pdf_page_header_html(
            contrato,
            "Detalle por PK",
            subtitle="Distribución mensual",
            prog_meta=prog_meta,
            fecha_gen=fecha_gen,
        )
        return f"""
<div class="page-cronograma">
  <div class="frame-double-outer"><div class="frame-double-inner">
    {header}
    <div style="font-size:7pt;color:{c['muted']};">Sin actividades programadas con costo en el alcance seleccionado.</div>
  </div></div>
</div>
"""

    out = ""
    for sec_label, rows in sections:
        by_pk = _group_detalle_by_pk(rows)
        month_slices = _chunk_list(month_keys, DETALLE_PK_MONTH_COLS) or [[]]
        for pk in sorted(by_pk.keys(), key=lambda x: (len(x), x)):
            pk_rows = by_pk[pk]
            for m_idx, mslice in enumerate(month_slices, start=1):
                head = "".join(
                    f'<th style="{th("text-align:right;font-size:5.5pt;")}">{_h(month_labels.get(mk, mk))}</th>'
                    for mk in mslice
                )
                body = ""
                for row in pk_rows:
                    dist = row.get("distribucion_mensual") or {}
                    ag_nombre = str(row.get("agrupador_nombre") or row.get("label") or "").strip()
                    wbs = str(row.get("codigo_wbs") or row.get("label") or "").strip()
                    cells = "".join(
                        f'<td style="{td("text-align:right;")}">${float(dist.get(mk, 0)):,.0f}</td>'
                        for mk in mslice
                    )
                    body += (
                        f"<tr>"
                        f'<td style="{td()}">{_h(ag_nombre)}</td>'
                        f'<td style="{td()}">'
                        f'<div>{_h(row.get("capitulo"))}</div>'
                        f'<div style="font-size:5.5pt;color:{c["muted"]};">{_h(wbs)}</div>'
                        f"</td>"
                        f'<td style="{td("text-align:right;font-weight:bold;")}">'
                        f'${float(row.get("costo_total") or 0):,.0f}</td>'
                        f"{cells}</tr>"
                    )
                m_cont = f" · meses {m_idx}/{len(month_slices)}" if len(month_slices) > 1 else ""
                header = _pdf_page_header_html(
                    contrato,
                    f"Detalle por PK — {_h(sec_label)}",
                    subtitle=f"PK {_h(pk)}{m_cont}",
                    prog_meta=prog_meta,
                    fecha_gen=fecha_gen,
                )
                out += f"""
<div class="page-cronograma">
  <div class="frame-double-outer">
    <div class="frame-double-inner">
      {header}
      <table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;">
        <thead><tr>
          <th style="{th('text-align:left;width:18%;')}">Agrupador</th>
          <th style="{th('text-align:left;width:10%;')}">Capítulo / WBS</th>
          <th style="{th('text-align:right;width:8%;')}">Costo total</th>
          {head}
        </tr></thead>
        <tbody>{body}</tbody>
      </table>
    </div>
  </div>
</div>
"""
    return out


def _render_curva_s_intro_pdf_page(
    contrato: dict,
    chart_block: str,
    *,
    fecha_gen: str,
    prog_meta: Optional[dict],
    page_num: int,
    total_pages: int,
) -> str:
    pag = f"Página {page_num} de {total_pages}" if total_pages > 1 else ""
    header = _pdf_page_header_html(
        contrato,
        "Curva S — Inversión acumulada",
        prog_meta=prog_meta,
        fecha_gen=fecha_gen,
        pag=pag,
    )
    return f"""
<div class="page-cronograma">
  <div class="frame-double-outer">
    <div class="frame-double-inner">
      {header}
      {chart_block}
    </div>
  </div>
</div>
"""


def build_curva_s_pdf_html(
    contrato: dict,
    data: dict,
    *,
    cronograma: Optional[List[dict]] = None,
    prog_meta: Optional[dict] = None,
    ppto_meta: Optional[dict] = None,
    cpm_export: Optional[dict] = None,
    resumen_ejecutivo: Optional[dict] = None,
    fecha_generacion: Optional[str] = None,
) -> str:
    """HTML multipágina: curva S + detalle PK + resumen ejecutivo + cronograma + CPM."""
    from informes import _fmt_informe_fecha_generacion

    meses = data.get("meses") or []
    fecha_gen = (fecha_generacion or "").strip() or _fmt_informe_fecha_generacion()
    c = PDF_CLR

    chart_block = _build_curva_s_chart_block(meses)

    meses_chunks = _chunk_list(meses, MESES_ROWS_PER_PDF_PAGE) or [[]]
    total_curva_pages = 1 + len(meses_chunks)
    curva_pages_html = _render_curva_s_intro_pdf_page(
        contrato,
        chart_block,
        fecha_gen=fecha_gen,
        prog_meta=prog_meta,
        page_num=1,
        total_pages=total_curva_pages,
    )
    for idx, chunk in enumerate(meses_chunks, start=2):
        curva_pages_html += _render_curva_s_meses_pdf_page(
            contrato,
            chunk,
            page_num=idx,
            total_pages=total_curva_pages,
            fecha_gen=fecha_gen,
            prog_meta=prog_meta,
        )

    detalle_pk_html = _render_detalle_pk_pdf_pages(
        contrato, data, meses, prog_meta=prog_meta, fecha_gen=fecha_gen
    )

    crono_pages_html = _render_cronograma_pk_pdf_pages(
        contrato, cronograma or [], cpm_export, prog_meta, fecha_gen
    )

    cpm_pages_html = _render_cpm_export_pdf_html(contrato, cpm_export, prog_meta, fecha_gen)

    resumen_html = ""
    if resumen_ejecutivo is not None:
        resumen_html = _render_resumen_ejecutivo_pdf_html(
            contrato, resumen_ejecutivo, prog_meta, fecha_gen
        )

    return f"""<!DOCTYPE html>
<html xmlns="http://www.w3.org/1999/xhtml"><head><meta charset="utf-8"/>
<style>
@page {{ size: A4 landscape; margin: 0.9cm; }}
body {{ font-family: Helvetica, Arial, sans-serif; font-size: 9pt; color: {c['text']}; margin: 0; padding: 0; }}
.page-cronograma {{ page-break-after: always; }}
.page-cronograma:last-child {{ page-break-after: auto; }}
.frame-double-outer {{ border: 1px solid #000; padding: 2px; min-height: 17.5cm; box-sizing: border-box; }}
.frame-double-inner {{ border: 0.5px solid #000; padding: 6px 8px; min-height: calc(17.5cm - 8px); box-sizing: border-box; }}
.frame-meses-outer {{ border: 0.5px solid #000; padding: 2px; min-height: 17.5cm; box-sizing: border-box; }}
.frame-meses-inner {{ border: 0.25px solid #000; padding: 5px 7px; min-height: calc(17.5cm - 6px); box-sizing: border-box; }}
</style></head><body>
{curva_pages_html}
{detalle_pk_html}
{resumen_html}
{crono_pages_html}
{cpm_pages_html}
</body></html>"""


EXCEL_COP_FMT = '"$"#,##0'
EXCEL_PCT_FMT = '0.0"%"'


def _xl_solid_fill(hex6: str) -> "PatternFill":
    from openpyxl.styles import PatternFill

    h = (hex6 or "").replace("#", "").upper()
    return PatternFill("solid", fgColor=h)


def _xl_thin_border() -> "Border":
    from openpyxl.styles import Border, Side

    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def _chart_png_bytes(meses: List[dict]) -> bytes:
    svg = _build_curva_s_chart_svg(meses)
    return base64.b64decode(_svg_to_png_b64(svg))


def _cronograma_row_label(r: dict) -> str:
    kind = r.get("kind")
    if kind == "pk":
        return f"PK {r.get('pk_id')}"
    if kind == "cap":
        return f"Capítulo {r.get('capitulo')}"
    if kind == "ag":
        return f"{r.get('codigo_wbs') or ''} · {r.get('nombre') or ''}".strip(" ·")
    if kind == "item":
        desc = (r.get("descripcion") or "").strip()
        label = str(r.get("item") or "")
        return f"    {label} · {desc}" if desc else f"    {label}"
    return ""


def _month_starts(d0: date, d1: date) -> List[date]:
    cur = date(d0.year, d0.month, 1)
    end = date(d1.year, d1.month, 1)
    out: List[date] = []
    while cur <= end:
        out.append(cur)
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)
    return out


def _month_end(d: date) -> date:
    return date(d.year, d.month, monthrange(d.year, d.month)[1])


def _gantt_cost_lookup(detalle_vigente: List[dict]) -> Dict[Tuple[str, str, int], float]:
    out: Dict[Tuple[str, str, int], float] = {}
    for r in detalle_vigente or []:
        ag_id = r.get("agrupador_id")
        if ag_id is None:
            continue
        key = (str(r.get("pk_id") or ""), str(r.get("capitulo") or ""), int(ag_id))
        out[key] = float(r.get("costo_total") or 0)
    return out


def _gantt_month_columns(meses_data: List[dict], span_dates: List[date]) -> List[Tuple[str, date, str]]:
    seen: Dict[str, Tuple[date, str]] = {}
    for m in meses_data or []:
        mk = str(m.get("mes") or "").strip()
        if not mk or mk in seen:
            continue
        y, mo = mk.split("-")
        seen[mk] = (date(int(y), int(mo), 1), str(m.get("mes_label") or _month_label(mk)))
    if span_dates:
        d0, d1 = min(span_dates), max(span_dates)
        for ms in _month_starts(d0, d1):
            mk = _month_key(ms)
            if mk not in seen:
                seen[mk] = (ms, _month_label(mk))
    return [(mk, seen[mk][0], seen[mk][1]) for mk in sorted(seen.keys())]


def _month_key_to_date(mk: str) -> Optional[date]:
    mk = (mk or "").strip()
    if not mk or "-" not in mk:
        return None
    try:
        y, mo = mk.split("-", 1)
        return date(int(y), int(mo), 1)
    except (TypeError, ValueError):
        return None


def _xl_linear_month_formula(
    row: int,
    col_letter: str,
    hdr_row: int,
    *,
    fi_col: str,
    ff_col: str,
    cost_col: str,
) -> str:
    """Distribución lineal por días calendario (misma lógica que _linear_monthly_distribution)."""
    return (
        f'=IF(OR(${fi_col}{row}="",${ff_col}{row}="",${cost_col}{row}=0),0,'
        f'MAX(0,MIN(${ff_col}{row},EOMONTH({col_letter}${hdr_row},0))'
        f'-MAX(${fi_col}{row},{col_letter}${hdr_row})+1)'
        f'/(${ff_col}{row}-${fi_col}{row}+1)*${cost_col}{row})'
    )


def _xl_gantt_month_formula(row: int, col_letter: str, hdr_row: int = 4) -> str:
    inner = _xl_linear_month_formula(
        row, col_letter, hdr_row, fi_col="B", ff_col="C", cost_col="F",
    )[1:]
    return f'=IF($E{row}<>"Agrupador WBS","",{inner})'


def _xl_detalle_pk_month_formula(row: int, col_letter: str, hdr_row: int = 4) -> str:
    return _xl_linear_month_formula(
        row, col_letter, hdr_row, fi_col="F", ff_col="G", cost_col="E",
    )


def _xl_curva_s_sumif_detalle(
    sheet: str,
    section: str,
    col_letter: str,
    first_row: int,
    last_row: int,
) -> str:
    return (
        f"=SUMIF('{sheet}'!$A${first_row}:$A${last_row},\"{section}\","
        f"'{sheet}'!{col_letter}${first_row}:{col_letter}${last_row})"
    )


def _xl_curva_s_running_sum(col_letter: str, row: int, first_row: int) -> str:
    return f"=SUM(${col_letter}${first_row}:{col_letter}{row})"


def _xl_curva_s_delta_pct(acum_col: str, base_col: str, row: int) -> str:
    return f"=IF({base_col}{row}=0,0,({acum_col}{row}-{base_col}{row})/{base_col}{row})"


def _xl_gantt_row_total_formula(row: int, first_month_col: int, last_month_col: int) -> str:
    from openpyxl.utils import get_column_letter

    fc = get_column_letter(first_month_col)
    lc = get_column_letter(last_month_col)
    return f'=IF($E{row}<>"Agrupador WBS","",SUM({fc}{row}:{lc}{row}))'


def _write_gantt_sheet(
    ws_g,
    *,
    flat: List[dict],
    meses_data: List[dict],
    detalle_vigente: List[dict],
    numero: str,
    prog_meta: Optional[dict],
    fecha_gen: str,
) -> Optional[dict]:
    """Escribe pestaña Gantt. Retorna metadatos para enlazar fórmulas en Curva S."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    ws_g.merge_cells("A1:F1")
    ws_g["A1"] = "Cronograma de Programación de Obra"
    ws_g["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws_g.merge_cells("A2:F2")
    ws_g["A2"] = (
        f"Contrato {numero} · Versión programación: {_prog_version_label(prog_meta)} · "
        f"Generado: {fecha_gen}"
    )
    ws_g["A2"].font = Font(size=9, color="64748B")
    ws_g["A3"] = (
        "Valores mensuales formulados: distribución lineal del costo programado según Inicio, Fin y días calendario."
    )
    ws_g["A3"].font = Font(size=8, italic=True, color="64748B")

    g_hdr = 4
    base_headers = ["Actividad / WBS", "Inicio", "Fin", "Días háb.", "Tipo", "Costo programado"]
    for col, h in enumerate(base_headers, start=1):
        cell = ws_g.cell(row=g_hdr, column=col, value=h)
        cell.font = Font(bold=True, color="1E3A8A", size=9)
        cell.fill = _xl_solid_fill("DBEAFE")
        cell.border = _xl_thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    span_dates: List[date] = []
    for row in flat:
        if row.get("kind") not in ("ag", "item"):
            continue
        fi = _parse_d(row.get("fecha_inicio"))
        ff = _parse_d(row.get("fecha_fin"))
        if fi:
            span_dates.append(fi)
        if ff:
            span_dates.append(ff)

    month_specs = _gantt_month_columns(meses_data, span_dates)
    if len(month_specs) > 48:
        month_specs = month_specs[:48]

    first_month_col = len(base_headers) + 1
    last_month_col = first_month_col + len(month_specs) - 1 if month_specs else first_month_col - 1
    total_col = last_month_col + 1 if month_specs else len(base_headers) + 1

    for i, (_mk, ms, lbl) in enumerate(month_specs):
        col = first_month_col + i
        c = ws_g.cell(row=g_hdr, column=col, value=ms)
        c.number_format = "mmm yyyy"
        c.font = Font(bold=True, color="1E3A8A", size=8)
        c.fill = _xl_solid_fill("DBEAFE")
        c.border = _xl_thin_border()
        c.alignment = Alignment(horizontal="center", text_rotation=90)

    if month_specs:
        tc = ws_g.cell(row=g_hdr, column=total_col, value="Total distribuido")
        tc.font = Font(bold=True, color="1E3A8A", size=9)
        tc.fill = _xl_solid_fill("DBEAFE")
        tc.border = _xl_thin_border()
        tc.alignment = Alignment(horizontal="center", wrap_text=True)

    cost_lookup = _gantt_cost_lookup(detalle_vigente)
    row_fills = {"pk": "E2E8F0", "cap": "F1F5F9", "ag": "E0F2FE", "item": "FFFFFF"}
    tipo_lbl = {"ag": "Agrupador WBS", "item": "Ítem"}

    gr = g_hdr + 1
    first_data_row = gr
    sheet_last_col = max(total_col, 6)

    for row in flat:
        kind = row.get("kind")
        if kind in ("pk", "cap"):
            ws_g.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=sheet_last_col)
            cell = ws_g.cell(row=gr, column=1, value=_cronograma_row_label(row))
            cell.font = Font(bold=True, size=9 if kind == "pk" else 8)
            cell.fill = _xl_solid_fill(row_fills[kind])
            gr += 1
            continue
        if kind not in ("ag", "item"):
            continue

        ws_g.cell(row=gr, column=1, value=_cronograma_row_label(row).strip())
        fi = _parse_d(row.get("fecha_inicio"))
        ff = _parse_d(row.get("fecha_fin"))
        ws_g.cell(row=gr, column=2, value=fi)
        ws_g.cell(row=gr, column=3, value=ff)
        if fi:
            ws_g.cell(row=gr, column=2).number_format = "yyyy-mm-dd"
        if ff:
            ws_g.cell(row=gr, column=3).number_format = "yyyy-mm-dd"
        dur = row.get("duracion_dias_habiles")
        ws_g.cell(row=gr, column=4, value=int(dur) if dur is not None else None)
        ws_g.cell(row=gr, column=5, value=tipo_lbl.get(kind, ""))

        if kind == "ag":
            ag_key = (
                str(row.get("pk_id") or ""),
                str(row.get("capitulo") or ""),
                int(row.get("agrupador_id") or 0),
            )
            ws_g.cell(row=gr, column=6, value=cost_lookup.get(ag_key, 0.0))
        ws_g.cell(row=gr, column=6).number_format = EXCEL_COP_FMT

        fill = _xl_solid_fill(row_fills[kind])
        for col in range(1, sheet_last_col + 1):
            c = ws_g.cell(row=gr, column=col)
            c.fill = fill
            c.border = _xl_thin_border()
            if col == 4:
                c.alignment = Alignment(horizontal="center")
            elif col >= 6:
                c.alignment = Alignment(horizontal="right")

        if kind == "ag" and month_specs:
            for i in range(len(month_specs)):
                col = first_month_col + i
                letter = get_column_letter(col)
                c = ws_g.cell(row=gr, column=col)
                c.value = _xl_gantt_month_formula(gr, letter, hdr_row=g_hdr)
                c.number_format = EXCEL_COP_FMT
            ws_g.cell(row=gr, column=total_col, value=_xl_gantt_row_total_formula(gr, first_month_col, last_month_col))
            ws_g.cell(row=gr, column=total_col).number_format = EXCEL_COP_FMT

        gr += 1

    last_data_row = gr - 1
    tr: Optional[int] = None

    if not flat:
        ws_g.merge_cells(start_row=g_hdr + 1, start_column=1, end_row=g_hdr + 1, end_column=sheet_last_col)
        ws_g.cell(row=g_hdr + 1, column=1, value="Sin actividades programadas en esta versión.").alignment = (
            Alignment(horizontal="center")
        )
    elif month_specs and last_data_row >= first_data_row:
        tr = gr
        ws_g.cell(row=tr, column=1, value="TOTAL GENERAL").font = Font(bold=True, color="1E40AF", size=10)
        ws_g.cell(row=tr, column=1).fill = _xl_solid_fill("DBEAFE")
        ws_g.cell(row=tr, column=6, value=(
            f'=SUMIF($E${first_data_row}:$E${last_data_row},"Agrupador WBS",'
            f'$F${first_data_row}:$F${last_data_row})'
        ))
        ws_g.cell(row=tr, column=6).number_format = EXCEL_COP_FMT
        ws_g.cell(row=tr, column=6).font = Font(bold=True)
        for i in range(len(month_specs)):
            col = first_month_col + i
            letter = get_column_letter(col)
            c = ws_g.cell(
                row=tr,
                column=col,
                value=(
                    f'=SUMIF($E${first_data_row}:$E${last_data_row},"Agrupador WBS",'
                    f'{letter}${first_data_row}:{letter}${last_data_row})'
                ),
            )
            c.number_format = EXCEL_COP_FMT
            c.font = Font(bold=True)
            c.fill = _xl_solid_fill("DBEAFE")
        ws_g.cell(
            row=tr,
            column=total_col,
            value=(
                f'=SUMIF($E${first_data_row}:$E${last_data_row},"Agrupador WBS",'
                f'{get_column_letter(total_col)}${first_data_row}:{get_column_letter(total_col)}${last_data_row})'
            ),
        )
        ws_g.cell(row=tr, column=total_col).number_format = EXCEL_COP_FMT
        ws_g.cell(row=tr, column=total_col).font = Font(bold=True)
        ws_g.cell(row=tr, column=total_col).fill = _xl_solid_fill("DBEAFE")
        for col in range(1, sheet_last_col + 1):
            ws_g.cell(row=tr, column=col).border = _xl_thin_border()

    ws_g.column_dimensions["A"].width = 44
    ws_g.column_dimensions["B"].width = 12
    ws_g.column_dimensions["C"].width = 12
    ws_g.column_dimensions["D"].width = 10
    ws_g.column_dimensions["E"].width = 14
    ws_g.column_dimensions["F"].width = 18
    for i in range(len(month_specs)):
        ws_g.column_dimensions[get_column_letter(first_month_col + i)].width = 14
    if month_specs:
        ws_g.column_dimensions[get_column_letter(total_col)].width = 16
    ws_g.freeze_panes = f"{get_column_letter(first_month_col)}{g_hdr + 1}"

    month_col_by_key = {mk: first_month_col + i for i, (mk, _ms, _lbl) in enumerate(month_specs)}
    total_row_num = tr if month_specs and last_data_row >= first_data_row and flat else None
    if not month_col_by_key:
        return None
    return {"total_row": total_row_num, "month_cols": month_col_by_key}


def load_curva_s_export_context(
    sb,
    contrato_id: int,
    *,
    baseline_id: Optional[str] = None,
    target_id: Optional[str] = None,
    version_ppto_id: Optional[str] = None,
    pk_ids: Optional[str] = None,
    tramos: Optional[str] = None,
) -> dict:
    """Contrato, curva S, cronograma, CPM y metadatos compartidos por PDF y Excel."""
    from prog_obra_service import build_cpm_export_data

    data = build_curva_s(
        sb,
        contrato_id,
        baseline_id=baseline_id,
        target_id=target_id,
        version_ppto_id=version_ppto_id,
        pk_ids=pk_ids,
        tramos=tramos,
    )
    crows = (
        sb.table("contratos")
        .select("id,numero,objeto,contratista,interventoria,logo_contratista")
        .eq("id", contrato_id)
        .limit(1)
        .execute()
        .data
        or [{}]
    )
    contrato = crows[0]
    pk_set = _resolve_export_pk_ids(sb, contrato_id, pk_ids, tramos)
    resolved_target = (data.get("target_id") or "").strip()
    prog_meta: dict = {}
    if resolved_target:
        prows = (
            sb.table("prog_versiones")
            .select(
                "numero_version,tipo,estado,fecha_inicio,fecha_fin,cpm_dirty,cpm_calculado_en"
            )
            .eq("id", resolved_target)
            .limit(1)
            .execute()
            .data
            or []
        )
        prog_meta = prows[0] if prows else {}
    ppto_meta: dict = {}
    ppto_id = (data.get("version_ppto_id") or "").strip()
    if ppto_id:
        from prog_obra_costos_presupuesto import assert_ppto_version_contrato

        ppto_meta = assert_ppto_version_contrato(sb, contrato_id, ppto_id)
    cronograma = (
        fetch_cronograma_pdf_tree(sb, resolved_target, contrato_id, pk_ids=pk_set)
        if resolved_target
        else []
    )
    cpm_export = (
        build_cpm_export_data(sb, resolved_target, contrato_id, pk_set)
        if resolved_target
        else {"resultados": [], "dependencias": []}
    )
    if resolved_target:
        from prog_obra_costos_presupuesto import _fetch_agrupadores_meta
        from prog_obra_service import _dep_nodo_label, listar_dependencias

        ag_meta_all = _fetch_agrupadores_meta(sb, contrato_id)
        deps_all = listar_dependencias(sb, resolved_target)
        cpm_export["dependencias_generales"] = [
            {
                **d,
                "origen_label": _dep_nodo_label(d, "orig", ag_meta_all),
                "destino_label": _dep_nodo_label(d, "dest", ag_meta_all),
                "tipo": (d.get("tipo") or "FS").strip().upper(),
                "lag_dias": int(d.get("lag_dias") or 0),
            }
            for d in deps_all
        ]
        cpm_export["dependencias_generales"].sort(
            key=lambda x: (
                str(x.get("pk_id_origen") or ""),
                str(x.get("capitulo_origen") or ""),
                str(x.get("origen_label") or ""),
            )
        )
    resumen_ejecutivo = (
        build_resumen_ejecutivo_data(
            sb,
            contrato_id,
            version_ppto_id=ppto_id or None,
            pk_ids=pk_set,
            cronograma=cronograma,
            cpm_export=cpm_export,
            prog_meta=prog_meta,
        )
        if resolved_target
        else None
    )
    return {
        "contrato": contrato,
        "data": data,
        "cronograma": cronograma,
        "cpm_export": cpm_export,
        "prog_meta": prog_meta,
        "ppto_meta": ppto_meta,
        "resumen_ejecutivo": resumen_ejecutivo,
    }


def _write_detalle_pk_excel_sheet(
    ws,
    data: dict,
    meses: List[dict],
    *,
    numero: str,
    prog_meta: Optional[dict],
    fecha_gen: str,
) -> Optional[dict]:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    det = data.get("detalle_pk") or {}
    baseline_rows = det.get("baseline") or []
    vigente_rows = det.get("vigente") or []
    month_keys = [str(m.get("mes") or "") for m in meses if m.get("mes")]
    if not month_keys:
        month_keys = sorted(
            {
                mk
                for row in (baseline_rows + vigente_rows)
                for mk in (row.get("distribucion_mensual") or {}).keys()
            }
        )
    month_labels = {str(m.get("mes") or ""): str(m.get("mes_label") or "") for m in meses}
    same_version = str(data.get("baseline_id") or "") == str(data.get("target_id") or "")
    sections: List[Tuple[str, List[dict]]] = []
    if baseline_rows:
        sections.append(("Baseline", baseline_rows))
    if vigente_rows and not same_version:
        sections.append(("Vigente", vigente_rows))
    elif vigente_rows and not baseline_rows:
        sections.append(("Vigente", vigente_rows))

    ws.merge_cells("A1:H1")
    ws["A1"] = "Detalle por PK — Distribución mensual"
    ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Contrato {numero} · Versión programación: {_prog_version_label(prog_meta)} · "
        f"Baseline: {data.get('baseline_id') or '—'} · Vigente: {data.get('target_id') or '—'} · "
        f"Generado: {fecha_gen}"
    )
    ws["A2"].font = Font(size=9, color="64748B")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A3"] = (
        "Valores mensuales formulados: distribución lineal del costo total según Inicio, Fin y días calendario."
    )
    ws["A3"].font = Font(size=8, italic=True, color="64748B")

    hdr_row = 4
    base_headers = [
        "Línea",
        "PK",
        "Capítulo",
        "Agrupador / actividad",
        "Costo total",
        "Inicio",
        "Fin",
    ]
    first_month_col = len(base_headers) + 1

    if not sections:
        ws.cell(row=hdr_row + 1, column=1, value="Sin actividades programadas con costo en el alcance.")
        return None

    for i, mk in enumerate(month_keys):
        col = first_month_col + i
        ms = _month_key_to_date(mk)
        c = ws.cell(row=hdr_row, column=col, value=ms if ms else month_labels.get(mk, mk))
        if ms:
            c.number_format = "mmm yyyy"
        c.font = Font(bold=True, color="1E3A8A", size=8)
        c.fill = _xl_solid_fill("DBEAFE")
        c.border = _xl_thin_border()
        c.alignment = Alignment(horizontal="center", text_rotation=90)

    for col, h in enumerate(base_headers, start=1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font = Font(bold=True, color="1E3A8A", size=9)
        c.fill = _xl_solid_fill("DBEAFE")
        c.border = _xl_thin_border()

    row = hdr_row + 1
    first_data_row = row
    has_baseline = False
    has_vigente = False
    for sec_label, sec_rows in sections:
        if sec_label == "Baseline":
            has_baseline = True
        if sec_label == "Vigente":
            has_vigente = True
        for pk, pk_rows in sorted(_group_detalle_by_pk(sec_rows).items(), key=lambda x: (len(x[0]), x[0])):
            for item in pk_rows:
                ag_lbl = str(item.get("agrupador_nombre") or item.get("label") or "").strip()
                ws.cell(row=row, column=1, value=sec_label)
                ws.cell(row=row, column=2, value=pk)
                ws.cell(row=row, column=3, value=item.get("capitulo"))
                ws.cell(row=row, column=4, value=ag_lbl)
                ws.cell(row=row, column=5, value=float(item.get("costo_total") or 0)).number_format = EXCEL_COP_FMT
                fi = _parse_d(item.get("fecha_inicio"))
                ff = _parse_d(item.get("fecha_fin"))
                if fi:
                    ws.cell(row=row, column=6, value=fi).number_format = "yyyy-mm-dd"
                if ff:
                    ws.cell(row=row, column=7, value=ff).number_format = "yyyy-mm-dd"
                for i, mk in enumerate(month_keys):
                    col = first_month_col + i
                    letter = get_column_letter(col)
                    cell = ws.cell(row=row, column=col)
                    cell.value = _xl_detalle_pk_month_formula(row, letter, hdr_row=hdr_row)
                    cell.number_format = EXCEL_COP_FMT
                for col in range(1, first_month_col + len(month_keys)):
                    ws.cell(row=row, column=col).border = _xl_thin_border()
                row += 1

    last_data_row = row - 1
    month_col_by_key = {
        mk: first_month_col + i for i, mk in enumerate(month_keys)
    }

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    for i in range(len(month_keys)):
        ws.column_dimensions[get_column_letter(first_month_col + i)].width = 12
    ws.freeze_panes = f"A{hdr_row + 1}"

    return {
        "sheet": ws.title,
        "hdr_row": hdr_row,
        "first_data_row": first_data_row,
        "last_data_row": last_data_row,
        "month_cols": month_col_by_key,
        "has_baseline": has_baseline,
        "has_vigente": has_vigente,
        "vigente_section": "Vigente" if has_vigente else ("Baseline" if has_baseline else ""),
        "baseline_section": "Baseline" if has_baseline else "",
    }


def _apply_curva_s_table_formulas(
    ws,
    meses: List[dict],
    detalle_meta: Optional[dict],
    *,
    hdr_row: int,
    first_table_row: int,
) -> int:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    money_cols = {2, 3, 4, 5, 6, 7}
    pct_cols = {8, 9}
    r = first_table_row

    if not meses:
        return r - 1

    sheet = (detalle_meta or {}).get("sheet", "Detalle PK")
    d_first = (detalle_meta or {}).get("first_data_row")
    d_last = (detalle_meta or {}).get("last_data_row")
    month_cols = (detalle_meta or {}).get("month_cols") or {}
    baseline_sec = (detalle_meta or {}).get("baseline_section") or "Baseline"
    vigente_sec = (detalle_meta or {}).get("vigente_section") or "Vigente"
    use_formulas = bool(detalle_meta and d_first and d_last and d_last >= d_first)

    for m in meses:
        mk = str(m.get("mes") or "")
        ws.cell(row=r, column=1, value=m.get("mes_label")).alignment = Alignment(horizontal="left")

        if use_formulas and mk in month_cols:
            letter = get_column_letter(month_cols[mk])
            if baseline_sec:
                ws.cell(row=r, column=2).value = _xl_curva_s_sumif_detalle(
                    sheet, baseline_sec, letter, d_first, d_last,
                )
            else:
                ws.cell(row=r, column=2, value=0)
            if vigente_sec:
                ws.cell(row=r, column=4).value = _xl_curva_s_sumif_detalle(
                    sheet, vigente_sec, letter, d_first, d_last,
                )
            else:
                ws.cell(row=r, column=4, value=0)
        else:
            ws.cell(row=r, column=2, value=float(m.get("baseline_mes") or 0))
            ws.cell(row=r, column=4, value=float(m.get("vigente_mes") or 0))

        if r == first_table_row:
            ws.cell(row=r, column=3).value = f"=B{r}"
            ws.cell(row=r, column=5).value = f"=D{r}"
            ws.cell(row=r, column=7).value = f"=F{r}"
        else:
            ws.cell(row=r, column=3).value = _xl_curva_s_running_sum("B", r, first_table_row)
            ws.cell(row=r, column=5).value = _xl_curva_s_running_sum("D", r, first_table_row)
            ws.cell(row=r, column=7).value = _xl_curva_s_running_sum("F", r, first_table_row)

        ws.cell(row=r, column=6, value=float(m.get("ejecutado_mes") or 0))
        ws.cell(row=r, column=8).value = _xl_curva_s_delta_pct("E", "C", r)
        ws.cell(row=r, column=9).value = _xl_curva_s_delta_pct("G", "C", r)

        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.border = _xl_thin_border()
            if col in money_cols:
                cell.number_format = EXCEL_COP_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col in pct_cols:
                cell.number_format = EXCEL_PCT_FMT
                cell.alignment = Alignment(horizontal="right")
        r += 1

    return r - 1


def _apply_curva_s_kpi_formulas(
    ws,
    detalle_meta: Optional[dict],
    *,
    last_table_row: int,
    first_table_row: int,
    presupuesto_fallback: float,
) -> None:
    from openpyxl.styles import Font

    if last_table_row < first_table_row:
        ws.cell(row=6, column=1, value=presupuesto_fallback).number_format = EXCEL_COP_FMT
        return

    sheet = (detalle_meta or {}).get("sheet", "Detalle PK")
    d_first = (detalle_meta or {}).get("first_data_row")
    d_last = (detalle_meta or {}).get("last_data_row")
    vigente_sec = (detalle_meta or {}).get("vigente_section") or "Vigente"
    use_formulas = bool(detalle_meta and d_first and d_last and d_last >= d_first)

    if use_formulas and vigente_sec:
        ws.cell(row=6, column=1).value = (
            f"=SUMIF('{sheet}'!$A${d_first}:$A${d_last},\"{vigente_sec}\","
            f"'{sheet}'!$E${d_first}:$E${d_last})"
        )
    else:
        ws.cell(row=6, column=1, value=presupuesto_fallback)
    ws.cell(row=6, column=1).number_format = EXCEL_COP_FMT

    ws.cell(row=6, column=2).value = f"=E{last_table_row}"
    ws.cell(row=6, column=2).number_format = EXCEL_COP_FMT
    ws.cell(row=6, column=3).value = f'=IF($A$6=0,"",TEXT(E{last_table_row}/$A$6,"0.0%"))'
    ws.cell(row=6, column=3).font = Font(size=9, color="64748B")

    ws.cell(row=6, column=4).value = f"=G{last_table_row}"
    ws.cell(row=6, column=4).number_format = EXCEL_COP_FMT
    ws.cell(row=6, column=5).value = f'=IF($A$6=0,"",TEXT(G{last_table_row}/$A$6,"0.0%"))'
    ws.cell(row=6, column=5).font = Font(size=9, color="64748B")

    ws.cell(row=6, column=6).value = f"=G{last_table_row}-E{last_table_row}"
    ws.cell(row=6, column=6).number_format = EXCEL_COP_FMT
    ws.cell(row=6, column=7).value = f'=IF(E{last_table_row}=0,"",TEXT(G{last_table_row}/E{last_table_row}-1,"0.0%"))'
    ws.cell(row=6, column=7).font = Font(size=9, color="64748B")


def _write_resumen_ejecutivo_excel_sheet(
    ws,
    resumen: Optional[dict],
    *,
    contrato: dict,
    data: dict,
    prog_meta: Optional[dict],
    ppto_meta: Optional[dict],
    fecha_gen: str,
) -> None:
    from openpyxl.styles import Alignment, Font

    ind = data.get("indicadores") or {}
    numero = contrato.get("numero") or contrato.get("id") or ""
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Resumen ejecutivo — Contrato {numero}"
    ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"{(contrato.get('objeto') or '').strip()} · "
        f"{(contrato.get('contratista') or '').strip()} · "
        f"Interventoría: {(contrato.get('interventoria') or '').strip()} · Generado: {fecha_gen}"
    )
    ws["A2"].font = Font(size=9, color="64748B")
    ws["A2"].alignment = Alignment(wrap_text=True)

    row = 4
    ws.cell(row=row, column=1, value="Indicadores Curva S").font = Font(bold=True, size=10)
    row += 1
    for label, key in [
        ("Presupuesto total", "presupuesto_total"),
        ("Programado a la fecha", "programado_a_fecha"),
        ("Ejecutado a la fecha", "ejecutado_a_fecha"),
        ("Desviación", "desviacion_valor"),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=9)
        cell = ws.cell(row=row, column=2, value=float(ind.get(key) or 0))
        cell.number_format = EXCEL_COP_FMT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Resumen por capítulo").font = Font(bold=True, size=10)
    row += 1
    headers = ["Capítulo", "Agrupadores WBS", "Inicio", "Fin", "Días hábiles"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="1E3A8A", size=9)
        c.fill = _xl_solid_fill("DBEAFE")
        c.border = _xl_thin_border()
    row += 1
    res = resumen or {}
    caps = res.get("capitulos") or []
    if not caps:
        ws.cell(row=row, column=1, value="Sin capítulos en el alcance.")
        row += 1
    for cap in caps:
        if cap.get("sin_programar"):
            vals = [cap.get("capitulo"), 0, "Sin programar", "", ""]
        else:
            vals = [
                cap.get("capitulo"),
                cap.get("num_agrupadores"),
                cap.get("fecha_inicio"),
                cap.get("fecha_fin"),
                cap.get("dias_habiles"),
            ]
        for col, val in enumerate(vals, start=1):
            ws.cell(row=row, column=col, value=val).border = _xl_thin_border()
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Ruta crítica").font = Font(bold=True, size=10)
    row += 1
    if not res.get("tiene_cpm"):
        ws.cell(row=row, column=1, value="No hay resultados CPM calculados para esta versión.")
        row += 2
    else:
        rc_headers = ["Agrupador WBS", "Inicio", "Fin", "Holgura"]
        for col, h in enumerate(rc_headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(bold=True, color="1E3A8A", size=9)
            c.fill = _xl_solid_fill("DBEAFE")
        row += 1
        criticos = res.get("criticos") or []
        if not criticos:
            ws.cell(row=row, column=1, value="Ningún agrupador en ruta crítica.")
            row += 1
        for cr in criticos:
            ws.cell(row=row, column=1, value=cr.get("nombre"))
            ws.cell(row=row, column=2, value=cr.get("fecha_inicio"))
            ws.cell(row=row, column=3, value=cr.get("fecha_fin"))
            ws.cell(row=row, column=4, value=cr.get("holgura"))
            row += 1
        row += 1
        ws.cell(row=row, column=1, value="Alertas de desfase").font = Font(bold=True, size=10)
        row += 1
        desf_headers = ["Agrupador WBS", "Fin calculada", "Días exceso"]
        for col, h in enumerate(desf_headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(bold=True, color="1E3A8A", size=9)
            c.fill = _xl_solid_fill("DBEAFE")
        row += 1
        desfases = res.get("desfases") or []
        if not desfases:
            ws.cell(row=row, column=1, value="Ningún agrupador supera la fecha fin de la versión.")
            row += 1
        for d in desfases:
            ws.cell(row=row, column=1, value=d.get("nombre"))
            ws.cell(row=row, column=2, value=d.get("fecha_fin"))
            ws.cell(row=row, column=3, value=d.get("dias_exceso"))
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Estado de la versión activa").font = Font(bold=True, size=10)
    row += 1
    ver = res.get("version") or {}
    for label, val in [
        ("Versión", ver.get("nombre") or _prog_version_label(prog_meta)),
        ("Estado", ver.get("estado")),
        ("Inicio cronograma", ver.get("fecha_inicio")),
        ("Fin cronograma", ver.get("fecha_fin")),
        ("Estado CPM", ver.get("cpm_estado") if res.get("tiene_cpm") else "Sin calcular"),
        ("Versión presupuesto", _ppto_version_label(ppto_meta)),
        ("Baseline programación", data.get("baseline_id")),
        ("Vigente programación", data.get("target_id")),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=9)
        ws.cell(row=row, column=2, value=val if val is not None else "—")
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14


def build_curva_s_xlsx_bytes(
    contrato: dict,
    data: dict,
    *,
    cronograma: Optional[List[dict]] = None,
    prog_meta: Optional[dict] = None,
    ppto_meta: Optional[dict] = None,
    cpm_export: Optional[dict] = None,
    resumen_ejecutivo: Optional[dict] = None,
    fecha_generacion: Optional[str] = None,
) -> bytes:
    """Excel: Curva S + detalle PK + resumen ejecutivo (+ Gantt/CPM opcionales)."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    ind = data.get("indicadores") or {}
    meses = data.get("meses") or []
    numero = contrato.get("numero") or contrato.get("id") or ""
    objeto = (contrato.get("objeto") or "").strip()
    contratista = (contrato.get("contratista") or "").strip()
    interventoria = (contrato.get("interventoria") or "").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Curva S"

    fecha_gen = (fecha_generacion or "").strip()
    if not fecha_gen:
        from informes import _fmt_informe_fecha_generacion

        fecha_gen = _fmt_informe_fecha_generacion()

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Curva S — Contrato {numero}"
    ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws.merge_cells("A2:I2")
    ws["A2"] = f"{objeto} · {contratista} · Interventoría: {interventoria}"
    ws["A2"].font = Font(size=9, color="64748B")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A3"] = (
        f"Baseline: {data.get('baseline_id') or '—'} · "
        f"Vigente: {data.get('target_id') or '—'} · Generado: {fecha_gen}"
    )
    ws["A3"].font = Font(size=8, color="64748B")

    kpi_labels = [
        (1, "Presupuesto total"),
        (2, "Programado a la fecha"),
        (4, "Ejecutado a la fecha"),
        (6, "Desviación"),
    ]
    for col, label in kpi_labels:
        c_l = ws.cell(row=5, column=col, value=label)
        c_l.font = Font(bold=True, size=8, color="64748B")
        c_l.fill = _xl_solid_fill("EFF6FF")

    ws_det = wb.create_sheet("Detalle PK")
    detalle_meta = _write_detalle_pk_excel_sheet(
        ws_det, data, meses, numero=str(numero), prog_meta=prog_meta, fecha_gen=fecha_gen
    )

    hdr_row = 20
    first_table_row = hdr_row + 1
    headers = [
        "Mes",
        "Base. mes",
        "Base. acum.",
        "Vig. mes",
        "Vig. acum.",
        "Ejec. mes",
        "Ejec. acum.",
        "Δ Vig.",
        "Δ Ejec.",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = Font(bold=True, color="1E3A8A", size=9)
        cell.fill = _xl_solid_fill("DBEAFE")
        cell.border = _xl_thin_border()
        cell.alignment = Alignment(horizontal="center" if col == 1 else "right", vertical="center")

    last_table_row = _apply_curva_s_table_formulas(
        ws,
        meses,
        detalle_meta,
        hdr_row=hdr_row,
        first_table_row=first_table_row,
    )
    _apply_curva_s_kpi_formulas(
        ws,
        detalle_meta,
        last_table_row=last_table_row,
        first_table_row=first_table_row,
        presupuesto_fallback=float(ind.get("presupuesto_total") or 0),
    )
    for col in (1, 2, 4, 6):
        ws.cell(row=6, column=col).font = Font(bold=True, size=10, color="2563EB")

    if meses:
        try:
            png = _chart_png_bytes(meses)
            img = XLImage(io.BytesIO(png))
            img.width = 690
            img.height = 180
            ws.add_image(img, "A8")
            ws.row_dimensions[8].height = 135
        except Exception:
            ws["A8"] = "Gráfico no disponible"

    ws_res = wb.create_sheet("Resumen ejecutivo")
    _write_resumen_ejecutivo_excel_sheet(
        ws_res,
        resumen_ejecutivo,
        contrato=contrato,
        data=data,
        prog_meta=prog_meta,
        ppto_meta=ppto_meta,
        fecha_gen=fecha_gen,
    )

    foot = last_table_row + 2 if last_table_row >= first_table_row else 22
    ws.cell(row=foot, column=1, value="Versión presupuesto activa").font = Font(bold=True, size=8, color="64748B")
    ws.cell(row=foot + 1, column=1, value=_ppto_version_label(ppto_meta))
    ws.cell(row=foot, column=5, value="Versión programación").font = Font(bold=True, size=8, color="64748B")
    ws.cell(row=foot + 1, column=5, value=_prog_version_label(prog_meta))
    ws.cell(row=foot, column=7, value="Generado").font = Font(bold=True, size=8, color="64748B")
    ws.cell(row=foot + 1, column=7, value=fecha_gen)

    ws.column_dimensions["A"].width = 12
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.freeze_panes = "A21"

    flat = _flatten_cronograma_rows(cronograma or [])
    detalle_vigente = (data.get("detalle_pk") or {}).get("vigente") or []
    if flat:
        ws_g = wb.create_sheet("Gantt")
        _write_gantt_sheet(
            ws_g,
            flat=flat,
            meses_data=meses,
            detalle_vigente=detalle_vigente,
            numero=str(numero),
            prog_meta=prog_meta,
            fecha_gen=fecha_gen,
        )

    export = cpm_export or {}
    resultados_cpm = export.get("resultados") or []
    dependencias_cpm = export.get("dependencias") or []
    if resultados_cpm:
        ws_cpm = wb.create_sheet("CPM")
        _write_cpm_excel_sheet(ws_cpm, resultados_cpm)
    if dependencias_cpm:
        ws_dep = wb.create_sheet("Dependencias")
        _write_dependencias_excel_sheet(ws_dep, dependencias_cpm)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

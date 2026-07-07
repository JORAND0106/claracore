"""
Exportación Excel (.xlsx) — módulo Contabilidad.
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from contabilidad_reportes import (
    EXPORT_TIPOS,
    MESES_ES,
    obtener_reporte,
    reporte_deducciones_tributarias,
    reporte_evolucion_mensual,
    reporte_ingresos_centro_costo,
    reporte_cuentas_especiales_historico,
    reporte_resumen_dashboard,
)
from contabilidad_service import get_cierre, list_transacciones

_FILL_HDR = PatternFill("solid", fgColor="0077B6")
_FONT_HDR = Font(bold=True, color="FFFFFF", size=10)
_FONT_TITLE = Font(bold=True, size=14, color="0077B6")
_SIDE = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
_AL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_AL_RIGHT = Alignment(horizontal="right", vertical="center")
_AL_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

_LOGO_BYTES: Optional[bytes] = None
_LOGO_RESOLVED = False


def _logo_claracore_bytes() -> Optional[bytes]:
    global _LOGO_BYTES, _LOGO_RESOLVED
    if _LOGO_RESOLVED:
        return _LOGO_BYTES
    _LOGO_RESOLVED = True
    base = Path(__file__).resolve().parent / "assets"
    for name in ("CLARA.CORE.png", "claracore-logo.png", "logo-claracore.png"):
        path = base / name
        if path.is_file():
            _LOGO_BYTES = path.read_bytes()
            return _LOGO_BYTES
    _LOGO_BYTES = None
    return None


def _write_sheet_header(ws, title: str, ncols: int) -> int:
    """Fila 1: logo ClaraCore (col A) + título. Devuelve la fila para subtítulo o tabla."""
    ws.row_dimensions[1].height = 52
    col_a = get_column_letter(1)
    ws.column_dimensions[col_a].width = max(float(ws.column_dimensions[col_a].width or 0), 18)

    logo_bytes = _logo_claracore_bytes()
    if logo_bytes:
        try:
            img = XLImage(io.BytesIO(logo_bytes))
            target_h = 44
            if img.height:
                img.width = int(img.width * (target_h / img.height))
            img.height = target_h
            ws.add_image(img, "A1")
        except Exception:
            ws.cell(row=1, column=1, value="ClaraCore")

    title_col = 2 if ncols > 1 else 1
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=title_col, value=title)
    cell.font = _FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return 3


def _write_title(ws, title: str, ncols: int) -> int:
    """Encabezado con logo ClaraCore."""
    return _write_sheet_header(ws, title, ncols)


def _safe_sheet_name(name: str, fallback: str = "Hoja") -> str:
    bad = set(r'[]:*?/\\')
    s = "".join(c for c in (name or fallback) if c not in bad)[:31]
    return s or fallback


def _style_header_row(ws, row_idx: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = _FILL_HDR
        cell.font = _FONT_HDR
        cell.alignment = _AL_CENTER
        cell.border = _BORDER


def _autosize_columns(ws, max_width: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_evolucion(ws, data: dict) -> None:
    ncols = 7
    row = _write_title(ws, "Evolución mensual — Ingresos vs Egresos", ncols)
    ws.cell(row=row, column=1, value=f"Desde {data.get('fecha_desde')} hasta {data.get('fecha_hasta')}")
    row += 2
    headers = [
        "Período", "Ingresos brutos", "Egresos brutos", "Ingresos neto",
        "Egresos neto", "Utilidad neta", "Transacciones",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, ncols)
    row += 1
    for s in data.get("series") or []:
        ws.append([
            s.get("periodo_label") or s.get("periodo"),
            s.get("ingresos_brutos"),
            s.get("egresos_brutos"),
            s.get("ingresos_neto"),
            s.get("egresos_neto"),
            s.get("utilidad_neta"),
            s.get("transacciones"),
        ])
        for c in range(2, 7):
            ws.cell(row=row, column=c).alignment = _AL_RIGHT
            ws.cell(row=row, column=c).number_format = '#,##0.00'
        row += 1
    _autosize_columns(ws)


def _sheet_centro_costo(ws, data: dict) -> None:
    ncols = 5
    row = _write_title(ws, "Distribución de ingresos por centro de costo", ncols)
    ws.cell(row=row, column=1, value=f"Total ingresos: {data.get('total_ingresos_brutos', 0):,.2f}")
    row += 2
    headers = ["Centro de costo", "Tipo", "Contrato ID", "Ingresos brutos", "% del total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, ncols)
    row += 1
    for it in data.get("items") or []:
        ws.append([
            it.get("label"),
            it.get("centro_costo_tipo"),
            it.get("contrato_id"),
            it.get("ingresos_brutos"),
            it.get("porcentaje"),
        ])
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '0.00"%"'
        row += 1
    _autosize_columns(ws)


def _sheet_cuentas_especiales(ws, data: dict) -> None:
    ncols = 7
    row = _write_title(ws, "Saldo acumulado — Cuentas especiales", ncols)
    row += 1
    headers = [
        "Período", "Operativa", "Cap. licenciamiento", "Cap. servicios",
        "Cap. total", "IVA neto", "Retenciones",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, ncols)
    row += 1
    for s in data.get("series") or []:
        ws.append([
            s.get("periodo_label") or s.get("periodo"),
            s.get("operativa"),
            s.get("capitalizacion_lic"),
            s.get("capitalizacion_srv"),
            s.get("capitalizacion_total"),
            s.get("impuestos_iva_neto"),
            s.get("impuestos_retencion"),
        ])
        for c in range(2, 8):
            ws.cell(row=row, column=c).number_format = '#,##0.00'
        row += 1
    _autosize_columns(ws)


def _sheet_deducciones(ws, data: dict) -> None:
    ncols = 7
    row = _write_title(ws, "Composición de deducciones tributarias", ncols)
    row += 1
    headers = [
        "Período", "Retención fuente", "IVA recaudado", "IVA pagado",
        "IVA neto", "Total deducciones", "",
    ]
    headers = headers[:ncols]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, ncols)
    row += 1
    for s in data.get("series") or []:
        ws.append([
            s.get("periodo_label") or s.get("periodo"),
            s.get("retencion_fuente"),
            s.get("iva_recaudado"),
            s.get("iva_pagado"),
            s.get("iva_neto"),
            s.get("total_deducciones"),
        ])
        for c in range(2, 7):
            ws.cell(row=row, column=c).number_format = '#,##0.00'
        row += 1
    _autosize_columns(ws)


def _sheet_transacciones(ws, txs: List[dict], meta: dict) -> None:
    ncols = 12
    row = _write_title(ws, "Libro de transacciones", ncols)
    ws.cell(row=row, column=1, value=f"Período: {meta.get('fecha_desde')} — {meta.get('fecha_hasta')}")
    row += 2
    headers = [
        "ID", "Fecha", "Tipo", "Bruto", "Retención", "IVA", "Neto",
        "Categoría", "Centro costo", "Contrato", "Origen", "Notas",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, ncols)
    row += 1
    for tx in txs:
        cat = (tx.get("categoria") or {}).get("nombre") or tx.get("categoria_id")
        contrato = tx.get("contrato")
        c_label = ""
        if contrato:
            c_label = f"{contrato.get('numero') or contrato.get('id')}"
        elif (tx.get("centro_costo_tipo") or "") == "empresa":
            c_label = "Empresa general"
        ws.append([
            tx.get("id"),
            tx.get("fecha"),
            tx.get("tipo"),
            tx.get("valor_bruto"),
            tx.get("retencion_fuente_valor"),
            tx.get("iva_valor"),
            tx.get("valor_neto"),
            cat,
            tx.get("centro_costo_tipo"),
            c_label,
            tx.get("origen"),
            (tx.get("notas") or "")[:200],
        ])
        for c in range(4, 8):
            ws.cell(row=row, column=c).number_format = '#,##0.00'
        row += 1
    _autosize_columns(ws)


def _sheet_cierre(ws, cierre: dict) -> None:
    ncols = 4
    anio, mes = int(cierre.get("anio") or 0), int(cierre.get("mes") or 0)
    titulo = f"Cierre mensual — {MESES_ES[mes] if 1 <= mes <= 12 else mes} {anio}"
    row = _write_title(ws, titulo, ncols)
    row += 1

    def _kv(label: str, value: Any, r: int) -> int:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        return r + 1

    pairs = [
        ("Estado", cierre.get("estado")),
        ("Ingresos brutos", cierre.get("ingresos_brutos")),
        ("Total deducciones", cierre.get("total_deducciones")),
        ("Total gastos", cierre.get("total_gastos")),
        ("Utilidad neta", cierre.get("utilidad_neta")),
        ("Flujo de caja neto", cierre.get("flujo_caja_neto")),
        ("Saldo operativa", cierre.get("saldo_operativa")),
        ("Saldo cap. licenciamiento", cierre.get("saldo_capitalizacion_lic")),
        ("Saldo cap. servicios", cierre.get("saldo_capitalizacion_srv")),
        ("IVA neto acumulado", cierre.get("saldo_impuestos_iva_neto")),
        ("Retenciones acumuladas", cierre.get("saldo_impuestos_retencion")),
        ("Notas contador", cierre.get("notas_contador")),
        ("Aprobado por", (cierre.get("aprobado_por") or {}).get("nombre")),
        ("Aprobado at", cierre.get("aprobado_at")),
        ("Firmado por", (cierre.get("firmado_por") or {}).get("nombre")),
        ("Firmado at", cierre.get("firmado_at")),
        ("Hash firma", cierre.get("firma_contenido_hash")),
    ]
    for label, val in pairs:
        row = _kv(label, val, row)
        if isinstance(val, (int, float)) and label != "Estado":
            ws.cell(row=row - 1, column=2).number_format = '#,##0.00'
    _autosize_columns(ws)


def build_export_xlsx(
    sb,
    tipo: str,
    *,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    anio: Optional[int] = None,
    cierre_id: Optional[int] = None,
) -> Tuple[bytes, str]:
    t = (tipo or "").strip().lower()
    if t not in EXPORT_TIPOS:
        raise ValueError(f"Tipo de exportación inválido: {tipo}")

    ts = datetime.now().strftime("%Y%m%d")
    wb = Workbook()
    wb.remove(wb.active)

    kwargs = {"fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta, "anio": anio}

    if t == "completo":
        ws1 = wb.create_sheet(_safe_sheet_name("Evolución"))
        _sheet_evolucion(ws1, reporte_evolucion_mensual(sb, **kwargs))
        ws2 = wb.create_sheet(_safe_sheet_name("Centros costo"))
        _sheet_centro_costo(ws2, reporte_ingresos_centro_costo(sb, **kwargs))
        ws3 = wb.create_sheet(_safe_sheet_name("Cuentas esp."))
        _sheet_cuentas_especiales(ws3, reporte_cuentas_especiales_historico(sb, **kwargs))
        ws4 = wb.create_sheet(_safe_sheet_name("Deducciones"))
        _sheet_deducciones(ws4, reporte_deducciones_tributarias(sb, **kwargs))
        res = list_transacciones(
            sb,
            fecha_desde=fecha_desde or (f"{anio}-01-01" if anio else None),
            fecha_hasta=fecha_hasta or (f"{anio}-12-31" if anio else None),
            estado="activa",
            limit=5000,
        )
        ws5 = wb.create_sheet(_safe_sheet_name("Transacciones"))
        _sheet_transacciones(ws5, res.get("items") or [], {
            "fecha_desde": fecha_desde or (f"{anio}-01-01" if anio else ""),
            "fecha_hasta": fecha_hasta or (f"{anio}-12-31" if anio else ""),
        })
        filename = f"ClaraCore_Contabilidad_Completo_{ts}.xlsx"
    elif t == "transacciones":
        res = list_transacciones(
            sb,
            fecha_desde=fecha_desde or (f"{anio}-01-01" if anio else None),
            fecha_hasta=fecha_hasta or (f"{anio}-12-31" if anio else None),
            estado="activa",
            limit=5000,
        )
        ws = wb.create_sheet(_safe_sheet_name("Transacciones"))
        _sheet_transacciones(ws, res.get("items") or [], {
            "fecha_desde": fecha_desde or "",
            "fecha_hasta": fecha_hasta or "",
        })
        filename = f"ClaraCore_Contabilidad_Transacciones_{ts}.xlsx"
    elif t == "cierre":
        if cierre_id is None:
            raise ValueError("cierre_id requerido para exportar cierre.")
        cierre = get_cierre(sb, int(cierre_id))
        ws = wb.create_sheet(_safe_sheet_name(f"Cierre {cierre.get('mes')}-{cierre.get('anio')}"))
        _sheet_cierre(ws, cierre)
        filename = f"ClaraCore_Cierre_{cierre.get('anio')}_{int(cierre.get('mes') or 0):02d}_{ts}.xlsx"
    elif t == "resumen":
        data = reporte_resumen_dashboard(sb, **kwargs)
        ws1 = wb.create_sheet("Evolución")
        _sheet_evolucion(ws1, data["evolucion_mensual"])
        ws2 = wb.create_sheet("Centros costo")
        _sheet_centro_costo(ws2, data["ingresos_centro_costo"])
        ws3 = wb.create_sheet("Cuentas esp.")
        _sheet_cuentas_especiales(ws3, data["cuentas_especiales"])
        ws4 = wb.create_sheet("Deducciones")
        _sheet_deducciones(ws4, data["deducciones_tributarias"])
        filename = f"ClaraCore_Contabilidad_Reportes_{ts}.xlsx"
    else:
        data = obtener_reporte(sb, t, **kwargs)
        ws = wb.create_sheet(_safe_sheet_name(t))
        if t == "evolucion-mensual":
            _sheet_evolucion(ws, data)
        elif t == "ingresos-centro-costo":
            _sheet_centro_costo(ws, data)
        elif t == "cuentas-especiales":
            _sheet_cuentas_especiales(ws, data)
        elif t == "deducciones-tributarias":
            _sheet_deducciones(ws, data)
        filename = f"ClaraCore_Contabilidad_{t.replace('-', '_')}_{ts}.xlsx"

    return _workbook_to_bytes(wb), filename

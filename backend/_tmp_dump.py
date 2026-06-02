import io, sys
from openpyxl import load_workbook

path = r"C:\Users\JORAND\Downloads\ClaraCore_4.ESPACIO PUBLICO_2026-06-02 (1).xlsx"
wb = load_workbook(path, data_only=True)
out = []
out.append(f"SHEETS: {wb.sheetnames}")
for sn in wb.sheetnames:
    ws = wb[sn]
    out.append(f"\n===== SHEET: {sn}  dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column} =====")
    for r in range(1, min(ws.max_row, 80) + 1):
        cells = []
        for c in range(1, min(ws.max_column, 10) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            cells.append(f"[{c}]{v}")
        if cells:
            out.append(f"r{r}: " + " | ".join(str(x) for x in cells))

with open("_tmp_dump.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(out))
print("done", len(out), "lines")

"""Informe Excel completo por capítulo (resumen ejecutivo + hoja por ítem)."""
from __future__ import annotations

import io
import math
import re
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pytz

_ITEM_XLSX_NCOLS = 11
_ITEM_DET_HEADERS = (
    "Registro",
    "Acta RPO",
    "PK_ID",
    "Abs Inicio",
    "Abs Final",
    "Long",
    "Ancho",
    "Espesor",
    "Cantidad",
    "Cant. Total",
    "Observaciones",
)
_ITEM_ACTA_PAIRS_PER_ROW = 5
_C_CANTIDAD = 9
_C_CANT_TOTAL = 10
_C_OBS = 11
_BALANCE_OBS = "Ajuste de Cantidades por Balance de Obra"


def _detail_row_from_sicoe(r: dict) -> dict:
    cant = float(r.get("cantidad") or 0)
    cant_tot = float(r.get("cantidad_total") if r.get("cantidad_total") is not None else cant)
    acta = r.get("acta_rpo_numero")
    if acta is None:
        acta = r.get("acta_rpo_id") or ""
    return {
        "registro": r.get("numero_registro") or r.get("id_pol") or "",
        "pk_id": r.get("pk_id") or r.get("pk_id_valor") or "",
        "abs_inicio": r.get("abs_inicio") or r.get("no_inicio") or "",
        "abs_final": r.get("abs_final") or r.get("no_final") or "",
        "longitud": r.get("longitud") or r.get("area_long_nod"),
        "ancho": r.get("ancho"),
        "espesor": r.get("espesor"),
        "cantidad": cant,
        "cant_total": cant_tot,
        "observaciones": (r.get("observacion") or "").strip(),
        "acta_rpo": str(acta).strip() if acta not in (None, "") else "—",
        "is_balance_adj": False,
    }


def _detail_balance_row(delta_cant: float) -> dict:
    """Fila de ajuste: solo cantidad total; reconcilia cobrado → Cant Final ClaraCore."""
    return {
        "registro": "",
        "acta_rpo": "",
        "pk_id": "",
        "abs_inicio": "",
        "abs_final": "",
        "longitud": "",
        "ancho": "",
        "espesor": "",
        "cantidad": "",
        "cant_total": round(float(delta_cant), 4),
        "observaciones": _BALANCE_OBS,
        "is_balance_adj": True,
    }


def _build_sicoe_detail_rows(sicoe_rows: List[dict], cant_final: float) -> List[dict]:
    """Detalle 100 % cobrado SICOE + fila de balance si Cant Final ≠ Σ cobrado."""
    detail = [_detail_row_from_sicoe(r) for r in (sicoe_rows or [])]
    cob_sum = round(sum(float(d.get("cant_total") or 0) for d in detail), 4)
    cc = round(float(cant_final or 0), 4)
    delta = round(cc - cob_sum, 4)
    if detail and abs(delta) > 1e-6:
        detail.append(_detail_balance_row(delta))
    return detail


def _sicoe_acta_groups(sicoe_rows: List[dict]) -> List[Tuple[str, float]]:
    agg: Dict[str, float] = defaultdict(float)
    for r in sicoe_rows or []:
        d = _detail_row_from_sicoe(r)
        lbl = d["acta_rpo"] or "—"
        agg[lbl] += float(d["cant_total"] or 0)
    return sorted(agg.items(), key=lambda x: (x[0] != "—", str(x[0])))


def _layout_item_sheet_rows(
    n_actas: int,
    n_detail: int,
    *,
    section_start: int = 9,
) -> Dict[str, int]:
    """``section_start``: primera fila libre tras encabezado + meta ítem + subtítulo."""
    pairs = _ITEM_ACTA_PAIRS_PER_ROW
    acta_blocks = max(1, math.ceil(max(n_actas, 1) / pairs))
    acta_rows = acta_blocks * 2
    acta_title = section_start
    acta_start = section_start + 1
    acta_end = acta_start + acta_rows - 1
    balance_row = acta_end + 1
    cant_final_row = acta_end + 2
    detail_hdr = acta_end + 4
    detail_first = detail_hdr + 1
    detail_last = detail_first + max(n_detail, 1) - 1
    total_row = detail_last + 1
    return {
        "acta_title": acta_title,
        "acta_start": acta_start,
        "acta_end": acta_end,
        "balance_row": balance_row,
        "cant_final_row": cant_final_row,
        "detail_hdr": detail_hdr,
        "detail_first": detail_first,
        "detail_last": detail_last,
        "total_row": total_row,
    }


def _write_item_acta_resumen_grid(
    ws,
    start_row: int,
    acta_groups: List[Tuple[str, float]],
    *,
    sumif_rng_acta: Optional[str],
    sumif_rng_cant: Optional[str],
    fill_hdr,
    fill_sub,
    fill_white,
    border_tbl,
    al_center,
    al_right,
    font_hdr,
    font_bold,
    fmt_cant,
):
    """Grid Acta RPO | Cant. Total repartido en hasta 10 columnas (5 pares por fila)."""
    from openpyxl.styles import Font

    pairs = _ITEM_ACTA_PAIRS_PER_ROW
    entries = acta_groups or [("—", 0.0)]
    idx = 0
    row = start_row
    cant_cells: List[str] = []

    while idx < len(entries):
        chunk = entries[idx : idx + pairs]
        for pi, (acta_lbl, _qty) in enumerate(chunk):
            c_acta = 1 + pi * 2
            c_cant = c_acta + 1
            for c, lbl in ((c_acta, "Acta RPO"), (c_cant, "Cant. Total")):
                cell = ws.cell(row=row, column=c, value=lbl)
                cell.fill = fill_hdr
                cell.font = font_hdr
                cell.alignment = al_center
                cell.border = border_tbl
            lbl_safe = str(acta_lbl).replace('"', '""')
            ws.cell(row=row + 1, column=c_acta, value=acta_lbl).alignment = al_center
            ws.cell(row=row + 1, column=c_acta).border = border_tbl
            ws.cell(row=row + 1, column=c_acta).font = Font(size=10, color="0A1628")
            if sumif_rng_acta and sumif_rng_cant:
                cant_val = f'=SUMIF({sumif_rng_acta},"{lbl_safe}",{sumif_rng_cant})'
            else:
                cant_val = round(float(_qty or 0), 4)
            cant_cell = ws.cell(row=row + 1, column=c_cant, value=cant_val)
            cant_cell.number_format = fmt_cant
            cant_cell.alignment = al_right
            cant_cell.border = border_tbl
            cant_cells.append(cant_cell.coordinate)
        for c in range(len(chunk) * 2 + 1, _ITEM_XLSX_NCOLS + 1):
            for rr in (row, row + 1):
                cell = ws.cell(row=rr, column=c)
                cell.border = border_tbl
                cell.fill = fill_sub if rr == row else fill_white
        idx += pairs
        row += 2
    return row, cant_cells


def _write_item_sheet_completo(
    main,
    ws,
    *,
    cap_raw: str,
    item: str,
    drill: dict,
    ppto_rows: List[dict],
    detail_rows: List[dict],
    sicoe_rows: List[dict],
    vista: str,
    meta: dict,
    version_lbl: str,
    gen_ts: str,
    descargado_por: str,
    listado_row: Optional[dict] = None,
):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tipo_label = (
        "Obra Ejecutada"
        if main.parse_dash_vista(vista) == main.DASH_VISTA_OBRA_EJECUTADA
        else "Presupuesto de Obra"
    )
    titulo = f"Ítem {item} · {cap_raw}"
    subtitulo = f"Soporte de cantidades · Obra cobrada (SICOE aprobada) · vista: {tipo_label}"
    cant_final_cc = round(float(drill.get("total_claracore_cant") or 0), 4)

    lp = listado_row or {}
    descripcion = (
        (lp.get("descripcion") or "").strip()
        or (drill.get("descripcion") or "").strip()
        or ((ppto_rows[0].get("descripcion") or "").strip() if ppto_rows else "")
    )
    und = (
        (lp.get("unidad") or "").strip()
        or (drill.get("unidad") or drill.get("und") or "").strip()
        or ((ppto_rows[0].get("und") or "").strip() if ppto_rows else "")
    )
    vlr = lp.get("precio_unitario")
    try:
        vlr_f = round(float(vlr or 0), 2)
    except (TypeError, ValueError):
        vlr_f = 0.0

    fill_hdr = PatternFill("solid", fgColor="4472C4")
    fill_sub = PatternFill("solid", fgColor="E6F4FA")
    fill_tot = PatternFill("solid", fgColor="111827")
    fill_zebra = PatternFill("solid", fgColor="F8FAFC")
    fill_white = PatternFill("solid", fgColor="FFFFFF")
    _side = Side(style="thin", color="FF7A7A7A")
    border_tbl = Border(left=_side, right=_side, top=_side, bottom=_side)
    al_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_right = Alignment(horizontal="right", vertical="center")
    al_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    al_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    font_hdr = Font(bold=True, color="FFFFFF", size=10)
    font_bold = Font(bold=True, color="0A1628", size=10)
    fmt_cant = main._XLSX_FMT_CANT
    L = get_column_letter

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    for c in "FGHI":
        ws.column_dimensions[c].width = 10
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 28

    section_start = main._xlsx_apply_informe_header(
        ws,
        titulo=titulo,
        ncols=_ITEM_XLSX_NCOLS,
        contrato_meta=meta,
        version_lbl=version_lbl,
        gen_ts=gen_ts,
        descargado_por=descargado_por,
        subtitulo_tabla=subtitulo,
        label_merge_end=2,
        item_meta={
            "capitulo": cap_raw,
            "item": item,
            "descripcion": descripcion,
            "und": und,
            "vlr_unitario": vlr_f,
        },
    )

    acta_groups = _sicoe_acta_groups(sicoe_rows)
    n_detail_rows = max(len(detail_rows), 1)
    layout = _layout_item_sheet_rows(
        len(acta_groups), n_detail_rows, section_start=section_start
    )

    ws.cell(row=layout["acta_title"], column=1, value="Resumen por Acta RPO (cobrado aprobado)").font = Font(
        bold=True, size=10, color="0A1628"
    )

    df = layout["detail_first"]
    dl = layout["detail_last"]
    # SUMIF sobre columna Acta RPO (B) y Cant. Total (J) del detalle SICOE (sin fila balance).
    sicoe_only_last = df + max(len(sicoe_rows), 1) - 1 if sicoe_rows else df
    if sicoe_rows:
        sumif_rng_acta = f"$B${df}:$B${sicoe_only_last}"
        sumif_rng_cant = f"${L(_C_CANT_TOTAL)}${df}:${L(_C_CANT_TOTAL)}${sicoe_only_last}"
    else:
        sumif_rng_acta = sumif_rng_cant = None

    _, cant_cells = _write_item_acta_resumen_grid(
        ws,
        layout["acta_start"],
        acta_groups,
        sumif_rng_acta=sumif_rng_acta,
        sumif_rng_cant=sumif_rng_cant,
        fill_hdr=fill_hdr,
        fill_sub=fill_sub,
        fill_white=fill_white,
        border_tbl=border_tbl,
        al_center=al_center,
        al_right=al_right,
        font_hdr=font_hdr,
        font_bold=font_bold,
        fmt_cant=fmt_cant,
    )

    br = layout["balance_row"]
    cfr = layout["cant_final_row"]
    ws.cell(row=br, column=1, value="Balance").font = font_bold
    ws.cell(row=br, column=1).border = border_tbl
    ws.cell(row=cfr, column=1, value="Cant Final").font = font_bold
    ws.cell(row=cfr, column=1).border = border_tbl

    cant_final_ref = f"B{cfr}"
    if cant_cells:
        acta_sum = "+".join(cant_cells)
        ws.cell(row=br, column=2, value=f"={acta_sum}-({cant_final_ref})")
    else:
        ws.cell(row=br, column=2, value=0)
    ws.cell(row=br, column=2).number_format = fmt_cant
    ws.cell(row=br, column=2).alignment = al_right
    ws.cell(row=br, column=2).border = border_tbl

    ws.cell(row=cfr, column=2, value=cant_final_cc)
    ws.cell(row=cfr, column=2).number_format = fmt_cant
    ws.cell(row=cfr, column=2).alignment = al_right
    ws.cell(row=cfr, column=2).border = border_tbl

    hdr_r = layout["detail_hdr"]
    for j, h in enumerate(_ITEM_DET_HEADERS, start=1):
        cell = ws.cell(row=hdr_r, column=j, value=h)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = al_center
        cell.border = border_tbl

    fill_balance = PatternFill("solid", fgColor="FFF7ED")
    _C_NUMERIC = (6, 7, 8, 9, 10)

    if not detail_rows:
        ws.merge_cells(start_row=df, start_column=1, end_row=df, end_column=_ITEM_XLSX_NCOLS)
        c0 = ws.cell(row=df, column=1, value="Sin registros cobrados (SICOE aprobado) para este ítem.")
        c0.font = Font(italic=True, color="666666")
        c0.alignment = al_center
    else:
        for i, drow in enumerate(detail_rows):
            ri = df + i
            is_adj = bool(drow.get("is_balance_adj"))
            vals = (
                drow.get("registro"),
                drow.get("acta_rpo"),
                drow.get("pk_id"),
                drow.get("abs_inicio"),
                drow.get("abs_final"),
                drow.get("longitud"),
                drow.get("ancho"),
                drow.get("espesor"),
                drow.get("cantidad") if not is_adj else "",
                drow.get("cant_total") if not is_adj else drow.get("cant_total"),
                drow.get("observaciones"),
            )
            zfill = fill_balance if is_adj else (fill_zebra if i % 2 else fill_white)
            for j, val in enumerate(vals, start=1):
                cell = ws.cell(row=ri, column=j, value=val if val != "" else None)
                cell.border = border_tbl
                cell.fill = zfill
                if is_adj and j != _C_CANT_TOTAL and j != _C_OBS:
                    continue
                if j in _C_NUMERIC and val not in ("", None):
                    cell.number_format = fmt_cant
                    cell.alignment = al_right
                elif j == _C_OBS:
                    cell.alignment = al_left_top
                elif not is_adj:
                    cell.alignment = al_left if j <= 3 else al_center

    tr = layout["total_row"]
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=_C_CANT_TOTAL - 1)
    ws.cell(row=tr, column=1, value="TOTAL ÍTEM").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=tr, column=1).alignment = al_left
    if detail_rows:
        tot_cell = ws.cell(row=tr, column=_C_CANT_TOTAL, value=f"=SUM({L(_C_CANT_TOTAL)}{df}:{L(_C_CANT_TOTAL)}{dl})")
    else:
        tot_cell = ws.cell(row=tr, column=_C_CANT_TOTAL, value=cant_final_cc)
    tot_cell.number_format = fmt_cant
    tot_cell.alignment = al_right
    for c in range(1, _ITEM_XLSX_NCOLS + 1):
        cell = ws.cell(row=tr, column=c)
        cell.fill = fill_tot
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.border = border_tbl

    table_top = 5
    table_bottom = tr
    main._xlsx_apply_outer_double_border(ws, table_top, table_bottom, 1, _ITEM_XLSX_NCOLS)

    footer = tr + 2
    ws.merge_cells(start_row=footer, start_column=1, end_row=footer, end_column=_ITEM_XLSX_NCOLS)
    ws.cell(
        row=footer,
        column=1,
        value="ClaraCore · Producto de gestión de obra · Todos los derechos reservados · claracore.co",
    ).font = Font(size=8, color="64748B", italic=True)
    ws.cell(row=footer, column=1).alignment = al_center
    ws.print_area = f"A1:{L(_ITEM_XLSX_NCOLS)}{footer}"


def build_dashboard_capitulo_completo_xlsx(
    main,
    contrato_id: int,
    capitulo: str,
    vista: str,
    current_user,
    *,
    items_sorted: List[str],
    by_item: Dict[str, List[dict]],
    item_filtro: Optional[str] = None,
    job_id: Optional[str] = None,
) -> Tuple[bytes, str]:
    from openpyxl import load_workbook

    cap_raw = (capitulo or "").strip()

    if job_id:
        main._export_job_set(job_id, progreso="resumen ejecutivo…")

    resumen_bytes, _ = main._build_dashboard_capitulo_resumen_ejecutivo_xlsx(
        contrato_id, cap_raw, vista, current_user, item_filtro=item_filtro
    )
    wb = load_workbook(io.BytesIO(resumen_bytes))

    drill_items = main._drill_agg_by_item(contrato_id, cap_raw, item_filtro, vista, current_user)
    drill_map = {main._dash_norm_item_key_py(r.get("item")): r for r in drill_items}

    item_labels: Dict[str, str] = {main._dash_norm_item_key_py(it): it for it in items_sorted}
    for r in drill_items:
        raw = (r.get("item") or "").strip()
        ik = main._dash_norm_item_key_py(raw)
        if ik and ik not in item_labels:
            item_labels[ik] = raw
    items_for_sheets = sorted(item_labels.values(), key=lambda x: str(x))

    meta = main._xlsx_load_contrato_export_meta(contrato_id)
    version_lbl = main._contrato_presupuesto_version_label(contrato_id)
    gen_ts = datetime.now(pytz.timezone("America/Bogota")).strftime("%d/%m/%Y %H:%M")
    descargado_por = main._calculo_usuario_label(current_user)
    listado_idx = main._listado_precios_index_por_item_norm(contrato_id, cap_raw)

    n_items = len(items_for_sheets)
    for idx, it in enumerate(items_for_sheets):
        if job_id:
            main._export_job_set(
                job_id,
                progreso=f"ítem {idx + 1}/{n_items}: {it[:40]}…",
            )
        drill = drill_map.get(main._dash_norm_item_key_py(it), {})
        listado_row = listado_idx.get(main._dash_norm_item_key_py(it))
        ppto_it = by_item.get(it) or []
        if not ppto_it:
            for k, rows in by_item.items():
                if main._dash_norm_item_key_py(k) == main._dash_norm_item_key_py(it):
                    ppto_it = rows
                    break
        sicoe_it = main._export_sicoe_obra_rows_capitulo(contrato_id, cap_raw, [it])
        sicoe_it = main._sicoe_registros_xlsx_sin_ids_internos(
            main._sicoe_enriquecer_registros_export(sicoe_it)
        )

        cant_final = round(float(drill.get("total_claracore_cant") or 0), 4)
        detail_rows = _build_sicoe_detail_rows(sicoe_it, cant_final)

        ws = wb.create_sheet(title=main._xlsx_safe_sheet_name(it, f"Item{idx + 1}"))
        _write_item_sheet_completo(
            main,
            ws,
            cap_raw=cap_raw,
            item=it,
            drill=drill,
            ppto_rows=ppto_it,
            detail_rows=detail_rows,
            sicoe_rows=sicoe_it,
            vista=vista,
            meta=meta,
            version_lbl=version_lbl,
            gen_ts=gen_ts,
            descargado_por=descargado_por,
            listado_row=listado_row,
        )

    for sh in wb.worksheets:
        sh.sheet_view.showGridLines = False

    if job_id:
        main._export_job_set(job_id, progreso="guardando archivo…")
    bio = io.BytesIO()
    wb.save(bio)
    safe_cap = re.sub(r"[^\w\-.]+", "_", cap_raw)[:40]
    fn = f"ClaraCore_{safe_cap}_{datetime.now(pytz.timezone('America/Bogota')).strftime('%Y-%m-%d')}.xlsx"
    return bio.getvalue(), fn
